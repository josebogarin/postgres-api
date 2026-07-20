"""
comparar_tbl_check.py  — sesion 59+
Compara puntajes del TBL MASTER (Excel basico) vs puntaje_detalle en BD.
Cubre TODOS los partidos finalizados en el Excel (grupos + R32 + lo que haya).
Items: H, I, J, K, L, M, N, O

Uso:
  cd "C:\proyecto FAST API"
  backend\.venv\Scripts\python.exe comparar_tbl_check.py "ruta\al\TBL CHECK.xlsx"
  (o sin argumento: busca el mas reciente en uploads/)
"""
import sys, pathlib, collections, datetime, re
import openpyxl, psycopg2
from openpyxl.styles import PatternFill, Font, Alignment

# ── Config BD ──────────────────────────────────────────────────────────────
BECBUC_DB = dict(host="localhost", port=5432, dbname="becbuc",
                 user="app_user", password="superpassword")
APP_DB    = dict(host="localhost", port=5432, dbname="app_db",
                 user="app_user", password="superpassword")
TORNEO_ID = 2

# ── Colores ──────────────────────────────────────────────────────────────
F_OK    = PatternFill("solid", fgColor="C6EFCE")   # verde  BD == Excel
F_PLUS  = PatternFill("solid", fgColor="C6EFCE")   # verde  BD > Excel
F_MINUS = PatternFill("solid", fgColor="FFC7CE")   # rojo   BD < Excel
F_HDR   = PatternFill("solid", fgColor="2B4DA1")
F_SUB   = PatternFill("solid", fgColor="4472C4")
F_GRAY  = PatternFill("solid", fgColor="D9D9D9")
WH      = Font(color="FFFFFF", bold=True)
BOLD    = Font(bold=True)
ITEMS   = ["H","I","J","K","L","M","N","O"]

def iv(v):
    try:    return int(v)
    except: return 0

# ── 1. Buscar Excel ─────────────────────────────────────────────────────────
excel_path = sys.argv[1] if len(sys.argv) > 1 else None
if not excel_path:
    uploads = (pathlib.Path(r"C:\Users\Jose Bogarin\AppData\Roaming\Claude"
                            r"\local-agent-mode-sessions")
               / "a9fdc79d-9227-450c-a0c1-27eafc601471"
               / "dfc0381f-d9d1-4349-b3fa-24cab5c5da8b"
               / "local_d09d3b3b-3380-4f77-80fa-069772ec423b" / "uploads")
    matches = sorted(uploads.glob("*TBL*CHECK*.xlsx")) if uploads.exists() else []
    if not matches:
        matches = sorted(uploads.glob("*TBL*.xlsx")) if uploads.exists() else []
    excel_path = str(matches[-1]) if matches else None

if not excel_path or not pathlib.Path(excel_path).exists():
    sys.exit("❌  No se encontro el Excel. Pasalo como argumento:\n"
             '   backend\\.venv\\Scripts\\python.exe comparar_tbl_check.py "ruta.xlsx"')

print(f"📂  {excel_path}")

# ── 2. Leer TBL MASTER ─────────────────────────────────────────────────────
wb   = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
ws   = wb["50- TBL MASTER"]

# excel_data[nf][alias] = {H,I,J,K,L,M,N,O,TOTAL}
excel_data   = collections.defaultdict(dict)
pids_pend    = set()

for row in ws.iter_rows(min_row=2, values_only=True):
    pid = row[1]
    if not pid or not str(pid).startswith("P"):
        continue
    try:
        nf = int(str(pid)[1:])
    except:
        continue

    estado = (row[54] or "").strip().upper()
    if estado != "FINALIZADO":
        pids_pend.add(nf)
        continue

    alias = str(row[9]).strip().lstrip("\xa0").lstrip().lower()
    if not alias:
        continue

    H  = iv(row[39]); I  = iv(row[40])
    J  = iv(row[41]); K  = iv(row[42])
    L  = iv(row[43]); M  = iv(row[44])
    N  = iv(row[45])
    O  = iv(row[46]) + iv(row[47])
    T  = iv(row[53])
    excel_data[nf][alias] = dict(H=H,I=I,J=J,K=K,L=L,M=M,N=N,O=O,TOTAL=T)

wb.close()

nf_fin = sorted(excel_data.keys())
all_aliases_ex = sorted({a for d in excel_data.values() for a in d})
print(f"✅  Excel: {len(nf_fin)} partidos finalizados  |  {len(all_aliases_ex)} apostadores")
print(f"    Rango: P{nf_fin[0]:03d} – P{nf_fin[-1]:03d}  |  pendientes: {sorted(pids_pend)[:6]}...")

# ── 3. Leer BD ──────────────────────────────────────────────────────────────
print("\n🔗  Conectando BD…")
conn_b = psycopg2.connect(**BECBUC_DB)
conn_a = psycopg2.connect(**APP_DB)
cur_b  = conn_b.cursor()
cur_a  = conn_a.cursor()

# alias map (apostador_id -> username)
cur_a.execute("SELECT id, LOWER(username) FROM users WHERE username IS NOT NULL")
id2alias = {r[0]: r[1] for r in cur_a.fetchall()}

# numero_fifa -> partido.id
cur_b.execute("""
    SELECT p.id, p.numero_fifa
    FROM partido p JOIN fase f ON p.fase_id=f.id
    WHERE f.torneo_id=%s AND p.numero_fifa IS NOT NULL
""", (TORNEO_ID,))
nf2pid = {r[1]: r[0] for r in cur_b.fetchall()}   # numero_fifa -> partido_db_id
pid2nf = {v: k for k, v in nf2pid.items()}

# Filtrar solo los partidos del scope
scope_pids = [nf2pid[nf] for nf in nf_fin if nf in nf2pid]
if not scope_pids:
    sys.exit("❌  No se encontraron los partidos en BD. Verificar numero_fifa.")

pids_sql = ",".join(str(p) for p in scope_pids)
cur_b.execute(f"""
    SELECT apostador_id, partido_id,
           COALESCE(pts_resultado,0),
           COALESCE(pts_marcador,0),
           COALESCE(pts_amarillas,0),
           COALESCE(pts_rojas,0),
           COALESCE(pts_var,0),
           COALESCE(pts_penales_partido,0),
           COALESCE(pts_minuto,0),
           COALESCE(pts_penales_tanda,0)
    FROM puntaje_detalle
    WHERE partido_id IN ({pids_sql})
""")

# bd_data[nf][alias] = {H,I,J,K,L,M,N,O,TOTAL}
bd_data = collections.defaultdict(dict)
for r in cur_b.fetchall():
    aid, part_id, H, I, J, K, L, M, N, O = r
    alias = id2alias.get(aid, f"id_{aid}")
    nf    = pid2nf.get(part_id)
    if nf is None: continue
    bd_data[nf][alias] = dict(H=H,I=I,J=J,K=K,L=L,M=M,N=N,O=O,TOTAL=H+I+J+K+L+M+N+O)

conn_b.close(); conn_a.close()

all_aliases_bd = sorted({a for d in bd_data.values() for a in d})
print(f"✅  BD:    {len(scope_pids)} partidos en scope  |  {len(all_aliases_bd)} apostadores")

# ── 4. Comparar ─────────────────────────────────────────────────────────────
# Unir aliases (algunos pueden diferir en minúsculas/espacios)
all_aliases = sorted(set(all_aliases_ex) | set(all_aliases_bd))

diffs = []   # [{nf, pid, alias, item, excel, bd, diff}]

alias_totals = {}   # alias -> {ex_T, bd_T, diff_T, per_item:{}}

for alias in all_aliases:
    item_ex  = {i: 0 for i in ITEMS}
    item_bd  = {i: 0 for i in ITEMS}
    for nf in nf_fin:
        ex = excel_data[nf].get(alias, {})
        bd = bd_data[nf].get(alias, {})
        for item in ITEMS:
            ev = ex.get(item, 0)
            bv = bd.get(item, 0)
            item_ex[item] += ev
            item_bd[item] += bv
            if ev != bv:
                diffs.append(dict(nf=nf, pid=f"P{nf:03d}", alias=alias,
                                  item=item, excel=ev, bd=bv, diff=bv-ev))

    ex_T = sum(item_ex.values())
    bd_T = sum(item_bd.values())
    alias_totals[alias] = dict(
        ex=ex_T, bd=bd_T, diff=bd_T-ex_T,
        per_item={i: item_bd[i]-item_ex[i] for i in ITEMS}
    )

# ── 5. Consola ──────────────────────────────────────────────────────────────
print(f"\n{'─'*80}")
print(f"  RESUMEN  (partidos scope: P{nf_fin[0]:03d}–P{nf_fin[-1]:03d})")
print(f"{'─'*80}")
print(f"  Total diffs items: {len(diffs)}")

from collections import Counter
item_cnt = Counter(d['item'] for d in diffs)
print(f"  Difs por item: " + "  ".join(f"{i}={item_cnt.get(i,0)}" for i in ITEMS))

n_ok   = sum(1 for t in alias_totals.values() if t['diff']==0)
n_diff = len(alias_totals) - n_ok
print(f"  Apostadores OK (TOTAL=0): {n_ok}/{len(alias_totals)}")
print(f"  Apostadores con diff:     {n_diff}/{len(alias_totals)}")

print(f"\n{'Apostador':<25} {'Excel':>7} {'BD':>7} {'DIFF':>7}")
print("  " + "─"*50)
for alias in sorted(alias_totals, key=lambda a: alias_totals[a]['diff']):
    t = alias_totals[alias]
    if t['diff'] != 0:
        print(f"  {alias:<23} {t['ex']:>7} {t['bd']:>7} {t['diff']:>+7}")

# ── 6. Excel de salida ──────────────────────────────────────────────────────
ts      = datetime.datetime.now().strftime("%Y%m%d_%H%M")
outname = f"becbuc_comparacion_tbl_{ts}.xlsx"
outwb   = openpyxl.Workbook()

# ── Hoja 1: RESUMEN ──────────────────────────────────────────────────────────
ws1 = outwb.active
ws1.title = "1-RESUMEN"

# Titulo
ws1.merge_cells("A1:I1")
c = ws1.cell(1,1, f"COMPARACIÓN TBL CHECK vs BD BECBUC  —  P{nf_fin[0]:03d}–P{nf_fin[-1]:03d}")
c.fill = F_HDR; c.font = WH; c.alignment = Alignment(horizontal="center")
ws1.merge_cells("A2:I2")
c = ws1.cell(2,1, f"Generado: {datetime.datetime.now():%Y-%m-%d %H:%M}  |  "
             f"Partidos finalizados: {len(nf_fin)}  |  Apostadores: {len(all_aliases)}")
c.fill = F_GRAY; c.alignment = Alignment(horizontal="center")

hdr = ["Pos","Apostador","Excel TOTAL","BD TOTAL","DIFF"] + [f"Δ{i}" for i in ITEMS]
for c_i, h in enumerate(hdr,1):
    cell = ws1.cell(4,c_i,h)
    cell.fill = F_SUB; cell.font = WH; cell.alignment = Alignment(horizontal="center")

sorted_aliases = sorted(alias_totals, key=lambda a: -alias_totals[a]['bd'])
for pos, alias in enumerate(sorted_aliases,1):
    t   = alias_totals[alias]
    row = [pos, alias, t['ex'], t['bd'], t['diff']]
    row += [t['per_item'][i] for i in ITEMS]
    for c_i, v in enumerate(row,1):
        cell = ws1.cell(4+pos, c_i, v)
        cell.alignment = Alignment(horizontal="center" if c_i!=2 else "left")
        if c_i == 5:
            if v > 0:   cell.fill = F_PLUS
            elif v < 0: cell.fill = F_MINUS
        elif c_i > 5:
            if v > 0:   cell.fill = F_PLUS
            elif v < 0: cell.fill = F_MINUS

ws1.column_dimensions["A"].width = 5
ws1.column_dimensions["B"].width = 26
for col in "CDEFGHIJKLMN":
    ws1.column_dimensions[col].width = 9

# ── Hoja 2: DIFERENCIAS por partido ──────────────────────────────────────────
ws2 = outwb.create_sheet("2-DIFERENCIAS")
hdr2 = ["Apostador","Partido","Item","Excel","BD","Diff"]
for c_i,h in enumerate(hdr2,1):
    cell=ws2.cell(1,c_i,h); cell.fill=F_HDR; cell.font=WH
    cell.alignment=Alignment(horizontal="center")

for r_i, d in enumerate(sorted(diffs, key=lambda x:(x['alias'],x['nf'],x['item'])),2):
    ws2.cell(r_i,1,d['alias'])
    ws2.cell(r_i,2,d['pid'])
    ws2.cell(r_i,3,d['item'])
    ws2.cell(r_i,4,d['excel'])
    ws2.cell(r_i,5,d['bd'])
    dc = ws2.cell(r_i,6,d['diff'])
    if d['diff']>0:   dc.fill=F_PLUS
    elif d['diff']<0: dc.fill=F_MINUS

for col,w in zip("ABCDEF",[26,8,6,8,8,8]):
    ws2.column_dimensions[col].width=w

# ── Hoja 3: MATRIZ items (alias × Δitem) ─────────────────────────────────────
ws3 = outwb.create_sheet("3-MATRIZ-ITEMS")
hdr3 = ["Apostador"] + [f"Δ{i}" for i in ITEMS] + ["ΔTOTAL"]
for c_i,h in enumerate(hdr3,1):
    cell=ws3.cell(1,c_i,h); cell.fill=F_SUB; cell.font=WH
    cell.alignment=Alignment(horizontal="center")

for r_i, alias in enumerate(sorted_aliases,2):
    t = alias_totals[alias]
    ws3.cell(r_i,1,alias)
    for c_i, item in enumerate(ITEMS,2):
        v = t['per_item'][item]
        cell = ws3.cell(r_i,c_i,v); cell.alignment=Alignment(horizontal="center")
        if v>0: cell.fill=F_PLUS
        elif v<0: cell.fill=F_MINUS
    dc = ws3.cell(r_i,len(ITEMS)+2, t['diff'])
    dc.alignment=Alignment(horizontal="center"); dc.font=BOLD
    if t['diff']>0: dc.fill=F_PLUS
    elif t['diff']<0: dc.fill=F_MINUS

ws3.column_dimensions["A"].width=26
for c_i in range(2, len(hdr3)+2):
    ws3.column_dimensions[chr(64+c_i)].width=9

# ── Hoja 4: DETALLE por partido (solo diferencias) ────────────────────────────
ws4 = outwb.create_sheet("4-DETALLE-DIFFS")
hdr4 = ["Partido","Fase","Alias","Item","Excel","BD","Diff"]
for c_i,h in enumerate(hdr4,1):
    cell=ws4.cell(1,c_i,h); cell.fill=F_HDR; cell.font=WH
    cell.alignment=Alignment(horizontal="center")

# Determinar fase por numero_fifa
def fase_de(nf):
    if nf <= 72: return "Grupos"
    if nf <= 88: return "R32"
    if nf <= 96: return "R16"
    if nf <= 100: return "4tos"
    if nf <= 102: return "Semis"
    return "Final/3P"

for r_i, d in enumerate(sorted(diffs, key=lambda x:(x['nf'],x['alias'],x['item'])),2):
    ws4.cell(r_i,1,d['pid'])
    ws4.cell(r_i,2,fase_de(d['nf']))
    ws4.cell(r_i,3,d['alias'])
    ws4.cell(r_i,4,d['item'])
    ws4.cell(r_i,5,d['excel'])
    ws4.cell(r_i,6,d['bd'])
    dc = ws4.cell(r_i,7,d['diff'])
    if d['diff']>0: dc.fill=F_PLUS
    elif d['diff']<0: dc.fill=F_MINUS

for col,w in zip("ABCDEFG",[8,8,26,6,8,8,8]):
    ws4.column_dimensions[col].width=w

outwb.save(outname)
print(f"\n💾  Excel: {outname}")
print("✅  Listo.")
