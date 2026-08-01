"""
comparar_excel_vs_bd_full.py
Compara BECBUC_verificacion.xlsx (Jun-23) contra puntaje_detalle actual.
Reporta diferencias por apostador.
"""
import os as _osp
_BASE = _osp.path.dirname(_osp.path.abspath(__file__))
import subprocess, sys
from datetime import datetime

EXCEL_PATH = _osp.path.join(_BASE, 'BECBUC_verificacion.xlsx')
LOG_PATH   = _osp.path.join(_BASE, 'comparar_excel_bd_log.txt')

def psql(sql, db="becbuc"):
    cmd = ["docker", "exec", "core-postgres", "psql",
           "-U", "app_user", "-d", db,
           "--tuples-only", "--no-align", "--field-separator=|", "-c", sql]
    r = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace")
    rows = [l.strip() for l in r.stdout.strip().splitlines() if l.strip()]
    return rows

def read_excel_ranking(path):
    """Lee hoja 🏆 Ranking del Excel. Retorna dict username -> {total, partidos, globales}"""
    try:
        import openpyxl
    except ImportError:
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
        import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # Buscar hoja ranking
    sheet = None
    for name in wb.sheetnames:
        if "ranking" in name.lower() or "🏆" in name:
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.worksheets[0]

    data = {}
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        if row[0] is None:
            break
        # Pos | Apostador | Total | Partidos | Globales | ...
        username = str(row[1]).strip().lower() if row[1] else ""
        total    = int(row[2]) if row[2] is not None else 0
        partidos = int(row[3]) if row[3] is not None else 0
        globales = int(row[4]) if row[4] is not None else 0
        data[username] = {"total": total, "partidos": partidos, "globales": globales, "pos_excel": int(row[0])}
    wb.close()
    return data

def read_bd_ranking():
    """Lee puntaje_detalle + puntaje_global + usernames de BD actual."""
    sql = """
    WITH pd_agg AS (
        SELECT apostador_id,
               COALESCE(SUM(pts_resultado),0)::INT       AS pts_H,
               COALESCE(SUM(pts_marcador),0)::INT        AS pts_I,
               COALESCE(SUM(pts_amarillas),0)::INT       AS pts_J,
               COALESCE(SUM(pts_rojas),0)::INT           AS pts_K,
               COALESCE(SUM(pts_var),0)::INT             AS pts_L,
               COALESCE(SUM(pts_penales_partido),0)::INT AS pts_M,
               COALESCE(SUM(pts_minuto),0)::INT          AS pts_N,
               COALESCE(SUM(pts_penales_tanda),0)::INT   AS pts_O,
               (COALESCE(SUM(pts_resultado),0)+
                COALESCE(SUM(pts_marcador),0)+
                COALESCE(SUM(pts_amarillas),0)+
                COALESCE(SUM(pts_rojas),0)+
                COALESCE(SUM(pts_var),0)+
                COALESCE(SUM(pts_penales_partido),0)+
                COALESCE(SUM(pts_minuto),0)+
                COALESCE(SUM(pts_penales_tanda),0))::INT AS total_partidos
        FROM puntaje_detalle
        WHERE torneo_id = 2
        GROUP BY apostador_id
    ),
    pg_agg AS (
        SELECT apostador_id,
               (COALESCE(pts_campeon,0)+COALESCE(pts_finalistas,0)+
                COALESCE(pts_goleador,0)+COALESCE(pts_peor_equipo,0)+
                COALESCE(pts_mayor_goleada,0)+COALESCE(pts_etapa_paraguay,0)+
                COALESCE(pts_goles_paraguay,0))::INT AS globales
        FROM puntaje_global
        WHERE torneo_id = 2
    ),
    nombres AS (
        SELECT DISTINCT ON (apostador_id) apostador_id, nombre_apostador
        FROM apuesta
        WHERE nombre_apostador IS NOT NULL
        ORDER BY apostador_id, id DESC
    )
    SELECT pd.apostador_id,
           LOWER(TRIM(n.nombre_apostador)) AS nombre,
           pd.total_partidos,
           COALESCE(pg.globales, 0) AS globales,
           (pd.total_partidos + COALESCE(pg.globales, 0)) AS total,
           pd.pts_H, pd.pts_I, pd.pts_J, pd.pts_K,
           pd.pts_L, pd.pts_M, pd.pts_N, pd.pts_O
    FROM pd_agg pd
    LEFT JOIN pg_agg pg ON pg.apostador_id = pd.apostador_id
    LEFT JOIN nombres n ON n.apostador_id = pd.apostador_id
    ORDER BY total DESC
    """
    rows = psql(sql)
    data = {}
    for r in rows:
        parts = r.split("|")
        if len(parts) < 13:
            continue
        aid = int(parts[0])
        nombre = parts[1].strip()
        data[aid] = {
            "nombre":         nombre,
            "total_partidos": int(parts[2] or 0),
            "globales":       int(parts[3] or 0),
            "total":          int(parts[4] or 0),
            "H": int(parts[5] or 0),
            "I": int(parts[6] or 0),
            "J": int(parts[7] or 0),
            "K": int(parts[8] or 0),
            "L": int(parts[9] or 0),
            "M": int(parts[10] or 0),
            "N": int(parts[11] or 0),
            "O": int(parts[12] or 0),
        }
    return data

def get_usernames():
    """Trae username -> apostador_id desde app_db via dblink."""
    sql = """
    SELECT u.uid::text, LOWER(TRIM(u.username)), LOWER(TRIM(u.nombre))
    FROM dblink('dbname=app_db user=app_user',
        'SELECT id, username, COALESCE(nombre,username) FROM users'
    ) AS u(uid INT, username TEXT, nombre TEXT)
    ORDER BY u.uid
    """
    rows = psql(sql)
    mapping = {}  # username -> apostador_id, nombre -> apostador_id
    for r in rows:
        parts = r.split("|")
        if len(parts) < 3:
            continue
        aid      = int(parts[0])
        username = parts[1].strip()
        nombre   = parts[2].strip()
        mapping[username] = aid
        if nombre not in mapping:
            mapping[nombre] = aid
    return mapping

def normalize(s):
    return s.lower().strip().replace(" ", "")

lines = []
def out(msg=""):
    print(msg)
    lines.append(msg)

def main():
    out(f"=== COMPARACION EXCEL vs BD - {datetime.now():%Y-%m-%d %H:%M:%S} ===")
    out()

    # 1. Leer Excel
    out("Leyendo Excel...")
    excel = read_excel_ranking(EXCEL_PATH)
    out(f"  Excel: {len(excel)} apostadores")

    # 2. Leer BD
    out("Consultando BD (puntaje_detalle + puntaje_global)...")
    bd = read_bd_ranking()
    out(f"  BD:    {len(bd)} apostadores con puntaje")

    # 3. Mapear username -> apostador_id
    out("Obteniendo usernames...")
    umap = get_usernames()

    # 4. Construir mapa username->bd_data
    bd_by_username = {}
    for aid, row in bd.items():
        nombre = row["nombre"]
        bd_by_username[nombre] = row
        bd_by_username[aid] = row

    # Agregar por username de app_db
    for uname, aid in umap.items():
        if aid in bd:
            bd_by_username[uname] = bd[aid]

    # 5. Comparar
    out()
    out("=" * 80)
    out(f"{'Apostador':<24} {'ExcelPart':>10} {'ExcelGlob':>10} {'ExcelTot':>9} {'BDPart':>7} {'BDGlob':>7} {'BDTot':>7} {'DIFF':>7}")
    out("=" * 80)

    ok_count   = 0
    diff_count = 0
    no_match   = []
    diffs_list = []

    for uname, ex_row in sorted(excel.items(), key=lambda x: x[1]["pos_excel"]):
        # Buscar en BD
        bd_row = bd_by_username.get(uname)
        if bd_row is None:
            # Intentar matching parcial
            for k, v in bd_by_username.items():
                if isinstance(k, str) and (normalize(uname) in normalize(k) or normalize(k) in normalize(uname)):
                    bd_row = v
                    break

        if bd_row is None:
            no_match.append(uname)
            out(f"  {uname:<24} Excel={ex_row['total']:>6}  -- SIN MATCH EN BD --")
            continue

        ex_total = ex_row["total"]
        bd_total = bd_row["total"]
        diff     = bd_total - ex_total

        if diff == 0:
            ok_count += 1
            out(f"  {uname:<24} {ex_row['partidos']:>10} {ex_row['globales']:>10} {ex_total:>9} {bd_row['total_partidos']:>7} {bd_row['globales']:>7} {bd_total:>7} {'OK':>7}")
        else:
            diff_count += 1
            diff_str = f"{diff:+d}"
            out(f"* {uname:<24} {ex_row['partidos']:>10} {ex_row['globales']:>10} {ex_total:>9} {bd_row['total_partidos']:>7} {bd_row['globales']:>7} {bd_total:>7} {diff_str:>7}  <-- DIFF")
            diffs_list.append({
                "username": uname,
                "ex_partidos": ex_row["partidos"],
                "ex_globales": ex_row["globales"],
                "ex_total":    ex_total,
                "bd_partidos": bd_row["total_partidos"],
                "bd_globales": bd_row["globales"],
                "bd_total":    bd_total,
                "diff":        diff,
                "bd_H": bd_row["H"], "bd_I": bd_row["I"], "bd_J": bd_row["J"],
                "bd_K": bd_row["K"], "bd_L": bd_row["L"], "bd_M": bd_row["M"],
                "bd_N": bd_row["N"], "bd_O": bd_row["O"],
            })

    out("=" * 80)
    out()
    out(f"RESUMEN: {ok_count} OK  |  {diff_count} DIFFS  |  {len(no_match)} sin match")
    if no_match:
        out(f"Sin match: {', '.join(no_match)}")

    if diffs_list:
        out()
        out("--- DETALLE DE DIFERENCIAS ---")
        for d in diffs_list:
            out(f"  [{d['username']}]")
            out(f"    Excel: partidos={d['ex_partidos']}  globales={d['ex_globales']}  TOTAL={d['ex_total']}")
            out(f"    BD:    partidos={d['bd_partidos']}  globales={d['bd_globales']}  TOTAL={d['bd_total']}")
            diff_part = d['bd_partidos'] - d['ex_partidos']
            diff_glob = d['bd_globales'] - d['ex_globales']
            out(f"    Diff:  partidos={diff_part:+d}  globales={diff_glob:+d}  TOTAL={d['diff']:+d}")
            out(f"    BD desglose: H={d['bd_H']} I={d['bd_I']} J={d['bd_J']} K={d['bd_K']} L={d['bd_L']} M={d['bd_M']} N={d['bd_N']} O={d['bd_O']}")
            out()

        out()
        out("CONCLUSION: El Excel esta DESACTUALIZADO respecto a la BD.")
        out("Ejecutar generar_excel_becbuc.py para regenerar con puntajes actuales.")
    else:
        out()
        out("✅ EXCEL y BD coinciden 100% en puntajes totales.")

    with open(LOG_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    out(f"\nLog: {LOG_PATH}")

if __name__ == "__main__":
    main()
