"""
Script combinado (usa asyncpg, que ya esta instalado en el venv):
  1. Pone pts_mayor_goleada = 0 en puntaje_global (directo a BD)
  2. Genera Excel de puntaje por item - Fase de Grupos + Peor equipo (D)
Ejecutar: python goleada_off_y_excel_grupos.py
"""
import asyncio, sys, datetime
import io
# Force UTF-8 output on Windows console
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
import asyncpg
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DSN = "postgresql://app_user:superpassword@localhost:5432/becbuc"

HDR_BG = "1e3a5f"; HDR_FG = "FFFFFF"
ITEM_BG = "0f2840"; ITEM_FG = "93c5fd"
AP_BG = "0d1b2a"; AP_BG2 = "111f30"
POS_FG = "fbbf24"
GLOB_BG = "1a2e1a"; GLOB_FG = "4ade80"
TOTAL_BG = "292524"; TOTAL_FG = "fde68a"
ZERO = "374151"

def cfill(h): return PatternFill("solid", fgColor=h)
def thin():
    s = Side(style="thin", color="1e293b")
    return Border(left=s, right=s, top=s, bottom=s)

async def main():
    print("Conectando a PostgreSQL...")
    conn = await asyncpg.connect(DSN)

    # ── PASO 1: Cero a pts_mayor_goleada ─────────────────────────────
    print("Poniendo pts_mayor_goleada = 0...")
    rows_upd = await conn.execute("""
        UPDATE puntaje_global
        SET pts_total = pts_total - pts_mayor_goleada,
            pts_mayor_goleada = 0
        WHERE pts_mayor_goleada > 0 AND torneo_id = 2
    """)
    print(f"  ✓ Actualizados: {rows_upd}")

    # ── PASO 2: Apostadores ───────────────────────────────────────────
    print("Cargando datos de grupos...")
    apostadores = await conn.fetch("""
        SELECT DISTINCT
            pd.apostador_id,
            COALESCE(
                (SELECT nombre_apostador FROM apuesta a2
                 WHERE a2.apostador_id = pd.apostador_id LIMIT 1),
                pd.apostador_id::text
            ) AS alias
        FROM puntaje_detalle pd
        JOIN partido p ON p.id = pd.partido_id
        JOIN fase f ON f.id = p.fase_id
        WHERE f.torneo_id = 2 AND f.tipo ILIKE 'grupo%'
        ORDER BY alias
    """)
    if not apostadores:
        print("ERROR: No hay datos en fase de grupos.")
        await conn.close(); sys.exit(1)

    ap_ids = [r["apostador_id"] for r in apostadores]
    alias_map = {r["apostador_id"]: r["alias"] for r in apostadores}
    id_list = ",".join(str(i) for i in ap_ids)

    # ── PASO 3: Puntaje por item (grupos) ─────────────────────────────
    pts_raw = await conn.fetch(f"""
        SELECT
            pd.apostador_id,
            COALESCE(SUM(pd.pts_resultado),0)                   AS h,
            COALESCE(SUM(pd.pts_marcador),0)                    AS i,
            COALESCE(SUM(pd.pts_amarillas),0)                   AS j,
            COALESCE(SUM(COALESCE(pd.pts_rojas,0)),0)           AS k,
            COALESCE(SUM(pd.pts_var),0)                         AS l,
            COALESCE(SUM(COALESCE(pd.pts_penales_partido,0)),0) AS m,
            COALESCE(SUM(pd.pts_minuto),0)                      AS n,
            COALESCE(SUM(COALESCE(pd.pts_penales_tanda,0)),0)   AS o,
            COALESCE(SUM(COALESCE(pd.pts_equipo,0)),0)          AS p_ko
        FROM puntaje_detalle pd
        JOIN partido pt ON pt.id = pd.partido_id
        JOIN fase f ON f.id = pt.fase_id
        WHERE f.torneo_id = 2 AND f.tipo ILIKE 'grupo%'
          AND pd.apostador_id IN ({id_list})
        GROUP BY pd.apostador_id
    """)
    pts_rows = {r["apostador_id"]: dict(r) for r in pts_raw}

    # ── PASO 4: P grupos ──────────────────────────────────────────────
    try:
        grp_raw = await conn.fetch(f"""
            SELECT apostador_id, COALESCE(aciertos,0) AS aciertos
            FROM apostador_clasificados
            WHERE torneo_id = 2 AND fase_tipo='grupo'
              AND apostador_id IN ({id_list})
        """)
        grp_p = {r["apostador_id"]: r["aciertos"] for r in grp_raw}
    except Exception as e:
        print(f"  (apostador_clasificados no disponible: {e})")
        grp_p = {}

    # ── PASO 5: D (peor equipo) ───────────────────────────────────────
    glob_raw = await conn.fetch(f"""
        SELECT apostador_id, COALESCE(pts_peor_equipo,0) AS d
        FROM puntaje_global
        WHERE torneo_id = 2 AND apostador_id IN ({id_list})
    """)
    glob_d = {r["apostador_id"]: r["d"] for r in glob_raw}

    await conn.close()

    # ── PASO 6: Armar filas ───────────────────────────────────────────
    ITEMS  = ["H","I","J","K","L","M","N","O","P"]
    LABELS = {"H":"Resultado","I":"Marcador exacto","J":"Amarillas",
              "K":"Rojas","L":"VAR","M":"Penales juego",
              "N":"Minuto gol","O":"Pen. tanda","P":"Clasifica R32"}

    rows_data = []
    for aid in ap_ids:
        p  = pts_rows.get(aid, {})
        gp = grp_p.get(aid, 0)
        row = {
            "alias": alias_map[aid],
            "H": p.get("h",0), "I": p.get("i",0), "J": p.get("j",0),
            "K": p.get("k",0), "L": p.get("l",0), "M": p.get("m",0),
            "N": p.get("n",0), "O": p.get("o",0),
            "P": p.get("p_ko",0) + gp,
            "D": glob_d.get(aid, 0),
        }
        row["total_grupos"] = sum(row[k] for k in ITEMS)
        row["total_con_d"]  = row["total_grupos"] + row["D"]
        rows_data.append(row)

    rows_data.sort(key=lambda x: -x["total_con_d"])

    # ── PASO 7: Excel ─────────────────────────────────────────────────
    print("Generando Excel...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Puntaje Grupos"
    ws.sheet_view.showGridLines = False

    headers = ["#","Apostador"] + ITEMS + ["D - Peor equipo","Total Grupos","TOTAL (con D)"]
    d_col     = headers.index("D - Peor equipo") + 1
    tg_col    = headers.index("Total Grupos") + 1
    total_col = headers.index("TOTAL (con D)") + 1

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = cfill(HDR_BG); c.font = Font(color=HDR_FG, bold=True, size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin()
    ws.cell(row=1, column=d_col).fill = cfill(GLOB_BG)
    ws.cell(row=1, column=d_col).font = Font(color=GLOB_FG, bold=True, size=10, name="Calibri")
    ws.cell(row=1, column=total_col).fill = cfill(TOTAL_BG)
    ws.cell(row=1, column=total_col).font = Font(color=TOTAL_FG, bold=True, size=11, name="Calibri")
    ws.row_dimensions[1].height = 28

    sub = ["",""] + [LABELS[k] for k in ITEMS] + ["","",""]
    for ci, s in enumerate(sub, 1):
        c = ws.cell(row=2, column=ci, value=s)
        c.fill = cfill(ITEM_BG); c.font = Font(color=ITEM_FG, size=8, italic=True, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()
    ws.row_dimensions[2].height = 14

    for ri, row in enumerate(rows_data, 3):
        bg = AP_BG if ri % 2 == 1 else AP_BG2
        def wc(ci, value, bold=False, color="C8D3E0", bg_ovr=None, align="center", size=10):
            c = ws.cell(row=ri, column=ci, value=value)
            c.fill = cfill(bg_ovr or bg)
            c.font = Font(color=(color if value != 0 else ZERO), bold=bold, size=size, name="Calibri")
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border = thin()
        wc(1, ri-2, bold=True, color=POS_FG)
        wc(2, row["alias"], align="left", bold=True, color="e2e8f0")
        for ki, k in enumerate(ITEMS, 3):
            v = row[k]
            clr = ("38bdf8" if k in ("H","I") else "4ade80") if v > 0 else ZERO
            wc(ki, v, color=clr)
        v_d = row["D"]
        wc(d_col, v_d, bg_ovr=GLOB_BG, color=(GLOB_FG if v_d > 0 else ZERO), bold=(v_d > 0))
        wc(tg_col, row["total_grupos"], color="94a3b8")
        wc(total_col, row["total_con_d"], bg_ovr=TOTAL_BG, color=TOTAL_FG, bold=True, size=11)
        ws.row_dimensions[ri].height = 18

    tr = len(rows_data) + 3
    for c in range(1, len(headers)+1):
        cel = ws.cell(row=tr, column=c)
        cel.fill = cfill("1c1917"); cel.font = Font(color="fbbf24", bold=True, size=10, name="Calibri")
        cel.border = thin(); cel.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=tr, column=2, value="TOTAL TODOS")
    for ki, k in enumerate(ITEMS, 3):
        ws.cell(row=tr, column=ki, value=sum(r[k] for r in rows_data))
    ws.cell(row=tr, column=d_col,     value=sum(r["D"] for r in rows_data))
    ws.cell(row=tr, column=tg_col,    value=sum(r["total_grupos"] for r in rows_data))
    ws.cell(row=tr, column=total_col, value=sum(r["total_con_d"] for r in rows_data))
    ws.row_dimensions[tr].height = 20

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 22
    for i in range(3, len(ITEMS)+3):
        ws.column_dimensions[get_column_letter(i)].width = 7
    ws.column_dimensions[get_column_letter(d_col)].width = 12
    ws.column_dimensions[get_column_letter(tg_col)].width = 10
    ws.column_dimensions[get_column_letter(total_col)].width = 12

    ts  = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    out = rf"C:\proyecto FAST API\puntaje_grupos_{ts}.xlsx"
    wb.save(out)
    print(f"\nExcel generado: {out}")
    print(f"   Apostadores: {len(rows_data)}")
    print(f"   Columnas: # | Apostador | H I J K L M N O P | D Peor equipo | Total Grupos | TOTAL")

if __name__ == "__main__":
    asyncio.run(main())
