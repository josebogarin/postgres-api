# -*- coding: utf-8 -*-
r"""
exportar_diffs_fin_torneo.py  (solo lectura)

Recomputa la auditoria de origen (puntos por item: Excel de cierre vs BD) y la
EXPORTA a un Excel:  becbuc_diferencias_puntaje_<YYYYMMDD_HHMM>.xlsx

Hojas:
  - Diferencias : una fila por (partido x apostador x item) que difiere, con
                  real / ap.Excel / ap.BD / ptsExcel / ptsBD / ORIGEN.
  - Resumen     : matriz item x fase + conteo por ORIGEN.
  - Leyenda     : que significa cada ORIGEN.

Uso:
  backend\.venv\Scripts\python.exe exportar_diffs_fin_torneo.py
"""
import sys, os, datetime
from collections import Counter, defaultdict

PHASES = {'grupos': (1, 72, False), '16avos': (73, 88, True), 'octavos': (89, 96, True),
          'cuartos': (97, 100, True), 'semis': (101, 102, True), 'final3p': (103, 104, True)}
ORDER = ['grupos', '16avos', 'octavos', 'cuartos', 'semis', 'final3p']
NF_SET, fase_de_nf = set(), {}
for f in ORDER:
    a, b, _ = PHASES[f]
    for n in range(a, b + 1):
        NF_SET.add(n); fase_de_nf[n] = f
es_ko_de = {f: PHASES[f][2] for f in ORDER}

BASE = os.path.dirname(os.path.abspath(__file__))
_OK = ('TORNEO CERRADO', 'FIN DE TORNEO', '20260720'); _BAD = ('CORRECCIONES', 'SEMIFINAL')
_c = [os.path.join(BASE, f) for f in os.listdir(BASE)
      if f.lower().endswith('.xlsx') and not f.startswith('~')
      and any(k in f.upper() for k in _OK) and not any(b in f.upper() for b in _BAD)]
if not _c:
    sys.exit("ERROR: no encontre el Excel de cierre en la raiz.")
EXCEL = sorted(_c, key=lambda p: os.path.getmtime(p), reverse=True)[0]
print(f"Excel: {os.path.basename(EXCEL)}")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    os.system(f'"{sys.executable}" -m pip install openpyxl --quiet')
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
try:
    import psycopg2, psycopg2.extras
except ImportError:
    os.system(f'"{sys.executable}" -m pip install psycopg2-binary --quiet'); import psycopg2, psycopg2.extras

CONN_BEC = "host=localhost port=5432 dbname=becbuc user=app_user password=superpassword"
CONN_APP = "host=localhost port=5432 dbname=app_db  user=app_user password=superpassword"
TID = 2

conn = psycopg2.connect(CONN_BEC); capp = psycopg2.connect(CONN_APP)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
cua = capp.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

cua.execute("SELECT id, username FROM users WHERE is_active=TRUE")
uname = {u['username'].lower(): u['id'] for u in cua.fetchall()}
def clean_alias(s):
    if not s: return ''
    return str(s).replace('\xa0', '').strip().upper().lstrip('@').replace('  ', ' ')
def find_uid(al):
    a = clean_alias(al)
    if not a: return None
    for k, v in uname.items():
        if k.upper().lstrip('@') == a: return v
    for k, v in uname.items():
        if a in k.upper() or k.upper() in a: return v
    return None

cur.execute("SELECT id, nombre, nombre_es FROM equipo")
ename = {r['id']: (r['nombre'] or r['nombre_es']) for r in cur.fetchall()}
def enm(x): return ename.get(x, x) if x is not None else 'None'

cur.execute("""SELECT pd.apostador_id, p.numero_fifa,
       COALESCE(pd.pts_resultado,0) h, COALESCE(pd.pts_marcador,0) i,
       COALESCE(pd.pts_amarillas,0) j, COALESCE(pd.pts_rojas,0) k,
       COALESCE(pd.pts_var,0) l, COALESCE(pd.pts_penales_partido,0) m,
       COALESCE(pd.pts_minuto,0) n, COALESCE(pd.pts_penales_tanda,0) o,
       COALESCE(pd.pts_equipo,0) p
       FROM puntaje_detalle pd JOIN partido p ON p.id=pd.partido_id
       JOIN fase f ON f.id=p.fase_id WHERE f.torneo_id=%s""", (TID,))
bd_pts = {(r['apostador_id'], r['numero_fifa']): r for r in cur.fetchall()}

cur.execute("""SELECT a.apostador_id, p.numero_fifa,
       a.pred_local, a.pred_visitante, a.pred_amarillas, a.pred_rojas, a.pred_var,
       a.pred_penales_partido, a.pred_minuto_gol,
       a.pred_penales_local_tanda, a.pred_penales_visitante_tanda, a.pred_equipo_clasifica
       FROM apuesta a JOIN partido p ON p.id=a.partido_id
       JOIN fase f ON f.id=p.fase_id WHERE f.torneo_id=%s""", (TID,))
bd_pred = {(r['apostador_id'], r['numero_fifa']): r for r in cur.fetchall()}

cur.execute("""SELECT p.numero_fifa, p.goles_local, p.goles_visitante, p.amarillas, p.rojas,
       p.decisiones_var, p.penales_partido, p.minuto_primer_gol,
       p.penales_local, p.penales_visitante, p.equipo_clasificado_id
       FROM partido p JOIN fase f ON f.id=p.fase_id WHERE f.torneo_id=%s""", (TID,))
off = {r['numero_fifa']: r for r in cur.fetchall()}

def sc(v):
    try:
        s = str(v).strip()
        if s in ('', '-', 'None'): return 0
        return int(float(s))
    except Exception:
        return 0
def si(v):
    try:
        s = str(v).strip()
        if s in ('', '-', 'None'): return None
        return int(float(s))
    except Exception:
        return None

C = dict(pid=2, alias=10, pl=13, pv=15, predJ=25, predK=26, predL=27, predM=28, predN=29,
         predOe1=30, predOe2=31, predQ=32,
         ptsH=40, ptsI=41, ptsJ=42, ptsK=43, ptsL=44, ptsM=45, ptsN=46, ptsOe1=47, ptsOe2=48, ptsP=51)
ITEM_DESC = {'H': 'Resultado', 'I': 'Marcador', 'J': 'Amarillas', 'K': 'Rojas', 'L': 'VAR',
             'M': 'Penales juego', 'N': 'Minuto 1er gol', 'O': 'Tanda', 'P': 'Equipo pasa'}
ALL_ITEMS = ['H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']

wb_in = openpyxl.load_workbook(EXCEL, data_only=True)
SHEET = next((s for s in wb_in.sheetnames if 'MASTER' in s.upper()), None)
ws = wb_in[SHEET]
for _cc in range(45, 62):
    _h = ws.cell(1, _cc).value
    if _h and 'CLASIFICAD' in str(_h).upper():
        C['ptsP'] = _cc; break

def origen(it, real, pex, pbd, ptse, ptsb):
    if it in ('K', 'M') and ptsb > ptse:
        return "BD lenient null->0 (por diseno; BD correcta)"
    if it == 'N':
        return "desempate minuto (decision org; BD correcta)"
    if it == 'O':
        return "quirk Excel tanda (BD correcta)"
    if it == 'P':
        return "escala reglamento BD vs plano Excel (BD correcta)"
    if it in ('H', 'I'):
        return "error del Excel (marcador/placeholder)" if str(real) == str(pex) else "REVISAR marcador"
    return "REVISAR"

tot = Counter(); dif = Counter(); detalles = []
for r in range(2, ws.max_row + 1):
    raw = ws.cell(r, C['pid']).value
    pid = str(raw).strip() if raw is not None else ''
    if not (len(pid) == 4 and pid[0] in 'Pp' and pid[1:].isdigit()):
        continue
    nf = int(pid[1:])
    if nf not in NF_SET:
        continue
    fase = fase_de_nf[nf]; es_ko = es_ko_de[fase]
    uid = find_uid(ws.cell(r, C['alias']).value)
    if not uid:
        continue
    b = bd_pts.get((uid, nf))
    if not b:
        continue
    alias = clean_alias(ws.cell(r, C['alias']).value)
    pr = bd_pred.get((uid, nf), {}); of = off.get(nf, {})
    exP = {'H': sc(ws.cell(r, C['ptsH']).value), 'I': sc(ws.cell(r, C['ptsI']).value),
           'J': sc(ws.cell(r, C['ptsJ']).value), 'K': sc(ws.cell(r, C['ptsK']).value),
           'L': sc(ws.cell(r, C['ptsL']).value), 'M': sc(ws.cell(r, C['ptsM']).value),
           'N': sc(ws.cell(r, C['ptsN']).value),
           'O': sc(ws.cell(r, C['ptsOe1']).value) + sc(ws.cell(r, C['ptsOe2']).value),
           'P': sc(ws.cell(r, C['ptsP']).value)}
    marc_real = f"{of.get('goles_local')}-{of.get('goles_visitante')}"
    marc_ex = f"{si(ws.cell(r, C['pl']).value)}-{si(ws.cell(r, C['pv']).value)}"
    marc_bd = f"{pr.get('pred_local')}-{pr.get('pred_visitante')}"
    REAL = {'H': marc_real, 'I': marc_real, 'J': of.get('amarillas'), 'K': of.get('rojas'),
            'L': of.get('decisiones_var'), 'M': of.get('penales_partido'), 'N': of.get('minuto_primer_gol'),
            'O': f"{of.get('penales_local')}-{of.get('penales_visitante')}", 'P': enm(of.get('equipo_clasificado_id'))}
    PEX = {'H': marc_ex, 'I': marc_ex, 'J': si(ws.cell(r, C['predJ']).value), 'K': si(ws.cell(r, C['predK']).value),
           'L': si(ws.cell(r, C['predL']).value), 'M': si(ws.cell(r, C['predM']).value), 'N': si(ws.cell(r, C['predN']).value),
           'O': f"{si(ws.cell(r, C['predOe1']).value)}-{si(ws.cell(r, C['predOe2']).value)}",
           'P': (str(ws.cell(r, C['predQ']).value).strip() if ws.cell(r, C['predQ']).value not in (None, '', '-') else 'None')}
    PBD = {'H': marc_bd, 'I': marc_bd, 'J': pr.get('pred_amarillas'), 'K': pr.get('pred_rojas'),
           'L': pr.get('pred_var'), 'M': pr.get('pred_penales_partido'), 'N': pr.get('pred_minuto_gol'),
           'O': f"{pr.get('pred_penales_local_tanda')}-{pr.get('pred_penales_visitante_tanda')}",
           'P': enm(pr.get('pred_equipo_clasifica'))}
    for it in ALL_ITEMS:
        if it == 'P' and not es_ko:
            continue
        vb = b[it.lower()]; ve = exP[it]
        tot[(it, fase)] += 1
        if ve != vb:
            dif[(it, fase)] += 1
            detalles.append((fase, it, ITEM_DESC[it], nf, alias, REAL[it], PEX[it], PBD[it], ve, vb,
                             origen(it, REAL[it], PEX[it], PBD[it], ve, vb)))

# ---------- Excel de salida ----------
wb = openpyxl.Workbook()
hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="1F4E78")
bd_ok = PatternFill("solid", fgColor="E2EFDA")
rev_fill = PatternFill("solid", fgColor="FCE4D6")

ws1 = wb.active; ws1.title = "Diferencias"
cols = ["FASE", "PARTIDO", "APOSTADOR", "ITEM", "CONCEPTO", "RESULTADO REAL",
        "APUESTA (Excel)", "APUESTA (BD)", "PTS Excel", "PTS BD", "ORIGEN"]
ws1.append(cols)
for c in range(1, len(cols) + 1):
    cell = ws1.cell(1, c); cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = Alignment(horizontal="center")
for fase, it, desc, nf, alias, real, pex, pbd, ptse, ptsb, org in sorted(detalles, key=lambda x: (ORDER.index(x[0]), x[3], x[1], x[4])):
    ws1.append([fase, f"P{nf:03d}", alias, it, desc, str(real), str(pex), str(pbd), ptse, ptsb, org])
    rr = ws1.max_row
    fill = bd_ok if 'BD correcta' in org or 'BD lenient' in org else rev_fill
    for c in range(1, len(cols) + 1):
        ws1.cell(rr, c).fill = fill
widths = [10, 9, 18, 6, 16, 14, 14, 14, 9, 8, 44]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws1.freeze_panes = "A2"

ws2 = wb.create_sheet("Resumen")
ws2.append(["Matriz item x fase (diferencias / comparaciones)"])
ws2.append(["ITEM"] + ORDER + ["TOTAL"])
for c in range(1, len(ORDER) + 2):
    ws2.cell(2, c).font = Font(bold=True)
for it in ALL_ITEMS:
    row = [it]; td = 0; tt = 0
    for f in ORDER:
        t = tot[(it, f)]; d = dif[(it, f)]; td += d; tt += t
        row.append(f"{d}/{t}" if t else "-")
    row.append(f"{td}/{tt}" if tt else "-")
    ws2.append(row)
ws2.append([])
ws2.append(["Por ORIGEN", "cantidad"])
ws2.cell(ws2.max_row, 1).font = Font(bold=True)
for org, n in Counter(d[10] for d in detalles).most_common():
    ws2.append([org, n])
ws2.append([])
ws2.append([f"TOTAL diferencias: {len(detalles)}"])
ws2.column_dimensions['A'].width = 48
for i in range(2, len(ORDER) + 3):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 11

ws3 = wb.create_sheet("Leyenda")
leyenda = [
    ["Contexto", ""],
    ["Se alinearon PREDICCIONES y RESULTADOS de la BD con el Excel de cierre.", ""],
    ["Por eso las diferencias que quedan son SOLO del algoritmo de puntaje.", ""],
    ["", ""],
    ["ORIGEN", "Significado"],
    ["BD lenient null->0 (por diseno; BD correcta)",
     "Prediccion en blanco = 0 (regla de la organizacion). Si el resultado real es 0, la BD suma el punto; el Excel no. BD correcta."],
    ["escala reglamento BD vs plano Excel (BD correcta)",
     "Item P (equipo que clasifica): la BD escala por fase (16avos=2, 8vos=4, 4tos=6, semis=8, final=12); el Excel puntea plano. BD correcta."],
    ["desempate minuto (decision org; BD correcta)",
     "Item N: la BD da el punto a TODOS los empatados en la distancia minima (decision organizacion)."],
    ["quirk Excel tanda (BD correcta)",
     "Item O: el Excel otorgo tanda donde no hubo definicion por penales. BD=0, correcta."],
    ["error del Excel (marcador/placeholder)",
     "El Excel no punteo un marcador correcto. BD correcta."],
    ["REVISAR",
     "No entra en un patron conocido: mirar a mano."],
]
for row in leyenda:
    ws3.append(row)
ws3.cell(5, 1).font = Font(bold=True); ws3.cell(5, 2).font = Font(bold=True)
ws3.column_dimensions['A'].width = 46
ws3.column_dimensions['B'].width = 90

ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
OUT = os.path.join(BASE, f"becbuc_diferencias_puntaje_{ts}.xlsx")
wb.save(OUT)
print(f"\nOK. Generado: {os.path.basename(OUT)}")
print(f"Ruta: {OUT}")
print(f"Diferencias exportadas: {len(detalles)}")
conn.close(); capp.close()
