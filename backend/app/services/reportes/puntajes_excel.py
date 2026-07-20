# -*- coding: utf-8 -*-
"""
puntajes_excel.py — Export Excel de puntajes (resumen + detalle) (Fase 3).
Movido desde apostador_bets.py. Lee las vistas v_copamundial_puntajes(_det).
"""
from datetime import datetime as _dt

from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import text


async def build_puntajes_export(db, torneo_id: int):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    ALT_FILL  = PatternFill("solid", fgColor="DCE6F1")
    NORM_FILL = PatternFill("solid", fgColor="FFFFFF")
    TOT_FILL  = PatternFill("solid", fgColor="FFF2CC")
    TOT_FONT  = Font(bold=True, name="Calibri", size=10)
    PTS_FILL  = PatternFill("solid", fgColor="E2EFDA")
    PTS_FONT  = Font(name="Calibri", size=10, color="375623")
    CELL_FONT = Font(name="Calibri", size=10)
    CENTER    = Alignment(horizontal="center", vertical="center")
    LEFT      = Alignment(horizontal="left",   vertical="center")
    _s        = Side(style="thin", color="BBBBBB")
    BORDER    = Border(left=_s, right=_s, top=_s, bottom=_s)

    def _is_pts(col):  return col.endswith("_pts") or col in ("total_partido","total_puntos","subtotal_partidos","subtotal_globales")
    def _is_tot(col):  return col in ("total_partido","total_puntos","subtotal_partidos","subtotal_globales")

    def _write_sheet(ws, rows, title):
        ws.title = title
        if not rows:
            ws.cell(1, 1, "Sin datos")
            return
        cols = list(rows[0].keys())
        for c, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=col.replace("_", " ").title())
            cell.font = HDR_FONT; cell.fill = HDR_FILL
            cell.alignment = CENTER; cell.border = BORDER
        for r, row in enumerate(rows, 2):
            alt = (r % 2 == 0)
            for c, col in enumerate(cols, 1):
                val = row[col]
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = BORDER
                cell.font   = TOT_FONT if _is_tot(col) else (PTS_FONT if _is_pts(col) else CELL_FONT)
                cell.alignment = CENTER if isinstance(val, (int, float)) else LEFT
                cell.fill   = TOT_FILL if _is_tot(col) else (PTS_FILL if _is_pts(col) else (ALT_FILL if alt else NORM_FILL))
        # ancho automático
        for c, col in enumerate(cols, 1):
            max_w = max(len(col), max((len(str(r[col] or "")) for r in rows), default=0))
            ws.column_dimensions[get_column_letter(c)].width = min(max_w + 3, 38)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    try:
        r1 = await db.execute(text("SELECT * FROM v_copamundial_puntajes"))
        rows1 = [dict(row._mapping) for row in r1]
    except Exception as ex:
        raise HTTPException(500, f"Error consultando vista resumen: {ex}")
    try:
        # Try to enrich detail view with numero_fifa (requires partido_id in view)
        r2 = await db.execute(text("""
            SELECT v.*, COALESCE(p.numero_fifa, 0) AS numero_fifa
            FROM v_copamundial_puntajes_det v
            LEFT JOIN partido p ON p.id = v.partido_id
        """))
        rows2_raw = [dict(row._mapping) for row in r2]
        # Prepend P# to the 'partido' column text if present
        for row in rows2_raw:
            nf = row.get("numero_fifa") or 0
            if nf and "partido" in row:
                row["partido"] = f"P{nf}  {row['partido']}"
        rows2 = rows2_raw
    except Exception:
        try:
            r2b = await db.execute(text("SELECT * FROM v_copamundial_puntajes_det"))
            rows2 = [dict(row._mapping) for row in r2b]
        except Exception as ex2:
            raise HTTPException(500, f"Error consultando vista detalle: {ex2}")

    wb = Workbook()
    ws1 = wb.active
    _write_sheet(ws1, rows1, "Resumen")
    ws2 = wb.create_sheet()
    _write_sheet(ws2, rows2, "Detalle")

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    ts = _dt.now().strftime("%Y%m%d_%H%M")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="becbuc_puntajes_copa_{ts}.xlsx"'},
    )
