# -*- coding: utf-8 -*-
"""
reimportar_predicciones_bonus.py [--apply]
Re-importa las PREDICCIONES DE BONUS de cada apostador desde la hoja '50- TBL MASTER'
hacia la BD (tabla apuesta), para TODAS las fases jugadas (grupos..semis, P001-P102),
dejandolas IDENTICAS al Excel.

Columnas actualizadas (Excel col 1-based -> campo apuesta):
  J-amarillas(25) -> pred_amarillas
  K-rojas(26)     -> pred_rojas
  L-var(27)       -> pred_var
  M-penales(28)   -> pred_penales_partido
  N-1er gol(29)   -> pred_minuto_gol
  O-tanda EQ1(30) -> pred_penales_local_tanda
  P-tanda EQ2(31) -> pred_penales_visitante_tanda
  Q-clasifica(32) -> pred_equipo_clasifica (nombre equipo -> equipo_id)

NO toca el marcador (pred_local/pred_visitante). Solo UPDATE de apuestas existentes.

SIN --apply: DRY RUN (muestra que cambiaria, no escribe).
CON --apply: escribe en la BD. Luego correr run_recalc_hasta_semis.bat.

Uso:
  backend\.venv\Scripts\python.exe reimportar_predicciones_bonus.py
  backend\.venv\Scripts\python.exe reimportar_predicciones_bonus.py --apply
"""
import sys, os
from collections import Counter

DO_APPLY = '--apply' in [a.lower() for a in sys.argv[1:]]
print(f"{'[APPLY]' if DO_APPLY else '[DRY RUN]'}")

BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = None
for f in os.listdir(BASE):
    if 'SEMIFINAL' in f.upper() and f.endswith('.xlsx'):
        EXCEL_FILE = os.path.join(BASE, f); break
if not EXCEL_FILE:
    sys.exit(f"ERROR: no se encontro Excel *SEMIFINAL*.xlsx en {BASE}")
print(f"Excel: {os.path.basename(EXCEL_FILE)}")

try:
    import openpyxl
except ImportError:
    os.system(f'"{sys.executable}" -m pip install openpyxl --quiet'); import openpyxl
try:
    import psycopg2, psycopg2.extras
except ImportError:
    os.system(f'"{sys.executable}" -m pip install psycopg2-binary --quiet'); import psycopg2, psycopg2.extras

CONN_BEC = "host=localhost port=5432 dbname=becbuc user=app_user password=superpassword"
CONN_APP = "host=localhost port=5432 dbname=app_db  user=app_user password=superpassword"
TID = 2
SHEET = '50- TBL MASTER'
FASES_OK = {'10- GRUPOS','20- DIECISEISAVOS','30- OCTAVOS','40- CUARTOS','50- SEMIFINAL'}

conn = psycopg2.connect(CONN_BEC); conn.autocommit = False
capp = psycopg2.connect(CONN_APP)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
cua = capp.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# alias -> apostador_id
cua.execute("SELECT id, username FROM users WHERE is_active=TRUE")
uname = {u['username'].lower(): u['id'] for u in cua.fetchall()}
def clean_alias(s):
    if not s: return ''
    return str(s).replace('\xa0','').strip().upper().lstrip('@').replace('  ',' ')
def find_uid(al):
    a = clean_alias(al)
    for k, v in uname.items():
        if k.upper().lstrip('@') == a: return v
    for k, v in uname.items():
        if a and (a in k.upper() or k.upper() in a): return v
    return None

# equipos para "quien clasifica"
cur.execute("SELECT id, nombre, nombre_es FROM equipo")
eid = {}
for eq in cur.fetchall():
    if eq['nombre']:    eid[eq['nombre'].upper().strip()] = eq['id']
    if eq['nombre_es']: eid[eq['nombre_es'].upper().strip()] = eq['id']
ALIASE = {'FRANCIA':'France','ESPANA':'Spain','INGLATERRA':'England','ARGENTINA':'Argentina',
    'MARRUECOS':'Morocco','BELGICA':'Belgium','NORUEGA':'Norway','SUIZA':'Switzerland',
    'COLOMBIA':'Colombia','MEXICO':'Mexico','BRASIL':'Brazil','PORTUGAL':'Portugal',
    'PARAGUAY':'Paraguay','CANADA':'Canada','EGIPTO':'Egypt','ESTADOS UNIDOS':'USA',
    'EE UU':'USA','EEUU':'USA','ALEMANIA':'Germany'}
def find_eid(nom):
    if not nom: return None
    k = str(nom).upper().strip()
    if k in ('','-'): return None
    if k in eid: return eid[k]
    a = ALIASE.get(k)
    if a and a.upper() in eid: return eid[a.upper()]
    for kk, vv in eid.items():
        if k in kk or kk in k: return vv
    return None
def to_int(v):
    try:
        s = str(v).strip()
        if s in ('','-','None'): return None
        return int(float(s))
    except: return None

# partidos numero_fifa -> id (1..102)
cur.execute("""
    SELECT p.numero_fifa, p.id FROM partido p JOIN fase f ON f.id=p.fase_id
    WHERE f.torneo_id=%s AND p.numero_fifa BETWEEN 1 AND 102
""", (TID,))
pid_by_nf = {r['numero_fifa']: r['id'] for r in cur.fetchall()}

# apuestas actuales: bonus preds por (apostador_id, partido_id)
cur.execute("""
    SELECT apostador_id, partido_id,
           pred_amarillas, pred_rojas, pred_var, pred_penales_partido, pred_minuto_gol,
           pred_penales_local_tanda, pred_penales_visitante_tanda, pred_equipo_clasifica
    FROM apuesta
    WHERE partido_id = ANY(%s)
""", (list(pid_by_nf.values()),))
cur_bets = {(r['apostador_id'], r['partido_id']): dict(r) for r in cur.fetchall()}

# columnas: (campo_bd, col_excel_1based, tipo)
CAMPOS = [
    ('pred_amarillas', 25, 'int'),
    ('pred_rojas', 26, 'int'),
    ('pred_var', 27, 'int'),
    ('pred_penales_partido', 28, 'int'),
    ('pred_minuto_gol', 29, 'int'),
    ('pred_penales_local_tanda', 30, 'int'),
    ('pred_penales_visitante_tanda', 31, 'int'),
    ('pred_equipo_clasifica', 32, 'equipo'),
]

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb[SHEET]

updates = []          # (apostador_id, partido_id, dict_valores)
por_campo = Counter()
diffs_sample = []
sin_alias = Counter(); sin_apuesta = 0; comparados = 0

for r in range(2, ws.max_row + 1):
    ftxt = str(ws.cell(r, 7).value)
    if ftxt not in FASES_OK:
        continue
    pid = ws.cell(r, 2).value
    if not (isinstance(pid, str) and pid.startswith('P')):
        continue
    nf = int(pid[1:])
    partido_id = pid_by_nf.get(nf)
    if not partido_id:
        continue
    uid = find_uid(ws.cell(r, 10).value)
    if not uid:
        sin_alias[clean_alias(ws.cell(r, 10).value)] += 1; continue
    cur_bet = cur_bets.get((uid, partido_id))
    if not cur_bet:
        sin_apuesta += 1; continue
    comparados += 1

    nuevos = {}
    for campo, col, tipo in CAMPOS:
        raw = ws.cell(r, col).value
        val = find_eid(raw) if tipo == 'equipo' else to_int(raw)
        nuevos[campo] = val
    # detectar diffs vs BD
    difs = [(c, cur_bet[c], nuevos[c]) for c, _, _ in CAMPOS if cur_bet[c] != nuevos[c]]
    if difs:
        for c, old, new in difs:
            por_campo[c] += 1
        if len(diffs_sample) < 40:
            diffs_sample.append((pid, clean_alias(ws.cell(r,10).value), difs))
        updates.append((uid, partido_id, nuevos))

print("\n" + "="*74)
print("RE-IMPORT PREDICCIONES DE BONUS (TBL MASTER -> apuesta)")
print("="*74)
print(f"Apuestas comparadas: {comparados}")
if sin_alias: print(f"Aliases sin match: {dict(sin_alias)}")
if sin_apuesta: print(f"Filas Excel sin apuesta en BD (omitidas): {sin_apuesta}")
print(f"Apuestas que cambiarian: {len(updates)}")
if por_campo:
    print("Cambios por campo:")
    for c, k in por_campo.most_common():
        print(f"   {c}: {k}")
if diffs_sample:
    print("\nMuestra (primeras 40):")
    for pid, al, difs in diffs_sample:
        ds = ", ".join(f"{c.replace('pred_','')}: {o}->{n}" for c, o, n in difs)
        print(f"   {pid} {al:<16} {ds}")

if not DO_APPLY:
    print("\n[DRY RUN] No se escribio nada. Para aplicar:")
    print("   reimportar_predicciones_bonus.py --apply")
    conn.close(); capp.close(); sys.exit(0)

print(f"\nAPLICANDO {len(updates)} updates...")
n = 0
for uid, partido_id, nuevos in updates:
    cur.execute("""
        UPDATE apuesta SET
            pred_amarillas=%(pred_amarillas)s, pred_rojas=%(pred_rojas)s, pred_var=%(pred_var)s,
            pred_penales_partido=%(pred_penales_partido)s, pred_minuto_gol=%(pred_minuto_gol)s,
            pred_penales_local_tanda=%(pred_penales_local_tanda)s,
            pred_penales_visitante_tanda=%(pred_penales_visitante_tanda)s,
            pred_equipo_clasifica=%(pred_equipo_clasifica)s
        WHERE apostador_id=%(uid)s AND partido_id=%(pid)s
    """, {**nuevos, 'uid': uid, 'pid': partido_id})
    n += cur.rowcount
conn.commit()
print(f"OK. Filas actualizadas: {n}")
print("\nAHORA recalcular: run_recalc_hasta_semis.bat")
conn.close(); capp.close()
