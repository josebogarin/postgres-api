# -*- coding: utf-8 -*-
"""
comparar_items_semis.py
Compara ITEM POR ITEM el RESULTADO OFICIAL de los partidos de SEMIFINAL (P101-P102)
entre el Excel (hoja '40- RESULTADOS OFICIALES') y la BD (tabla partido).

Items comparados por partido:
  goles_local, goles_visitante, J-amarillas, K-rojas, L-VAR, M-penales(juego),
  N-minuto_1er_gol, tanda EQ1(local), tanda EQ2(visitante), equipo que pasa.

Nota: valor 99 en tanda (Excel) = 'sin tanda' -> se compara como NULL.
Solo lectura: NO modifica la BD.

Uso:
  backend\\.venv\\Scripts\\python.exe comparar_items_semis.py
"""
import sys, os

BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = None
for f in os.listdir(BASE):
    if 'SEMIFINAL' in f.upper() and f.endswith('.xlsx'):
        EXCEL_FILE = os.path.join(BASE, f); break
if not EXCEL_FILE:
    sys.exit(f"ERROR: no se encontro Excel *SEMIFINAL*.xlsx en {BASE}")
print(f"Excel: {os.path.basename(EXCEL_FILE)}")

try:
    import openpyxl
except ImportError:
    os.system(f'"{sys.executable}" -m pip install openpyxl --quiet'); import openpyxl
try:
    import psycopg2, psycopg2.extras
except ImportError:
    os.system(f'"{sys.executable}" -m pip install psycopg2-binary --quiet'); import psycopg2, psycopg2.extras

CONN_BEC = "host=localhost port=5432 dbname=becbuc user=app_user password=superpassword"
TORNEO_ID = 2
SHEET = '40- RESULTADOS OFICIALES'
NF_MIN, NF_MAX = 101, 102

conn = psycopg2.connect(CONN_BEC)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# -- equipos para 'quien pasa' ----------------------------------------------
cur.execute("SELECT id, nombre, nombre_es FROM equipo")
equipo_id_by_nombre = {}
equipo_nombre_by_id = {}
for eq in cur.fetchall():
    equipo_nombre_by_id[eq['id']] = eq['nombre'] or eq['nombre_es']
    if eq['nombre']:    equipo_id_by_nombre[eq['nombre'].upper().strip()] = eq['id']
    if eq['nombre_es']: equipo_id_by_nombre[eq['nombre_es'].upper().strip()] = eq['id']
EQUIPO_ALIAS = {
    'FRANCIA':'France','ESPAÑA':'Spain','ESPANA':'Spain',
    'INGLATERRA':'England','ARGENTINA':'Argentina',
}
def find_equipo_id(nombre_excel):
    if not nombre_excel: return None
    key = str(nombre_excel).upper().strip()
    if key in equipo_id_by_nombre: return equipo_id_by_nombre[key]
    alt = EQUIPO_ALIAS.get(key)
    if alt and alt.upper().strip() in equipo_id_by_nombre:
        return equipo_id_by_nombre[alt.upper().strip()]
    for k, v in equipo_id_by_nombre.items():
        if key in k or k in key: return v
    return None

# -- partidos BD ------------------------------------------------------------
cur.execute("""
    SELECT p.numero_fifa, el.nombre AS local, ev.nombre AS visit, p.estado,
           p.goles_local, p.goles_visitante,
           p.amarillas, p.rojas, p.decisiones_var, p.penales_partido,
           p.minuto_primer_gol, p.penales_local, p.penales_visitante,
           p.equipo_clasificado_id
    FROM partido p JOIN fase f ON f.id=p.fase_id
    LEFT JOIN equipo el ON el.id=p.equipo_local_id
    LEFT JOIN equipo ev ON ev.id=p.equipo_visitante_id
    WHERE f.torneo_id=%s AND p.numero_fifa BETWEEN %s AND %s
    ORDER BY p.numero_fifa
""", (TORNEO_ID, NF_MIN, NF_MAX))
bd = {f"P{r['numero_fifa']:03d}": dict(r) for r in cur.fetchall()}

def to_int(v):
    try:
        s = str(v).strip()
        if s in ('', '-', 'None'): return None
        return int(float(s))
    except: return None
def tanda(v):           # 99 = sin tanda -> NULL
    x = to_int(v)
    return None if x == 99 else x

# columnas 1-based del Excel (hoja RESULTADOS OFICIALES)
COL = dict(pid=1, eq1=10, gl=12, gv=14, eq2=15, J=30, K=31, L=32, M=33, N=34,
           tl=35, tv=36, pasa=37)

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb[SHEET]
excel = {}
for r in range(2, ws.max_row + 1):
    pid = ws.cell(r, COL['pid']).value
    if pid in ('P101', 'P102'):
        excel[pid] = {
            'goles_local': to_int(ws.cell(r, COL['gl']).value),
            'goles_visitante': to_int(ws.cell(r, COL['gv']).value),
            'amarillas': to_int(ws.cell(r, COL['J']).value),
            'rojas': to_int(ws.cell(r, COL['K']).value),
            'decisiones_var': to_int(ws.cell(r, COL['L']).value),
            'penales_partido': to_int(ws.cell(r, COL['M']).value),
            'minuto_primer_gol': to_int(ws.cell(r, COL['N']).value),
            'penales_local': tanda(ws.cell(r, COL['tl']).value),
            'penales_visitante': tanda(ws.cell(r, COL['tv']).value),
            'equipo_clasificado_id': find_equipo_id(ws.cell(r, COL['pasa']).value),
            '_pasa_txt': ws.cell(r, COL['pasa']).value,
        }

LABELS = [
    ('goles_local', 'Goles local (H)'),
    ('goles_visitante', 'Goles visitante (H)'),
    ('amarillas', 'J- Amarillas'),
    ('rojas', 'K- Rojas'),
    ('decisiones_var', 'L- VAR'),
    ('penales_partido', 'M- Penales juego'),
    ('minuto_primer_gol', 'N- Minuto 1er gol'),
    ('penales_local', 'O- Tanda local'),
    ('penales_visitante', 'O- Tanda visitante'),
    ('equipo_clasificado_id', 'P- Equipo que pasa'),
]

print("\n" + "="*70)
print("COMPARACION ITEMS OFICIALES SEMIFINAL: Excel vs BD (partido)")
print("="*70)

total_diffs = 0
for pid in ('P101', 'P102'):
    ex = excel.get(pid); db = bd.get(pid)
    if not ex:
        print(f"\n{pid}: no esta en la hoja RESULTADOS OFICIALES del Excel."); continue
    if not db:
        print(f"\n{pid}: no esta en la BD."); continue
    print(f"\n{pid}: {db['local']} {db['goles_local']}-{db['goles_visitante']} {db['visit']} [{db['estado']}]")
    print(f"  {'ITEM':<22}{'EXCEL':<16}{'BD':<16}{'ESTADO'}")
    print("  " + "-"*58)
    for campo, label in LABELS:
        ve = ex[campo]; vb = db[campo]
        if campo == 'equipo_clasificado_id':
            se = ex['_pasa_txt']
            sb = equipo_nombre_by_id.get(vb, vb)
            ok = (ve == vb)
            print(f"  {label:<22}{str(se):<16}{str(sb):<16}{'OK' if ok else '<-- DIFERENTE'}")
        else:
            ok = (ve == vb)
            print(f"  {label:<22}{str(ve):<16}{str(vb):<16}{'OK' if ok else '<-- DIFERENTE'}")
        if not ok:
            total_diffs += 1

print("\n" + "="*70)
if total_diffs == 0:
    print("RESULTADO: TODOS LOS ITEMS COINCIDEN (Excel == BD). 0 diferencias.")
else:
    print(f"RESULTADO: {total_diffs} DIFERENCIA(S). Revisar filas marcadas '<-- DIFERENTE'.")
print("="*70)

conn.close()
