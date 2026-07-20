"""
importar_r32_excel.py  — Sesion 52
Importa pronosticos R32 (P073-P088) desde el Excel consolidado
y verifica puntajes de fase de grupos por apostador.

Ejecutar:
  run_importar_r32.bat               <-- dry run (solo verifica)
  run_importar_r32.bat --import      <-- importa pronosticos R32
"""

import sys, os, glob

BASE = os.path.dirname(os.path.abspath(__file__))

EXCEL_FILE = None
for f in os.listdir(BASE):
    fu = f.upper()
    if ('16AVOS' in fu or 'CONSOLIDADOS' in fu) and f.endswith('.xlsx'):
        EXCEL_FILE = os.path.join(BASE, f)
        break

if not EXCEL_FILE:
    print("ERROR: No se encontro el Excel.")
    print(f"Copia el archivo '...CONSOLIDADOS 16AVOS.xlsx' a: {BASE}")
    sys.exit(1)

print(f"Excel: {os.path.basename(EXCEL_FILE)}")

try:
    import openpyxl
except ImportError:
    os.system(f'"{sys.executable}" -m pip install openpyxl --quiet')
    import openpyxl

try:
    import psycopg2, psycopg2.extras
except ImportError:
    os.system(f'"{sys.executable}" -m pip install psycopg2-binary --quiet')
    import psycopg2, psycopg2.extras

CONN_BEC = "host=localhost port=5432 dbname=becbuc user=app_user password=superpassword"
CONN_APP = "host=localhost port=5432 dbname=app_db  user=app_user password=superpassword"
TORNEO_ID = 2
DO_IMPORT = '--import' in sys.argv

print("Conectando a BD...")
try:
    conn_bec = psycopg2.connect(CONN_BEC)
    conn_app = psycopg2.connect(CONN_APP)
except Exception as e:
    print(f"ERROR de conexion: {e}")
    print("Asegurate de que Docker este corriendo: docker start core-postgres")
    sys.exit(1)

cur_bec = conn_bec.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
cur_app = conn_app.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

cur_app.execute("SELECT id, username FROM users WHERE is_active=TRUE ORDER BY id")
users = cur_app.fetchall()
user_by_ulow = {u['username'].lower(): u['id'] for u in users}

cur_bec.execute("""
    SELECT DISTINCT a.apostador_id
    FROM apuesta a
    JOIN partido p ON p.id = a.partido_id
    JOIN fase f ON f.id = p.fase_id
    WHERE f.torneo_id = %s
    ORDER BY a.apostador_id
""", (TORNEO_ID,))
bec_ids = {r['apostador_id'] for r in cur_bec.fetchall()}

# Obtener usernames desde app_db para los IDs que tienen apuestas
cur_app.execute("SELECT id, username, nombre as nombre_completo FROM users WHERE is_active=TRUE ORDER BY id")
all_users = cur_app.fetchall()
bd_apostadores = [u for u in all_users if u['id'] in bec_ids]
print(f"Apostadores en BD: {len(bd_apostadores)}")

# username.lower() -> apostador_id
apostador_to_id = {u['username'].lower(): u['id'] for u in bd_apostadores}
# nombre_completo.upper() -> apostador_id
nombre_to_id = {}
for u in bd_apostadores:
    nc = (u.get('nombre_completo') or '').upper().strip()
    if nc:
        nombre_to_id[nc] = u['id']

def clean_alias(s):
    if not s: return ''
    return str(s).strip().upper().lstrip('@').replace('\xa0','').replace('  ',' ')

def find_apostador_id(nombre_excel, alias_excel):
    alias_c = clean_alias(alias_excel)
    nombre_c = nombre_excel.upper().strip() if nombre_excel else ''
    for k, v in apostador_to_id.items():
        if k.upper().lstrip('@') == alias_c:
            return v
    if nombre_c and nombre_c in nombre_to_id:
        return nombre_to_id[nombre_c]
    for candidate in [alias_c.lower(), nombre_c.lower()]:
        if candidate in user_by_ulow:
            uid = user_by_ulow[candidate]
            if uid in bec_ids:
                return uid
    for k, v in apostador_to_id.items():
        if alias_c and (alias_c in k.upper() or k.upper() in alias_c):
            return v
    return None

print("Leyendo Excel...")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb['pronosticos']

cur_bec.execute("""
    SELECT p.id, p.numero_fifa,
           el.nombre AS local_nombre, ev.nombre AS visitante_nombre
    FROM partido p
    JOIN fase f ON f.id = p.fase_id
    LEFT JOIN equipo el ON el.id = p.equipo_local_id
    LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
    WHERE f.torneo_id = %s AND p.numero_fifa BETWEEN 73 AND 88
    ORDER BY p.numero_fifa
""", (TORNEO_ID,))
partidos_r32 = {f"P{r['numero_fifa']:03d}": dict(r) for r in cur_bec.fetchall()}
print(f"Partidos R32 en BD ({len(partidos_r32)}):")
for k, v in sorted(partidos_r32.items()):
    print(f"  {k} id={v['id']}: {v['local_nombre']} vs {v['visitante_nombre']}")

r32_preds = []
unmatched = {}

def to_int(v):
    try: return int(v) if v is not None else None
    except: return None

for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    if row[3] != '20- DIECISEISAVOS': continue
    nombre = str(row[5]).strip() if row[5] else ''
    alias  = str(row[6]).strip() if row[6] else ''
    pid_str = str(row[1]).strip() if row[1] else ''
    uid = find_apostador_id(nombre, alias)
    if not uid:
        key = f"{clean_alias(alias)}/{nombre}"
        unmatched[key] = unmatched.get(key, 0) + 1
        continue
    partido_db = partidos_r32.get(pid_str)
    if not partido_db:
        continue
    r32_preds.append({
        'apostador_id':                 uid,
        'partido_id':                   partido_db['id'],
        'pred_local':                   to_int(row[9]),
        'pred_visitante':               to_int(row[11]),
        'pred_amarillas':               to_int(row[21]),
        'pred_rojas':                   to_int(row[22]),
        'pred_var':                     to_int(row[23]),
        'pred_penales_partido':         to_int(row[24]),
        'pred_minuto_gol':              to_int(row[25]),
        'pred_penales_local_tanda':     to_int(row[26]),
        'pred_penales_visitante_tanda': to_int(row[27]),
    })

print(f"\nPronosticos R32 resueltos: {len(r32_preds)}")
if unmatched:
    print(f"Aliases sin match ({len(unmatched)}):")
    for k, cnt in sorted(unmatched.items()):
        print(f"  {k} ({cnt} filas)")
else:
    print("Aliases sin match: ninguno")

if not DO_IMPORT:
    print("\n[DRY Run] Muestra de pronosticos que se importarian:")
    for p in r32_preds[:8]:
        aid = p['apostador_id']
        alias_bd = next((r['username'] for r in bd_apostadores if r['id']==aid), str(aid))
        pid_name = next((k for k,v in partidos_r32.items() if v['id']==p['partido_id']), str(p['partido_id']))
        print(f"  {alias_bd:<20} {pid_name} pred={p['pred_local']}-{p['pred_visitante']} amar={p['pred_amarillas']} O1={p['pred_penales_local_tanda']} O2={p['pred_penales_visitante_tanda']}")
    print(f"\nTotal: {len(r32_preds)} apuestas en {len(partidos_r32)} partidos R32")
    print("Para importar: run_importar_r32.bat --import")
else:
    print(f"\nImportando {len(r32_preds)} pronosticos R32...")
    upserted = 0
    errors = 0
    for pred in r32_preds:
        try:
            cur_bec.execute("""
                INSERT INTO apuesta (
                    apostador_id, partido_id,
                    pred_local, pred_visitante,
                    pred_amarillas, pred_rojas, pred_var,
                    pred_penales_partido, pred_minuto_gol,
                    pred_penales_local_tanda, pred_penales_visitante_tanda
                ) VALUES (
                    %(apostador_id)s, %(partido_id)s,
                    %(pred_local)s, %(pred_visitante)s,
                    %(pred_amarillas)s, %(pred_rojas)s, %(pred_var)s,
                    %(pred_penales_partido)s, %(pred_minuto_gol)s,
                    %(pred_penales_local_tanda)s, %(pred_penales_visitante_tanda)s
                )
                ON CONFLICT (apostador_id, partido_id) DO UPDATE SET
                    pred_local                   = EXCLUDED.pred_local,
                    pred_visitante               = EXCLUDED.pred_visitante,
                    pred_amarillas               = EXCLUDED.pred_amarillas,
                    pred_rojas                   = EXCLUDED.pred_rojas,
                    pred_var                     = EXCLUDED.pred_var,
                    pred_penales_partido         = EXCLUDED.pred_penales_partido,
                    pred_minuto_gol              = EXCLUDED.pred_minuto_gol,
                    pred_penales_local_tanda     = EXCLUDED.pred_penales_local_tanda,
                    pred_penales_visitante_tanda = EXCLUDED.pred_penales_visitante_tanda
            """, pred)
            upserted += 1
        except Exception as e:
            print(f"  ERROR aid={pred['apostador_id']} pid={pred['partido_id']}: {e}")
            conn_bec.rollback()
            errors += 1
    conn_bec.commit()
    print(f"\nImportados: {upserted} | Errores: {errors}")

print("\n" + "="*62)
print("VERIFICACION PUNTAJES GRUPOS — Excel vs BD")
print("="*62)

excel_totals = {}
excel_count  = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    if row[3] != '10- GRUPOS': continue
    alias = clean_alias(row[6])
    pts   = int(row[36] or 0)
    excel_totals[alias] = excel_totals.get(alias, 0) + pts
    excel_count[alias]  = excel_count.get(alias, 0) + 1

cur_bec.execute("""
    SELECT pd.apostador_id,
           SUM(
               COALESCE(pd.pts_resultado,0)  + COALESCE(pd.pts_marcador,0)  +
               COALESCE(pd.pts_amarillas,0)  + COALESCE(pd.pts_rojas,0)     +
               COALESCE(pd.pts_var,0)        + COALESCE(pd.pts_penales_partido,0) +
               COALESCE(pd.pts_minuto,0)
           ) AS total_bd
    FROM puntaje_detalle pd
    JOIN partido p ON p.id = pd.partido_id
    JOIN fase f ON f.id = p.fase_id
    WHERE f.torneo_id = %s AND f.tipo ILIKE '%%grupo%%'
    GROUP BY pd.apostador_id
""", (TORNEO_ID,))
bd_totals = {r['apostador_id']: int(r['total_bd'] or 0) for r in cur_bec.fetchall()}

print(f"\n{'ALIAS':<27} {'EXCEL':>7} {'FILAS':>6} {'BD':>7} {'DIFF':>7}")
print("-" * 62)
diffs = []
matched = 0

for alias, excel_pts in sorted(excel_totals.items(), key=lambda x: -x[1]):
    uid   = find_apostador_id('', alias)
    bd_pts = bd_totals.get(uid, 0) if uid else 0
    diff   = excel_pts - bd_pts
    n      = excel_count.get(alias, 0)
    ok     = "" if uid else " (sin match)"
    marker = " DIFF!" if abs(diff) > 5 else ""
    if uid: matched += 1
    print(f"  {alias:<25} {excel_pts:>7} {n:>6} {bd_pts:>7} {diff:>+7}{marker}{ok}")
    if abs(diff) > 5 and uid:
        diffs.append((alias, excel_pts, bd_pts, diff, uid))

print(f"\nResumen: {matched}/{len(excel_totals)} aliases matcheados")
if diffs:
    print(f"\n{len(diffs)} con diferencia > 5 pts:")
    for alias, ex, bd, d, uid in sorted(diffs, key=lambda x: abs(x[3]), reverse=True):
        bdname = next((r['username'] for r in bd_apostadores if r['id']==uid), '?')
        print(f"  {alias} (BD:{bdname}): Excel={ex} BD={bd} diff={d:+}")
else:
    print("Todos los puntajes coinciden (diff <= 5 pts)")

cur_bec.close(); cur_app.close()
conn_bec.close(); conn_app.close()
print("\nListo.")
