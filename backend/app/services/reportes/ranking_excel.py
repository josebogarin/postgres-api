# -*- coding: utf-8 -*-
"""
ranking_excel.py — Excel de ranking con desglose por fase (Fase 3c).
Movido desde apostador_bets.py. Hoja 'Puntaje general' + una hoja por fase.
"""
from fastapi import HTTPException
from fastapi.responses import StreamingResponse

from sqlalchemy import text

from app.db.session import engine as _app_engine


async def build_ranking_export(torneo_id: int, current, db):
    import io
    from datetime import datetime as _dt
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Shared helpers ──────────────────────────────────────────────────────
    def _hf(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    HDR_GRN = _hf("1A6B45"); HDR_BLU = _hf("2D4A6B"); HDR_SUB = _hf("3A5A7C")
    FILL_GRN = _hf("D6F0E0"); FILL_GRY = _hf("F5F5F5"); FILL_WHT = _hf("FFFFFF")
    FNT_W9B = Font(color="FFFFFF", bold=True, size=9)
    FNT_9B  = Font(bold=True, size=9)
    FNT_9   = Font(size=9)
    AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL = Alignment(horizontal="left", vertical="center")
    _s = Side(style="thin", color="CCCCCC")
    BDR = Border(left=_s, right=_s, top=_s, bottom=_s)

    def _cell(ws, r, c, v=None, fill=None, font=None, align=None):
        cell = ws.cell(r, c)
        if v is not None:
            cell.value = v
        cell.fill   = fill or FILL_WHT
        cell.font   = font or FNT_9
        cell.alignment = align or AC
        cell.border = BDR
        return cell

    def _vv(x):
        return "" if x is None else x

    def _marc(l, v):
        return "" if l is None and v is None else f"{_vv(l)}-{_vv(v)}"

    # Icons for each scoring item
    ICON = {
        "H": "Resultado", "I": "Exacto",   "J": "Amarillas", "K": "Rojas",
        "L": "VAR",       "M": "Pen. Pdo", "N": "1er Gol",   "O": "Pen. Tanda",
        "G": "Globales",  "T": "Total",
    }

    # ── 1. Ranking summary ──────────────────────────────────────────────────
    try:
        r_rk = await db.execute(text("""
            SELECT pd.apostador_id,
                   COALESCE(SUM(pd.pts_resultado),       0)::int AS cat_h,
                   COALESCE(SUM(pd.pts_marcador),        0)::int AS cat_i,
                   COALESCE(SUM(pd.pts_amarillas),       0)::int AS cat_j,
                   COALESCE(SUM(pd.pts_rojas),           0)::int AS cat_k,
                   COALESCE(SUM(pd.pts_var),             0)::int AS cat_l,
                   COALESCE(SUM(pd.pts_penales_partido), 0)::int AS cat_m,
                   COALESCE(SUM(pd.pts_minuto),          0)::int AS cat_n,
                   COALESCE(SUM(pd.pts_penales_tanda),   0)::int AS cat_o,
                   COALESCE(SUM(pd.pts_total),           0)::int AS pts_partidos
            FROM puntaje_detalle pd WHERE pd.torneo_id = :tid
            GROUP BY pd.apostador_id
        """), {"tid": torneo_id})
        rk_rows = [dict(r) for r in r_rk.mappings()]
    except Exception:
        await db.rollback()
        rk_rows = []

    try:
        r_glob = await db.execute(text(
            "SELECT apostador_id, puntos_total AS pts_globales FROM puntaje_global WHERE torneo_id = :tid"
        ), {"tid": torneo_id})
        glob_map = {r["apostador_id"]: r["pts_globales"] for r in r_glob.mappings()}
    except Exception:
        await db.rollback()
        glob_map = {}

    try:
        r_clas = await db.execute(text("""
            SELECT apostador_id, pts_obtenidos AS pts_grupos_p
            FROM apostador_clasificados WHERE torneo_id = :tid AND fase_tipo = 'grupo'
        """), {"tid": torneo_id})
        clas_map = {r["apostador_id"]: int(r["pts_grupos_p"] or 0) for r in r_clas.mappings()}
    except Exception:
        await db.rollback()
        clas_map = {}

    for row in rk_rows:
        row["pts_globales"] = glob_map.get(row["apostador_id"], 0) or 0
        row["pts_grupos_p"] = clas_map.get(row["apostador_id"], 0) or 0
        row["pts_total"]    = row["pts_partidos"] + row["pts_globales"] + row["pts_grupos_p"]

    # ── 2. Detail rows ──────────────────────────────────────────────────────
    try:
        r_det = await db.execute(text("""
            SELECT pd.apostador_id, pd.partido_id,
                   f.nombre AS fase, COALESCE(f.orden, 0) AS fase_orden,
                   COALESCE(el.nombre_es, el.nombre) AS equipo_local,
                   COALESCE(ev.nombre_es, ev.nombre) AS equipo_visitante,
                   p.goles_local, p.goles_visitante,
                   COALESCE(p.amarillas,      0) AS real_amarillas,
                   COALESCE(p.rojas,          0) AS real_rojas,
                   COALESCE(p.decisiones_var, 0) AS real_var,
                   p.penales_partido          AS real_pen_partido,
                   p.minuto_primer_gol        AS real_minuto,
                   p.penales_local            AS real_pen_local,
                   p.penales_visitante        AS real_pen_visitante,
                   COALESCE(p.numero_fifa, 0) AS numero_fifa,
                   f.tipo AS fase_tipo,
                   a.pred_local, a.pred_visitante, a.pred_amarillas,
                   a.pred_rojas, a.pred_var, a.pred_penales_partido,
                   a.pred_minuto_gol, a.pred_penales_local_tanda, a.pred_penales_visitante_tanda,
                   a.pred_equipo_clasifica,
                   COALESCE(epred.nombre_es, epred.nombre, '') AS pred_equipo_clasifica_nm,
                   COALESCE(ereal.nombre_es, ereal.nombre, '') AS real_equipo_clasifica_nm,
                   COALESCE(pd.pts_resultado,       0)::int AS pts_h,
                   COALESCE(pd.pts_marcador,        0)::int AS pts_i,
                   COALESCE(pd.pts_amarillas,       0)::int AS pts_j,
                   COALESCE(pd.pts_rojas,           0)::int AS pts_k,
                   COALESCE(pd.pts_var,             0)::int AS pts_l,
                   COALESCE(pd.pts_penales_partido, 0)::int AS pts_m,
                   COALESCE(pd.pts_minuto,          0)::int AS pts_n,
                   COALESCE(pd.pts_penales_tanda,   0)::int AS pts_o,
                   COALESCE(pd.pts_equipo,          0)::int AS pts_p,
                   COALESCE(pd.pts_total,           0)::int AS pts_total
            FROM puntaje_detalle pd
            JOIN partido p  ON p.id  = pd.partido_id
            JOIN fase    f  ON f.id  = p.fase_id
            LEFT JOIN equipo el    ON el.id    = p.equipo_local_id
            LEFT JOIN equipo ev    ON ev.id    = p.equipo_visitante_id
            LEFT JOIN apuesta a ON a.apostador_id = pd.apostador_id
                               AND a.partido_id   = pd.partido_id
            LEFT JOIN equipo epred ON epred.id = a.pred_equipo_clasifica
            LEFT JOIN equipo ereal ON ereal.id = p.equipo_clasificado_id
            WHERE pd.torneo_id = :tid
            ORDER BY fase_orden, pd.partido_id, pd.apostador_id
        """), {"tid": torneo_id})
        det_rows = [dict(r) for r in r_det.mappings()]
    except Exception:
        await db.rollback()
        det_rows = []

    # Resolve usernames
    all_ids = list({r["apostador_id"] for r in rk_rows} | {r["apostador_id"] for r in det_rows})
    nombre_map: dict = {}
    if all_ids:
        async with _app_engine.connect() as conn:
            nr = await conn.execute(text("SELECT id, username FROM users WHERE id = ANY(:ids)"), {"ids": all_ids})
            nombre_map = {r["id"]: r["username"] for r in nr.mappings()}

    for row in rk_rows:
        row["nombre"] = nombre_map.get(row["apostador_id"], f"?{row['apostador_id']}")
    for row in det_rows:
        row["nombre"] = nombre_map.get(row["apostador_id"], f"?{row['apostador_id']}")

    rk_rows.sort(key=lambda r: -r["pts_total"])

    # ── 3. Build workbook ───────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    # ══════════════════════════════════════════════════════════════════════
    # FICHA 1 — Por Ítem
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet("Por Ítem")
    hdr1 = ["#", "Apostador",
            ICON["H"], ICON["I"], ICON["J"], ICON["K"],
            ICON["L"], ICON["M"], ICON["N"], ICON["O"],
            ICON["G"], ICON["T"]]
    wdths1 = [4, 18, 6, 6, 6, 6, 6, 6, 6, 6, 6, 8]
    for ci, (h, w) in enumerate(zip(hdr1, wdths1), 1):
        _cell(ws1, 1, ci, h, HDR_GRN, FNT_W9B, AC)
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[1].height = 22
    ws1.freeze_panes = "A2"
    for ri, row in enumerate(rk_rows, 2):
        fill = FILL_GRN if ri == 2 else (FILL_GRY if ri % 2 == 0 else FILL_WHT)
        vals = [ri - 1, row["nombre"],
                row.get("cat_h", 0), row.get("cat_i", 0),
                row.get("cat_j", 0), row.get("cat_k", 0),
                row.get("cat_l", 0), row.get("cat_m", 0),
                row.get("cat_n", 0), row.get("cat_o", 0),
                row["pts_globales"], row["pts_total"]]
        for ci, v in enumerate(vals, 1):
            _cell(ws1, ri, ci, v, fill,
                  FNT_9B if ci in (1, len(vals)) else FNT_9,
                  AL if ci == 2 else AC)
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(hdr1))}1"

    # ══════════════════════════════════════════════════════════════════════
    # FICHA 2 — Por Fase
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Por Fase")
    # Build ordered phase list
    fase_order: list = []
    fase_seen: set = set()
    for row in sorted(det_rows, key=lambda r: (r["fase_orden"], r["fase"] or "")):
        fn = row["fase"] or "Sin fase"
        if fn not in fase_seen:
            fase_order.append(fn)
            fase_seen.add(fn)
    # Pivot: apostador → {fase: pts_total}
    fase_pts: dict = {}
    for row in det_rows:
        aid = row["apostador_id"]
        fn  = row["fase"] or "Sin fase"
        fase_pts.setdefault(aid, {})
        fase_pts[aid][fn] = fase_pts[aid].get(fn, 0) + row["pts_total"]
    hdr2 = ["Apostador"] + fase_order + [ICON["T"]]
    wdths2 = [18] + [max(10, min(18, len(f))) for f in fase_order] + [8]
    for ci, (h, w) in enumerate(zip(hdr2, wdths2), 1):
        fill2 = HDR_GRN if ci in (1, len(hdr2)) else HDR_BLU
        _cell(ws2, 1, ci, h, fill2, FNT_W9B, AC)
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[1].height = 24
    ws2.freeze_panes = "A2"
    for ri, row in enumerate(rk_rows, 2):
        fill = FILL_GRN if ri == 2 else (FILL_GRY if ri % 2 == 0 else FILL_WHT)
        fp = fase_pts.get(row["apostador_id"], {})
        fase_vals = [fp.get(fn, 0) for fn in fase_order]
        all_vals = [row["nombre"]] + fase_vals + [sum(fase_vals)]
        for ci, v in enumerate(all_vals, 1):
            _cell(ws2, ri, ci, v, fill,
                  FNT_9B if ci in (1, len(all_vals)) else FNT_9,
                  AL if ci == 1 else AC)
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(hdr2))}1"

    # ══════════════════════════════════════════════════════════════════════
    # FICHA 3 — Por Partido
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Por Partido")
    # Layout (2-row header):
    # 1  2    3     4       5         6       7      8        9
    # Ap P#  Fase  Local  Visitante Pronóst Real  ICON[H]  ICON[I]
    # [J: 10,11,12] [K:13,14,15] [L:16,17,18] [M:19,20,21] [N:22,23,24]
    # [O: 25,26,27,28,29]  Total=30
    C_AP=1; C_FA=2; C_PN=3; C_LO=4; C_VI=5
    C_PR=6; C_RE=7; C_H=8; C_I=9
    C_J=10; C_K=13; C_L=16; C_M=19; C_N=22
    C_OL=25; C_OV=28; C_P=31; C_TOT=34

    def _gh3(c1, c2, label, fill):
        if c2 > c1:
            ws3.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        _cell(ws3, 1, c1, label, fill, FNT_W9B, AC)

    def _sh3(c, label):
        _cell(ws3, 2, c, label, HDR_SUB, FNT_W9B, AC)

    _gh3(C_AP, C_AP, "Apostador",  HDR_GRN)
    _gh3(C_FA, C_FA, "Fase",       HDR_GRN)
    _gh3(C_PN, C_PN, "P#",         HDR_GRN)
    _gh3(C_LO, C_LO, "Local",      HDR_GRN)
    _gh3(C_VI, C_VI, "Visitante",  HDR_GRN)
    _gh3(C_PR, C_RE, "Marcador",   HDR_GRN)
    _gh3(C_H,  C_H,  ICON["H"],    HDR_BLU)
    _gh3(C_I,  C_I,  ICON["I"],    HDR_BLU)
    for base, key in [(C_J,"J"),(C_K,"K"),(C_L,"L"),(C_M,"M"),(C_N,"N")]:
        _gh3(base, base+2, ICON[key], HDR_BLU)
    _gh3(C_OL, C_OL+2, "Tanda Local",     HDR_BLU)
    _gh3(C_OV, C_OV+2, "Tanda Visitante", HDR_BLU)
    _gh3(C_P,  C_P+2,  "País que Pasa",    HDR_BLU)
    _gh3(C_TOT, C_TOT, ICON["T"],          HDR_GRN)

    for c, lbl in [(C_AP,"Apostador"),(C_FA,"Fase"),(C_PN,"P#"),
                   (C_LO,"Local"),(C_VI,"Visitante"),
                   (C_PR,"Pronóst."),(C_RE,"Real"),
                   (C_H,"Pts"),(C_I,"Pts")]:
        _sh3(c, lbl)
    for base in [C_J,C_K,C_L,C_M,C_N]:
        for off, lbl in enumerate(["Pred","Real","Pts"]): _sh3(base+off, lbl)
    for base in [C_OL, C_OV, C_P]:
        for off, lbl in enumerate(["Pred","Real","Pts"]): _sh3(base+off, lbl)
    _sh3(C_TOT, ICON["T"])

    ws3.row_dimensions[1].height = 22
    ws3.row_dimensions[2].height = 20
    ws3.freeze_panes = "A3"

    for c, w in [(C_AP,16),(C_FA,14),(C_PN,5),(C_LO,16),(C_VI,16),
                 (C_PR,9),(C_RE,9),(C_H,5),(C_I,5)]:
        ws3.column_dimensions[get_column_letter(c)].width = w
    for base in [C_J,C_K,C_L,C_M,C_N]:
        for off in range(3): ws3.column_dimensions[get_column_letter(base+off)].width = 6
    for base in [C_OL, C_OV]:
        for off in range(3): ws3.column_dimensions[get_column_letter(base+off)].width = 6
    for off in range(2): ws3.column_dimensions[get_column_letter(C_P+off)].width = 14
    ws3.column_dimensions[get_column_letter(C_P+2)].width = 6
    ws3.column_dimensions[get_column_letter(C_TOT)].width = 7
    ws3.auto_filter.ref = f"A2:{get_column_letter(C_TOT)}2"

    # ── Sort: apostador → fase_orden → numero_fifa ─────────────────────────────
    _FASE_TIPO_ORDER = {
        'grupo': 0, 'ronda32': 1, 'ronda16': 2,
        'cuartos': 3, 'semis': 4, 'tercer_puesto': 5, 'final': 6
    }
    det_sorted = sorted(det_rows, key=lambda r: (
        r["nombre"],
        _FASE_TIPO_ORDER.get(r.get("fase_tipo",""), r.get("fase_orden", 99)),
        r.get("numero_fifa") or 0
    ))

    _FASE_LABELS = {
        'grupo': 'Grupos', 'ronda32': 'R32', 'ronda16': 'Octavos',
        'cuartos': 'Cuartos', 'semis': 'Semis', 'tercer_puesto': '3er Puesto', 'final': 'Final'
    }
    HDR_APOS   = _hf("1A3A5C")   # dark blue — apostador header row
    HDR_FASE   = _hf("2D5A7C")   # medium blue — phase separator row
    FILL_FADE1 = _hf("EAF4FB")   # very light blue — odd data row within block
    FILL_FADE2 = _hf("F5FAFD")   # near-white — even data row within block

    left_align_cols = {C_AP, C_FA, C_LO, C_VI, C_P, C_P+1}

    # Build pts_total per apostador from rk_rows for the section header
    _apos_pts = {r["nombre"]: r.get("pts_total", 0) for r in rk_rows}

    ri = 3   # current Excel row (rows 1-2 are header)
    prev_nombre = None
    prev_fase   = None
    block_row   = 0   # alternating color counter within a block

    for fr in det_sorted:
        nombre   = fr["nombre"]
        fase_typ = fr.get("fase_tipo", "")
        fase_lbl = _FASE_LABELS.get(fase_typ, fr.get("fase", fase_typ))
        pnum     = fr.get("numero_fifa") or 0

        # ── Apostador section header ────────────────────────────────────────
        if nombre != prev_nombre:
            apos_pts = _apos_pts.get(nombre, 0)
            # Merge all columns for the header row
            ws3.merge_cells(start_row=ri, start_column=1,
                            end_row=ri, end_column=C_TOT - 1)
            c_h = ws3.cell(ri, 1)
            c_h.value   = nombre
            c_h.fill    = HDR_APOS
            c_h.font    = Font(color="FFFFFF", bold=True, size=10)
            c_h.alignment = Alignment(horizontal="left", vertical="center")
            c_h.border  = BDR
            # Total pts at the end
            ct = ws3.cell(ri, C_TOT)
            ct.value = apos_pts
            ct.fill  = HDR_APOS
            ct.font  = Font(color="FFFFFF", bold=True, size=10)
            ct.alignment = AC
            ct.border = BDR
            ri += 1
            prev_nombre = nombre
            prev_fase   = None
            block_row   = 0

        # ── Phase separator ─────────────────────────────────────────────────
        if fase_lbl != prev_fase:
            ws3.merge_cells(start_row=ri, start_column=1,
                            end_row=ri, end_column=C_TOT)
            cf = ws3.cell(ri, 1)
            cf.value = f"  {fase_lbl}"
            cf.fill  = HDR_FASE
            cf.font  = Font(color="DDEEEE", bold=True, size=8, italic=True)
            cf.alignment = Alignment(horizontal="left", vertical="center")
            cf.border = BDR
            ri += 1
            prev_fase = fase_lbl
            block_row = 0

        # ── Data row ─────────────────────────────────────────────────────────
        block_row += 1
        fill = FILL_FADE1 if block_row % 2 == 1 else FILL_FADE2

        def _d3(c, v, bold=False, fov=None, _ri=ri, _fill=fill):
            _cell(ws3, _ri, c, v, fov or _fill,
                  FNT_9B if bold else FNT_9,
                  AL if c in left_align_cols else AC)

        _d3(C_AP, nombre, bold=True)
        _d3(C_PN, f"P{pnum}" if pnum else "")
        _d3(C_FA, fr.get("fase",""))
        _d3(C_LO, fr.get("equipo_local",""))
        _d3(C_VI, fr.get("equipo_visitante",""))
        _d3(C_PR, _marc(fr.get("pred_local"), fr.get("pred_visitante")))
        _d3(C_RE, _marc(fr.get("goles_local"), fr.get("goles_visitante")))
        _d3(C_H,  fr["pts_h"],  bold=fr["pts_h"]>0)
        _d3(C_I,  fr["pts_i"],  bold=fr["pts_i"]>0)
        for base, pk, rk_, ptk in [
            (C_J,"pred_amarillas","real_amarillas","pts_j"),
            (C_K,"pred_rojas","real_rojas","pts_k"),
            (C_L,"pred_var","real_var","pts_l"),
            (C_M,"pred_penales_partido","real_pen_partido","pts_m"),
            (C_N,"pred_minuto_gol","real_minuto","pts_n"),
        ]:
            _d3(base,    _vv(fr.get(pk)))
            _d3(base+1,  _vv(fr.get(rk_)))
            _d3(base+2,  fr.get(ptk,0), bold=fr.get(ptk,0)>0)
        # OL / OV / P — solo si hubo tanda de penales (empate en KO)
        _hubo_tanda = fr.get("real_pen_local") is not None
        if _hubo_tanda:
            _pl  = fr.get("pred_penales_local_tanda")
            _rl  = fr.get("real_pen_local")
            _pv_ = fr.get("pred_penales_visitante_tanda")
            _rv  = fr.get("real_pen_visitante")
            _pts_o = fr.get("pts_o", 0)
            _hit_ol = _pl is not None and _rl is not None and str(_pl) == str(_rl) and str(_rl) != ''
            _hit_ov = _pv_ is not None and _rv is not None and str(_pv_) == str(_rv) and str(_rv) != ''
            _hits = (1 if _hit_ol else 0) + (1 if _hit_ov else 0)
            if _hits == 0:   _pts_ol, _pts_ov = 0, 0
            elif _hits == 1: _pts_ol = _pts_o if _hit_ol else 0; _pts_ov = _pts_o if _hit_ov else 0
            else:            _pts_ol = _pts_o // 2; _pts_ov = _pts_o - _pts_ol
            _d3(C_OL,   _vv(_pl));  _d3(C_OL+1, _vv(_rl));  _d3(C_OL+2, _pts_ol, bold=_pts_ol>0)
            _d3(C_OV,   _vv(_pv_)); _d3(C_OV+1, _vv(_rv));  _d3(C_OV+2, _pts_ov, bold=_pts_ov>0)
            _d3(C_P,    fr.get("pred_equipo_clasifica_nm","") or "")
            _d3(C_P+1,  fr.get("real_equipo_clasifica_nm","") or "")
            _d3(C_P+2,  fr.get("pts_p",0), bold=fr.get("pts_p",0)>0)
        else:
            for _c in [C_OL,C_OL+1,C_OL+2,C_OV,C_OV+1,C_OV+2,C_P,C_P+1,C_P+2]:
                _d3(_c, "")
        pts_t = fr.get("pts_total", 0)
        _d3(C_TOT, pts_t, bold=True, fov=FILL_GRN if pts_t > 0 else None)
        ri += 1

    # ══════════════════════════════════════════════════════════════════════
    # FICHA 4 — Clasificados por Fase
    # ══════════════════════════════════════════════════════════════════════
    # 4a. Grupos — desde apostador_clasificados
    try:
        r_clas_grupo = await db.execute(text("""
            SELECT apostador_id, COALESCE(aciertos, 0)::int AS aciertos
            FROM apostador_clasificados
            WHERE torneo_id = :tid AND fase_tipo = 'grupo'
        """), {"tid": torneo_id})
        clas_grupo_map = {r["apostador_id"]: int(r["aciertos"] or 0) for r in r_clas_grupo.mappings()}
    except Exception:
        await db.rollback()
        clas_grupo_map = {}

    # 4b. KO — desde puntaje_detalle.pts_equipo
    try:
        r_clas_ko = await db.execute(text("""
            SELECT pd.apostador_id, f.tipo AS fase_tipo,
                   COUNT(*) FILTER (WHERE pd.pts_equipo > 0) AS aciertos,
                   COUNT(DISTINCT pd.partido_id) AS total_partidos
            FROM puntaje_detalle pd
            JOIN partido p ON p.id = pd.partido_id
            JOIN fase f ON f.id = p.fase_id
            WHERE pd.torneo_id = :tid
              AND LOWER(f.tipo) NOT LIKE 'grupo%'
              AND p.equipo_clasificado_id IS NOT NULL
            GROUP BY pd.apostador_id, f.tipo
        """), {"tid": torneo_id})
        clas_ko_rows = [dict(r) for r in r_clas_ko.mappings()]
    except Exception:
        await db.rollback()
        clas_ko_rows = []

    # 4c. Total de equipos clasificados por fase KO (denominadores)
    try:
        r_tot_ko = await db.execute(text("""
            SELECT f.tipo AS fase_tipo, COUNT(DISTINCT p.id) AS total_partidos
            FROM partido p
            JOIN fase f ON f.id = p.fase_id
            WHERE f.torneo_id = :tid
              AND LOWER(f.tipo) NOT LIKE 'grupo%'
              AND p.equipo_clasificado_id IS NOT NULL
              AND p.estado = 'finalizado'
            GROUP BY f.tipo
        """), {"tid": torneo_id})
        totales_ko_map = {r["fase_tipo"]: int(r["total_partidos"] or 0) for r in r_tot_ko.mappings()}
    except Exception:
        await db.rollback()
        totales_ko_map = {}

    _FASE_TIPO_ORDER4 = ["grupo", "ronda32", "ronda16", "cuartos", "semis", "tercer_puesto", "final"]
    _FASE_LABELS4 = {
        "grupo": "Grupos\n(de 32)", "ronda32": "R32\n(de 16)", "ronda16": "Octavos\n(de 8)",
        "cuartos": "Cuartos\n(de 4)", "semis": "Semis\n(de 2)",
        "tercer_puesto": "3er Puesto\n(de 1)", "final": "Campeón\n(de 1)",
    }
    _FASE_DENOM4 = {"grupo": 32, "ronda32": 16, "ronda16": 8, "cuartos": 4,
                    "semis": 2, "tercer_puesto": 1, "final": 1}

    # Construir mapa {aid: {fase_tipo: aciertos}}
    clas_map4: dict = {}
    for row in rk_rows:
        aid = row["apostador_id"]
        clas_map4[aid] = {"grupo": clas_grupo_map.get(aid, 0)}
    for row in clas_ko_rows:
        aid = row["apostador_id"]
        if aid not in clas_map4:
            clas_map4[aid] = {}
        clas_map4[aid][row["fase_tipo"]] = int(row["aciertos"] or 0)

    # Solo incluir fases con datos (al menos un partido finalizado o aciertos > 0)
    fases_con_datos4 = [
        ft for ft in _FASE_TIPO_ORDER4
        if ft == "grupo" or totales_ko_map.get(ft, 0) > 0
        or any(clas_map4.get(r["apostador_id"], {}).get(ft, 0) > 0 for r in rk_rows)
    ]

    ws4 = wb.create_sheet("Clasificados por Fase")
    hdr4 = ["#", "Apostador"] + [_FASE_LABELS4.get(ft, ft) for ft in fases_con_datos4] + ["Total"]
    wdths4 = [4, 18] + [12] * len(fases_con_datos4) + [8]
    for ci, (h, w) in enumerate(zip(hdr4, wdths4), 1):
        _cell(ws4, 1, ci, h, HDR_GRN, FNT_W9B, AC)
        ws4.column_dimensions[get_column_letter(ci)].width = w
    ws4.row_dimensions[1].height = 30
    ws4.freeze_panes = "A2"

    # Ordenar por total de aciertos desc
    def _total_clas4(row):
        aid = row["apostador_id"]
        return sum(clas_map4.get(aid, {}).get(ft, 0) for ft in fases_con_datos4)

    rk_rows_clas4 = sorted(rk_rows, key=lambda r: -_total_clas4(r))

    for ri4, row in enumerate(rk_rows_clas4, 2):
        aid   = row["apostador_id"]
        fill4 = FILL_GRN if ri4 == 2 else (FILL_GRY if ri4 % 2 == 0 else FILL_WHT)
        aciertos_vals = []
        for ft in fases_con_datos4:
            ac  = clas_map4.get(aid, {}).get(ft, 0)
            tot = _FASE_DENOM4.get(ft, totales_ko_map.get(ft, 0))
            aciertos_vals.append((ac, tot))
        total_ac = sum(a for a, _ in aciertos_vals)
        all_v4 = [ri4 - 1, row["nombre"]] + [f"{a}/{t}" if t > 0 else str(a) for a, t in aciertos_vals] + [total_ac]
        for ci, v in enumerate(all_v4, 1):
            _cell(ws4, ri4, ci, v, fill4,
                  FNT_9B if ci in (1, len(all_v4)) else FNT_9,
                  AL if ci == 2 else AC)
    ws4.auto_filter.ref = f"A1:{get_column_letter(len(hdr4))}1"

    # ══════════════════════════════════════════════════════════════════════
    # FICHA 5 — Puntaje Detallado por Apostador
    # ══════════════════════════════════════════════════════════════════════
    try:
        r_pts_fase = await db.execute(text("""
            SELECT pd.apostador_id, f.tipo AS fase_tipo,
                   COALESCE(SUM(pd.pts_resultado),       0)::int AS ph,
                   COALESCE(SUM(pd.pts_marcador),        0)::int AS pi,
                   COALESCE(SUM(pd.pts_amarillas),       0)::int AS pj,
                   COALESCE(SUM(pd.pts_rojas),           0)::int AS pk,
                   COALESCE(SUM(pd.pts_var),             0)::int AS pl,
                   COALESCE(SUM(pd.pts_penales_partido), 0)::int AS pm,
                   COALESCE(SUM(pd.pts_minuto),          0)::int AS pn,
                   COALESCE(SUM(pd.pts_penales_tanda),   0)::int AS po,
                   COALESCE(SUM(pd.pts_equipo),          0)::int AS pp,
                   COALESCE(SUM(pd.pts_total),           0)::int AS pts_fase
            FROM puntaje_detalle pd
            JOIN partido p ON p.id = pd.partido_id
            JOIN fase f ON f.id = p.fase_id
            WHERE pd.torneo_id = :tid
            GROUP BY pd.apostador_id, f.tipo
        """), {"tid": torneo_id})
        pts_fase_rows = [dict(r) for r in r_pts_fase.mappings()]
    except Exception:
        await db.rollback()
        pts_fase_rows = []

    # {aid: {fase_tipo: {ph..pp, pts_fase}}}
    pts_fase_map5: dict = {}
    for row in pts_fase_rows:
        pts_fase_map5.setdefault(row["apostador_id"], {})[row["fase_tipo"]] = row

    _FASE_TIPO_ORDER5 = ["grupo", "ronda32", "ronda16", "cuartos", "semis", "tercer_puesto", "final"]
    _FASE_LABELS5 = {
        "grupo": "Grupos", "ronda32": "R32", "ronda16": "Octavos",
        "cuartos": "Cuartos", "semis": "Semis", "tercer_puesto": "3er Puesto", "final": "Final",
    }
    fases_con_datos5 = [
        ft for ft in _FASE_TIPO_ORDER5
        if any(pts_fase_map5.get(r["apostador_id"], {}).get(ft) for r in rk_rows)
    ]

    _ITEMS5   = ["H",  "I",    "J",    "K",    "L",   "M",       "N",      "O",         "P",     "Tot"]
    _ILBLS5   = {"H":"Res","I":"Exac","J":"Amar","K":"Rojas","L":"VAR","M":"PenPdo","N":"1erGol","O":"PenTanda","P":"Clásif","Tot":"Total"}
    _IKEYS5   = {"H":"ph","I":"pi","J":"pj","K":"pk","L":"pl","M":"pm","N":"pn","O":"po","P":"pp","Tot":"pts_fase"}

    _F5_FILLS = [_hf("1A4A7A"), _hf("1A5A3A"), _hf("4A2A8A"), _hf("7A3A1A"),
                 _hf("1A6A6A"), _hf("5A4A1A"), _hf("3A1A5A")]

    # Construir mapa de columnas
    C5_IDX = 1; C5_NOM = 2
    fase5_col: dict = {}   # {fase_tipo: col_inicio}
    cur5 = 3
    for ft in fases_con_datos5:
        fase5_col[ft] = cur5
        cur5 += len(_ITEMS5)
    C5_GLOB  = cur5;     cur5 += 1
    C5_GRPP  = cur5;     cur5 += 1
    C5_TOTAL = cur5

    ws5 = wb.create_sheet("Puntaje Detallado")

    # Anchos de columna
    ws5.column_dimensions[get_column_letter(C5_IDX)].width = 4
    ws5.column_dimensions[get_column_letter(C5_NOM)].width = 18
    for ft in fases_con_datos5:
        cs5 = fase5_col[ft]
        for i in range(len(_ITEMS5)):
            ws5.column_dimensions[get_column_letter(cs5 + i)].width = 6
    ws5.column_dimensions[get_column_letter(C5_GLOB)].width  = 8
    ws5.column_dimensions[get_column_letter(C5_GRPP)].width  = 8
    ws5.column_dimensions[get_column_letter(C5_TOTAL)].width = 8

    # Fila 1: cabeceras de fase (merged)
    _cell(ws5, 1, C5_IDX, "#",          HDR_GRN, FNT_W9B, AC)
    _cell(ws5, 1, C5_NOM, "Apostador",  HDR_GRN, FNT_W9B, AC)
    for fi5, ft in enumerate(fases_con_datos5):
        cs5 = fase5_col[ft]
        ce5 = cs5 + len(_ITEMS5) - 1
        ff5 = _F5_FILLS[fi5 % len(_F5_FILLS)]
        if ce5 > cs5:
            ws5.merge_cells(start_row=1, start_column=cs5, end_row=1, end_column=ce5)
        _cell(ws5, 1, cs5, _FASE_LABELS5.get(ft, ft), ff5, FNT_W9B, AC)
    _cell(ws5, 1, C5_GLOB,  "Globales\n(A-G)", HDR_BLU, FNT_W9B, AC)
    _cell(ws5, 1, C5_GRPP,  "Clásif.\nGrupos",  HDR_BLU, FNT_W9B, AC)
    _cell(ws5, 1, C5_TOTAL, "TOTAL",             HDR_GRN, FNT_W9B, AC)
    ws5.row_dimensions[1].height = 28

    # Fila 2: sub-ítems
    _cell(ws5, 2, C5_IDX, "#",          HDR_SUB, FNT_W9B, AC)
    _cell(ws5, 2, C5_NOM, "Apostador",  HDR_SUB, FNT_W9B, AC)
    for ft in fases_con_datos5:
        cs5 = fase5_col[ft]
        for i5, itm in enumerate(_ITEMS5):
            _cell(ws5, 2, cs5 + i5, _ILBLS5[itm], HDR_SUB, FNT_W9B, AC)
    _cell(ws5, 2, C5_GLOB,  "Pts", HDR_SUB, FNT_W9B, AC)
    _cell(ws5, 2, C5_GRPP,  "Pts", HDR_SUB, FNT_W9B, AC)
    _cell(ws5, 2, C5_TOTAL, "Pts", HDR_SUB, FNT_W9B, AC)
    ws5.row_dimensions[2].height = 20
    ws5.freeze_panes = "A3"

    for ri5, row in enumerate(rk_rows, 3):
        aid   = row["apostador_id"]
        fill5 = FILL_GRN if ri5 == 3 else (FILL_GRY if ri5 % 2 == 0 else FILL_WHT)
        _cell(ws5, ri5, C5_IDX, ri5 - 2,        fill5, FNT_9B, AC)
        _cell(ws5, ri5, C5_NOM, row["nombre"],   fill5, FNT_9,  AL)
        for ft in fases_con_datos5:
            cs5 = fase5_col[ft]
            fd5 = pts_fase_map5.get(aid, {}).get(ft, {})
            for i5, itm in enumerate(_ITEMS5):
                v5 = int(fd5.get(_IKEYS5[itm], 0) or 0)
                is_tot = (itm == "Tot")
                fill_v = FILL_GRN if (is_tot and v5 > 0) else fill5
                _cell(ws5, ri5, cs5 + i5, v5, fill_v, FNT_9B if (is_tot and v5 > 0) else FNT_9, AC)
        _cell(ws5, ri5, C5_GLOB,  row.get("pts_globales", 0), fill5, FNT_9,  AC)
        _cell(ws5, ri5, C5_GRPP,  row.get("pts_grupos_p", 0), fill5, FNT_9,  AC)
        _cell(ws5, ri5, C5_TOTAL, row.get("pts_total",    0), fill5, FNT_9B, AC)
    ws5.auto_filter.ref = f"A2:{get_column_letter(C5_TOTAL)}2"

    # ── Save & stream ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = _dt.now().strftime("%Y%m%d_%H%M")
    fname = f"becbuc_ranking_torneo{torneo_id}_{ts}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# GET /bets/exportar-puntajes/{torneo_id}
# Excel con dos hojas: v_copamundial_puntajes + v_copamundial_puntajes_det
# ─────────────────────────────────────────────────────────────────────────────
