# -*- coding: utf-8 -*-
"""
comparar_resultados_todos.py
Compara el RESULTADO de TODOS los partidos: marcador (grupos y KO) +
quien pasa (SOLO KO; en grupos el pase es por tabla de posiciones).
entre el Excel (hoja '40- RESULTADOS OFICIALES') y la BD (tabla partido).
Lista un renglon por partido (P001..P104) marcando OK / DIF. Solo lectura.

Uso:
  backend\.venv\Scripts\python.exe comparar_resultados_todos.py
"""
import sys, os
BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = None
for f in os.listdir(BASE):
    if 'SEMIFINAL' in f.upper() and f.endswith('.xlsx'):
        EXCEL_FILE = os.path.join(BASE, f); break
if not EXCEL_FILE:
    sys.exit(f"ERROR: no se encontro Excel *SEMIFINAL*.xlsx en {BASE}")
print(f"Excel: {os.path.basename(EXCEL_FILE)}")

try:
    import openpyxl
except ImportError:
    os.system(f'"{sys.executable}" -m pip install openpyxl --quiet'); import openpyxl
try:
    import psycopg2, psycopg2.extras
except ImportError:
    os.system(f'"{sys.executable}" -m pip install psycopg2-binary --quiet'); import psycopg2, psycopg2.extras

CONN = "host=localhost port=5432 dbname=becbuc user=app_user password=superpassword"
TID = 2
conn = psycopg2.connect(CONN); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

cur.execute("SELECT id, nombre, nombre_es FROM equipo")
eid = {}; ename = {}
for eq in cur.fetchall():
    ename[eq['id']] = eq['nombre'] or eq['nombre_es']
    if eq['nombre']:    eid[eq['nombre'].upper().strip()] = eq['id']
    if eq['nombre_es']: eid[eq['nombre_es'].upper().strip()] = eq['id']
ALIAS = {'FRANCIA':'France','ESPANA':'Spain','INGLATERRA':'England','ARGENTINA':'Argentina',
    'MARRUECOS':'Morocco','BELGICA':'Belgium','NORUEGA':'Norway','SUIZA':'Switzerland',
    'COLOMBIA':'Colombia','MEXICO':'Mexico','BRASIL':'Brazil','PORTUGAL':'Portugal',
    'PARAGUAY':'Paraguay','CANADA':'Canada','EGIPTO':'Egypt','ESTADOS UNIDOS':'USA',
    'EE UU':'USA','EEUU':'USA','ALEMANIA':'Germany'}
def find_eid(nom):
    if not nom: return None
    k = str(nom).upper().strip()
    if k in eid: return eid[k]
    a = ALIAS.get(k)
    if a and a.upper() in eid: return eid[a.upper()]
    for kk, vv in eid.items():
        if k in kk or kk in k: return vv
    return None
def ti(v):
    try:
        s = str(v).strip()
        if s in ('','-','None'): return None
        return int(float(s))
    except: return None

# Excel official all matches
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb['40- RESULTADOS OFICIALES']
C = dict(pid=1, gl=12, gv=14, pasa=37)
excel = {}
for r in range(2, ws.max_row + 1):
    pid = ws.cell(r, C['pid']).value
    if isinstance(pid, str) and pid.startswith('P'):
        excel[pid] = (ti(ws.cell(r, C['gl']).value), ti(ws.cell(r, C['gv']).value),
                      find_eid(ws.cell(r, C['pasa']).value), ws.cell(r, C['pasa']).value)

# BD all matches
cur.execute("""
    SELECT p.numero_fifa, p.goles_local AS gl, p.goles_visitante AS gv,
           p.equipo_clasificado_id AS pasa, p.estado,
           el.nombre AS local, ev.nombre AS visit
    FROM partido p JOIN fase f ON f.id=p.fase_id
    LEFT JOIN equipo el ON el.id=p.equipo_local_id
    LEFT JOIN equipo ev ON ev.id=p.equipo_visitante_id
    WHERE f.torneo_id=%s AND p.numero_fifa BETWEEN 1 AND 104
    ORDER BY p.numero_fifa
""", (TID,))
bd = {f"P{r['numero_fifa']:03d}": dict(r) for r in cur.fetchall()}

print("\n" + "="*86)
print(f"{'PART':<6}{'PARTIDO':<34}{'EXCEL':<10}{'BD':<10}{'PASA(Excel/BD)':<22}EST")
print("="*86)
diffs_marc = 0; diffs_pasa = 0; comparados = 0; solo_bd = 0
for n in range(1, 105):
    pid = f"P{n:03d}"
    ex = excel.get(pid); db = bd.get(pid)
    if not db:
        continue
    nombre = f"{(db['local'] or '?')[:15]} vs {(db['visit'] or '?')[:14]}"
    if not ex:
        print(f"{pid:<6}{nombre:<34}{'--':<10}{str(db['gl'])+'-'+str(db['gv']):<10}{'(no en Excel)':<22}{db['estado'][:4]}")
        continue
    comparados += 1
    es_ko = n >= 73   # grupos (P001-P072) pasan por tabla, no por partido
    ex_marc = f"{ex[0]}-{ex[1]}"; bd_marc = f"{db['gl']}-{db['gv']}"
    marc_ok = (ex[0] == db['gl'] and ex[1] == db['gv'])
    pasa_ok = True
    if es_ko:
        pasa_ok = (ex[2] == db['pasa'])
        if (ex[2] is not None or db['pasa'] is not None) and not pasa_ok:
            diffs_pasa += 1
    if not marc_ok: diffs_marc += 1
    flag = '' if (marc_ok and pasa_ok) else '  <<'
    pasa_txt = (f"{(ex[3] or '-')}/{ename.get(db['pasa'],'-')}") if es_ko else '(grupos: por tabla)'
    if marc_ok and pasa_ok:
        estado = 'OK'
    else:
        estado = ('MARC ' if not marc_ok else '') + ('PASA' if not pasa_ok else '')
    print(f"{pid:<6}{nombre:<34}{ex_marc:<10}{bd_marc:<10}{pasa_txt:<22}{estado}{flag}")

print("="*86)
print(f"Comparados: {comparados}  |  Diferencias de marcador: {diffs_marc}  |  Diferencias de 'quien pasa': {diffs_pasa}")
if diffs_marc == 0 and diffs_pasa == 0:
    print("RESULTADO: TODOS los marcadores y clasificados coinciden Excel == BD.")
else:
    print("RESULTADO: hay diferencias (ver renglones marcados con << ).")
conn.close()
