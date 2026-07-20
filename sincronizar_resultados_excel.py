# -*- coding: utf-8 -*-
r"""
sincronizar_resultados_excel.py  [--apply]

Alinea el RESULTADO OFICIAL de cada partido en la BD (tabla partido) con la hoja
'40- RESULTADOS OFICIALES' del Excel de cierre, para poder analizar SOLO el
algoritmo de puntaje (con predicciones + resultados identicos, cualquier diff de
puntos que quede es del algoritmo).

Campos alineados (Excel -> partido):
  goles_local, goles_visitante, amarillas, rojas, decisiones_var (VAR),
  penales_partido, minuto_primer_gol, penales_local, penales_visitante (tanda)

Autodetecta las columnas por su encabezado y las IMPRIME (verificar antes de aplicar).
NUNCA sobrescribe con vacio: si el Excel no trae dato, deja el de la BD.

SIN --apply : DRY RUN (solo muestra diferencias).
CON --apply : escribe en partido. Luego: run_reabrir_y_recalcular.bat

Uso:
  backend\.venv\Scripts\python.exe sincronizar_resultados_excel.py
  backend\.venv\Scripts\python.exe sincronizar_resultados_excel.py --apply
"""
import sys, os, re
_a = [a.lower() for a in sys.argv[1:]]
DO_APPLY = '--apply' in _a
INCLUIR_TANDA = '--incluir-tanda' in _a   # por defecto NO se alinea la tanda (None = sin definicion por penales)
print(f"[{'APPLY - escribe en partido' if DO_APPLY else 'DRY RUN - no escribe'}]"
      + ('' if INCLUIR_TANDA else '   (tanda EXCLUIDA: el Excel usa 0/99 centinela; None de la BD es lo correcto)'))

BASE = os.path.dirname(os.path.abspath(__file__))
_OK = ('TORNEO CERRADO', 'FIN DE TORNEO', '20260720'); _BAD = ('CORRECCIONES', 'SEMIFINAL')
_c = [os.path.join(BASE, f) for f in os.listdir(BASE)
      if f.lower().endswith('.xlsx') and not f.startswith('~')
      and any(k in f.upper() for k in _OK) and not any(b in f.upper() for b in _BAD)]
if not _c:
    sys.exit("ERROR: no encontre el Excel de cierre en la raiz.")
EXCEL = sorted(_c, key=lambda p: os.path.getmtime(p), reverse=True)[0]
print(f"Excel: {os.path.basename(EXCEL)}")

try:
    import openpyxl
except ImportError:
    os.system(f'"{sys.executable}" -m pip install openpyxl --quiet'); import openpyxl
try:
    import psycopg2, psycopg2.extras
except ImportError:
    os.system(f'"{sys.executable}" -m pip install psycopg2-binary --quiet'); import psycopg2, psycopg2.extras

CONN_BEC = "host=localhost port=5432 dbname=becbuc user=app_user password=superpassword"
TID = 2
FLAG_VAR = {63, 64, 65, 76, 79, 83, 104}

def to_int(v):
    try:
        s = str(v).strip()
        if s in ('', '-', 'None', 'nan'): return None
        return int(float(s))
    except Exception:
        return None

wb = openpyxl.load_workbook(EXCEL, data_only=True)
SHEET = next((s for s in wb.sheetnames if 'RESULTADOS OFICIALES' in s.upper()), None) \
        or next((s for s in wb.sheetnames if 'OFICIAL' in s.upper()), None)
if not SHEET:
    sys.exit("ERROR: no encontre la hoja '40- RESULTADOS OFICIALES'.")
ws = wb[SHEET]
print(f"Hoja: '{SHEET}'  ({ws.max_row}x{ws.max_column})")

# fila de encabezado = la de mas texto en las primeras 5
hdr_row, best = 1, -1
for r in range(1, 6):
    cnt = sum(1 for c in range(1, ws.max_column + 1) if isinstance(ws.cell(r, c).value, str))
    if cnt > best: best, hdr_row = cnt, r
headers = {c: str(ws.cell(hdr_row, c).value).strip()
           for c in range(1, ws.max_column + 1)
           if ws.cell(hdr_row, c).value not in (None, '')}
print(f"Fila de encabezado: {hdr_row}")
print("Encabezados:")
for c, h in headers.items():
    print(f"   col {c:>2}: {h}")

def find_col(*keys, avoid=()):
    for c, h in headers.items():
        u = h.upper()
        if any(k in u for k in keys) and not any(a in u for a in avoid):
            return c
    return None
def find_all(*keys, avoid=()):
    out = []
    for c, h in headers.items():
        u = h.upper()
        if any(k in u for k in keys) and not any(a in u for a in avoid):
            out.append(c)
    return out

col_pid  = find_col('ID PARTIDO', 'NRO PARTIDO', 'N PARTIDO', 'PARTIDO', avoid=('EQUIPO', 'GOL'))
goles    = find_all('GOLES', avoid=('GOLEAD',))            # [local, visitante]
col_gl   = goles[0] if len(goles) >= 1 else None
col_gv   = goles[1] if len(goles) >= 2 else None
col_amar = find_col('AMARILLA')
col_roja = find_col('ROJA', avoid=('AMARILLA',))
col_var  = find_col('VAR')
col_penj = find_col('PENALES', 'PENAL', avoid=('TANDA',))
col_min  = find_col('1ER GOL', 'PRIMER GOL', 'MINUTO', avoid=('GOLES',))
tandas   = find_all('TANDA')                                # [local, visitante]
col_tl   = tandas[0] if len(tandas) >= 1 else None
col_tv   = tandas[1] if len(tandas) >= 2 else None

# (campo_bd, col_excel, etiqueta)
MAP = [
    ('goles_local', col_gl, 'goles_local'),
    ('goles_visitante', col_gv, 'goles_visit'),
    ('amarillas', col_amar, 'amarillas'),
    ('rojas', col_roja, 'rojas'),
    ('decisiones_var', col_var, 'VAR'),
    ('penales_partido', col_penj, 'penales_juego'),
    ('minuto_primer_gol', col_min, 'minuto_1er_gol'),
]
if INCLUIR_TANDA:
    MAP += [('penales_local', col_tl, 'tanda_local'),
            ('penales_visitante', col_tv, 'tanda_visit')]
print("\nColumnas detectadas (verificar!):")
for bd_f, col, lbl in MAP:
    print(f"   {lbl:<16} -> col {col}  ({headers.get(col, '-') if col else 'NO DETECTADA'})")
if not col_pid:
    sys.exit("\n*** No detecte la columna de ID PARTIDO. Revisar encabezados y avisame. ***")

# BD
conn = psycopg2.connect(CONN_BEC); conn.autocommit = False
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
cur.execute("""SELECT p.id, p.numero_fifa, p.goles_local, p.goles_visitante, p.amarillas,
                      p.rojas, p.decisiones_var, p.penales_partido, p.minuto_primer_gol,
                      p.penales_local, p.penales_visitante
               FROM partido p JOIN fase f ON f.id=p.fase_id WHERE f.torneo_id=%s""", (TID,))
bd = {r['numero_fifa']: r for r in cur.fetchall()}

def pid_to_nf(v):
    if v is None: return None
    m = re.match(r'[Pp]?(\d{1,3})$', str(v).strip())
    return int(m.group(1)) if m else None

changes = []   # (nf, campo_bd, lbl, excel, bd)
from collections import Counter
by_field = Counter()
for r in range(hdr_row + 1, ws.max_row + 1):
    nf = pid_to_nf(ws.cell(r, col_pid).value)
    if nf is None or nf not in bd:
        continue
    b = bd[nf]
    for bd_f, col, lbl in MAP:
        if not col: continue
        ex = to_int(ws.cell(r, col).value)
        if ex is None:             # el Excel no trae dato -> no tocar
            continue
        # descartar centinelas del Excel (98/99/999 = "sin dato", 99 tanda = null)
        if lbl in ('amarillas', 'rojas', 'VAR', 'penales_juego') and ex >= 20:
            continue
        if lbl == 'minuto_1er_gol' and (ex >= 130 or ex in (98, 99, 998, 999)):
            continue
        if lbl in ('tanda_local', 'tanda_visit') and (ex >= 20 or ex in (0, 98, 99)):
            continue
        if ex != b[bd_f]:
            changes.append((nf, bd_f, lbl, ex, b[bd_f]))
            by_field[lbl] += 1

print("\n" + "=" * 78)
print("DIFERENCIAS DE RESULTADO (Excel oficial -> BD)  -- lo que se alinearia")
print("=" * 78)
print(f"{'PART':<6}{'campo':<16}{'Excel':>8}{'BD':>8}   nota")
for nf, bd_f, lbl, ex, bv in sorted(changes, key=lambda x: (x[1], x[0])):
    nota = ''
    if lbl == 'VAR' and nf in FLAG_VAR: nota = '<== VAR flag auditoria'
    if nf == 104: nota = (nota + '  OJO final (correccion sesion70)').strip()
    print(f"P{nf:<5}{lbl:<16}{str(ex):>8}{str(bv):>8}   {nota}")
print("\nResumen por campo:")
for k, v in by_field.most_common():
    print(f"   {k}: {v}")
print(f"TOTAL cambios de resultado: {len(changes)}")

if not DO_APPLY:
    print("\n[DRY RUN] No se escribio nada.")
    print("Aplicar:  sincronizar_resultados_excel.py --apply")
    conn.close(); sys.exit(0)

# APPLY
print("\nAPLICANDO cambios de resultado...")
per_partido = {}
for nf, bd_f, lbl, ex, bv in changes:
    per_partido.setdefault(nf, {})[bd_f] = ex
n = 0
for nf, sets in per_partido.items():
    setclause = ', '.join(f"{c}=%({c})s" for c in sets)
    cur.execute(f"""UPDATE partido SET {setclause}
                    WHERE id=(SELECT p.id FROM partido p JOIN fase f ON f.id=p.fase_id
                              WHERE f.torneo_id=%(tid)s AND p.numero_fifa=%(nf)s)""",
                {**sets, 'tid': TID, 'nf': nf})
    n += cur.rowcount
conn.commit()
print(f"Partidos actualizados: {n}  ({len(changes)} campos)")
print("\nOK. Resultados alineados al Excel.")
print("SIGUIENTE PASO:  aplicar predicciones (verificar_apuestas_fin_torneo.py --apply)")
print("                 luego  run_reabrir_y_recalcular.bat")
conn.close()
