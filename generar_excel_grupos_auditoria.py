"""
generar_excel_grupos_auditoria.py
Genera Excel de auditoria de fase de grupos con totales por item H-O + Bonus P (R32).
Ejecutar: backend\.venv\Scripts\python generar_excel_grupos_auditoria.py
"""
import os as _osp
_BASE = _osp.path.dirname(_osp.path.abspath(__file__))
import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

BECBUC_DSN  = "host=localhost dbname=becbuc user=app_user"
APP_DB_DSN  = "host=localhost dbname=app_db user=app_user"
TORNEO_ID   = 2
OUT_PATH    = _osp.path.join(_BASE, 'auditoria_grupos_becbuc.xlsx')

# ── Colores ──────────────────────────────────────────────────────────────────
C_HEADER_BG  = "1F2937"   # gris oscuro
C_HEADER_FG  = "F9FAFB"   # blanco
C_PLENO      = "D1FAE5"   # verde claro (marcador exacto)
C_ACIERTO    = "FEF3C7"   # amarillo (resultado correcto)
C_FALLO      = "FEE2E2"   # rojo claro (fallo)
C_BONUS_BG   = "1E3A5F"   # azul oscuro (bonus P)
C_BONUS_FG   = "93C5FD"   # azul claro texto
C_TOTAL_BG   = "374151"   # gris (total)
C_TOTAL_FG   = "F9FAFB"
C_SUBTOT_BG  = "111827"
C_SUBTOT_FG  = "D1D5DB"
C_HIT_TEAM   = "065F46"   # verde oscuro fondo equipo acertado
C_MISS_TEAM  = "374151"   # gris equipo fallado
C_HIT_TXT    = "A7F3D0"
C_MISS_TXT   = "6B7280"

thin = Side(style="thin", color="374151")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell_style(ws, row, col, value=None, bold=False, bg=None, fg=None,
               align="center", wrap=False, border_on=True, fmt=None, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=fg or "000000",
                  size=10, italic=italic)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border_on:
        c.border = border
    if fmt:
        c.number_format = fmt
    return c

def header(ws, row, col, txt, w=None):
    c = cell_style(ws, row, col, txt, bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG)
    if w:
        ws.column_dimensions[get_column_letter(col)].width = w
    return c

# ── DB queries ────────────────────────────────────────────────────────────────
def fetch_data():
    bc = psycopg2.connect(BECBUC_DSN)
    bc.autocommit = True
    cur = bc.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # 1. Nombres desde app_db
    try:
        ap = psycopg2.connect(APP_DB_DSN)
        ap.autocommit = True
        ac = ap.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        ac.execute("""
            SELECT u.id, u.username AS apostador
            FROM users u
            JOIN user_roles ur ON ur.user_id = u.id
            JOIN roles ro ON ro.id = ur.role_id
            WHERE ro.name = 'apostador' AND u.is_active = TRUE
            ORDER BY u.username
        """)
        apostadores = {r["id"]: r["apostador"] for r in ac.fetchall()}
        ap.close()
    except Exception:
        apostadores = {}

    # 2. Totales por item por apostador (fase grupos)
    cur.execute("""
        SELECT
            pd.apostador_id,
            COALESCE(SUM(pd.pts_resultado),       0)::int AS H,
            COALESCE(SUM(pd.pts_marcador),        0)::int AS I,
            COALESCE(SUM(pd.pts_amarillas),       0)::int AS J,
            COALESCE(SUM(pd.pts_rojas),           0)::int AS K,
            COALESCE(SUM(pd.pts_var),             0)::int AS L,
            COALESCE(SUM(pd.pts_penales_partido), 0)::int AS M,
            COALESCE(SUM(pd.pts_minuto),          0)::int AS N,
            COALESCE(SUM(pd.pts_penales_tanda),   0)::int AS O,
            COUNT(*)::int AS partidos,
            COUNT(*) FILTER (WHERE pd.pts_marcador > 0)::int AS plenos,
            COUNT(*) FILTER (WHERE pd.pts_resultado > 0 AND pd.pts_marcador = 0)::int AS aciertos
        FROM puntaje_detalle pd
        JOIN partido p ON p.id = pd.partido_id
        JOIN fase f ON f.id = p.fase_id
        WHERE f.torneo_id = %s AND LOWER(f.tipo) LIKE '%%grupo%%'
        GROUP BY pd.apostador_id
    """, (TORNEO_ID,))
    totales = {r["apostador_id"]: dict(r) for r in cur.fetchall()}

    # 3. Bonus P grupos (apostador_clasificados)
    cur.execute("""
        SELECT apostador_id, aciertos, pts_obtenidos, pts_por_acierto,
               equipos_pronosticados, equipos_reales
        FROM apostador_clasificados
        WHERE torneo_id = %s AND fase_tipo = 'grupo'
    """, (TORNEO_ID,))
    bonusp = {r["apostador_id"]: dict(r) for r in cur.fetchall()}

    # 4. Detalle partido × apostador
    cur.execute("""
        SELECT
            pd.apostador_id,
            p.numero_fifa,
            COALESCE(el.nombre_es, el.nombre) AS local,
            p.goles_local   AS gl,
            p.goles_visitante AS gv,
            COALESCE(ev.nombre_es, ev.nombre) AS visitante,
            a.pred_local    AS pl,
            a.pred_visitante AS pv,
            a.pred_amarillas, a.pred_rojas, a.pred_var,
            a.pred_penales_partido, a.pred_minuto_gol,
            p.amarillas, p.rojas, p.decisiones_var,
            p.penales_partido, p.minuto_primer_gol,
            pd.pts_resultado  AS H,
            pd.pts_marcador   AS I,
            pd.pts_amarillas  AS J,
            pd.pts_rojas      AS K,
            pd.pts_var        AS L,
            pd.pts_penales_partido AS M,
            pd.pts_minuto     AS N,
            pd.pts_penales_tanda AS O,
            f.nombre AS fase_nombre
        FROM puntaje_detalle pd
        JOIN partido p ON p.id = pd.partido_id
        JOIN fase f ON f.id = p.fase_id
        LEFT JOIN equipo el ON el.id = p.equipo_local_id
        LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
        LEFT JOIN apuesta a ON a.partido_id = p.id AND a.apostador_id = pd.apostador_id
        WHERE f.torneo_id = %s AND LOWER(f.tipo) LIKE '%%grupo%%'
        ORDER BY p.numero_fifa, pd.apostador_id
    """, (TORNEO_ID,))
    detalle = cur.fetchall()

    # 5. Equipos R32 con nombres
    cur.execute("""
        SELECT id, aciertos FROM apostador_clasificados
        WHERE torneo_id = %s AND fase_tipo = 'grupo' LIMIT 1
    """, (TORNEO_ID,))
    sample = cur.fetchone()

    cur.execute("""
        SELECT DISTINCT e.id, COALESCE(e.nombre_es, e.nombre) AS nombre
        FROM (
            SELECT unnest(equipos_reales) AS eid
            FROM apostador_clasificados
            WHERE torneo_id = %s AND fase_tipo = 'grupo' LIMIT 1
        ) sub
        JOIN equipo e ON e.id = sub.eid
        ORDER BY nombre
    """, (TORNEO_ID,))
    equipos_r32 = [(r["id"], r["nombre"]) for r in cur.fetchall()]

    bc.close()
    return apostadores, totales, bonusp, detalle, equipos_r32


# ── Sheet 1: Resumen Grupos ───────────────────────────────────────────────────
def build_resumen(wb, apostadores, totales, bonusp):
    ws = wb.create_sheet("Resumen Grupos")
    ws.freeze_panes = "C3"

    ITEMS = ["H", "I", "J", "K", "L", "M", "N", "O"]
    ITEM_LABEL = {
        "H": "⚽ Resultado",
        "I": "🎯 Marcador exacto",
        "J": "🟨 Amarillas",
        "K": "🟥 Rojas",
        "L": "📺 VAR",
        "M": "🥅 Penales partido",
        "N": "⏱ Minuto gol",
        "O": "🏆 Penales tanda",
    }

    # Título
    ws.merge_cells("A1:P1")
    t = ws["A1"]
    t.value = "BECBUC 2026 · Auditoría Fase de Grupos · Puntajes por Ítem + Bonus Equipos Clasificados (R32)"
    t.font = Font(name="Arial", bold=True, size=13, color="F9FAFB")
    t.fill = PatternFill("solid", start_color="0F172A")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Headers fila 2
    item_hdrs = [ITEM_LABEL[k] for k in ITEMS]
    cols   = ["#", "Apostador"] + item_hdrs + ["Subtotal\nPartidos", "Plenos", "Aciertos",
                                                 "🏅 Bonus P\n(Equipos R32)", "Aciertos/32",
                                                 "TOTAL\nGRUPOS"]
    widths = [4, 22] + [14]*8 + [11, 8, 8, 12, 10, 12]
    for ci, (lbl, w) in enumerate(zip(cols, widths), 1):
        c = header(ws, 2, ci, lbl, w)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 36

    # Construir ranking
    ranked = []
    for aid, alias in apostadores.items():
        t_row = totales.get(aid, {})
        bp    = bonusp.get(aid, {})
        sub   = sum(t_row.get(k, 0) for k in ITEMS)
        bonus = bp.get("pts_obtenidos", 0) or 0
        total = sub + bonus
        ranked.append((aid, alias, t_row, bp, sub, bonus, total))
    ranked.sort(key=lambda x: -x[6])

    fills_pos = ["FFD700", "C0C0C0", "CD7F32"]  # oro, plata, bronce

    for i, (aid, alias, t_row, bp, sub, bonus, total) in enumerate(ranked, 1):
        r = i + 2
        ws.row_dimensions[r].height = 18

        # Pos
        pos_fill = fills_pos[i-1] if i <= 3 else None
        cell_style(ws, r, 1, i, bold=(i<=3), bg=pos_fill, fg="000000" if pos_fill else None)

        # Nombre
        cell_style(ws, r, 2, alias, bold=(i<=3), align="left",
                   bg="1F2937" if i % 2 == 0 else None,
                   fg="F9FAFB" if i % 2 == 0 else "E5E7EB")

        # Items H-O
        for ci, item in enumerate(ITEMS, 3):
            v = t_row.get(item, 0) or 0
            bg = None
            if v > 0:
                bg = "D1FAE5" if item == "I" else "EFF6FF"
            cell_style(ws, r, ci, v or None, fmt="0",
                       bg=bg, fg="065F46" if (item=="I" and v>0) else None)

        # Subtotal partidos
        cell_style(ws, r, 11, sub, bold=True, fmt="0",
                   bg=C_SUBTOT_BG, fg=C_SUBTOT_FG)

        # Plenos / Aciertos
        cell_style(ws, r, 12, t_row.get("plenos", 0) or 0, fmt="0")
        cell_style(ws, r, 13, t_row.get("aciertos", 0) or 0, fmt="0")

        # Bonus P
        cell_style(ws, r, 14, bonus if bonus else None, bold=True, fmt="0",
                   bg=C_BONUS_BG if bonus else None,
                   fg=C_BONUS_FG if bonus else None)

        # Aciertos/32
        aciertos32 = bp.get("aciertos", 0) or 0
        cell_style(ws, r, 15, f"{aciertos32}/32",
                   fg="34D399" if aciertos32 >= 27 else ("FBBF24" if aciertos32 >= 25 else None))

        # Total
        cell_style(ws, r, 16, total, bold=True, fmt="0",
                   bg=C_TOTAL_BG, fg=C_TOTAL_FG)

    # Totales fila
    last = len(ranked) + 3
    ws.row_dimensions[last].height = 20
    cell_style(ws, last, 1, "", bg=C_HEADER_BG)
    cell_style(ws, last, 2, "TOTALES", bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG, align="left")
    for ci in range(3, 17):
        col_letter = get_column_letter(ci)
        ws.cell(last, ci).value = f"=SUM({col_letter}3:{col_letter}{last-1})"
        ws.cell(last, ci).font = Font(name="Arial", bold=True, color=C_HEADER_FG, size=10)
        ws.cell(last, ci).fill = PatternFill("solid", start_color=C_HEADER_BG)
        ws.cell(last, ci).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(last, ci).border = border
        ws.cell(last, ci).number_format = "0"

    ws.auto_filter.ref = f"A2:{get_column_letter(16)}{last-1}"
    return ws


# ── Sheet 2: Detalle Partidos ─────────────────────────────────────────────────
def build_detalle(wb, apostadores, detalle):
    ws = wb.create_sheet("Detalle Partidos")
    ws.freeze_panes = "D3"

    ITEMS = ["H", "I", "J", "K", "L", "M", "N", "O"]
    ITEM_LABEL = {
        "H": "⚽ Resultado",
        "I": "🎯 Marcador exacto",
        "J": "🟨 Amarillas",
        "K": "🟥 Rojas",
        "L": "📺 VAR",
        "M": "🥅 Penales partido",
        "N": "⏱ Minuto gol",
        "O": "🏆 Penales tanda",
    }
    item_hdrs = [ITEM_LABEL[k] for k in ITEMS]
    hdr_cols = ["Apostador", "P#", "Fase", "Local", "GL", "GV", "Visitante",
                "Pred.L", "Pred.V"] + item_hdrs + ["Total Partido"]
    widths   = [22, 5, 20, 22, 5, 5, 22, 7, 7] + [14]*8 + [11]

    ws.merge_cells("A1:U1")
    t = ws["A1"]
    t.value = "BECBUC 2026 · Detalle Partidos Fase de Grupos"
    t.font = Font(name="Arial", bold=True, size=12, color="F9FAFB")
    t.fill = PatternFill("solid", start_color="0F172A")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    for ci, (lbl, w) in enumerate(zip(hdr_cols, widths), 1):
        header(ws, 2, ci, lbl, w)

    row = 3
    prev_num = None
    for d in detalle:
        aid = d["apostador_id"]
        if aid not in apostadores:
            continue

        bg_alt = "1A2332" if d["numero_fifa"] != prev_num and prev_num is not None else None
        prev_num = d["numero_fifa"]

        pts_items = [d.get(k) or 0 for k in ITEMS]
        total_p   = sum(pts_items)

        vals = [
            apostadores.get(aid, f"uid{aid}"),
            d["numero_fifa"],
            d["fase_nombre"],
            d["local"],
            d["gl"],
            d["gv"],
            d["visitante"],
            d["pl"],
            d["pv"],
        ] + pts_items + [total_p]

        for ci, v in enumerate(vals, 1):
            bg = None
            fg = None
            item_idx = ci - 10  # 0-based index into ITEMS (cols 10-17)
            if 0 <= item_idx < 8 and v:
                item = ITEMS[item_idx]
                if item == "I" and v > 0:
                    bg, fg = C_PLENO, "065F46"
                elif item == "H" and v > 0:
                    bg, fg = C_ACIERTO, "78350F"
                elif v > 0:
                    bg = "EFF6FF"
            if ci == len(vals) and v > 0:
                bg, fg = C_SUBTOT_BG, C_SUBTOT_FG

            cell_style(ws, row, ci, v if v != 0 else None,
                       align="left" if ci in (1, 3, 4, 7) else "center",
                       bg=bg, fg=fg, fmt="0" if ci >= 10 else None)
        row += 1

    ws.auto_filter.ref = f"A2:{get_column_letter(len(hdr_cols))}2"
    return ws


# ── Sheet 3: Grupos P (Lista de 32) ──────────────────────────────────────────
def build_grupos_p(wb, apostadores, bonusp, equipos_r32):
    ws = wb.create_sheet("Grupos P (R32)")
    ws.freeze_panes = "C3"

    # Título
    n_eq = len(equipos_r32)
    n_cols = 4 + n_eq
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    t = ws["A1"]
    t.value = "BECBUC 2026 · Bonus P — Equipos pronosticados a R32 (✅ Acertado | ❌ Fallado)"
    t.font = Font(name="Arial", bold=True, size=12, color="F9FAFB")
    t.fill = PatternFill("solid", start_color="0F172A")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Fila 2 — encabezados
    header(ws, 2, 1, "#",          4)
    header(ws, 2, 2, "Apostador",  22)
    header(ws, 2, 3, "Aciertos",   10)
    header(ws, 2, 4, "Bonus P",    9)
    for ci, (eid, enombre) in enumerate(equipos_r32, 5):
        c = ws.cell(row=2, column=ci, value=enombre)
        c.font = Font(name="Arial", bold=True, size=8, color=C_HEADER_FG)
        c.fill = PatternFill("solid", start_color=C_BONUS_BG)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                text_rotation=90, wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = 4
    ws.row_dimensions[2].height = 90

    # Construir ranking por aciertos desc
    ranked_bp = sorted(
        [(aid, alias) for aid, alias in apostadores.items() if aid in bonusp],
        key=lambda x: -(bonusp[x[0]].get("aciertos", 0) or 0)
    )

    # Set de equipos reales
    real_set = set(equipos_r32[0][0] for _ in [0]) if equipos_r32 else set()
    if bonusp:
        sample_bp = next(iter(bonusp.values()))
        real_set = set(sample_bp.get("equipos_reales") or [])

    fills_pos = ["FFD700", "C0C0C0", "CD7F32"]

    for i, (aid, alias) in enumerate(ranked_bp, 1):
        r = i + 2
        bp = bonusp[aid]
        aciertos = bp.get("aciertos", 0) or 0
        pts      = bp.get("pts_obtenidos", 0) or 0
        pred_set = set(bp.get("equipos_pronosticados") or [])

        pos_fill = fills_pos[i-1] if i <= 3 else None
        cell_style(ws, r, 1, i, bold=(i<=3), bg=pos_fill, fg="000000" if pos_fill else None)
        cell_style(ws, r, 2, alias, bold=(i<=3), align="left",
                   bg="1F2937" if i % 2 == 0 else None,
                   fg="F9FAFB" if i % 2 == 0 else "E5E7EB")
        cell_style(ws, r, 3, aciertos, bold=True, fmt="0",
                   fg="34D399" if aciertos >= 27 else ("FBBF24" if aciertos >= 25 else "F87171"))
        cell_style(ws, r, 4, pts, bold=True, fmt="0", bg=C_BONUS_BG, fg=C_BONUS_FG)

        for ci, (eid, _) in enumerate(equipos_r32, 5):
            hit = eid in pred_set
            cell_style(ws, r, ci, "✅" if hit else "❌",
                       bg=C_HIT_TEAM if hit else None,
                       fg=C_HIT_TXT if hit else C_MISS_TXT,
                       bold=hit)

    # Fila frecuencias (cuántos acertaron cada equipo)
    last = len(ranked_bp) + 3
    ws.row_dimensions[last].height = 18
    cell_style(ws, last, 1, "", bg=C_HEADER_BG)
    cell_style(ws, last, 2, "Apostadores que acertaron →", bold=True, bg=C_HEADER_BG,
               fg=C_HEADER_FG, align="left")
    cell_style(ws, last, 3, "", bg=C_HEADER_BG)
    cell_style(ws, last, 4, "", bg=C_HEADER_BG)
    for ci, (eid, _) in enumerate(equipos_r32, 5):
        cnt = sum(
            1 for (aid, _) in ranked_bp
            if eid in set(bonusp[aid].get("equipos_pronosticados") or [])
        )
        pct = cnt / max(len(ranked_bp), 1) * 100
        c = ws.cell(row=last, column=ci, value=cnt)
        c.font = Font(name="Arial", bold=True, size=9,
                      color="34D399" if pct >= 80 else ("FBBF24" if pct >= 50 else "F87171"))
        c.fill = PatternFill("solid", start_color="0F172A")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        c.number_format = "0"

    return ws


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    print(f"[{datetime.now():%H:%M:%S}] Cargando datos...")
    apostadores, totales, bonusp, detalle, equipos_r32 = fetch_data()

    print(f"  Apostadores: {len(apostadores)}")
    print(f"  Con bonus P: {len(bonusp)}")
    print(f"  Partidos detalle: {len(detalle)}")
    print(f"  Equipos R32: {len(equipos_r32)}")

    wb = Workbook()
    wb.remove(wb.active)  # eliminar hoja default

    print("Generando hojas...")
    build_resumen(wb, apostadores, totales, bonusp)
    build_detalle(wb, apostadores, detalle)
    build_grupos_p(wb, apostadores, bonusp, equipos_r32)

    wb.save(OUT_PATH)
    print(f"\n✅ Excel generado: {OUT_PATH}")
    print(f"   Hojas: {wb.sheetnames}")
    os.startfile(OUT_PATH)

if __name__ == "__main__":
    main()
