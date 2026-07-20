#!/usr/bin/env python3
"""
VERIFICACION / IMPORTACION: Pronósticos Excel vs Base de Datos BECBUC
Hoja fuente: 50- TBL MASTER

Secciones:
  A) Fase de grupos (P001-P072): compara pred_local/visitante + bonus
  B) Globales (P111-P118):       compara/importa apuesta_global A-G

Mapeo globales Excel → BD:
  P111  CAMPEON DEL MUNDO           → pred_campeon_id + pred_finalista1_id
  P112  EL OTRO FINALISTA           → pred_finalista2_id
  P113  GOLEADOR DEL MUNDIAL        → pred_goleador (texto)
  P114  PEOR EQUIPO DEL MUNDIAL     → pred_peor_equipo_id
  P115  HASTA QUE FASE LLEGA PARAGUAY → pred_etapa_paraguay (texto)
  P116  CUANTOS GOLES ANOTARA PARAGUAY → pred_goles_paraguay (int)
  P117  MAYOR GOLEADA - GOLES GANADOR  → pred_goleada_ganador (int)
  P118  MAYOR GOLEADA - GOLES PERDEDOR → pred_goleada_perdedor (int)

Uso:
  Verificar grupos (por defecto):
      .venv\Scripts\python ..\comparar_pronosticos.py

  Ver globales de Excel:
      .venv\Scripts\python ..\comparar_pronosticos.py --globales

  Importar globales a BD:
      .venv\Scripts\python ..\comparar_pronosticos.py --importar-globales

  Combinado (grupos + globales):
      .venv\Scripts\python ..\comparar_pronosticos.py --globales --grupos

Ejecutar desde C:\proyecto FAST API\backend
"""
import sys, asyncio, json
from pathlib import Path
from collections import defaultdict

EXCEL_PATH = (sys.argv[1] if len(sys.argv) > 1 and not sys.argv[1].startswith('--')
    else r"C:\proyecto FAST API\documentacion\20260611_2000- TBL CONSOLIDADA PRONOSTICOS ok.xlsx")

DB_BECBUC  = "postgresql://app_user:superpassword@localhost:5432/becbuc"
DB_APP     = "postgresql://app_user:superpassword@localhost:5432/app_db"
REPORT_OUT = Path(r"C:\proyecto FAST API\documentacion\verificacion_pronosticos.json")
TORNEO_ID  = 2  # Copa del Mundo 2026

ARGS = set(sys.argv[1:])
MODO_GRUPOS   = '--grupos'    in ARGS or '--globales' not in ARGS  # default=True
MODO_GLOBALES = '--globales'  in ARGS or '--importar-globales' in ARGS
IMPORTAR_GL   = '--importar-globales' in ARGS

# ─────────────────────────────────────────────────────────────────────────────
# Campos grupo
# ─────────────────────────────────────────────────────────────────────────────
CAMPOS = [
    ('gl',      'Goles L'),
    ('gv',      'Goles V'),
    ('amar',    'Amarillas'),
    ('rojas',   'Rojas'),
    ('var',     'VAR'),
    ('pen',     'Penales'),
    ('min_gol', 'Min.gol'),
]

# Mapeo P-numero → campo en apuesta_global
PID_A_CAMPO = {
    111: 'campeon',      # pred_campeon_id (equipo)
    112: 'finalista2',   # pred_finalista2_id (el otro finalista)
    113: 'goleador',     # pred_goleador (texto)
    114: 'peor_equipo',  # pred_peor_equipo_id (equipo)
    115: 'etapa_py',     # pred_etapa_paraguay (texto)
    116: 'goles_py',     # pred_goles_paraguay (int)
    117: 'goleada_g',    # pred_goleada_ganador (int)
    118: 'goleada_p',    # pred_goleada_perdedor (int)
}

# Campos que son texto libre (no necesitan equipo.id)
CAMPOS_TEXTO = {'goleador', 'etapa_py'}
# Campos que son enteros
CAMPOS_INT   = {'goles_py', 'goleada_g', 'goleada_p'}
# Campos que necesitan equipo.id
CAMPOS_EQUIPO = {'campeon', 'finalista2', 'peor_equipo'}

# Normalización de nombres de equipos (español Excel → como puede aparecer en BD)
NOMBRE_ES_NORM = {
    'españa':           ['spain', 'españa', 'espana'],
    'alemania':         ['germany', 'alemania', 'deutschland'],
    'brasil':           ['brazil', 'brasil'],
    'paises bajos':     ['netherlands', 'paises bajos', 'holland', 'holanda'],
    'marruecos':        ['morocco', 'marruecos'],
    'argentina':        ['argentina'],
    'france':           ['france', 'francia'],
    'francia':          ['france', 'francia'],
    'inglaterra':       ['england', 'inglaterra'],
    'estados unidos':   ['usa', 'united states', 'estados unidos'],
    'eeuu':             ['usa', 'united states'],
    'corea del sur':    ['south korea', 'korea republic', 'corea del sur'],
    'japon':            ['japan', 'japon', 'japón'],
    'japón':            ['japan', 'japon', 'japón'],
    'australia':        ['australia'],
    'suecia':           ['sweden', 'suecia'],
    'dinamarca':        ['denmark', 'dinamarca'],
    'belgica':          ['belgium', 'belgica', 'bélgica'],
    'bélgica':          ['belgium', 'belgica', 'bélgica'],
    'portugal':         ['portugal'],
    'italia':           ['italy', 'italia'],
    'mexico':           ['mexico', 'méxico'],
    'méxico':           ['mexico', 'méxico'],
    'colombia':         ['colombia'],
    'ecuador':          ['ecuador'],
    'peru':             ['peru', 'perú'],
    'perú':             ['peru', 'perú'],
    'chile':            ['chile'],
    'uruguay':          ['uruguay'],
    'bolivia':          ['bolivia'],
    'venezuela':        ['venezuela'],
    'paraguay':         ['paraguay'],
    'canada':           ['canada', 'canadá'],
    'canadá':           ['canada', 'canadá'],
    'senegal':          ['senegal'],
    'ghana':            ['ghana'],
    'nigeria':          ['nigeria'],
    'camerun':          ['cameroon', 'camerun', 'camerún'],
    'camerún':          ['cameroon', 'camerun', 'camerún'],
    'sudafrica':        ['south africa', 'sudafrica', 'sudáfrica'],
    'sudáfrica':        ['south africa', 'sudafrica', 'sudáfrica'],
    'egipto':           ['egypt', 'egipto'],
    'argelia':          ['algeria', 'argelia'],
    'tunez':            ['tunisia', 'tunez', 'túnez'],
    'túnez':            ['tunisia', 'tunez', 'túnez'],
    'suiza':            ['switzerland', 'suiza'],
    'austria':          ['austria'],
    'croacia':          ['croatia', 'croacia'],
    'polonia':          ['poland', 'polonia'],
    'hungria':          ['hungary', 'hungria', 'hungría'],
    'hungría':          ['hungary', 'hungria', 'hungría'],
    'rumania':          ['romania', 'rumania'],
    'rumanía':          ['romania', 'rumania'],
    'eslovaquia':       ['slovakia', 'eslovaquia'],
    'eslovenia':        ['slovenia', 'eslovenia'],
    'serbia':           ['serbia'],
    'turquia':          ['turkey', 'turquia', 'türkiye'],
    'turquía':          ['turkey', 'turquia', 'türkiye'],
    'grecia':           ['greece', 'grecia'],
    'iran':             ['iran', 'irán'],
    'irán':             ['iran', 'irán'],
    'irak':             ['iraq', 'irak'],
    'arabia saudita':   ['saudi arabia', 'arabia saudita'],
    'qatar':            ['qatar'],
    'emiratos':         ['uae', 'emiratos árabes', 'united arab emirates'],
    'australia':        ['australia'],
    'nueva zelanda':    ['new zealand', 'nueva zelanda'],
    'curazao':          ['curaçao', 'curacao', 'curazao'],
    'jamaica':          ['jamaica'],
    'costa rica':       ['costa rica'],
    'panama':           ['panama', 'panamá'],
    'panamá':           ['panama', 'panamá'],
    'honduras':         ['honduras'],
    'guatemala':        ['guatemala'],
    'el salvador':      ['el salvador'],
}

def _norm_team_name(name: str) -> str:
    """Normaliza nombre de equipo para búsqueda."""
    return name.strip().lower()

def equiv(a, b):
    """None y 0 son equivalentes."""
    if a == b: return True
    return (a is None or a == 0) and (b is None or b == 0)

# ─────────────────────────────────────────────────────────────────────────────
# 1. Leer Excel — grupos
# ─────────────────────────────────────────────────────────────────────────────
def leer_excel_grupos(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['50- TBL MASTER']
    rows = list(ws.iter_rows(values_only=True))

    data = {}
    aliases = {}

    for r in rows[1:]:
        if not r[1]: continue
        fase = str(r[6] or '')
        if '10-' not in fase and 'GRUPO' not in fase.upper(): continue

        alias = str(r[9] or '').replace('\xa0', '').strip()
        nombre = str(r[8] or '').strip()
        pid_str = str(r[1] or '').strip()
        try:
            pid_int = int(pid_str[1:])
        except (ValueError, IndexError):
            continue

        alias_low = alias.lower()
        aliases[alias_low] = (alias, nombre)
        data[(alias_low, pid_int)] = {
            'alias': alias, 'nombre': nombre,
            'pid': pid_str, 'pid_int': pid_int,
            'gl': r[12], 'gv': r[14],
            'amar': r[23], 'rojas': r[24],
            'var': r[25], 'pen': r[26], 'min_gol': r[27],
        }

    return data, aliases

# ─────────────────────────────────────────────────────────────────────────────
# 2. Leer Excel — globales (P111-P118)
# ─────────────────────────────────────────────────────────────────────────────
def leer_excel_globales(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['50- TBL MASTER']
    rows = list(ws.iter_rows(values_only=True))

    # {alias_low: {campo_key: valor_raw, 'alias': str, 'nombre': str}}
    globales = {}

    for r in rows[1:]:
        if not r[1]: continue
        pid_str = str(r[1] or '').strip()
        try:
            pid_int = int(pid_str[1:])
        except (ValueError, IndexError):
            continue

        if pid_int not in PID_A_CAMPO:
            continue

        alias = str(r[9] or '').replace('\xa0', '').strip()
        nombre = str(r[8] or '').strip()
        valor = r[10]   # columna EQUIPO 1

        alias_low = alias.lower()
        if alias_low not in globales:
            globales[alias_low] = {'alias': alias, 'nombre': nombre}

        campo = PID_A_CAMPO[pid_int]
        globales[alias_low][campo] = valor

    return globales

# ─────────────────────────────────────────────────────────────────────────────
# 3. Leer BD
# ─────────────────────────────────────────────────────────────────────────────
async def leer_bd_grupos():
    import asyncpg

    conn_app = await asyncpg.connect(DB_APP)
    users = await conn_app.fetch("SELECT id, username FROM users ORDER BY id")
    await conn_app.close()

    uid_by_username  = {u['username'].lower(): (u['id'], u['username']) for u in users}
    username_by_uid  = {u['id']: u['username'].lower() for u in users}

    conn = await asyncpg.connect(DB_BECBUC)
    rows = await conn.fetch("""
        SELECT
            a.apostador_id,
            p.id                 AS partido_id,
            a.pred_local,
            a.pred_visitante,
            a.pred_amarillas,
            a.pred_rojas,
            a.pred_var,
            a.pred_penales_partido,
            a.pred_minuto_gol
        FROM apuesta a
        JOIN partido p ON p.id  = a.partido_id
        JOIN fase    f ON f.id  = p.fase_id
        WHERE f.tipo = 'grupo'
        ORDER BY a.apostador_id, p.id
    """)
    await conn.close()

    bd_data = {}
    unknown_uids = set()
    for r in rows:
        uid = r['apostador_id']
        uname_low = username_by_uid.get(uid)
        if uname_low is None:
            unknown_uids.add(uid)
            continue
        key = (uname_low, r['partido_id'])
        bd_data[key] = {
            'gl':    r['pred_local'],
            'gv':    r['pred_visitante'],
            'amar':  r['pred_amarillas'],
            'rojas': r['pred_rojas'],
            'var':   r['pred_var'],
            'pen':   r['pred_penales_partido'],
            'min_gol': r['pred_minuto_gol'],
        }

    return bd_data, uid_by_username, unknown_uids


async def leer_bd_globales_y_equipos():
    """Retorna (equipos_dict, globales_bd_dict, uid_by_username)."""
    import asyncpg

    conn_app = await asyncpg.connect(DB_APP)
    users    = await conn_app.fetch("SELECT id, username FROM users ORDER BY id")
    await conn_app.close()
    uid_by_username = {u['username'].lower(): (u['id'], u['username']) for u in users}
    uid_by_id       = {u['id']: u['username'].lower() for u in users}

    conn = await asyncpg.connect(DB_BECBUC)

    # Equipos: nombre y nombre_es para mapeo
    equipos_rows = await conn.fetch(
        "SELECT id, nombre, nombre_es FROM equipo ORDER BY id"
    )

    # Globales existentes en BD
    gl_rows = await conn.fetch(
        """
        SELECT apostador_id,
               pred_campeon_id, pred_finalista1_id, pred_finalista2_id,
               pred_goleador, pred_peor_equipo_id,
               pred_goleada_ganador, pred_goleada_perdedor,
               pred_etapa_paraguay, pred_goles_paraguay
        FROM apuesta_global
        WHERE torneo_id = $1
        """,
        TORNEO_ID
    )

    await conn.close()

    # {nombre_low → id}  — acepta nombre y nombre_es
    equipos = {}
    equipos_por_id = {}
    for eq in equipos_rows:
        equipos_por_id[eq['id']] = eq['nombre']
        for col in ('nombre', 'nombre_es'):
            v = eq[col]
            if v:
                k = _norm_team_name(v)
                if k not in equipos:
                    equipos[k] = eq['id']

    # {apostador_low → dict campos}
    globales_bd = {}
    for r in gl_rows:
        uname = uid_by_id.get(r['apostador_id'])
        if uname:
            globales_bd[uname] = dict(r)

    return equipos, equipos_por_id, globales_bd, uid_by_username


# ─────────────────────────────────────────────────────────────────────────────
# Mapeo nombre equipo Excel → equipo.id
# ─────────────────────────────────────────────────────────────────────────────
def resolver_equipo_id(nombre_excel, equipos_dict):
    """
    Intenta mapear nombre de equipo del Excel (español) a equipo.id en BD.
    Retorna (equipo_id, matched_name) o (None, motivo).
    """
    if not nombre_excel:
        return None, 'vacío'
    n = _norm_team_name(str(nombre_excel))

    # 1) Búsqueda directa (si en BD ya está en español o inglés)
    if n in equipos_dict:
        return equipos_dict[n], n

    # 2) Buscar via alias conocidos
    aliases = NOMBRE_ES_NORM.get(n, [])
    for alias in aliases:
        if alias in equipos_dict:
            return equipos_dict[alias], alias

    # 3) Búsqueda parcial (substring)
    for eq_name, eq_id in equipos_dict.items():
        if n in eq_name or eq_name in n:
            return eq_id, f'~{eq_name}'

    return None, f'NO ENCONTRADO: "{nombre_excel}"'


# ─────────────────────────────────────────────────────────────────────────────
# 4. Comparar grupos
# ─────────────────────────────────────────────────────────────────────────────
def comparar_grupos(excel_data, bd_data, uid_by_username):
    ok, diffs, faltantes, sin_user = 0, [], [], []

    for (alias_low, pid_int), ex in excel_data.items():
        if alias_low not in uid_by_username:
            sin_user.append({'alias': ex['alias'], 'nombre': ex['nombre'], 'pid': ex['pid']})
            continue

        bd_key = (alias_low, pid_int)
        if bd_key not in bd_data:
            faltantes.append({'alias': ex['alias'], 'pid': ex['pid']})
            continue

        bd = bd_data[bd_key]
        row_diffs = []
        for campo, label in CAMPOS:
            ev = ex[campo]
            bv = bd[campo]
            if not equiv(ev, bv):
                row_diffs.append(f"{label}: excel={ev} ≠ bd={bv}")

        if row_diffs:
            diffs.append({'alias': ex['alias'], 'pid': ex['pid'], 'diff': row_diffs})
        else:
            ok += 1

    return ok, diffs, faltantes, sin_user


# ─────────────────────────────────────────────────────────────────────────────
# 5. Mostrar/importar globales
# ─────────────────────────────────────────────────────────────────────────────
def build_global_payload(gl_excel, equipos_dict, sin_resolver):
    """Convierte los valores raw del Excel al formato de apuesta_global."""
    payload = {}
    err = []

    # Campeón → pred_campeon_id + pred_finalista1_id
    campeon_raw = gl_excel.get('campeon')
    if campeon_raw:
        eq_id, why = resolver_equipo_id(campeon_raw, equipos_dict)
        if eq_id:
            payload['pred_campeon_id']    = eq_id
            payload['pred_finalista1_id'] = eq_id   # campeón es también finalista
        else:
            err.append(f"Campeón: {why}")
            sin_resolver.add(str(campeon_raw))

    # Otro finalista → pred_finalista2_id
    fin2_raw = gl_excel.get('finalista2')
    if fin2_raw:
        eq_id, why = resolver_equipo_id(fin2_raw, equipos_dict)
        if eq_id:
            payload['pred_finalista2_id'] = eq_id
        else:
            err.append(f"Finalista2: {why}")
            sin_resolver.add(str(fin2_raw))

    # Goleador (texto libre)
    if gl_excel.get('goleador'):
        payload['pred_goleador'] = str(gl_excel['goleador']).strip()

    # Peor equipo → pred_peor_equipo_id
    peor_raw = gl_excel.get('peor_equipo')
    if peor_raw:
        eq_id, why = resolver_equipo_id(peor_raw, equipos_dict)
        if eq_id:
            payload['pred_peor_equipo_id'] = eq_id
        else:
            err.append(f"Peor equipo: {why}")
            sin_resolver.add(str(peor_raw))

    # Etapa Paraguay (texto)
    if gl_excel.get('etapa_py') is not None:
        v = str(gl_excel['etapa_py']).strip()
        # Normalizar: "4tos" → "cuartos"
        _ETAPA_MAP = {
            '4tos': 'cuartos', 'cuartos': 'cuartos', 'qf': 'cuartos',
            'grupos': 'grupo', 'grupo': 'grupo', 'group stage': 'grupo',
            '32avos': 'ronda32', 'ronda32': 'ronda32',
            '16avos': 'ronda16', 'octavos': 'ronda16', 'ronda16': 'ronda16',
            'semis': 'semis', 'semifinales': 'semis', 'semifinal': 'semis',
            'final': 'final', 'finalista': 'final', '3er puesto': 'final',
            'tercer puesto': 'final',
        }
        payload['pred_etapa_paraguay'] = _ETAPA_MAP.get(v.lower(), v)

    # Goles Paraguay (int)
    if gl_excel.get('goles_py') is not None:
        try:
            payload['pred_goles_paraguay'] = int(gl_excel['goles_py'])
        except (ValueError, TypeError):
            pass

    # Mayor goleada ganador (int)
    if gl_excel.get('goleada_g') is not None:
        try:
            payload['pred_goleada_ganador'] = int(gl_excel['goleada_g'])
        except (ValueError, TypeError):
            pass

    # Mayor goleada perdedor (int)
    if gl_excel.get('goleada_p') is not None:
        try:
            payload['pred_goleada_perdedor'] = int(gl_excel['goleada_p'])
        except (ValueError, TypeError):
            pass

    return payload, err


def mostrar_globales(globales_excel, equipos_dict, globales_bd,
                     uid_by_username, equipos_por_id):
    """Muestra tabla de globales del Excel por apostador, comparando con BD."""
    SEP = '─' * 90
    print(f"\n{'='*90}")
    print("  APUESTAS GLOBALES — A-G por apostador (Excel → BD)")
    print(f"{'='*90}\n")

    sin_user = []
    sin_resolver_global = set()
    resumen_diff = []

    headers = ['ALIAS', 'CAMPEÓN', 'FIN2', 'GOLEADOR', 'PEOR EQ', 'ETAPA PY', 'GL.PY', 'GOLEADA G', 'GOLEADA P']
    widths  = [20, 14, 14, 22, 14, 10, 6, 9, 9]

    def cell(v, w): return str(v or '—')[:w].ljust(w)

    header_line = ' | '.join(cell(h, w) for h, w in zip(headers, widths))
    print(header_line)
    print(SEP)

    # Nombre del equipo a partir de id
    def eq_name(eq_id):
        if eq_id is None: return '—'
        return equipos_por_id.get(eq_id, f'id={eq_id}')

    for alias_low in sorted(globales_excel.keys()):
        gl = globales_excel[alias_low]
        alias  = gl['alias']
        nombre = gl['nombre']

        if alias_low not in uid_by_username:
            sin_user.append(alias)
            continue

        payload, err_eq = build_global_payload(gl, equipos_dict, sin_resolver_global)
        bd = globales_bd.get(alias_low, {})

        # Línea Excel
        row_xl = [
            alias,
            gl.get('campeon', '—'),
            gl.get('finalista2', '—'),
            gl.get('goleador', '—'),
            gl.get('peor_equipo', '—'),
            gl.get('etapa_py', '—'),
            gl.get('goles_py', '—'),
            gl.get('goleada_g', '—'),
            gl.get('goleada_p', '—'),
        ]
        print(' | '.join(cell(v, w) for v, w in zip(row_xl, widths)))

        # Diferencias vs BD
        diffs_lin = []
        if bd:
            check = [
                ('pred_campeon_id',       payload.get('pred_campeon_id'),       bd.get('pred_campeon_id'),
                 lambda v: eq_name(v)),
                ('pred_finalista2_id',    payload.get('pred_finalista2_id'),    bd.get('pred_finalista2_id'),
                 lambda v: eq_name(v)),
                ('pred_goleador',         payload.get('pred_goleador','').lower(),
                 (bd.get('pred_goleador') or '').lower(), lambda v: v),
                ('pred_peor_equipo_id',   payload.get('pred_peor_equipo_id'),   bd.get('pred_peor_equipo_id'),
                 lambda v: eq_name(v)),
                ('pred_etapa_paraguay',   payload.get('pred_etapa_paraguay'),   bd.get('pred_etapa_paraguay'),
                 lambda v: v),
                ('pred_goles_paraguay',   payload.get('pred_goles_paraguay'),   bd.get('pred_goles_paraguay'),
                 lambda v: v),
                ('pred_goleada_ganador',  payload.get('pred_goleada_ganador'),  bd.get('pred_goleada_ganador'),
                 lambda v: v),
                ('pred_goleada_perdedor', payload.get('pred_goleada_perdedor'), bd.get('pred_goleada_perdedor'),
                 lambda v: v),
            ]
            for campo, xv, bv, fmt in check:
                if xv != bv:
                    diffs_lin.append(f"  ⚠ {campo}: excel={fmt(xv)} ≠ bd={fmt(bv)}")
            if diffs_lin:
                resumen_diff.append((alias, diffs_lin))
        else:
            print(f"  ↳ Sin registro en apuesta_global BD")

        if err_eq:
            for e in err_eq:
                print(f"  ✗ {e}")

    print(SEP)

    if sin_user:
        print(f"\nAlias sin usuario BD ({len(sin_user)}): {', '.join(sin_user)}\n")

    if sin_resolver_global:
        print(f"\nEquipos NO resueltos (agregar al diccionario NOMBRE_ES_NORM):")
        for n in sorted(sin_resolver_global):
            print(f"  '{n}'")

    if resumen_diff:
        print(f"\nDIFERENCIAS EXCEL vs BD ({len(resumen_diff)} apostadores):")
        for alias, diffs in resumen_diff:
            print(f"\n  [{alias}]")
            for d in diffs:
                print(d)
    else:
        print("\n✅ Sin diferencias entre Excel y BD (para apostadores con registro en apuesta_global)")

    return sin_resolver_global


async def importar_globales(globales_excel, equipos_dict, uid_by_username):
    """Hace upsert de los globales del Excel en apuesta_global."""
    import asyncpg
    conn = await asyncpg.connect(DB_BECBUC)

    UPSERT = """
        INSERT INTO apuesta_global
            (apostador_id, torneo_id,
             pred_campeon_id, pred_finalista1_id, pred_finalista2_id,
             pred_goleador, pred_peor_equipo_id,
             pred_goleada_ganador, pred_goleada_perdedor,
             pred_etapa_paraguay, pred_goles_paraguay,
             creado_en, actualizado_en)
        VALUES
            ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11,
             NOW(), NOW())
        ON CONFLICT (apostador_id, torneo_id) DO UPDATE SET
             pred_campeon_id      = EXCLUDED.pred_campeon_id,
             pred_finalista1_id   = EXCLUDED.pred_finalista1_id,
             pred_finalista2_id   = EXCLUDED.pred_finalista2_id,
             pred_goleador        = EXCLUDED.pred_goleador,
             pred_peor_equipo_id  = EXCLUDED.pred_peor_equipo_id,
             pred_goleada_ganador = EXCLUDED.pred_goleada_ganador,
             pred_goleada_perdedor= EXCLUDED.pred_goleada_perdedor,
             pred_etapa_paraguay  = EXCLUDED.pred_etapa_paraguay,
             pred_goles_paraguay  = EXCLUDED.pred_goles_paraguay,
             actualizado_en       = NOW()
    """

    ok = 0
    skipped = 0
    sin_user = 0
    sin_resolver_set = set()

    for alias_low, gl in globales_excel.items():
        if alias_low not in uid_by_username:
            sin_user += 1
            continue

        uid = uid_by_username[alias_low][0]
        payload, err = build_global_payload(gl, equipos_dict, sin_resolver_set)

        if err:
            print(f"  [{gl['alias']}] Errores en equipos: {err}")
            skipped += 1
            continue

        try:
            await conn.execute(UPSERT,
                uid, TORNEO_ID,
                payload.get('pred_campeon_id'),
                payload.get('pred_finalista1_id'),
                payload.get('pred_finalista2_id'),
                payload.get('pred_goleador'),
                payload.get('pred_peor_equipo_id'),
                payload.get('pred_goleada_ganador'),
                payload.get('pred_goleada_perdedor'),
                payload.get('pred_etapa_paraguay'),
                payload.get('pred_goles_paraguay'),
            )
            ok += 1
        except Exception as e:
            print(f"  [{gl['alias']}] ERROR upsert: {e}")
            skipped += 1

    await conn.close()

    print(f"\n  Importados OK : {ok}")
    print(f"  Saltados      : {skipped}")
    print(f"  Sin usuario   : {sin_user}")
    if sin_resolver_set:
        print(f"\n  Equipos NO resueltos (revisar NOMBRE_ES_NORM):")
        for n in sorted(sin_resolver_set):
            print(f"    '{n}'")

    return ok


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
async def main():
    SEP = '=' * 60

    print(f"\n{SEP}")
    print("  BECBUC — Verificación / Importación de Pronósticos")
    print(f"{SEP}")
    print(f"  Modo grupos   : {'SÍ' if MODO_GRUPOS else 'NO'}")
    print(f"  Modo globales : {'SÍ' if MODO_GLOBALES else 'NO'}")
    print(f"  Importar BD   : {'SÍ' if IMPORTAR_GL else 'NO'}")
    print(f"{SEP}\n")

    print(f"Leyendo Excel:\n  {EXCEL_PATH}\n")
    try:
        excel_grupos   = leer_excel_grupos(EXCEL_PATH)   if MODO_GRUPOS   else ({}, {})
        excel_globales = leer_excel_globales(EXCEL_PATH) if MODO_GLOBALES else {}
    except FileNotFoundError:
        print("ERROR: archivo Excel no encontrado."); return

    if MODO_GRUPOS:
        print(f"  Grupos  : {len(excel_grupos[0]):,} registros | {len(excel_grupos[1])} apostadores")
    if MODO_GLOBALES:
        print(f"  Globales: {len(excel_globales)} apostadores\n")

    # ── Globales ──────────────────────────────────────────────────────────────
    if MODO_GLOBALES:
        print("Conectando BD para globales y equipos...")
        try:
            equipos, equipos_por_id, globales_bd, uid_by_username = \
                await leer_bd_globales_y_equipos()
        except Exception as e:
            import traceback
            print(f"ERROR BD: {e}"); traceback.print_exc(); return

        print(f"  Equipos en BD       : {len(equipos)}")
        print(f"  Apostadores con BD  : {len(globales_bd)}")
        print(f"  Usuarios en app_db  : {len(uid_by_username)}\n")

        # Mostrar tabla globales
        mostrar_globales(excel_globales, equipos, globales_bd,
                         uid_by_username, equipos_por_id)

        if IMPORTAR_GL:
            print(f"\n{SEP}")
            print("  IMPORTANDO GLOBALES A BD...")
            print(f"{SEP}\n")
            n = await importar_globales(excel_globales, equipos, uid_by_username)
            print(f"\n{'='*60}")
            print(f"  ✅ Importación completada: {n} apostadores actualizados")
            print(f"{'='*60}\n")
            return

    # ── Grupos ────────────────────────────────────────────────────────────────
    if MODO_GRUPOS:
        excel_data, aliases = excel_grupos
        print("Conectando BD para grupos...")
        try:
            bd_data, uid_by_username, unknown_uids = await leer_bd_grupos()
        except Exception as e:
            import traceback
            print(f"ERROR BD: {e}"); traceback.print_exc(); return

        print(f"  {len(bd_data):,} registros en BD\n")
        if unknown_uids:
            print(f"  AVISO: {len(unknown_uids)} apostador_id sin usuario: {sorted(unknown_uids)}\n")

        bd_sample = list(bd_data.keys())[:3]
        xl_sample = list(excel_data.keys())[:3]
        print(f"  Muestra claves BD   : {bd_sample}")
        print(f"  Muestra claves Excel: {xl_sample}\n")

        ok, diffs, faltantes, sin_user = comparar_grupos(excel_data, bd_data, uid_by_username)

        print(f"{SEP}")
        print("  RESULTADO — GRUPOS")
        print(f"{SEP}")
        print(f"  Coinciden exactamente  : {ok:5,}  de  {len(excel_data)-len(sin_user):,}")
        print(f"  Con diferencias        : {len(diffs):5,}")
        print(f"  Apuestas faltantes BD  : {len(faltantes):5,}")
        print(f"  Alias sin usuario BD   : {len(set(s['alias'] for s in sin_user)):5,}")
        print(f"{SEP}\n")

        if sin_user:
            print("ALIAS SIN USUARIO EN BD:")
            seen = set()
            for s in sin_user:
                if s['alias'] not in seen:
                    print(f"  '{s['alias']}'  ({s['nombre']})")
                    seen.add(s['alias'])
            print()

        if diffs:
            print(f"DIFERENCIAS ({len(diffs)} registros):")
            by_user = defaultdict(list)
            for d in diffs: by_user[d['alias']].append(d)
            for alias, items in sorted(by_user.items()):
                print(f"\n  [{alias}]  {len(items)} partido/s:")
                for d in items[:6]:
                    print(f"    {d['pid']}: {' | '.join(d['diff'])}")
                if len(items) > 6: print(f"    ... y {len(items)-6} mas")

        if faltantes:
            print(f"\nAPUESTAS FALTANTES EN BD ({len(faltantes)}):")
            by_user = defaultdict(list)
            for f in faltantes: by_user[f['alias']].append(f['pid'])
            for alias, pids in sorted(by_user.items()):
                print(f"  {alias}: {len(pids)} partidos -> {sorted(pids)[:5]}")

        REPORT_OUT.parent.mkdir(exist_ok=True)
        REPORT_OUT.write_text(json.dumps({
            'resumen': {'ok': ok, 'diferencias': len(diffs),
                        'faltantes_bd': len(faltantes), 'sin_usuario': len(sin_user)},
            'sin_usuario': sin_user, 'diferencias': diffs, 'faltantes': faltantes,
        }, ensure_ascii=False, indent=2, default=str))
        print(f"\nReporte: {REPORT_OUT}")


if __name__ == '__main__':
    asyncio.run(main())
