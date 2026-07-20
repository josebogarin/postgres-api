"""
debug_import_excel.py
Importa las columnas clave del Excel consolidado a una tabla auxiliar
en becbuc y verifica la correspondencia con las apuestas reales.

Uso:
    cd "C:\proyecto FAST API"
    python debug_import_excel.py "ruta\al\archivo.xlsx"
"""

import sys
import re
import openpyxl
import psycopg2

# ── Config BD ────────────────────────────────────────────────────────────────
BECBUC_DSN = "host=localhost port=5432 dbname=becbuc user=app_user password=superpassword"
APP_DSN    = "host=localhost port=5432 dbname=app_db  user=app_user password=superpassword"
TORNEO_ID  = 2

# ── Normalización alias (igual que el backend) ────────────────────────────────
def norm(s):
    return (s or "").replace("\xa0", "").strip().lower()

# ── Leer Excel ────────────────────────────────────────────────────────────────
def sheetjs_headers(raw):
    """Simula deduplicación de SheetJS para columnas duplicadas."""
    seen = {}
    result = []
    for h in raw:
        h = str(h) if h is not None else "__EMPTY"
        if h in seen:
            seen[h] += 1
            result.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            result.append(h)
    return result

def leer_excel(path):
    print(f"Leyendo {path}...")
    wb = openpyxl.load_workbook(path, data_only=True)
    # Buscar hoja TBL MASTER explícitamente
    hoja_nombre = None
    for n in wb.sheetnames:
        if re.search(r'tbl[\s_-]*master|master', n, re.IGNORECASE):
            hoja_nombre = n
            break
    if not hoja_nombre:
        # fallback: hoja con más filas
        hoja_nombre = max(wb.sheetnames, key=lambda n: wb[n].max_row)
    print(f"Hoja seleccionada: {hoja_nombre} (de {wb.sheetnames})")
    ws = wb[hoja_nombre]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    headers = sheetjs_headers(rows[0])
    print(f"Columnas ({len(headers)}): {headers[:20]}...")

    # Mapear campos clave
    idx = {h: i for i, h in enumerate(headers)}
    ALIAS_COL    = "ALIAS"
    PID_COL      = "ID PARTIDO"
    GL_COL       = "GOLES"      # local
    GV_COL       = "GOLES_1"    # visitante (SheetJS deduplica)
    FASE_COL     = "FASE"

    missing = [c for c in [ALIAS_COL, PID_COL, GL_COL, GV_COL, FASE_COL] if c not in idx]
    if missing:
        print(f"ADVERTENCIA: columnas no encontradas: {missing}")
        print(f"Columnas disponibles: {headers}")

    registros = []
    for r in rows[1:]:
        fase = str(r[idx.get(FASE_COL, -1)] or "")
        if "GRUPO" not in fase.upper() and "10-" not in fase.upper():
            continue
        alias  = norm(str(r[idx.get(ALIAS_COL, -1)] or ""))
        pid    = str(r[idx.get(PID_COL, -1)] or "").strip()
        gl     = r[idx.get(GL_COL, -1)]
        gv     = r[idx.get(GV_COL, -1)]
        if not alias or not pid:
            continue
        registros.append((alias, pid, gl, gv))

    print(f"Registros de grupos leidos del Excel: {len(registros)}")
    return registros

# ── Crear tabla auxiliar y cargar datos ──────────────────────────────────────
def cargar_en_bd(registros):
    con = psycopg2.connect(BECBUC_DSN)
    cur = con.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS debug_excel_import (
            alias        TEXT,
            partido_num  TEXT,
            gl_excel     INT,
            gv_excel     INT
        )
    """)
    cur.execute("TRUNCATE debug_excel_import")
    cur.executemany(
        "INSERT INTO debug_excel_import VALUES (%s,%s,%s,%s)",
        registros
    )
    con.commit()
    print(f"Tabla debug_excel_import cargada con {len(registros)} filas.")
    cur.close()
    con.close()

# ── Comparar Excel vs apuesta real ───────────────────────────────────────────
def comparar(torneo_id):
    # Obtener usernames desde app_db
    app_con = psycopg2.connect(APP_DSN)
    app_cur = app_con.cursor()
    app_cur.execute("SELECT id, username FROM users WHERE is_active = TRUE")
    user_map = {norm(username): uid for uid, username in app_cur.fetchall()}
    app_cur.close()
    app_con.close()
    print(f"Usuarios en app_db: {len(user_map)}")

    con = psycopg2.connect(BECBUC_DSN)
    cur = con.cursor()

    # partido_map: num_seq -> partido_id (igual que el backend)
    cur.execute("""
        SELECT p.id, ROW_NUMBER() OVER (ORDER BY f.orden, p.id)::int AS num_seq
        FROM partido p JOIN fase f ON f.id = p.fase_id
        WHERE p.torneo_id = %s AND f.tipo = 'grupo'
        ORDER BY f.orden, p.id
    """, (torneo_id,))
    partido_map = {num_seq: pid for pid, num_seq in cur.fetchall()}
    print(f"Partidos de grupos en BD: {len(partido_map)}")

    # Leer tabla auxiliar
    cur.execute("SELECT alias, partido_num, gl_excel, gv_excel FROM debug_excel_import ORDER BY alias, partido_num")
    filas = cur.fetchall()

    print(f"\n{'ALIAS':<20} {'PID':<6} {'GL_Excel':>8} {'GV_Excel':>8} {'GL_BD':>6} {'GV_BD':>6} {'MATCH':>6} {'PROBLEMA'}")
    print("-" * 90)

    no_match    = []
    user_miss   = []
    partido_miss= []
    ok_count    = 0

    for alias, pid_str, gl_ex, gv_ex in filas:
        # Resolver partido
        raw = re.sub(r'[Pp]0*', '', pid_str).strip() or '0'
        try:
            num_int = int(raw)
        except ValueError:
            num_int = 0
        partido_id = partido_map.get(num_int)

        # Resolver usuario
        user_id = user_map.get(alias)

        if not user_id:
            user_miss.append(alias)
            if alias == 'cherem':
                print(f"{'CHEREM':<20} {'⚠ usuario NO ENCONTRADO en app_db':}")
            continue
        if not partido_id:
            partido_miss.append((alias, pid_str))
            continue

        # Leer apuesta real
        cur.execute("""
            SELECT pred_local, pred_visitante
            FROM apuesta
            WHERE apostador_id = %s AND partido_id = %s
        """, (user_id, partido_id))
        row = cur.fetchone()

        if row is None:
            gl_bd, gv_bd = None, None
            match = "SIN_APUESTA"
        else:
            gl_bd, gv_bd = row
            match = "OK" if (gl_bd == gl_ex and gv_bd == gv_ex) else "DIFF"

        # Solo mostrar cherem o diferencias
        if alias == 'cherem' or match in ('DIFF', 'SIN_APUESTA'):
            flag = "✅" if match == "OK" else ("❌" if match == "DIFF" else "⚠")
            print(f"{alias:<20} {pid_str:<6} {str(gl_ex):>8} {str(gv_ex):>8} {str(gl_bd):>6} {str(gv_bd):>6} {flag:>6} {match}")
        if match == "OK":
            ok_count += 1
        elif match == "DIFF":
            no_match.append((alias, pid_str, gl_ex, gv_ex, gl_bd, gv_bd))

    print(f"\n{'='*90}")
    print(f"OK: {ok_count}  |  DIFF: {len(no_match)}  |  Sin apuesta: suma de lo anterior")
    if user_miss:
        print(f"\nUsuarios NO encontrados en app_db ({len(set(user_miss))}): {sorted(set(user_miss))}")
    if partido_miss:
        print(f"\nPartidos NO encontrados en BD ({len(partido_miss)}): {partido_miss[:10]}")
    if no_match:
        print(f"\nPrimeras diferencias:")
        for x in no_match[:20]:
            print(f"  {x[0]} P{x[1]}: Excel={x[2]}-{x[3]} BD={x[4]}-{x[5]}")

    cur.close()
    con.close()

# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python debug_import_excel.py <ruta_excel>")
        print('Ejemplo: python debug_import_excel.py "20260611_2000- TBL CONSOLIDADA.xlsx"')
        sys.exit(1)

    path = sys.argv[1]
    registros = leer_excel(path)
    cargar_en_bd(registros)
    comparar(TORNEO_ID)
