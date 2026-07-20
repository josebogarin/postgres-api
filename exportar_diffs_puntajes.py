# -*- coding: utf-8 -*-
"""
exportar_diffs_puntajes.py
Compara los PUNTAJES por apostador x partido, item por item (Excel '50- TBL MASTER'
vs BD puntaje_detalle) y exporta un Excel con cada DIFERENCIA, mostrando para el item:
  Resultado real (oficial) | Apuesta Excel | Apuesta BD | Pts.Excel | Pts.BD | Causa
Hojas: "Leyenda" (que significa cada letra) + "Diferencias" (detalle) + "Resumen".
Salida: becbuc_diffs_puntajes_<YYYYMMDD_HHMM>.xlsx
Solo lectura. Requiere psycopg2 + openpyxl.
"""
import sys, os, datetime
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    os.system(f'"{sys.executable}" -m pip install openpyxl --quiet')
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
try:
    import psycopg2, psycopg2.extras
except ImportError:
    os.system(f'"{sys.executable}" -m pip install psycopg2-binary --quiet'); import psycopg2, psycopg2.extras

PHASES = {
    'grupos':  ('10- GRUPOS',        1,  72, False),
    'r32':     ('20- DIECISEISAVOS', 73,  88, True),
    'octavos': ('30- OCTAVOS',       89,  96, True),
    'cuartos': ('40- CUARTOS',       97, 100, True),
    'semis':   ('50- SEMIFINAL',    101, 102, True),
}
ORDER = ['grupos', 'r32', 'octavos', 'cuartos', 'semis']
FASE_TXT_SET = {PHASES[f][0] for f in ORDER}
fase_de_nf = {}
for f in ORDER:
    _, a, b, _ = PHASES[f]
    for n in range(a, b + 1): fase_de_nf[n] = f

BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = None
_cands = [os.path.join(BASE, f) for f in os.listdir(BASE)
          if f.lower().endswith('.xlsx')
          and any(k in f.upper() for k in ('CORRECCIONES', 'SEMIFINAL', 'TBL PARA CARGAR'))]
def _rank(p):
    u = os.path.basename(p).upper()
    return (0 if 'CORRECCIONES' in u else (1 if 'SEMIFINAL' in u else 2), -os.path.getmtime(p))
if _cands: EXCEL_FILE = sorted(_cands, key=_rank)[0]
if not EXCEL_FILE: sys.exit(f"ERROR: no se encontro Excel en {BASE}")
print(f"Excel: {os.path.basename(EXCEL_FILE)}")

CONN_BEC = "host=localhost port=5432 dbname=becbuc user=app_user password=superpassword"
CONN_APP = "host=localhost port=5432 dbname=app_db  user=app_user password=superpassword"
TID = 2
SHEET = '50- TBL MASTER'

conn = psycopg2.connect(CONN_BEC); capp = psycopg2.connect(CONN_APP)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
cua = capp.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

cua.execute("SELECT id, username FROM users WHERE is_active=TRUE")
uname = {u['username'].lower(): u['id'] for u in cua.fetchall()}
def clean_alias(s):
    if not s: return ''
    return str(s).replace('\xa0','').strip().upper().lstrip('@').replace('  ',' ')
def find_uid(al):
    a = clean_alias(al)
    for k, v in uname.items():
        if k.upper().lstrip('@') == a: return v
    for k, v in uname.items():
        if a and (a in k.upper() or k.upper() in a): return v
    return None

cur.execute("SELECT id, nombre, nombre_es FROM equipo")
ename = {}
for eq in cur.fetchall(): ename[eq['id']] = eq['nombre'] or eq['nombre_es']
def enm(x): return ename.get(x, x) if x is not None else 'None'

cur.execute("""SELECT p.numero_fifa FROM partido p JOIN fase f ON f.id=p.fase_id
               LEFT JOIN equipo el ON el.id=p.equipo_local_id
               LEFT JOIN equipo ev ON ev.id=p.equipo_visitante_id
               WHERE f.torneo_id=%s AND ('Paraguay' IN (el.nombre, ev.nombre)
                  OR 'PARAGUAY' IN (upper(el.nombre_es), upper(ev.nombre_es)))""", (TID,))
paraguay_nfs = {r['numero_fifa'] for r in cur.fetchall()}
print(f"Partidos de Paraguay: {sorted(paraguay_nfs)}")

cur.execute("""SELECT pd.apostador_id, p.numero_fifa,
       COALESCE(pd.pts_resultado,0) AS h, COALESCE(pd.pts_marcador,0) AS i,
       COALESCE(pd.pts_amarillas,0) AS j, COALESCE(pd.pts_rojas,0) AS k,
       COALESCE(pd.pts_var,0) AS l, COALESCE(pd.pts_penales_partido,0) AS m,
       COALESCE(pd.pts_minuto,0) AS n, COALESCE(pd.pts_penales_tanda,0) AS o,
       COALESCE(pd.pts_equipo,0) AS p
       FROM puntaje_detalle pd JOIN partido p ON p.id=pd.partido_id
       JOIN fase f ON f.id=p.fase_id WHERE f.torneo_id=%s""", (TID,))
bd_pts = {(r['apostador_id'], r['numero_fifa']): r for r in cur.fetchall()}

# predicciones BD completas
cur.execute("""SELECT a.apostador_id, p.numero_fifa,
       a.pred_local, a.pred_visitante, a.pred_amarillas, a.pred_rojas, a.pred_var,
       a.pred_penales_partido, a.pred_minuto_gol, a.pred_penales_local_tanda,
       a.pred_penales_visitante_tanda, a.pred_equipo_clasifica
       FROM apuesta a JOIN partido p ON p.id=a.partido_id JOIN fase f ON f.id=p.fase_id
       WHERE f.torneo_id=%s""", (TID,))
bd_pred = {(r['apostador_id'], r['numero_fifa']): r for r in cur.fetchall()}

# resultado oficial completo
cur.execute("""SELECT p.numero_fifa, p.goles_local, p.goles_visitante, p.amarillas, p.rojas,
       p.decisiones_var, p.penales_partido, p.minuto_primer_gol,
       p.penales_local, p.penales_visitante, p.equipo_clasificado_id
       FROM partido p JOIN fase f ON f.id=p.fase_id WHERE f.torneo_id=%s""", (TID,))
off = {r['numero_fifa']: r for r in cur.fetchall()}

def sc(v):
    try:
        s = str(v).strip()
        return 0 if s in ('','-','None') else int(float(s))
    except: return 0
def sv(v):
    if v is None: return 'None'
    s = str(v).strip()
    return s if s not in ('','-') else 'None'
def par(a, b):  # par de valores "a-b"
    return f"{sv(a)}-{sv(b)}"

C = dict(pid=2, fase=7, alias=10, pl=13, pv=15,
         predJ=25, predK=26, predL=27, predM=28, predN=29, predOe1=30, predOe2=31, predQ=32,
         ptsH=40, ptsI=41, ptsJ=42, ptsK=43, ptsL=44, ptsM=45, ptsN=46, ptsOe1=47, ptsOe2=48, ptsP=51)
ITEM_DESC = {'H':'Resultado (gana/empata/pierde)','I':'Marcador exacto','J':'Amarillas',
             'K':'Rojas','L':'VAR','M':'Penales en el juego','N':'Minuto 1er gol',
             'O':'Tanda de penales','P':'Equipo que clasifica'}
ALL_ITEMS = ['H','I','J','K','L','M','N','O','P']

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True, read_only=True)
ws = wb[SHEET]
hdr = None
for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
    hdr = row; break
if hdr:
    for idx, h in enumerate(hdr, start=1):
        if h and 'CLASIFICAD' in str(h).upper() and idx >= 45:
            C['ptsP'] = idx; break
print(f"Col puntos P detectada: {C['ptsP']}")

def real_val(item, o):
    if o is None: return ''
    if item in ('H','I'): return par(o['goles_local'], o['goles_visitante'])
    if item == 'J': return sv(o['amarillas'])
    if item == 'K': return sv(o['rojas'])
    if item == 'L': return sv(o['decisiones_var'])
    if item == 'M': return sv(o['penales_partido'])
    if item == 'N': return sv(o['minuto_primer_gol'])
    if item == 'O': return par(o['penales_local'], o['penales_visitante'])
    if item == 'P': return enm(o['equipo_clasificado_id'])
    return ''
def pred_excel_val(item, row):
    if item in ('H','I'): return par(row[C['pl']-1], row[C['pv']-1])
    if item == 'J': return sv(row[C['predJ']-1])
    if item == 'K': return sv(row[C['predK']-1])
    if item == 'L': return sv(row[C['predL']-1])
    if item == 'M': return sv(row[C['predM']-1])
    if item == 'N': return sv(row[C['predN']-1])
    if item == 'O': return par(row[C['predOe1']-1], row[C['predOe2']-1])
    if item == 'P': return sv(row[C['predQ']-1])
    return ''
def pred_bd_val(item, pr):
    if pr is None: return ''
    if item in ('H','I'): return par(pr['pred_local'], pr['pred_visitante'])
    if item == 'J': return sv(pr['pred_amarillas'])
    if item == 'K': return sv(pr['pred_rojas'])
    if item == 'L': return sv(pr['pred_var'])
    if item == 'M': return sv(pr['pred_penales_partido'])
    if item == 'N': return sv(pr['pred_minuto_gol'])
    if item == 'O': return par(pr['pred_penales_local_tanda'], pr['pred_penales_visitante_tanda'])
    if item == 'P': return enm(pr['pred_equipo_clasifica'])
    return ''

def causa(item, nf, ptsEx, ptsBD, predBD_clasif):
    if item == 'N':
        return "N desempate minuto (decision org): BD da 1pt a todos los mas cercanos"
    if item == 'O':
        return "O error Excel: tanda otorgada en semis sin definicion por penales"
    if item in ('K','M'):
        return "regla NULL->0 de la BD: pred en blanco tratada como 0 (lenient)" if ptsBD > ptsEx else "revisar"
    if item == 'I':
        return "error Excel (marcador/placeholder)"
    if item == 'P':
        if nf in paraguay_nfs and ptsEx > 0 and ptsBD == 2*ptsEx:
            return "Paraguay x2 en item P (BD correcta, reglamento)"
        if predBD_clasif in (None,'None','') and ptsEx > ptsBD:
            return "hueco BD: pred_equipo_clasifica NULL (BD da 0)"
        if ptsBD > ptsEx:
            return "error Excel: no otorgo P a prediccion correcta"
        if ptsBD < ptsEx:
            return "revisar: Excel otorgo P de mas / pred BD distinta"
        return "revisar"
    return "revisar"

filas = []
for row in ws.iter_rows(min_row=2, values_only=True):
    pid = row[C['pid']-1]
    if not (isinstance(pid, str) and pid.startswith('P')): continue
    if str(row[C['fase']-1]) not in FASE_TXT_SET: continue
    try: nf = int(pid[1:])
    except: continue
    fase = fase_de_nf.get(nf); es_ko = PHASES[fase][3] if fase else False
    uid = find_uid(row[C['alias']-1])
    if uid is None: continue
    alias = clean_alias(row[C['alias']-1])
    b = bd_pts.get((uid, nf))
    if not b: continue
    o = off.get(nf); pr = bd_pred.get((uid, nf))
    PEX = {'H':sc(row[C['ptsH']-1]),'I':sc(row[C['ptsI']-1]),'J':sc(row[C['ptsJ']-1]),
           'K':sc(row[C['ptsK']-1]),'L':sc(row[C['ptsL']-1]),'M':sc(row[C['ptsM']-1]),
           'N':sc(row[C['ptsN']-1]),'O':sc(row[C['ptsOe1']-1])+sc(row[C['ptsOe2']-1]),
           'P':sc(row[C['ptsP']-1])}
    PBD = {'H':b['h'],'I':b['i'],'J':b['j'],'K':b['k'],'L':b['l'],'M':b['m'],
           'N':b['n'],'O':b['o'],'P':b['p']}
    predBD_clasif = enm(pr['pred_equipo_clasifica']) if pr else 'None'
    for it in ALL_ITEMS:
        if it == 'P' and not es_ko: continue
        ve = PEX[it]; vb = PBD[it]
        if ve == vb: continue
        rv = real_val(it, o); pex = pred_excel_val(it, row); pbd = pred_bd_val(it, pr)
        cau = causa(it, nf, ve, vb, predBD_clasif)
        filas.append((fase, it, ITEM_DESC[it], pid, alias, rv, pex, pbd, ve, vb, cau))

print(f"Total diferencias: {len(filas)}")

wbo = openpyxl.Workbook()
hf = Font(bold=True, color="FFFFFF"); hb = PatternFill("solid", fgColor="305496")
tit = Font(bold=True, size=13)

# ---- Hoja Leyenda ----
wsL = wbo.active; wsL.title = "Leyenda"
wsL.append(["BECBUC - Comparacion de puntajes Excel (TBL MASTER) vs BD"]); wsL['A1'].font = tit
wsL.append([])
wsL.append(["ITEMS DE PUNTAJE POR PARTIDO"]); wsL['A3'].font = Font(bold=True)
wsL.append(["Letra","Concepto","Puntos (por fase)","Notas"])
for c in range(1,5): wsL.cell(4,c).font = hf; wsL.cell(4,c).fill = hb
leyenda = [
    ("H","Resultado (gana / empata / pierde)","4 / 6 / 8 / 10 / 12 / 14 / 20","Acertar el signo del partido."),
    ("I","Marcador exacto","8 / 12 / 16 / 20 / 24 / 28 / 40","Acertar los goles exactos."),
    ("J","Amarillas (total del partido)","1","1 pt si coincide el numero."),
    ("K","Rojas","1","1 pt si coincide. Pred en blanco en BD = 0 (lenient)."),
    ("L","Decisiones VAR","1","1 pt si coincide."),
    ("M","Penales cobrados en el juego","1","1 pt si coincide. Pred en blanco en BD = 0 (lenient)."),
    ("N","Minuto del 1er gol","1","1 pt al/los mas cercanos. BD premia a TODOS los empatados."),
    ("O","Tanda de penales (por equipo)","2 por equipo","Solo KO definidos por penales."),
    ("P","Equipo que clasifica","2 / 4 / 6 / 8 / 10 / 12","Solo KO. Paraguay NO duplica este item."),
]
for r in leyenda: wsL.append(list(r))
wsL.append([])
wsL.append(["Nota: fases = Grupos / 16avos / Octavos / Cuartos / Semis / 3er puesto / Final.",])
wsL.append(["Paraguay: DOBLE puntaje en todos los items del partido EXCEPTO P (equipo que clasifica).",])
wsL.append([])
wsL.append(["CAUSAS DE DIFERENCIA (columna Causa en la hoja Diferencias)"]); wsL.cell(wsL.max_row,1).font = Font(bold=True)
causas_txt = [
    ("regla NULL->0 de la BD","La prediccion en blanco se trata como 0; si el oficial es 0, la BD suma 1 pt y el Excel 0. BD lenient por diseno."),
    ("N desempate minuto","La BD da 1 pt a TODOS los apostadores empatados en la distancia minima al minuto real (decision organizacion)."),
    ("O error Excel","El Excel otorgo tanda en semis que se definieron en tiempo reglamentario (sin penales). BD = 0, correcta."),
    ("error Excel: no otorgo P","El apostador acerto el equipo que clasifica pero el Excel dio 0. BD correcta."),
    ("Paraguay x2 en item P","Diferencia historica ya corregida: el item P no se duplica para Paraguay."),
]
wsL.append(["Causa","Explicacion"])
for c in range(1,3): wsL.cell(wsL.max_row,c).font = hf; wsL.cell(wsL.max_row,c).fill = hb
for r in causas_txt: wsL.append(list(r))
wsL.column_dimensions['A'].width = 26
wsL.column_dimensions['B'].width = 70
wsL.column_dimensions['C'].width = 24
wsL.column_dimensions['D'].width = 52

# ---- Hoja Diferencias ----
ws1 = wbo.create_sheet("Diferencias")
hd = ["Fase","Item","Concepto","Partido","Apostador","Resultado real","Apuesta Excel","Apuesta BD","Pts.Excel","Pts.BD","Causa"]
ws1.append(hd)
for c in range(1, len(hd)+1):
    ws1.cell(1, c).font = hf; ws1.cell(1, c).fill = hb
IT_ORDER = {it:i for i,it in enumerate(ALL_ITEMS)}
for f in sorted(filas, key=lambda x:(IT_ORDER[x[1]], x[3], x[4])):
    ws1.append(list(f))
widths = [10,6,26,9,20,16,16,16,10,8,54]
for i,w in enumerate(widths, start=1):
    ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws1.freeze_panes = "A2"

# ---- Hoja Resumen ----
ws2 = wbo.create_sheet("Resumen")
from collections import Counter
cnt = Counter((f[1], f[10]) for f in filas)
ws2.append(["Item","Causa","Cantidad"])
for c in range(1,4): ws2.cell(1,c).font = hf; ws2.cell(1,c).fill = hb
for (it, cau), n in sorted(cnt.items(), key=lambda x:(IT_ORDER[x[0][0]], -x[1])):
    ws2.append([it, cau, n])
ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 62
ws2.column_dimensions['C'].width = 10
ws2.append([]); ws2.append(["TOTAL","", len(filas)])

ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
out = os.path.join(BASE, f"becbuc_diffs_puntajes_{ts}.xlsx")
wbo.save(out)
print(f"\nGuardado: {os.path.basename(out)}")
print("\nResumen item x causa:")
for (it, cau), n in sorted(cnt.items(), key=lambda x:(IT_ORDER[x[0][0]], -x[1])):
    print(f"  {it}: {n:>4}  {cau}")
conn.close(); capp.close()
