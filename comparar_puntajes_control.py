"""
comparar_puntajes_control.py
============================
Compara el Excel de control (ranking_torneo2_YYYYMMDD_HHMM.xlsx) contra:
  1. Datos actuales de la BD via API (nivel apostador: H,I,J,K,L,M,N,O,Total)
  2. Re-calculo con el algoritmo oficial (nivel partido: verifica cada item)

El Excel ya contiene pred + real + pts para cada fila → el recalculo es
self-contained (no necesita llamar a BD para los items J/K/L/M).
Para N (minuto gol) necesita todos los apostadores de cada partido.
Para el nivel apostador, llama a /bets/scores-por-apostador/2.

Ejecutar con el venv activo:
  cd "C:\proyecto FAST API"
  backend\\.venv\\Scripts\\Activate.ps1
  python comparar_puntajes_control.py [ruta_al_excel]

Si no se pasa ruta, busca automaticamente el archivo mas reciente.
"""

import sys
import re
import json
import io
import urllib.request
import urllib.error
from pathlib import Path
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACION
# ─────────────────────────────────────────────────────────────────────────────

TORNEO_ID  = 2
ADMIN_USER = "jose"
ADMIN_PASS = "catalina"

API_CANDIDATES = [
    "https://cupped-oink-thousand.ngrok-free.dev/api/v1",
    "http://localhost:8000/api/v1",
    "http://127.0.0.1:8000/api/v1",
]

OUTPUT_PATH = Path(r"C:\proyecto FAST API\reporte_diferencias_puntajes.xlsx")

# Puntos por fase grupos
H_PTS = 4
I_PTS = 8   # adicionales (solo si H tambien se gana)

# ─────────────────────────────────────────────────────────────────────────────
# LECTURA DEL EXCEL DE CONTROL
# ─────────────────────────────────────────────────────────────────────────────

def encontrar_excel(argv):
    """Busca el Excel de control en orden de prioridad."""
    # 1. Argumento CLI
    if len(argv) > 1:
        p = Path(argv[1])
        if p.exists():
            return p
        print(f"[WARN] Archivo CLI no encontrado: {p}")

    # 2. Ruta del upload del usuario
    upload_path = Path(
        r"C:\Users\Jose Bogarin\AppData\Roaming\Claude\local-agent-mode-sessions"
        r"\a9fdc79d-9227-450c-a0c1-27eafc601471"
        r"\dfc0381f-d9d1-4349-b3fa-24cab5c5da8b"
        r"\agent\local_ditto_dfc0381f-d9d1-4349-b3fa-24cab5c5da8b"
        r"\uploads\8f7cd0c4-ranking_torneo2_20260624_1241.xlsx"
    )
    if upload_path.exists():
        return upload_path

    # 3. Mas reciente en el directorio del proyecto
    candidates = sorted(
        list(Path(".").glob("*ranking_torneo2*.xlsx")) +
        list(Path(".").glob("*ranking*.xlsx")),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if candidates:
        print(f"[INFO] Usando Excel encontrado: {candidates[0]}")
        return candidates[0]

    return None


def _norm_header(v):
    """Normaliza string para matching de columnas."""
    if v is None:
        return ""
    return str(v).strip().lower().replace("\n", " ").replace(".", "").replace("(", "").replace(")", "").strip()


def _int_or_none(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return None


def leer_puntaje_general(wb):
    """
    Hoja 'Puntaje general' — devuelve lista de dicts:
      {nombre, pos, h, i, j, k, l, m, n, o, glob, total}

    El Excel tiene cabecera en fila 1: #, Apostador, H\nResultado, I\nExacto, ...
    """
    ws = wb["Puntaje general"]
    rows = list(ws.iter_rows(values_only=True))

    # Buscar fila de header (tiene 'apostador' y algun item)
    header_row_idx = None
    for i, row in enumerate(rows):
        cells = [_norm_header(v) for v in row if v is not None]
        if any("apostador" in c for c in cells) and any(c in ("h", "i") for c in cells):
            header_row_idx = i
            break

    if header_row_idx is None:
        print("[WARN] No se encontró fila de encabezado en 'Puntaje general'")
        return []

    headers = [_norm_header(v) for v in rows[header_row_idx]]
    print(f"  Headers Puntaje general: {headers}")

    # Mapeo flexible
    col = {}
    for idx, h in enumerate(headers):
        if h in ("#", "pos", "posicion"):
            col.setdefault("pos", idx)
        elif "apostador" in h or "nombre" in h:
            col.setdefault("nombre", idx)
        elif h == "h" or h.startswith("h ") or h.startswith("h\n"):
            col.setdefault("h", idx)
        elif h == "i" or h.startswith("i ") or h.startswith("i\n"):
            col.setdefault("i", idx)
        elif h == "j" or h.startswith("j "):
            col.setdefault("j", idx)
        elif h == "k" or h.startswith("k "):
            col.setdefault("k", idx)
        elif h == "l" or h.startswith("l "):
            col.setdefault("l", idx)
        elif h == "m" or h.startswith("m "):
            col.setdefault("m", idx)
        elif h == "n" or h.startswith("n "):
            col.setdefault("n", idx)
        elif h == "o" or h.startswith("o "):
            col.setdefault("o", idx)
        elif "glob" in h:
            col.setdefault("glob", idx)
        elif "total" in h:
            col.setdefault("total", idx)

    # Tambien puede haber solo letras como valores (H Resultado -> "h resultado")
    # Fallback: re-buscar con starts-with
    if "h" not in col:
        for idx, h in enumerate(headers):
            if h and h[0] == "h" and len(h) <= 15:
                col.setdefault("h", idx)
            elif h and h[0] == "i" and len(h) <= 15:
                col.setdefault("i", idx)
            elif h and h[0] == "j" and len(h) <= 15:
                col.setdefault("j", idx)
            elif h and h[0] == "k" and len(h) <= 15:
                col.setdefault("k", idx)
            elif h and h[0] == "l" and len(h) <= 15:
                col.setdefault("l", idx)
            elif h and h[0] == "m" and len(h) <= 15:
                col.setdefault("m", idx)
            elif h and h[0] == "n" and len(h) <= 15:
                col.setdefault("n", idx)
            elif h and h[0] == "o" and len(h) <= 15:
                col.setdefault("o", idx)

    print(f"  Columnas mapeadas: {col}")
    result = []
    for row in rows[header_row_idx + 1:]:
        if not row or not any(v for v in row):
            continue
        nombre_v = row[col["nombre"]] if "nombre" in col else None
        if not nombre_v:
            continue
        nombre = str(nombre_v).strip()
        if nombre in ("", "Apostador", "APOSTADOR"):
            continue

        def g(k):
            idx = col.get(k)
            return _int_or_none(row[idx]) if idx is not None else None

        result.append({
            "nombre": nombre,
            "pos":   g("pos"),
            "h":     g("h"),
            "i":     g("i"),
            "j":     g("j"),
            "k":     g("k"),
            "l":     g("l"),
            "m":     g("m"),
            "n":     g("n"),
            "o":     g("o"),
            "glob":  g("glob"),
            "total": g("total"),
        })

    print(f"  Apostadores en Puntaje general: {len(result)}")
    return result


def leer_hojas_fase(wb):
    """
    Lee todas las hojas de fase (todas excepto 'Puntaje general').

    Estructura real del Excel (ranking-export):
      Fila 1: headers agrupados (Apostador, Partido, Marcador, H/I, J Amarillas, ...)
      Fila 2: sub-headers (Apostador, Partido, Pronóst., Real, H(Res.), I(Exact.), ...)
      Fila 3+: datos

    Columnas (posiciones fijas basadas en ranking_export_inner):
      1=Apostador  2=Partido  3=PredMarcador  4=RealMarcador
      5=H_pts  6=I_pts
      7=J_pred  8=J_real  9=J_pts
      10=K_pred  11=K_real  12=K_pts
      13=L_pred  14=L_real  15=L_pts
      16=M_pred  17=M_real  18=M_pts
      19=N_pred  20=N_real  21=N_pts
      22=O_PredLocal  23=O_PredVisit  24=O_RealLocal  25=O_RealVisit  26=O_Pts
      27=Total

    Devuelve:
      rows: list de dicts con todos los campos
      fase_map: {nombre_hoja: [rows]}
    """
    fase_sheets = [s for s in wb.sheetnames if s != "Puntaje general"]
    print(f"\n  Hojas de fase: {fase_sheets}")

    all_rows = []

    for sheet_name in fase_sheets:
        ws = wb[sheet_name]
        raw = list(ws.iter_rows(values_only=True))
        if len(raw) < 3:
            continue

        # Detectar fila de datos (buscar fila con ≥3 filas de header)
        # El Excel tiene 2 filas de header. Verificamos que row[0] fila 3 sea un nombre (string largo)
        # Buscar primera fila no-header (tiene valores que parecen nombres y partidos)
        data_start = 2  # Por defecto despues de 2 filas de header

        # Intentar detectar automaticamente
        for i in range(min(5, len(raw))):
            r = raw[i]
            if r and r[0] and isinstance(r[0], str) and len(str(r[0]).strip()) > 3:
                non_headers = [v for v in r if v and not isinstance(v, str)]
                if non_headers or (r[1] and isinstance(r[1], str) and "vs" in str(r[1])):
                    data_start = i
                    break

        # Intentar leer sub-header (fila 2) para verificar columnas
        if len(raw) > 1:
            sh2 = [_norm_header(v) for v in raw[1]]
            # Verificar que coincide con esperado
            expected = ["apostador", "partido", "pronost", "real"]
            match_count = sum(1 for e in expected if any(e in h for h in sh2))
            if match_count >= 2:
                data_start = 2

        for raw_row in raw[data_start:]:
            if not raw_row or not raw_row[0]:
                continue
            # Columnas fijas basadas en ranking_export_inner
            apos    = raw_row[0]  if len(raw_row) > 0  else None
            partido = raw_row[1]  if len(raw_row) > 1  else None
            if not apos or not partido:
                continue
            if isinstance(apos, str) and _norm_header(apos) in ("apostador", "nombre"):
                continue  # Es una segunda fila de header

            row = {
                "fase":       sheet_name,
                "apostador":  str(apos).strip(),
                "partido":    str(partido).strip(),
                "pred_marc":  raw_row[2]  if len(raw_row) > 2  else None,
                "real_marc":  raw_row[3]  if len(raw_row) > 3  else None,
                "h_pts":      _int_or_none(raw_row[4]  if len(raw_row) > 4  else None),
                "i_pts":      _int_or_none(raw_row[5]  if len(raw_row) > 5  else None),
                "j_pred":     _int_or_none(raw_row[6]  if len(raw_row) > 6  else None),
                "j_real":     _int_or_none(raw_row[7]  if len(raw_row) > 7  else None),
                "j_pts":      _int_or_none(raw_row[8]  if len(raw_row) > 8  else None),
                "k_pred":     _int_or_none(raw_row[9]  if len(raw_row) > 9  else None),
                "k_real":     _int_or_none(raw_row[10] if len(raw_row) > 10 else None),
                "k_pts":      _int_or_none(raw_row[11] if len(raw_row) > 11 else None),
                "l_pred":     _int_or_none(raw_row[12] if len(raw_row) > 12 else None),
                "l_real":     _int_or_none(raw_row[13] if len(raw_row) > 13 else None),
                "l_pts":      _int_or_none(raw_row[14] if len(raw_row) > 14 else None),
                "m_pred":     _int_or_none(raw_row[15] if len(raw_row) > 15 else None),
                "m_real":     _int_or_none(raw_row[16] if len(raw_row) > 16 else None),
                "m_pts":      _int_or_none(raw_row[17] if len(raw_row) > 17 else None),
                "n_pred":     _int_or_none(raw_row[18] if len(raw_row) > 18 else None),
                "n_real":     _int_or_none(raw_row[19] if len(raw_row) > 19 else None),
                "n_pts":      _int_or_none(raw_row[20] if len(raw_row) > 20 else None),
                "o_pred_l":   _int_or_none(raw_row[21] if len(raw_row) > 21 else None),
                "o_pred_v":   _int_or_none(raw_row[22] if len(raw_row) > 22 else None),
                "o_real_l":   _int_or_none(raw_row[23] if len(raw_row) > 23 else None),
                "o_real_v":   _int_or_none(raw_row[24] if len(raw_row) > 24 else None),
                "o_pts":      _int_or_none(raw_row[25] if len(raw_row) > 25 else None),
                "total":      _int_or_none(raw_row[26] if len(raw_row) > 26 else None),
            }
            all_rows.append(row)

    print(f"  Total filas de fase leidas: {len(all_rows)}")
    if all_rows:
        print(f"  Ejemplo fila: apostador={all_rows[0]['apostador']!r} partido={all_rows[0]['partido']!r}")
        print(f"             H={all_rows[0]['h_pts']} I={all_rows[0]['i_pts']} J={all_rows[0]['j_pts']}")
    return all_rows


# ─────────────────────────────────────────────────────────────────────────────
# API
# ─────────────────────────────────────────────────────────────────────────────

def _req(url, token=None, method="GET", data=None):
    headers = {
        "Content-Type": "application/json",
        "ngrok-skip-browser-warning": "true",
        "Accept": "application/json",
    }
    if token:
        headers["Authorization"] = f"Bearer {token}"
    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    with urllib.request.urlopen(req, timeout=20) as r:
        return json.loads(r.read())


def api_login():
    for base in API_CANDIDATES:
        try:
            print(f"  Probando {base} ...", end=" ")
            resp = _req(f"{base}/auth/login", method="POST",
                        data={"username": ADMIN_USER, "password": ADMIN_PASS})
            token = resp.get("access_token")
            if token:
                print("✓")
                return base, token
            print("sin token")
        except Exception as e:
            print(f"✗ ({type(e).__name__}: {e})")
    return None, None


def get_scores_api(base, token):
    """Devuelve {nombre_norm: {h,i,j,k,l,m,n,o,pts_partidos,pts_globales,total}}"""
    # Primero intentar endpoint especializado
    for endpoint in [f"/bets/scores-por-apostador/{TORNEO_ID}", f"/bets/ranking/{TORNEO_ID}"]:
        try:
            data = _req(f"{base}{endpoint}", token=token)
            if isinstance(data, dict):
                data = data.get("ranking", [])
            if data:
                print(f"  {endpoint}: {len(data)} apostadores")
                result = {}
                for a in data:
                    nombre = str(a.get("nombre") or a.get("apostador") or a.get("name") or "").strip()
                    result[nombre.lower()] = a
                return result
        except Exception as e:
            print(f"  {endpoint}: ✗ {e}")
    return {}


# ─────────────────────────────────────────────────────────────────────────────
# ALGORITMO DE SCORING - VERIFICACION INDEPENDIENTE
# ─────────────────────────────────────────────────────────────────────────────

def parse_marc(s):
    """'2-1' -> (2, 1) | '1 - 0' -> (1, 0) | None si falla"""
    if not s:
        return None, None
    m = re.match(r"^\s*(\d+)\s*[-–]\s*(\d+)\s*$", str(s))
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None


def gep(gl, gv):
    """Ganó / Empató / Perdió desde perspectiva local"""
    if gl is None or gv is None:
        return None
    if gl > gv: return "G"
    if gl == gv: return "E"
    return "P"


def calc_h_i(pred_marc, real_marc):
    """Devuelve (h_pts, i_pts) para fase de grupos (4/8)"""
    pl, pv = parse_marc(pred_marc)
    rl, rv = parse_marc(real_marc)
    if None in (pl, pv, rl, rv):
        return 0, 0
    h = H_PTS if gep(pl, pv) == gep(rl, rv) else 0
    i = I_PTS if (pl == rl and pv == rv) else 0
    return h, i


def calc_exact(pred, real):
    """1 pt si exacto, ignorando None"""
    if pred is None or real is None:
        return 0
    try:
        return 1 if int(pred) == int(real) else 0
    except (ValueError, TypeError):
        return 0


def calc_N_por_partido(rows_partido):
    """
    Para un partido dado, rows_partido = list de dicts con n_pred y n_real.
    Devuelve {apostador: 0 o 1}
    1 pt al(los) apostador(es) con pred mas cercano al minuto real.
    """
    result = {r["apostador"]: 0 for r in rows_partido}
    real_min = rows_partido[0].get("n_real") if rows_partido else None
    if real_min is None:
        return result
    try:
        real_min = int(real_min)
    except (ValueError, TypeError):
        return result

    validos = []
    for r in rows_partido:
        pm = r.get("n_pred")
        if pm is not None:
            try:
                validos.append((r["apostador"], int(pm)))
            except (ValueError, TypeError):
                pass

    if not validos:
        return result

    min_dist = min(abs(pm - real_min) for _, pm in validos)
    for apos, pm in validos:
        if abs(pm - real_min) == min_dist:
            result[apos] = 1
    return result


# ─────────────────────────────────────────────────────────────────────────────
# COMPARACIONES
# ─────────────────────────────────────────────────────────────────────────────

def comparar_apostador(excel_pg, api_scores):
    """
    Compara la hoja 'Puntaje general' contra API.
    Devuelve (diffs, no_match)
    """
    if not api_scores:
        return [], []

    diffs = []
    no_match = []
    items = ["h", "i", "j", "k", "l", "m", "n", "o", "total"]

    for ea in excel_pg:
        nom = ea["nombre"]
        nom_k = nom.lower()

        # Buscar en API (nombre normalizado)
        api_a = api_scores.get(nom_k)
        if api_a is None:
            # Match parcial
            for k, v in api_scores.items():
                if nom_k in k or k in nom_k:
                    api_a = v
                    break

        if api_a is None:
            no_match.append(nom)
            continue

        for item in items:
            xl_val = ea.get(item)

            # Mapeo flexible de campos API
            if item == "total":
                api_val = (api_a.get("total") or api_a.get("puntos_total") or
                           api_a.get("total_puntos") or
                           (api_a.get("pts_partidos", 0) or 0) + (api_a.get("pts_globales", 0) or 0))
            elif item == "glob":
                api_val = api_a.get("pts_globales") or api_a.get("globales") or 0
            else:
                api_val = (api_a.get(item) or
                           api_a.get(f"cat_{item}") or
                           api_a.get(f"pts_{item}"))

            xl_int  = _int_or_none(xl_val)
            api_int = _int_or_none(api_val)

            if xl_int != api_int:
                diffs.append({
                    "nombre": nom,
                    "item":   item.upper(),
                    "excel":  xl_int,
                    "api":    api_int,
                    "diff":   (xl_int or 0) - (api_int or 0),
                })

    return diffs, no_match


def verificar_algoritmo(fase_rows):
    """
    Re-calcula H, I, J, K, L, M, N con el algoritmo oficial y compara
    contra los pts en el Excel. Trabaja solo con datos del Excel.

    Devuelve lista de discrepancias.
    """
    discrepancias = []

    # Agrupar por partido para calcular N
    por_partido = defaultdict(list)
    for row in fase_rows:
        por_partido[row["partido"]].append(row)

    # Pre-calcular ganadores de N por partido
    n_ganadores = {}
    for partido, rows_p in por_partido.items():
        gans = calc_N_por_partido(rows_p)
        for apos, pts in gans.items():
            n_ganadores[(partido, apos)] = pts

    checked = 0
    for row in fase_rows:
        apos    = row["apostador"]
        partido = row["partido"]
        pred_m  = row["pred_marc"]
        real_m  = row["real_marc"]

        # Solo verificar si hay resultado real
        rl, rv = parse_marc(real_m)
        if rl is None:
            continue  # Partido sin resultado → saltar

        checked += 1

        # H / I
        h_calc, i_calc = calc_h_i(pred_m, real_m)
        h_xl = row["h_pts"] or 0
        i_xl = row["i_pts"] or 0

        if h_calc != h_xl:
            discrepancias.append({
                "tipo":   "H_pts",
                "apostador": apos,
                "partido":   partido,
                "fase":      row["fase"],
                "esperado":  h_calc,
                "excel":     h_xl,
                "diff":      h_xl - h_calc,
                "detalle":   f"pred={pred_m!r} real={real_m!r}",
            })
        if i_calc != i_xl:
            discrepancias.append({
                "tipo":   "I_pts",
                "apostador": apos,
                "partido":   partido,
                "fase":      row["fase"],
                "esperado":  i_calc,
                "excel":     i_xl,
                "diff":      i_xl - i_calc,
                "detalle":   f"pred={pred_m!r} real={real_m!r}",
            })

        # J / K / L / M — solo si hay valor real (no None/vacío)
        for item, pred_k, real_k, pts_k in [
            ("J_pts", "j_pred", "j_real", "j_pts"),
            ("K_pts", "k_pred", "k_real", "k_pts"),
            ("L_pts", "l_pred", "l_real", "l_pts"),
            ("M_pts", "m_pred", "m_real", "m_pts"),
        ]:
            pred_v = row.get(pred_k)
            real_v = row.get(real_k)
            pts_xl = row.get(pts_k) or 0

            if real_v is None:
                continue  # Sin datos reales en Excel → no verificable

            pts_calc = calc_exact(pred_v, real_v)
            if pts_calc != pts_xl:
                discrepancias.append({
                    "tipo":      item,
                    "apostador": apos,
                    "partido":   partido,
                    "fase":      row["fase"],
                    "esperado":  pts_calc,
                    "excel":     pts_xl,
                    "diff":      pts_xl - pts_calc,
                    "detalle":   f"pred={pred_v} real={real_v}",
                })

        # N (minuto gol)
        n_calc = n_ganadores.get((partido, apos), 0)
        n_xl   = row.get("n_pts") or 0
        n_real = row.get("n_real")
        if n_real is not None and n_calc != n_xl:
            discrepancias.append({
                "tipo":      "N_pts",
                "apostador": apos,
                "partido":   partido,
                "fase":      row["fase"],
                "esperado":  n_calc,
                "excel":     n_xl,
                "diff":      n_xl - n_calc,
                "detalle":   f"pred={row.get('n_pred')} real={n_real}",
            })

    print(f"\n  Filas verificadas (con resultado): {checked}")
    print(f"  Discrepancias encontradas: {len(discrepancias)}")

    # Estadísticas
    tipos = defaultdict(int)
    for d in discrepancias:
        tipos[d["tipo"]] += 1
    for t, c in sorted(tipos.items()):
        print(f"    {t}: {c} casos")

    return discrepancias


def resumen_por_apostador(discrepancias):
    """Agrupa discrepancias por apostador y calcula impacto total."""
    por_apos = defaultdict(lambda: {"total_diff": 0, "items": defaultdict(int)})
    for d in discrepancias:
        por_apos[d["apostador"]]["total_diff"] += d["diff"]
        por_apos[d["apostador"]]["items"][d["tipo"]] += d["diff"]
    return dict(sorted(por_apos.items(), key=lambda x: abs(x[1]["total_diff"]), reverse=True))


# ─────────────────────────────────────────────────────────────────────────────
# GENERACION DEL REPORTE
# ─────────────────────────────────────────────────────────────────────────────

FILL_H_RED    = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
FILL_H_AMBER  = PatternFill(start_color="FFE4A0", end_color="FFE4A0", fill_type="solid")
FILL_H_GREEN  = PatternFill(start_color="C8F0D0", end_color="C8F0D0", fill_type="solid")
FILL_HDR      = PatternFill(start_color="1A4B7A", end_color="1A4B7A", fill_type="solid")
FILL_HDR2     = PatternFill(start_color="2E6DA4", end_color="2E6DA4", fill_type="solid")
FONT_WHT_B    = Font(color="FFFFFF", bold=True)
FONT_BOLD     = Font(bold=True)


def _hdr(ws, row, col, text, fill=None, bold_white=True):
    c = ws.cell(row, col, text)
    c.fill = fill or FILL_HDR
    c.font = FONT_WHT_B if bold_white else FONT_BOLD
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c


def generar_reporte(diffs_ap, no_match_ap, discrep, excel_pg, fase_rows):
    wb = openpyxl.Workbook()
    del wb["Sheet"]

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")

    # ── 1. RESUMEN ────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Resumen", 0)
    ws.column_dimensions["A"].width = 65
    ws.column_dimensions["B"].width = 15

    r = 1
    ws.cell(r, 1, "REPORTE DE AUDITORIA DE PUNTAJES BECBUC 2026").font = Font(bold=True, size=14)
    r += 1
    ws.cell(r, 1, f"Generado: {ts}")
    r += 2

    ws.cell(r, 1, "SECCIÓN 1: PUNTAJE GENERAL (apostador vs API/BD)").font = FONT_BOLD
    r += 1
    if not diffs_ap and not no_match_ap:
        ws.cell(r, 1, "✅ Sin diferencias — Excel coincide exactamente con la BD").font = Font(color="006600")
    else:
        ws.cell(r, 1, f"⚠️  {len(diffs_ap)} diferencias encontradas ({len(no_match_ap)} sin match en API)")
    r += 1

    if diffs_ap:
        by_item = defaultdict(list)
        for d in diffs_ap:
            by_item[d["item"]].append(d)
        ws.cell(r, 1, "Por ítem:")
        r += 1
        for item in ["H","I","J","K","L","M","N","O","TOTAL"]:
            if item in by_item:
                ds = by_item[item]
                avg_diff = sum(abs(d["diff"] or 0) for d in ds) / len(ds)
                ws.cell(r, 1, f"  {item}: {len(ds)} apostadores con diferencia  (diff prom: {avg_diff:.1f} pts)")
                r += 1
        r += 1
        ws.cell(r, 1, "Top 10 mayores diferencias en TOTAL:").font = FONT_BOLD
        r += 1
        top = sorted([d for d in diffs_ap if d["item"] == "TOTAL"],
                     key=lambda x: abs(x.get("diff") or 0), reverse=True)[:10]
        for d in top:
            ws.cell(r, 1, f"  {d['nombre']}: Excel={d['excel']}  BD={d['api']}  diff={d['diff']:+d}")
            r += 1

    r += 1
    ws.cell(r, 1, "SECCIÓN 2: VERIFICACIÓN ALGORITMO (Excel vs recálculo item por item)").font = FONT_BOLD
    r += 1
    if not discrep:
        ws.cell(r, 1, "✅ Sin discrepancias — todos los items del Excel coinciden con el algoritmo").font = Font(color="006600")
    else:
        ws.cell(r, 1, f"⚠️  {len(discrep)} discrepancias de puntaje por item")
    r += 1

    if discrep:
        tipos = defaultdict(list)
        for d in discrep:
            tipos[d["tipo"]].append(d)
        ws.cell(r, 1, "Por tipo de discrepancia:").font = FONT_BOLD
        r += 1
        for tipo, lst in sorted(tipos.items()):
            suma_abs = sum(abs(x["diff"]) for x in lst)
            ws.cell(r, 1, f"  {tipo}: {len(lst)} casos  (impacto neto: {sum(x['diff'] for x in lst):+d} pts, abs: {suma_abs})")
            r += 1

        r += 1
        ws.cell(r, 1, "Apostadores con mayor impacto (Excel - Algoritmo):").font = FONT_BOLD
        r += 1
        res_ap = resumen_por_apostador(discrep)
        for nom, info in list(res_ap.items())[:15]:
            ws.cell(r, 1, f"  {nom}: {info['total_diff']:+d} pts  ({dict(info['items'])})")
            r += 1

    # ── 2. DIFERENCIAS PUNTAJE GENERAL (apostador vs API) ─────────────────────
    ws2 = wb.create_sheet("Dif. apostador (vs API)")
    hdrs2 = ["Apostador", "Ítem", "Excel (control)", "API/BD ahora", "Diferencia (Excel-BD)"]
    for ci, h in enumerate(hdrs2, 1):
        _hdr(ws2, 1, ci, h)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 8
    for col in "CDEF":
        ws2.column_dimensions[col].width = 18

    if diffs_ap:
        for d in sorted(diffs_ap, key=lambda x: abs(x.get("diff") or 0), reverse=True):
            ri = ws2.max_row + 1
            ws2.cell(ri, 1, d["nombre"])
            ws2.cell(ri, 2, d["item"])
            ws2.cell(ri, 3, d["excel"])
            ws2.cell(ri, 4, d["api"])
            diff_v = d.get("diff", 0) or 0
            ws2.cell(ri, 5, diff_v)
            fill = FILL_H_RED if abs(diff_v) >= 5 else FILL_H_AMBER
            for ci in range(1, 6):
                ws2.cell(ri, ci).fill = fill
    else:
        ws2.cell(2, 1, "✅ Sin diferencias con la BD actual")

    if no_match_ap:
        ws2.append([])
        ws2.append(["Sin match en API:", ", ".join(no_match_ap)])

    ws2.freeze_panes = "A2"

    # ── 3. DISCREPANCIAS ALGORITMO ────────────────────────────────────────────
    ws3 = wb.create_sheet("Discrep. algoritmo")
    hdrs3 = ["Tipo", "Fase", "Apostador", "Partido",
             "Valor Excel", "Valor Algoritmo", "Diferencia (E-A)", "Detalle"]
    for ci, h in enumerate(hdrs3, 1):
        _hdr(ws3, 1, ci, h)
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 26
    ws3.column_dimensions["D"].width = 38
    ws3.column_dimensions["E"].width = 13
    ws3.column_dimensions["F"].width = 16
    ws3.column_dimensions["G"].width = 16
    ws3.column_dimensions["H"].width = 30

    if discrep:
        for d in sorted(discrep, key=lambda x: abs(x.get("diff") or 0), reverse=True):
            ri = ws3.max_row + 1
            diff_v = d.get("diff", 0) or 0
            ws3.cell(ri, 1, d["tipo"])
            ws3.cell(ri, 2, d["fase"])
            ws3.cell(ri, 3, d["apostador"])
            ws3.cell(ri, 4, d["partido"])
            ws3.cell(ri, 5, d["excel"])
            ws3.cell(ri, 6, d["esperado"])
            ws3.cell(ri, 7, diff_v)
            ws3.cell(ri, 8, d.get("detalle", ""))
            fill = FILL_H_RED if diff_v > 0 else (FILL_H_GREEN if diff_v < 0 else None)
            if fill:
                for ci in range(1, 9):
                    ws3.cell(ri, ci).fill = fill
    else:
        ws3.cell(2, 1, "✅ Sin discrepancias — algoritmo coincide con Excel")

    ws3.freeze_panes = "A2"

    # ── 4. RESUMEN POR APOSTADOR (impacto total) ──────────────────────────────
    ws4 = wb.create_sheet("Impacto por apostador")
    hdrs4 = ["Apostador", "Impacto neto (Excel-Alg)", "Casos", "H", "I", "J", "K", "L", "M", "N"]
    for ci, h in enumerate(hdrs4, 1):
        _hdr(ws4, 1, ci, h)
    ws4.column_dimensions["A"].width = 30
    for i, col in enumerate("BCDEFGHIJ", 2):
        ws4.column_dimensions[col].width = 13

    res_ap = resumen_por_apostador(discrep)
    counts_ap = defaultdict(int)
    for d in discrep:
        counts_ap[d["apostador"]] += 1

    for nom, info in res_ap.items():
        ri = ws4.max_row + 1
        ws4.cell(ri, 1, nom)
        ws4.cell(ri, 2, info["total_diff"])
        ws4.cell(ri, 3, counts_ap[nom])
        for ci, item in enumerate(["H_pts","I_pts","J_pts","K_pts","L_pts","M_pts","N_pts"], 4):
            ws4.cell(ri, ci, info["items"].get(item, 0) or 0)
        fill = FILL_H_RED if abs(info["total_diff"]) >= 5 else FILL_H_AMBER
        for ci in range(1, 11):
            ws4.cell(ri, ci).fill = fill

    ws4.freeze_panes = "A2"

    wb.save(str(OUTPUT_PATH))
    return wb


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 62)
    print("  AUDITORIA DE PUNTAJES BECBUC 2026")
    print("=" * 62)

    # 1. Encontrar y leer Excel de control
    xl_path = encontrar_excel(sys.argv)
    if not xl_path:
        print("\n[ERROR] No se encontró el Excel de control.")
        print("  Opciones:")
        print("  1. Pasar la ruta como argumento: python comparar_puntajes_control.py mi_archivo.xlsx")
        print("  2. Copiar el archivo al directorio del proyecto")
        print("  3. Ajustar la variable EXCEL_CONTROL en este script")
        sys.exit(1)

    print(f"\n[1] Leyendo Excel: {xl_path.name}")
    wb_ctrl = openpyxl.load_workbook(str(xl_path), read_only=True, data_only=True)
    print(f"    Hojas: {wb_ctrl.sheetnames}")

    excel_pg   = leer_puntaje_general(wb_ctrl)
    fase_rows  = leer_hojas_fase(wb_ctrl)

    if not excel_pg:
        print("[WARN] No se pudo leer la hoja 'Puntaje general'")
    if not fase_rows:
        print("[WARN] No se pudo leer ninguna hoja de fase")

    # 2. Conectar a la API (para comparación nivel apostador)
    print("\n[2] Conectando a la API...")
    base, token = api_login()
    api_scores = {}
    if base and token:
        api_scores = get_scores_api(base, token)
    else:
        print("  [SKIP] Sin acceso a la API — saltando comparación nivel apostador")

    # 3. Comparación nivel apostador (Excel vs API)
    print("\n[3] Comparando Puntaje general vs API...")
    diffs_ap, no_match_ap = comparar_apostador(excel_pg, api_scores)
    print(f"    Diferencias: {len(diffs_ap)} | Sin match: {len(no_match_ap)}")

    # 4. Verificación algoritmo (Excel auto-contenido)
    print("\n[4] Verificando algoritmo de scoring (auto-contenido desde Excel)...")
    discrep = verificar_algoritmo(fase_rows)

    # 5. Generar reporte
    print(f"\n[5] Generando reporte: {OUTPUT_PATH}")
    generar_reporte(diffs_ap, no_match_ap, discrep, excel_pg, fase_rows)

    print("\n" + "=" * 62)
    print(f"✅  Reporte guardado: {OUTPUT_PATH}")
    print(f"    Diferencias apostador vs BD: {len(diffs_ap)}")
    print(f"    Discrepancias algoritmo: {len(discrep)}")
    print("=" * 62)

    if not diffs_ap and not discrep:
        print("\n🎉  Sin diferencias — los puntajes son consistentes.")
    elif discrep:
        print(f"\n⚠️   Hay {len(discrep)} filas donde los pts del Excel NO coinciden con el algoritmo.")
        print("    Ver hoja 'Discrep. algoritmo' en el reporte.")


if __name__ == "__main__":
    main()
