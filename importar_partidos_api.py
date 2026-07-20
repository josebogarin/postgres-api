"""
importar_partidos_api.py
Lee apuestas de partido (P001-P104) de la hoja '50- TBL MASTER' del Excel consolidado
y las importa llamando al endpoint POST /bets/importar-apuestas-grupos/{torneo_id}.

Uso:
    python importar_partidos_api.py <ruta_excel> [--torneo-id 2] [--dry-run] [--crear-usuarios]

Ejemplo:
    python importar_partidos_api.py "20260611_2000- TBL CONSOLIDADA PRONOSTICOS ok.xlsx"
    python importar_partidos_api.py "...xlsx" --dry-run
    python importar_partidos_api.py "...xlsx" --crear-usuarios

Columnas relevantes de TBL MASTER:
  col  0: ID MASTER
  col  1: ID PARTIDO  (P001..P118)
  col  2: FECHA
  col  3: HORA
  col  4: CIUDAD
  col  5: NOMBRE PARTIDO
  col  6: FASE        (10- GRUPOS / 70- CLASIFICADOS / 80- GLOBALES)
  col  7: GRUPO
  col  8: NOMBRE      (nombre completo apostador)
  col  9: ALIAS       (username en el sistema)
  col 10: EQUIPO 1    (local)
  col 11: NOM CORTO1
  col 12: GOLES       (prediccion local)
  col 13: vs
  col 14: GOLES       (prediccion visitante)
  col 15: EQUIPO 2
  ...
  col 23: J-AMARILLAS (prediccion)
  col 24: K-ROJAS     (prediccion)
  col 25: L-VAR       (prediccion)
  col 26: M-PENALES   (prediccion)
  col 27: N-1ER GOL   (prediccion minuto)
"""

import sys
import os
import json
import argparse
import unicodedata
import urllib.request
import urllib.error
import openpyxl

API_BASE   = "http://localhost:8000/api/v1"
ADMIN_USER = "admin"
ADMIN_PASS = "faute1964"
HOJA       = "50- TBL MASTER"
TORNEO_ID  = None   # None = auto-detect desde /torneo/activas

# Solo importar filas con FASE de grupos
FASES_GRUPO = {"10- grupos", "grupos", "10-grupos"}


def norm(s: str) -> str:
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return s.lower()


def limpiar_alias(v) -> str:
    if v is None:
        return ""
    return str(v).replace("\xa0", "").strip()


def to_int(v, default=None):
    """Convierte a int; si no puede, retorna default."""
    if v is None:
        return default
    try:
        f = float(str(v).strip())
        return int(f)
    except (ValueError, TypeError):
        return default


def parse_partido_num(v) -> int | None:
    """'P001' → 1, '1' → 1, 1 → 1. Retorna None si no puede."""
    if v is None:
        return None
    s = str(v).strip()
    # Formato Pxxx o pxxx
    if s[0].upper() == "P":
        s = s[1:].lstrip("0") or "0"
    try:
        return int(s)
    except ValueError:
        return None


def api_post(path, body, token=None):
    url = API_BASE + path
    data = json.dumps(body, ensure_ascii=False).encode("utf-8")
    headers = {"Content-Type": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    req = urllib.request.Request(url, data=data, headers=headers, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            return json.loads(r.read().decode("utf-8")), r.status
    except urllib.error.HTTPError as e:
        body_err = e.read().decode("utf-8")
        try:
            return json.loads(body_err), e.code
        except Exception:
            return {"detail": body_err}, e.code


def api_get(path, token=None):
    url = API_BASE + path
    headers = {}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    req = urllib.request.Request(url, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            return json.loads(r.read().decode("utf-8")), r.status
    except urllib.error.HTTPError as e:
        body_err = e.read().decode("utf-8")
        try:
            return json.loads(body_err), e.code
        except Exception:
            return {"detail": body_err}, e.code


def login():
    resp, code = api_post("/auth/login", {"username": ADMIN_USER, "password": ADMIN_PASS})
    if code != 200 or "access_token" not in resp:
        raise RuntimeError(f"Login fallido ({code}): {resp}")
    return resp["access_token"]


def get_active_torneo_id(token: str) -> int:
    """Retorna el ID del torneo activo. Falla si no hay ninguno."""
    resp, code = api_get("/torneo/activas", token)
    if code == 200 and isinstance(resp, list) and resp:
        tid = resp[0].get("id") or resp[0].get("torneo_id")
        nombre = resp[0].get("nombre", "")
        print(f"  Torneo activo detectado: '{nombre}' (id={tid})")
        return int(tid)
    raise RuntimeError(
        f"No se pudo detectar torneo activo (HTTP {code}). "
        "Pasá --torneo-id manualmente."
    )


def leer_partidos(path: str, hoja: str = HOJA) -> list[dict]:
    """
    Retorna lista de dicts con los campos para ImportRowIn.
    Solo incluye filas de fase de grupos (P001..P104 approx)
    con al menos uno de goles_local o goles_visitante no nulo.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if hoja not in wb.sheetnames:
        raise RuntimeError(f"Hoja '{hoja}' no encontrada. Hojas: {wb.sheetnames}")

    ws = wb[hoja]
    rows_out = []
    skipped_fase = 0
    skipped_sin_goles = 0
    skipped_sin_alias = 0
    skipped_sin_num = 0

    for row in ws.iter_rows(values_only=True):
        # col 1: ID PARTIDO
        pid_raw = row[1] if len(row) > 1 else None
        partido_num = parse_partido_num(pid_raw)

        if partido_num is None or partido_num < 1:
            continue  # fila vacía o header

        # Solo partidos de grupos (P001-P104 aprox); globales son P111-P118
        if partido_num > 104:
            continue

        # col 6: FASE — filtrar solo grupos
        fase_raw = str(row[6]).strip() if len(row) > 6 and row[6] else ""
        fase_n = norm(fase_raw)
        if fase_n and "grupo" not in fase_n and "10-" not in fase_raw:
            skipped_fase += 1
            continue

        # col 9: ALIAS
        alias = limpiar_alias(row[9] if len(row) > 9 else None).lower()
        if not alias:
            skipped_sin_alias += 1
            continue

        # col 12: goles local (prediccion), col 14: goles visitante
        gl = to_int(row[12] if len(row) > 12 else None)
        gv = to_int(row[14] if len(row) > 14 else None)

        if gl is None and gv is None:
            skipped_sin_goles += 1
            continue

        # Bonus cols
        amarillas = to_int(row[23] if len(row) > 23 else None)
        rojas     = to_int(row[24] if len(row) > 24 else None)
        var       = to_int(row[25] if len(row) > 25 else None)
        penales   = to_int(row[26] if len(row) > 26 else None)
        minuto    = to_int(row[27] if len(row) > 27 else None)

        rows_out.append({
            "apostador":            alias,
            "partido_num":          partido_num,
            "goles_local":          gl if gl is not None else 0,
            "goles_visitante":      gv if gv is not None else 0,
            "pred_amarillas":       amarillas,
            "pred_rojas":           rojas,
            "pred_var":             var,
            "pred_penales_partido": penales,
            "pred_minuto_gol":      minuto,
        })

    wb.close()
    print(f"  Leídas {len(rows_out)} filas válidas")
    if skipped_fase:      print(f"  Saltadas por fase (no grupos): {skipped_fase}")
    if skipped_sin_goles: print(f"  Saltadas sin goles: {skipped_sin_goles}")
    if skipped_sin_alias: print(f"  Saltadas sin alias: {skipped_sin_alias}")
    if skipped_sin_num:   print(f"  Saltadas sin partido_num: {skipped_sin_num}")
    return rows_out


def main():
    parser = argparse.ArgumentParser(description="Importar apuestas de partido desde TBL MASTER")
    parser.add_argument("excel",            help="Ruta al Excel consolidado")
    parser.add_argument("--torneo-id",      type=int, default=None,
                        help="ID del torneo (default: auto desde /torneo/activas)")
    parser.add_argument("--dry-run",        action="store_true", help="Solo parsea, no importa")
    parser.add_argument("--crear-usuarios", action="store_true", help="Crea usuarios faltantes")
    parser.add_argument("--hoja",           default=None, help=f"Nombre de hoja (default: '{HOJA}')")
    args = parser.parse_args()

    hoja = args.hoja or HOJA

    if not os.path.exists(args.excel):
        print(f"ERROR: No existe {args.excel}")
        sys.exit(1)

    print(f"Leyendo {args.excel} (hoja: {hoja}) ...")
    rows = leer_partidos(args.excel, hoja)
    print(f"Total filas a importar: {len(rows)}")

    if not rows:
        print("Sin datos para importar.")
        sys.exit(0)

    # Resumen por apostador
    from collections import Counter
    por_ap = Counter(r["apostador"] for r in rows)
    print(f"\nApostadores detectados ({len(por_ap)}):")
    for alias, cnt in sorted(por_ap.items()):
        print(f"  {alias}: {cnt} apuestas")

    if args.dry_run:
        print("\n[DRY RUN] No se envía nada al servidor.")
        print("\nPrimeras 5 filas:")
        for r in rows[:5]:
            print(f"  {r}")
        sys.exit(0)

    print(f"\nLogin como admin ...")
    token = login()
    print("  Token OK")

    # Auto-detectar torneo activo si no se especificó
    torneo_id = args.torneo_id
    if torneo_id is None:
        torneo_id = get_active_torneo_id(token)

    # Confirmar antes de importar
    resp_confirm = input(f"\n¿Importar {len(rows)} filas al torneo {torneo_id}? (s/n): ")
    if resp_confirm.strip().lower() not in ("s", "si", "sí", "y", "yes"):
        print("Cancelado.")
        sys.exit(0)

    # Llamar al endpoint
    url_path = f"/bets/importar-apuestas-grupos/{torneo_id}"
    if args.crear_usuarios:
        url_path += "?crear_usuarios=true"

    print(f"\nEnviando {len(rows)} filas a {url_path} ...")

    resp, code = api_post(url_path, rows, token)

    if code == 200 and resp.get("ok"):
        print(f"\n✓ Importación exitosa:")
        print(f"  Creadas/actualizadas: {resp.get('creadas', '?')}")
        if resp.get("advertencias"):
            print("  Advertencias:")
            for w in resp["advertencias"]:
                print(f"    ⚠ {w}")
        if resp.get("usuarios_creados"):
            print(f"  Usuarios creados: {resp['usuarios_creados']}")
        if resp.get("errores"):
            print(f"  Errores ({len(resp['errores'])}):")
            for e in resp["errores"][:10]:  # máx 10
                print(f"    ✗ {e}")
    else:
        print(f"\n✗ Error {code}:")
        print(f"  {resp.get('detail', resp)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
