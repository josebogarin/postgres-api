# -*- coding: utf-8 -*-
"""
diag_var_L.py
Diagnostica las diferencias del item L (VAR) entre Excel (TBL MASTER) y BD.
Para cada apostador x partido con L distinto, imprime:
  VAR oficial(Excel) | VAR oficial(BD decisiones_var) | predVar(Excel) | predVar(BD)
  | L Excel | pts_var BD | CAUSA
y un resumen por causa. Solo lectura.
"""
import sys, os
BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = None
for f in os.listdir(BASE):
    if 'SEMIFINAL' in f.upper() and f.endswith('.xlsx'):
        EXCEL_FILE = os.path.join(BASE, f); break
if not EXCEL_FILE: sys.exit("no se encontro Excel *SEMIFINAL*.xlsx")
print(f"Excel: {os.path.basename(EXCEL_FILE)}")
try:
    import openpyxl
except ImportError:
    os.system(f'"{sys.executable}" -m pip install openpyxl --quiet'); import openpyxl
try:
    import psycopg2, psycopg2.extras
except ImportError:
    os.system(f'"{sys.executable}" -m pip install psycopg2-binary --quiet'); import psycopg2, psycopg2.extras
from collections import Counter

CONN_BEC = "host=localhost port=5432 dbname=becbuc user=app_user password=superpassword"
CONN_APP = "host=localhost port=5432 dbname=app_db  user=app_user password=superpassword"
TID = 2
conn = psycopg2.connect(CONN_BEC); capp = psycopg2.connect(CONN_APP)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
cua = capp.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

def ti(v):
    try:
        s=str(v).strip()
        if s in ('','-','None'): return None
        return int(float(s))
    except: return None

# alias -> uid
cua.execute("SELECT id, username FROM users WHERE is_active=TRUE")
uname = {u['username'].lower(): u['id'] for u in cua.fetchall()}
def ca(s):
    if not s: return ''
    return str(s).replace('\xa0','').strip().upper().lstrip('@').replace('  ',' ')
def uid_of(al):
    a=ca(al)
    for k,v in uname.items():
        if k.upper().lstrip('@')==a: return v
    for k,v in uname.items():
        if a and (a in k.upper() or k.upper() in a): return v
    return None

# BD: VAR oficial por numero_fifa + es_paraguay
cur.execute("""
    SELECT p.numero_fifa, p.decisiones_var AS var_bd,
           (el.nombre='Paraguay' OR ev.nombre='Paraguay') AS py
    FROM partido p JOIN fase f ON f.id=p.fase_id
    LEFT JOIN equipo el ON el.id=p.equipo_local_id
    LEFT JOIN equipo ev ON ev.id=p.equipo_visitante_id
    WHERE f.torneo_id=%s AND p.numero_fifa BETWEEN 1 AND 102
""", (TID,))
var_bd={}; py={}
for r in cur.fetchall():
    var_bd[r['numero_fifa']]=r['var_bd']; py[r['numero_fifa']]=r['py']

# BD: pred_var + pts_var por (uid, numero_fifa)
cur.execute("""
    SELECT a.apostador_id, p.numero_fifa, a.pred_var,
           COALESCE(pd.pts_var,0) AS pts_var
    FROM apuesta a JOIN partido p ON p.id=a.partido_id
    JOIN fase f ON f.id=p.fase_id
    LEFT JOIN puntaje_detalle pd ON pd.partido_id=a.partido_id AND pd.apostador_id=a.apostador_id
    WHERE f.torneo_id=%s AND p.numero_fifa BETWEEN 1 AND 102
""", (TID,))
bd={(r['apostador_id'], r['numero_fifa']): r for r in cur.fetchall()}

# Excel: VAR oficial (RESULTADOS OFICIALES) + pred_var/L por apostador (TBL MASTER)
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws1 = wb['40- RESULTADOS OFICIALES']
var_ex={}
for r in ws1.iter_rows(min_row=2, values_only=True):
    if isinstance(r[0],str) and r[0].startswith('P'):
        var_ex[int(r[0][1:])]=ti(r[31])
ws2 = wb['50- TBL MASTER']

causa = Counter(); filas=[]
for r in ws2.iter_rows(min_row=2, values_only=True):
    if not (isinstance(r[1],str) and r[1].startswith('P')): continue
    nf=int(r[1][1:])
    if nf<1 or nf>102: continue
    uid=uid_of(r[9])
    if not uid: continue
    key=(uid,nf); b=bd.get(key)
    if not b: continue
    pv_ex=ti(r[26]); l_ex=ti(r[43]) or 0
    pv_bd=b['pred_var']; pts_bd=b['pts_var']
    if l_ex==pts_bd: continue           # sin diferencia
    voex=var_ex.get(nf); vobd=b and var_bd.get(nf)
    # clasificar causa
    if voex!=vobd:
        c='VAR oficial BD != Excel'
    elif (pv_ex or 0)!=(pv_bd or 0):
        c='prediccion difiere'
    elif pv_bd is None and (vobd or 0)==0:
        c='regla NULL->0 (blanco=0, oficial 0) -> BD suma'
    else:
        c='otro'
    causa[c]+=1
    if len(filas)<40:
        filas.append((r[1], ca(r[9]), voex, vobd, pv_ex, pv_bd, l_ex, pts_bd, c))

print("\n"+"="*100)
print("DIAGNOSTICO ITEM L (VAR): diferencias Excel(TBL MASTER) vs BD(pts_var)")
print("="*100)
print(f"{'PART':<6}{'ALIAS':<16}{'VARof.Ex':>9}{'VARof.BD':>9}{'predEx':>8}{'predBD':>8}{'Lexcel':>8}{'ptsBD':>7}   CAUSA")
for pid,al,voex,vobd,pex,pbd,lex,pbd2,c in filas:
    print(f"{pid:<6}{al:<16}{str(voex):>9}{str(vobd):>9}{str(pex):>8}{str(pbd):>8}{lex:>8}{pbd2:>7}   {c}")
if len(filas)==40: print("  ... (muestra 40)")
print("\nRESUMEN POR CAUSA:")
for c,n in causa.most_common():
    print(f"   {n:>4}  {c}")
print(f"   TOTAL diferencias L: {sum(causa.values())}")
conn.close(); capp.close()
