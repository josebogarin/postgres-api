# -*- coding: utf-8 -*-
r"""
comparar_puntajes_fin_torneo.py

AUDITORIA DE ORIGEN (como la sesion 69): compara, item por item (H..P) y por
apostador x partido, los PUNTOS del Excel de fin de torneo (hoja TBL MASTER,
columnas de puntos ya calculadas) contra los PUNTOS de la BD (puntaje_detalle).

Para cada DIFERENCIA imprime:
   real (resultado oficial) | ap.Excel | ap.BD | ptsExcel | ptsBD | ORIGEN

ORIGEN = clasificacion del motivo (Excel vs BD), segun las reglas confirmadas:
   K/M con BD>Excel .... "BD lenient (null->0, por diseno)"   -> BD correcta
   N ................... "desempate minuto (decision organizacion)" -> BD correcta
   O ................... "quirk Excel (tanda sin penales)"      -> BD correcta
   P ................... "escala reglamento BD vs plano Excel"   -> BD correcta
   H/I con real==apuesta pero pts distinto ... "error del Excel (marcador)"
   otros .............. "REVISAR"

Cubre TODO el torneo: grupos (P001-072), 16avos, octavos, cuartos, semis,
final + 3er puesto (P103-P104).  SOLO LECTURA (no escribe nada).

Uso:
  backend\.venv\Scripts\python.exe comparar_puntajes_fin_torneo.py
"""
import sys, os
from collections import Counter, defaultdict

# nf_ini, nf_fin, es_ko
PHASES = {
    'grupos':  (1,   72, False),
    '16avos':  (73,  88, True),
    'octavos': (89,  96, True),
    'cuartos': (97, 100, True),
    'semis':   (101,102, True),
    'final3p': (103,104, True),
}
ORDER = ['grupos', '16avos', 'octavos', 'cuartos', 'semis', 'final3p']
NF_SET, fase_de_nf = set(), {}
for f in ORDER:
    a, b, _ = PHASES[f]
    for n in range(a, b + 1):
        NF_SET.add(n); fase_de_nf[n] = f
es_ko_de = {f: PHASES[f][2] for f in ORDER}

BASE = os.path.dirname(os.path.abspath(__file__))
# ESTRICTO: solo el Excel de FIN DE TORNEO (o ruta .xlsx pasada como argumento).
# Nunca hace fallback a otros .xlsx del root para no usar un archivo viejo por error.
_arg = next((a for a in sys.argv[1:] if not a.startswith('--') and a.lower().endswith('.xlsx')), None)
if _arg and os.path.exists(_arg):
    EXCEL_FILE = _arg
else:
    _OK = ('TORNEO CERRADO', 'FIN DE TORNEO', '20260720')
    _BAD = ('CORRECCIONES', 'SEMIFINAL')
    _cands = [os.path.join(BASE, f) for f in os.listdir(BASE)
              if f.lower().endswith('.xlsx') and not f.startswith('~')
              and any(k in f.upper() for k in _OK) and not any(b in f.upper() for b in _BAD)]
    if not _cands:
        sys.exit("ERROR: no encontre el Excel de cierre ('...torneo cerrado.xlsx' / '20260720...') en la raiz.\n"
                 "Copialo a C:\\proyecto FAST API\\ y saca del root el viejo 'con correcciones'.")
    EXCEL_FILE = sorted(_cands, key=lambda p: os.path.getmtime(p), reverse=True)[0]
print(f"Excel: {os.path.basename(EXCEL_FILE)}")

try:
    import openpyxl
except ImportError:
    os.system(f'"{sys.executable}" -m pip install openpyxl --quiet'); import openpyxl
try:
    import psycopg2, psycopg2.extras
except ImportError:
    os.system(f'"{sys.executable}" -m pip install psycopg2-binary --quiet'); import psycopg2, psycopg2.extras

CONN_BEC = "host=localhost port=5432 dbname=becbuc user=app_user password=superpassword"
CONN_APP = "host=localhost port=5432 dbname=app_db  user=app_user password=superpassword"
TID = 2

conn = psycopg2.connect(CONN_BEC); capp = psycopg2.connect(CONN_APP)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
cua = capp.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

cua.execute("SELECT id, username FROM users WHERE is_active=TRUE")
uname = {u['username'].lower(): u['id'] for u in cua.fetchall()}
def clean_alias(s):
    if not s: return ''
    return str(s).replace('\xa0','').strip().upper().lstrip('@').replace('  ',' ')
def find_uid(al):
    a = clean_alias(al)
    if not a: return None
    for k, v in uname.items():
        if k.upper().lstrip('@') == a: return v
    for k, v in uname.items():
        if a in k.upper() or k.upper() in a: return v
    return None

cur.execute("SELECT id, nombre, nombre_es FROM equipo")
ename = {}
for eq in cur.fetchall():
    ename[eq['id']] = eq['nombre'] or eq['nombre_es']
def enm(x): return ename.get(x, x) if x is not None else 'None'

cur.execute("""SELECT pd.apostador_id, p.numero_fifa,
       COALESCE(pd.pts_resultado,0) h, COALESCE(pd.pts_marcador,0) i,
       COALESCE(pd.pts_amarillas,0) j, COALESCE(pd.pts_rojas,0) k,
       COALESCE(pd.pts_var,0) l, COALESCE(pd.pts_penales_partido,0) m,
       COALESCE(pd.pts_minuto,0) n, COALESCE(pd.pts_penales_tanda,0) o,
       COALESCE(pd.pts_equipo,0) p
       FROM puntaje_detalle pd JOIN partido p ON p.id=pd.partido_id
       JOIN fase f ON f.id=p.fase_id WHERE f.torneo_id=%s""", (TID,))
bd_pts = {(r['apostador_id'], r['numero_fifa']): r for r in cur.fetchall()}

cur.execute("""SELECT a.apostador_id, p.numero_fifa,
       a.pred_local, a.pred_visitante, a.pred_amarillas, a.pred_rojas, a.pred_var,
       a.pred_penales_partido, a.pred_minuto_gol,
       a.pred_penales_local_tanda, a.pred_penales_visitante_tanda, a.pred_equipo_clasifica
       FROM apuesta a JOIN partido p ON p.id=a.partido_id
       JOIN fase f ON f.id=p.fase_id WHERE f.torneo_id=%s""", (TID,))
bd_pred = {(r['apostador_id'], r['numero_fifa']): r for r in cur.fetchall()}

cur.execute("""SELECT p.numero_fifa, p.goles_local, p.goles_visitante, p.amarillas, p.rojas,
       p.decisiones_var, p.penales_partido, p.minuto_primer_gol,
       p.penales_local, p.penales_visitante, p.equipo_clasificado_id
       FROM partido p JOIN fase f ON f.id=p.fase_id WHERE f.torneo_id=%s""", (TID,))
off = {r['numero_fifa']: r for r in cur.fetchall()}

def sc(v):
    try:
        s = str(v).strip()
        if s in ('','-','None'): return 0
        return int(float(s))
    except: return 0
def si(v):
    try:
        s = str(v).strip()
        if s in ('','-','None'): return None
        return int(float(s))
    except: return None

# columnas Excel (1-based) - iguales a comparar_puntajes_items.py (semifinal master)
C = dict(pid=2, alias=10, pl=13, pv=15,
         predJ=25, predK=26, predL=27, predM=28, predN=29, predOe1=30, predOe2=31, predQ=32,
         ptsH=40, ptsI=41, ptsJ=42, ptsK=43, ptsL=44, ptsM=45, ptsN=46, ptsOe1=47, ptsOe2=48, ptsP=51)
ITEM_DESC = {'H':'Resultado','I':'Marcador','J':'Amarillas','K':'Rojas','L':'VAR',
             'M':'Penales juego','N':'Minuto 1er gol','O':'Tanda','P':'Equipo pasa'}
ALL_ITEMS = ['H','I','J','K','L','M','N','O','P']

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
SHEET = next((s for s in wb.sheetnames if 'MASTER' in s.upper()), None)
if not SHEET:
    sys.exit("ERROR: no hallada la hoja TBL MASTER en el Excel.")
ws = wb[SHEET]
print(f"Hoja: '{SHEET}'  ({ws.max_row}x{ws.max_column})")

# auto-detect col puntos P (clasificados) por encabezado (cambia entre versiones)
for _c in range(45, 62):
    _h = ws.cell(1, _c).value
    if _h and 'CLASIFICAD' in str(_h).upper():
        C['ptsP'] = _c; break
print(f"Col puntos P (clasificados) detectada: {C['ptsP']}")
print("Cabeceras columnas de puntos (verificar que correspondan a H..O):")
for it, col in [('H',C['ptsH']),('I',C['ptsI']),('J',C['ptsJ']),('K',C['ptsK']),
                ('L',C['ptsL']),('M',C['ptsM']),('N',C['ptsN']),('O-1',C['ptsOe1']),
                ('O-2',C['ptsOe2']),('P',C['ptsP'])]:
    print(f"   {it:<3} col {col:>2}: {ws.cell(1, col).value}")

def origen(it, real, pex, pbd, ptse, ptsb):
    if it in ('K','M') and ptsb > ptse:
        return "BD lenient null->0 (por diseno; BD correcta)"
    if it == 'N':
        return "desempate minuto (decision org; BD correcta)"
    if it == 'O':
        return "quirk Excel tanda (BD correcta)"
    if it == 'P':
        return "escala reglamento BD vs plano Excel (BD correcta)"
    if it in ('H','I'):
        return "error del Excel (marcador/placeholder)" if str(real) == str(pex) else "REVISAR marcador"
    return "REVISAR"

tot = Counter(); dif = Counter(); detalles = []
for r in range(2, ws.max_row + 1):
    raw = ws.cell(r, C['pid']).value
    pid = str(raw).strip() if raw is not None else ''
    if not (len(pid) == 4 and pid[0] in 'Pp' and pid[1:].isdigit()): continue
    nf = int(pid[1:])
    if nf not in NF_SET: continue
    fase = fase_de_nf[nf]; es_ko = es_ko_de[fase]
    uid = find_uid(ws.cell(r, C['alias']).value)
    if not uid: continue
    b = bd_pts.get((uid, nf))
    if not b: continue
    alias = clean_alias(ws.cell(r, C['alias']).value)
    pr = bd_pred.get((uid, nf), {}); of = off.get(nf, {})

    exP = {'H': sc(ws.cell(r,C['ptsH']).value), 'I': sc(ws.cell(r,C['ptsI']).value),
           'J': sc(ws.cell(r,C['ptsJ']).value), 'K': sc(ws.cell(r,C['ptsK']).value),
           'L': sc(ws.cell(r,C['ptsL']).value), 'M': sc(ws.cell(r,C['ptsM']).value),
           'N': sc(ws.cell(r,C['ptsN']).value),
           'O': sc(ws.cell(r,C['ptsOe1']).value) + sc(ws.cell(r,C['ptsOe2']).value),
           'P': sc(ws.cell(r,C['ptsP']).value)}
    marc_real = f"{of.get('goles_local')}-{of.get('goles_visitante')}"
    marc_ex = f"{si(ws.cell(r,C['pl']).value)}-{si(ws.cell(r,C['pv']).value)}"
    marc_bd = f"{pr.get('pred_local')}-{pr.get('pred_visitante')}"
    REAL = {'H':marc_real,'I':marc_real,'J':of.get('amarillas'),'K':of.get('rojas'),
            'L':of.get('decisiones_var'),'M':of.get('penales_partido'),
            'N':of.get('minuto_primer_gol'),
            'O':f"{of.get('penales_local')}-{of.get('penales_visitante')}",
            'P':enm(of.get('equipo_clasificado_id'))}
    PEX = {'H':marc_ex,'I':marc_ex,'J':si(ws.cell(r,C['predJ']).value),'K':si(ws.cell(r,C['predK']).value),
           'L':si(ws.cell(r,C['predL']).value),'M':si(ws.cell(r,C['predM']).value),
           'N':si(ws.cell(r,C['predN']).value),
           'O':f"{si(ws.cell(r,C['predOe1']).value)}-{si(ws.cell(r,C['predOe2']).value)}",
           'P':(str(ws.cell(r,C['predQ']).value).strip() if ws.cell(r,C['predQ']).value not in (None,'','-') else 'None')}
    PBD = {'H':marc_bd,'I':marc_bd,'J':pr.get('pred_amarillas'),'K':pr.get('pred_rojas'),
           'L':pr.get('pred_var'),'M':pr.get('pred_penales_partido'),'N':pr.get('pred_minuto_gol'),
           'O':f"{pr.get('pred_penales_local_tanda')}-{pr.get('pred_penales_visitante_tanda')}",
           'P':enm(pr.get('pred_equipo_clasifica'))}

    for it in ALL_ITEMS:
        if it == 'P' and not es_ko: continue
        vb = b[it.lower()]; ve = exP[it]
        tot[(it, fase)] += 1
        if ve != vb:
            dif[(it, fase)] += 1
            detalles.append((fase, it, pid, alias, REAL[it], PEX[it], PBD[it], ve, vb,
                             origen(it, REAL[it], PEX[it], PBD[it], ve, vb)))

# ---------- REPORTE ----------
print("\n" + "="*112)
print("AUDITORIA DE ORIGEN - PUNTAJES POR ITEM (Excel fin de torneo vs BD) -- solo diferencias")
print("="*112)
print("\nRESUMEN item x fase (diferencias / comparaciones):")
hdr = f"{'ITEM':<6}" + "".join(f"{f[:8]:>11}" for f in ORDER) + f"{'TOTAL':>13}"
print(hdr); print("-"*len(hdr))
item_tot = Counter(); item_dif = Counter()
for it in ALL_ITEMS:
    line = f"{it:<6}"; hay = False
    for f in ORDER:
        t = tot[(it,f)]; d = dif[(it,f)]; item_tot[it]+=t; item_dif[it]+=d
        line += (f"{'-':>11}" if t == 0 else f"{(str(d)+'/'+str(t)):>11}")
        if t: hay = True
    line += f"{((str(item_dif[it])+'/'+str(item_tot[it])) if item_tot[it] else '-'):>13}"
    if hay: print(line)

print("\n" + "="*112)
print("DETALLE DE DIFERENCIAS  (real = resultado oficial)")
print("="*112)
by = defaultdict(list)
for d in detalles: by[d[1]].append(d)
for it in ALL_ITEMS:
    if it not in by: continue
    lst = by[it]
    print(f"\n### Item {it} ({ITEM_DESC[it]}) - {len(lst)} diferencia(s)")
    print(f"  {'FASE':<9}{'PART':<6}{'APOSTADOR':<16}{'real':>8} {'ap.Excel':>10} {'ap.BD':>10} {'ptsEx':>6}{'ptsBD':>6}  ORIGEN")
    for fase, _, pid, alias, real, pex, pbd, ptse, ptsb, org in sorted(lst, key=lambda x:(x[0],x[2],x[3]))[:150]:
        print(f"  {fase:<9}{pid:<6}{alias:<16}{str(real):>8} {str(pex):>10} {str(pbd):>10} {ptse:>6}{ptsb:>6}  {org}")
    if len(lst) > 150:
        print(f"  ... (+{len(lst)-150} mas)")

print("\n" + "="*112)
org_count = Counter(d[9] for d in detalles)
print(f"TOTAL diferencias: {len(detalles)}")
print("Por ORIGEN:")
for org, n in org_count.most_common():
    print(f"   {n:>4}  {org}")
print("\nItems marcados 'REVISAR' o 'error del Excel' son los que ameritan mirar a mano.")
print("El resto son casos donde la BD es la fuente correcta (reglamento / decision org).")

try: conn.close(); capp.close()
except Exception: pass
