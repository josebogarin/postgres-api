# -*- coding: utf-8 -*-
r"""
diag_var_real.py  (solo lectura)

Compara el RESULTADO OFICIAL de cada partido entre el Excel de cierre
(hoja '40- RESULTADOS OFICIALES') y la BD (tabla partido), para los items de
"dato de partido": VAR (decisiones_var), amarillas y rojas. Sirve para decidir
el item L (VAR): quien tiene la cifra correcta.

Autodetecta las columnas de la hoja por su encabezado e imprime SOLO las
diferencias, resaltando los partidos que la auditoria marco en VAR
(P063, P064, P065, P076, P079, P083, P104).

Uso:
  backend\.venv\Scripts\python.exe diag_var_real.py
"""
import sys, os, re
BASE = os.path.dirname(os.path.abspath(__file__))
_OK = ('TORNEO CERRADO', 'FIN DE TORNEO', '20260720'); _BAD = ('CORRECCIONES', 'SEMIFINAL')
_c = [os.path.join(BASE, f) for f in os.listdir(BASE)
      if f.lower().endswith('.xlsx') and not f.startswith('~')
      and any(k in f.upper() for k in _OK) and not any(b in f.upper() for b in _BAD)]
if not _c:
    sys.exit("ERROR: no encontre el Excel de cierre en la raiz.")
EXCEL = sorted(_c, key=lambda p: os.path.getmtime(p), reverse=True)[0]
print(f"Excel: {os.path.basename(EXCEL)}")

try:
    import openpyxl
except ImportError:
    os.system(f'"{sys.executable}" -m pip install openpyxl --quiet'); import openpyxl
try:
    import psycopg2, psycopg2.extras
except ImportError:
    os.system(f'"{sys.executable}" -m pip install psycopg2-binary --quiet'); import psycopg2, psycopg2.extras

CONN_BEC = "host=localhost port=5432 dbname=becbuc user=app_user password=superpassword"
TID = 2
FLAG = {63, 64, 65, 76, 79, 83, 104}   # partidos que la auditoria marco en VAR

def to_int(v):
    try:
        s = str(v).strip()
        if s in ('', '-', 'None', 'nan'): return None
        return int(float(s))
    except Exception:
        return None

wb = openpyxl.load_workbook(EXCEL, data_only=True)
SHEET = next((s for s in wb.sheetnames if 'RESULTADOS OFICIALES' in s.upper()), None)
if not SHEET:
    SHEET = next((s for s in wb.sheetnames if 'OFICIAL' in s.upper()), None)
if not SHEET:
    sys.exit("ERROR: no encontre la hoja '40- RESULTADOS OFICIALES'.")
ws = wb[SHEET]
print(f"Hoja: '{SHEET}'  ({ws.max_row}x{ws.max_column})")

# localizar fila de encabezado (la que tenga mas texto) en las primeras 5 filas
hdr_row = 1; best = -1
for r in range(1, 6):
    cnt = sum(1 for c in range(1, ws.max_column + 1) if isinstance(ws.cell(r, c).value, str))
    if cnt > best: best = cnt; hdr_row = r
print(f"Fila de encabezado detectada: {hdr_row}")
print("Encabezados no vacios:")
headers = {}
for c in range(1, ws.max_column + 1):
    h = ws.cell(hdr_row, c).value
    if h is not None and str(h).strip():
        headers[c] = str(h).strip()
        print(f"   col {c:>2}: {h}")

def find_col(*keys, avoid=()):
    for c, h in headers.items():
        u = h.upper()
        if any(k in u for k in keys) and not any(a in u for a in avoid):
            return c
    return None

col_pid  = find_col('ID PARTIDO', 'N PARTIDO', 'NRO', 'PARTIDO', avoid=('EQUIPO',))
col_var  = find_col('VAR')
col_amar = find_col('AMARILLA')
col_roja = find_col('ROJA', avoid=('AMARILLA',))
col_penj = find_col('PENALES', 'PENAL', avoid=('TANDA',))
print(f"\nColumnas detectadas -> pid={col_pid}  VAR={col_var}  amarillas={col_amar}  rojas={col_roja}  penales_juego={col_penj}")
if not col_pid or not col_var:
    print("\n*** No pude autodetectar pid/VAR. Revisar encabezados de arriba y avisame la columna. ***")

# BD
conn = psycopg2.connect(CONN_BEC)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
cur.execute("""SELECT p.numero_fifa, p.decisiones_var, p.amarillas, p.rojas, p.penales_partido
               FROM partido p JOIN fase f ON f.id=p.fase_id WHERE f.torneo_id=%s""", (TID,))
bd = {r['numero_fifa']: r for r in cur.fetchall()}

def pid_to_nf(v):
    if v is None: return None
    m = re.match(r'[Pp]?(\d{1,3})$', str(v).strip())
    return int(m.group(1)) if m else None

print("\n" + "=" * 78)
print("DIFERENCIAS DE DATO OFICIAL (Excel hoja RESULTADOS vs BD partido)")
print("=" * 78)
print(f"{'PART':<6}{'campo':<12}{'Excel':>8}{'BD':>8}   flag")
diffs = {'VAR': 0, 'amarillas': 0, 'rojas': 0, 'penales_juego': 0}
seen = 0
for r in range(hdr_row + 1, ws.max_row + 1):
    nf = pid_to_nf(ws.cell(r, col_pid).value) if col_pid else None
    if nf is None or nf not in bd:
        continue
    seen += 1
    b = bd[nf]
    checks = [('VAR', col_var, 'decisiones_var'),
              ('amarillas', col_amar, 'amarillas'),
              ('rojas', col_roja, 'rojas'),
              ('penales_juego', col_penj, 'penales_partido')]
    for lbl, col, bdkey in checks:
        if not col: continue
        ex = to_int(ws.cell(r, col).value)
        bv = b[bdkey]
        if ex != bv:
            diffs[lbl] += 1
            flag = ' <== VAR flag' if (lbl == 'VAR' and nf in FLAG) else ''
            print(f"P{nf:<5}{lbl:<12}{str(ex):>8}{str(bv):>8}{flag}")

print("\n" + "-" * 78)
print(f"Partidos comparados: {seen}")
print("Diferencias por campo:")
for k, v in diffs.items():
    print(f"   {k}: {v}")
print("\nNOTA: 'VAR' aca es el DATO OFICIAL del partido (decisiones de VAR), no la apuesta.")
print("Si el Excel y la BD difieren, hay que decidir cual es el oficial. Cambiar la BD")
print("moveria los puntos del item L de todos los apostadores en esos partidos.")
conn.close()
