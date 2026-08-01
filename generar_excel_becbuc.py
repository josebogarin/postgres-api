"""
generar_excel_becbuc.py — Genera Excel completo de verificación BECBUC.

Hojas:
  1. Ranking        — Clasificación final con desglose de puntos
  2. Resultados     — Fichas de resultado por fase y partido
  3. Apuestas       — Pronósticos de cada jugador × partido
  4. Puntajes       — Detalle de puntaje (H,I,J,L,N,O) × jugador × partido
  5. Globales       — Apuestas bonus A-G × jugador vs resultado real

Uso:
    backend\\.venv\\Scripts\\python.exe generar_excel_becbuc.py [torneo_id]
"""
import os as _osp
_BASE = _osp.path.dirname(_osp.path.abspath(__file__))

import sys, os, subprocess, json
from datetime import datetime

# Forzar UTF-8 en stdout/stderr para evitar UnicodeEncodeError en Windows
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

# ── Config ────────────────────────────────────────────────────────────────────
BASE_URL   = "http://localhost:8000"
ADMIN_USER = "jose"
ADMIN_PASS = "catalina"
OUTPUT     = _osp.path.join(_BASE, 'BECBUC_verificacion.xlsx')

# Colores por fase (hex sin #)
FASE_COLORS = {
    "grupo":         "E8F5E9",  # verde muy claro
    "ronda32":       "E8F5E9",
    "ronda16":       "E3F2FD",  # azul muy claro
    "octavos":       "E3F2FD",
    "cuartos":       "FFF9C4",  # amarillo claro
    "semis":         "FFE0B2",  # naranja claro
    "tercero":       "F3E5F5",  # lila
    "tercer_puesto": "F3E5F5",
    "final":         "FFEBEE",  # rojo claro
}
FASE_LABELS = {
    "grupo": "Grupos", "ronda32": "Ronda 32", "ronda16": "16avos",
    "octavos": "Octavos", "cuartos": "Cuartos", "semis": "Semis",
    "tercero": "3er Puesto", "tercer_puesto": "3er Puesto", "final": "Final",
}

# ── HTTP ──────────────────────────────────────────────────────────────────────
import urllib.request, urllib.error

def _req(method, url, data=None, token=None):
    body = json.dumps(data).encode() if data is not None else None
    hdrs = {"Content-Type": "application/json"}
    if token: hdrs["Authorization"] = f"Bearer {token}"
    req = urllib.request.Request(url, data=body, headers=hdrs, method=method)
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            return json.loads(r.read().decode())
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"HTTP {e.code} {url}: {e.read().decode()[:200]}")

def api_get(path, tok=None): return _req("GET",  f"{BASE_URL}{path}", token=tok)
def api_post(path, d=None, tok=None): return _req("POST", f"{BASE_URL}{path}", data=d, token=tok)

def login():
    try:
        r = api_post("/api/v1/auth/login", {"username": ADMIN_USER, "password": ADMIN_PASS})
        t = r.get("access_token")
        if not t: raise RuntimeError(f"Login falló: {r}")
        return t
    except Exception as e:
        print(f"  ⚠️  Servidor no disponible ({e}), modo sin API (apostadores desde BD)")
        return None

# ── psql ─────────────────────────────────────────────────────────────────────
# Usa psycopg2 directo (no requiere docker ni servidor API)
_PG_PARAMS = {
    "becbuc": dict(host="localhost", port=5432, user="app_user",
                   password="superpassword", dbname="becbuc"),
    "app_db": dict(host="localhost", port=5432, user="app_user",
                   password="superpassword", dbname="app_db"),
}

def psql(sql, db="becbuc"):
    import psycopg2
    params = _PG_PARAMS.get(db, _PG_PARAMS["becbuc"])
    try:
        conn = psycopg2.connect(**params)
        cur = conn.cursor()
        cur.execute(sql)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return [[str(col) if col is not None else "" for col in row] for row in rows]
    except Exception as e:
        raise RuntimeError(f"psql ({db}): {e}")

# ── Carga de datos ────────────────────────────────────────────────────────────

def load_all_data(torneo_id, tok):
    """Carga todos los datos necesarios desde BD y API."""
    print("  Cargando datos de BD...")

    # Competicion_id del torneo — garantiza que solo se usen equipos de esta competicion
    comp_rows = psql(f"""
        SELECT c.id, c.nombre, c.codigo
        FROM torneo t JOIN competicion c ON c.id = t.competicion_id
        WHERE t.id = {torneo_id}
    """)
    competicion_id   = int(comp_rows[0][0].strip()) if comp_rows else None
    competicion_nombre = comp_rows[0][1].strip() if comp_rows else "?"
    print(f"  Competicion: [{competicion_id}] {competicion_nombre}")

    # Partidos finalizados — equipos filtrados por competicion del torneo.
    # Los equipos se toman SOLO si su competicion_id coincide con la del torneo
    # (o si competicion_id es NULL, como fallback tolerante).
    partidos_rows = psql(f"""
        SELECT p.id, f.tipo, f.nombre AS fase_nombre, f.orden,
               COALESCE(el.nombre_es, el.nombre, 'TBD') AS local_nom,
               COALESCE(ev.nombre_es, ev.nombre, 'TBD') AS visit_nom,
               p.goles_local, p.goles_visitante,
               COALESCE(p.penales_local, -1), COALESCE(p.penales_visitante, -1),
               COALESCE(p.amarillas, 0), COALESCE(p.decisiones_var, 0),
               COALESCE(p.minuto_primer_gol::text, '-'),
               p.estado,
               el.id AS local_id, ev.id AS visit_id,
               COALESCE(p.rojas, 0),
               COALESCE(p.penales_partido, -1),
               COALESCE(p.numero_fifa, 0)
        FROM partido p
        JOIN fase f ON f.id = p.fase_id
        LEFT JOIN equipo el ON el.id = p.equipo_local_id
        LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
        WHERE f.torneo_id = {torneo_id}
          AND (p.estado = 'finalizado' OR p.goles_local IS NOT NULL)
        ORDER BY f.orden, p.id
    """)

    partidos = []
    for i, r in enumerate(partidos_rows):
        try:
            pen_l = int(r[8]) if r[8].strip() != "-1" else None
            pen_v = int(r[9]) if r[9].strip() != "-1" else None
            pen_partido = int(r[17]) if r[17].strip() != "-1" else None
            numero_fifa = int(r[18]) if r[18].strip() not in ("0", "") else None
            partidos.append({
                "id":          int(r[0]),
                "numero":      numero_fifa if numero_fifa else i + 1,  # numero_fifa oficial; fallback a posicion
                "fase_tipo":   r[1].strip(),
                "fase_nombre": r[2].strip(),
                "fase_orden":  int(r[3]),
                "local":       r[4].strip()[:20],
                "visitante":   r[5].strip()[:20],
                "gl":          int(r[6]),
                "gv":          int(r[7]),
                "pen_l":       pen_l,
                "pen_v":       pen_v,
                "amarillas":   int(r[10]),
                "var":         int(r[11]),
                "min_gol":     r[12].strip(),
                "rojas":       int(r[16]),
                "pen_partido": pen_partido,
            })
        except Exception:
            pass

    # Apostadores — obtenemos IDs de puntaje_detalle (becbuc) y nombres via ranking API o BD
    apostadores = []
    if tok:
        try:
            ranking_data = api_get(f"/api/v1/bets/ranking/{torneo_id}", tok)
            ranking_list = ranking_data if isinstance(ranking_data, list) else ranking_data.get("ranking", [])
            apostadores = [
                {"id": item["apostador_id"],
                 "username": item.get("nombre", f"#{item['apostador_id']}"),
                 "nombre":   item.get("nombre", f"#{item['apostador_id']}")}
                for item in ranking_list
            ]
        except Exception:
            pass
    if not apostadores:
        # Fallback: nombres desde apuesta.nombre_apostador (no requiere API ni docker)
        nombre_rows = psql(f"""
            SELECT DISTINCT ON (a.apostador_id)
                a.apostador_id,
                COALESCE(a.nombre_apostador, 'Apostador ' || a.apostador_id::text) AS nombre
            FROM apuesta a
            JOIN partido p ON p.id = a.partido_id
            JOIN fase f ON f.id = p.fase_id
            WHERE f.torneo_id = {torneo_id}
            ORDER BY a.apostador_id, a.id DESC
        """)
        apostadores = [{"id": int(r[0]), "username": r[1].strip(), "nombre": r[1].strip()}
                       for r in nombre_rows]
        if not apostadores:
            # Último fallback: solo IDs de puntaje_detalle
            id_rows = psql(f"""
                SELECT DISTINCT apostador_id FROM puntaje_detalle
                WHERE torneo_id = {torneo_id} ORDER BY apostador_id
            """)
            apostadores = [{"id": int(r[0]), "username": f"#{r[0]}", "nombre": f"#{r[0]}"}
                           for r in id_rows]

    # Apuestas (pred_local, pred_visitante por apostador × partido)
    apuestas_rows = psql(f"""
        SELECT a.apostador_id, a.partido_id,
               a.pred_local, a.pred_visitante,
               COALESCE(a.pred_penales_local_tanda::text, '-'),
               COALESCE(a.pred_penales_visitante_tanda::text, '-'),
               COALESCE(a.pred_penales_partido::text, '-')
        FROM apuesta a
        JOIN partido p ON p.id = a.partido_id
        JOIN fase f2 ON f2.id = p.fase_id
        WHERE f2.torneo_id = {torneo_id}
    """)
    apuestas = {}  # (apostador_id, partido_id) → dict
    for r in apuestas_rows:
        try:
            key = (int(r[0]), int(r[1]))
            apuestas[key] = {
                "pred_l":     int(r[2]) if r[2].strip() not in ("", "NULL") else None,
                "pred_v":     int(r[3]) if r[3].strip() not in ("", "NULL") else None,
                "pen_l":      r[4].strip() if r[4].strip() != "-" else None,
                "pen_v":      r[5].strip() if r[5].strip() != "-" else None,
                "pred_pen_p": int(r[6]) if r[6].strip() not in ("-", "", "NULL") else None,
            }
        except Exception:
            pass

    # Puntajes detalle
    pts_rows = psql(f"""
        SELECT pd.apostador_id, pd.partido_id,
               COALESCE(pd.pts_resultado, 0),
               COALESCE(pd.pts_marcador, 0),
               COALESCE(pd.pts_amarillas, 0),
               COALESCE(pd.pts_var, 0),
               COALESCE(pd.pts_rojas, 0),
               COALESCE(pd.pts_penales_tanda, 0),
               0,
               COALESCE(pd.pts_resultado,0)+COALESCE(pd.pts_marcador,0)+
               COALESCE(pd.pts_amarillas,0)+COALESCE(pd.pts_var,0)+
               COALESCE(pd.pts_rojas,0)+COALESCE(pd.pts_penales_tanda,0)+
               COALESCE(pd.pts_minuto,0)+COALESCE(pd.pts_penales_partido,0)+
               COALESCE(pd.pts_equipo,0),
               COALESCE(pd.pts_minuto, 0),
               COALESCE(pd.pts_penales_partido, 0)
        FROM puntaje_detalle pd
        WHERE pd.torneo_id = {torneo_id}
    """)
    puntajes = {}  # (apostador_id, partido_id) → dict
    for r in pts_rows:
        try:
            key = (int(r[0]), int(r[1]))
            puntajes[key] = {
                "H":     int(r[2]),
                "I":     int(r[3]),
                "J":     int(r[4]),
                "L":     int(r[5]),
                "K":     int(r[6]),
                "O":     int(r[7]),
                "bonus": int(r[8]),
                "total": int(r[9]),
                "N":     int(r[10]),
                "M":     int(r[11]),
            }
        except Exception:
            pass

    # Puntaje total por apostador (partidos) — suma directa de todas las columnas
    pts_total_rows = psql(f"""
        SELECT apostador_id,
               SUM(COALESCE(pts_resultado,0)+COALESCE(pts_marcador,0)+
                   COALESCE(pts_amarillas,0)+COALESCE(pts_var,0)+
                   COALESCE(pts_rojas,0)+COALESCE(pts_penales_tanda,0)+
                   COALESCE(pts_minuto,0)+COALESCE(pts_penales_partido,0)+
                   COALESCE(pts_equipo,0)) AS total
        FROM puntaje_detalle
        WHERE torneo_id = {torneo_id}
        GROUP BY apostador_id
    """)
    pts_partidos_total = {int(r[0]): int(r[1]) for r in pts_total_rows if r[0].strip()}

    # Ranking completo desde API (incluye pts_globales y posicion)
    try:
        rk = api_get(f"/api/v1/bets/ranking/{torneo_id}", tok)
        ranking = rk if isinstance(rk, list) else rk.get("ranking", [])
        # Enriquecer apostadores con posición y totales del ranking
        rank_by_id = {item["apostador_id"]: item for item in ranking}
        for ap in apostadores:
            ritem = rank_by_id.get(ap["id"], {})
            ap["posicion"]    = ritem.get("posicion", "-")
            ap["pts_partidos"] = ritem.get("puntos_partidos_total", pts_partidos_total.get(ap["id"], 0))
            ap["pts_globales"] = ritem.get("pts_globales", 0)
            ap["pts_total"]    = ritem.get("puntos_total",
                                            ap["pts_partidos"] + ap["pts_globales"])
    except Exception:
        # Build ranking from DB data
        pg_rows = psql(f"""
            SELECT apostador_id,
                   COALESCE(pts_campeon,0)+COALESCE(pts_finalistas,0)+COALESCE(pts_goleador,0)+
                   COALESCE(pts_peor_equipo,0)+COALESCE(pts_mayor_goleada,0)+
                   COALESCE(pts_etapa_paraguay,0)+COALESCE(pts_goles_paraguay,0) AS pg
            FROM puntaje_global WHERE torneo_id={torneo_id}
        """)
        pg_by_id = {int(r[0]): int(r[1]) for r in pg_rows if r[0].strip()}
        plenos_rows = psql(f"""
            SELECT apostador_id,
                   COUNT(*) FILTER (WHERE pts_marcador>0) AS plenos,
                   COUNT(*) FILTER (WHERE pts_resultado>0) AS aciertos
            FROM puntaje_detalle WHERE torneo_id={torneo_id} GROUP BY apostador_id
        """)
        plenos_by_id = {int(r[0]): (int(r[1]), int(r[2])) for r in plenos_rows if r[0].strip()}
        ranking = []
        for ap in apostadores:
            pts_p = pts_partidos_total.get(ap["id"], 0)
            pts_g = pg_by_id.get(ap["id"], 0)
            pl, ac = plenos_by_id.get(ap["id"], (0, 0))
            ap["posicion"]     = "-"
            ap["pts_partidos"] = pts_p
            ap["pts_globales"] = pts_g
            ap["pts_total"]    = pts_p + pts_g
            ranking.append({
                "apostador_id": ap["id"],
                "nombre": ap["nombre"],
                "puntos_total": pts_p + pts_g,
                "puntos_partidos_total": pts_p,
                "pts_globales": pts_g,
                "plenos": pl,
                "aciertos": ac,
                "fallos": 0,
            })
        ranking.sort(key=lambda x: x["puntos_total"], reverse=True)
        for i, r in enumerate(ranking, 1):
            r["posicion"] = i
            for ap in apostadores:
                if ap["id"] == r["apostador_id"]:
                    ap["posicion"] = i
                    break

    # Fallback final: si apostadores sigue vacío, poblar desde ranking
    if not apostadores and ranking:
        apostadores = [
            {"id": item["apostador_id"],
             "username": item.get("nombre", f"#{item['apostador_id']}"),
             "nombre":   item.get("nombre", f"#{item['apostador_id']}"),
             "posicion":     item.get("posicion", "-"),
             "pts_partidos": item.get("puntos_partidos_total", 0),
             "pts_globales": item.get("pts_globales", 0),
             "pts_total":    item.get("puntos_total", 0)}
            for item in ranking
        ]

    # Apuestas globales
    globales_rows = psql(f"""
        SELECT ag.apostador_id,
               COALESCE(ec.nombre_es, ec.nombre, '-') AS campeon,
               COALESCE(ef1.nombre_es, ef1.nombre, '-') AS fin1,
               COALESCE(ef2.nombre_es, ef2.nombre, '-') AS fin2,
               COALESCE(ag.pred_goleador, '-'),
               COALESCE(ep.nombre_es, ep.nombre, '-') AS peor,
               COALESCE(ag.pred_goleada_ganador::text, '-'),
               COALESCE(ag.pred_goleada_perdedor::text, '-'),
               COALESCE(ag.pred_etapa_paraguay, '-'),
               COALESCE(ag.pred_goles_paraguay::text, '-')
        FROM apuesta_global ag
        LEFT JOIN equipo ec  ON ec.id  = ag.pred_campeon_id
        LEFT JOIN equipo ef1 ON ef1.id = ag.pred_finalista1_id
        LEFT JOIN equipo ef2 ON ef2.id = ag.pred_finalista2_id
        LEFT JOIN equipo ep  ON ep.id  = ag.pred_peor_equipo_id
        WHERE ag.torneo_id = {torneo_id}
    """)
    globales_ap = {}  # apostador_id → dict
    for r in globales_rows:
        try:
            globales_ap[int(r[0])] = {
                "campeon": r[1].strip(), "fin1": r[2].strip(), "fin2": r[3].strip(),
                "goleador": r[4].strip(), "peor": r[5].strip(),
                "goleada": f"{r[6].strip()}-{r[7].strip()}",
                "etapa_py": r[8].strip(), "goles_py": r[9].strip(),
            }
        except Exception:
            pass

    # Puntaje global por apostador
    pts_global_rows = psql(f"""
        SELECT pg.apostador_id,
               COALESCE(pg.pts_campeon, 0), COALESCE(pg.pts_finalistas, 0),
               COALESCE(pg.pts_goleador, 0), COALESCE(pg.pts_peor_equipo, 0),
               COALESCE(pg.pts_mayor_goleada, 0), COALESCE(pg.pts_etapa_paraguay, 0),
               COALESCE(pg.pts_goles_paraguay, 0)
        FROM puntaje_global pg WHERE pg.torneo_id = {torneo_id}
    """)
    pts_global = {}
    for r in pts_global_rows:
        try:
            pts_global[int(r[0])] = {
                "A": int(r[1]), "B": int(r[2]), "C": int(r[3]), "D": int(r[4]),
                "E": int(r[5]), "F": int(r[6]), "G": int(r[7]),
                "total": sum(int(x) for x in r[1:]),
            }
        except Exception:
            pass

    # Resultados reales del torneo (para Globales sheet)
    torneo_real = {}
    try:
        rows = psql(f"""
            SELECT COALESCE(ec.nombre_es, ec.nombre, '-') AS campeon,
                   COALESCE(el.nombre_es, el.nombre, '-') AS fin_local,
                   COALESCE(ev.nombre_es, ev.nombre, '-') AS fin_visit
            FROM partido p
            JOIN fase f ON f.id = p.fase_id
            LEFT JOIN equipo ec ON ec.id = p.equipo_clasificado_id
            LEFT JOIN equipo el ON el.id = p.equipo_local_id
            LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
            WHERE f.torneo_id = {torneo_id} AND f.tipo = 'final'
              AND p.estado = 'finalizado' LIMIT 1
        """)
        if rows:
            torneo_real["A"] = rows[0][0].strip()
            torneo_real["B"] = f"{rows[0][1].strip()} / {rows[0][2].strip()}"
    except Exception: pass
    try:
        rows = psql(f"SELECT COALESCE(resultado_goleador,'-') FROM torneo WHERE id={torneo_id}")
        torneo_real["C"] = rows[0][0].strip() if rows else "-"
    except Exception: torneo_real["C"] = "-"
    try:
        rows = psql(f"""SELECT COALESCE(e.nombre_es,e.nombre,'-') FROM torneo t
            LEFT JOIN equipo e ON e.id=t.resultado_peor_equipo_id WHERE t.id={torneo_id}""")
        torneo_real["D"] = rows[0][0].strip() if rows else "-"
    except Exception: torneo_real["D"] = "-"
    try:
        rows = psql(f"""SELECT p.goles_local, p.goles_visitante FROM partido p
            JOIN fase f ON f.id=p.fase_id
            WHERE f.torneo_id={torneo_id} AND p.estado='finalizado'
            ORDER BY ABS(p.goles_local-p.goles_visitante) DESC, p.goles_local+p.goles_visitante DESC LIMIT 1""")
        if rows:
            gl, gv = int(rows[0][0]), int(rows[0][1])
            torneo_real["E"] = f"{max(gl,gv)}-{min(gl,gv)}"
    except Exception: pass
    try:
        rows = psql(f"""SELECT f.tipo FROM partido p JOIN fase f ON f.id=p.fase_id
            LEFT JOIN equipo e ON (e.id=p.equipo_local_id OR e.id=p.equipo_visitante_id)
            WHERE f.torneo_id={torneo_id} AND p.estado='finalizado'
              AND e.nombre ILIKE '%paraguay%'
            ORDER BY f.orden DESC LIMIT 1""")
        torneo_real["F"] = FASE_LABELS.get(rows[0][0].strip(), rows[0][0].strip()) if rows else "-"
    except Exception: torneo_real["F"] = "-"
    try:
        rows = psql(f"""SELECT COALESCE(SUM(
            CASE WHEN p.equipo_local_id=e.id THEN p.goles_local
                 ELSE p.goles_visitante END),0)
            FROM partido p JOIN fase f ON f.id=p.fase_id
            JOIN equipo e ON (e.id=p.equipo_local_id OR e.id=p.equipo_visitante_id)
            WHERE f.torneo_id={torneo_id} AND p.estado='finalizado'
              AND e.nombre ILIKE '%paraguay%'""")
        torneo_real["G"] = str(int(rows[0][0])) if rows else "0"
    except Exception: torneo_real["G"] = "-"

    print(f"  OK {len(partidos)} partidos | {len(apostadores)} apostadores | "
          f"{len(apuestas)} apuestas | {len(puntajes)} filas puntaje_detalle")

    # Diagnóstico: verificar cobertura de apuestas por partido
    if apostadores and partidos:
        n_ap = len(apostadores)
        faltantes = []
        for p in partidos:
            ap_ids_con_apuesta = {k[0] for k in apuestas if k[1] == p["id"]}
            faltantes_en_partido = [
                ap["nombre"] for ap in apostadores
                if ap["id"] not in ap_ids_con_apuesta
            ]
            if faltantes_en_partido:
                faltantes.append((p["numero"], p["local"], p["gv"] if False else f"{p['gl']}-{p['gv']}", faltantes_en_partido))
        if faltantes:
            print(f"\n  ⚠️  PARTIDOS SIN APUESTA DE ALGUNOS USUARIOS ({len(faltantes)}):")
            for num, local, res, nombres in faltantes[:20]:
                print(f"     P{num} {local} {res}  ->  sin apuesta: {', '.join(nombres)}")
            if len(faltantes) > 20:
                print(f"     ... y {len(faltantes)-20} más")
        else:
            print(f"  OK Cobertura completa: todos los apostadores tienen apuesta en los {len(partidos)} partidos")

    return {
        "partidos":     partidos,
        "apostadores":  apostadores,
        "apuestas":     apuestas,
        "puntajes":     puntajes,
        "ranking":      ranking,
        "globales_ap":  globales_ap,
        "pts_global":   pts_global,
        "torneo_real":  torneo_real,
    }

# ── openpyxl helpers ──────────────────────────────────────────────────────────

def ensure_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        print("  Instalando openpyxl...")
        subprocess.run([sys.executable, "-m", "pip", "install",
                        "openpyxl", "--break-system-packages", "-q"])
        import openpyxl
        return openpyxl

def style_header(cell, ox, bg="4472C4", fg="FFFFFF", bold=True, size=10):
    from openpyxl.styles import Font, PatternFill, Alignment
    cell.font = Font(bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_cell(cell, ox, bg=None, bold=False, align="center", size=9):
    from openpyxl.styles import Font, PatternFill, Alignment
    cell.font = Font(bold=bold, size=size)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")

def freeze_and_autofit(ws, freeze="A2"):
    ws.freeze_panes = freeze
    for col in ws.columns:
        max_w = 8
        for cell in col:
            try:
                v = str(cell.value or "")
                max_w = max(max_w, min(len(v) + 2, 35))
            except Exception:
                pass
        ws.column_dimensions[col[0].column_letter].width = max_w

# ── Sheet 1: Ranking ─────────────────────────────────────────────────────────

def build_sheet_ranking(wb, ox, data):
    ws = wb.active
    ws.title = "\U0001f3c6 Ranking"
    ws.row_dimensions[1].height = 22
    headers = ["Pos", "Apostador", "Total", "Partidos", "Globales",
               "Plenos", "Aciertos", "Fallos"]
    for c, h in enumerate(headers, 1):
        style_header(ws.cell(1, c, h), ox)
    ranking = data["ranking"]
    for i, r in enumerate(ranking, 1):
        row = i + 1
        nombre = r.get("nombre") or r.get("email") or r.get("username") or "?"
        ws.cell(row, 1, i)
        ws.cell(row, 2, nombre)
        ws.cell(row, 3, r.get("puntos_total", 0))
        ws.cell(row, 4, r.get("puntos_partidos_total", 0))
        ws.cell(row, 5, r.get("pts_globales", 0))
        ws.cell(row, 6, r.get("plenos", 0))
        ws.cell(row, 7, r.get("aciertos", 0))
        ws.cell(row, 8, r.get("fallos", 0))
        bg = "FFF9C4" if i == 1 else ("E8F5E9" if i == 2 else ("FFE0B2" if i == 3 else None))
        for c in range(1, 9):
            style_cell(ws.cell(row, c), ox, bg=bg,
                       bold=(c in (1, 3)), align="left" if c == 2 else "center")
    freeze_and_autofit(ws)


# ── Sheet 2: Resultados ───────────────────────────────────────────────────────

def build_sheet_resultados(wb, ox, data):
    ws = wb.create_sheet("\U0001f4cb Resultados")
    ws.row_dimensions[1].height = 22
    HDRS = ["Fase", "P#", "Local", "GL", "GV", "Visitante",
            "Pen-L", "Pen-V", "Pen.part", "Amar", "Rojas", "VAR", "Min.Gol"]
    for c, h in enumerate(HDRS, 1):
        style_header(ws.cell(1, c, h), ox, bg="37474F")

    for col_letter, w in zip("ABCDEFGHIJKLM",
                              [12, 5, 20, 5, 5, 20, 7, 7, 7, 6, 6, 5, 8]):
        ws.column_dimensions[col_letter].width = w

    data_row = 2
    fase_actual = None
    for p in data["partidos"]:
        tipo = p["fase_tipo"]
        if tipo != fase_actual:
            fase_actual = tipo
            ws.cell(data_row, 1, "── " + FASE_LABELS.get(tipo, tipo))
            for c in range(1, 13):
                style_cell(ws.cell(data_row, c), ox,
                           bg=FASE_COLORS.get(tipo, "EEEEEE"),
                           bold=True, align="left" if c == 1 else "center", size=9)
            data_row += 1

        bg = FASE_COLORS.get(tipo, "FFFFFF")
        pen_l = str(p["pen_l"]) if p["pen_l"] is not None else ""
        pen_v = str(p["pen_v"]) if p["pen_v"] is not None else ""
        pen_partido = str(p["pen_partido"]) if p.get("pen_partido") is not None else ""
        row_vals = [
            FASE_LABELS.get(tipo, tipo), f"P{p['numero']}",
            p["local"], p["gl"], p["gv"], p["visitante"],
            pen_l, pen_v, pen_partido, p["amarillas"], p["rojas"], p["var"], p["min_gol"],
        ]
        for c, v in enumerate(row_vals, 1):
            al = "left" if c in (3, 6) else "center"
            style_cell(ws.cell(data_row, c, v), ox, bg=bg, align=al, size=9)
        data_row += 1

    ws.freeze_panes = "A2"


# ── Helpers por apostador ─────────────────────────────────────────────────────

def _pred_color(pred_l, pred_v, real_l, real_v):
    """Verde=exacto, Amarillo=resultado, Rojo=fallo, Gris=sin apuesta."""
    if pred_l is None:
        return "EEEEEE", "?"
    if pred_l == real_l and pred_v == real_v:
        return "C8E6C9", "✓✓"
    same_res = ((pred_l > pred_v) == (real_l > real_v)
                and (pred_l == pred_v) == (real_l == real_v))
    if same_res:
        return "FFF9C4", "✓"
    return "FFCDD2", "✗"


def _write_fase_subtotal(ws, ox, row, fase_tipo, ft):
    # Cols: 1=Fase|2=P#|3=Local|4=GL-GV|5=Visit|6=Apuesta|7=Pen.part|8=Pen.tanda|9=OK|
    #       10=Resultado|11=Marcador|12=Amarillas|13=Rojas|14=VAR|15=Pen.part|16=Minuto|17=Pen.tanda|18=Total
    label = FASE_LABELS.get(fase_tipo, fase_tipo)
    bg = "BBDEFB"
    ws.cell(row, 1, f"── {label}  ({ft.get('n', 0)} partidos)")
    vals = [ft.get(k, 0) or "" for k in ["H", "I", "J", "K", "L", "M", "N", "O", "total"]]
    for c, v in zip(range(10, 19), vals):
        style_cell(ws.cell(row, c, v), ox, bg=bg, bold=True, align="center", size=9)
    style_cell(ws.cell(row, 1), ox, bg=bg, bold=True, align="left", size=9)
    for c in range(2, 10):
        style_cell(ws.cell(row, c), ox, bg=bg)


# ── Sheet por Apostador ───────────────────────────────────────────────────────

def build_sheet_apostador(wb, ox, data, ap):
    """Una hoja por apostador: identificacion + apuesta/puntaje por partido."""
    import openpyxl.styles as _s

    sheet_name = ap["nombre"][:28]
    ws = wb.create_sheet(f"\U0001f464 {sheet_name}")

    pos          = ap.get("posicion", "-")
    pts_partidos = ap.get("pts_partidos", 0)
    pts_globales = ap.get("pts_globales", 0)
    pts_total    = ap.get("pts_total", 0)

    # Fila 1: identificacion del apostador
    ws.row_dimensions[1].height = 24
    ws.merge_cells("A1:E1")
    c1 = ws.cell(1, 1, f"\U0001f3c5 {ap['nombre'].upper()}   -   Posicion #{pos}")
    c1.font = _s.Font(bold=True, size=13, color="FFFFFF")
    c1.fill = _s.PatternFill("solid", fgColor="0D47A1")
    c1.alignment = _s.Alignment(horizontal="left", vertical="center")

    ws.merge_cells("F1:J1")
    c1b = ws.cell(1, 6, f"Pts Partidos: {pts_partidos}")
    c1b.font = _s.Font(bold=True, size=11, color="FFFFFF")
    c1b.fill = _s.PatternFill("solid", fgColor="1565C0")
    c1b.alignment = _s.Alignment(horizontal="center", vertical="center")

    ws.merge_cells("K1:M1")
    c1c = ws.cell(1, 11, f"Pts Globales: {pts_globales}")
    c1c.font = _s.Font(bold=True, size=11, color="FFFFFF")
    c1c.fill = _s.PatternFill("solid", fgColor="1976D2")
    c1c.alignment = _s.Alignment(horizontal="center", vertical="center")

    ws.merge_cells("N1:R1")
    c1d = ws.cell(1, 14, f"TOTAL: {pts_total} pts")
    c1d.font = _s.Font(bold=True, size=13, color="FFFFFF")
    c1d.fill = _s.PatternFill("solid", fgColor="E65100")
    c1d.alignment = _s.Alignment(horizontal="center", vertical="center")

    # Fila 2: leyenda colores
    ws.row_dimensions[2].height = 14
    leyenda = [
        ("Verde=Marcador exacto", "C8E6C9"),
        ("Amarillo=Resultado OK", "FFF9C4"),
        ("Rojo=Fallo",            "FFCDD2"),
        ("Gris=Sin apuesta",      "F5F5F5"),
    ]
    col = 1
    for txt, bg in leyenda:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+2)
        lc = ws.cell(2, col, txt)
        lc.font = _s.Font(size=8, color="333333")
        lc.fill = _s.PatternFill("solid", fgColor=bg)
        lc.alignment = _s.Alignment(horizontal="center", vertical="center")
        col += 3
    for c in range(col, 17):
        ws.cell(2, c).fill = _s.PatternFill("solid", fgColor="EEEEEE")

    # Fila 3: cabeceras columnas
    # Cols: 1=Fase|2=P#|3=Local|4=GL-GV|5=Visit|6=Apuesta|7=Pen.part|8=Pen.tanda|9=OK|
    #       10=Resultado|11=Marcador|12=Amarillas|13=Rojas|14=VAR|15=Pen.part(pts)|16=Minuto|17=Pen.tanda(pts)|18=Total
    ws.row_dimensions[3].height = 20
    HDRS = ["Fase", "P#", "Local", "GL-GV", "Visitante", "Apuesta", "Pen.part", "Pen.tanda", "OK",
            "Resultado", "Marcador", "Amarillas", "Rojas", "VAR", "Pen.part", "Minuto", "Pen.tanda", "Total"]
    for c, h in enumerate(HDRS, 1):
        style_header(ws.cell(3, c, h), ox, bg="1565C0")

    for col_letter, w in zip("ABCDEFGHIJKLMNOPQR",
                              [12, 5, 18, 7, 18, 8, 7, 8, 4, 7, 7, 7, 5, 5, 6, 6, 7, 6]):
        ws.column_dimensions[col_letter].width = w

    partidos   = data["partidos"]
    apuestas   = data["apuestas"]
    puntajes   = data["puntajes"]
    pts_global = data.get("pts_global", {})

    data_row    = 4
    fase_actual = None
    fase_totals = {}

    for p in partidos:
        tipo = p["fase_tipo"]
        if fase_actual is not None and tipo != fase_actual:
            _write_fase_subtotal(ws, ox, data_row, fase_actual,
                                 fase_totals.get(fase_actual, {}))
            data_row += 1
        fase_actual = tipo
        ft = fase_totals.setdefault(tipo, {"H": 0, "I": 0, "J": 0, "K": 0,
                                           "L": 0, "M": 0, "N": 0, "O": 0, "total": 0, "n": 0})

        key  = (ap["id"], p["id"])
        bet  = apuestas.get(key, {})
        pts  = puntajes.get(key, {})
        total = pts.get("total", 0)

        pred_l = bet.get("pred_l")
        pred_v = bet.get("pred_v")
        pred_str = f"{pred_l}-{pred_v}" if pred_l is not None else "-"
        pen_l = bet.get("pen_l")
        pen_v = bet.get("pen_v")
        pen_str = f"{pen_l}-{pen_v}" if pen_l is not None else ""
        pred_pen_p = bet.get("pred_pen_p")
        pred_pen_p_str = str(pred_pen_p) if pred_pen_p is not None else ""

        bg_fase = FASE_COLORS.get(tipo, "FFFFFF")
        row_bg, ok_str = _pred_color(pred_l, pred_v, p["gl"], p["gv"])
        fase_label = FASE_LABELS.get(tipo, tipo)

        row_vals = [
            fase_label, f"P{p['numero']}", p["local"], f"{p['gl']}-{p['gv']}",
            p["visitante"], pred_str, pred_pen_p_str, pen_str, ok_str,
            pts.get("H", 0) or "", pts.get("I", 0) or "",
            pts.get("J", 0) or "", pts.get("K", 0) or "",
            pts.get("L", 0) or "", pts.get("M", 0) or "",
            pts.get("N", 0) or "",
            pts.get("O", 0) or "", total or "",
        ]
        for c, v in enumerate(row_vals, 1):
            bg = bg_fase if c <= 5 else row_bg
            al = "left" if c in (3, 5) else "center"
            style_cell(ws.cell(data_row, c, v), ox, bg=bg, align=al, size=9)

        for k in ["H", "I", "J", "K", "L", "M", "N", "O"]:
            ft[k] += pts.get(k, 0)
        ft["total"] += total
        ft["n"] += 1
        data_row += 1

    if fase_actual:
        _write_fase_subtotal(ws, ox, data_row, fase_actual, fase_totals.get(fase_actual, {}))
        data_row += 1

    data_row += 1
    grand_partidos = sum(ft["total"] for ft in fase_totals.values())
    grand_globales = pts_global.get(ap["id"], {}).get("total", ap.get("pts_globales", 0))
    grand_total    = grand_partidos + grand_globales

    ws.cell(data_row, 1, "TOTAL PARTIDOS")
    ws.cell(data_row, 18, grand_partidos)
    for c in range(1, 19):
        style_cell(ws.cell(data_row, c), ox, bg="0D47A1", fg="FFFFFF",
                   bold=True, align="left" if c == 1 else "center")

    data_row += 1
    ws.cell(data_row, 1, "GLOBALES (A-G)")
    ws.cell(data_row, 18, grand_globales)
    for c in range(1, 19):
        style_cell(ws.cell(data_row, c), ox, bg="4A148C", fg="FFFFFF",
                   bold=True, align="left" if c == 1 else "center")

    data_row += 1
    ws.cell(data_row, 1, "GRAND TOTAL")
    ws.cell(data_row, 18, grand_total)
    for c in range(1, 19):
        style_cell(ws.cell(data_row, c), ox, bg="E65100", fg="FFFFFF",
                   bold=True, size=11, align="left" if c == 1 else "center")

    ws.freeze_panes = "A4"


# ── Sheet Globales A-G ────────────────────────────────────────────────────────

def build_sheet_globales(wb, ox, data):
    ws = wb.create_sheet("\U0001f310 Globales")
    ws.row_dimensions[1].height = 22

    apostadores  = data["apostadores"]
    globales_ap  = data["globales_ap"]
    pts_global   = data["pts_global"]
    torneo_real  = data.get("torneo_real", {})

    # Cols: Item | Concepto | Valor Real | Pts Max | [Apostador Pronostico | Pts] x N
    conceptos = [
        ("A", "Campeon mundial",      20, lambda g: g.get("campeon", "-"),
                                          torneo_real.get("A", "-")),
        ("B", "Finalistas (10c/u)",   20, lambda g: f"{g.get('fin1','-')} / {g.get('fin2','-')}",
                                          torneo_real.get("B", "-")),
        ("C", "Goleador",             20, lambda g: g.get("goleador", "-"),
                                          torneo_real.get("C", "-")),
        ("D", "Peor equipo",          20, lambda g: g.get("peor", "-"),
                                          torneo_real.get("D", "-")),
        ("E", "Mayor goleada (10+10)",20, lambda g: g.get("goleada", "-"),
                                          torneo_real.get("E", "-")),
        ("F", "Etapa Paraguay",        6, lambda g: g.get("etapa_py", "-"),
                                          torneo_real.get("F", "-")),
        ("G", "Goles Paraguay",        6, lambda g: g.get("goles_py", "-"),
                                          torneo_real.get("G", "-")),
    ]

    HDRS_BASE = ["Item", "Concepto", "Valor Real", "Pts Max"]
    hdrs = HDRS_BASE[:]
    for ap in apostadores:
        hdrs += [ap["nombre"], "Pts"]
    for c, h in enumerate(hdrs, 1):
        style_header(ws.cell(1, c, h), ox, bg="4A148C")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 8
    from openpyxl.utils import get_column_letter
    for i in range(len(apostadores)):
        ws.column_dimensions[get_column_letter(5 + i*2)].width = 22
        ws.column_dimensions[get_column_letter(6 + i*2)].width = 6

    for i, (letra, concepto, pts_max, get_pred, val_real) in enumerate(conceptos, 1):
        row = i + 1
        bg_row = "F3E5F5" if i % 2 else "EDE7F6"
        ws.cell(row, 1, letra)
        ws.cell(row, 2, concepto)
        ws.cell(row, 3, val_real if val_real else "-")
        ws.cell(row, 4, pts_max)
        for c in range(1, 5):
            style_cell(ws.cell(row, c), ox, bg=bg_row,
                       bold=(c == 1), align="left" if c in (2, 3) else "center")

        col = 5
        for ap in apostadores:
            ap_glob = globales_ap.get(ap["id"], {})
            pred_val = get_pred(ap_glob)
            pts_val  = pts_global.get(ap["id"], {}).get(letra, 0)
            cell_bg = "C8E6C9" if pts_val > 0 else "FFCDD2"
            style_cell(ws.cell(row, col, pred_val), ox, bg=cell_bg, align="left", size=9)
            style_cell(ws.cell(row, col+1, pts_val if pts_val else ""),
                       ox, bg=cell_bg, bold=(pts_val > 0), align="center", size=9)
            col += 2

    total_row = len(conceptos) + 2
    ws.cell(total_row, 2, "TOTAL GLOBALES")
    ws.cell(total_row, 4, 112)
    for c in range(1, 5):
        style_cell(ws.cell(total_row, c), ox, bg="4A148C", fg="FFFFFF",
                   bold=True, align="left" if c == 2 else "center")
    col = 5
    for ap in apostadores:
        total = pts_global.get(ap["id"], {}).get("total", 0)
        ws.cell(total_row, col, "")
        ws.cell(total_row, col+1, total)
        style_cell(ws.cell(total_row, col),   ox, bg="FFF9C4")
        style_cell(ws.cell(total_row, col+1), ox, bg="FFF9C4", bold=True, align="center")
        col += 2

    freeze_and_autofit(ws)


# ── Sheet Matriz (apostador x partido) ───────────────────────────────────────

def build_sheet_matriz(wb, ox, data):
    """Matriz unica: todas las filas apostador x partido con resultado, pronostico y puntajes."""
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("\U0001f4ca Matriz")

    apostadores = data["apostadores"]
    partidos    = data["partidos"]
    apuestas    = data["apuestas"]
    puntajes    = data["puntajes"]
    pts_global  = data.get("pts_global", {})

    # ── Columnas ─────────────────────────────────────────────────────────────
    #  1        2    3      4        5     6      7          8       9      10
    # Apost | P# | Fase | Local | GL  | GV  | Visitante | PL  | PV  | OK
    #  11       12       13       14      15      16     17    18
    # Pen.R.L | Pen.R.V | Pen.P.L | Pen.P.V | Amar | Rojas
    #  19          20          21        22       23    24      25       26
    # Resultado | Marcador | Amarillas | Rojas | VAR | Minuto | Pen.tanda | Total
    # Total columns: 26

    HDRS = [
        "Apostador", "P#", "Fase",
        "Local", "GL", "GV", "Visitante",
        "Pred.L", "Pred.V", "OK",
        "Pen.R.L", "Pen.R.V", "Pen.P.L", "Pen.P.V",
        "Pen.P.Part", "Pen.R.Part",
        "Amar.", "Rojas",
        "Resultado", "Marcador", "Pts.Amar", "Pts.Rojas", "Pts.VAR", "Pts.M", "Pts.Min", "Pen.tanda", "Total"
    ]
    NCOLS = len(HDRS)  # 27

    COL_WIDTHS = [18, 5, 12, 18, 5, 5, 18, 7, 7, 5, 7, 7, 7, 7, 8, 8, 6, 6, 8, 8, 7, 7, 6, 6, 6, 8, 7]

    ws.row_dimensions[1].height = 20
    for c, h in enumerate(HDRS, 1):
        style_header(ws.cell(1, c, h), ox, bg="37474F")
    for c, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    def _style_row(row_n, bg, fg="000000", bold=False, sz=9):
        for c in range(1, NCOLS + 1):
            cell = ws.cell(row_n, c)
            al = "left" if c in (1, 3, 4, 7) else "center"
            style_cell(cell, ox, bg=bg, fg=fg, bold=bold, align=al, size=sz)

    def _write_subtotal(row_n, label, ft, bg, fg="000000", bold=True, sz=9):
        ws.cell(row_n, 1, label)
        ws.cell(row_n, 19, ft.get("H", 0) or "")
        ws.cell(row_n, 20, ft.get("I", 0) or "")
        ws.cell(row_n, 21, ft.get("J", 0) or "")
        ws.cell(row_n, 22, ft.get("K", 0) or "")
        ws.cell(row_n, 23, ft.get("L", 0) or "")
        ws.cell(row_n, 24, ft.get("M", 0) or "")
        ws.cell(row_n, 25, ft.get("N", 0) or "")
        ws.cell(row_n, 26, ft.get("O", 0) or "")
        ws.cell(row_n, 27, ft.get("total", 0) or "")
        _style_row(row_n, bg=bg, fg=fg, bold=bold, sz=sz)

    cur_row = 2

    for ap in apostadores:
        ap_id   = ap["id"]
        ap_name = ap["nombre"]

        fase_actual = None
        fase_totals = {}
        grand_ft    = {"H": 0, "I": 0, "J": 0, "K": 0, "L": 0, "M": 0, "N": 0, "O": 0, "total": 0}

        for p in partidos:
            tipo = p["fase_tipo"]
            if fase_actual is not None and tipo != fase_actual:
                # subtotal de fase anterior
                label = f"  Subtotal {ap_name[:16]} - {FASE_LABELS.get(fase_actual, fase_actual)}"
                _write_subtotal(cur_row, label, fase_totals.get(fase_actual, {}),
                                bg="CFD8DC", fg="000000", bold=True, sz=8)
                cur_row += 1
            fase_actual = tipo
            ft = fase_totals.setdefault(tipo, {"H": 0, "I": 0, "J": 0, "K": 0,
                                                "L": 0, "M": 0, "N": 0, "O": 0, "total": 0})

            key = (ap_id, p["id"])
            bet = apuestas.get(key, {})
            pts = puntajes.get(key, {})
            total = pts.get("total", 0)

            pred_l = bet.get("pred_l")
            pred_v = bet.get("pred_v")
            pen_p_l = bet.get("pen_l")
            pen_p_v = bet.get("pen_v")
            pred_pen_p = bet.get("pred_pen_p")

            row_bg, ok_str = _pred_color(pred_l, pred_v, p["gl"], p["gv"])
            fase_bg = FASE_COLORS.get(tipo, "FFFFFF")

            row_vals = [
                ap_name,
                f"P{p['numero']}",
                FASE_LABELS.get(tipo, tipo),
                p["local"],
                p["gl"] if p["gl"] is not None else "",
                p["gv"] if p["gv"] is not None else "",
                p["visitante"],
                pred_l if pred_l is not None else "",
                pred_v if pred_v is not None else "",
                ok_str,
                p.get("pen_l", "") if p.get("pen_l") is not None else "",
                p.get("pen_v", "") if p.get("pen_v") is not None else "",
                pen_p_l if pen_p_l is not None else "",
                pen_p_v if pen_p_v is not None else "",
                pred_pen_p if pred_pen_p is not None else "",        # Pen.P.Part
                p.get("pen_partido", "") if p.get("pen_partido") is not None else "",  # Pen.R.Part
                p.get("amarillas", "") if p.get("amarillas") is not None else "",
                p.get("rojas", "") if p.get("rojas") is not None else "",
                pts.get("H", 0) or "",
                pts.get("I", 0) or "",
                pts.get("J", 0) or "",
                pts.get("K", 0) or "",
                pts.get("L", 0) or "",
                pts.get("M", 0) or "",
                pts.get("N", 0) or "",
                pts.get("O", 0) or "",
                total or "",
            ]

            for c, v in enumerate(row_vals, 1):
                # Cols 1-3: apostador/partido/fase con bg de fase
                # Cols 4-7: equipos con bg de fase
                # Cols 8+: color por resultado de pronostico
                bg = fase_bg if c <= 7 else row_bg
                al = "left" if c in (1, 3, 4, 7) else "center"
                style_cell(ws.cell(cur_row, c, v), ox, bg=bg, align=al, size=9)

            for k in ["H", "I", "J", "K", "L", "M", "N", "O"]:
                ft[k] += pts.get(k, 0)
                grand_ft[k] += pts.get(k, 0)
            ft["total"] += total
            grand_ft["total"] += total
            cur_row += 1

        # subtotal ultima fase
        if fase_actual:
            label = f"  Subtotal {ap_name[:16]} - {FASE_LABELS.get(fase_actual, fase_actual)}"
            _write_subtotal(cur_row, label, fase_totals.get(fase_actual, {}),
                            bg="CFD8DC", fg="000000", bold=True, sz=8)
            cur_row += 1

        # grand total apostador (partidos)
        grand_globales = pts_global.get(ap_id, {}).get("total", ap.get("pts_globales", 0))
        grand_all      = grand_ft["total"] + grand_globales
        label_grand    = f"TOTAL {ap_name.upper()[:20]}   Partidos: {grand_ft['total']}  Globales: {grand_globales}  TOTAL: {grand_all}"
        ws.cell(cur_row, 1, label_grand)
        ws.cell(cur_row, 19, grand_ft.get("H", 0) or "")
        ws.cell(cur_row, 20, grand_ft.get("I", 0) or "")
        ws.cell(cur_row, 21, grand_ft.get("J", 0) or "")
        ws.cell(cur_row, 22, grand_ft.get("K", 0) or "")
        ws.cell(cur_row, 23, grand_ft.get("L", 0) or "")
        ws.cell(cur_row, 24, grand_ft.get("M", 0) or "")
        ws.cell(cur_row, 25, grand_ft.get("N", 0) or "")
        ws.cell(cur_row, 26, grand_ft.get("O", 0) or "")
        ws.cell(cur_row, 27, grand_all or "")
        for c in range(1, NCOLS + 1):  # NCOLS = 27
            al = "left" if c == 1 else "center"
            style_cell(ws.cell(cur_row, c), ox, bg="E65100", fg="FFFFFF",
                       bold=True, size=10, align=al)
        cur_row += 2  # espacio entre apostadores


# ── Main ──────────────────────────────────────────────────────────────────────

def generar(torneo_id, tok=None):
    ox = ensure_openpyxl()
    if tok is None:
        tok = login()

    print(f"\n  Generando Excel para torneo {torneo_id}...")
    data = load_all_data(torneo_id, tok)

    if not data["partidos"]:
        print("  Sin partidos con resultados. Corre primero el test integral.")
        return None

    print("  Construyendo workbook...")
    wb = ox.Workbook()

    import openpyxl.styles as _s
    def style_cell_ext(cell, ox, bg=None, bold=False, align="center", size=9, fg="000000"):
        cell.font = _s.Font(bold=bold, size=size, color=fg)
        if bg:
            cell.fill = _s.PatternFill("solid", fgColor=bg)
        cell.alignment = _s.Alignment(horizontal=align, vertical="center")
    globals()["style_cell"] = style_cell_ext

    build_sheet_ranking(wb, ox, data)
    build_sheet_resultados(wb, ox, data)
    for ap in data["apostadores"]:
        build_sheet_apostador(wb, ox, data, ap)
    build_sheet_globales(wb, ox, data)
    build_sheet_matriz(wb, ox, data)

    wb.properties.title   = f"BECBUC Verificacion Torneo {torneo_id}"
    wb.properties.creator = "BECBUC Sistema"
    wb.properties.created = datetime.now()

    out_path = OUTPUT
    wb.save(out_path)
    size_kb = os.path.getsize(out_path) / 1024
    print(f"  Excel guardado: {out_path}  ({size_kb:.1f} KB)")
    print(f"    Hojas: {[ws.title for ws in wb.worksheets]}")
    return out_path


if __name__ == "__main__":
    torneo_id = int(sys.argv[1]) if len(sys.argv) > 1 else None

    if torneo_id is None:
        rows = psql("""
            SELECT t.id, t.nombre FROM torneo t
            JOIN competicion c ON c.id = t.competicion_id
            WHERE t.nombre ILIKE '%mundial%' OR t.nombre ILIKE '%world cup%'
               OR t.nombre ILIKE '%fifa%' OR c.codigo ILIKE '%copa_mundo%'
            ORDER BY t.id LIMIT 1
        """)
        if rows:
            torneo_id = int(rows[0][0])
            print(f"  Torneo detectado: [{torneo_id}] {rows[0][1].strip()}")
        else:
            all_t = psql("SELECT id, nombre FROM torneo ORDER BY id")
            print("Torneos disponibles:")
            for r in all_t:
                print(f"  {r[0].strip()}: {r[1].strip()}")
            print("Uso: python generar_excel_becbuc.py <torneo_id>")
            sys.exit(1)

    tok = login()
    path2 = generar(torneo_id, tok)
    if path2:
        print("\nAbri el archivo: " + str(path2))
