"""
auditoria_jugador.py
====================
Genera un Excel de auditoría BECBUC con dos hojas:
  1. "Por partido"  → detalle partido a partido para un apostador
  2. "Ranking"      → totales A-G de todos los apostadores ordenados

Uso:
    python auditoria_jugador.py [nombre_apostador] [torneo_id] [output.xlsx]

Defaults:
    nombre_apostador = patito
    torneo_id        = 2
    output           = auditoria_<nombre>.xlsx
"""

import sys
import os

# Forzar UTF-8 en stdout para evitar errores de encoding en Windows
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ─── Configuración ────────────────────────────────────────────────────────────
NOMBRE_APOSTADOR = sys.argv[1] if len(sys.argv) > 1 else "patito"
TORNEO_ID        = int(sys.argv[2]) if len(sys.argv) > 2 else 2
OUTPUT_PATH      = sys.argv[3] if len(sys.argv) > 3 else None

if OUTPUT_PATH is None:
    script_dir  = os.path.dirname(os.path.abspath(__file__))
    OUTPUT_PATH = os.path.join(script_dir, f"auditoria_{NOMBRE_APOSTADOR}.xlsx")

PG_APP = dict(host="localhost", port=5432, user="app_user",
              password="superpassword", dbname="app_db")
PG_BEC = dict(host="localhost", port=5432, user="app_user",
              password="superpassword", dbname="becbuc")

# ─── Imports ──────────────────────────────────────────────────────────────────
try:
    import psycopg2, psycopg2.extras
except ImportError:
    sys.exit("ERROR: psycopg2 no disponible. Usar el venv del backend:\n"
             r'  "C:\proyecto FAST API\backend\.venv\Scripts\python.exe" auditoria_jugador.py')
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: openpyxl no disponible.")


# ─── Helpers ──────────────────────────────────────────────────────────────────
def connect(cfg):
    return psycopg2.connect(**cfg, cursor_factory=psycopg2.extras.RealDictCursor)

def fetchall(conn, sql, params=None):
    with conn.cursor() as cur:
        cur.execute(sql, params or ())
        return cur.fetchall()

def fill(hex_color):
    return PatternFill("solid", fgColor="FF" + hex_color)

def bd(color="BFBFBF", thick=False):
    s = Side(style="medium" if thick else "thin", color="FF" + color)
    return Border(left=s, right=s, top=s, bottom=s)

AL  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
AC  = Alignment(horizontal="center", vertical="center")
AR  = Alignment(horizontal="right",  vertical="center")

# Colores
C_HEADER     = fill("1F3864")   # azul marino cabecera
C_SUBHDR     = fill("2F5496")   # azul medio sub-cabecera
C_ROW_ODD    = fill("FFFFFF")
C_ROW_EVEN   = fill("F2F2F2")
C_TOTAL      = fill("1F3864")
C_GREEN      = fill("C6EFCE")   # ranking top
C_YELLOW     = fill("FFEB9C")   # ranking medio
C_RED        = fill("FFC7CE")   # ranking bajo

F_WHITE_B    = Font(bold=True,  color="FFFFFF", size=9)
F_DARK_B     = Font(bold=True,  color="1F3864", size=9)
F_DARK       = Font(bold=False, color="1F3864", size=9)
F_TOTAL      = Font(bold=True,  color="FFFFFF", size=10)

DASH = "-"

def val(v):
    """Muestra el número si >0, guión si 0."""
    return v if (v and v > 0) else DASH


# ══════════════════════════════════════════════════════════════════════════════
# 1. BUSCAR APOSTADOR
# ══════════════════════════════════════════════════════════════════════════════
print(f"Buscando apostador '{NOMBRE_APOSTADOR}' en app_db …")
with connect(PG_APP) as conn_app:
    rows = fetchall(conn_app,
        "SELECT id, username, nombre FROM users "
        "WHERE LOWER(username)=LOWER(%s) OR LOWER(nombre)=LOWER(%s)",
        (NOMBRE_APOSTADOR, NOMBRE_APOSTADOR))
    if not rows:
        rows = fetchall(conn_app,
            "SELECT id, username, nombre FROM users "
            "WHERE LOWER(username) ILIKE %s OR LOWER(nombre) ILIKE %s",
            (f"%{NOMBRE_APOSTADOR.lower()}%", f"%{NOMBRE_APOSTADOR.lower()}%"))

if not rows:
    sys.exit(f"ERROR: No se encontró ningún usuario con nombre '{NOMBRE_APOSTADOR}'.")

usuario = rows[0]
UID  = usuario["id"]
ALIAS = (usuario["nombre"] or usuario["username"] or NOMBRE_APOSTADOR).upper()
print(f"  → {ALIAS} (id={UID})")


# ══════════════════════════════════════════════════════════════════════════════
# 2. QUERY: DETALLE POR PARTIDO (apostador seleccionado)
# ══════════════════════════════════════════════════════════════════════════════
SQL_DETALLE = """
SELECT
    COALESCE(p.numero_fifa, p.id) AS num_fifa,
    COALESCE(el.nombre_es, el.nombre) AS local,
    COALESCE(ev.nombre_es, ev.nombre) AS visitante,
    p.goles_local  AS gl,
    p.goles_visitante AS gv,
    p.estado,
    -- Predicción
    a.pred_local,
    a.pred_visitante,
    -- Puntos por ítem
    COALESCE(pd.pts_resultado,       0)::int AS pts_h,
    COALESCE(pd.pts_marcador,        0)::int AS pts_i,
    COALESCE(pd.pts_amarillas,       0)::int AS pts_j,
    COALESCE(pd.pts_rojas,           0)::int AS pts_k,
    COALESCE(pd.pts_var,             0)::int AS pts_l,
    (COALESCE(pd.pts_penales_partido,0) + COALESCE(pd.pts_penales_tanda,0))::int AS pts_f,
    COALESCE(pd.pts_minuto,          0)::int AS pts_n,
    COALESCE(pd.pts_total,           0)::int AS pts_total
FROM partido p
JOIN fase f   ON f.id = p.fase_id
JOIN equipo el ON el.id = p.equipo_local_id
JOIN equipo ev ON ev.id = p.equipo_visitante_id
LEFT JOIN apuesta a
       ON a.partido_id = p.id AND a.apostador_id = %(uid)s
LEFT JOIN puntaje_detalle pd
       ON pd.partido_id = p.id AND pd.apostador_id = %(uid)s
WHERE f.torneo_id = %(tid)s
ORDER BY f.orden, p.jornada NULLS LAST, p.fecha NULLS LAST, p.id
"""

# ══════════════════════════════════════════════════════════════════════════════
# 3. QUERY: TOTALES DE TODOS LOS APOSTADORES (ranking)
# ══════════════════════════════════════════════════════════════════════════════
SQL_RANKING = """
WITH pd_agg AS (
    SELECT
        pd.apostador_id,
        SUM(COALESCE(pd.pts_resultado,       0))::int AS pts_h,
        SUM(COALESCE(pd.pts_marcador,        0))::int AS pts_i,
        SUM(COALESCE(pd.pts_amarillas,       0))::int AS pts_j,
        SUM(COALESCE(pd.pts_rojas,           0))::int AS pts_k,
        SUM(COALESCE(pd.pts_var,             0))::int AS pts_l,
        SUM(COALESCE(pd.pts_penales_partido, 0) +
            COALESCE(pd.pts_penales_tanda,   0))::int AS pts_f,
        SUM(COALESCE(pd.pts_minuto,          0))::int AS pts_n,
        SUM(COALESCE(pd.pts_total,           0))::int AS pts_part
    FROM puntaje_detalle pd
    JOIN partido p ON p.id = pd.partido_id
    JOIN fase    f ON f.id = p.fase_id
    WHERE f.torneo_id = %(tid)s
    GROUP BY pd.apostador_id
),
pg_agg AS (
    SELECT apostador_id, COALESCE(SUM(pts_total), 0)::int AS pts_glob
    FROM puntaje_global
    WHERE torneo_id = %(tid)s
    GROUP BY apostador_id
),
nombres AS (
    SELECT apostador_id, MAX(nombre_apostador) AS alias
    FROM apuesta a
    JOIN partido p ON p.id = a.partido_id
    JOIN fase    f ON f.id = p.fase_id
    WHERE f.torneo_id = %(tid)s
    GROUP BY apostador_id
)
SELECT
    COALESCE(n.alias, pd.apostador_id::text) AS alias,
    COALESCE(pd.pts_h,    0) AS pts_h,
    COALESCE(pd.pts_i,    0) AS pts_i,
    COALESCE(pd.pts_j,    0) AS pts_j,
    COALESCE(pd.pts_k,    0) AS pts_k,
    COALESCE(pd.pts_l,    0) AS pts_l,
    COALESCE(pd.pts_f,    0) AS pts_f,
    COALESCE(pd.pts_n,    0) AS pts_n,
    COALESCE(pd.pts_part, 0) + COALESCE(pg.pts_glob, 0) AS total
FROM pd_agg pd
LEFT JOIN pg_agg   pg ON pg.apostador_id = pd.apostador_id
LEFT JOIN nombres  n  ON n.apostador_id  = pd.apostador_id
ORDER BY total DESC
"""


# ══════════════════════════════════════════════════════════════════════════════
# 4. FETCH DATA
# ══════════════════════════════════════════════════════════════════════════════
with connect(PG_BEC) as conn_bec:
    partidos = fetchall(conn_bec, SQL_DETALLE, {"uid": UID, "tid": TORNEO_ID})
    ranking_raw = fetchall(conn_bec, SQL_RANKING, {"tid": TORNEO_ID})

# Si nombre_apostador está vacío para algún apostador, completar desde app_db
uid_sin_nombre = [r["alias"] for r in ranking_raw if not r.get("alias") or str(r["alias"]).isdigit()]
name_map = {}
if uid_sin_nombre:
    try:
        uid_ids = [int(r["alias"]) for r in ranking_raw
                   if r.get("alias") and str(r["alias"]).isdigit()]
        if uid_ids:
            with connect(PG_APP) as conn_app2:
                name_rows = fetchall(conn_app2,
                    f"SELECT id, COALESCE(nombre, username) AS alias FROM users "
                    f"WHERE id IN ({','.join(str(x) for x in uid_ids)})")
            name_map = {r["id"]: r["alias"] for r in name_rows}
    except Exception:
        pass

ranking = []
for r in ranking_raw:
    alias = r.get("alias") or ""
    if str(alias).isdigit():
        alias = name_map.get(int(alias), alias)
    ranking.append({
        "alias": str(alias).upper(),
        "pts_h": r["pts_h"], "pts_i": r["pts_i"], "pts_j": r["pts_j"],
        "pts_k": r["pts_k"], "pts_l": r["pts_l"], "pts_f": r["pts_f"],
        "pts_n": r["pts_n"], "total": r["total"],
    })

print(f"  → {len(partidos)} partidos en BD, {len(ranking)} apostadores en ranking")


# ══════════════════════════════════════════════════════════════════════════════
# 5. CONSTRUIR EXCEL
# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ─── HOJA 1: POR PARTIDO ──────────────────────────────────────────────────────
ws1 = wb.active
_sheet1_title = f"Por partido {ALIAS}"
ws1.title = _sheet1_title[:31]  # Excel max 31 chars
ws1.freeze_panes = "A2"
ws1.sheet_view.showGridLines = False

# Columnas: TXT PRONOSTICO | TXT RESULTADO FINAL | A | B | C | D | E | F | G | TOTAL
COL_PRED   = 1
COL_RES    = 2
COL_A      = 3   # pts_h  (resultado)
COL_B      = 4   # pts_i  (exacto)
COL_C      = 5   # pts_j  (amarillas)
COL_D      = 6   # pts_k  (rojas)
COL_E      = 7   # pts_l  (VAR)
COL_F      = 8   # pts_f  (penales partido+tanda)
COL_G      = 9   # pts_n  (1er gol)
COL_TOT    = 10

# Anchos
ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 30
for c in range(3, 11):
    ws1.column_dimensions[get_column_letter(c)].width = 13

HEADERS1 = [
    "TXT PRONOSTICO",
    "TXT RESULTADO FINAL",
    "A- GANA-EMPATA-\nPIERDE",
    "B- RESULTADO\nEXACTO",
    "C-\nAMARILLAS",
    "D- ROJAS",
    "E- VAR",
    "F- PENALES",
    "G- 1ER GOL",
    "TOTAL",
]

# Cabecera
ws1.row_dimensions[1].height = 36
for col, hdr in enumerate(HEADERS1, 1):
    c = ws1.cell(1, col, hdr)
    c.fill      = C_HEADER
    c.font      = F_WHITE_B
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    c.border    = bd("FFFFFF", thick=True)

# Filas de datos
accum = dict(h=0, i=0, j=0, k=0, l=0, f=0, n=0, tot=0)
for row_idx, p in enumerate(partidos, 2):
    ws1.row_dimensions[row_idx].height = 16
    bg = C_ROW_EVEN if row_idx % 2 == 0 else C_ROW_ODD

    num = p["num_fifa"] or row_idx - 1
    local     = (p["local"]     or "").upper()
    visitante = (p["visitante"] or "").upper()

    # TXT predicción
    if p["pred_local"] is not None and p["pred_visitante"] is not None:
        txt_pred = f"P{num:03d}: {local} {p['pred_local']} vs {visitante} {p['pred_visitante']}"
    else:
        txt_pred = f"P{num:03d}: {local} ? vs ? {visitante}"

    # TXT resultado
    if p["gl"] is not None and p["gv"] is not None:
        txt_res = f"P{num:03d}: {local} {p['gl']} vs {visitante} {p['gv']}"
    else:
        txt_res = f"P{num:03d}: {local} - vs - {visitante}"

    row_vals = [
        txt_pred,
        txt_res,
        val(p["pts_h"]),
        val(p["pts_i"]),
        val(p["pts_j"]),
        val(p["pts_k"]),
        val(p["pts_l"]),
        val(p["pts_f"]),
        val(p["pts_n"]),
        val(p["pts_total"]),
    ]

    for col, v in enumerate(row_vals, 1):
        c = ws1.cell(row_idx, col, v)
        c.fill   = bg
        c.border = bd()
        c.font   = F_DARK
        if col <= 2:
            c.alignment = AL
        else:
            c.alignment = AC

    # Acumular
    accum["h"]   += p["pts_h"]
    accum["i"]   += p["pts_i"]
    accum["j"]   += p["pts_j"]
    accum["k"]   += p["pts_k"]
    accum["l"]   += p["pts_l"]
    accum["f"]   += p["pts_f"]
    accum["n"]   += p["pts_n"]
    accum["tot"] += p["pts_total"]

# Fila TOTAL GENERAL
tr = len(partidos) + 2
ws1.row_dimensions[tr].height = 20
total_vals = [
    "Total general", "",
    val(accum["h"]), val(accum["i"]), val(accum["j"]),
    val(accum["k"]), val(accum["l"]), val(accum["f"]),
    val(accum["n"]), val(accum["tot"]),
]
for col, v in enumerate(total_vals, 1):
    c = ws1.cell(tr, col, v)
    c.fill   = C_TOTAL
    c.font   = F_TOTAL
    c.border = bd("FFFFFF", thick=True)
    c.alignment = AC if col > 2 else AL

# Merge "Total general" across first 2 cols
ws1.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=2)


# ─── HOJA 2: RANKING GENERAL ──────────────────────────────────────────────────
ws2 = wb.create_sheet("Ranking general")
ws2.freeze_panes = "A2"
ws2.sheet_view.showGridLines = False

ws2.column_dimensions["A"].width = 6   # RANK
ws2.column_dimensions["B"].width = 22  # ALIAS
for c in range(3, 11):
    ws2.column_dimensions[get_column_letter(c)].width = 12

HEADERS2 = [
    "RANK", "ALIAS",
    "A- GANA-EMPATA-\nPIERDE",
    "B- RESULTADO\nEXACTO",
    "C-\nAMARILLAS",
    "D- ROJAS",
    "E- VAR",
    "F- PENALES",
    "G- 1ER GOL",
    "TOTAL\nPUNTOS",
]

ws2.row_dimensions[1].height = 36
for col, hdr in enumerate(HEADERS2, 1):
    c = ws2.cell(1, col, hdr)
    c.fill      = C_HEADER
    c.font      = F_WHITE_B
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    c.border    = bd("FFFFFF", thick=True)

n_players = len(ranking)
top_cut   = max(1, n_players // 3)
bot_cut   = max(1, n_players - n_players // 3)

for row_idx, r in enumerate(ranking, 2):
    rank = row_idx - 1
    ws2.row_dimensions[row_idx].height = 16

    if rank <= top_cut:
        bg = C_GREEN
    elif rank <= bot_cut:
        bg = C_YELLOW
    else:
        bg = C_RED

    alias = (r.get("alias") or "").upper()

    row_vals = [
        rank,
        alias,
        val(r["pts_h"]),
        val(r["pts_i"]),
        val(r["pts_j"]),
        val(r["pts_k"]),
        val(r["pts_l"]),
        val(r["pts_f"]),
        val(r["pts_n"]),
        r["total"] if r["total"] else DASH,
    ]

    for col, v in enumerate(row_vals, 1):
        c = ws2.cell(row_idx, col, v)
        c.fill   = bg
        c.border = bd()
        c.font   = F_DARK_B if col in (1, 2, 10) else F_DARK
        c.alignment = AL if col == 2 else AC


# ─── 6. Guardar ───────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f"\nExcel guardado en: {OUTPUT_PATH}")
print(f"\nRESUMEN — {ALIAS}")
print(f"  A  Gana/Empata/Pierde: {accum['h']:>5}")
print(f"  B  Exacto:             {accum['i']:>5}")
print(f"  C  Amarillas:          {accum['j']:>5}")
print(f"  D  Rojas:              {accum['k']:>5}")
print(f"  E  VAR:                {accum['l']:>5}")
print(f"  F  Penales:            {accum['f']:>5}")
print(f"  G  1er Gol:            {accum['n']:>5}")
print(f"  ────────────────────────────")
print(f"     TOTAL:              {accum['tot']:>5}")
print(f"\nRanking: {n_players} apostadores incluidos")
