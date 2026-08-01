"""
comparar_excel_bd.py
====================
Compara puntajes del Excel BECBUC_contexto.xlsx vs puntaje_detalle en BD.
Para cada apostador × partido finalizado identifica diferencias en A-G y total.
Genera: comparar_bd_excel.xlsx con resumen y detalle.
"""
import os as _osp
_BASE = _osp.path.dirname(_osp.path.abspath(__file__))
import sys, io, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

try:
    import psycopg2, psycopg2.extras
except ImportError:
    sys.exit("ERROR: psycopg2 no disponible.")
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: openpyxl no disponible.")

# ── Config ─────────────────────────────────────────────────────────────────────
XLSX_SRC = (r"C:\Users\Jose Bogarin\AppData\Roaming\Claude\local-agent-mode-sessions"
            r"\a9fdc79d-9227-450c-a0c1-27eafc601471\dfc0381f-d9d1-4349-b3fa-24cab5c5da8b"
            r"\agent\local_ditto_dfc0381f-d9d1-4349-b3fa-24cab5c5da8b\uploads"
            r"\83901ff8-BECBUC_contexto.xlsx")
TORNEO_ID = 2
OUTPUT    = _osp.path.join(_BASE, 'comparar_bd_excel.xlsx')
PG_BEC    = dict(host="localhost", port=5432, user="app_user",
                 password="superpassword", dbname="becbuc")
PG_APP    = dict(host="localhost", port=5432, user="app_user",
                 password="superpassword", dbname="app_db")

def connect(cfg):
    return psycopg2.connect(**cfg, cursor_factory=psycopg2.extras.RealDictCursor)

def num(v):
    try: return int(v) if v is not None else 0
    except: return 0

def fill(h): return PatternFill("solid", fgColor="FF"+h)
def bd(c="BFBFBF", thick=False):
    s = Side(style="medium" if thick else "thin", color="FF"+c)
    return Border(left=s, right=s, top=s, bottom=s)
def cell(ws, r, c, v, f=None, font=None, aln=None, brd=None):
    cl = ws.cell(r, c, v)
    if f:    cl.fill   = f
    if font: cl.font   = font
    if aln:  cl.alignment = aln
    if brd:  cl.border = brd
    return cl

AC = Alignment(horizontal="center", vertical="center")
AL = Alignment(horizontal="left",   vertical="center")
C_HDR  = fill("1F3864"); C_OK  = fill("C6EFCE")
C_DIF  = fill("FFC7CE"); C_WARN= fill("FFEB9C")
C_GREY = fill("E8E8E8"); C_WHITE = fill("FFFFFF")
FW = Font(bold=True, color="FFFFFFFF", size=9)
FD = Font(bold=False,color="FF1F3864", size=9)
FB = Font(bold=True, color="FF1F3864", size=9)

# ── 1. Leer Excel ──────────────────────────────────────────────────────────────
print("Leyendo Excel fuente...")
wb_src = load_workbook(XLSX_SRC)
ws_res = wb_src['RESULTADOS OFICIALES']
ws_mst = wb_src['MASTER']

# Resultados oficiales
res_xls = {}
for row in ws_res.iter_rows(min_row=2, values_only=True):
    pid = str(row[0] or '').strip()
    if not pid.startswith('P'): continue
    res_xls[pid] = dict(
        gl=num(row[11]), gv=num(row[13]),
        amar=num(row[22]), rojas=num(row[23]), var=num(row[24]),
        pen=num(row[25]), min_gol=num(row[26]),
        estado=str(row[29] or '').strip()
    )

# MASTER: predicciones + puntajes por apostador
# key: (nombre_normalizado, pid)
master_data = {}   # {nombre_norm: {pid: {...}}}
alias_map   = {}   # {nombre_norm: alias}

def norm_nombre(n):
    return re.sub(r'\s+', ' ', str(n or '').strip().upper())

for row in ws_mst.iter_rows(min_row=2, values_only=True):
    pid    = str(row[1] or '').strip()
    nombre = norm_nombre(row[8])
    alias  = str(row[9] or '').replace('\xa0','').strip().upper()
    if not pid.startswith('P') or not nombre: continue
    # fusionar duplicados (Alevo 46+72)
    if nombre not in master_data:
        master_data[nombre] = {}
        alias_map[nombre]   = alias
    master_data[nombre][pid] = dict(
        pred_l=num(row[12]), pred_v=num(row[14]),
        j=num(row[24]), k=num(row[25]), l=num(row[26]),
        m=num(row[27]), n=num(row[28]),
        pts_a=num(row[29]), pts_b=num(row[30]),
        pts_c=num(row[31]), pts_d=num(row[32]),
        pts_e=num(row[33]), pts_f=num(row[34]),
        pts_g=num(row[35]), total=num(row[36]),
        estado=str(row[37] or '').strip()
    )

pids_finalizados = {p for p,d in res_xls.items() if d['estado']=='FINALIZADO'}
print(f"  Apostadores en Excel: {len(master_data)}")
print(f"  Partidos finalizados: {len(pids_finalizados)}")

# ── 2. Leer puntaje_detalle de BD ──────────────────────────────────────────────
print("Leyendo puntaje_detalle de BD...")
SQL_PD = """
SELECT
    ap.nombre || ' ' || COALESCE(ap.first_name,'') AS nombre_raw,
    UPPER(TRIM(COALESCE(ap.nombre, ap.username))) AS nombre_norm,
    a.apostador_id,
    LPAD(ROW_NUMBER() OVER (PARTITION BY a.apostador_id ORDER BY p.numero_partido_fifa), 3, '0') AS pid_seq,
    'P' || LPAD(p.numero_partido_fifa::text, 3, '0') AS pid,
    COALESCE(pd.pts_resultado,0)       AS pts_a,
    COALESCE(pd.pts_marcador,0)        AS pts_b,
    COALESCE(pd.pts_amarillas,0)       AS pts_c,
    COALESCE(pd.pts_rojas,0)           AS pts_d,
    COALESCE(pd.pts_var,0)             AS pts_e,
    COALESCE(pd.pts_penales_partido,0)+COALESCE(pd.pts_penales_tanda,0) AS pts_f,
    COALESCE(pd.pts_minuto,0)          AS pts_g,
    COALESCE(pd.pts_total,0)           AS total_bd,
    a.pred_local                       AS pred_l,
    a.pred_visitante                   AS pred_v,
    p.goles_local                      AS real_l,
    p.goles_visitante                  AS real_v,
    p.estado                           AS estado_p
FROM puntaje_detalle pd
JOIN apuesta a   ON a.id = pd.apuesta_id   -- ajustar si key es diferente
JOIN partido p   ON p.id = pd.partido_id
JOIN fase    f   ON f.id = p.fase_id
JOIN app_db.public.users ap ON ap.id = pd.apostador_id
WHERE f.torneo_id = %(tid)s
ORDER BY pd.apostador_id, p.numero_partido_fifa
"""

# Alternativa sin join en apuesta (si apuesta_id es el join):
SQL_PD2 = """
SELECT
    UPPER(TRIM(COALESCE(ap.nombre, ap.username))) AS nombre_norm,
    'P' || LPAD(p.numero_partido_fifa::text, 3, '0') AS pid,
    COALESCE(pd.pts_resultado,0)       AS pts_a,
    COALESCE(pd.pts_marcador,0)        AS pts_b,
    COALESCE(pd.pts_amarillas,0)       AS pts_c,
    COALESCE(pd.pts_rojas,0)           AS pts_d,
    COALESCE(pd.pts_var,0)             AS pts_e,
    COALESCE(pd.pts_penales_partido,0)+COALESCE(pd.pts_penales_tanda,0) AS pts_f,
    COALESCE(pd.pts_minuto,0)          AS pts_g,
    COALESCE(pd.pts_total,0)           AS total_bd,
    a.pred_local                       AS pred_l,
    a.pred_visitante                   AS pred_v,
    p.goles_local                      AS real_l,
    p.goles_visitante                  AS real_v,
    pd.apostador_id
FROM puntaje_detalle pd
JOIN partido p   ON p.id = pd.partido_id
JOIN fase    f   ON f.id = p.fase_id
JOIN apuesta a   ON a.partido_id = pd.partido_id AND a.apostador_id = pd.apostador_id
WHERE f.torneo_id = %(tid)s
  AND p.estado = 'finalizado'
  AND p.goles_local IS NOT NULL
ORDER BY pd.apostador_id, p.numero_partido_fifa
"""

bd_data = {}   # {nombre_norm: {pid: {...}}}
id_nombre = {} # {apostador_id: nombre_norm}

try:
    conn_bec = connect(PG_BEC)
    # Necesitamos los nombres desde app_db también
    conn_app = connect(PG_APP)
    with conn_app.cursor() as cur:
        cur.execute("SELECT id, UPPER(TRIM(COALESCE(nombre, username))) AS nombre_norm FROM users")
        for r in cur.fetchall():
            id_nombre[r['id']] = r['nombre_norm']
    conn_app.close()

    with conn_bec.cursor() as cur:
        cur.execute(SQL_PD2, {'tid': TORNEO_ID})
        rows = cur.fetchall()

    for r in rows:
        aid = r['apostador_id']
        nombre_norm = id_nombre.get(aid, str(aid))
        pid = r['pid']
        if nombre_norm not in bd_data:
            bd_data[nombre_norm] = {}
        bd_data[nombre_norm][pid] = dict(
            pts_a=r['pts_a'], pts_b=r['pts_b'], pts_c=r['pts_c'],
            pts_d=r['pts_d'], pts_e=r['pts_e'], pts_f=r['pts_f'],
            pts_g=r['pts_g'], total=r['total_bd'],
            pred_l=r['pred_l'], pred_v=r['pred_v'],
            real_l=r['real_l'], real_v=r['real_v']
        )
    conn_bec.close()
    print(f"  Apostadores en BD: {len(bd_data)}")
    print(f"  Filas puntaje_detalle: {len(rows)}")
except Exception as e:
    print(f"ERROR BD: {e}")
    import traceback; traceback.print_exc()
    sys.exit(1)

# ── 3. Matching nombres Excel ↔ BD ────────────────────────────────────────────
# Construir mapa Excel nombre → BD nombre
def best_match(xls_name, bd_names):
    if xls_name in bd_names: return xls_name
    # Intentar partial match
    for bn in bd_names:
        if xls_name in bn or bn in xls_name: return bn
        # match por primera palabra significativa
        xw = xls_name.split()
        bw = bn.split()
        if xw and bw and xw[0]==bw[0] and xw[-1]==bw[-1]: return bn
    return None

nombre_map = {}   # {xls_nombre: bd_nombre}
unmatched  = []
for xn in master_data:
    bm = best_match(xn, bd_data.keys())
    if bm:
        nombre_map[xn] = bm
    else:
        unmatched.append(xn)
        nombre_map[xn] = None

print(f"\nMatcheo nombres:")
for xn, bn in nombre_map.items():
    status = "✓" if bn else "✗ SIN MATCH"
    if not bn: print(f"  {status} XLS='{xn}'")
print(f"  Matcheados: {len([v for v in nombre_map.values() if v])}")
print(f"  Sin match:  {len(unmatched)}")

# ── 4. Comparación ─────────────────────────────────────────────────────────────
print("\nComparando...")

# Estructura: por apostador → por pid → dict con difs
COLS = ['pts_a','pts_b','pts_c','pts_d','pts_e','pts_f','pts_g','total']
NAMES= ['A','B','C','D','E','F','G','TOT']

resumen   = []   # {apostador, tot_xls, tot_bd, dif, pid_diffs:int, ...}
diferencias= []  # {apostador, pid, col, xls_val, bd_val, dif, es_resultado}

for xn in sorted(master_data.keys()):
    bn = nombre_map.get(xn)
    alias = alias_map.get(xn, xn)
    xls_apos = master_data[xn]
    bd_apos  = bd_data.get(bn, {}) if bn else {}

    tot_xls = sum(d['total'] for d in xls_apos.values()
                  if d.get('estado','')=='FINALIZADO')
    tot_bd  = sum(d['total'] for d in bd_apos.values())

    col_sums_xls = {c: sum(d[c] for d in xls_apos.values() if d.get('estado','')=='FINALIZADO')
                    for c in COLS}
    col_sums_bd  = {c: sum(d[c] for d in bd_apos.values()) for c in COLS}

    pid_dif_count = 0
    res_errors    = []   # partidos donde A o B difieren (posible error resultado BD)
    bonus_difs    = []   # partidos donde solo C-G difieren

    for pid in pids_finalizados:
        dx = xls_apos.get(pid)
        db = bd_apos.get(pid)
        if not dx: continue

        for col, nm in zip(COLS, NAMES):
            vx = dx.get(col, 0)
            vb = db.get(col, 0) if db else 0
            if vx != vb:
                diferencias.append(dict(
                    apostador=alias, nombre=xn, pid=pid,
                    col=nm, xls=vx, bd=vb, dif=vb-vx,
                    pred_xls=f"{dx.get('pred_l','?')}-{dx.get('pred_v','?')}",
                    real_xls=f"{res_xls[pid]['gl']}-{res_xls[pid]['gv']}",
                    pred_bd=f"{(db or {}).get('pred_l','?')}-{(db or {}).get('pred_v','?')}",
                    real_bd=f"{(db or {}).get('real_l','?')}-{(db or {}).get('real_v','?')}"
                ))
                pid_dif_count += 1

        # Detectar si A o B difieren → posible error resultado
        if db:
            a_ok = dx.get('pts_a',0)==db.get('pts_a',0)
            b_ok = dx.get('pts_b',0)==db.get('pts_b',0)
            if not a_ok or not b_ok:
                res_errors.append(pid)

    resumen.append(dict(
        alias=alias, nombre=xn,
        tot_xls=tot_xls, tot_bd=tot_bd, dif=tot_bd-tot_xls,
        col_xls=col_sums_xls, col_bd=col_sums_bd,
        pid_dif=pid_dif_count, res_errors=res_errors,
        matched=bool(bn)
    ))

print(f"  Filas diferencias: {len(diferencias)}")

# ── 5. Analisis de errores por partido ────────────────────────────────────────
# Para cada partido: cuántos apostadores tienen A o B diferente
pid_impact = {}
for d in diferencias:
    if d['col'] in ('A','B'):
        k = (d['pid'], d['col'])
        if k not in pid_impact:
            pid_impact[k] = {'count':0,'examples':[], 'real_xls':'','real_bd':''}
        pid_impact[k]['count'] += 1
        pid_impact[k]['real_xls'] = d['real_xls']
        pid_impact[k]['real_bd']  = d['real_bd']
        if len(pid_impact[k]['examples'])<3:
            pid_impact[k]['examples'].append(f"{d['apostador']}:{d['pred_xls']}")

print("\n=== PARTIDOS CON ERRORES DE RESULTADO (A o B difieren) ===")
for (pid,col), data in sorted(pid_impact.items()):
    print(f"  {pid} col={col}: {data['count']} apostadores afectados | "
          f"XLS={data['real_xls']} BD={data['real_bd']} | ej: {data['examples']}")

# ── 6. Generar Excel de comparación ───────────────────────────────────────────
print("\nGenerando Excel de comparación...")
wb_out = Workbook()

# Hoja 1: Resumen por apostador
ws1 = wb_out.active
ws1.title = "Resumen apostadores"
ws1.freeze_panes = "A2"
ws1.sheet_view.showGridLines = False

hdrs = ['ALIAS','A_XLS','A_BD','B_XLS','B_BD','C_XLS','C_BD',
        'D_XLS','D_BD','E_XLS','E_BD','F_XLS','F_BD','G_XLS','G_BD',
        'TOT_XLS','TOT_BD','DIF','PART_DIFS','ERR_RESULT']
ws1.row_dimensions[1].height = 36
col_widths = [22,6,6,6,6,6,6,6,6,6,6,6,6,6,6,8,8,8,8,12]
for i,(h,w) in enumerate(zip(hdrs,col_widths),1):
    ws1.column_dimensions[get_column_letter(i)].width = w
    c = ws1.cell(1,i,h)
    c.fill=C_HDR; c.font=FW; c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    c.border=bd("FFFFFF",True)

resumen_sorted = sorted(resumen, key=lambda x: -abs(x['dif']))
for ri, r in enumerate(resumen_sorted, 2):
    ws1.row_dimensions[ri].height = 16
    bg = C_DIF if abs(r['dif'])>10 else (C_WARN if abs(r['dif'])>0 else C_OK)
    vals = [
        r['alias'],
        r['col_xls']['pts_a'], r['col_bd']['pts_a'],
        r['col_xls']['pts_b'], r['col_bd']['pts_b'],
        r['col_xls']['pts_c'], r['col_bd']['pts_c'],
        r['col_xls']['pts_d'], r['col_bd']['pts_d'],
        r['col_xls']['pts_e'], r['col_bd']['pts_e'],
        r['col_xls']['pts_f'], r['col_bd']['pts_f'],
        r['col_xls']['pts_g'], r['col_bd']['pts_g'],
        r['tot_xls'], r['tot_bd'],
        r['dif'], r['pid_dif'],
        ','.join(r['res_errors'][:5]) if r['res_errors'] else '✓'
    ]
    for ci, v in enumerate(vals,1):
        c = ws1.cell(ri, ci, v)
        c.fill = bg
        c.font = FB if ci in (1,16,17,18) else FD
        c.alignment = AL if ci==1 else AC
        c.border = bd()

# Hoja 2: Diferencias por partido
ws2 = wb_out.create_sheet("Diferencias por partido")
ws2.freeze_panes = "A2"
ws2.sheet_view.showGridLines = False

hdrs2 = ['PID','CONCEPTO','APOSTADORES_AFECTADOS','REAL_EXCEL','REAL_BD',
         'DIAGNOSTICO']
col_w2 = [6,10,22,10,10,40]
ws2.row_dimensions[1].height = 28
for i,(h,w) in enumerate(zip(hdrs2,col_w2),1):
    ws2.column_dimensions[get_column_letter(i)].width = w
    c = ws2.cell(1,i,h)
    c.fill=C_HDR; c.font=FW; c.alignment=AC; c.border=bd("FFFFFF",True)

r2 = 2
# Ordenar por cantidad de afectados
pid_col_sorted = sorted(pid_impact.items(), key=lambda x: -x[1]['count'])
for (pid,col), data in pid_col_sorted:
    r_xls = data['real_xls']; r_bd = data['real_bd']
    if r_xls == r_bd:
        diag = f"Resultado coincide ({r_xls}), diferencia de puntaje en col {col}"
        bg2 = C_WARN
    else:
        diag = f"RESULTADO DISTINTO: Excel={r_xls} BD={r_bd} → posible error en BD"
        bg2 = C_DIF
    ws2.row_dimensions[r2].height = 22
    for ci,v in enumerate([pid, col, data['count'], r_xls, r_bd, diag],1):
        c = ws2.cell(r2,ci,v)
        c.fill=bg2; c.font=FD; c.border=bd()
        c.alignment = AL if ci in(5,) else AC
    r2 += 1

# También partidos con diferencias en C-G (no A/B)
from collections import defaultdict
pid_cfg = defaultdict(lambda: {'count':0,'cols':set()})
for d in diferencias:
    if d['col'] not in ('A','B','TOT'):
        pid_cfg[d['pid']]['count'] += 1
        pid_cfg[d['pid']]['cols'].add(d['col'])
for pid, data in sorted(pid_cfg.items(), key=lambda x:-x[1]['count']):
    if pid in [k[0] for k in pid_impact]: continue
    ws2.row_dimensions[r2].height = 18
    diag = f"Solo bonus: cols {sorted(data['cols'])}"
    for ci,v in enumerate([pid,'C-F',data['count'],
                            res_xls[pid]['gl'] if pid in res_xls else '?',
                            '-', diag],1):
        c = ws2.cell(r2,ci,v)
        c.fill=C_WARN; c.font=FD; c.border=bd(); c.alignment=AC
    r2+=1

# Hoja 3: Detalle por apostador y partido (solo diferencias A o B)
ws3 = wb_out.create_sheet("Detalle resultado erroneo")
ws3.freeze_panes = "A2"
ws3.sheet_view.showGridLines = False
hdrs3 = ['APOSTADOR','PID','PRED_XLS','PRED_BD','REAL_XLS','REAL_BD',
         'A_XLS','A_BD','B_XLS','B_BD','TOT_XLS','TOT_BD','DIF']
col_w3 = [20,6,8,8,8,8,6,6,6,6,8,8,8]
ws3.row_dimensions[1].height = 28
for i,(h,w) in enumerate(zip(hdrs3,col_w3),1):
    ws3.column_dimensions[get_column_letter(i)].width = w
    c = ws3.cell(1,i,h)
    c.fill=C_HDR; c.font=FW; c.alignment=AC; c.border=bd("FFFFFF",True)

r3=2
ab_difs = [d for d in diferencias if d['col'] in ('A','B')]
# Agrupar por apostador+pid
seen = set()
for d in sorted(ab_difs, key=lambda x:(x['pid'],x['apostador'])):
    key=(d['apostador'],d['pid'])
    if key in seen: continue
    seen.add(key)
    dx = master_data[d['nombre']].get(d['pid'],{})
    bn = nombre_map.get(d['nombre'])
    db = bd_data.get(bn,{}).get(d['pid'],{}) if bn else {}
    ws3.row_dimensions[r3].height=15
    diag_color = C_DIF if d['real_xls']!=d['real_bd'] else C_WARN
    for ci,v in enumerate([d['apostador'],d['pid'],
                            d['pred_xls'],d['pred_bd'],
                            d['real_xls'],d['real_bd'],
                            dx.get('pts_a',0), db.get('pts_a',0),
                            dx.get('pts_b',0), db.get('pts_b',0),
                            dx.get('total',0), db.get('total',0),
                            db.get('total',0)-dx.get('total',0)],1):
        c = ws3.cell(r3,ci,v)
        c.fill=diag_color; c.font=FD; c.border=bd()
        c.alignment = AL if ci==1 else AC
    r3+=1

wb_out.save(OUTPUT)
print(f"\nExcel guardado: {OUTPUT}")
print(f"  Hoja 1: Resumen {len(resumen)} apostadores")
print(f"  Hoja 2: {r2-2} partidos con diferencias")
print(f"  Hoja 3: {r3-2} filas detalle resultado erróneo")

# Resumen en consola
print("\n=== TOP DIFERENCIAS POR APOSTADOR ===")
for r in resumen_sorted[:15]:
    err_str = f"  err_A/B: {r['res_errors']}" if r['res_errors'] else ""
    print(f"  {r['alias']:<22} XLS={r['tot_xls']:>4} BD={r['tot_bd']:>4} dif={r['dif']:>+5}  pids_dif={r['pid_dif']}{err_str}")
