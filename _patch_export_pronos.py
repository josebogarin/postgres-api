# -*- coding: utf-8 -*-
"""Exportar pronosticos (Excel) + indicador de completados por fase abierta. Portal + Movil + Backend."""
import ast, re, shutil, subprocess, os
from datetime import datetime
B="/sessions/stoic-busy-euler/mnt/proyecto FAST API"
if not os.path.exists(B): B=r"C:\proyecto FAST API"
HTMLP=B+"/backend/static/BECBUC-portal.html"
HTMLM=B+"/backend/static/BECBUC-movil.html"
PY=B+"/backend/app/api/v1/endpoints/apostador_bets.py"
BKP=B+"/_backups"; os.makedirs(BKP,exist_ok=True)

def vhtml(p):
    raw=open(p,'rb').read()
    if not raw.rstrip().endswith(b'</html>'): return False,'falta </html>'
    s=re.findall(rb'<script>([\s\S]*?)</script>',raw)
    if not s: return False,'sin <script>'
    r=subprocess.run(['node','--check'],input=s[-1],capture_output=True)
    return (r.returncode==0),('OK' if r.returncode==0 else r.stderr.decode(errors='replace')[:200])
def vpy(p):
    try: ast.parse(open(p,encoding='utf-8').read()); return True,'OK'
    except SyntaxError as e: return False,f'line {e.lineno}: {e.msg}'
def apply(path,repls,verifier):
    src=open(path,encoding='utf-8').read()
    for old,new in repls:
        c=src.count(old)
        if c!=1: raise SystemExit(f'count={c} en {os.path.basename(path)}: {old[:70]!r}')
        src=src.replace(old,new,1)
    b=os.path.join(BKP,os.path.basename(path)+'.'+datetime.now().strftime('%Y%m%d_%H%M%S')+'.bak')
    shutil.copy2(path,b); open(path,'w',encoding='utf-8').write(src)
    ok,msg=verifier(path)
    if not ok: shutil.copy2(b,path); raise SystemExit(f'VERIFY FALLO {os.path.basename(path)}: {msg} -> restaurado')
    print(f'OK {os.path.basename(path)} ({msg})')

# ═══════════ BACKEND ═══════════
BACK = r'''@router.get("/exportar-pronosticos/{torneo_id}",
            summary="Excel de pronosticos por apostador de todas las fases abiertas")
async def exportar_pronosticos(torneo_id: int, db: DBSession, current: CurrentAdmin):
    import io
    from datetime import datetime as _dt
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Apostadores validos (rol apostador) + nombres desde app_db
    async with _app_engine.connect() as conn:
        ar = await conn.execute(text("""
            SELECT u.id, u.username, u.nombre FROM users u
            JOIN user_roles ur ON ur.user_id = u.id
            JOIN roles ro ON ro.id = ur.role_id
            WHERE ro.name = 'apostador' AND u.is_active = TRUE
        """))
        ainfo = {row["id"]: (row["username"], row["nombre"]) for row in ar.mappings()}
    ids = list(ainfo.keys())

    r = await db.execute(text("""
        SELECT a.apostador_id, p.numero_fifa,
               COALESCE(f.nombre, f.tipo) AS fase, COALESCE(f.orden, 0) AS fase_orden,
               COALESCE(el.nombre_es, el.nombre) AS local,
               COALESCE(ev.nombre_es, ev.nombre) AS visitante,
               p.goles_local, p.goles_visitante, p.estado,
               a.pred_local, a.pred_visitante,
               a.pred_amarillas, a.pred_rojas, a.pred_var, a.pred_penales_partido,
               a.pred_minuto_gol, a.pred_penales_local_tanda, a.pred_penales_visitante_tanda,
               COALESCE(ecl.nombre_es, ecl.nombre) AS clasifica_nom
        FROM apuesta a
        JOIN partido p ON p.id = a.partido_id
        JOIN fase f ON f.id = p.fase_id
        JOIN equipo el ON el.id = p.equipo_local_id
        JOIN equipo ev ON ev.id = p.equipo_visitante_id
        LEFT JOIN equipo ecl ON ecl.id = a.pred_equipo_clasifica
        WHERE f.torneo_id = :tid
          AND COALESCE(f.bloqueada, FALSE) = FALSE
          AND a.pred_local IS NOT NULL
          AND a.apostador_id = ANY(:ids)
    """), {"tid": torneo_id, "ids": ids})
    rows = [dict(x) for x in r.mappings()]
    for row in rows:
        info = ainfo.get(row["apostador_id"], ("?", ""))
        row["usuario"] = info[0] or ("?%s" % row["apostador_id"])
        row["nombre_real"] = info[1] or ""
    rows.sort(key=lambda x: ((x["usuario"] or "").lower(), x["fase_orden"] or 0, x["numero_fifa"] or 0))

    def _hf(c): return PatternFill("solid", fgColor=c)
    HDR=_hf("1A6B45"); F_A=_hf("EAF6EF"); F_B=_hf("FFFFFF")
    FW=Font(color="FFFFFF", bold=True, size=9); F9=Font(size=9); F9B=Font(bold=True, size=9)
    AC=Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL=Alignment(horizontal="left", vertical="center")
    _s=Side(style="thin", color="CCCCCC"); BDR=Border(left=_s, right=_s, top=_s, bottom=_s)

    wb=Workbook(); ws=wb.active; ws.title="Pronosticos fases abiertas"
    headers=["No Partido","Fase","Usuario","Nombre","Local","Visitante","Resultado",
             "Pron. Local","Pron. Visit","Amarillas (J)","Rojas (K)","VAR (L)",
             "Pen. juego (M)","Min. 1er gol (N)","Tanda Local (Ol)","Tanda Visit (Ov)","Clasifica (P)"]
    widths=[9,16,14,22,18,18,10,10,10,10,9,8,12,13,12,12,16]
    for ci,(h,w) in enumerate(zip(headers,widths),1):
        c=ws.cell(1,ci,h); c.fill=HDR; c.font=FW; c.alignment=AC; c.border=BDR
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[1].height=26; ws.freeze_panes="A2"

    def _v(x): return "" if x is None else x
    def _res(gl, gv, estado):
        if estado == "finalizado" and gl is not None and gv is not None:
            return "%s-%s" % (gl, gv)
        return "-"

    color_idx=0; prev_user=None
    for ri,row in enumerate(rows,2):
        if row["usuario"]!=prev_user:
            color_idx ^= 1; prev_user=row["usuario"]
        fill=F_A if color_idx==0 else F_B
        vals=["P%03d" % row["numero_fifa"] if row["numero_fifa"] else "",
              _v(row["fase"]), _v(row["usuario"]), _v(row["nombre_real"]),
              _v(row["local"]), _v(row["visitante"]),
              _res(row["goles_local"], row["goles_visitante"], row["estado"]),
              _v(row["pred_local"]), _v(row["pred_visitante"]),
              _v(row["pred_amarillas"]), _v(row["pred_rojas"]), _v(row["pred_var"]),
              _v(row["pred_penales_partido"]), _v(row["pred_minuto_gol"]),
              _v(row["pred_penales_local_tanda"]), _v(row["pred_penales_visitante_tanda"]),
              _v(row["clasifica_nom"])]
        for ci,v in enumerate(vals,1):
            c=ws.cell(ri,ci,v); c.fill=fill
            c.font=F9B if ci==3 else F9
            c.alignment=AL if ci in (2,3,4,5,6) else AC; c.border=BDR
    ws.auto_filter.ref="A1:%s1" % get_column_letter(len(headers))

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    ts=_dt.now().strftime("%Y%m%d_%H%M")
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="becbuc_pronosticos_fases_abiertas_%s.xlsx"' % ts})


@router.get("/pronosticos-completados/{torneo_id}",
            summary="Por fase abierta: apostadores que completaron todas sus apuestas")
async def pronosticos_completados(torneo_id: int, db: DBSession, current: CurrentAdmin) -> dict:
    async with _app_engine.connect() as conn:
        ar = await conn.execute(text("""
            SELECT u.id FROM users u
            JOIN user_roles ur ON ur.user_id = u.id
            JOIN roles ro ON ro.id = ur.role_id
            WHERE ro.name = 'apostador' AND u.is_active = TRUE
        """))
        ids = [row["id"] for row in ar.mappings()]
    total_apost = len(ids)

    r = await db.execute(text("""
        WITH fa AS (
            SELECT f.id, COALESCE(f.nombre, f.tipo) AS fase, f.tipo,
                   COALESCE(f.orden, 0) AS orden, COUNT(p.id) AS total_partidos
            FROM fase f JOIN partido p ON p.fase_id = f.id
            WHERE f.torneo_id = :tid AND COALESCE(f.bloqueada, FALSE) = FALSE
            GROUP BY f.id, f.nombre, f.tipo, f.orden
        ),
        ap AS (
            SELECT p.fase_id, a.apostador_id, COUNT(DISTINCT a.partido_id) AS n
            FROM apuesta a JOIN partido p ON p.id = a.partido_id
            JOIN fa ON fa.id = p.fase_id
            WHERE a.pred_local IS NOT NULL AND a.apostador_id = ANY(:ids)
            GROUP BY p.fase_id, a.apostador_id
        )
        SELECT fa.fase, fa.tipo, fa.orden, fa.total_partidos,
               COUNT(ap.apostador_id) FILTER (WHERE ap.n >= fa.total_partidos) AS completados,
               COUNT(ap.apostador_id) AS con_alguna
        FROM fa LEFT JOIN ap ON ap.fase_id = fa.id
        GROUP BY fa.fase, fa.tipo, fa.orden, fa.total_partidos
        ORDER BY fa.orden
    """), {"tid": torneo_id, "ids": ids})
    fases = []
    for x in r.mappings():
        fases.append({"fase": x["fase"], "tipo": x["tipo"],
                      "total_partidos": x["total_partidos"],
                      "completados": x["completados"] or 0,
                      "con_alguna": x["con_alguna"] or 0,
                      "total_apostadores": total_apost})
    return {"total_apostadores": total_apost, "fases": fases}


'''
back_anchor = '@router.get("/exportar-puntajes/{torneo_id}", summary="Excel puntajes: resumen + detalle por partido")'

# ═══════════ PORTAL ═══════════
p_card_anchor = "        <!-- Apostadores sin apuestas completas -->"
p_card_new = ('''        <!-- Apuestas completas por fase abierta + exportar pronosticos -->
        <div id="monCompletadosWrap" style="display:none;margin-bottom:14px;">
          <div class="mon-card">
            <div class="mon-card-title" style="display:flex;align-items:center;gap:8px;">
              <i class="ti ti-checklist"></i> Apuestas completas por fase abierta
              <button class="mon-act-btn primary" onclick="exportarPronosticos()" style="margin-left:auto;padding:4px 12px;font-size:.7rem;"><i class="ti ti-file-spreadsheet"></i> Exportar pronósticos</button>
            </div>
            <div id="monCompletadosList" style="max-height:220px;overflow-y:auto;"><div style="color:var(--muted);font-size:.78rem;text-align:center;padding:12px;">Cargando…</div></div>
          </div>
        </div>

'''
              + p_card_anchor)

p_js_anchor = "async function exportarPuntajes() {"
p_js_new = r'''async function exportarPronosticos(){
  if(!_betTorneoId){ alert('Seleccioná un torneo primero.'); return; }
  const btn=event&&event.target?event.target.closest('button'):null;
  const orig=btn?btn.innerHTML:'';
  if(btn){ btn.disabled=true; btn.innerHTML='<i class="ti ti-loader-2"></i> Generando…'; }
  try{
    await _downloadFile(`/api/v1/bets/exportar-pronosticos/${_betTorneoId}`, `becbuc_pronosticos.xlsx`);
  }catch(e){ alert('Error al exportar: '+e.message); }
  finally{ if(btn){ btn.disabled=false; btn.innerHTML=orig||'<i class="ti ti-file-spreadsheet"></i> Exportar pronósticos'; } }
}

async function loadCompletadosFase(){
  const wrap=document.getElementById('monCompletadosWrap');
  const list=document.getElementById('monCompletadosList');
  if(!wrap||!list||!_betTorneoId) return;
  try{
    const r=await fetch(`/api/v1/bets/pronosticos-completados/${_betTorneoId}`,{headers:{Authorization:`Bearer ${token}`}});
    if(!r.ok) throw new Error('Error '+r.status);
    const d=await r.json(); const fases=d.fases||[]; const tot=d.total_apostadores||0;
    if(!fases.length){ wrap.style.display='none'; return; }
    wrap.style.display='block';
    list.innerHTML=fases.map(f=>{
      const pct=tot?Math.round(100*(f.completados||0)/tot):0;
      const col=pct>=100?'#34d399':pct>=50?'#fbbf24':'#f87171';
      return `<div style="display:flex;align-items:center;gap:8px;padding:7px 4px;border-bottom:1px solid rgba(255,255,255,.06)">
        <span style="flex:1;font-size:.82rem;color:#e2e8f0">${esc(f.fase)}</span>
        <span style="font-size:.72rem;color:var(--muted)">${f.total_partidos} part.</span>
        <span style="font-weight:800;color:${col};font-size:.9rem">${f.completados}/${tot}</span>
        <span style="font-size:.7rem;color:var(--muted)">completaron</span>
      </div>`;
    }).join('');
  }catch(e){ wrap.style.display='none'; }
}

async function exportarPuntajes() {'''

p_hook_anchor = ("  // Partidos de hoy (reemplaza panel sin-apuestas)\n"
                 "  loadPartidosHoyMon();")
p_hook_new = ("  // Partidos de hoy (reemplaza panel sin-apuestas)\n"
              "  loadPartidosHoyMon();\n"
              "  loadCompletadosFase();")

# ═══════════ MOVIL ═══════════
m_sec_anchor = ("  // Fases bloqueo\n"
                "  h+=`<div style=\"margin-top:14px\">\n"
                "    <div style=\"font-weight:800;font-size:.85rem;margin-bottom:6px\">🔒 Bloqueo por Fase</div>")
m_sec_new = ("  // Pronosticos fases abiertas + completados\n"
             "  h+=`<div style=\"margin-top:14px\">\n"
             "    <div style=\"font-weight:800;font-size:.85rem;margin-bottom:6px\">📋 Pronósticos (fases abiertas)</div>\n"
             "    <div id=\"amCompletados\" style=\"display:flex;flex-direction:column;gap:6px;margin-bottom:8px\"><span style=\"color:var(--muted);font-size:.75rem\">Cargando…</span></div>\n"
             "    <button onclick=\"exportarPronosticosM(this)\" style=\"${bs};background:linear-gradient(135deg,#2563eb,#1d4ed8)\">📥 Exportar pronósticos (Excel)</button>\n"
             "  </div>`;\n"
             "  // Fases bloqueo\n"
             "  h+=`<div style=\"margin-top:14px\">\n"
             "    <div style=\"font-weight:800;font-size:.85rem;margin-bottom:6px\">🔒 Bloqueo por Fase</div>")

m_hook_anchor = ("  _loadFasesBloqueoM();\n"
                 "  loadAdminPanel();\n"
                 "}")
m_hook_new = ("  _loadFasesBloqueoM();\n"
              "  loadAdminPanel();\n"
              "  _loadCompletadosFaseM();\n"
              "}")

m_js_anchor = "async function exportarPuntajesM(btn){"
m_js_new = r'''async function exportarPronosticosM(btn){
  if(!_torneoId){alert('Seleccioná un torneo primero.');return;}
  const orig=btn?btn.textContent:'';
  if(btn){btn.disabled=true;btn.textContent='⏳…';}
  try{ await _downloadFileM(`/api/v1/bets/exportar-pronosticos/${_torneoId}`, `becbuc_pronosticos.xlsx`); }
  catch(e){alert(`Error al exportar: ${e.message}`);}
  finally{if(btn){btn.disabled=false;btn.textContent=orig||'📥 Exportar pronósticos (Excel)';}}
}
async function _loadCompletadosFaseM(){
  const el=document.getElementById('amCompletados');
  if(!el||!_torneoId) return;
  try{
    const r=await fetch(`/api/v1/bets/pronosticos-completados/${_torneoId}`,{headers:{Authorization:`Bearer ${token}`}});
    if(!r.ok) throw new Error('Error '+r.status);
    const d=await r.json(); const fases=d.fases||[]; const tot=d.total_apostadores||0;
    if(!fases.length){ el.innerHTML='<span style="color:var(--muted);font-size:.75rem">Sin fases abiertas.</span>'; return; }
    el.innerHTML=fases.map(f=>{
      const pct=tot?Math.round(100*(f.completados||0)/tot):0;
      const col=pct>=100?'#34d399':pct>=50?'#fbbf24':'#f87171';
      return `<div style="display:flex;align-items:center;gap:8px;background:var(--bg2);border-radius:8px;padding:8px 10px">
        <span style="flex:1;font-size:.8rem;color:var(--text)">${f.fase}</span>
        <span style="font-size:.7rem;color:var(--muted)">${f.total_partidos} part.</span>
        <span style="font-weight:800;color:${col};font-size:.85rem">${f.completados}/${tot}</span></div>`;
    }).join('');
  }catch(e){ el.innerHTML='<span style="color:var(--muted);font-size:.75rem">No disponible.</span>'; }
}
async function exportarPuntajesM(btn){'''

apply(PY,    [(back_anchor, BACK+back_anchor)], vpy)
apply(HTMLP, [(p_card_anchor, p_card_new), (p_js_anchor, p_js_new), (p_hook_anchor, p_hook_new)], vhtml)
apply(HTMLM, [(m_sec_anchor, m_sec_new), (m_hook_anchor, m_hook_new), (m_js_anchor, m_js_new)], vhtml)
print("TODO OK.")
