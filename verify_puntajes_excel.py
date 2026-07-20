# -*- coding: utf-8 -*-
"""verify_puntajes_excel.py — descarga /exportar-puntajes y vuelca su estructura
(hojas + dimensiones) para comparar antes/despues de mover el generador."""
import io
import json
import sys
import urllib.request

API = "http://localhost:8000"
TID = 2


def login():
    body = json.dumps({"username": "jose", "password": "catalina"}).encode()
    req = urllib.request.Request(API + "/api/v1/auth/login", data=body,
                                 headers={"Content-Type": "application/json"}, method="POST")
    d = json.loads(urllib.request.urlopen(req, timeout=30).read())
    return d.get("access_token") or d.get("token")


def main():
    tok = login()
    req = urllib.request.Request(API + f"/api/v1/bets/exportar-puntajes/{TID}",
                                 headers={"Authorization": f"Bearer {tok}"})
    data = urllib.request.urlopen(req, timeout=120).read()
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True)
    print(f"HOJAS={len(wb.sheetnames)}")
    for ws in wb.worksheets:
        print(f"  SHEET '{ws.title}'  dims={ws.max_row}x{ws.max_column}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("ERROR:", e); sys.exit(1)
