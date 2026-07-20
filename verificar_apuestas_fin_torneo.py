# -*- coding: utf-8 -*-
r"""
verificar_apuestas_fin_torneo.py  [--apply] [ruta_excel_opcional]

Verifica que TODAS las apuestas de la BD (torneo 2) coincidan con el Excel de
fin de torneo (hoja "50- TBL MASTER"). Si difieren o faltan, con --apply deja la
BD IDENTICA al Excel (el Excel manda).

Cubre:
  - Marcador + bonus por partido  (filas P001..P104):
      col 13 -> pred_local     col 15 -> pred_visitante
      col 25 -> pred_amarillas (J)   col 26 -> pred_rojas (K)
      col 27 -> pred_var (L)         col 28 -> pred_penales_partido (M)
      col 29 -> pred_minuto_gol (N)  col 30 -> pred_penales_local_tanda (O-L)
      col 31 -> pred_penales_visitante_tanda (O-V)  col 32 -> pred_equipo_clasifica (Q)
  - Globales A-G  (filas P111..P118, valor en col 11):
      P111 campeon   P112 otro finalista   P113 goleador   P114 peor equipo
      P115 etapa PY  P116 goles PY          P117 goleada ganador  P118 goleada perdedor

SIN --apply : DRY RUN (solo reporta; imprime cabeceras y muestra diferencias).
CON --apply : escribe (upsert apuesta + apuesta_global).  Luego: run_reabrir_y_recalcular.bat

Uso:
  backend\.venv\Scripts\python.exe verificar_apuestas_fin_torneo.py
  backend\.venv\Scripts\python.exe verificar_apuestas_fin_torneo.py --apply
"""
import sys, os, unicodedata
from collections import Counter

args = [a for a in sys.argv[1:]]
DO_APPLY = any(a.lower() == '--apply' for a in args)
FORCE_ALL = any(a.lower() == '--force-all' for a in args)
PATH_ARG = next((a for a in args if not a.startswith('--')), None)
_modo = (('APPLY FORZADO' if FORCE_ALL else 'APPLY SEGURO') if DO_APPLY else 'DRY RUN - no escribe')
print(f"[{_modo}]")

BASE = os.path.dirname(os.path.abspath(__file__))

def find_excel():
    # ESTRICTO: solo el Excel de FIN DE TORNEO (o una ruta pasada como argumento).
    # Nunca hace fallback a otros .xlsx del root para no usar un archivo viejo por error.
    if PATH_ARG and os.path.exists(PATH_ARG):
        return PATH_ARG
    OK = ('TORNEO CERRADO', 'FIN DE TORNEO', '20260720')
    BAD = ('CORRECCIONES', 'SEMIFINAL')
    cands = [f for f in os.listdir(BASE)
             if f.lower().endswith('.xlsx') and not f.startswith('~')
             and any(k in f.upper() for k in OK) and not any(b in f.upper() for b in BAD)]
    if not cands:
        return None
    cands.sort(key=lambda f: os.path.getmtime(os.path.join(BASE, f)), reverse=True)
    return os.path.join(BASE, cands[0])

EXCEL = find_excel()
if not EXCEL:
    sys.exit("ERROR: no encontre el Excel de fin de torneo en la raiz del proyecto.\n"
             "Copia '...FIN DE TORNEO.xlsx' a C:\\proyecto FAST API\\ o pasa la ruta como argumento.")
print(f"Excel: {os.path.basename(EXCEL)}")

try:
    import openpyxl
except ImportError:
    os.system(f'"{sys.executable}" -m pip install openpyxl --quiet'); import openpyxl
try:
    import psycopg2, psycopg2.extras
except ImportError:
    os.system(f'"{sys.executable}" -m pip install psycopg2-binary --quiet'); import psycopg2, psycopg2.extras

CONN_BEC = "host=localhost port=5432 dbname=becbuc user=app_user password=superpassword"
CONN_APP = "host=localhost port=5432 dbname=app_db  user=app_user password=superpassword"
TID = 2

# ---- columnas (1-based) de la hoja TBL MASTER ----
C_PID, C_ALIAS = 2, 10
C_PL, C_PV = 13, 15
BONUS_COLS = [   # (campo_bd, col, tipo)
    ('pred_amarillas', 25, 'int'),
    ('pred_rojas', 26, 'int'),
    ('pred_var', 27, 'int'),
    ('pred_penales_partido', 28, 'int'),
    ('pred_minuto_gol', 29, 'int'),
    ('pred_penales_local_tanda', 30, 'int'),
    ('pred_penales_visitante_tanda', 31, 'int'),
    ('pred_equipo_clasifica', 32, 'equipo'),
]
# globales: fila pid -> campo logico ; valor en columna 11
GLOB_PID = {
    'P111': 'campeon', 'P112': 'finalista2', 'P113': 'goleador', 'P114': 'peor',
    'P115': 'etapa_py', 'P116': 'goles_py', 'P117': 'goleada_gan', 'P118': 'goleada_per',
}
C_GLOB_VAL = 11
FASE_NORM = {
    'grupo': 'grupos', 'grupos': 'grupos', '16avos': '16avos', 'dieciseisavos': '16avos',
    '8vos': '8vos', 'octavos': '8vos', '4tos': '4tos', 'cuartos': '4tos',
    'semi': 'semis', 'semifinal': 'semis', 'semis': 'semis', '3p': '3p', 'tercer puesto': '3p',
    'final': 'final', 'campeon': 'final',
}

# ---------------- helpers ----------------
def to_int(v):
    try:
        s = str(v).strip()
        if s in ('', '-', 'None', 'nan'): return None
        return int(float(s))
    except Exception:
        return None

def clean_alias(s):
    if not s: return ''
    return str(s).replace('\xa0', '').strip().upper().lstrip('@').replace('  ', ' ')

def norm(s):
    s = unicodedata.normalize('NFKD', str(s).strip()).encode('ascii', 'ignore').decode()
    return s.lower()

# ---------------- BD ----------------
print("Conectando a la BD...")
try:
    conn = psycopg2.connect(CONN_BEC); conn.autocommit = False
    capp = psycopg2.connect(CONN_APP)
except Exception as e:
    sys.exit(f"ERROR conexion: {e}\n(Docker corriendo? docker start core-postgres)")
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
cua = capp.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

cua.execute("SELECT id, username FROM users WHERE is_active=TRUE")
uname = {u['username'].lower(): u['id'] for u in cua.fetchall()}
def find_uid(al):
    a = clean_alias(al)
    if not a: return None
    for k, v in uname.items():
        if k.upper().lstrip('@') == a: return v
    for k, v in uname.items():
        if a in k.upper() or k.upper() in a: return v
    return None

cur.execute("SELECT id, nombre, nombre_es FROM equipo")
eid = {}
for eq in cur.fetchall():
    if eq['nombre']:    eid[eq['nombre'].upper().strip()] = eq['id']
    if eq['nombre_es']: eid[eq['nombre_es'].upper().strip()] = eq['id']
EQ_ALIAS = {'FRANCIA':'France','ESPANA':'Spain','ESPAÑA':'Spain','INGLATERRA':'England',
    'ARGENTINA':'Argentina','MARRUECOS':'Morocco','BELGICA':'Belgium','BÉLGICA':'Belgium',
    'NORUEGA':'Norway','SUIZA':'Switzerland','COLOMBIA':'Colombia','MEXICO':'Mexico',
    'MÉXICO':'Mexico','BRASIL':'Brazil','PORTUGAL':'Portugal','PARAGUAY':'Paraguay',
    'CANADA':'Canada','CANADÁ':'Canada','EGIPTO':'Egypt','ESTADOS UNIDOS':'USA',
    'EE UU':'USA','EEUU':'USA','ALEMANIA':'Germany','IRAK':'Iraq','SUECIA':'Sweden',
    'CROACIA':'Croatia','JAPON':'Japan','JAPÓN':'Japan','AUSTRIA':'Austria'}
def find_eid(nom):
    if not nom: return None
    k = str(nom).upper().strip()
    if k in ('', '-'): return None
    if k in eid: return eid[k]
    a = EQ_ALIAS.get(k)
    if a and a.upper() in eid: return eid[a.upper()]
    for kk, vv in eid.items():
        if k in kk or kk in k: return vv
    return None

cur.execute("""SELECT p.numero_fifa, p.id FROM partido p JOIN fase f ON f.id=p.fase_id
               WHERE f.torneo_id=%s AND p.numero_fifa BETWEEN 1 AND 104""", (TID,))
pid_by_nf = {r['numero_fifa']: r['id'] for r in cur.fetchall()}
print(f"Partidos 1..104 en BD: {len(pid_by_nf)}")

APU_COLS = ['pred_local','pred_visitante','pred_amarillas','pred_rojas','pred_var',
            'pred_penales_partido','pred_minuto_gol','pred_penales_local_tanda',
            'pred_penales_visitante_tanda','pred_equipo_clasifica']
cur.execute(f"""SELECT apostador_id, partido_id, {', '.join(APU_COLS)}
                FROM apuesta WHERE partido_id = ANY(%s)""", (list(pid_by_nf.values()),))
db_apu = {(r['apostador_id'], r['partido_id']): dict(r) for r in cur.fetchall()}
print(f"Apuestas por partido en BD: {len(db_apu)}")

GCOLS = ['pred_campeon_id','pred_finalista1_id','pred_finalista2_id','pred_goleador',
         'pred_peor_equipo_id','pred_goleada_ganador','pred_goleada_perdedor',
         'pred_etapa_paraguay','pred_goles_paraguay']
db_glob = {}
try:
    cur.execute(f"SELECT apostador_id, {', '.join(GCOLS)} FROM apuesta_global WHERE torneo_id=%s", (TID,))
    db_glob = {r['apostador_id']: dict(r) for r in cur.fetchall()}
    print(f"Apuestas globales en BD: {len(db_glob)}")
except Exception as e:
    conn.rollback(); print(f"(no pude leer apuesta_global: {e})")

# ---------------- Excel ----------------
wb = openpyxl.load_workbook(EXCEL, data_only=True)
print("\nHojas del Excel:")
for s in wb.sheetnames:
    ws0 = wb[s]
    print(f"   '{s}'  ({ws0.max_row} filas x {ws0.max_column} cols)")
SHEET = next((s for s in wb.sheetnames if 'MASTER' in s.upper()), None)
if not SHEET:
    sys.exit("ERROR: no encontre la hoja 'TBL MASTER' en el Excel.")
ws = wb[SHEET]
print(f"\nUsando hoja: '{SHEET}'")
print("Cabecera (fila 1) de columnas clave:")
for c in [C_PID, C_ALIAS, C_PL, C_PV] + [col for _, col, _ in BONUS_COLS]:
    print(f"   col {c:>2}: {ws.cell(1, c).value}")

# ---------------- comparacion marcador+bonus ----------------
to_insert, to_update = [], []
diff_by_field = Counter()
diff_sample = []
sin_alias = Counter(); partidos_sin_match = set(); comparados = 0; sin_bet_excel = 0

for r in range(2, ws.max_row + 1):
    raw_pid = ws.cell(r, C_PID).value
    pid = str(raw_pid).strip() if raw_pid is not None else ''
    if not (len(pid) == 4 and pid[0] in ('P', 'p') and pid[1:].isdigit()):
        continue
    nf = int(pid[1:])
    if not (1 <= nf <= 104):
        continue
    partido_id = pid_by_nf.get(nf)
    if not partido_id:
        partidos_sin_match.add(pid); continue
    uid = find_uid(ws.cell(r, C_ALIAS).value)
    if not uid:
        sin_alias[clean_alias(ws.cell(r, C_ALIAS).value)] += 1; continue

    pl, pv = to_int(ws.cell(r, C_PL).value), to_int(ws.cell(r, C_PV).value)
    bonus = {}
    for campo, col, tipo in BONUS_COLS:
        raw = ws.cell(r, col).value
        bonus[campo] = find_eid(raw) if tipo == 'equipo' else to_int(raw)
    # fila sin ninguna carga -> el apostador no aposto ese partido
    if pl is None and pv is None and all(v is None for v in bonus.values()):
        sin_bet_excel += 1; continue

    excel_vals = {'pred_local': pl if pl is not None else 0,
                  'pred_visitante': pv if pv is not None else 0, **bonus}
    db_bet = db_apu.get((uid, partido_id))
    if db_bet is None:
        to_insert.append((uid, partido_id, excel_vals, pid, clean_alias(ws.cell(r, C_ALIAS).value), nf))
        diff_by_field['(apuesta nueva)'] += 1
        continue
    comparados += 1
    difs = [(c, db_bet[c], excel_vals[c]) for c in APU_COLS if db_bet[c] != excel_vals[c]]
    if difs:
        for c, _, _ in difs:
            diff_by_field[c] += 1
        if len(diff_sample) < 40:
            diff_sample.append((pid, clean_alias(ws.cell(r, C_ALIAS).value), difs))
        to_update.append((uid, partido_id, difs, pid, clean_alias(ws.cell(r, C_ALIAS).value), nf))

# ---------------- comparacion globales ----------------
glob_excel = {}   # uid -> {campo_logico: valor_raw}
glob_alias_sin = Counter()
for r in range(2, ws.max_row + 1):
    raw_pid = ws.cell(r, C_PID).value
    pid = str(raw_pid).strip() if raw_pid is not None else ''
    if pid not in GLOB_PID:
        continue
    uid = find_uid(ws.cell(r, C_ALIAS).value)
    if not uid:
        glob_alias_sin[clean_alias(ws.cell(r, C_ALIAS).value)] += 1; continue
    val = ws.cell(r, C_GLOB_VAL).value
    if val is None or str(val).strip() == '':
        continue
    glob_excel.setdefault(uid, {})[GLOB_PID[pid]] = val

def build_glob_row(d):
    campeon = find_eid(d.get('campeon'))
    fin2 = find_eid(d.get('finalista2'))
    etapa_raw = d.get('etapa_py')
    etapa = FASE_NORM.get(norm(etapa_raw), str(etapa_raw).strip()) if etapa_raw else None
    return {
        'pred_campeon_id': campeon,
        'pred_finalista1_id': campeon,          # A: campeon es tambien finalista1 (igual que importador previo)
        'pred_finalista2_id': fin2,
        'pred_goleador': (str(d['goleador']).strip() if d.get('goleador') not in (None, '') else None),
        'pred_peor_equipo_id': find_eid(d.get('peor')),
        'pred_goleada_ganador': to_int(d.get('goleada_gan')),
        'pred_goleada_perdedor': to_int(d.get('goleada_per')),
        'pred_etapa_paraguay': etapa,
        'pred_goles_paraguay': to_int(d.get('goles_py')),
    }

glob_to_write = []   # (uid, row_dict, diffs)
glob_diff_field = Counter()
for uid, d in glob_excel.items():
    newrow = build_glob_row(d)
    cur_row = db_glob.get(uid)
    if cur_row is None:
        glob_to_write.append((uid, newrow, [('(global nueva)', None, None)]))
        glob_diff_field['(global nueva)'] += 1
        continue
    difs = [(c, cur_row.get(c), newrow[c]) for c in GCOLS if cur_row.get(c) != newrow[c]]
    if difs:
        for c, _, _ in difs:
            glob_diff_field[c] += 1
        glob_to_write.append((uid, newrow, difs))

# ---------------- reporte ----------------
uid2name = {v: k for k, v in uname.items()}
print("\n" + "=" * 74)
print("VERIFICACION APUESTAS  (Excel fin de torneo  vs  BD torneo 2)")
print("=" * 74)
print(f"Apuestas comparadas (existen en ambos): {comparados}")
print(f"Filas Excel sin carga (apostador no aposto, skip): {sin_bet_excel}")
if sin_alias: print(f"Aliases del Excel sin match en users: {dict(sin_alias)}")
if partidos_sin_match: print(f"Partidos del Excel no hallados en BD: {sorted(partidos_sin_match)}")
print(f"\n-- MARCADOR + BONUS --")
print(f"Apuestas a INSERTAR (estan en Excel, faltan en BD): {len(to_insert)}")
print(f"Apuestas a ACTUALIZAR (difieren): {len(to_update)}")
if diff_by_field:
    print("Diferencias por campo:")
    for c, k in diff_by_field.most_common():
        print(f"   {c}: {k}")
if diff_sample:
    print("\nMuestra de diferencias (primeras 40):")
    for pid, al, difs in diff_sample:
        ds = ", ".join(f"{c.replace('pred_','')}: {o}->{n}" for c, o, n in difs)
        print(f"   {pid} {al:<16} {ds}")
if to_insert:
    print("\nMuestra apuestas a INSERTAR (primeras 20):")
    for uid, ppid, vals, pid, al, nf in to_insert[:20]:
        print(f"   {pid} {al:<16} {vals['pred_local']}-{vals['pred_visitante']}")

print(f"\n-- GLOBALES A-G --")
print(f"Apostadores con globales en Excel: {len(glob_excel)}")
if glob_alias_sin: print(f"Aliases globales sin match: {dict(glob_alias_sin)}")
print(f"Globales a escribir (nuevas o distintas): {len(glob_to_write)}")
if glob_diff_field:
    print("Diferencias globales por campo:")
    for c, k in glob_diff_field.most_common():
        print(f"   {c}: {k}")
    print("\nMuestra globales (primeras 20):")
    for uid, newrow, difs in glob_to_write[:20]:
        nm = uid2name.get(uid, str(uid))
        ds = ", ".join(f"{c.replace('pred_','')}:{o}->{n}" for c, o, n in difs[:5])
        print(f"   {nm:<16} {ds}")

TANDA = {'pred_penales_local_tanda', 'pred_penales_visitante_tanda'}
safe_rows = 0; safe_fields = 0; tanda_ruido = 0
safe_field_count = Counter()
for uid, ppid, difs, pid, al, nf in to_update:
    w = [(c, new) for c, old, new in difs if c not in TANDA and new is not None]
    tanda_ruido += sum(1 for c, old, new in difs if c in TANDA)
    if w:
        safe_rows += 1; safe_fields += len(w)
        for c, new in w:
            safe_field_count[c] += 1

print("\n" + "-" * 74)
print(f"RESULTADO blanket ('BD=Excel'): {len(to_insert)} insert, {len(to_update)} update, {len(glob_to_write)} globales.")
print("\nMODO SEGURO (lo que hace --apply por defecto):")
print(f"  Actualiza {safe_rows} apuestas ({safe_fields} campos) = solo predicciones reales:")
for c, k in safe_field_count.most_common():
    print(f"      {c}: {k}")
print(f"  OMITE tanda None->0 en grupos (~{tanda_ruido} campos, sin efecto en scoring).")
print(f"  OMITE {len(to_insert)} inserts sentinela P103/P104 (apostadores sin cargar final/3P).")
print(f"  OMITE {len(glob_to_write)} globales (el Excel viene en blanco -> borraria peor-equipo/campeon validos).")
print("  Nunca sobrescribe un valor de la BD con vacio.")
print("Blanket literal (NO recomendado): agregar --force-all")

if not DO_APPLY:
    print("\n[DRY RUN] No se escribio nada.")
    print("Aplicar seguro:  verificar_apuestas_fin_torneo.py --apply")
    conn.close(); capp.close(); sys.exit(0)

# ---------------- APPLY ----------------
print(f"\nAPLICANDO ({'FORZADO: todo incl. globales/tanda/inserts' if FORCE_ALL else 'SEGURO: solo predicciones reales'})...")
UPSERT = f"""
INSERT INTO apuesta (apostador_id, partido_id, {', '.join(APU_COLS)})
VALUES (%(apostador_id)s, %(partido_id)s, {', '.join('%(' + c + ')s' for c in APU_COLS)})
ON CONFLICT (apostador_id, partido_id) DO UPDATE SET
  {', '.join(f'{c}=EXCLUDED.{c}' for c in APU_COLS)}
"""
GUPSERT = f"""
INSERT INTO apuesta_global (torneo_id, apostador_id, {', '.join(GCOLS)})
VALUES (%(torneo_id)s, %(apostador_id)s, {', '.join('%(' + c + ')s' for c in GCOLS)})
ON CONFLICT (torneo_id, apostador_id) DO UPDATE SET
  {', '.join(f'{c}=EXCLUDED.{c}' for c in GCOLS)}, updated_at=NOW()
"""

# UPDATES: dinamico, solo campos que cambian
upd_rows = upd_fields = 0
for uid, ppid, difs, pid, al, nf in to_update:
    if FORCE_ALL:
        writes = {c: new for c, old, new in difs}
    else:
        writes = {c: new for c, old, new in difs if c not in TANDA and new is not None}
    if not writes:
        continue
    setclause = ', '.join(f"{c}=%({c})s" for c in writes)
    try:
        cur.execute(f"UPDATE apuesta SET {setclause} WHERE apostador_id=%(uid)s AND partido_id=%(pid)s",
                    {**writes, 'uid': uid, 'pid': ppid})
        upd_rows += 1; upd_fields += len(writes)
    except Exception as e:
        conn.rollback(); print(f"  ERROR update {al} {pid}: {e}")
conn.commit()
print(f"Apuestas actualizadas: {upd_rows} filas ({upd_fields} campos)")

# INSERTS: solo en modo forzado; en seguro se omiten (sentinela P103/P104)
if FORCE_ALL:
    ins = 0
    for uid, ppid, vals, pid, al, nf in to_insert:
        try:
            cur.execute(UPSERT, {'apostador_id': uid, 'partido_id': ppid, **vals}); ins += 1
        except Exception as e:
            conn.rollback(); print(f"  ERROR insert {al} {pid}: {e}")
    conn.commit()
    print(f"Apuestas insertadas: {ins}")
else:
    print(f"Inserts OMITIDOS (sentinela P103/P104): {len(to_insert)}")

# GLOBALES: solo en modo forzado; en seguro se preservan las de la BD
if FORCE_ALL:
    gw = 0
    for uid, newrow, difs in glob_to_write:
        try:
            cur.execute(GUPSERT, {'torneo_id': TID, 'apostador_id': uid, **newrow}); gw += 1
        except Exception as e:
            conn.rollback(); print(f"  ERROR global uid={uid}: {e}")
    conn.commit()
    print(f"Globales escritas: {gw}")
else:
    print(f"Globales OMITIDAS: {len(glob_to_write)} (el Excel viene en blanco; se preserva la BD)")

print("\nOK. " + ("BD = Excel (forzado)." if FORCE_ALL else "Predicciones reales sincronizadas; globales/tanda de la BD intactas."))
print("SIGUIENTE PASO:  run_reabrir_y_recalcular.bat")
conn.close(); capp.close()
