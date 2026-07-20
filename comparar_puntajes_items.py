# -*- coding: utf-8 -*-
"""
comparar_puntajes_items.py <fase>
Analisis FINO de los PUNTAJES OTORGADOS por apostador x partido, item por item,
entre el Excel (hoja '50- TBL MASTER') y la BD (puntaje_detalle).

Para cada DIFERENCIA imprime:
  resultado oficial | apuesta Excel | apuesta BD | puntos Excel | puntos BD

fase = grupos | r32 | octavos | cuartos | semis | todas   (default: todas)
Solo lectura.
"""
import sys, os
from collections import Counter, defaultdict

PHASES = {
    'grupos':  ('10- GRUPOS',        1,  72, False),
    'r32':     ('20- DIECISEISAVOS', 73,  88, True),
    'octavos': ('30- OCTAVOS',       89,  96, True),
    'cuartos': ('40- CUARTOS',       97, 100, True),
    'semis':   ('50- SEMIFINAL',    101, 102, True),
}
ORDER = ['grupos', 'r32', 'octavos', 'cuartos', 'semis']
arg = (sys.argv[1].lower() if len(sys.argv) > 1 else 'todas')
if arg in ('todas', 'all', 'todo'): fases_run = ORDER
elif arg in PHASES: fases_run = [arg]
else: sys.exit(f"fase invalida '{arg}'. Opciones: {', '.join(ORDER)}, todas")

BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = None
# Prioridad: 'CORRECCIONES' > 'SEMIFINAL' > 'TBL PARA CARGAR'. Entre matches, el mas reciente.
_cands = [os.path.join(BASE, f) for f in os.listdir(BASE)
          if f.lower().endswith('.xlsx')
          and any(k in f.upper() for k in ('CORRECCIONES', 'SEMIFINAL', 'TBL PARA CARGAR'))]
def _rank(p):
    u = os.path.basename(p).upper()
    pref = 0 if 'CORRECCIONES' in u else (1 if 'SEMIFINAL' in u else 2)
    return (pref, -os.path.getmtime(p))
if _cands:
    EXCEL_FILE = sorted(_cands, key=_rank)[0]
if not EXCEL_FILE:
    sys.exit(f"ERROR: no se encontro Excel (CORRECCIONES/SEMIFINAL/TBL) en {BASE}")
print(f"Excel: {os.path.basename(EXCEL_FILE)}")

try:
    import openpyxl
except ImportError:
    os.system(f'"{sys.executable}" -m pip install openpyxl --quiet'); import openpyxl
try:
    import psycopg2, psycopg2.extras
except ImportError:
    os.system(f'"{sys.executable}" -m pip install psycopg2-binary --quiet'); import psycopg2, psycopg2.extras

CONN_BEC = "host=localhost port=5432 dbname=becbuc user=app_user password=superpassword"
CONN_APP = "host=localhost port=5432 dbname=app_db  user=app_user password=superpassword"
TID = 2
SHEET = '50- TBL MASTER'
FASE_TXT_SET = {PHASES[f][0] for f in fases_run}
NF_SET = set(); fase_de_nf = {}
for f in fases_run:
    _, a, b, _ = PHASES[f]
    for n in range(a, b + 1): NF_SET.add(n); fase_de_nf[n] = f

conn = psycopg2.connect(CONN_BEC); capp = psycopg2.connect(CONN_APP)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
cua = capp.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

cua.execute("SELECT id, username FROM users WHERE is_active=TRUE")
uname = {u['username'].lower(): u['id'] for u in cua.fetchall()}
def clean_alias(s):
    if not s: return ''
    return str(s).replace('\xa0','').strip().upper().lstrip('@').replace('  ',' ')
def find_uid(al):
    a = clean_alias(al)
    for k, v in uname.items():
        if k.upper().lstrip('@') == a: return v
    for k, v in uname.items():
        if a and (a in k.upper() or k.upper() in a): return v
    return None

# equipos (nombre)
cur.execute("SELECT id, nombre, nombre_es FROM equipo")
ename = {}
for eq in cur.fetchall():
    ename[eq['id']] = eq['nombre'] or eq['nombre_es']
def enm(x):
    return ename.get(x, x) if x is not None else 'None'

# BD puntaje_detalle (puntos)
cur.execute("""
    SELECT pd.apostador_id, p.numero_fifa,
           COALESCE(pd.pts_resultado,0) AS h, COALESCE(pd.pts_marcador,0) AS i,
           COALESCE(pd.pts_amarillas,0) AS j, COALESCE(pd.pts_rojas,0) AS k,
           COALESCE(pd.pts_var,0) AS l, COALESCE(pd.pts_penales_partido,0) AS m,
           COALESCE(pd.pts_minuto,0) AS n, COALESCE(pd.pts_penales_tanda,0) AS o,
           COALESCE(pd.pts_equipo,0) AS p
    FROM puntaje_detalle pd JOIN partido p ON p.id=pd.partido_id
    JOIN fase f ON f.id=p.fase_id WHERE f.torneo_id=%s
""", (TID,))
bd_pts = {(r['apostador_id'], r['numero_fifa']): r for r in cur.fetchall()}

# BD apuesta (predicciones)
cur.execute("""
    SELECT a.apostador_id, p.numero_fifa,
           a.pred_local, a.pred_visitante, a.pred_amarillas, a.pred_rojas, a.pred_var,
           a.pred_penales_partido, a.pred_minuto_gol,
           a.pred_penales_local_tanda, a.pred_penales_visitante_tanda, a.pred_equipo_clasifica
    FROM apuesta a JOIN partido p ON p.id=a.partido_id
    JOIN fase f ON f.id=p.fase_id WHERE f.torneo_id=%s
""", (TID,))
bd_pred = {(r['apostador_id'], r['numero_fifa']): r for r in cur.fetchall()}

# partido (resultado oficial)
cur.execute("""
    SELECT p.numero_fifa, p.goles_local, p.goles_visitante, p.amarillas, p.rojas,
           p.decisiones_var, p.penales_partido, p.minuto_primer_gol,
           p.penales_local, p.penales_visitante, p.equipo_clasificado_id
    FROM partido p JOIN fase f ON f.id=p.fase_id WHERE f.torneo_id=%s
""", (TID,))
off = {r['numero_fifa']: r for r in cur.fetchall()}

def sc(v):
    try:
        s = str(v).strip()
        if s in ('','-','None'): return 0
        return int(float(s))
    except: return 0
def si(v):  # to int o None
    try:
        s = str(v).strip()
        if s in ('','-','None'): return None
        return int(float(s))
    except: return None

# Excel cols (1-based): predicciones y puntos
C = dict(pid=2, fase=7, alias=10, pl=13, pv=15,
         predJ=25, predK=26, predL=27, predM=28, predN=29, predOe1=30, predOe2=31, predQ=32,
         ptsH=40, ptsI=41, ptsJ=42, ptsK=43, ptsL=44, ptsM=45, ptsN=46, ptsOe1=47, ptsOe2=48, ptsP=51)

ITEM_DESC = {'H':'Resultado','I':'Marcador','J':'Amarillas','K':'Rojas','L':'VAR',
             'M':'Penales juego','N':'Minuto 1er gol','O':'Tanda','P':'Equipo pasa'}
ALL_ITEMS = ['H','I','J','K','L','M','N','O','P']

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb[SHEET]

# Auto-detectar la columna de puntos P (CLASIFICADOS) por su encabezado en fila 1,
# porque cambia de posicion entre versiones del Excel (viejo=51, corregido=49).
for _c in range(45, 61):
    _h = ws.cell(1, _c).value
    if _h and 'CLASIFICAD' in str(_h).upper():
        C['ptsP'] = _c
        break
print(f"Col puntos P (clasificados) detectada: {C['ptsP']}")

tot = Counter(); dif = Counter()
detalles = []          # (fase, item, pid, alias, real, predEx, predBD, ptsEx, ptsBD)

for r in range(2, ws.max_row + 1):
    if str(ws.cell(r, C['fase']).value) not in FASE_TXT_SET: continue
    pid = ws.cell(r, C['pid']).value
    if not (isinstance(pid, str) and pid.startswith('P')): continue
    nf = int(pid[1:])
    if nf not in NF_SET: continue
    fase = fase_de_nf[nf]; es_ko = PHASES[fase][3]
    uid = find_uid(ws.cell(r, C['alias']).value)
    if not uid: continue
    b = bd_pts.get((uid, nf))
    if not b: continue
    alias = clean_alias(ws.cell(r, C['alias']).value)
    pr = bd_pred.get((uid, nf), {})
    of = off.get(nf, {})

    # puntos Excel por item
    exP = {
        'H': sc(ws.cell(r, C['ptsH']).value), 'I': sc(ws.cell(r, C['ptsI']).value),
        'J': sc(ws.cell(r, C['ptsJ']).value), 'K': sc(ws.cell(r, C['ptsK']).value),
        'L': sc(ws.cell(r, C['ptsL']).value), 'M': sc(ws.cell(r, C['ptsM']).value),
        'N': sc(ws.cell(r, C['ptsN']).value),
        'O': sc(ws.cell(r, C['ptsOe1']).value) + sc(ws.cell(r, C['ptsOe2']).value),
        'P': sc(ws.cell(r, C['ptsP']).value),
    }
    # resultado / apuesta excel / apuesta bd por item
    marc_real = f"{of.get('goles_local')}-{of.get('goles_visitante')}"
    marc_ex   = f"{si(ws.cell(r,C['pl']).value)}-{si(ws.cell(r,C['pv']).value)}"
    marc_bd   = f"{pr.get('pred_local')}-{pr.get('pred_visitante')}"
    REAL = {
        'H': marc_real, 'I': marc_real,
        'J': of.get('amarillas'), 'K': of.get('rojas'), 'L': of.get('decisiones_var'),
        'M': of.get('penales_partido'), 'N': of.get('minuto_primer_gol'),
        'O': f"{of.get('penales_local')}-{of.get('penales_visitante')}",
        'P': enm(of.get('equipo_clasificado_id')),
    }
    PEX = {
        'H': marc_ex, 'I': marc_ex,
        'J': si(ws.cell(r,C['predJ']).value), 'K': si(ws.cell(r,C['predK']).value),
        'L': si(ws.cell(r,C['predL']).value), 'M': si(ws.cell(r,C['predM']).value),
        'N': si(ws.cell(r,C['predN']).value),
        'O': f"{si(ws.cell(r,C['predOe1']).value)}-{si(ws.cell(r,C['predOe2']).value)}",
        'P': (str(ws.cell(r,C['predQ']).value).strip() if ws.cell(r,C['predQ']).value not in (None,'','-') else 'None'),
    }
    PBD = {
        'H': marc_bd, 'I': marc_bd,
        'J': pr.get('pred_amarillas'), 'K': pr.get('pred_rojas'), 'L': pr.get('pred_var'),
        'M': pr.get('pred_penales_partido'), 'N': pr.get('pred_minuto_gol'),
        'O': f"{pr.get('pred_penales_local_tanda')}-{pr.get('pred_penales_visitante_tanda')}",
        'P': enm(pr.get('pred_equipo_clasifica')),
    }

    for it in ALL_ITEMS:
        if it == 'P' and not es_ko: continue
        vb = b[it.lower()]; ve = exP[it]
        tot[(it, fase)] += 1
        if ve != vb:
            dif[(it, fase)] += 1
            detalles.append((fase, it, pid, alias, REAL[it], PEX[it], PBD[it], ve, vb))

# ---------- REPORTE ----------
print("\n" + "="*100)
print("VERIFICACION PUNTAJES POR ITEM  (Excel TBL MASTER vs BD)  -- solo diferencias")
print("="*100)

# matriz resumen
print("\nRESUMEN item x fase (diferencias / comparaciones):")
hdr = f"{'ITEM':<6}" + "".join(f"{f[:8]:>13}" for f in fases_run) + f"{'TOTAL':>15}"
print(hdr); print("-"*len(hdr))
item_tot = Counter(); item_dif = Counter()
for it in ALL_ITEMS:
    line = f"{it:<6}"; hay=False
    for f in fases_run:
        t=tot[(it,f)]; d=dif[(it,f)]; item_tot[it]+=t; item_dif[it]+=d
        line += (f"{'-':>13}" if t==0 else f"{(str(d)+'/'+str(t)):>13}")
        if t: hay=True
    line += f"{((str(item_dif[it])+'/'+str(item_tot[it])) if item_tot[it] else '-'):>15}"
    if hay: print(line)

# detalle enriquecido por item
print("\n" + "="*100)
print("DETALLE DE DIFERENCIAS  (real = resultado oficial)")
print("="*100)
by = defaultdict(list)
for d in detalles: by[d[1]].append(d)
for it in ALL_ITEMS:
    if it not in by: continue
    lst = by[it]
    print(f"\n### Item {it} ({ITEM_DESC[it]}) - {len(lst)} diferencia(s)")
    print(f"  {'FASE':<8}{'PART':<6}{'APOSTADOR':<18}{'real':>8}   {'ap.Excel':>10}   {'ap.BD':>10}   {'ptsEx':>6}{'ptsBD':>7}")
    for fase, _, pid, alias, real, pex, pbd, ptse, ptsb in sorted(lst, key=lambda x:(x[0],x[2],x[3]))[:120]:
        print(f"  {fase:<8}{pid:<6}{alias:<18}{str(real):>8}   {str(pex):>10}   {str(pbd):>10}   {ptse:>6}{ptsb:>7}")
    if len(lst) > 120:
        print(f"  ... (+{len(lst)-120} mas)")

print("\n" + "="*100)
print(f"TOTAL diferencias: {len(detalles)}")
NOTA = {
    'N': ' (desempate minuto - decision organizacion)',
    'O': ' (tanda: quirk Excel en semis sin penales)',
    'P': ' (escala: Excel plano 1 vs BD reglamento; decision pendiente)',
}
for it in ALL_ITEMS:
    n = item_dif[it]
    if n == 0:
        continue
    print(f"  {it}: {n}{NOTA.get(it,'')}")

try:
    conn.close(); capp.close()
except Exception:
    pass
