# -*- coding: utf-8 -*-
"""
auditoria_excel.py — Generacion del Workbook de auditoria/transparencia (Fase 3).

Movido desde apostador_bets.py (God file). Construye el Excel unico usado por
las salidas de export (Auditoria, Transparencia, Puntos por fase).
Comportamiento identico al original.
"""
from collections import defaultdict
from datetime import datetime, timezone

from sqlalchemy import text

from app.db.session import engine as _app_engine

# Constantes de fase (copiadas de apostador_bets para no acoplar al endpoint)
_PHASE_ORDER = ["grupo", "ronda32", "ronda16", "cuartos", "semis", "tercer_puesto", "final"]
_PHASE_LABELS_FULL = {
    "grupo": "Fase de grupos", "ronda32": "Ronda de 32", "ronda16": "Octavos de final",
    "cuartos": "Cuartos de final", "semis": "Semifinales",
    "tercer_puesto": "Tercer puesto", "final": "Final",
}


async def build_auditoria_workbook(db, torneo_id: int):
    """Construye el Workbook ÚNICO de auditoría usado por TODAS las salidas de
    export (Auditoría, Transparencia, Puntos por fase).

    Estructura:
      - Hoja 1 'Puntaje general': lista de apostadores con la sumatoria de sus
        puntajes a la fecha (marcador + bonus partido + bonus de terceros),
        rankeada de mayor a menor.
      - Una hoja por fase, empezando por 'Fase de grupos'. Dentro, ordenado por
        grupo y por partido; en cada partido los apostadores se agrupan en
        Pleno (marcador exacto) / Ganador (acertó resultado) / Cero acierto.
        A nivel de apostador se muestra el puntaje del partido (marcador + bonus
        por ítem) sumarizando el Total.

    Devuelve (Workbook, torneo_nombre).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    tr = await db.execute(text("SELECT nombre FROM torneo WHERE id=:tid"), {"tid": torneo_id})
    t = tr.one_or_none()
    torneo_nombre = t[0] if t else f"Torneo {torneo_id}"

    # Orden jerárquico del bracket (izquierda→derecha) por número de partido FIFA.
    # MISMO orden visual que pronosticos/resultado (_renderBracketTree).
    KO_BRACKET_ORDER = {
        "ronda32":       [74, 77, 73, 75, 83, 84, 81, 82, 76, 78, 79, 80, 86, 88, 85, 87],
        "ronda16":       [89, 90, 93, 94, 91, 92, 95, 96],
        "cuartos":       [97, 98, 99, 100],
        "semis":         [101, 102],
        "tercer_puesto": [103],
        "final":         [104],
    }
    try:
        from app.services.ko_scoring import build_num_maps
        _maps = await build_num_maps(db, torneo_id)
        _pid2num = _maps.get("pid2num", {})
    except Exception:
        _pid2num = {}

    def _bracket_pos(tipo, pid):
        order = KO_BRACKET_ORDER.get(tipo)
        if not order:
            return None
        num = _pid2num.get(pid)
        try:
            return order.index(num)
        except (ValueError, TypeError):
            return None

    rm = await db.execute(
        text("""
            SELECT pd.apostador_id, pd.fase_id, pd.fase_tipo, pd.fase_nombre,
                   pd.partido_id, pd.multiplicador,
                   pd.pred_local, pd.pred_visitante, pd.real_local, pd.real_visitante,
                   pd.pts_marcador_base, pd.pts_marcador,
                   COALESCE(pd.pts_resultado, 0)         AS pts_resultado,
                   pd.pts_minuto, pd.pts_amarillas, pd.pts_var,
                   COALESCE(pd.pts_rojas, 0)             AS pts_rojas,
                   COALESCE(pd.pts_penales_partido, 0)   AS pts_penales_partido,
                   COALESCE(pd.pts_penales_tanda, 0)     AS pts_penales_tanda,
                   pd.pts_bonus, pd.pts_total,
                   p.jornada,
                   COALESCE(el.nombre_es, el.nombre) AS local_nombre,
                   COALESCE(ev.nombre_es, ev.nombre) AS visit_nombre
            FROM puntaje_detalle pd
            LEFT JOIN partido p ON p.id = pd.partido_id
            LEFT JOIN equipo el ON el.id = p.equipo_local_id
            LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
            WHERE pd.torneo_id = :tid
            ORDER BY pd.fase_id, p.jornada NULLS LAST, pd.partido_id, pd.apostador_id
        """),
        {"tid": torneo_id},
    )
    detalle = [dict(row) for row in rm.mappings()]

    # Nombres de TODOS los apostadores activos (aunque tengan 0 puntos)
    ids_detalle = {d["apostador_id"] for d in detalle}
    async with _app_engine.connect() as conn:
        ar = await conn.execute(text("""
            SELECT u.id, u.username FROM users u
            JOIN user_roles ur ON ur.user_id = u.id
            JOIN roles ro ON ro.id = ur.role_id
            WHERE ro.name = 'apostador' AND u.is_active = TRUE"""))
        user_map = {row["id"]: row["username"] for row in ar.mappings()}
        extra = [i for i in ids_detalle if i not in user_map]
        if extra:
            ur = await conn.execute(text("SELECT id, username FROM users WHERE id = ANY(:ids)"), {"ids": extra})
            for row in ur.mappings():
                user_map[row["id"]] = row["username"]

    # Puntajes globales A-G por apostador
    pg_r = await db.execute(
        text("SELECT * FROM puntaje_global WHERE torneo_id = :tid"),
        {"tid": torneo_id},
    )
    pts_glob_map: dict = {}
    for row in pg_r.mappings():
        pts_glob_map[row["apostador_id"]] = dict(row)

    # Apuestas globales A-G por apostador (para hoja Globales)
    ag_r = await db.execute(
        text("SELECT * FROM apuesta_global WHERE torneo_id = :tid"),
        {"tid": torneo_id},
    )
    apuesta_glob_map: dict = {}
    for row in ag_r.mappings():
        apuesta_glob_map[row["apostador_id"]] = dict(row)

    # ── Estilos ──
    HDR_FILL    = PatternFill("solid", start_color="1a3a5c")
    GRP_FILL    = PatternFill("solid", start_color="243044")
    ALT_FILL    = PatternFill("solid", start_color="1e2535")
    PART_FILL   = PatternFill("solid", start_color="0f2336")
    GRPHDR_FILL = PatternFill("solid", start_color="33240f")
    PLENO_FILL  = PatternFill("solid", start_color="1e4d2b")
    GANA_FILL   = PatternFill("solid", start_color="3a3410")
    CERO_FILL   = PatternFill("solid", start_color="3a1414")
    TOP_FILL    = PatternFill("solid", start_color="1e4d2b")
    W_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
    N_FONT     = Font(name="Calibri", color="E0E0E0", size=9)
    PLENO_FONT = Font(name="Calibri", color="46d17f", bold=True, size=9)
    GANA_FONT  = Font(name="Calibri", color="fbbf24", bold=True, size=9)
    CERO_FONT  = Font(name="Calibri", color="fc7c7c", bold=True, size=9)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(horizontal="left",   vertical="center")
    thin   = Side(style="thin", color="2e3540")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Hoja 1: Puntaje general ──
    # Columnas: H=resultado, I=marcador exacto, J=amarillas, K=rojas, L=VAR, N=minuto, M=penales partido
    _CAT_KEYS = ["pts_resultado", "pts_marcador", "pts_amarillas", "pts_rojas",
                 "pts_var", "pts_minuto", "pts_penales_partido"]
    sub: dict[int, dict] = defaultdict(lambda: {k: 0 for k in _CAT_KEYS})
    for d in detalle:
        s = sub[d["apostador_id"]]
        for k in _CAT_KEYS:
            s[k] += (d.get(k) or 0)
    gen_cols = ["#", "Apostador",
                "H\nResultado", "I\nExacto", "J\nAmar.", "K\nRojas", "L\nVAR", "N\nMinuto", "M\nPen.P.",
                "Sub", "Glob", "Total"]
    gen_w    = [4, 26, 7, 7, 7, 7, 7, 7, 7, 8, 8, 9]
    TITLE_FILL = PatternFill("solid", start_color="1a2840")
    HDR_CAT_FONT = Font(name="Calibri", color="94a3b8", bold=True, size=8)
    ws_g = wb.create_sheet("Puntaje general")
    ws_g.sheet_view.showGridLines = False
    for i, w in enumerate(gen_w, start=1):
        ws_g.column_dimensions[chr(64 + i)].width = w
    ws_g.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(gen_cols))
    ws_g.cell(1, 1, f"Puntaje general — {torneo_nombre}").font = Font(
        name="Calibri", color="E05020", bold=True, size=13)
    ws_g["A2"] = f"Generado: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M UTC')}"
    ws_g["A2"].font = Font(name="Calibri", color="888888", size=9)
    # Leyenda de columnas fila 3
    leyenda = [
        "", "",
        "H=Resultado (gana/pierde/empata)", "I=Marcador exacto", "J=Amarillas exactas",
        "K=Rojas exactas", "L=Decisiones VAR", "N=Minuto 1er gol", "M=Penales en el partido",
        "Sub=H+I+J+K+L+N+M", "Glob=A-G globales", "Total=Sub+Glob"
    ]
    for col, txt in enumerate(leyenda, start=1):
        c = ws_g.cell(3, col, txt)
        c.font = Font(name="Calibri", color="64748b", italic=True, size=7)
    ws_g.row_dimensions[4].height = 28
    for col, h in enumerate(gen_cols, start=1):
        c = ws_g.cell(4, col, h)
        c.font = HDR_CAT_FONT if col > 2 else W_FONT
        c.fill = HDR_FILL; c.alignment = CENTER; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    filas = []
    for uid, nombre in user_map.items():
        s = sub.get(uid, {k: 0 for k in _CAT_KEYS})
        pg = pts_glob_map.get(uid, {})
        globales = pg.get("pts_total") or 0
        sub_pts = sum(s[k] for k in _CAT_KEYS)
        total = sub_pts + globales
        filas.append((nombre, s, globales, sub_pts, total))
    filas.sort(key=lambda x: (-x[4], x[0].lower()))
    for idx, (nombre, s, globales, sub_pts, total) in enumerate(filas, start=1):
        ri = idx + 4
        vals = [idx, nombre,
                s["pts_resultado"] or "", s["pts_marcador"] or "",
                s["pts_amarillas"] or "", s["pts_rojas"] or "",
                s["pts_var"] or "", s["pts_minuto"] or "",
                s["pts_penales_partido"] or "",
                sub_pts or "", globales or "", total]
        fill = TOP_FILL if idx == 1 and total > 0 else (ALT_FILL if ri % 2 == 0 else GRP_FILL)
        for col, val in enumerate(vals, start=1):
            c = ws_g.cell(ri, col, val)
            c.font = N_FONT; c.fill = fill; c.border = BORDER
            c.alignment = LEFT if col == 2 else CENTER

    # ── Hojas por fase ──
    # por_fase[fase_tipo][grupo/fase_nombre][partido_id] -> filas de apostadores
    por_fase: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for d in detalle:
        por_fase[d["fase_tipo"]][d["fase_nombre"]][d["partido_id"]].append(d)
    fases = [ft for ft in _PHASE_ORDER if ft in por_fase] + \
            [ft for ft in por_fase if ft not in _PHASE_ORDER]

    cols   = ["Apostador", "Pronóstico", "H\nRes.", "I\nExact.", "J\nAmar.", "K\nRojas", "L\nVAR", "N\nMin.", "M\nPen.P.", "O\nP.Tanda", "Total"]
    cols_w = [26, 11, 7, 7, 7, 7, 7, 7, 7, 8, 8]
    cat_meta = {
        3: ("✅ PLENO — marcador exacto",   PLENO_FILL, PLENO_FONT),
        1: ("➕ GANADOR — acertó resultado", GANA_FILL,  GANA_FONT),
        0: ("✗ CERO ACIERTO",               CERO_FILL,  CERO_FONT),
    }

    for ft in fases:
        nombre_fase = _PHASE_LABELS_FULL.get(ft, ft)
        ws = wb.create_sheet(nombre_fase[:31])
        ws.sheet_view.showGridLines = False
        for i, w in enumerate(cols_w, start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        ws.cell(1, 1, f"{nombre_fase} — {torneo_nombre}").font = Font(
            name="Calibri", color="E05020", bold=True, size=13)
        ri = 3
        es_grupo = ft == "grupo"
        for grupo_nombre in sorted(por_fase[ft].keys()):
            partidos = por_fase[ft][grupo_nombre]
            if es_grupo:
                ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=len(cols))
                c = ws.cell(ri, 1, f"▣ {grupo_nombre}")
                c.font = Font(name="Calibri", color="FFD27F", bold=True, size=12)
                c.fill = GRPHDR_FILL; c.alignment = LEFT
                ri += 2

            def _pkey(pid):
                if not es_grupo:
                    bp = _bracket_pos(ft, pid)
                    if bp is not None:
                        return (0, bp)
                r0 = partidos[pid][0]
                return (r0.get("jornada") or 0, pid)

            for pid in sorted(partidos.keys(), key=_pkey):
                rows = partidos[pid]
                d0 = rows[0]
                partido = f"{d0.get('local_nombre') or '?'} vs {d0.get('visit_nombre') or '?'}"
                real = (f"{d0['real_local']}-{d0['real_visitante']}"
                        if d0["real_local"] is not None else "—")
                jor = f"J{d0['jornada']} · " if d0.get("jornada") else ""
                ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=len(cols))
                mult = d0['multiplicador'] or 1
                py_marker = "   🇵🇾 PARAGUAY x2" if mult > 1 else ""
                c = ws.cell(ri, 1, f"⚽ {jor}{partido}   ·   Real: {real}   ·   x{mult}{py_marker}")
                c.font = Font(name="Calibri", color="FFD27F" if mult == 1 else "7EE0A0", bold=True, size=11)
                c.fill = PART_FILL; c.alignment = LEFT
                ri += 1
                ws.row_dimensions[ri].height = 28
                for col, h in enumerate(cols, start=1):
                    c = ws.cell(ri, col, h)
                    c.font = W_FONT; c.fill = HDR_FILL; c.border = BORDER
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ri += 1

                buckets = {3: [], 1: [], 0: []}
                for d in rows:
                    base = d["pts_marcador_base"]
                    buckets[3 if base == 3 else (1 if base == 1 else 0)].append(d)

                for cat in (3, 1, 0):
                    grp = buckets[cat]
                    if not grp:
                        continue
                    label, cfill, cfont = cat_meta[cat]
                    ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=len(cols))
                    c = ws.cell(ri, 1, f"{label}   ({len(grp)})")
                    c.font = cfont; c.fill = cfill; c.alignment = LEFT; c.border = BORDER
                    ri += 1
                    grp.sort(key=lambda d: (-(d["pts_total"] or 0),
                                            user_map.get(d["apostador_id"], "").lower()))
                    for d in grp:
                        nombre = user_map.get(d["apostador_id"], f"Usuario {d['apostador_id']}")
                        pred = (f"{d['pred_local']}-{d['pred_visitante']}"
                                if d["pred_local"] is not None else "—")
                        pts_total_row = (
                            (d.get("pts_resultado") or 0) +
                            (d.get("pts_marcador") or 0) +
                            (d.get("pts_amarillas") or 0) +
                            (d.get("pts_rojas") or 0) +
                            (d.get("pts_var") or 0) +
                            (d.get("pts_minuto") or 0) +
                            (d.get("pts_penales_partido") or 0) +
                            (d.get("pts_penales_tanda") or 0)
                        )
                        vals = [nombre, pred,
                                d.get("pts_resultado") or None,
                                d.get("pts_marcador") or None,
                                d.get("pts_amarillas") or None,
                                d.get("pts_rojas") or None,
                                d.get("pts_var") or None,
                                d.get("pts_minuto") or None,
                                d.get("pts_penales_partido") or None,
                                d.get("pts_penales_tanda") or None,
                                pts_total_row or None]
                        fill = ALT_FILL if ri % 2 == 0 else GRP_FILL
                        for col, val in enumerate(vals, start=1):
                            c = ws.cell(ri, col, val if val is not None else "")
                            c.font = N_FONT; c.fill = fill; c.border = BORDER
                            c.alignment = LEFT if col == 1 else CENTER
                        ri += 1
                ri += 1  # separador entre partidos

    # ── Hoja Globales A-G ──
    GLOB_FILL   = PatternFill("solid", start_color="1a2840")
    GLOB_HDR    = PatternFill("solid", start_color="1e3a5c")
    GLOB_PT_FNT = Font(name="Calibri", color="46d17f", bold=True, size=9)
    glob_cols   = ["Apostador",
                   "A · Campeón", "B · Fin.1", "B · Fin.2", "C · Goleador",
                   "D · Peor eq.", "E · Gol.G", "E · Gol.P",
                   "F · Etapa Py", "G · Goles Py",
                   "Pts A", "Pts B", "Pts C", "Pts D", "Pts E", "Pts F", "Pts G", "Total"]
    glob_w      = [22, 14, 14, 14, 16, 14, 7, 7, 13, 10, 6, 6, 6, 6, 6, 6, 6, 7]
    ws_gl = wb.create_sheet("Globales")
    ws_gl.sheet_view.showGridLines = False
    for i, w in enumerate(glob_w, start=1):
        ws_gl.column_dimensions[chr(64 + i)].width = w
    ws_gl.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(glob_cols))
    ws_gl.cell(1, 1, f"Pronósticos Globales A-G — {torneo_nombre}").font = Font(
        name="Calibri", color="818cf8", bold=True, size=13)
    ws_gl["A2"] = f"Generado: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M UTC')}"
    ws_gl["A2"].font = Font(name="Calibri", color="888888", size=9)
    for col, h in enumerate(glob_cols, start=1):
        c = ws_gl.cell(4, col, h)
        c.font = W_FONT; c.fill = GLOB_HDR; c.alignment = CENTER; c.border = BORDER

    # Mapeo equipo_id -> nombre (para los selectores)
    eq_r = await db.execute(
        text("SELECT id, COALESCE(nombre_es, nombre) AS nombre FROM equipo ORDER BY nombre"),
        {},
    )
    eq_map = {row["id"]: row["nombre"] for row in eq_r.mappings()}

    glob_filas = []
    for uid, nombre in user_map.items():
        ag = apuesta_glob_map.get(uid, {})
        pg = pts_glob_map.get(uid, {})
        if not ag and not pg:
            continue
        def _eq(eid): return eq_map.get(eid, f"#{eid}") if eid else "—"
        def _pt(k): v = pg.get(k); return v if v else ""
        row_vals = [
            nombre,
            _eq(ag.get("pred_campeon_id")),
            _eq(ag.get("pred_finalista1_id")),
            _eq(ag.get("pred_finalista2_id")),
            ag.get("pred_goleador") or "—",
            _eq(ag.get("pred_peor_equipo_id")),
            ag.get("pred_goleada_ganador") if ag.get("pred_goleada_ganador") is not None else "—",
            ag.get("pred_goleada_perdedor") if ag.get("pred_goleada_perdedor") is not None else "—",
            ag.get("pred_etapa_paraguay") or "—",
            ag.get("pred_goles_paraguay") if ag.get("pred_goles_paraguay") is not None else "—",
            _pt("pts_campeon"), _pt("pts_finalistas"), _pt("pts_goleador"), _pt("pts_peor_equipo"),
            _pt("pts_goleada"), _pt("pts_etapa_paraguay"), _pt("pts_goles_paraguay"),
            pg.get("pts_total") or "",
        ]
        glob_filas.append((pg.get("pts_total") or 0, nombre, row_vals))

    glob_filas.sort(key=lambda x: (-x[0], x[1].lower()))
    for idx, (_, _, row_vals) in enumerate(glob_filas, start=1):
        ri = idx + 4
        fill = ALT_FILL if ri % 2 == 0 else GLOB_FILL
        for col, val in enumerate(row_vals, start=1):
            c = ws_gl.cell(ri, col, val)
            is_pts_col = col >= len(glob_cols) - 7  # últimas 8 = pts cols
            c.font = GLOB_PT_FNT if (is_pts_col and val) else N_FONT
            c.fill = fill; c.border = BORDER
            c.alignment = LEFT if col == 1 else CENTER

    if not glob_filas:
        ws_gl.merge_cells(start_row=5, start_column=1, end_row=5, end_column=len(glob_cols))
        c = ws_gl.cell(5, 1, "Sin pronósticos globales registrados")
        c.font = N_FONT; c.fill = GLOB_FILL; c.alignment = CENTER

    # ── Hoja "pronosticos" ──
    from openpyxl.utils import get_column_letter
    PRON_FONT = Font(name="Calibri", size=9)
    PRON_COLS = [
        "ID", "ID PARTIDO", "NOMBRE PARTIDO", "FASE", "GRUPO",
        "NOMBRE", "ALIAS",
        "EQUIPO 1", "NOM CORTO1",
        "GOLES", "vs", "GOLES",
        "EQUIPO 2", "NOM CORTO2",
        "GANADOR", "EMPATE 1", "EMPATE 2", "PERDEDOR",
        "TXT PRONOSTICO", "TXT EMPATE RESULTADO FINAL", "TXT RESULTADO FINAL",
        "J- AMARILLAS", "K- ROJAS", "L- VAR", "M- PENALES", "N- 1ER GOL",
        "O- TANDA PENALES A", "P- TANDA PENALES B", "Q- QUIEN CLASIFICA",
        "A- GANA-EMPATA-PIERDE", "B- RESULTADO EXACTO",
        "C- AMARILLAS", "D- ROJAS", "E- VAR", "F- PENALES", "G- 1ER GOL",
        "H- TANDA PENALES",
        "TOTAL", "ESTADO PARTIDO",
    ]
    PRON_WIDTHS = [
        6, 10, 58, 16, 8,
        22, 22,
        24, 10,
        7, 4, 7,
        24, 10,
        20, 20, 20, 20,
        48, 46, 46,
        13, 9, 8, 10, 10,
        16, 16, 20,
        22, 20,
        12, 9, 7, 10, 10,
        14,
        8, 14,
    ]

    _FASE_LABELS_ES = {
        "grupo": "GRUPOS",
        "ronda32": "DIECISEISAVOS",
        "ronda16": "OCTAVOS",
        "cuartos": "CUARTOS",
        "semis": "SEMIFINAL",
        "tercer_puesto": "TERCER PUESTO",
        "final": "FINAL",
    }

    try:
        pron_r = await db.execute(
            text("""
                SELECT
                    a.apostador_id,
                    a.partido_id,
                    COALESCE(p.numero_fifa, 0)              AS numero_fifa,
                    f.tipo                                  AS fase_tipo,
                    f.nombre                                AS fase_nombre,
                    COALESCE(f.orden, 0)                    AS fase_orden,
                    COALESCE(el.nombre_es, el.nombre)       AS local_nombre,
                    COALESCE(UPPER(LEFT(el.codigo_iso, 3)),
                             UPPER(LEFT(COALESCE(el.nombre_es, el.nombre), 3))) AS local_short,
                    COALESCE(ev.nombre_es, ev.nombre)       AS visit_nombre,
                    COALESCE(UPPER(LEFT(ev.codigo_iso, 3)),
                             UPPER(LEFT(COALESCE(ev.nombre_es, ev.nombre), 3))) AS visit_short,
                    a.pred_local,
                    a.pred_visitante,
                    p.goles_local                           AS real_local,
                    p.goles_visitante                       AS real_visitante,
                    p.estado,
                    COALESCE(a.pred_amarillas, 0)           AS pred_amarillas,
                    COALESCE(a.pred_rojas, 0)               AS pred_rojas,
                    COALESCE(a.pred_var, 0)                 AS pred_var,
                    COALESCE(a.pred_penales_partido, 0)     AS pred_penales_partido,
                    a.pred_minuto_gol,
                    a.pred_penales_local_tanda,
                    a.pred_penales_visitante_tanda,
                    a.pred_equipo_clasifica,
                    COALESCE(pd.pts_resultado, 0)           AS pts_resultado,
                    COALESCE(pd.pts_marcador, 0)            AS pts_marcador,
                    COALESCE(pd.pts_amarillas, 0)           AS pts_amarillas,
                    COALESCE(pd.pts_rojas, 0)               AS pts_rojas,
                    COALESCE(pd.pts_var, 0)                 AS pts_var,
                    COALESCE(pd.pts_penales_partido, 0)     AS pts_penales_partido,
                    COALESCE(pd.pts_minuto, 0)              AS pts_minuto,
                    COALESCE(pd.pts_penales_tanda, 0)       AS pts_penales_tanda
                FROM apuesta a
                JOIN partido p  ON p.id  = a.partido_id
                JOIN fase    f  ON f.id  = p.fase_id
                JOIN equipo  el ON el.id = p.equipo_local_id
                JOIN equipo  ev ON ev.id = p.equipo_visitante_id
                LEFT JOIN puntaje_detalle pd
                    ON  pd.apostador_id = a.apostador_id
                    AND pd.partido_id   = a.partido_id
                    AND pd.torneo_id    = :tid
                WHERE f.torneo_id = :tid
                  AND p.estado = 'finalizado'
                ORDER BY COALESCE(p.numero_fifa, 9999), a.apostador_id
            """),
            {"tid": torneo_id},
        )
        pron_rows = [dict(row) for row in pron_r.mappings()]
    except Exception:
        await db.rollback()
        pron_rows = []

    ws_pr = wb.create_sheet("becbuc audit")
    ws_pr.sheet_view.showGridLines = False
    for i, w in enumerate(PRON_WIDTHS, start=1):
        ws_pr.column_dimensions[get_column_letter(i)].width = w

    for col, h in enumerate(PRON_COLS, start=1):
        c = ws_pr.cell(1, col, h)
        c.font      = PRON_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_pr.auto_filter.ref = f"A1:{get_column_letter(len(PRON_COLS))}1"
    ws_pr.row_dimensions[1].height = 22

    def _pron_txt(num, loc, gl, vis, gv):
        num_s = f"P{num:03d}: " if num else ""
        gl_s  = str(gl) if gl is not None else "?"
        gv_s  = str(gv) if gv is not None else "?"
        return f"{num_s}{loc} {gl_s} vs {gv_s} {vis}"

    def _fase_label_es(tipo, nombre, orden):
        label = _FASE_LABELS_ES.get(tipo, (nombre or tipo).upper())
        return f"{orden}- {label}" if orden else label

    def _grupo_letra(tipo, nombre):
        if tipo != "grupo":
            return ""
        parts = (nombre or "").strip().split()
        return parts[-1].upper() if parts else ""

    for seq, row in enumerate(pron_rows, start=1):
        ri  = seq + 1
        num = row["numero_fifa"] or 0
        id_partido = f"P{num:03d}" if num else "P???"
        alias = user_map.get(row["apostador_id"], f"U{row['apostador_id']}")
        loc   = row["local_nombre"]  or ""
        vis   = row["visit_nombre"]  or ""
        loc_s = (row["local_short"]  or loc[:3] or "").upper()
        vis_s = (row["visit_short"]  or vis[:3] or "").upper()
        pl = row["pred_local"]
        pv = row["pred_visitante"]
        rl = row["real_local"]
        rv = row["real_visitante"]

        if pl is not None and pv is not None:
            if pl > pv:
                ganador, perdedor, empate1, empate2 = loc, vis, "-", "-"
            elif pl < pv:
                ganador, perdedor, empate1, empate2 = vis, loc, "-", "-"
            else:
                # Empate en marcador
                if row["fase_tipo"] == "grupo":
                    # En grupos el empate es resultado final
                    ganador, perdedor, empate1, empate2 = "-", "-", loc, vis
                else:
                    # En KO: 90 min -> prórroga (120 min) -> penales
                    # No existe empate final; clasifica quien indique pred_equipo_clasifica
                    empate1, empate2 = loc, vis  # ambos equipos empataron
                    clas_id = row.get("pred_equipo_clasifica")
                    if clas_id:
                        clas_nom = eq_map.get(clas_id, "")
                        otro_nom = vis if clas_nom == loc else loc
                        ganador, perdedor = clas_nom, otro_nom
                    else:
                        ganador, perdedor = "-", "-"
        else:
            ganador = perdedor = empate1 = empate2 = "-"

        txt_prono   = _pron_txt(num, loc, pl, vis, pv)
        txt_real    = _pron_txt(num, loc, rl, vis, rv) if rl is not None else "-"
        txt_emp_res = txt_real if (rl is not None and rv is not None and rl == rv) else "-"

        clasifica_nom = eq_map.get(row.get("pred_equipo_clasifica") or 0, "") or ""

        pts_total_row = (
            (row.get("pts_resultado") or 0) +
            (row.get("pts_marcador")  or 0) +
            (row.get("pts_amarillas") or 0) +
            (row.get("pts_rojas")     or 0) +
            (row.get("pts_var")       or 0) +
            (row.get("pts_penales_partido") or 0) +
            (row.get("pts_minuto")    or 0) +
            (row.get("pts_penales_tanda") or 0)
        )

        vals = [
            seq,
            id_partido,
            f"{id_partido}- {loc} vs {vis}",
            _fase_label_es(row["fase_tipo"], row["fase_nombre"], row["fase_orden"]),
            _grupo_letra(row["fase_tipo"], row["fase_nombre"]),
            alias,
            alias,
            loc,
            loc_s,
            pl if pl is not None else "",
            "vs",
            pv if pv is not None else "",
            vis,
            vis_s,
            ganador,
            empate1,
            empate2,
            perdedor,
            txt_prono,
            txt_emp_res,
            txt_real,
            row.get("pred_amarillas") if row.get("pred_amarillas") else "",
            row.get("pred_rojas")     if row.get("pred_rojas")     else "",
            row.get("pred_var")       if row.get("pred_var")       else "",
            row.get("pred_penales_partido") if row.get("pred_penales_partido") else "",
            row.get("pred_minuto_gol") if row.get("pred_minuto_gol") else "",
            row.get("pred_penales_local_tanda")     if row.get("pred_penales_local_tanda")     else "",
            row.get("pred_penales_visitante_tanda") if row.get("pred_penales_visitante_tanda") else "",
            clasifica_nom,
            row.get("pts_resultado") or "",
            row.get("pts_marcador")  or "",
            row.get("pts_amarillas") or "",
            row.get("pts_rojas")     or "",
            row.get("pts_var")       or "",
            row.get("pts_penales_partido") or "",
            row.get("pts_minuto")    or "",
            row.get("pts_penales_tanda") or "",
            pts_total_row if pts_total_row else "",
            row.get("estado") or "",
        ]

        for col, val in enumerate(vals, start=1):
            c = ws_pr.cell(ri, col, val)
            c.font = PRON_FONT

    if not wb.sheetnames:
        wb.create_sheet("Sin datos")
    return wb, torneo_nombre
