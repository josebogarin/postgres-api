"""
Endpoints del módulo de apuestas para apostadores.

GET   /bets/partidos/{torneo_id}        → partidos pendientes de grupo
GET   /bets/mis-apuestas/{torneo_id}    → apuestas del usuario autenticado
POST  /bets/apuestas                    → crear o actualizar una apuesta
GET   /bets/periodo/{torneo_id}         → período de apuestas configurado
PATCH /bets/periodo/{torneo_id}         → configurar período (solo admin)
GET   /bets/grupos/{torneo_id}          → standings de grupos para simulación
GET   /bets/mi-bracket/{torneo_id}      → bracket personal simulado
GET   /bets/ranking/{torneo_id}         → ranking de puntos del torneo
GET   /bets/auditoria/{torneo_id}       → lista de snapshots de auditoría
POST  /bets/auditoria/{torneo_id}       → generar Excel de auditoría (admin)
GET   /bets/auditoria/download/{id}     → descargar snapshot
GET   /bets/mensajes/{torneo_id}        → mensajes del admin visibles para todos
POST  /bets/mensajes/{torneo_id}        → crear mensaje (solo admin)
DELETE /bets/mensajes/{torneo_id}/{id}  → eliminar mensaje (solo admin)
"""

import json
import logging
import os
import random
from collections import defaultdict
from datetime import datetime, timezone

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import CurrentUser, CurrentAdmin, OptionalCurrentUser, BECBUCSession as DBSession
from app.db.session import engine as _app_engine

from app.services.bracket_service import (
    simular_standings_usuario,
    seleccionar_mejores_terceros,
    armar_ronda32,
    propagar_ko_usuario,
    _sort_grupo as _sort_grupo_fifa,   # H2H + fair_play + FIFA ranking
)
from app.services import ko_scoring
from app.services.ko_scoring import PHASE_MULT, TIPO_NUM_RANGE, KO_FEEDERS  # noqa: F401
from app.services.scoring import registry as scoring_registry
from app.services.scoring.calculator import ScoringCalculator

router = APIRouter()

# ── Online tracking (in-memory, se resetea al reiniciar servidor) ─────────────
import time as _time

_ONLINE_TTL = 120  # segundos para considerar "activo"

# {user_id: (timestamp_float, source_str)}  source: "web" | "movil"
_online_users: dict[int, tuple[float, str]] = {}


def _is_online(user_id: int) -> str | None:
    """Retorna la fuente ('web'|'movil') si el usuario está activo, None si no."""
    entry = _online_users.get(user_id)
    if entry and (_time.monotonic() - entry[0]) < _ONLINE_TTL:
        return entry[1]
    return None


# ── Schemas ─────────────────────────────────────────────────────────────────

class ApuestaIn(BaseModel):
    partido_id:      int
    pred_local:      int
    pred_visitante:  int
    pred_minuto_gol: int | None = None   # 1-90, None = no predicó
    pred_amarillas:  int | None = None   # 0+,   None = no predicó
    pred_var:        int | None = None   # 0+,   None = no predicó
    pred_penales:    bool | None = None  # legacy bool (no borrar)
    pred_rojas:                   int | None = None  # K: tarjetas rojas
    pred_penales_partido:         int | None = None  # M: penales sancionados en partido
    pred_penales_local_tanda:     int | None = None  # O: goles local en tanda
    pred_penales_visitante_tanda: int | None = None  # O: goles visitante en tanda


class PeriodoIn(BaseModel):
    apuesta_inicio: datetime | None = None
    apuesta_fin:    datetime | None = None


class MensajeIn(BaseModel):
    titulo:    str
    contenido: str


# ── Helper: período de apuestas ──────────────────────────────────────────────

async def _get_periodo(db: DBSession, torneo_id: int) -> dict:
    """Devuelve {apuesta_inicio, apuesta_fin, abierto} para el torneo."""
    try:
        r = await db.execute(
            text("SELECT apuesta_inicio, apuesta_fin FROM torneo WHERE id = :tid"),
            {"tid": torneo_id}
        )
        row = r.one_or_none()
    except Exception:
        # Columnas no migradas aún → período abierto por defecto
        return {"apuesta_inicio": None, "apuesta_fin": None, "abierto": True}

    if not row:
        raise HTTPException(404, "Torneo no encontrado")

    inicio, fin = row[0], row[1]
    now = datetime.now(timezone.utc)
    abierto = True
    if inicio and now < inicio:
        abierto = False
    if fin and now > fin:
        abierto = False
    return {
        "apuesta_inicio": inicio.isoformat() if inicio else None,
        "apuesta_fin":    fin.isoformat()    if fin    else None,
        "abierto":        abierto,
    }


async def _check_admin(current: CurrentUser) -> bool:
    async with _app_engine.connect() as conn:
        ur = await conn.execute(
            text("""
                SELECT r.name FROM users u
                JOIN user_roles ur2 ON ur2.user_id = u.id
                JOIN roles r ON r.id = ur2.role_id
                WHERE u.id = :uid AND r.name IN ('admin','superadmin')
                LIMIT 1
            """),
            {"uid": current.id}
        )
        return ur.one_or_none() is not None


async def _check_superadmin(current: CurrentUser) -> bool:
    async with _app_engine.connect() as conn:
        ur = await conn.execute(
            text("""
                SELECT r.name FROM users u
                JOIN user_roles ur2 ON ur2.user_id = u.id
                JOIN roles r ON r.id = ur2.role_id
                WHERE u.id = :uid AND r.name = 'superadmin'
                LIMIT 1
            """),
            {"uid": current.id}
        )
        return ur.one_or_none() is not None


_audit_logger = logging.getLogger("becbuc.audit")


async def _audit_log(action: str, resource: str, *, current=None,
                     details: dict | None = None, status_code: int = 200,
                     method: str = "EVENT", path: str = "", resource_id: str | None = None):
    """Escribe un registro explícito en audit_logs (app_db) con detalle JSON.

    Usado para eventos de dominio (simulación, avance de fase, reset, intentos de
    modificar fases encerradas). No interrumpe el flujo si falla el registro.
    """
    uid = getattr(current, "id", None)
    uemail = getattr(current, "email", None) or getattr(current, "username", None)
    try:
        async with _app_engine.begin() as conn:
            await conn.execute(
                text("""
                    INSERT INTO audit_logs
                        (user_id, user_email, action, resource, resource_id,
                         method, path, status_code, ip_address, details)
                    VALUES
                        (:uid, :uemail, :action, :resource, :rid,
                         :method, :path, :status, NULL, CAST(:details AS json))
                """),
                {"uid": uid, "uemail": uemail, "action": action, "resource": resource,
                 "rid": resource_id, "method": method, "path": path,
                 "status": status_code,
                 "details": json.dumps(details or {}, ensure_ascii=False, default=str)},
            )
    except Exception:
        _audit_logger.exception("No se pudo escribir audit_log: %s/%s", resource, action)


# ── Endpoints ────────────────────────────────────────────────────────────────

@router.post("/heartbeat", summary="Registra presencia del usuario (online indicator)")
async def heartbeat(source: str = "web", current_user: OptionalCurrentUser = None):
    """source: 'web' | 'movil'. TTL = _ONLINE_TTL segundos. Auth opcional."""
    if current_user is not None:
        _online_users[current_user.id] = (_time.monotonic(), source)
    return {"ok": True}


@router.get("/partidos/{torneo_id}", summary="Partidos de grupo pendientes de jugar")
async def partidos_pendientes(torneo_id: int, db: DBSession) -> list[dict]:
    r = await db.execute(
        text("""
            SELECT
                p.id, p.fase_id, p.jornada, p.fecha, p.estado,
                f.nombre  AS fase_nombre,
                el.id     AS local_id,
                el.nombre AS local_nombre,
                el.nombre_es AS local_nombre_es,
                el.logo_url  AS local_logo,
                ev.id     AS visit_id,
                ev.nombre AS visit_nombre,
                ev.nombre_es AS visit_nombre_es,
                ev.logo_url  AS visit_logo
            FROM partido p
            JOIN fase  f  ON f.id  = p.fase_id  AND f.tipo = 'grupo'
            JOIN equipo el ON el.id = p.equipo_local_id
            JOIN equipo ev ON ev.id = p.equipo_visitante_id
            WHERE p.torneo_id = :tid
              AND p.estado IN ('programado', 'aplazado')
            ORDER BY f.nombre, p.jornada, p.fecha NULLS LAST
        """),
        {"tid": torneo_id}
    )
    return [dict(row) for row in r.mappings()]


@router.get("/mis-apuestas/{torneo_id}", summary="Apuestas del usuario en el torneo")
async def mis_apuestas(
    torneo_id: int,
    current: CurrentUser,
    db: DBSession,
    for_apostador_id: int = None,
) -> list[dict]:
    # Admin puede ver apuestas de otro apostador
    target_id = current.id
    if for_apostador_id and await _check_admin(current):
        target_id = for_apostador_id
    r = await db.execute(
        text("""
            SELECT
                a.id, a.partido_id, a.pred_local, a.pred_visitante,
                a.pred_ganador, a.puntos,
                COALESCE(a.pred_minuto_gol, NULL) AS pred_minuto_gol,
                COALESCE(a.pred_amarillas,  NULL) AS pred_amarillas,
                COALESCE(a.pred_var,        NULL) AS pred_var,
                a.pred_penales,
                COALESCE(a.pred_rojas,                   NULL) AS pred_rojas,
                COALESCE(a.pred_penales_local_tanda,     NULL) AS pred_penales_local_tanda,
                COALESCE(a.pred_penales_visitante_tanda, NULL) AS pred_penales_visitante_tanda,
                COALESCE(a.pred_penales_partido,          NULL) AS pred_penales_partido,
                COALESCE(a.puntos_bonus,    0)    AS puntos_bonus,
                a.updated_at,
                p.goles_local, p.goles_visitante, p.estado AS partido_estado,
                COALESCE(p.minuto_primer_gol, NULL) AS minuto_primer_gol,
                COALESCE(p.amarillas,         NULL) AS amarillas,
                COALESCE(p.decisiones_var,    NULL) AS decisiones_var,
                el.nombre AS local_nombre, el.nombre_es AS local_nombre_es,
                ev.nombre AS visit_nombre, ev.nombre_es AS visit_nombre_es,
                f.nombre  AS fase_nombre
            FROM apuesta a
            JOIN partido p  ON p.id  = a.partido_id
            JOIN fase    f  ON f.id  = p.fase_id
            JOIN equipo el  ON el.id = p.equipo_local_id
            JOIN equipo ev  ON ev.id = p.equipo_visitante_id
            WHERE a.apostador_id = :uid
              AND p.torneo_id    = :tid
            ORDER BY f.nombre, p.jornada, p.fecha NULLS LAST
        """),
        {"uid": target_id, "tid": torneo_id}
    )
    return [dict(row) for row in r.mappings()]


@router.get("/periodo/{torneo_id}", summary="Período de apuestas del torneo")
async def get_periodo(torneo_id: int, db: DBSession) -> dict:
    return await _get_periodo(db, torneo_id)


@router.patch("/periodo/{torneo_id}", summary="Configurar período de apuestas (admin)")
async def set_periodo(torneo_id: int, body: PeriodoIn, current: CurrentUser, db: DBSession) -> dict:
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")
    await db.execute(
        text("UPDATE torneo SET apuesta_inicio=:ini, apuesta_fin=:fin WHERE id=:tid"),
        {"ini": body.apuesta_inicio, "fin": body.apuesta_fin, "tid": torneo_id}
    )
    await db.commit()
    return await _get_periodo(db, torneo_id)


@router.post("/apuestas", summary="Crear o actualizar apuesta")
async def upsert_apuesta(body: ApuestaIn, current: CurrentUser, db: DBSession) -> dict:
    # Verificar partido y obtener numero_fifa
    r = await db.execute(
        text("SELECT estado, torneo_id, COALESCE(numero_fifa, 0) AS numero_fifa FROM partido WHERE id = :pid"),
        {"pid": body.partido_id}
    )
    row = r.one_or_none()
    if not row:
        raise HTTPException(404, "Partido no encontrado")
    if row[0] == "finalizado":
        raise HTTPException(400, "El partido ya finalizó, no se pueden modificar apuestas")

    # Verificar período de apuestas
    periodo = await _get_periodo(db, row[1])
    if not periodo["abierto"]:
        fin = periodo.get("apuesta_fin")
        msg = (f"El período de apuestas cerró el {fin}"
               if fin else "El período de apuestas aún no ha comenzado")
        raise HTTPException(400, msg)

    numero_fifa = row[2] or None
    nombre_apost = getattr(current, "username", None) or getattr(current, "nombre", None) or str(current.id)

    await db.execute(
        text("""
            INSERT INTO apuesta
                (apostador_id, partido_id, nombre_apostador, numero_fifa,
                 pred_local, pred_visitante,
                 pred_minuto_gol, pred_amarillas, pred_var, pred_penales,
                 pred_rojas, pred_penales_partido, pred_penales_local_tanda, pred_penales_visitante_tanda)
            VALUES
                (:uid, :pid, :nombre, :nfifa,
                 :pl, :pv, :pmg, :pam, :pvar, :ppen,
                 :projas, :ppp, :pltanda, :pvtanda)
            ON CONFLICT (apostador_id, partido_id) DO UPDATE SET
                nombre_apostador             = EXCLUDED.nombre_apostador,
                numero_fifa                  = EXCLUDED.numero_fifa,
                pred_local                   = EXCLUDED.pred_local,
                pred_visitante               = EXCLUDED.pred_visitante,
                pred_minuto_gol              = EXCLUDED.pred_minuto_gol,
                pred_amarillas               = EXCLUDED.pred_amarillas,
                pred_var                     = EXCLUDED.pred_var,
                pred_penales                 = EXCLUDED.pred_penales,
                pred_rojas                   = EXCLUDED.pred_rojas,
                pred_penales_partido         = EXCLUDED.pred_penales_partido,
                pred_penales_local_tanda     = EXCLUDED.pred_penales_local_tanda,
                pred_penales_visitante_tanda = EXCLUDED.pred_penales_visitante_tanda,
                updated_at                   = NOW()
        """),
        {
            "uid":    current.id,
            "pid":    body.partido_id,
            "nombre": nombre_apost,
            "nfifa":  numero_fifa,
            "pl":     body.pred_local,
            "pv":     body.pred_visitante,
            "pmg":    body.pred_minuto_gol,
            "pam":    body.pred_amarillas,
            "pvar":   body.pred_var,
            "ppen":   body.pred_penales,
            "projas": body.pred_rojas,
            "ppp":    body.pred_penales_partido,
            "pltanda": body.pred_penales_local_tanda,
            "pvtanda": body.pred_penales_visitante_tanda,
        }
    )
    await db.commit()
    return {"ok": True, "partido_id": body.partido_id, "numero_fifa": numero_fifa,
            "pred": f"{body.pred_local} - {body.pred_visitante}"}


@router.post("/resetear-apuestas/{torneo_id}",
             summary="Resetear (anular) todas las apuestas del usuario para el torneo (sin restricciones de estado/periodo)")
async def resetear_apuestas(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """Reset completo del usuario para el torneo: pone en NULL todos los campos de
    predicción (incluyendo bonus y tanda), borra puntaje_detalle, apuesta_global y
    puntaje_global. No verifica estado de partido ni período de apuestas.
    """
    uid = current.id
    tid = torneo_id
    partido_ids_subq = """
        SELECT p.id FROM partido p
        JOIN fase f ON f.id = p.fase_id
        WHERE f.torneo_id = :tid
    """

    # 1. NULL todas las predicciones de apuesta (campos base + bonus + tanda)
    r = await db.execute(
        text(f"""
            UPDATE apuesta SET
                pred_local                   = NULL,
                pred_visitante               = NULL,
                pred_penales                 = NULL,
                pred_minuto_gol              = NULL,
                pred_amarillas               = NULL,
                pred_var                     = NULL,
                pred_rojas                   = NULL,
                pred_penales_partido         = NULL,
                pred_penales_local_tanda     = NULL,
                pred_penales_visitante_tanda = NULL,
                puntos                       = 0,
                puntos_bonus                 = 0,
                updated_at                   = NOW()
            WHERE apostador_id = :uid
              AND partido_id IN ({partido_ids_subq})
        """),
        {"uid": uid, "tid": tid},
    )

    # 2. Borrar puntaje_detalle del usuario (afecta ranking y transparencia)
    await db.execute(
        text(f"""
            DELETE FROM puntaje_detalle
            WHERE apostador_id = :uid
              AND partido_id IN ({partido_ids_subq})
        """),
        {"uid": uid, "tid": tid},
    )

    # 3. Borrar pronósticos globales A-G del usuario
    await db.execute(
        text("DELETE FROM apuesta_global WHERE apostador_id = :uid AND torneo_id = :tid"),
        {"uid": uid, "tid": tid},
    )

    # 4. Borrar puntaje global calculado del usuario
    await db.execute(
        text("DELETE FROM puntaje_global WHERE apostador_id = :uid AND torneo_id = :tid"),
        {"uid": uid, "tid": tid},
    )

    # 5. Borrar ítems de auditoría del usuario (partido + global)
    try:
        await db.execute(
            text(f"""
                DELETE FROM puntaje_item
                WHERE apostador_id = :uid
                  AND (
                    partido_id IN ({partido_ids_subq})
                    OR (torneo_id = :tid AND partido_id IS NULL)
                  )
            """),
            {"uid": uid, "tid": tid},
        )
    except Exception:
        pass  # tabla puede no existir aún

    await db.commit()
    return {"ok": True, "actualizadas": r.rowcount}


@router.get("/grupos/{torneo_id}", summary="Standings de grupos para simulación")
async def grupos_standings(torneo_id: int, db: DBSession) -> list[dict]:
    from math import comb

    # Refrescar el cuadro (PJ/Pts/posición) desde los resultados reales en cada carga
    try:
        await _recalc_participacion(db, torneo_id)
        await db.commit()
    except Exception:
        await db.rollback()

    rf = await db.execute(
        text("SELECT id, nombre FROM fase WHERE torneo_id=:tid AND tipo='grupo' AND nombre NOT ILIKE '%mejores%' ORDER BY nombre"),
        {"tid": torneo_id}
    )
    fases = [dict(r) for r in rf.mappings()]

    result = []
    for fase in fases:
        fid = fase["id"]

        rs = await db.execute(
            text("""
                SELECT e.id AS equipo_id,
                       COALESCE(e.nombre_es, e.nombre) AS nombre,
                       e.logo_url,
                       e.fifa_ranking,
                       pa.pj, pa.pg, pa.pe, pa.pp, pa.gf, pa.gc,
                       (pa.gf - pa.gc) AS gd,
                       pa.pts, pa.clasifica, pa.posicion,
                       COALESCE(pa.fair_play_pts, 0)       AS fair_play_pts,
                       COALESCE(pa.amarillas, 0)            AS amarillas,
                       COALESCE(pa.rojas_directas, 0)       AS rojas_directas,
                       COALESCE(pa.rojas_doble_amarilla, 0) AS rojas_doble_amarilla
                FROM participacion pa
                JOIN equipo e ON e.id = pa.equipo_id
                WHERE pa.fase_id = :fid
                ORDER BY pa.pts DESC, (pa.gf - pa.gc) DESC, pa.gf DESC
            """),
            {"fid": fid}
        )
        standings = [dict(r) for r in rs.mappings()]

        rp = await db.execute(
            text("""
                SELECT
                    p.id, p.estado, p.jornada, p.fecha,
                    p.equipo_local_id     AS local_id,
                    p.equipo_visitante_id AS visit_id,
                    p.goles_local, p.goles_visitante,
                    COALESCE(p.amarillas, NULL)          AS amarillas,
                    COALESCE(p.decisiones_var, NULL)     AS decisiones_var,
                    COALESCE(p.minuto_primer_gol, NULL)  AS minuto_primer_gol,
                    COALESCE(p.rojas, NULL)              AS rojas,
                    COALESCE(p.penales_partido, NULL)    AS penales_partido,
                    p.minuto_actual,
                    COALESCE(p.numero_fifa, 0)           AS numero_fifa,
                    el.nombre    AS local_nombre,
                    el.nombre_es AS local_nombre_es,
                    el.logo_url  AS local_logo,
                    ev.nombre    AS visit_nombre,
                    ev.nombre_es AS visit_nombre_es,
                    ev.logo_url  AS visit_logo
                FROM partido p
                JOIN equipo el ON el.id = p.equipo_local_id
                JOIN equipo ev ON ev.id = p.equipo_visitante_id
                WHERE p.fase_id = :fid
                ORDER BY p.jornada, p.fecha NULLS LAST
            """),
            {"fid": fid}
        )
        partidos = [dict(r) for r in rp.mappings()]

        n_equipos = len(standings)
        esperados = comb(n_equipos, 2) if n_equipos >= 2 else 0
        partidos_faltantes = max(0, esperados - len(partidos))

        result.append({
            "fase_id":            fid,
            "fase_nombre":        fase["nombre"],
            "standings":          standings,
            "partidos":           partidos,
            "partidos_esperados": esperados,
            "partidos_faltantes": partidos_faltantes,
        })
    return result


@router.get("/mejores-terceros-provisorios/{torneo_id}", summary="Ranking provisional de los 12 terceros (criterio FIFA)")
async def mejores_terceros_provisorios(torneo_id: int, db: DBSession) -> dict:
    """
    Lee los standings reales actuales y rankea los 12 terceros aplicando el criterio
    FIFA 2026: pts → DG → GF → fair_play_pts → fifa_ranking → grupo.
    Devuelve los 8 provisoriamente clasificados y los 4 afuera, con metadatos
    (partidos jugados, pendientes, margen al corte).
    """
    # ── Leer standings reales de todos los grupos ─────────────────────────────
    rf = await db.execute(
        text("SELECT id, nombre FROM fase WHERE torneo_id=:tid AND tipo='grupo' AND nombre NOT ILIKE '%mejores%' ORDER BY nombre"),
        {"tid": torneo_id},
    )
    fases = [dict(r) for r in rf.mappings()]

    if not fases:
        return {"terceros": [], "clasificados": [], "eliminados": [], "corte": None,
                "grupos_completos": 0, "grupos_totales": 0}

    terceros: list[dict] = []
    grupos_con_datos = 0

    for fase in fases:
        fid  = fase["id"]
        letra = fase["nombre"].replace("Grupo ", "").replace("Group ", "").strip()

        rs = await db.execute(
            text("""
                SELECT e.id AS equipo_id,
                       COALESCE(e.nombre_es, e.nombre) AS nombre,
                       e.logo_url, e.codigo_iso,
                       e.fifa_ranking,
                       pa.pj, pa.gf, pa.gc,
                       (pa.gf - pa.gc) AS gd,
                       pa.pts, pa.posicion,
                       COALESCE((
                           SELECT SUM(
                               CASE WHEN p2.equipo_local_id = e.id
                                    THEN COALESCE(p2.local_amarillas,0)   + COALESCE(p2.local_rojas,0)*3
                                    ELSE COALESCE(p2.visitante_amarillas,0) + COALESCE(p2.visitante_rojas,0)*3
                               END)
                           FROM partido p2
                           JOIN fase f2 ON f2.id = p2.fase_id
                                       AND f2.tipo = 'grupo'
                                       AND f2.nombre NOT ILIKE '%mejores%'
                           WHERE p2.torneo_id = :tid
                             AND p2.estado = 'finalizado'
                             AND (p2.equipo_local_id = e.id OR p2.equipo_visitante_id = e.id)
                       ), 0) AS fair_play_pts
                FROM participacion pa
                JOIN equipo e ON e.id = pa.equipo_id
                WHERE pa.fase_id = :fid
                ORDER BY pa.posicion ASC NULLS LAST,
                         pa.pts DESC, (pa.gf - pa.gc) DESC, pa.gf DESC
            """),
            {"fid": fid, "tid": torneo_id},
        )
        equipos = [dict(r) for r in rs.mappings()]
        if not equipos:
            continue

        # Partidos del grupo para saber cuántos faltan + overlay en_juego
        rp = await db.execute(
            text("""
                SELECT p.id, p.estado,
                       p.equipo_local_id, p.equipo_visitante_id,
                       p.goles_local, p.goles_visitante
                FROM partido p WHERE p.fase_id = :fid
            """),
            {"fid": fid},
        )
        partidos_grupo = [dict(r) for r in rp.mappings()]
        total_p     = len(partidos_grupo)
        finalizados = sum(1 for p in partidos_grupo if p["estado"] == "finalizado")
        en_juego_g  = [p for p in partidos_grupo if p["estado"] == "en_juego"
                       and p["goles_local"] is not None]
        pendientes  = total_p - finalizados

        # Overlay provisional: ajustar standings con partidos en_juego
        if en_juego_g:
            eq_map = {e["equipo_id"]: e for e in equipos}
            for lp in en_juego_g:
                gl = lp["goles_local"]  or 0
                gv = lp["goles_visitante"] or 0
                loc_id = lp["equipo_local_id"]
                vis_id = lp["equipo_visitante_id"]
                if gl > gv:   pts_l, pts_v = 3, 0
                elif gl < gv: pts_l, pts_v = 0, 3
                else:         pts_l, pts_v = 1, 1
                for eid, dpts, dgf, dgc in [
                    (loc_id, pts_l, gl, gv),
                    (vis_id, pts_v, gv, gl),
                ]:
                    if eid in eq_map:
                        eq_map[eid]["pts"] = (eq_map[eid].get("pts") or 0) + dpts
                        eq_map[eid]["pj"]  = (eq_map[eid].get("pj")  or 0) + 1
                        eq_map[eid]["gf"]  = (eq_map[eid].get("gf")  or 0) + dgf
                        eq_map[eid]["gc"]  = (eq_map[eid].get("gc")  or 0) + dgc
                        eq_map[eid]["gd"]  = eq_map[eid]["gf"] - eq_map[eid]["gc"]
            # Re-ordenar con standings overlay antes de elegir tercero
            equipos = sorted(eq_map.values(), key=lambda e: (
                -(e.get("pts") or 0),
                -(e.get("gd")  or 0),
                -(e.get("gf")  or 0),
                (e.get("fair_play_pts") or 0),
                (e.get("fifa_ranking")  or 9999),
            ))

        # Tercer lugar: índice 2 en el orden actual
        if len(equipos) >= 3:
            tercer = equipos[2]
            terceros.append({
                **tercer,
                "grupo":      letra,
                "pendientes": pendientes,
                "pj":         tercer.get("pj") or 0,
                "provisorio": len(en_juego_g) > 0,
            })
            grupos_con_datos += 1

    if not terceros:
        return {"terceros": [], "clasificados": [], "eliminados": [], "corte": None,
                "grupos_completos": 0, "grupos_totales": len(fases)}

    # ── Dedup: si un equipo aparece en más de un grupo (dato corrupto), quedarse con el de mayor pts ──
    seen: dict[int, dict] = {}
    for t in terceros:
        eid = t["equipo_id"]
        if eid not in seen or (t.get("pts") or 0) > (seen[eid].get("pts") or 0):
            seen[eid] = t
    terceros = list(seen.values())

    # ── Aplicar criterio FIFA ─────────────────────────────────────────────────
    terceros.sort(key=lambda e: (
        -(e.get("pts") or 0),
        -(e.get("gd")  or 0),
        -(e.get("gf")  or 0),
        (e.get("fair_play_pts") or 0),
        (e.get("fifa_ranking")  or 9999),
        e["grupo"],
    ))

    # Margen de puntos entre 8º y 9º (corte)
    corte_pts = terceros[7]["pts"] if len(terceros) >= 8 else None
    margen    = None
    if len(terceros) >= 9 and corte_pts is not None:
        margen = (terceros[7].get("pts") or 0) - (terceros[8].get("pts") or 0)

    grupos_completos = sum(1 for t in terceros if t["pendientes"] == 0)

    for i, t in enumerate(terceros):
        t["rank"]    = i + 1
        t["dentro"]  = i < 8
        # ¿Cuántos puntos le faltan para entrar (negativo = ya está dentro)
        t["diff_corte"] = ((t.get("pts") or 0) - (corte_pts or 0)) if corte_pts is not None else 0

    return {
        "terceros":         terceros,
        "clasificados":     [t for t in terceros if t["dentro"]],
        "eliminados":       [t for t in terceros if not t["dentro"]],
        "corte_pts":        corte_pts,
        "margen":           margen,          # pts entre 8º y 9º
        "grupos_totales":   len(fases),
        "grupos_completos": grupos_completos,
    }


@router.get("/bracket-real/{torneo_id}", summary="Bracket real (resultados oficiales) por número FIFA")
async def bracket_real(torneo_id: int, db: DBSession) -> dict:
    """Devuelve los partidos KO reales mapeados por número FIFA (74..104),
    con equipos reales (ya avanzados por _avanzar_bracket), marcador, penales,
    estado y ganador. Sin auth: lectura pública para el dashboard."""
    maps = await ko_scoring.build_num_maps(db, torneo_id)
    pid2num = maps.get("pid2num", {})
    num2tipo = maps.get("num2tipo", {})
    if not pid2num:
        return {"partidos": []}

    rp = await db.execute(
        text("""
            SELECT p.id, p.estado, p.fecha,
                   p.equipo_local_id     AS local_id,
                   p.equipo_visitante_id AS visit_id,
                   p.goles_local, p.goles_visitante,
                   p.penales_local, p.penales_visitante,
                   el.nombre AS local_nombre, el.nombre_es AS local_nombre_es, el.logo_url AS local_logo,
                   COALESCE(el.codigo_iso, '') AS local_iso,
                   ev.nombre AS visit_nombre, ev.nombre_es AS visit_nombre_es, ev.logo_url AS visit_logo,
                   COALESCE(ev.codigo_iso, '') AS visit_iso
            FROM partido p
            JOIN fase f ON f.id = p.fase_id
            LEFT JOIN equipo el ON el.id = p.equipo_local_id
            LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
            WHERE f.torneo_id = :tid AND f.tipo <> 'grupo'
        """),
        {"tid": torneo_id},
    )

    out = []
    for r in rp.mappings():
        num = pid2num.get(r["id"])
        if not num:
            continue
        gl, gv = r["goles_local"], r["goles_visitante"]
        pl, pv = r["penales_local"], r["penales_visitante"]
        fin = r["estado"] == "finalizado"
        ganador = None
        if fin and gl is not None and gv is not None:
            if gl > gv:
                ganador = "local"
            elif gv > gl:
                ganador = "visitante"
            elif pl is not None and pv is not None and pl != pv:
                ganador = "local" if pl > pv else "visitante"
        # Serializar fecha a ISO string con "Z" (UTC) para que JS convierta a hora local CR
        fecha_iso = r["fecha"].strftime("%Y-%m-%dT%H:%M:%SZ") if r["fecha"] else None
        out.append({
            "num":        num,
            "tipo":       num2tipo.get(num),
            "finalizado": fin,
            "en_vivo":    r["estado"] == "en_juego",
            "ganador":    ganador,
            "gl": gl, "gv": gv, "pen_l": pl, "pen_v": pv,
            "fecha":      fecha_iso,
            "provisional": r["local_id"] is None or r["visit_id"] is None,
            "local": ({"id": r["local_id"],
                       "nombre": r["local_nombre_es"] or r["local_nombre"],
                       "logo_url": r["local_logo"],
                       "iso": r["local_iso"]} if r["local_id"] else None),
            "visitante": ({"id": r["visit_id"],
                           "nombre": r["visit_nombre_es"] or r["visit_nombre"],
                           "logo_url": r["visit_logo"],
                           "iso": r["visit_iso"]} if r["visit_id"] else None),
        })
    return {"partidos": out}


@router.get("/mi-bracket/{torneo_id}", summary="Bracket personal simulado del apostador (R32 → Final)")
async def mi_bracket(
    torneo_id: int,
    current: CurrentUser,
    db: DBSession,
    for_apostador_id: int = None,
) -> dict:
    # Admin puede ver bracket de otro apostador
    target_id = current.id
    if for_apostador_id and await _check_admin(current):
        target_id = for_apostador_id
    standings = await simular_standings_usuario(db, target_id, torneo_id)
    if not standings:
        raise HTTPException(404, "No se encontraron grupos para este torneo")
    mejores, eliminados = seleccionar_mejores_terceros(standings)
    r32_bracket = armar_ronda32(standings, mejores)

    # Propagación completa KO (R32 → Final) usando predicciones del apostador.
    # Mismo algoritmo que el bracket real (ko_scoring); ganadores se derivan de
    # pred_local vs pred_visitante; empate/sin predicción → FIFA ranking.
    try:
        ko_bracket = await propagar_ko_usuario(db, target_id, torneo_id, r32_bracket)
    except Exception:
        # Fallback: build minimal bracket from partido_ids only (sin equipos ni predicciones).
        # Esto ocurre si pred_penales aún no fue migrada u otro error en propagación.
        # El frontend puede igual mostrar los inputs de carga (solo falta el nombre del equipo).
        try:
            from app.services.ko_scoring import build_num_maps as _build_num_maps
            maps = await _build_num_maps(db, torneo_id)
            ko_bracket = []
            for num, pid in maps["num2pid"].items():
                tipo = maps["num2tipo"].get(num, "")
                ko_bracket.append({
                    "num": num, "tipo": tipo, "partido_id": pid,
                    "local_id": None, "visit_id": None,
                    "local": None, "visitante": None,
                    "pred_gl": None, "pred_gv": None, "pred_penales": None,
                    "winner_id": None, "loser_id": None,
                })
        except Exception:
            ko_bracket = []

    return {
        "grupos_simulados": list(standings.values()),
        "mejores_terceros": mejores,
        "terceros_eliminados": eliminados,
        "ronda_32": r32_bracket,
        "ko_bracket": ko_bracket,   # Lista ordenada R32→Final con equipos propagados
    }


@router.get("/stats/{torneo_id}", summary="Estadísticas generales del torneo (para KPIs extra)")
async def stats_torneo(torneo_id: int, db: DBSession) -> dict:
    """
    KPIs rápidos del torneo:
    - total_apostadores, con_apuestas, total_pronosticos
    - pts_lider, pts_ultimo, lider_nombre, ultimo_nombre
    - fase_activa, pts_max_fase (puntos posibles en fase activa)
    """
    total_apostadores = 0
    user_map: dict[int, str] = {}

    async with _app_engine.connect() as conn:
        r = await conn.execute(text("""
            SELECT u.id, u.username FROM users u
            JOIN user_roles ur ON ur.user_id = u.id
            JOIN roles ro ON ro.id = ur.role_id
            WHERE ro.name = 'apostador' AND u.is_active = TRUE
        """))
        for row in r:
            user_map[row[0]] = row[1]
        total_apostadores = len(user_map)

    # Ranking usando puntaje_detalle + puntaje_global
    try:
        rq = await db.execute(
            text("""
                SELECT
                    pd.apostador_id,
                    COALESCE(SUM(pd.pts_resultado),0) + COALESCE(SUM(pd.pts_marcador),0) +
                    COALESCE(SUM(pd.pts_amarillas),0) + COALESCE(SUM(pd.pts_rojas),0) +
                    COALESCE(SUM(pd.pts_var),0)       + COALESCE(SUM(pd.pts_minuto),0) +
                    COALESCE(SUM(pd.pts_penales_partido),0) + COALESCE(SUM(pd.pts_penales_tanda),0)
                        AS pts_partidos
                FROM puntaje_detalle pd
                WHERE pd.torneo_id = :tid
                GROUP BY pd.apostador_id
            """),
            {"tid": torneo_id}
        )
        pts_map = {row[0]: int(row[1] or 0) for row in rq}
    except Exception:
        await db.rollback()
        pts_map = {}

    # Globales
    try:
        gq = await db.execute(
            text("SELECT apostador_id, pts_total FROM puntaje_global WHERE torneo_id = :tid"),
            {"tid": torneo_id}
        )
        for row in gq:
            pts_map[row[0]] = pts_map.get(row[0], 0) + int(row[1] or 0)
    except Exception:
        await db.rollback()

    # Total pronósticos guardados (apuestas con al menos un score)
    try:
        tq = await db.execute(
            text("""
                SELECT COUNT(*) FROM apuesta a
                JOIN partido p ON p.id = a.partido_id
                WHERE p.torneo_id = :tid
                  AND (a.pred_local IS NOT NULL OR a.pred_visitante IS NOT NULL)
            """),
            {"tid": torneo_id}
        )
        total_pronosticos = int(tq.scalar() or 0)
    except Exception:
        await db.rollback()
        total_pronosticos = 0

    # Con apuestas = apostadores que tienen al menos 1 pronóstico
    try:
        cq = await db.execute(
            text("""
                SELECT COUNT(DISTINCT a.apostador_id) FROM apuesta a
                JOIN partido p ON p.id = a.partido_id
                WHERE p.torneo_id = :tid
                  AND (a.pred_local IS NOT NULL OR a.pred_visitante IS NOT NULL)
            """),
            {"tid": torneo_id}
        )
        con_apuestas = int(cq.scalar() or 0)
    except Exception:
        await db.rollback()
        con_apuestas = 0

    # Lider / último
    lider_nombre, lider_pts, ultimo_nombre, ultimo_pts = "—", 0, "—", 0
    if pts_map:
        sorted_pts = sorted(pts_map.items(), key=lambda x: -x[1])
        lid_id, lider_pts   = sorted_pts[0]
        ult_id, ultimo_pts  = sorted_pts[-1]
        lider_nombre  = user_map.get(lid_id,  f"#{lid_id}")
        ultimo_nombre = user_map.get(ult_id,  f"#{ult_id}")

    # Fase activa (nombre de la próxima fase con partidos pendientes)
    try:
        fq = await db.execute(
            text("""
                SELECT f.nombre, f.tipo FROM fase f
                JOIN partido p ON p.fase_id = f.id
                WHERE f.torneo_id = :tid AND p.estado != 'finalizado'
                ORDER BY p.fecha NULLS LAST
                LIMIT 1
            """),
            {"tid": torneo_id}
        )
        fr = fq.first()
        fase_activa = fr[0] if fr else "—"
    except Exception:
        await db.rollback()
        fase_activa = "—"

    # Puntos máximos posibles = max_pts(fase) × partidos finalizados por fase
    # Copa del Mundo 2026: H+I+J+K+L+M+N (+O si KO con tanda)
    _MAX_PTS_FASE = {
        "grupos":   17,   # H(4)+I(8)+J+K+L+M+N
        "16avos":   27,   # H(6)+I(12)+J+K+L+M+N+O(4)
        "8avos":    33,   # H(8)+I(16)+J+K+L+M+N+O(4)
        "4tos":     39,   # H(10)+I(20)+J+K+L+M+N+O(4)
        "semi":     45,   # H(12)+I(24)+J+K+L+M+N+O(4)
        "tercero":  47,   # H(14)+I(28)+J+K+L+M+N (sin tanda)
        "final":    69,   # H(20)+I(40)+J+K+L+M+N+O(4)
    }
    pts_max_posibles = 0
    try:
        mq = await db.execute(
            text("""
                SELECT f.tipo, COUNT(p.id)
                FROM partido p
                JOIN fase f ON f.id = p.fase_id
                WHERE f.torneo_id = :tid AND p.estado = 'finalizado'
                GROUP BY f.tipo
            """),
            {"tid": torneo_id}
        )
        for row in mq:
            fase_tipo = row[0] or "grupos"
            count     = int(row[1] or 0)
            pts_max_posibles += _MAX_PTS_FASE.get(fase_tipo, 17) * count
    except Exception:
        await db.rollback()
        pts_max_posibles = 0

    return {
        "total_apostadores":  total_apostadores,
        "con_apuestas":       con_apuestas,
        "sin_apuestas":       max(0, total_apostadores - con_apuestas),
        "total_pronosticos":  total_pronosticos,
        "lider_nombre":       lider_nombre,
        "lider_pts":          lider_pts,
        "ultimo_nombre":      ultimo_nombre,
        "ultimo_pts":         ultimo_pts,
        "fase_activa":        fase_activa,
        "pts_max_posibles":   pts_max_posibles,
    }


@router.get("/apostadores", summary="Lista de usuarios con rol apostador (público)")
async def listar_apostadores() -> list[dict]:
    """Devuelve todos los usuarios activos con rol 'apostador'."""
    async with _app_engine.connect() as conn:
        r = await conn.execute(
            text("""
                SELECT u.id, u.username
                FROM users u
                JOIN user_roles ur ON ur.user_id = u.id
                JOIN roles ro ON ro.id = ur.role_id
                WHERE ro.name = 'apostador'
                  AND u.is_active = TRUE
                ORDER BY u.username
            """)
        )
        return [{"id": row[0], "username": row[1]} for row in r]


@router.get("/mis-partidos/{torneo_id}", summary="Partidos del torneo con puntajes por categoría del apostador")
async def mis_partidos(
    torneo_id: int,
    current: CurrentUser,
    db: DBSession,
    for_apostador_id: int | None = None,
) -> list[dict]:
    """
    Retorna cada partido del torneo con:
    - Datos del partido (estado, minuto_actual, equipos, goles reales, predicciones)
    - Puntaje por categoría (H-P) desde puntaje_detalle
    - Subtotal por partido
    Admin puede ver datos de otro apostador via for_apostador_id.
    """
    target_id = current.id
    if for_apostador_id and await _check_admin(current):
        target_id = for_apostador_id
    def _sql_partidos(minuto_col: str) -> str:
        return f"""
            SELECT
                p.id                                        AS partido_id,
                p.jornada,
                p.fecha,
                p.estado,
                p.goles_local,
                p.goles_visitante,
                p.amarillas,
                p.rojas,
                p.decisiones_var,
                p.minuto_primer_gol,
                p.penales_local,
                p.penales_visitante,
                COALESCE(p.penales_partido, 0)              AS penales_partido,
                {minuto_col}                                AS minuto_actual,
                f.nombre                                    AS fase_nombre,
                f.tipo                                      AS fase_tipo,
                f.orden                                     AS fase_orden,
                el.id                                       AS local_id,
                COALESCE(el.nombre_es, el.nombre)           AS local_nombre,
                el.logo_url                                 AS local_logo,
                ev.id                                       AS visit_id,
                COALESCE(ev.nombre_es, ev.nombre)           AS visit_nombre,
                ev.logo_url                                 AS visit_logo,
                a.pred_local,
                a.pred_visitante,
                a.pred_minuto_gol,
                a.pred_amarillas,
                a.pred_var,
                a.pred_rojas,
                a.pred_penales,
                a.pred_penales_partido,
                a.pred_penales_local_tanda,
                a.pred_penales_visitante_tanda,
                COALESCE(a.puntos, 0)                       AS puntos_base,
                COALESCE(a.puntos_bonus, 0)                 AS puntos_bonus,
                pd.pts_resultado,
                pd.pts_marcador,
                pd.pts_amarillas,
                pd.pts_var,
                pd.pts_minuto,
                pd.pts_rojas,
                pd.pts_penales_partido,
                pd.pts_penales_tanda,
                COALESCE(pd.pts_equipo, 0)                  AS pts_equipo,
                pd.pts_bonus,
                pd.pts_total
            FROM partido p
            JOIN fase    f  ON f.id  = p.fase_id
            JOIN equipo  el ON el.id = p.equipo_local_id
            JOIN equipo  ev ON ev.id = p.equipo_visitante_id
            LEFT JOIN apuesta a ON a.partido_id = p.id AND a.apostador_id = :uid
            LEFT JOIN puntaje_detalle pd ON pd.partido_id = p.id AND pd.apostador_id = :uid
            LEFT JOIN monitor_partido_estado mpe ON mpe.partido_id = p.id
            WHERE p.torneo_id = :tid
            ORDER BY f.orden, p.jornada NULLS LAST, p.fecha NULLS LAST, p.id
        """
    _qparams = {"uid": target_id, "tid": torneo_id}
    try:
        r = await db.execute(text(_sql_partidos("p.minuto_actual")), _qparams)
    except Exception as _e:
        if "minuto_actual" in str(_e):
            await db.rollback()
            r = await db.execute(text(_sql_partidos("NULL::smallint")), _qparams)
        else:
            raise
    rows = [dict(row) for row in r.mappings()]
    # Calcular subtotal partido = resultado + marcador + bonus (si puntaje_detalle no disponible)
    for row in rows:
        if row.get("pts_total") is None:
            row["pts_total"] = (row.get("puntos_base") or 0) + (row.get("puntos_bonus") or 0)
        # minuto_display
        estado = row.get("estado") or ""
        minuto = row.get("minuto_actual")
        if estado == "finalizado":
            row["minuto_display"] = "Concluido"
        elif estado == "programado":
            row["minuto_display"] = "No iniciado"
        elif minuto is not None:
            row["minuto_display"] = str(int(minuto))
        else:
            row["minuto_display"] = estado.capitalize() if estado else "—"
    return rows


@router.get("/puntaje-items/{torneo_id}", summary="Auditoría granular: una fila por ítem de puntuación (H-P partido, A-G global)")
async def get_puntaje_items(
    torneo_id: int,
    partido_id: int | None = None,
    apostador_id: int | None = None,
    categoria: str | None = None,   # 'partido' | 'global'
    item: str | None = None,         # H-P o A-G
    current: CurrentUser = None,
    db: DBSession = None,
) -> list[dict]:
    """
    Devuelve los ítems de auditoría de puntaje.
    - Apostadores: ven solo sus propios ítems.
    - Admin: pueden filtrar por apostador_id (o ver todos si omiten el filtro).
    Filtros opcionales: partido_id, categoria ('partido'/'global'), item ('H'-'P'/'A'-'G').
    """
    from fastapi import Query
    is_admin = await _check_admin(current)
    uid = apostador_id if (is_admin and apostador_id) else current.id

    conditions = ["pi.torneo_id = :tid"]
    params: dict = {"tid": torneo_id}

    if not is_admin:
        # Usuario normal solo ve sus propios datos
        conditions.append("pi.apostador_id = :uid")
        params["uid"] = current.id
    elif apostador_id:
        conditions.append("pi.apostador_id = :uid")
        params["uid"] = apostador_id

    if partido_id is not None:
        conditions.append("pi.partido_id = :pid")
        params["pid"] = partido_id

    if categoria:
        conditions.append("pi.categoria = :cat")
        params["cat"] = categoria

    if item:
        conditions.append("pi.item = :item")
        params["item"] = item.upper()

    where = " AND ".join(conditions)
    try:
        r = await db.execute(
            text(f"""
                SELECT pi.id, pi.torneo_id, pi.partido_id, pi.apostador_id,
                       pi.categoria, pi.item,
                       pi.fase_tipo, pi.fase_nombre, pi.fecha_partido,
                       pi.local_nombre, pi.visit_nombre,
                       pi.resultado, pi.apuesta, pi.puntaje, pi.multiplicador,
                       pi.updated_at
                FROM puntaje_item pi
                WHERE {where}
                ORDER BY pi.apostador_id, pi.partido_id NULLS LAST, pi.item
            """),
            params,
        )
        return [dict(row) for row in r.mappings()]
    except Exception as e:
        raise HTTPException(500, f"Error al leer puntaje_item: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# GET /bets/ranking-partido/{torneo_id}?partido_id=N
# Ranking de apostadores para un partido específico (desde puntaje_detalle).
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/ranking-partido/{torneo_id}")
async def ranking_partido(torneo_id: int, partido_id: int, db: DBSession) -> list[dict]:
    """Ranking de apostadores para un partido concreto usando puntaje_detalle."""
    try:
        r = await db.execute(
            text("""
                SELECT
                    pd.apostador_id,
                    pd.pts_resultado,
                    pd.pts_marcador,
                    COALESCE(pd.pts_amarillas,       0)::int AS pts_amarillas,
                    COALESCE(pd.pts_rojas,           0)::int AS pts_rojas,
                    COALESCE(pd.pts_var,             0)::int AS pts_var,
                    COALESCE(pd.pts_minuto,          0)::int AS pts_minuto,
                    COALESCE(pd.pts_penales_tanda,   0)::int AS pts_penales_tanda,
                    COALESCE(pd.pts_equipo,          0)::int AS pts_equipo,
                    pd.pts_total,
                    a.pred_local,
                    a.pred_visitante,
                    a.pred_amarillas      AS ap_amarillas,
                    a.pred_rojas          AS ap_rojas,
                    a.pred_var            AS ap_var,
                    a.pred_minuto_gol     AS ap_minuto,
                    p.goles_local,
                    p.goles_visitante,
                    p.amarillas           AS real_amarillas,
                    p.decisiones_var      AS real_var,
                    p.minuto_primer_gol   AS real_minuto,
                    el.nombre             AS equipo_local,
                    ev.nombre             AS equipo_visitante
                FROM puntaje_detalle pd
                JOIN apuesta a  ON a.apostador_id = pd.apostador_id AND a.partido_id = pd.partido_id
                JOIN partido p  ON p.id = pd.partido_id
                LEFT JOIN equipo el ON el.id = p.equipo_local_id
                LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
                WHERE pd.torneo_id = :tid AND pd.partido_id = :pid
                ORDER BY pd.pts_total DESC, pd.apostador_id
            """),
            {"tid": torneo_id, "pid": partido_id},
        )
        rows = [dict(row) for row in r.mappings()]
    except Exception as e:
        await db.rollback()
        raise HTTPException(500, f"Error ranking-partido: {e}")

    if not rows:
        return []

    # Resolver nombres de usuario desde app_db
    ids = [row["apostador_id"] for row in rows]
    async with _app_engine.connect() as conn:
        ur = await conn.execute(
            text("SELECT id, username FROM users WHERE id = ANY(:ids)"),
            {"ids": ids},
        )
        user_map = {row["id"]: row["username"] for row in ur.mappings()}

    for i, row in enumerate(rows):
        row["nombre"]   = user_map.get(row["apostador_id"], f"Usuario {row['apostador_id']}")
        row["posicion"] = i + 1

    return rows


# ─────────────────────────────────────────────────────────────────────────────
# GET /bets/ranking-detalle/{torneo_id}
# Flat apostador×partido data for client-side Excel generation
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/ranking-detalle/{torneo_id}", summary="Detalle plano apostador×partido para Excel")
async def ranking_detalle(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """Todos los puntajes apostador×partido en formato plano + ranking summary."""
    try:
        r = await db.execute(
            text("""
                SELECT
                    pd.partido_id,
                    pd.apostador_id,
                    p.goles_local, p.goles_visitante,
                    COALESCE(el.nombre_es, el.nombre) AS equipo_local,
                    COALESCE(ev.nombre_es, ev.nombre) AS equipo_visitante,
                    f.nombre AS fase,
                    COALESCE(pd.pts_resultado,     0)::int AS pts_resultado,
                    COALESCE(pd.pts_marcador,      0)::int AS pts_marcador,
                    COALESCE(pd.pts_amarillas,     0)::int AS pts_amarillas,
                    COALESCE(pd.pts_rojas,         0)::int AS pts_rojas,
                    COALESCE(pd.pts_var,           0)::int AS pts_var,
                    COALESCE(pd.pts_minuto,        0)::int AS pts_minuto,
                    COALESCE(pd.pts_penales_tanda, 0)::int AS pts_penales_tanda,
                    COALESCE(pd.pts_total,         0)::int AS pts_total,
                    a.pred_local, a.pred_visitante
                FROM puntaje_detalle pd
                JOIN partido p  ON p.id = pd.partido_id
                JOIN fase f     ON f.id = p.fase_id
                LEFT JOIN equipo el ON el.id = p.equipo_local_id
                LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
                LEFT JOIN apuesta a ON a.apostador_id = pd.apostador_id AND a.partido_id = pd.partido_id
                WHERE pd.torneo_id = :tid
                ORDER BY pd.partido_id, pd.apostador_id
            """),
            {"tid": torneo_id},
        )
        rows = [dict(row) for row in r.mappings()]
    except Exception as e:
        await db.rollback()
        raise HTTPException(500, f"Error ranking-detalle: {e}")

    # Resolve apostador names from app_db
    ids = list({row["apostador_id"] for row in rows})
    nombre_map: dict = {}
    if ids:
        async with _app_engine.connect() as conn:
            ur = await conn.execute(
                text("SELECT id, username FROM users WHERE id = ANY(:ids)"),
                {"ids": ids},
            )
            nombre_map = {row["id"]: row["username"] for row in ur.mappings()}

    for row in rows:
        row["nombre"] = nombre_map.get(row["apostador_id"], f"Usuario {row['apostador_id']}")

    return {"detalle": rows}


# ─────────────────────────────────────────────────────────────────────────────
# GET /bets/ranking-export/{torneo_id}
# Excel de ranking: hoja "Puntaje general" + una hoja por fase (AutoFilter)
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/ranking-export/{torneo_id}", summary="Excel de ranking con desglose por fase")
async def ranking_export(torneo_id: int, current: CurrentUser, db: DBSession):
    """Genera workbook Excel con:
    - Hoja 'Puntaje general': tabla ranking con cols H/I/J/K/L/M/N/O + Glob + Total
    - Una hoja por fase: tabla plana apostador×partido con AutoFilter
    """
    import io
    import traceback as _tb
    from datetime import datetime as _dt
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    try:
      return await _ranking_export_inner(torneo_id, current, db)
    except Exception as _ex:
      raise HTTPException(status_code=500, detail=f"ranking-export error: {_ex}\n{_tb.format_exc()}")

async def _ranking_export_inner(torneo_id: int, current, db):
    import io
    from datetime import datetime as _dt
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── 1. Ranking summary (para "Puntaje general") ────────────────────────
    try:
        r_rk = await db.execute(
            text("""
                SELECT pd.apostador_id,
                       COALESCE(SUM(pd.pts_resultado),        0)::int AS cat_h,
                       COALESCE(SUM(pd.pts_marcador),         0)::int AS cat_i,
                       COALESCE(SUM(pd.pts_amarillas),        0)::int AS cat_j,
                       COALESCE(SUM(pd.pts_rojas),            0)::int AS cat_k,
                       COALESCE(SUM(pd.pts_var),              0)::int AS cat_l,
                       COALESCE(SUM(pd.pts_penales_partido),  0)::int AS cat_m,
                       COALESCE(SUM(pd.pts_minuto),           0)::int AS cat_n,
                       COALESCE(SUM(pd.pts_penales_tanda),    0)::int AS cat_o,
                       COALESCE(SUM(pd.pts_total),            0)::int AS pts_partidos
                FROM puntaje_detalle pd
                WHERE pd.torneo_id = :tid
                GROUP BY pd.apostador_id
            """),
            {"tid": torneo_id},
        )
        rk_rows = [dict(row) for row in r_rk.mappings()]
    except Exception:
        await db.rollback()
        rk_rows = []

    # Fetch usernames from app_db (puntaje_detalle.apostador_id → users)
    apostador_ids = list({row["apostador_id"] for row in rk_rows})
    nombre_map: dict = {}
    if apostador_ids:
        async with _app_engine.connect() as conn:
            nr = await conn.execute(
                text("SELECT id, username FROM users WHERE id = ANY(:ids)"),
                {"ids": apostador_ids},
            )
            nombre_map = {row["id"]: row["username"] for row in nr.mappings()}
    # Globales
    try:
        r_glob = await db.execute(
            text("SELECT apostador_id, puntos_total AS pts_globales FROM puntaje_global WHERE torneo_id = :tid"),
            {"tid": torneo_id},
        )
        glob_map = {row["apostador_id"]: row["pts_globales"] for row in r_glob.mappings()}
    except Exception:
        await db.rollback()
        glob_map = {}

    # Clasificados grupos (item P — no incluido en puntaje_detalle para grupos)
    try:
        r_clas = await db.execute(
            text("""
                SELECT apostador_id, pts_obtenidos AS pts_grupos_p
                FROM apostador_clasificados
                WHERE torneo_id = :tid AND fase_tipo = 'grupo'
            """),
            {"tid": torneo_id},
        )
        clas_map = {row["apostador_id"]: int(row["pts_grupos_p"] or 0)
                    for row in r_clas.mappings()}
    except Exception:
        await db.rollback()
        clas_map = {}

    for row in rk_rows:
        row["pts_globales"] = glob_map.get(row["apostador_id"], 0) or 0
        row["pts_grupos_p"] = clas_map.get(row["apostador_id"], 0) or 0
        row["pts_total"] = row["pts_partidos"] + row["pts_globales"] + row["pts_grupos_p"]
    rk_rows.sort(key=lambda r: -r["pts_total"])

    # ── 2. Detalle por fase (para hojas por fase) ───────────────────────────
    try:
        r_det = await db.execute(
            text("""
                SELECT
                    pd.apostador_id,
                    pd.partido_id,
                    f.nombre  AS fase,
                    COALESCE(f.orden, 0) AS fase_orden,
                    COALESCE(el.nombre_es, el.nombre) AS equipo_local,
                    COALESCE(ev.nombre_es, ev.nombre) AS equipo_visitante,
                    p.goles_local,
                    p.goles_visitante,
                    COALESCE(p.amarillas,       0) AS real_amarillas,
                    COALESCE(p.rojas,           0) AS real_rojas,
                    COALESCE(p.decisiones_var,  0) AS real_var,
                    p.penales_partido              AS real_pen_partido,
                    p.minuto_primer_gol            AS real_minuto,
                    p.penales_local                AS real_pen_local,
                    p.penales_visitante            AS real_pen_visitante,
                    COALESCE(p.numero_fifa, 0) AS numero_fifa,
                    a.pred_local,
                    a.pred_visitante,
                    a.pred_amarillas,
                    a.pred_rojas,
                    a.pred_var,
                    a.pred_penales_partido,
                    a.pred_minuto_gol,
                    a.pred_penales_local_tanda,
                    a.pred_penales_visitante_tanda,
                    COALESCE(pd.pts_resultado,        0)::int AS pts_h,
                    COALESCE(pd.pts_marcador,         0)::int AS pts_i,
                    COALESCE(pd.pts_amarillas,        0)::int AS pts_j,
                    COALESCE(pd.pts_rojas,            0)::int AS pts_k,
                    COALESCE(pd.pts_var,              0)::int AS pts_l,
                    COALESCE(pd.pts_penales_partido,  0)::int AS pts_m,
                    COALESCE(pd.pts_minuto,           0)::int AS pts_n,
                    COALESCE(pd.pts_penales_tanda,    0)::int AS pts_o,
                    COALESCE(pd.pts_total,            0)::int AS pts_total
                FROM puntaje_detalle pd
                JOIN partido p  ON p.id  = pd.partido_id
                JOIN fase    f  ON f.id  = p.fase_id
                LEFT JOIN equipo el ON el.id = p.equipo_local_id
                LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
                LEFT JOIN apuesta a  ON a.apostador_id = pd.apostador_id
                                    AND a.partido_id   = pd.partido_id
                WHERE pd.torneo_id = :tid
                ORDER BY fase_orden, pd.partido_id, pd.apostador_id
            """),
            {"tid": torneo_id},
        )
        det_rows = [dict(row) for row in r_det.mappings()]
    except Exception:
        await db.rollback()
        det_rows = []

    # Resolve all usernames from app_db (combine ids from both queries)
    all_ids = list({row["apostador_id"] for row in rk_rows} | {row["apostador_id"] for row in det_rows})
    nombre_map: dict = {}
    if all_ids:
        async with _app_engine.connect() as conn:
            nr = await conn.execute(
                text("SELECT id, username FROM users WHERE id = ANY(:ids)"),
                {"ids": all_ids},
            )
            nombre_map = {row["id"]: row["username"] for row in nr.mappings()}

    for row in rk_rows:
        row["nombre"] = nombre_map.get(row["apostador_id"], f"Apostador {row['apostador_id']}")
    for row in det_rows:
        row["nombre"] = nombre_map.get(row["apostador_id"], f"Apostador {row['apostador_id']}")

    # Group by fase
    from collections import OrderedDict
    fases: dict = OrderedDict()
    for row in det_rows:
        fn = row["fase"] or "Sin fase"
        if fn not in fases:
            fases[fn] = []
        fases[fn].append(row)

    # ── 3. Build workbook ───────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Styles ──
    def _hdr_fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    HDR_GREEN  = _hdr_fill("1A6B45")
    HDR_ITEM   = _hdr_fill("2D4A6B")
    HDR_SUB    = _hdr_fill("3A5A7C")
    FONT_WHITE = Font(color="FFFFFF", bold=True, size=9)
    FONT_BOLD  = Font(bold=True, size=9)
    FONT_NORM  = Font(size=9)
    ALIGN_C    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_L    = Alignment(horizontal="left", vertical="center")

    def _thin_border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    BORDER = _thin_border()

    FILL_GREEN  = _hdr_fill("D6F0E0")  # exacto
    FILL_YELLOW = _hdr_fill("FFF9C4")  # resultado ok
    FILL_GRAY   = _hdr_fill("F5F5F5")  # alterno
    FILL_WHITE  = _hdr_fill("FFFFFF")

    # ── Hoja "Puntaje general" ──────────────────────────────────────────────
    ws_pg = wb.create_sheet("Puntaje general")
    CATS_PG = [
        ("H", "Resultado"),
        ("I", "Exacto"),
        ("J", "Amarillas"),
        ("K", "Rojas"),
        ("L", "VAR"),
        ("M", "Pen. partido"),
        ("N", "Minuto"),
        ("O", "Pen. tanda"),
    ]
    pg_headers = ["#", "Apostador"] + [f"{c[0]}\n{c[1]}" for c in CATS_PG] + ["Glob.", "Total"]
    col_widths_pg = [4, 18] + [8] * len(CATS_PG) + [7, 8]

    for ci, (h, w) in enumerate(zip(pg_headers, col_widths_pg), 1):
        cell = ws_pg.cell(1, ci, h)
        cell.fill = HDR_GREEN
        cell.font = FONT_WHITE
        cell.alignment = ALIGN_C
        cell.border = BORDER
        ws_pg.column_dimensions[get_column_letter(ci)].width = w
    ws_pg.row_dimensions[1].height = 28

    for ri, row in enumerate(rk_rows, 2):
        fill = FILL_GREEN if ri == 2 else (FILL_GRAY if ri % 2 == 0 else FILL_WHITE)
        vals = [
            ri - 1,
            row["nombre"],
            row.get("cat_h", 0),
            row.get("cat_i", 0),
            row.get("cat_j", 0),
            row.get("cat_k", 0),
            row.get("cat_l", 0),
            row.get("cat_m", 0),
            row.get("cat_n", 0),
            row.get("cat_o", 0),
            row["pts_globales"],
            row["pts_total"],
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws_pg.cell(ri, ci, v)
            cell.fill = fill
            cell.font = FONT_BOLD if ci in (1, len(vals)) else FONT_NORM
            cell.alignment = ALIGN_C if ci != 2 else ALIGN_L
            cell.border = BORDER
    ws_pg.freeze_panes = "A2"
    ws_pg.auto_filter.ref = f"A1:{get_column_letter(len(pg_headers))}1"

    # ── Hojas por fase ──────────────────────────────────────────────────────
    # Columns: Apostador | Partido | Pred | Real | H-Pts | I-Pts |
    #          J-Pred | J-Real | J-Pts | K-Pred | K-Real | K-Pts |
    #          L-Pred | L-Real | L-Pts | M-Pred | M-Real | M-Pts |
    #          N-Pred | N-Real | N-Pts | O-Pred-L | O-Pred-V | O-Real-L | O-Real-V | O-Pts | Total

    ITEM_GROUPS = [
        # (label, pred_key_or_func, real_key_or_func, pts_key)
        # H & I share marcador cols — handled specially
        ("J\nAmarillas", "pred_amarillas", "real_amarillas", "pts_j"),
        ("K\nRojas",     "pred_rojas",     "real_rojas",     "pts_k"),
        ("L\nVAR",       "pred_var",       "real_var",       "pts_l"),
        ("M\nPen.juego", "pred_penales_partido", "real_pen_partido", "pts_m"),
        ("N\nMinuto",    "pred_minuto_gol", "real_minuto",   "pts_n"),
    ]

    def _v(x):
        return "" if x is None else x

    def _marcador(local, visitante):
        if local is None and visitante is None:
            return ""
        return f"{_v(local)}-{_v(visitante)}"

    for fase_nombre, fase_filas in fases.items():
        # Truncate sheet name to 31 chars (Excel limit)
        sheet_name = fase_nombre[:31]
        ws = wb.create_sheet(sheet_name)

        # Header row 1: group labels (merged)
        # Cols: 1=Apostador 2=Partido 3=Pred 4=Real 5=H 6=I  [J:7,8,9] [K:10,11,12] [L:13,14,15] [M:16,17,18] [N:19,20,21] [O:22,23,24,25,26] Total=27
        COL_APOS   = 1
        COL_PART   = 2
        COL_PRED   = 3
        COL_REAL   = 4
        COL_H      = 5
        COL_I      = 6
        # J=7,8,9; K=10,11,12; L=13,14,15; M=16,17,18; N=19,20,21
        # O=22,23,24,25 (Pred-L,Pred-V,Real-L,Real-V), O-Pts=25, wait...
        # Let me re-count: 7,8,9 for J; 10,11,12 for K; 13,14,15 for L; 16,17,18 for M; 19,20,21 for N; 22,23,24,25,26 for O(4cols+pts); 27=Total
        N_SIMPLE = len(ITEM_GROUPS)  # 5
        COL_O_BASE = COL_H + 2 + N_SIMPLE * 3  # 5 + 2 + 5*3 = 22
        COL_O_PRED_L  = COL_O_BASE
        COL_O_PRED_V  = COL_O_BASE + 1
        COL_O_REAL_L  = COL_O_BASE + 2
        COL_O_REAL_V  = COL_O_BASE + 3
        COL_O_PTS     = COL_O_BASE + 4
        COL_TOTAL     = COL_O_BASE + 5

        # Row 1: group headers (merged cells)
        # Apostador(1), Partido(2), Marcador(3-4 merged + 5+6), H-I section
        def _gh(ws, c1, c2, label, fill):
            if c2 > c1:
                ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
            cell = ws.cell(1, c1)
            cell.value = label
            cell.fill = fill
            cell.font = FONT_WHITE
            cell.alignment = ALIGN_C
            cell.border = BORDER

        def _rh(ws, r, c, label, fill):
            cell = ws.cell(r, c, label)
            cell.fill = fill
            cell.font = FONT_WHITE
            cell.alignment = ALIGN_C
            cell.border = BORDER

        # Row 1 group headers
        _gh(ws, 1, 1, "Apostador",  HDR_GREEN)
        _gh(ws, 2, 2, "Partido",    HDR_GREEN)
        _gh(ws, 3, 4, "Marcador",   HDR_GREEN)
        _gh(ws, 5, 6, "H / I",      HDR_ITEM)
        for gi, (lbl, *_) in enumerate(ITEM_GROUPS):
            c = COL_H + 2 + gi * 3  # 7, 10, 13, 16, 19
            _gh(ws, c, c + 2, lbl, HDR_ITEM)
        _gh(ws, COL_O_BASE, COL_O_PTS, "O\nPen.tanda", HDR_ITEM)
        _gh(ws, COL_TOTAL, COL_TOTAL, "Total", HDR_GREEN)

        # Row 2: sub-headers
        _rh(ws, 2, 1, "Apostador",  HDR_GREEN)
        _rh(ws, 2, 2, "Partido",    HDR_GREEN)
        _rh(ws, 2, 3, "Pronóst.",   HDR_SUB)
        _rh(ws, 2, 4, "Real",       HDR_SUB)
        _rh(ws, 2, 5, "H (Res.)",   HDR_SUB)
        _rh(ws, 2, 6, "I (Exact.)", HDR_SUB)
        for gi, _ in enumerate(ITEM_GROUPS):
            c = COL_H + 2 + gi * 3
            for ci2, lbl2 in enumerate(["Pronóst.", "Real", "Pts"]):
                _rh(ws, 2, c + ci2, lbl2, HDR_SUB)
        _rh(ws, 2, COL_O_PRED_L, "P.Local",  HDR_SUB)
        _rh(ws, 2, COL_O_PRED_V, "P.Visit.", HDR_SUB)
        _rh(ws, 2, COL_O_REAL_L, "R.Local",  HDR_SUB)
        _rh(ws, 2, COL_O_REAL_V, "R.Visit.", HDR_SUB)
        _rh(ws, 2, COL_O_PTS,    "Pts",      HDR_SUB)
        _rh(ws, 2, COL_TOTAL,    "Total",    HDR_GREEN)

        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 22
        ws.freeze_panes = "A3"

        # Col widths
        ws.column_dimensions[get_column_letter(1)].width = 16
        ws.column_dimensions[get_column_letter(2)].width = 22
        ws.column_dimensions[get_column_letter(3)].width = 9
        ws.column_dimensions[get_column_letter(4)].width = 9
        ws.column_dimensions[get_column_letter(5)].width = 7
        ws.column_dimensions[get_column_letter(6)].width = 7
        for gi in range(N_SIMPLE):
            for off in range(3):
                c = COL_H + 2 + gi * 3 + off
                ws.column_dimensions[get_column_letter(c)].width = 7
        for c in range(COL_O_BASE, COL_TOTAL + 1):
            ws.column_dimensions[get_column_letter(c)].width = 8
        ws.column_dimensions[get_column_letter(COL_TOTAL)].width = 7

        # AutoFilter on row 2
        ws.auto_filter.ref = f"A2:{get_column_letter(COL_TOTAL)}2"

        # Data rows
        for ri, frow in enumerate(fase_filas, 3):
            _pnum = frow.get('numero_fifa') or 0
            _pnum_prefix = f"P{_pnum}  " if _pnum else ""
            partido_lbl = f"{_pnum_prefix}{frow.get('equipo_local','')} vs {frow.get('equipo_visitante','')}"
            pred_marc = _marcador(frow.get("pred_local"), frow.get("pred_visitante"))
            real_marc = _marcador(frow.get("goles_local"), frow.get("goles_visitante"))
            row_fill = FILL_GRAY if ri % 2 == 0 else FILL_WHITE

            def _dc(c, v, fill=None, bold=False, center=True, _ws=ws, _ri=ri, _rf=row_fill):
                cell = _ws.cell(_ri, c)
                cell.value = v
                cell.fill = fill or _rf
                cell.font = Font(bold=bold, size=9)
                cell.alignment = ALIGN_C if center else ALIGN_L
                cell.border = BORDER

            _dc(1, frow["nombre"], center=False)
            _dc(2, partido_lbl, center=False)
            _dc(3, pred_marc)
            _dc(4, real_marc)
            _dc(5, frow["pts_h"])
            _dc(6, frow["pts_i"])

            for gi, (_, pred_key, real_key, pts_key) in enumerate(ITEM_GROUPS):
                c = COL_H + 2 + gi * 3
                _dc(c,     _v(frow.get(pred_key)))
                _dc(c + 1, _v(frow.get(real_key)))
                _dc(c + 2, frow.get(pts_key, 0), bold=(frow.get(pts_key, 0) > 0))

            _dc(COL_O_PRED_L, _v(frow.get("pred_penales_local_tanda")))
            _dc(COL_O_PRED_V, _v(frow.get("pred_penales_visitante_tanda")))
            _dc(COL_O_REAL_L, _v(frow.get("real_pen_local")))
            _dc(COL_O_REAL_V, _v(frow.get("real_pen_visitante")))
            _dc(COL_O_PTS,    frow.get("pts_o", 0), bold=(frow.get("pts_o", 0) > 0))
            pts_t = frow.get("pts_total", 0)
            _dc(COL_TOTAL, pts_t, fill=FILL_GREEN if pts_t > 0 else row_fill, bold=True)

    # ── Save & stream ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = _dt.now().strftime("%Y%m%d_%H%M")
    fname = f"ranking_torneo{torneo_id}_{ts}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# GET /bets/exportar-puntajes/{torneo_id}
# Excel con dos hojas: v_copamundial_puntajes + v_copamundial_puntajes_det
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/exportar-puntajes/{torneo_id}", summary="Excel puntajes: resumen + detalle por partido")
async def exportar_puntajes(torneo_id: int, db: DBSession, current: CurrentUser):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    ALT_FILL  = PatternFill("solid", fgColor="DCE6F1")
    NORM_FILL = PatternFill("solid", fgColor="FFFFFF")
    TOT_FILL  = PatternFill("solid", fgColor="FFF2CC")
    TOT_FONT  = Font(bold=True, name="Calibri", size=10)
    PTS_FILL  = PatternFill("solid", fgColor="E2EFDA")
    PTS_FONT  = Font(name="Calibri", size=10, color="375623")
    CELL_FONT = Font(name="Calibri", size=10)
    CENTER    = Alignment(horizontal="center", vertical="center")
    LEFT      = Alignment(horizontal="left",   vertical="center")
    _s        = Side(style="thin", color="BBBBBB")
    BORDER    = Border(left=_s, right=_s, top=_s, bottom=_s)

    def _is_pts(col):  return col.endswith("_pts") or col in ("total_partido","total_puntos","subtotal_partidos","subtotal_globales")
    def _is_tot(col):  return col in ("total_partido","total_puntos","subtotal_partidos","subtotal_globales")

    def _write_sheet(ws, rows, title):
        ws.title = title
        if not rows:
            ws.cell(1, 1, "Sin datos")
            return
        cols = list(rows[0].keys())
        for c, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=col.replace("_", " ").title())
            cell.font = HDR_FONT; cell.fill = HDR_FILL
            cell.alignment = CENTER; cell.border = BORDER
        for r, row in enumerate(rows, 2):
            alt = (r % 2 == 0)
            for c, col in enumerate(cols, 1):
                val = row[col]
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = BORDER
                cell.font   = TOT_FONT if _is_tot(col) else (PTS_FONT if _is_pts(col) else CELL_FONT)
                cell.alignment = CENTER if isinstance(val, (int, float)) else LEFT
                cell.fill   = TOT_FILL if _is_tot(col) else (PTS_FILL if _is_pts(col) else (ALT_FILL if alt else NORM_FILL))
        # ancho automático
        for c, col in enumerate(cols, 1):
            max_w = max(len(col), max((len(str(r[col] or "")) for r in rows), default=0))
            ws.column_dimensions[get_column_letter(c)].width = min(max_w + 3, 38)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    try:
        r1 = await db.execute(text("SELECT * FROM v_copamundial_puntajes"))
        rows1 = [dict(row._mapping) for row in r1]
    except Exception as ex:
        raise HTTPException(500, f"Error consultando vista resumen: {ex}")
    try:
        # Try to enrich detail view with numero_fifa (requires partido_id in view)
        r2 = await db.execute(text("""
            SELECT v.*, COALESCE(p.numero_fifa, 0) AS numero_fifa
            FROM v_copamundial_puntajes_det v
            LEFT JOIN partido p ON p.id = v.partido_id
        """))
        rows2_raw = [dict(row._mapping) for row in r2]
        # Prepend P# to the 'partido' column text if present
        for row in rows2_raw:
            nf = row.get("numero_fifa") or 0
            if nf and "partido" in row:
                row["partido"] = f"P{nf}  {row['partido']}"
        rows2 = rows2_raw
    except Exception:
        try:
            r2b = await db.execute(text("SELECT * FROM v_copamundial_puntajes_det"))
            rows2 = [dict(row._mapping) for row in r2b]
        except Exception as ex2:
            raise HTTPException(500, f"Error consultando vista detalle: {ex2}")

    wb = Workbook()
    ws1 = wb.active
    _write_sheet(ws1, rows1, "Resumen")
    ws2 = wb.create_sheet()
    _write_sheet(ws2, rows2, "Detalle")

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    ts = _dt.now().strftime("%Y%m%d_%H%M")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="puntajes_copa_{ts}.xlsx"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# GET /bets/live-analytics/{partido_id}
# Analytics en tiempo real: apostadores agrupados por pronóstico + items H-O
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/live-analytics/{partido_id}", summary="Analytics en vivo de un partido")
async def live_analytics(
    partido_id: int,
    db:          DBSession,
    current:     CurrentUser,
) -> dict:
    """
    Retorna grupos de apostadores por pronóstico de marcador + análisis por ítem.
    Visible para todos los usuarios autenticados.
    """
    # ── Partido ──────────────────────────────────────────────────────────────
    r = await db.execute(text("""
        SELECT p.id, p.estado,
               p.goles_local, p.goles_visitante,
               p.minuto_actual,
               COALESCE(p.amarillas,      0) AS amarillas,
               COALESCE(p.decisiones_var, 0) AS decisiones_var,
               COALESCE(p.rojas,          0) AS rojas,
               p.minuto_primer_gol,
               p.penales_local,
               p.penales_visitante,
               el.nombre     AS nombre_local,
               el.nombre_es  AS nombre_local_es,
               COALESCE(el.codigo_iso,'') AS local_iso,
               el.logo_url   AS local_logo,
               ev.nombre     AS nombre_visitante,
               ev.nombre_es  AS nombre_visitante_es,
               COALESCE(ev.codigo_iso,'') AS visita_iso,
               ev.logo_url   AS visita_logo,
               f.tipo        AS fase_tipo,
               f.nombre      AS fase_nombre
        FROM partido p
        JOIN equipo el ON el.id = p.equipo_local_id
        JOIN equipo ev ON ev.id = p.equipo_visitante_id
        JOIN fase   f  ON f.id  = p.fase_id
        WHERE p.id = :pid
    """), {"pid": partido_id})
    row = r.mappings().first()
    if not row:
        raise HTTPException(404, "Partido no encontrado")
    partido = dict(row)

    # ── Apuestas del partido ─────────────────────────────────────────────────
    ra = await db.execute(text("""
        SELECT a.apostador_id,
               a.pred_local, a.pred_visitante,
               a.pred_amarillas,
               a.pred_var,
               a.pred_minuto_gol,
               a.pred_rojas,
               a.pred_penales_local_tanda,
               a.pred_penales_visitante_tanda
        FROM apuesta a
        WHERE a.partido_id = :pid
          AND a.pred_local IS NOT NULL
    """), {"pid": partido_id})
    apuestas_raw = [dict(r) for r in ra.mappings()]

    if not apuestas_raw:
        return {
            "partido":           dict(partido),
            "total_apostadores": 0,
            "grupos_marcador":   [],
            "items":             [],
        }

    # ── Nombres desde app_db ─────────────────────────────────────────────────
    ids = [a["apostador_id"] for a in apuestas_raw]
    async with _app_engine.connect() as conn:
        ur = await conn.execute(
            text("SELECT id, username FROM users WHERE id = ANY(:ids)"),
            {"ids": ids},
        )
        user_map = {row["id"]: row["username"] for row in ur.mappings()}
    for a in apuestas_raw:
        a["nombre"] = user_map.get(a["apostador_id"], f"#{a['apostador_id']}")

    # ── Grupos por marcador ──────────────────────────────────────────────────
    gl = partido["goles_local"]
    gv = partido["goles_visitante"]
    hay_score = gl is not None and gv is not None

    def _res(lo, vi):
        if lo > vi:  return "local"
        if vi > lo:  return "visitante"
        return "empate"

    grupos: dict[tuple, list] = {}
    for a in apuestas_raw:
        grupos.setdefault((a["pred_local"], a["pred_visitante"]), []).append(a)

    grupos_marcador = []
    for (pl, pv), aps in grupos.items():
        es_exacto    = hay_score and pl == gl and pv == gv
        es_resultado = hay_score and _res(pl, pv) == _res(gl, gv)
        grupos_marcador.append({
            "marcador":       f"{pl}–{pv}",
            "pred_local":     pl,
            "pred_visitante": pv,
            "es_exacto":      es_exacto,
            "es_resultado":   es_resultado,
            "apostadores":    [{"id": a["apostador_id"], "nombre": a["nombre"]} for a in aps],
            "count":          len(aps),
        })
    grupos_marcador.sort(key=lambda g: (not g["es_exacto"], not g["es_resultado"], -g["count"]))

    # ── Items ────────────────────────────────────────────────────────────────
    items = []
    total = len(apuestas_raw)

    def _chip(a):
        return {"id": a["apostador_id"], "nombre": a["nombre"]}

    # H — Resultado
    if hay_score:
        rr = _res(gl, gv)
        ok = [a for a in apuestas_raw if _res(a["pred_local"], a["pred_visitante"]) == rr]
        no = [a for a in apuestas_raw if _res(a["pred_local"], a["pred_visitante"]) != rr]
        items.append({"codigo": "H", "label": "Resultado", "real": rr,
                      "aciertos": [_chip(a) for a in ok],
                      "otros": [{**_chip(a), "pred": f"{a['pred_local']}–{a['pred_visitante']}"} for a in no],
                      "total": total})

    # I — Marcador exacto
    if hay_score:
        ok = [a for a in apuestas_raw if a["pred_local"] == gl and a["pred_visitante"] == gv]
        no = [a for a in apuestas_raw if not (a["pred_local"] == gl and a["pred_visitante"] == gv)]
        items.append({"codigo": "I", "label": "Marcador exacto", "real": f"{gl}–{gv}",
                      "aciertos": [_chip(a) for a in ok],
                      "otros": [{**_chip(a), "pred": f"{a['pred_local']}–{a['pred_visitante']}"} for a in no],
                      "total": total})

    # J — Amarillas
    am = partido["amarillas"]
    with_pred = [a for a in apuestas_raw if a["pred_amarillas"] is not None]
    ok = [a for a in with_pred if a["pred_amarillas"] == am]
    no = [a for a in with_pred if a["pred_amarillas"] != am]
    sp = [a for a in apuestas_raw if a["pred_amarillas"] is None]
    items.append({"codigo": "J", "label": "Amarillas", "real": am,
                  "aciertos": [_chip(a) for a in ok],
                  "otros": [{**_chip(a), "pred": a["pred_amarillas"]} for a in no],
                  "sin_pred": [_chip(a) for a in sp], "total": total})

    # K — Rojas
    ro = partido["rojas"]
    with_pred = [a for a in apuestas_raw if a["pred_rojas"] is not None]
    ok = [a for a in with_pred if a["pred_rojas"] == ro]
    no = [a for a in with_pred if a["pred_rojas"] != ro]
    sp = [a for a in apuestas_raw if a["pred_rojas"] is None]
    items.append({"codigo": "K", "label": "Tarjetas rojas", "real": ro,
                  "aciertos": [_chip(a) for a in ok],
                  "otros": [{**_chip(a), "pred": a["pred_rojas"]} for a in no],
                  "sin_pred": [_chip(a) for a in sp], "total": total})

    # L — VAR
    var = partido["decisiones_var"]
    with_pred = [a for a in apuestas_raw if a["pred_var"] is not None]
    ok = [a for a in with_pred if a["pred_var"] == var]
    no = [a for a in with_pred if a["pred_var"] != var]
    sp = [a for a in apuestas_raw if a["pred_var"] is None]
    items.append({"codigo": "L", "label": "Decisiones VAR", "real": var,
                  "aciertos": [_chip(a) for a in ok],
                  "otros": [{**_chip(a), "pred": a["pred_var"]} for a in no],
                  "sin_pred": [_chip(a) for a in sp], "total": total})

    # N — Minuto primer gol
    mg = partido["minuto_primer_gol"]
    if mg is not None:
        with_pred = [a for a in apuestas_raw if a["pred_minuto_gol"] is not None]
        ranking_n = sorted(with_pred, key=lambda a: abs(a["pred_minuto_gol"] - mg))
        sp = [a for a in apuestas_raw if a["pred_minuto_gol"] is None]
        items.append({"codigo": "N", "label": "Minuto primer gol", "real": mg,
                      "ranking": [{**_chip(a), "pred": a["pred_minuto_gol"],
                                   "diff": abs(a["pred_minuto_gol"] - mg)} for a in ranking_n],
                      "sin_pred": [_chip(a) for a in sp], "total": total})

    # O — Tanda de penales (solo si hubo)
    pl_t = partido["penales_local"]
    pv_t = partido["penales_visitante"]
    if pl_t is not None and pv_t is not None:
        with_pred = [a for a in apuestas_raw
                     if a["pred_penales_local_tanda"] is not None
                     and a["pred_penales_visitante_tanda"] is not None]
        ok = [a for a in with_pred
              if a["pred_penales_local_tanda"] == pl_t and a["pred_penales_visitante_tanda"] == pv_t]
        no = [a for a in with_pred
              if not (a["pred_penales_local_tanda"] == pl_t and a["pred_penales_visitante_tanda"] == pv_t)]
        sp = [a for a in apuestas_raw
              if a["pred_penales_local_tanda"] is None or a["pred_penales_visitante_tanda"] is None]
        items.append({"codigo": "O", "label": "Tanda de penales", "real": f"{pl_t}–{pv_t}",
                      "aciertos": [_chip(a) for a in ok],
                      "otros": [{**_chip(a),
                                  "pred": f"{a['pred_penales_local_tanda']}–{a['pred_penales_visitante_tanda']}"}
                                 for a in no],
                      "sin_pred": [_chip(a) for a in sp], "total": total})

    return {
        "partido":           dict(partido),
        "total_apostadores": total,
        "grupos_marcador":   grupos_marcador,
        "items":             items,
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /bets/partidos-finalizados/{torneo_id}
# Lista de partidos finalizados con su numero de orden (para selector de UI).
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/partidos-finalizados/{torneo_id}")
async def partidos_finalizados(torneo_id: int, db: DBSession) -> list[dict]:
    """Devuelve los partidos finalizados del torneo con su número de orden (ROW_NUMBER)."""
    try:
        r = await db.execute(
            text("""
                WITH numerados AS (
                    SELECT
                        p.id,
                        ROW_NUMBER() OVER (ORDER BY f.orden, p.id) AS numero,
                        COALESCE(el.nombre_es, el.nombre) AS equipo_local,
                        COALESCE(ev.nombre_es, ev.nombre) AS equipo_visitante,
                        el.logo_url AS local_logo,
                        ev.logo_url AS visit_logo,
                        p.goles_local,
                        p.goles_visitante,
                        p.amarillas,
                        p.rojas,
                        p.decisiones_var,
                        p.penales_partido,
                        f.nombre   AS fase,
                        p.estado
                    FROM partido p
                    JOIN fase f ON f.id = p.fase_id
                    LEFT JOIN equipo el ON el.id = p.equipo_local_id
                    LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
                    WHERE f.torneo_id = :tid
                )
                SELECT * FROM numerados
                WHERE estado = 'finalizado'
                ORDER BY numero
            """),
            {"tid": torneo_id},
        )
        return [dict(row) for row in r.mappings()]
    except Exception as e:
        await db.rollback()
        raise HTTPException(500, f"Error partidos-finalizados: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# GET /bets/resultados-partidos/{torneo_id}
# Partidos finalizados con TODOS sus items (amarillas, rojas, var, minuto, penales)
# ordenados por fecha DESC (el ultimo partido jugado primero). Para becbuc-live.html.
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/scores-por-apostador/{torneo_id}")
async def scores_por_apostador(torneo_id: int, db: DBSession):
    """Puntaje por apostador desde puntaje_detalle + puntaje_global. Para comparacion con Excel."""
    rows = await db.execute(text("""
        SELECT
            pd.apostador_id,
            an.nombre_apostador AS nombre,
            SUM(COALESCE(pd.pts_resultado,0))        AS h,
            SUM(COALESCE(pd.pts_marcador,0))         AS i,
            SUM(COALESCE(pd.pts_amarillas,0))        AS j,
            SUM(COALESCE(pd.pts_rojas,0))            AS k,
            SUM(COALESCE(pd.pts_var,0))              AS l,
            SUM(COALESCE(pd.pts_penales_partido,0))  AS m,
            SUM(COALESCE(pd.pts_minuto,0))           AS n,
            SUM(COALESCE(pd.pts_penales_tanda,0))    AS o,
            SUM(COALESCE(pd.pts_resultado,0)+COALESCE(pd.pts_marcador,0)+
                COALESCE(pd.pts_amarillas,0)+COALESCE(pd.pts_rojas,0)+
                COALESCE(pd.pts_var,0)+COALESCE(pd.pts_penales_partido,0)+
                COALESCE(pd.pts_minuto,0)+COALESCE(pd.pts_penales_tanda,0)) AS pts_partidos
        FROM puntaje_detalle pd
        LEFT JOIN (
            SELECT DISTINCT ON (apostador_id) apostador_id, nombre_apostador
            FROM apuesta WHERE nombre_apostador IS NOT NULL
            ORDER BY apostador_id, id DESC
        ) an ON an.apostador_id=pd.apostador_id
        WHERE pd.torneo_id=:tid
        GROUP BY pd.apostador_id, an.nombre_apostador
        ORDER BY pts_partidos DESC
    """), {"tid": torneo_id})
    partidos_data = [dict(r._mapping) for r in rows.fetchall()]

    glob_rows = await db.execute(text("""
        SELECT apostador_id,
               COALESCE(pts_campeon,0)+COALESCE(pts_finalistas,0)+COALESCE(pts_goleador,0)+
               COALESCE(pts_peor_equipo,0)+COALESCE(pts_mayor_goleada,0)+
               COALESCE(pts_etapa_paraguay,0)+COALESCE(pts_goles_paraguay,0) AS globales
        FROM puntaje_global WHERE torneo_id=:tid
    """), {"tid": torneo_id})
    glob_map = {int(r[0]): int(r[1]) for r in glob_rows.fetchall()}

    result = []
    for row in partidos_data:
        aid   = int(row["apostador_id"])
        pts_p = int(row["pts_partidos"])
        pts_g = glob_map.get(aid, 0)
        result.append({
            "apostador_id": aid,
            "nombre":       row["nombre"] or f"id={aid}",
            "h": int(row["h"]), "i": int(row["i"]), "j": int(row["j"]),
            "k": int(row["k"]), "l": int(row["l"]), "m": int(row["m"]),
            "n": int(row["n"]), "o": int(row["o"]),
            "pts_partidos": pts_p,
            "pts_globales": pts_g,
            "total": pts_p + pts_g,
        })
    return result


@router.get("/resultados-partidos/{torneo_id}")
async def resultados_partidos(torneo_id: int, db: DBSession) -> list[dict]:
    """Partidos finalizados con todos sus items, orden fecha DESC (mas reciente primero)."""
    try:
        r = await db.execute(
            text("""
                SELECT
                    p.id,
                    f.nombre                              AS fase,
                    f.tipo                                AS fase_tipo,
                    p.fecha,
                    COALESCE(el.nombre_es, el.nombre)     AS equipo_local,
                    COALESCE(ev.nombre_es, ev.nombre)     AS equipo_visitante,
                    COALESCE(el.codigo_iso, '')           AS bandera_local,
                    COALESCE(ev.codigo_iso, '')           AS bandera_visitante,
                    p.goles_local,
                    p.goles_visitante,
                    p.penales_local,
                    p.penales_visitante,
                    COALESCE(p.amarillas, 0)              AS amarillas,
                    COALESCE(p.rojas, 0)                  AS rojas,
                    COALESCE(p.decisiones_var, 0)         AS decisiones_var,
                    COALESCE(p.penales_partido, 0)        AS penales_partido,
                    p.minuto_primer_gol
                FROM partido p
                JOIN fase f ON f.id = p.fase_id
                LEFT JOIN equipo el ON el.id = p.equipo_local_id
                LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
                WHERE f.torneo_id = :tid
                  AND p.estado = 'finalizado'
                ORDER BY p.fecha DESC NULLS LAST, f.orden DESC, p.id DESC
            """),
            {"tid": torneo_id},
        )
        out = []
        for row in r.mappings():
            d = dict(row)
            if d.get("fecha") and hasattr(d["fecha"], "strftime"):
                d["fecha"] = d["fecha"].strftime("%Y-%m-%dT%H:%M:%SZ")
            out.append(d)
        return out
    except Exception as e:
        await db.rollback()
        raise HTTPException(500, f"Error resultados-partidos: {e}")


async def _auto_lock_completed_grupos(db, torneo_id: int) -> int:
    """Bloquea fases de grupos donde todos los partidos están finalizados. Retorna N fases bloqueadas."""
    try:
        r = await db.execute(text("""
            UPDATE fase SET bloqueada = TRUE
            WHERE torneo_id = :tid
              AND tipo ILIKE 'grupo%'
              AND COALESCE(bloqueada, FALSE) = FALSE
              AND (SELECT COUNT(*) FROM partido p WHERE p.fase_id = fase.id AND p.estado != 'finalizado') = 0
              AND (SELECT COUNT(*) FROM partido p WHERE p.fase_id = fase.id) > 0
        """), {"tid": torneo_id})
        return r.rowcount if hasattr(r, 'rowcount') else 0
    except Exception:
        return 0


@router.get("/ranking/{torneo_id}", summary="Ranking de apostadores en el torneo con desglose por categoría")
async def ranking(torneo_id: int, db: DBSession) -> list[dict]:
    _ITEMS = [
        "pts_resultado", "pts_marcador", "pts_amarillas", "pts_rojas",
        "pts_var", "pts_minuto", "pts_penales_partido", "pts_penales_tanda", "pts_equipo"
    ]
    _SUM_EXPR = " + ".join(f"COALESCE(pd.{c}, 0)" for c in _ITEMS)

    # ── Puntajes totales y por item desde puntaje_detalle (fuente única) ──────
    try:
        rd = await db.execute(
            text(f"""
                SELECT
                    pd.apostador_id,
                    COALESCE(SUM({_SUM_EXPR}), 0)::int AS puntos_partidos_total,
                    COUNT(DISTINCT pd.partido_id)::int  AS apuestas_total,
                    SUM(CASE WHEN COALESCE(pd.pts_marcador,0) > 0 THEN 1 ELSE 0 END)::int AS plenos,
                    SUM(CASE WHEN COALESCE(pd.pts_resultado,0) > 0
                              AND COALESCE(pd.pts_marcador,0) = 0
                             THEN 1 ELSE 0 END)::int AS aciertos,
                    SUM(CASE WHEN COALESCE(pd.pts_resultado,0) = 0
                              AND COALESCE(pd.pts_marcador,0) = 0
                             THEN 1 ELSE 0 END)::int AS fallos,
                    COALESCE(SUM(pd.pts_resultado),                   0)::int AS cat_resultado,
                    COALESCE(SUM(pd.pts_marcador),                    0)::int AS cat_marcador,
                    COALESCE(SUM(pd.pts_amarillas),                   0)::int AS cat_amarillas,
                    COALESCE(SUM(COALESCE(pd.pts_rojas,0)),           0)::int AS cat_rojas,
                    COALESCE(SUM(pd.pts_var),                         0)::int AS cat_var,
                    COALESCE(SUM(pd.pts_minuto),                      0)::int AS cat_minuto,
                    COALESCE(SUM(COALESCE(pd.pts_penales_partido,0)), 0)::int AS cat_penales_partido,
                    COALESCE(SUM(COALESCE(pd.pts_penales_tanda,0)),   0)::int AS cat_penales_tanda,
                    COALESCE(SUM(COALESCE(pd.pts_equipo,0)),          0)::int AS cat_equipo
                FROM puntaje_detalle pd
                WHERE pd.torneo_id = :tid
                GROUP BY pd.apostador_id
            """),
            {"tid": torneo_id},
        )
        rows = [dict(row) for row in rd.mappings()]
    except Exception:
        await db.rollback()
        rows = []

    # ── Puntajes globales A-G ─────────────────────────────────────────────────
    try:
        rg = await db.execute(
            text("SELECT apostador_id, pts_total FROM puntaje_global WHERE torneo_id = :tid"),
            {"tid": torneo_id},
        )
        global_pts = {row["apostador_id"]: row["pts_total"] for row in rg.mappings()}
    except Exception:
        global_pts = {}

    # ── Puntos P grupos (apostador_clasificados) ──────────────────────────────
    try:
        rc = await db.execute(
            text("""
                SELECT apostador_id, COALESCE(pts_obtenidos, 0) AS pts_grupos_p
                FROM apostador_clasificados
                WHERE torneo_id = :tid AND fase_tipo = 'grupo'
            """),
            {"tid": torneo_id},
        )
        clas_pts = {row["apostador_id"]: int(row["pts_grupos_p"] or 0) for row in rc.mappings()}
    except Exception:
        clas_pts = {}

    # ── Puntos D (peor equipo) por separado ───────────────────────────────────
    try:
        rd2 = await db.execute(
            text("SELECT apostador_id, COALESCE(pts_peor_equipo, 0) AS pts_d FROM puntaje_global WHERE torneo_id = :tid"),
            {"tid": torneo_id},
        )
        peor_equipo_pts = {row["apostador_id"]: int(row["pts_d"] or 0) for row in rd2.mappings()}
    except Exception:
        peor_equipo_pts = {}

    # ── Desglose por fase ─────────────────────────────────────────────────────
    try:
        rf = await db.execute(
            text(f"""
                SELECT
                    pd.apostador_id,
                    f.id AS fase_id,
                    f.tipo AS fase_tipo,
                    f.nombre AS fase_nombre,
                    COALESCE(SUM({_SUM_EXPR}), 0)::int AS pts_fase
                FROM puntaje_detalle pd
                JOIN partido p ON p.id = pd.partido_id
                JOIN fase f ON f.id = p.fase_id
                WHERE pd.torneo_id = :tid
                GROUP BY pd.apostador_id, f.id, f.tipo, f.nombre
                ORDER BY pd.apostador_id, f.id
            """),
            {"tid": torneo_id},
        )
        fases_raw = [dict(row) for row in rf.mappings()]
    except Exception:
        fases_raw = []

    fases_by_uid: dict[int, list] = {}
    for fr in fases_raw:
        uid = fr["apostador_id"]
        fases_by_uid.setdefault(uid, []).append({
            "tipo": fr["fase_tipo"],
            "nombre": fr["fase_nombre"],
            "pts": fr["pts_fase"],
        })

    # ── Todos los apostadores activos ─────────────────────────────────────────
    async with _app_engine.connect() as conn:
        ar = await conn.execute(
            text("""
                SELECT u.id, u.username
                FROM users u
                JOIN user_roles ur ON ur.user_id = u.id
                JOIN roles ro ON ro.id = ur.role_id
                WHERE ro.name = 'apostador' AND u.is_active = TRUE
            """)
        )
        apostadores_all = {row["id"]: row["username"] for row in ar.mappings()}
        ids = [row["apostador_id"] for row in rows]
        user_map = dict(apostadores_all)
        if ids:
            ur = await conn.execute(
                text("SELECT id, username FROM users WHERE id = ANY(:ids)"),
                {"ids": ids}
            )
            for row in ur.mappings():
                user_map[row["id"]] = row["username"]

    _ZERO_CATS = {
        "cat_resultado": 0, "cat_marcador": 0, "cat_amarillas": 0,
        "cat_rojas": 0, "cat_var": 0, "cat_minuto": 0,
        "cat_penales_partido": 0, "cat_penales_tanda": 0, "cat_equipo": 0,
    }

    present = {row["apostador_id"] for row in rows}
    for uid in apostadores_all:
        if uid not in present:
            rows.append({
                "apostador_id": uid,
                "puntos_partidos_total": 0, "apuestas_total": 0,
                "plenos": 0, "aciertos": 0, "fallos": 0,
            })

    for row in rows:
        uid = row["apostador_id"]
        row["nombre"]                = user_map.get(uid, f"Usuario {uid}")
        pts_globales                 = global_pts.get(uid, 0) or 0
        pts_partidos                 = row.get("puntos_partidos_total") or 0
        pts_grupos_p                 = clas_pts.get(uid, 0) or 0
        pts_peor_equipo_d            = peor_equipo_pts.get(uid, 0) or 0
        row["pts_globales"]          = pts_globales
        row["pts_grupos_p"]          = pts_grupos_p
        row["pts_peor_equipo_d"]     = pts_peor_equipo_d
        row["puntos_total"]          = pts_partidos + pts_globales + pts_grupos_p
        row["puntos_partidos"]       = pts_partidos
        row["puntos_partidos_total"] = pts_partidos
        row.update({k: row.get(k, 0) for k in _ZERO_CATS})
        row["fases"]                 = fases_by_uid.get(uid, [])
        row.setdefault("plenos",  0)
        row.setdefault("aciertos", 0)
        row.setdefault("fallos",  0)
        row["bonus_count"]           = 0
        row["puntos_bonus_partido"]  = 0
        row["online_source"]         = _is_online(uid)

    rows.sort(key=lambda r: (-(r.get("puntos_total") or 0),
                             -(r.get("apuestas_total") or 0),
                             (r.get("nombre") or "").lower()))
    return rows


# ── Simulación y puntajes ────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# Helper: resetear puntajes de TODOS los apostadores de un torneo
# Borra puntaje_detalle y puntaje_global; pone puntos/puntos_bonus = 0 en apuesta.
# ─────────────────────────────────────────────────────────────────────────────
async def _reset_puntajes_todos(db: AsyncSession, torneo_id: int) -> dict:
    """
    Vuelve a cero TODOS los puntajes calculados del torneo:
      - apuesta.puntos = 0, puntos_bonus = 0  (para todos los partidos del torneo)
      - DELETE puntaje_detalle WHERE torneo_id
      - DELETE puntaje_global  WHERE torneo_id
    Retorna {"apuestas_zeroed": N, "detalle_borradas": N, "global_borradas": N}
    """
    # 1. Cero en apuesta.puntos / puntos_bonus
    r_a = await db.execute(
        text("""
            UPDATE apuesta
            SET puntos = 0, puntos_bonus = 0
            WHERE partido_id IN (
                SELECT id FROM partido WHERE torneo_id = :tid
            )
        """),
        {"tid": torneo_id},
    )
    apuestas_zeroed = r_a.rowcount or 0

    # 2. Borrar puntaje_detalle (incluye bonus calculados)
    detalle_borradas = 0
    try:
        r_d = await db.execute(
            text("DELETE FROM puntaje_detalle WHERE torneo_id = :tid"),
            {"tid": torneo_id},
        )
        detalle_borradas = r_d.rowcount or 0
    except Exception as _e:
        log.warning("reset_puntajes.detalle_skip", error=str(_e))
        # NO rollback — no cancelar el UPDATE apuesta que ya corrió

    # 3. Borrar puntaje_global (A-G)
    global_borradas = 0
    try:
        r_g = await db.execute(
            text("DELETE FROM puntaje_global WHERE torneo_id = :tid"),
            {"tid": torneo_id},
        )
        global_borradas = r_g.rowcount or 0
    except Exception as _e:
        log.warning("reset_puntajes.global_skip", error=str(_e))
        # NO rollback

    # 4. Borrar puntaje_item completo del torneo
    items_borradas = 0
    try:
        r_i = await db.execute(
            text("DELETE FROM puntaje_item WHERE torneo_id = :tid"),
            {"tid": torneo_id},
        )
        items_borradas = r_i.rowcount or 0
    except Exception as _e:
        log.warning("reset_puntajes.items_skip", error=str(_e))

    return {
        "apuestas_zeroed":  apuestas_zeroed,
        "detalle_borradas": detalle_borradas,
        "global_borradas":  global_borradas,
        "items_borradas":   items_borradas,
    }


@router.post("/reset-puntajes/{torneo_id}",
             summary="Resetear TODOS los puntajes calculados del torneo (admin)")
async def reset_puntajes(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """
    Vuelve a cero los puntajes de TODOS los apostadores del torneo.
    No borra los pronósticos (apuesta.pred_*) — solo los puntos calculados.
    Útil para recalcular desde cero cuando se modifican reglas de scoring.
    """
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    summary = await _reset_puntajes_todos(db, torneo_id)
    await db.commit()
    await _audit_log("reset:puntajes", "bets", current=current, method="POST",
                     path=f"/api/v1/bets/reset-puntajes/{torneo_id}",
                     resource_id=str(torneo_id),
                     details={**summary, "evento": "reset puntajes todos los apostadores"})
    return {"ok": True, "torneo_id": torneo_id, **summary}


@router.post("/reset-grupos-admin/{torneo_id}",
             summary="Reset completo de grupos para TODOS los apostadores (admin)")
async def reset_grupos_admin(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """
    Resetea pronósticos Y puntajes de TODOS los apostadores del torneo:
      - apuesta.pred_* = NULL, puntos = 0, puntos_bonus = 0
      - puntaje_detalle: DELETE
      - puntaje_global: DELETE
      - apuesta_global: DELETE (pronósticos A-G)
    Equivalente al resetear-apuestas pero aplicado a todos los apostadores.
    Usar desde el panel admin cuando se quiere reiniciar la fase de grupos.
    """
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    partido_ids_subq = "SELECT id FROM partido WHERE torneo_id = :tid"

    # 1. NULL todos los pronósticos de todos los apostadores del torneo
    r_pred = await db.execute(
        text(f"""
            UPDATE apuesta SET
                pred_local                   = NULL,
                pred_visitante               = NULL,
                pred_penales                 = NULL,
                pred_minuto_gol              = NULL,
                pred_amarillas               = NULL,
                pred_var                     = NULL,
                pred_rojas                   = NULL,
                pred_penales_partido         = NULL,
                pred_penales_local_tanda     = NULL,
                pred_penales_visitante_tanda = NULL,
                puntos                       = 0,
                puntos_bonus                 = 0,
                updated_at                   = NOW()
            WHERE partido_id IN ({partido_ids_subq})
        """),
        {"tid": torneo_id},
    )
    apuestas = r_pred.rowcount or 0

    # 2. Puntaje detalle y global
    pts = await _reset_puntajes_todos(db, torneo_id)

    # 3. Pronósticos globales A-G
    r_ag = await db.execute(
        text("DELETE FROM apuesta_global WHERE torneo_id = :tid"),
        {"tid": torneo_id},
    )
    globales = r_ag.rowcount or 0

    await db.commit()
    await _audit_log("reset:grupos_admin", "bets", current=current, method="POST",
                     path=f"/api/v1/bets/reset-grupos-admin/{torneo_id}",
                     resource_id=str(torneo_id),
                     details={"apuestas": apuestas, "globales": globales, **pts})
    return {
        "ok": True,
        "torneo_id": torneo_id,
        "apuestas_reseteadas": apuestas,
        "globales_borrados": globales,
        **pts,
    }


@router.post("/simular-resultados/{torneo_id}", summary="Simular resultados aleatorios de grupos (admin, solo pruebas)")
async def simular_resultados(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    r = await db.execute(
        text("""
            SELECT p.id FROM partido p
            JOIN fase f ON f.id = p.fase_id AND f.tipo = 'grupo'
            WHERE p.torneo_id = :tid
        """),
        {"tid": torneo_id},
    )
    partido_ids = [row[0] for row in r]
    if not partido_ids:
        raise HTTPException(404, "No hay partidos de grupo para este torneo")

    # Resetear puntajes calculados antes de simular (resultado cambia → scores quedan stale)
    await _reset_puntajes_todos(db, torneo_id)

    # Bloqueo: si el torneo ya avanzó a fases KO con resultados, la fase de grupos
    # está encerrada y no se puede re-simular (corrompería el bracket).
    rko = await db.execute(
        text("""SELECT COUNT(*)::int FROM partido p JOIN fase f ON f.id=p.fase_id
                WHERE p.torneo_id=:tid AND f.tipo<>'grupo' AND p.estado='finalizado'"""),
        {"tid": torneo_id})
    ko_finalizados = rko.scalar() or 0
    if ko_finalizados > 0:
        await _audit_log("simulacion:bloqueada", "bets", current=current,
                         status_code=409, method="POST",
                         path=f"/api/v1/bets/simular-resultados/{torneo_id}",
                         resource_id=str(torneo_id),
                         details={"motivo": "fase de grupos encerrada",
                                  "ko_finalizados": ko_finalizados})
        raise HTTPException(409, "La fase de grupos está encerrada: ya hay resultados en fases posteriores. Usá 'Reset total' (superadmin) para reiniciar.")

    for pid in partido_ids:
        gl = random.choices([0, 1, 2, 3], weights=[18, 38, 30, 14])[0]
        gv = random.choices([0, 1, 2, 3], weights=[18, 38, 30, 14])[0]
        # Bonus por partido (solo si la simulación los debe poblar para pruebas)
        minuto = random.randint(1, 90) if (gl + gv) > 0 else None
        amarillas = random.choices([0, 1, 2, 3, 4, 5, 6], weights=[5, 12, 22, 26, 18, 11, 6])[0]
        var = random.choices([0, 1], weights=[65, 35])[0]
        await db.execute(
            text("""
                UPDATE partido
                SET goles_local=:gl, goles_visitante=:gv,
                    minuto_primer_gol=:minuto, amarillas=:amar, rojas=:rojas, decisiones_var=:var,
                    penales_partido=:pen_part, estado='finalizado'
                WHERE id=:pid
            """),
            {"gl": gl, "gv": gv, "minuto": minuto, "amar": amarillas, "rojas": random.choices([0,1,2], weights=[78,18,4])[0], "var": var, "pen_part": random.choices([0,1,2], weights=[70,22,8])[0], "pid": pid},
        )

    await _recalc_participacion(db, torneo_id)
    # Avanzar bracket: R32 toma sus equipos desde los standings de grupos
    maps = await ko_scoring.build_num_maps(db, torneo_id)
    await _avanzar_bracket(db, torneo_id, maps)
    await db.commit()
    await _audit_log("simulacion:grupos", "bets", current=current, method="POST",
                     path=f"/api/v1/bets/simular-resultados/{torneo_id}",
                     resource_id=str(torneo_id),
                     details={"evento": "simular resultados de grupos",
                              "partidos_actualizados": len(partido_ids),
                              "avance_bracket": "R32 desde standings"})
    return {"ok": True, "partidos_actualizados": len(partido_ids)}


@router.post("/simular-fase/{fase_id}", summary="Simular resultados aleatorios de una fase (admin, solo pruebas)")
async def simular_fase(fase_id: int, current: CurrentUser, db: DBSession) -> dict:
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    rf = await db.execute(
        text("SELECT id, nombre, tipo, torneo_id, COALESCE(bloqueada, FALSE) AS bloqueada FROM fase WHERE id = :fid"),
        {"fid": fase_id},
    )
    fase = rf.mappings().first()
    if not fase:
        raise HTTPException(404, "Fase no encontrada")
    es_grupo = fase["tipo"] == "grupo"

    # Bloqueo: respetar el bloqueo manual del admin
    if fase.get("bloqueada"):
        await _audit_log("simulacion:bloqueada", "bets", current=current,
                         status_code=409, method="POST",
                         path=f"/api/v1/bets/simular-fase/{fase_id}",
                         resource_id=str(fase_id),
                         details={"motivo": "fase bloqueada manualmente", "fase": fase["nombre"]})
        raise HTTPException(409, f"La fase '{fase['nombre']}' está bloqueada. Desbloquéala en Configuración → Fases para simular.")

    r = await db.execute(
        text("SELECT id FROM partido WHERE fase_id = :fid"),
        {"fid": fase_id},
    )
    partido_ids = [row[0] for row in r]
    if not partido_ids:
        raise HTTPException(404, "La fase no tiene partidos")

    for pid in partido_ids:
        gl = random.choices([0, 1, 2, 3], weights=[18, 38, 30, 14])[0]
        gv = random.choices([0, 1, 2, 3], weights=[18, 38, 30, 14])[0]
        minuto = random.randint(1, 90) if (gl + gv) > 0 else None
        amarillas = random.choices([0, 1, 2, 3, 4, 5, 6], weights=[5, 12, 22, 26, 18, 11, 6])[0]
        var = random.choices([0, 1], weights=[65, 35])[0]
        # En eliminatorias no puede haber empate: definir por penales
        pen_l = pen_v = None
        if not es_grupo and gl == gv:
            pen_l, pen_v = random.choice([(4, 2), (5, 4), (3, 1), (5, 3), (4, 5), (2, 4), (1, 3)])
        await db.execute(
            text("""
                UPDATE partido
                SET goles_local=:gl, goles_visitante=:gv,
                    minuto_primer_gol=:minuto, amarillas=:amar, decisiones_var=:var,
                    rojas=:rojas, penales_partido=:pen_part, penales_local=:pl, penales_visitante=:pv,
                    estado='finalizado'
                WHERE id=:pid
            """),
            {"gl": gl, "gv": gv, "minuto": minuto, "amar": amarillas, "rojas": random.choices([0,1,2], weights=[78,18,4])[0], "var": var,
             "pen_part": random.choices([0,1,2], weights=[70,22,8])[0], "pl": pen_l, "pv": pen_v, "pid": pid},
        )

    if es_grupo:
        await _recalc_participacion(db, fase["torneo_id"])
    # Avanzar bracket: propaga ganadores a la(s) fase(s) KO siguiente(s)
    maps = await ko_scoring.build_num_maps(db, fase["torneo_id"])
    await _avanzar_bracket(db, fase["torneo_id"], maps)
    await db.commit()
    await _audit_log("simulacion:fase", "bets", current=current, method="POST",
                     path=f"/api/v1/bets/simular-fase/{fase_id}",
                     resource_id=str(fase_id),
                     details={"evento": "simular fase", "fase": fase["nombre"],
                              "tipo": fase["tipo"],
                              "partidos_actualizados": len(partido_ids),
                              "avance_bracket": True})
    return {"ok": True, "fase": fase["nombre"], "partidos_actualizados": len(partido_ids)}


@router.post("/reset-fase/{fase_id}", summary="Reiniciar resultados de una fase a 'programado' (admin, solo pruebas)")
async def reset_fase(fase_id: int, current: CurrentUser, db: DBSession) -> dict:
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    rf = await db.execute(
        text("SELECT id, nombre, tipo, torneo_id, COALESCE(bloqueada, FALSE) AS bloqueada FROM fase WHERE id = :fid"),
        {"fid": fase_id},
    )
    fase = rf.mappings().first()
    if not fase:
        raise HTTPException(404, "Fase no encontrada")

    # Bloqueo: respetar el bloqueo manual del admin
    if fase.get("bloqueada"):
        await _audit_log("reset-fase:bloqueada", "bets", current=current,
                         status_code=409, method="POST",
                         path=f"/api/v1/bets/reset-fase/{fase_id}",
                         resource_id=str(fase_id),
                         details={"motivo": "fase bloqueada manualmente", "fase": fase["nombre"]})
        raise HTTPException(409, f"La fase '{fase['nombre']}' está bloqueada. Desbloquéala en Configuración → Fases para reiniciar.")

    r = await db.execute(
        text("SELECT id FROM partido WHERE fase_id = :fid"),
        {"fid": fase_id},
    )
    partido_ids = [row[0] for row in r]
    if not partido_ids:
        raise HTTPException(404, "La fase no tiene partidos")

    # Reiniciar partidos: estado programado + limpiar resultados y bonus
    await db.execute(
        text("""
            UPDATE partido
            SET goles_local=NULL, goles_visitante=NULL,
                minuto_primer_gol=NULL, amarillas=NULL, rojas=NULL, decisiones_var=NULL,
                penales_partido=NULL, penales_local=NULL, penales_visitante=NULL,
                goles_local_prorroga=NULL, goles_visitante_prorroga=NULL,
                estado='programado'
            WHERE fase_id=:fid
        """),
        {"fid": fase_id},
    )
    # Limpiar puntajes de apuestas de esos partidos (las predicciones se conservan)
    await db.execute(
        text("UPDATE apuesta SET puntos=0, puntos_bonus=0 WHERE partido_id = ANY(:pids)"),
        {"pids": partido_ids},
    )

    if fase["tipo"] == "grupo":
        await _recalc_participacion(db, fase["torneo_id"])
    await db.commit()
    await _audit_log("reset:fase", "bets", current=current, method="POST",
                     path=f"/api/v1/bets/reset-fase/{fase_id}",
                     resource_id=str(fase_id),
                     details={"evento": "reiniciar fase", "fase": fase["nombre"],
                              "partidos_reiniciados": len(partido_ids)})
    return {"ok": True, "fase": fase["nombre"], "partidos_reiniciados": len(partido_ids)}


@router.post("/reset-torneo/{torneo_id}", summary="Reset total: borra resultados y pronósticos del torneo (superadmin)")
async def reset_torneo(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    if not await _check_superadmin(current):
        raise HTTPException(403, "Solo el superadmin puede reiniciar el torneo")

    part_subq = "SELECT id FROM partido WHERE torneo_id=:tid"

    # 1) Borrar todas las apuestas (predicciones) del torneo
    #    IMPORTANTE: usar subquery — ANY(:list) no funciona en asyncpg con listas Python
    rd = await db.execute(
        text(f"DELETE FROM apuesta WHERE partido_id IN ({part_subq})"),
        {"tid": torneo_id},
    )
    apuestas_borradas = rd.rowcount or 0

    # 2) Resetear scores de todos los partidos a NULL/programado
    await db.execute(
        text("""
            UPDATE partido
            SET goles_local          = NULL,
                goles_visitante      = NULL,
                minuto_primer_gol    = NULL,
                amarillas            = NULL,
                decisiones_var       = NULL,
                penales_local        = NULL,
                penales_visitante    = NULL,
                estado               = 'programado'
            WHERE torneo_id = :tid
        """),
        {"tid": torneo_id},
    )
    # Columnas opcionales: verificar cuáles existen ANTES de actualizar
    # (evita que una columna faltante ponga la transacción en estado ABORTED)
    _opt_cols_check = await db.execute(text("""
        SELECT column_name FROM information_schema.columns
        WHERE table_name='partido' AND table_schema='public'
          AND column_name = ANY(ARRAY['rojas','penales_partido','equipo_clasificado_id',
                                       'goles_local_prorroga','goles_visitante_prorroga',
                                       'minuto_actual'])
    """))
    _opt_cols_exist = {row[0] for row in _opt_cols_check}
    for _col in _opt_cols_exist:
        await db.execute(
            text(f"UPDATE partido SET {_col}=NULL WHERE torneo_id=:tid"),
            {"tid": torneo_id},
        )

    # 3) Resetear standings de grupos (participacion) a cero
    await db.execute(
        text(f"""
            UPDATE participacion
            SET pj=0, pg=0, pe=0, pp=0, gf=0, gc=0, pts=0, clasifica=FALSE
            WHERE fase_id IN (SELECT id FROM fase WHERE torneo_id=:tid AND tipo='grupo')
        """),
        {"tid": torneo_id},
    )

    # 4) Borrar puntaje_detalle (no tiene torneo_id, acceder via partido_id)
    await db.execute(
        text(f"DELETE FROM puntaje_detalle WHERE partido_id IN ({part_subq})"),
        {"tid": torneo_id},
    )

    # 5) Borrar pronósticos globales A-G y sus puntajes
    await db.execute(text("DELETE FROM apuesta_global WHERE torneo_id=:tid"), {"tid": torneo_id})
    await db.execute(text("DELETE FROM puntaje_global  WHERE torneo_id=:tid"), {"tid": torneo_id})

    # 5b) Borrar puntaje_item (partido vía partido_id IN subq + globales vía torneo_id)
    try:
        await db.execute(
            text(f"DELETE FROM puntaje_item WHERE partido_id IN ({part_subq})"),
            {"tid": torneo_id},
        )
        await db.execute(
            text("DELETE FROM puntaje_item WHERE torneo_id=:tid AND partido_id IS NULL"),
            {"tid": torneo_id},
        )
    except Exception:
        pass  # tabla puede no existir aún

    # 6) Resetear bracket KO a TBD
    try:
        await _resetear_ko_a_tbd(db, torneo_id)
    except Exception as _e:
        log.warning("reset_torneo.ko_tbd_skip", error=str(_e))

    # 7) Commit antes de recrear placeholders (libera locks)
    await db.commit()

    # 8) Recrear filas apuesta vacías (puntos=0) para cada apostador × partido de grupos
    #    Permite que los apostadores aparezcan en ranking aunque aún no hayan apostado
    placeholders = 0
    rg = await db.execute(
        text("""
            SELECT p.id FROM partido p
            JOIN fase f ON f.id = p.fase_id AND f.tipo = 'grupo'
            WHERE p.torneo_id = :tid
        """),
        {"tid": torneo_id},
    )
    grupo_pids = [row[0] for row in rg]
    if grupo_pids:
        async with _app_engine.connect() as conn:
            ar = await conn.execute(text("""
                SELECT u.id FROM users u
                JOIN user_roles ur ON ur.user_id = u.id
                JOIN roles ro ON ro.id = ur.role_id
                WHERE ro.name = 'apostador' AND u.is_active = TRUE
            """))
            apostador_ids = [row[0] for row in ar]
        for uid in apostador_ids:
            res = await db.execute(
                text("""
                    INSERT INTO apuesta (apostador_id, partido_id, puntos, puntos_bonus)
                    SELECT :uid, pid, 0, 0
                    FROM unnest(CAST(:pids AS bigint[])) AS pid
                    ON CONFLICT (apostador_id, partido_id)
                    DO UPDATE SET puntos=0, puntos_bonus=0
                """),
                {"uid": uid, "pids": grupo_pids},
            )
            placeholders += res.rowcount or 0
        await db.commit()

    # 9) Recalcular standings reales desde cero (confirmar todo en 0)
    try:
        await _recalc_participacion(db, torneo_id)
        await db.commit()
    except Exception:
        await db.rollback()

    await _audit_log("reset:torneo", "bets", current=current, method="POST",
                     path=f"/api/v1/bets/reset-torneo/{torneo_id}",
                     resource_id=str(torneo_id),
                     details={"evento": "RESET TOTAL del torneo (superadmin)",
                              "apuestas_borradas": apuestas_borradas,
                              "placeholders_recreados": placeholders})
    return {
        "ok": True,
        "torneo_id":         torneo_id,
        "apuestas_borradas": apuestas_borradas,
        "placeholders":      placeholders,
        "msg":               "Reset completo: resultados, puntajes, bracket y pronósticos borrados.",
    }


@router.api_route("/verificar-registros/{torneo_id}", methods=["GET", "POST"],
                  summary="Verifica (y opcionalmente repara) que todos los apostadores tengan registros de apuesta completos")
async def verificar_registros(torneo_id: int, current: CurrentUser, db: DBSession,
                              reparar: bool = False) -> dict:
    """Algoritmo de consistencia de registros de apuesta.

    Para garantizar que CADA apostador activo esté habilitado para apostar
    (incluso después de un reset, donde se borran las apuestas), verifica que
    exista una fila en `apuesta` por cada partido de la fase de grupos para
    cada apostador. Con ?reparar=true (o POST), crea las filas faltantes como
    placeholders (predicción NULL = registro habilitado, pendiente de completar).
    """
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    # 1) Partidos de la fase de grupos del torneo (set universal a apostar)
    rp = await db.execute(
        text("""
            SELECT p.id FROM partido p
            JOIN fase f ON f.id = p.fase_id AND f.tipo = 'grupo'
            WHERE p.torneo_id = :tid
            ORDER BY p.id
        """),
        {"tid": torneo_id},
    )
    partido_ids = [row[0] for row in rp]
    total_partidos = len(partido_ids)
    if total_partidos == 0:
        raise HTTPException(404, "No hay partidos de grupo en este torneo")

    # 2) Apostadores activos (rol 'apostador') desde app_db
    async with _app_engine.connect() as conn:
        ar = await conn.execute(text("""
            SELECT u.id, u.username FROM users u
            JOIN user_roles ur ON ur.user_id = u.id
            JOIN roles ro ON ro.id = ur.role_id
            WHERE ro.name = 'apostador' AND u.is_active = TRUE
            ORDER BY u.username
        """))
        apostadores = {row["id"]: row["username"] for row in ar.mappings()}

    # 3) Registros existentes por apostador para esos partidos
    re_ = await db.execute(
        text("""
            SELECT apostador_id, COUNT(*)::int AS n
            FROM apuesta
            WHERE partido_id = ANY(:pids)
            GROUP BY apostador_id
        """),
        {"pids": partido_ids},
    )
    existentes = {row["apostador_id"]: row["n"] for row in re_.mappings()}

    # 4) Diagnóstico + (opcional) reparación
    detalle = []
    creados_total = 0
    for uid, uname in apostadores.items():
        tiene = existentes.get(uid, 0)
        faltan = total_partidos - tiene
        creados = 0
        if reparar and faltan > 0:
            res = await db.execute(
                text("""
                    INSERT INTO apuesta (apostador_id, partido_id)
                    SELECT :uid, pid FROM unnest(CAST(:pids AS bigint[])) AS pid
                    ON CONFLICT (apostador_id, partido_id) DO NOTHING
                """),
                {"uid": uid, "pids": partido_ids},
            )
            creados = res.rowcount or 0
            creados_total += creados
        detalle.append({
            "apostador_id": uid, "nombre": uname,
            "registros": tiene, "esperados": total_partidos,
            "faltantes": faltan, "creados": creados,
            "completo": (tiene + creados) >= total_partidos,
        })

    if reparar:
        await db.commit()
        await _audit_log("registros:reparacion", "bets", current=current, method="POST",
                         path=f"/api/v1/bets/verificar-registros/{torneo_id}",
                         resource_id=str(torneo_id),
                         details={"evento": "habilitación de registros de apuesta",
                                  "registros_creados": creados_total,
                                  "apostadores": len(apostadores)})

    incompletos = [d for d in detalle if not d["completo"]]
    return {
        "ok": True,
        "torneo_id": torneo_id,
        "reparado": reparar,
        "total_apostadores": len(apostadores),
        "partidos_grupo": total_partidos,
        "registros_creados": creados_total,
        "apostadores_incompletos": len(incompletos),
        "detalle": sorted(detalle, key=lambda x: (x["completo"], x["nombre"].lower())),
    }


@router.get("/monitor-dashboard/{torneo_id}", summary="Dashboard completo para panel de mando admin")
async def monitor_dashboard(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """
    Un solo call que devuelve todo lo necesario para el panel de mando del admin:
      - stats: total apostadores, líder, último
      - top5: top 5 del ranking con puntos
      - sin_apuestas: apostadores con apuestas incompletas en fases abiertas
      - partido_activo: partidos en ventana activa (±5 min / 150 min)
      - fases: estado de todas las fases (total, finalizados, bloqueada)
      - mensajes_recientes: últimos 3 mensajes enviados
      - torneo: nombre, api_season, api_league_id, período
    """
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    result: dict = {}

    # ── Torneo info ───────────────────────────────────────────────────────────
    try:
        r_t = await db.execute(
            text("""
                SELECT t.id, t.nombre, t.api_season, t.apuesta_inicio, t.apuesta_fin,
                       c.nombre AS comp_nombre, c.api_league_id
                FROM torneo t
                LEFT JOIN competicion c ON c.id = t.competicion_id
                WHERE t.id = :tid
            """),
            {"tid": torneo_id},
        )
        tor = dict(r_t.mappings().first() or {})
        for f in ("apuesta_inicio", "apuesta_fin"):
            if tor.get(f) and hasattr(tor[f], "isoformat"):
                tor[f] = tor[f].isoformat()
        result["torneo"] = tor
    except Exception:
        await db.rollback()
        result["torneo"] = {}

    # ── Fases estado ──────────────────────────────────────────────────────────
    try:
        r_f = await db.execute(
            text("""
                SELECT f.id, f.nombre, f.tipo, f.orden,
                       COALESCE(f.bloqueada, FALSE) AS bloqueada,
                       COUNT(p.id) AS total,
                       SUM(CASE WHEN p.estado='finalizado' THEN 1 ELSE 0 END) AS finalizados
                FROM fase f
                LEFT JOIN partido p ON p.fase_id = f.id AND p.torneo_id = :tid
                WHERE f.torneo_id = :tid
                GROUP BY f.id ORDER BY f.orden, f.nombre
            """),
            {"tid": torneo_id},
        )
        result["fases"] = [dict(row) for row in r_f.mappings()]
    except Exception:
        await db.rollback()
        result["fases"] = []

    # ── Partido activo ────────────────────────────────────────────────────────
    try:
        r_pa = await db.execute(
            text("""
                SELECT p.id, p.fecha, p.estado, f.tipo AS fase_tipo, f.nombre AS fase_nombre,
                       COALESCE(el.nombre_es, el.nombre) AS local_nombre,
                       COALESCE(ev.nombre_es, ev.nombre) AS visitante_nombre,
                       p.goles_local, p.goles_visitante
                FROM partido p
                JOIN fase f ON f.id = p.fase_id
                LEFT JOIN equipo el ON el.id = p.equipo_local_id
                LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
                WHERE p.torneo_id = :tid
                  AND p.estado != 'finalizado'
                  AND p.fecha IS NOT NULL
                  AND p.fecha <= (NOW() AT TIME ZONE 'UTC') + INTERVAL '30 minutes'
                  AND p.fecha >= (NOW() AT TIME ZONE 'UTC') - INTERVAL '150 minutes'
                ORDER BY p.fecha
            """),
            {"tid": torneo_id},
        )
        partidos_activos = []
        for row in r_pa.mappings():
            d = dict(row)
            if d.get("fecha") and hasattr(d["fecha"], "strftime"):
                d["fecha"] = d["fecha"].strftime("%Y-%m-%dT%H:%M:%SZ")
            partidos_activos.append(d)
        result["partidos_activos"] = partidos_activos
    except Exception:
        await db.rollback()
        result["partidos_activos"] = []

    # ── Apostadores sin apuestas completas ────────────────────────────────────
    try:
        # Partidos en fases abiertas (no bloqueadas, no finalizados)
        r_pend = await db.execute(
            text("""
                SELECT f.tipo AS fase_tipo, f.nombre AS fase_nombre,
                       COUNT(p.id) AS partidos_pendientes
                FROM partido p
                JOIN fase f ON f.id = p.fase_id
                WHERE p.torneo_id = :tid
                  AND p.estado != 'finalizado'
                  AND COALESCE(f.bloqueada, FALSE) = FALSE
                GROUP BY f.id ORDER BY f.orden, f.nombre
            """),
            {"tid": torneo_id},
        )
        fases_abiertas_info = [dict(r) for r in r_pend.mappings()]
        total_pendientes = sum(f["partidos_pendientes"] for f in fases_abiertas_info)

        # Apuestas realizadas en esos partidos por apostador
        r_apost = await db.execute(
            text("""
                SELECT a.apostador_id, COUNT(*) AS apuestas_hechas
                FROM apuesta a
                JOIN partido p ON p.id = a.partido_id
                JOIN fase f ON f.id = p.fase_id
                WHERE p.torneo_id = :tid
                  AND p.estado != 'finalizado'
                  AND COALESCE(f.bloqueada, FALSE) = FALSE
                GROUP BY a.apostador_id
            """),
            {"tid": torneo_id},
        )
        apuestas_map = {row["apostador_id"]: row["apuestas_hechas"] for row in r_apost.mappings()}
    except Exception:
        await db.rollback()
        fases_abiertas_info = []
        total_pendientes    = 0
        apuestas_map        = {}

    # Cruzar con usuarios (app_db)
    sin_apuestas = []
    if total_pendientes > 0:
        async with _app_engine.connect() as conn:
            ar = await conn.execute(
                text("""
                    SELECT u.id, u.username
                    FROM users u
                    JOIN user_roles ur ON ur.user_id = u.id
                    JOIN roles ro ON ro.id = ur.role_id
                    WHERE ro.name = 'apostador' AND u.is_active = TRUE
                    ORDER BY u.username
                """)
            )
            todos = [dict(r) for r in ar.mappings()]
        for u in todos:
            hechas = apuestas_map.get(u["id"], 0)
            if hechas < total_pendientes:
                sin_apuestas.append({
                    "nombre":    u["username"],
                    "hechas":    hechas,
                    "pendientes": total_pendientes,
                    "pct":       round(100 * hechas / total_pendientes) if total_pendientes else 0,
                })
        sin_apuestas.sort(key=lambda x: x["hechas"])
    result["sin_apuestas"]        = sin_apuestas
    result["fases_abiertas_info"] = fases_abiertas_info
    result["total_pendientes"]    = total_pendientes

    # ── Top 5 ranking ─────────────────────────────────────────────────────────
    try:
        r_rk = await db.execute(
            text("""
                SELECT a.apostador_id,
                       (COALESCE(SUM(a.puntos),0) + COALESCE(SUM(a.puntos_bonus),0))::int AS pts_partidos,
                       SUM(CASE WHEN p.estado='finalizado' AND a.puntos>0 AND a.puntos%3=0  THEN 1 ELSE 0 END)::int AS plenos,
                       SUM(CASE WHEN p.estado='finalizado' AND a.puntos>0 AND a.puntos%3<>0 THEN 1 ELSE 0 END)::int AS aciertos
                FROM apuesta a
                JOIN partido p ON p.id = a.partido_id
                WHERE p.torneo_id = :tid
                GROUP BY a.apostador_id
                ORDER BY pts_partidos DESC
                LIMIT 5
            """),
            {"tid": torneo_id},
        )
        top_rows = [dict(r) for r in r_rk.mappings()]
        # Globales
        r_glob = await db.execute(
            text("SELECT apostador_id, pts_total FROM puntaje_global WHERE torneo_id = :tid"),
            {"tid": torneo_id},
        )
        glob_map = {r["apostador_id"]: r["pts_total"] for r in r_glob.mappings()}
        for row in top_rows:
            row["pts_globales"] = glob_map.get(row["apostador_id"], 0) or 0
            row["pts_total"]    = row["pts_partidos"] + row["pts_globales"]

        top_ids = [r["apostador_id"] for r in top_rows]
        if top_ids:
            async with _app_engine.connect() as conn:
                ur = await conn.execute(
                    text("SELECT id, username FROM users WHERE id = ANY(:ids)"),
                    {"ids": top_ids},
                )
                name_map = {r["id"]: r["username"] for r in ur.mappings()}
            for row in top_rows:
                row["nombre"] = name_map.get(row["apostador_id"], f"#{row['apostador_id']}")

        top_rows.sort(key=lambda r: -r["pts_total"])
        result["top5"] = top_rows
    except Exception:
        await db.rollback()
        result["top5"] = []

    # ── Stats generales ───────────────────────────────────────────────────────
    try:
        r_stats = await db.execute(
            text("""
                SELECT
                  COUNT(*)::int                                                               AS total_apuestas,
                  SUM(CASE WHEN p.estado='finalizado' AND a.puntos>0 AND a.puntos%3=0  THEN 1 ELSE 0 END)::int AS plenos,
                  SUM(CASE WHEN p.estado='finalizado' AND a.puntos>0 AND a.puntos%3<>0 THEN 1 ELSE 0 END)::int AS aciertos,
                  SUM(CASE WHEN p.estado='finalizado' AND a.puntos=0  THEN 1 ELSE 0 END)::int AS fallos,
                  SUM(CASE WHEN p.estado='finalizado' THEN 1 ELSE 0 END)::int                AS apuestas_resueltas
                FROM apuesta a
                JOIN partido p ON p.id = a.partido_id
                WHERE p.torneo_id = :tid
            """),
            {"tid": torneo_id},
        )
        stats = dict(r_stats.mappings().first() or {})
        # Total apostadores activos
        async with _app_engine.connect() as conn:
            r_ap = await conn.execute(
                text("""
                    SELECT COUNT(*)::int AS total FROM users u
                    JOIN user_roles ur ON ur.user_id = u.id
                    JOIN roles r ON r.id = ur.role_id
                    WHERE r.name='apostador' AND u.is_active=TRUE
                """)
            )
            stats["total_apostadores"] = r_ap.scalar() or 0
        result["stats"] = stats
    except Exception:
        await db.rollback()
        result["stats"] = {}

    # ── Mensajes recientes ────────────────────────────────────────────────────
    try:
        r_msg = await db.execute(
            text("""
                SELECT id, titulo, contenido, created_at
                FROM mensaje_admin
                WHERE torneo_id = :tid
                ORDER BY created_at DESC LIMIT 3
            """),
            {"tid": torneo_id},
        )
        mensajes = []
        for row in r_msg.mappings():
            d = dict(row)
            if d.get("created_at") and hasattr(d["created_at"], "isoformat"):
                d["created_at"] = d["created_at"].isoformat()
            mensajes.append(d)
        result["mensajes_recientes"] = mensajes
    except Exception:
        await db.rollback()
        result["mensajes_recientes"] = []

    return result


@router.get("/partidos-hoy/{torneo_id}", summary="Partidos del día actual con marcador (sin auth)")
async def partidos_hoy(torneo_id: int, db: DBSession, fecha: str = None) -> list[dict]:
    """
    Devuelve todos los partidos del día (parámetro fecha=YYYY-MM-DD, por defecto hoy)
    con estado, marcador y datos de equipo. Sin autenticación — lectura pública.
    """
    from datetime import date, timedelta, datetime
    try:
        target = date.fromisoformat(fecha) if fecha else date.today()
    except ValueError:
        target = date.today()
    # Ventana amplia: cubre cualquier timezone del mundo (UTC-12 a UTC+14).
    # IMPORTANTE: convertir a datetime ANTES de aplicar timedelta con horas;
    # date - timedelta(hours=N) ignora las horas (solo usa .days), dejando
    # una ventana de 1 día que excluye partidos nocturnos en UTC.
    target_dt = datetime(target.year, target.month, target.day)
    desde = target_dt - timedelta(hours=14)   # cubre UTC+14 (línea de fecha)
    hasta = target_dt + timedelta(hours=38)   # cubre UTC-12 + 2h de margen

    rp = await db.execute(
        text("""
            SELECT p.id, p.estado, f.tipo AS fase_tipo,
                   p.fecha,
                   p.goles_local, p.goles_visitante,
                   COALESCE(el.nombre_es, el.nombre) AS local_nombre,
                   el.logo_url AS local_logo,
                   COALESCE(ev.nombre_es, ev.nombre) AS visit_nombre,
                   ev.logo_url AS visit_logo
            FROM partido p
            JOIN fase f ON f.id = p.fase_id
            LEFT JOIN equipo el ON el.id = p.equipo_local_id
            LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
            WHERE p.torneo_id = :tid
              AND p.fecha >= :desde AND p.fecha < :hasta
            ORDER BY p.fecha
        """),
        {"tid": torneo_id, "desde": desde, "hasta": hasta},
    )
    result = []
    for r in rp.mappings():
        d = dict(r)
        if d.get("fecha") and hasattr(d["fecha"], "isoformat"):
            d["fecha"] = d["fecha"].isoformat()
        result.append(d)
    return result


@router.post("/finalizar-partido/{partido_id}", summary="Admin: finalizar partido manualmente y recalcular puntajes (emergencia)")
async def finalizar_partido_manual(
    partido_id: int,
    current: CurrentUser,
    db: DBSession,
    goles_local: int = Query(..., ge=0, le=30, description="Goles del equipo local"),
    goles_visitante: int = Query(..., ge=0, le=30, description="Goles del equipo visitante"),
    penales_local: int | None = Query(None, ge=0, le=30, description="Penales tanda local (opcional, solo si hubo tanda)"),
    penales_visitante: int | None = Query(None, ge=0, le=30, description="Penales tanda visitante (opcional)"),
) -> dict:
    """
    Fuerza el cierre de un partido cuando la API externa queda trabada.
    1. Actualiza partido: goles, estado='finalizado', penales tanda, equipo_clasificado_id.
    2. Recalcula standings (PJ/PG/PE/PP/GF/GC/Pts).
    3. Avanza bracket (asigna ganadores a siguiente fase).
    4. Recalcula puntajes de apostadores.
    Solo admin/superadmin.
    """
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    # ── Cargar partido ────────────────────────────────────────────────────────
    rp = await db.execute(
        text("""
            SELECT p.id, p.torneo_id, p.fase_id, p.equipo_local_id, p.equipo_visitante_id,
                   p.estado, p.goles_local, p.goles_visitante,
                   COALESCE(el.nombre_es, el.nombre) AS local_nombre,
                   COALESCE(ev.nombre_es, ev.nombre) AS visitante_nombre,
                   f.tipo AS fase_tipo
            FROM partido p
            JOIN fase f ON f.id = p.fase_id
            LEFT JOIN equipo el ON el.id = p.equipo_local_id
            LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
            WHERE p.id = :pid
        """),
        {"pid": partido_id},
    )
    row = rp.mappings().fetchone()
    if not row:
        raise HTTPException(404, f"Partido {partido_id} no encontrado")

    torneo_id = row["torneo_id"]
    fase_tipo  = row["fase_tipo"]

    # ── Determinar equipo_clasificado_id (KO: ganador por goles o penales tanda) ──
    equipo_clasificado_id = None
    if fase_tipo != "grupo":
        if goles_local != goles_visitante:
            equipo_clasificado_id = row["equipo_local_id"] if goles_local > goles_visitante else row["equipo_visitante_id"]
        elif penales_local is not None and penales_visitante is not None:
            equipo_clasificado_id = row["equipo_local_id"] if penales_local > penales_visitante else row["equipo_visitante_id"]

    # ── UPDATE partido ────────────────────────────────────────────────────────
    await db.execute(
        text("""
            UPDATE partido
            SET goles_local           = :gl,
                goles_visitante       = :gv,
                estado                = 'finalizado',
                penales_local         = :pl,
                penales_visitante     = :pv,
                equipo_clasificado_id = COALESCE(:ecid, equipo_clasificado_id),
                minuto_actual         = NULL
            WHERE id = :pid
        """),
        {
            "gl":   goles_local,
            "gv":   goles_visitante,
            "pl":   penales_local,
            "pv":   penales_visitante,
            "ecid": equipo_clasificado_id,
            "pid":  partido_id,
        },
    )
    await db.commit()

    # ── Cadena de recálculo ───────────────────────────────────────────────────
    recap: dict = {}

    # 1. Standings de grupos
    try:
        await _recalc_participacion(db, torneo_id)
        await db.commit()
        recap["standings_ok"] = True
    except Exception as e:
        recap["standings_error"] = str(e)

    # 2. Avanzar bracket KO
    try:
        maps = await ko_scoring.build_num_maps(db, torneo_id)
        await _avanzar_bracket(db, torneo_id, maps)
        await db.commit()
        recap["bracket_ok"] = True
    except Exception as e:
        recap["bracket_error"] = str(e)

    # 3. Calcular puntajes apostadores
    try:
        from app.services.scoring.registry import get_engine as _get_engine
        from app.services.scoring.calculator import ScoringCalculator
        rc = await db.execute(
            text("SELECT c.codigo FROM competicion c JOIN torneo t ON t.competicion_id=c.id WHERE t.id=:tid"),
            {"tid": torneo_id},
        )
        comp_row = rc.mappings().fetchone()
        engine = _get_engine(comp_row["codigo"] if comp_row else None)
        result = await ScoringCalculator(db).calculate(torneo_id, engine)
        await ScoringCalculator(db).calculate_global(torneo_id, engine)
        await db.commit()
        recap["puntajes_ok"] = True
        recap["puntajes_procesados"] = result.get("apostadores_procesados", 0)
    except Exception as e:
        recap["puntajes_error"] = str(e)

    await _audit_log(
        f"finalizar_partido_manual:{partido_id}:{goles_local}-{goles_visitante}",
        "bets", current=current, method="POST",
        detail={"partido_id": partido_id, "goles_local": goles_local,
                "goles_visitante": goles_visitante, "penales_local": penales_local,
                "penales_visitante": penales_visitante, "equipo_clasificado_id": equipo_clasificado_id},
    )

    return {
        "ok": True,
        "partido_id": partido_id,
        "partido": f"{row['local_nombre']} {goles_local}–{goles_visitante} {row['visitante_nombre']}",
        "estado": "finalizado",
        "equipo_clasificado_id": equipo_clasificado_id,
        **recap,
    }


@router.get("/hay-partido-activo/{torneo_id}", summary="¿Hay algún partido activo o por empezar? (para sync automático)")
async def hay_partido_activo(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """
    Devuelve {activo: bool, partidos: [...]} para que el sync automático
    decida si debe correr o no.

    Un partido es "activo" si:
      - Empezó hace entre 15 y 210 minutos (sync arranca al min 15 del primer tiempo)
      - Y su estado NO es 'finalizado'

    La ventana empieza en +15 min para evitar calls a la API antes de que haya
    datos de eventos disponibles. 210 min cubre 90 min regulares + alargue + penales.

    Usado por sync_auto.py: si activo=false, el script sale sin hacer ninguna llamada a API-Football.
    """
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    r = await db.execute(
        text("""
            SELECT p.id, p.fecha, p.estado,
                   COALESCE(el.nombre_es, el.nombre) AS local_nombre,
                   COALESCE(ev.nombre_es, ev.nombre) AS visitante_nombre,
                   f.tipo AS fase_tipo
            FROM partido p
            JOIN fase f ON f.id = p.fase_id
            LEFT JOIN equipo el ON el.id = p.equipo_local_id
            LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
            WHERE p.torneo_id = :tid
              AND p.estado != 'finalizado'
              AND p.fecha IS NOT NULL
              AND (
                -- Ventana temporal normal (min 15 a 300 desde inicio)
                (p.fecha <= (NOW() AT TIME ZONE 'UTC') - INTERVAL '15 minutes'
                 AND p.fecha >= (NOW() AT TIME ZONE 'UTC') - INTERVAL '300 minutes')
                OR
                -- Partido ya marcado en_juego en BD: sincronizar siempre hasta que finalice
                p.estado = 'en_juego'
              )
            ORDER BY p.fecha
        """),
        {"tid": torneo_id},
    )
    partidos = []
    for row in r.mappings():
        d = dict(row)
        if d.get("fecha"):
            d["fecha"] = d["fecha"].isoformat() if hasattr(d["fecha"], "isoformat") else str(d["fecha"])
        partidos.append(d)

    return {
        "activo":   len(partidos) > 0,
        "partidos": partidos,
        "total":    len(partidos),
    }


@router.get("/partidos-en-vivo/{torneo_id}", summary="Partidos activos hoy con datos live (score, cards, VAR…)")
async def partidos_en_vivo(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """
    Devuelve los partidos de hoy + en juego ahora + finalizados en las últimas 3h,
    con todos los campos actualizados por el sync (score, amarillas, rojas, VAR,
    minuto_primer_gol, penales tanda, estado, minuto_actual).
    Diseñado para el tab "En Vivo" con auto-refresh cada 30s en el UI.
    """
    import json as _json

    r = await db.execute(
        text("""
            SELECT
                p.id, p.fecha, p.estado,
                NULL::int AS numero,
                p.goles_local, p.goles_visitante,
                p.penales_local, p.penales_visitante,
                p.amarillas, p.rojas, p.decisiones_var,
                p.penales_partido,
                p.minuto_primer_gol,
                p.minuto_actual,
                COALESCE(el.nombre_es, el.nombre) AS local_nombre,
                el.logo_url  AS local_logo,
                COALESCE(ev.nombre_es, ev.nombre) AS visitante_nombre,
                ev.logo_url  AS visitante_logo,
                f.nombre     AS fase_nombre,
                f.tipo       AS fase_tipo,
                p.eventos_api::text      AS eventos_api_raw,
                p.estadisticas_api::text AS estadisticas_api_raw,
                -- estado del monitor (para detectar HT / descanso)
                mpe.estado_interno AS estado_monitor,
                mpe.api_status_raw,
                -- predicción del apostador actual
                ap.pred_local, ap.pred_visitante,
                ap.pred_penales_local_tanda, ap.pred_penales_visitante_tanda,
                ap.pred_amarillas, ap.pred_rojas, ap.pred_var, ap.pred_minuto_gol,
                ap.pred_penales_partido,
                -- puntaje calculado para este partido
                pd.pts_resultado, pd.pts_marcador,
                pd.pts_amarillas, pd.pts_rojas, pd.pts_var, pd.pts_minuto,
                pd.pts_penales_tanda,
                (COALESCE(pd.pts_resultado,0) + COALESCE(pd.pts_marcador,0) +
                 COALESCE(pd.pts_amarillas,0) + COALESCE(pd.pts_rojas,0) +
                 COALESCE(pd.pts_var,0)       + COALESCE(pd.pts_minuto,0) +
                 COALESCE(pd.pts_penales_tanda,0)) AS pts_total
            FROM partido p
            JOIN fase f ON f.id = p.fase_id
            LEFT JOIN equipo el ON el.id = p.equipo_local_id
            LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
            LEFT JOIN apuesta ap ON ap.partido_id = p.id AND ap.apostador_id = :uid
            LEFT JOIN puntaje_detalle pd ON pd.partido_id = p.id AND pd.apostador_id = :uid
            LEFT JOIN monitor_partido_estado mpe ON mpe.partido_id = p.id
            WHERE p.torneo_id = :tid
              AND p.fecha IS NOT NULL
              AND (
                  p.estado = 'en_juego'
                  OR (p.estado != 'finalizado'
                      AND p.fecha BETWEEN (NOW() AT TIME ZONE 'UTC') - INTERVAL '4 hours'
                                      AND (NOW() AT TIME ZONE 'UTC') + INTERVAL '20 hours')
                  OR (p.estado = 'finalizado'
                      AND p.fecha >= (NOW() AT TIME ZONE 'UTC') - INTERVAL '3 hours')
              )
            ORDER BY
                CASE p.estado
                    WHEN 'en_juego'   THEN 1
                    WHEN 'programado' THEN 2
                    WHEN 'finalizado' THEN 3
                    ELSE 4
                END,
                p.fecha
        """),
        {"tid": torneo_id, "uid": current.id},
    )
    partidos = []
    for row in r.mappings():
        d = dict(row)
        if d.get("fecha") and hasattr(d["fecha"], "isoformat"):
            d["fecha"] = d["fecha"].isoformat()
        # Parsear JSON raw de eventos y estadísticas
        raw_ev = d.pop("eventos_api_raw", None)
        raw_st = d.pop("estadisticas_api_raw", None)
        try:
            d["eventos_api"] = _json.loads(raw_ev) if raw_ev else []
        except Exception:
            d["eventos_api"] = []
        try:
            d["estadisticas_api"] = _json.loads(raw_st) if raw_st else []
        except Exception:
            d["estadisticas_api"] = []
        partidos.append(d)

    return {"partidos": partidos, "total": len(partidos)}


@router.get("/noticias", summary="Noticias del Mundial desde Google News RSS (proxy backend)")
async def get_noticias(
    q: str = "Copa del Mundo 2026 FIFA",
    hl: str = "es",
    gl: str = "MX",
    current: CurrentUser = None,
) -> dict:
    """
    Proxy server-side para Google News RSS.
    Evita problemas CORS del navegador. Devuelve lista de items.
    """
    import xml.etree.ElementTree as ET
    import httpx

    ceid = f"{gl}:{hl}"
    url = f"https://news.google.com/rss/search?q={q}&hl={hl}&gl={gl}&ceid={ceid}"
    headers = {
        "User-Agent": "Mozilla/5.0 (compatible; BECBUC-News/1.0)",
        "Accept": "application/rss+xml, application/xml, text/xml",
    }
    try:
        async with httpx.AsyncClient(timeout=15, follow_redirects=True) as client:
            resp = await client.get(url, headers=headers)
            resp.raise_for_status()
            xml_text = resp.text
    except Exception as e:
        raise HTTPException(502, f"Error al obtener noticias: {e}")

    try:
        root = ET.fromstring(xml_text)
        channel = root.find("channel")
        if channel is None:
            return {"items": [], "error": "RSS sin canal"}
        items = []
        for item in channel.findall("item")[:25]:
            title_raw = (item.findtext("title") or "").strip()
            link  = item.findtext("link") or ""
            date  = item.findtext("pubDate") or ""
            # Google News title format: "Article Title - Source Name"
            parts = title_raw.rsplit(" - ", 1)
            title  = parts[0].strip() if len(parts) > 1 else title_raw
            source = parts[1].strip() if len(parts) > 1 else ""
            items.append({"title": title, "source": source, "link": link, "date": date})
        return {"items": items, "query": q}
    except ET.ParseError as e:
        raise HTTPException(502, f"Error al parsear RSS: {e}")


@router.get("/fases-estado/{torneo_id}", summary="Estado de todas las fases del torneo (para consola admin)")
async def fases_estado(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    r = await db.execute(
        text("""
            SELECT
                f.id, f.nombre, f.tipo, f.orden,
                COALESCE(f.visible_apostador, TRUE) AS visible_apostador,
                COUNT(p.id)::int AS total,
                SUM(CASE WHEN p.estado='finalizado' THEN 1 ELSE 0 END)::int AS finalizados,
                SUM(CASE WHEN p.estado!='finalizado' THEN 1 ELSE 0 END)::int AS pendientes
            FROM fase f
            LEFT JOIN partido p ON p.fase_id = f.id
            WHERE f.torneo_id = :tid
            GROUP BY f.id, f.nombre, f.tipo, f.orden, f.visible_apostador
            ORDER BY f.orden, f.nombre
        """),
        {"tid": torneo_id},
    )
    fases = []
    for row in r.mappings():
        d = dict(row)
        d["concluida"] = d["total"] > 0 and d["pendientes"] == 0
        d["sin_partidos"] = d["total"] == 0
        fases.append(d)

    # Marcar cuál es la "fase activa" sugerida: la primera no concluida con partidos,
    # respetando que la anterior esté concluida.
    fase_activa_id = None
    prev_ok = True
    for f in fases:
        if f["sin_partidos"]:
            continue
        if prev_ok and not f["concluida"]:
            fase_activa_id = f["id"]
            break
        prev_ok = f["concluida"]

    return {"torneo_id": torneo_id, "fases": fases, "fase_activa_sugerida": fase_activa_id}


@router.get("/fases-bloqueo/{torneo_id}",
            summary="Estado de bloqueo manual por fase (admin)")
async def get_fases_bloqueo(torneo_id: int, current: CurrentUser, db: DBSession) -> list[dict]:
    """Devuelve todas las fases del torneo con su estado de bloqueo manual e inicio/fin calculados."""
    r = await db.execute(
        text("""
            SELECT f.id, f.nombre, f.tipo, f.orden,
                   COALESCE(f.bloqueada, FALSE) AS bloqueada,
                   COUNT(p.id)::int AS total,
                   SUM(CASE WHEN p.estado='finalizado' THEN 1 ELSE 0 END)::int AS finalizados,
                   MIN(p.fecha) AS fecha_inicio,
                   MAX(p.fecha) AS fecha_fin
            FROM fase f
            LEFT JOIN partido p ON p.fase_id = f.id
            WHERE f.torneo_id = :tid
            GROUP BY f.id, f.nombre, f.tipo, f.orden
            ORDER BY f.orden, f.nombre
        """),
        {"tid": torneo_id},
    )
    rows = []
    for row in r.mappings():
        d = dict(row)
        # Serializar fechas
        d["fecha_inicio"] = d["fecha_inicio"].isoformat() if d.get("fecha_inicio") else None
        d["fecha_fin"] = d["fecha_fin"].isoformat() if d.get("fecha_fin") else None
        rows.append(d)
    return rows


@router.patch("/fases-bloqueo-grupo/{torneo_id}",
              summary="Bloquear/desbloquear todas las fases de grupo del torneo (admin)")
async def set_fases_grupo_bloqueo(torneo_id: int, body: dict, current: CurrentUser, db: DBSession) -> dict:
    """Actualiza bloqueada=True/False para TODAS las fases de tipo 'grupo' del torneo."""
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")
    bloqueada = bool(body.get("bloqueada", False))
    await db.execute(
        text("UPDATE fase SET bloqueada = :b WHERE torneo_id = :tid AND tipo = 'grupo'"),
        {"b": bloqueada, "tid": torneo_id},
    )
    await db.commit()
    await _audit_log("fase:bloqueo_grupo", "bets", current=current, method="PATCH",
                     path=f"/api/v1/bets/fases-bloqueo-grupo/{torneo_id}",
                     resource_id=str(torneo_id),
                     details={"torneo_id": torneo_id, "bloqueada": bloqueada})
    return {"ok": True, "torneo_id": torneo_id, "bloqueada": bloqueada}


@router.patch("/fases-bloqueo/{fase_id}",
              summary="Actualizar bloqueo manual de una fase (admin)")
async def set_fase_bloqueo(fase_id: int, body: dict, current: CurrentUser, db: DBSession) -> dict:
    """Actualiza el campo bloqueada de una fase. body: {bloqueada: bool}"""
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")
    bloqueada = bool(body.get("bloqueada", False))
    r = await db.execute(
        text("UPDATE fase SET bloqueada = :b WHERE id = :fid RETURNING id, nombre, bloqueada"),
        {"b": bloqueada, "fid": fase_id},
    )
    row = r.mappings().first()
    if not row:
        raise HTTPException(404, "Fase no encontrada")
    await db.commit()
    await _audit_log("fase:bloqueo", "bets", current=current, method="PATCH",
                     path=f"/api/v1/bets/fases-bloqueo/{fase_id}",
                     resource_id=str(fase_id),
                     details={"fase": row["nombre"], "bloqueada": bloqueada})
    return {"ok": True, "fase_id": fase_id, "nombre": row["nombre"], "bloqueada": row["bloqueada"]}


@router.get("/fases-apuesta-estado/{torneo_id}",
            summary="Estado de apuesta por fase (bloqueada/habilitada) para cualquier usuario")
async def fases_apuesta_estado(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """Para cada fase indica si está habilitada o bloqueada para cargar apuestas.
    Bloqueada si: fase.bloqueada=true (manual admin) O fase concluida.
    La fecha del primer partido se muestra como referencia informativa.
    """
    r = await db.execute(
        text("""
            SELECT
                f.id, f.nombre, f.tipo, f.orden,
                COALESCE(f.bloqueada, FALSE) AS bloqueada_manual,
                COUNT(p.id)::int                                              AS total,
                SUM(CASE WHEN p.estado='finalizado' THEN 1 ELSE 0 END)::int  AS finalizados,
                MIN(p.fecha)                                                  AS fecha_inicio_fase,
                MIN(p.fecha) - INTERVAL '5 hours'                            AS fecha_corte
            FROM fase f
            LEFT JOIN partido p ON p.fase_id = f.id
            WHERE f.torneo_id = :tid
            GROUP BY f.id, f.nombre, f.tipo, f.orden
            ORDER BY f.orden, f.nombre
        """),
        {"tid": torneo_id},
    )
    rows = [dict(x) for x in r.mappings()]

    def _fmt(dt) -> str | None:
        if dt is None:
            return None
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return f"{dt.day}/{dt.month}/{dt.year} {dt.strftime('%H:%M')}"

    fases = []
    for d in rows:
        total       = d["total"] or 0
        finalizados = d["finalizados"] or 0
        concluida   = total > 0 and finalizados == total
        bloqueada_manual = d["bloqueada_manual"]
        fecha_corte  = d["fecha_corte"]
        fecha_inicio = d["fecha_inicio_fase"]

        # Normalizar tz
        if fecha_corte is not None and fecha_corte.tzinfo is None:
            fecha_corte = fecha_corte.replace(tzinfo=timezone.utc)
        if fecha_inicio is not None and fecha_inicio.tzinfo is None:
            fecha_inicio = fecha_inicio.replace(tzinfo=timezone.utc)

        if bloqueada_manual:
            bloqueada = True
            motivo    = "Bloqueada manualmente por el administrador"
            icono     = "lock"
        elif concluida:
            bloqueada = True
            motivo    = "Fase concluida — período de apuesta cerrado"
            icono     = "lock"
        else:
            bloqueada = False
            motivo    = f"Abierto · primer partido {_fmt(fecha_inicio)}" if fecha_inicio else "Habilitada para cargar apuestas"
            icono     = "lock-open"

        fases.append({
            "fase_id":       d["id"],
            "nombre":        d["nombre"],
            "tipo":          d["tipo"],
            "orden":         d["orden"],
            "total":         total,
            "finalizados":   finalizados,
            "concluida":     concluida,
            "fecha_inicio":  fecha_inicio.isoformat() if fecha_inicio else None,
            "fecha_corte":   fecha_corte.isoformat()  if fecha_corte  else None,
            "fecha_corte_fmt": _fmt(fecha_corte),
            "bloqueada":     bloqueada,
            "motivo":        motivo,
            "icono":         icono,
        })

    fase_habilitada_id = next((f["fase_id"] for f in fases if not f["bloqueada"]), None)
    return {"torneo_id": torneo_id, "fases": fases, "fase_habilitada_id": fase_habilitada_id}


async def _calc_standings_reales(db: DBSession, torneo_id: int) -> dict:
    """Calcula standings de grupos desde los resultados reales (goles en tabla partido).
    También acumula tarjetas por equipo (fair_play_pts) si las columnas existen."""
    # Garantizar columnas de tarjetas por equipo (idempotente)
    for _col in [
        "ALTER TABLE partido ADD COLUMN IF NOT EXISTS local_amarillas     INT",
        "ALTER TABLE partido ADD COLUMN IF NOT EXISTS visitante_amarillas INT",
        "ALTER TABLE partido ADD COLUMN IF NOT EXISTS local_rojas         INT",
        "ALTER TABLE partido ADD COLUMN IF NOT EXISTS visitante_rojas     INT",
        "ALTER TABLE participacion ADD COLUMN IF NOT EXISTS fair_play_pts INT DEFAULT 0",
    ]:
        try:
            await db.execute(text(_col))
        except Exception:
            pass
    r_teams = await db.execute(
        text("""
            SELECT f.id AS fase_id, f.nombre AS fase_nombre,
                   e.id AS equipo_id, e.fifa_ranking,
                   COALESCE(e.nombre_es, e.nombre) AS nombre,
                   e.logo_url,
                   COALESCE(e.codigo_iso, '') AS codigo_iso
            FROM participacion pa
            JOIN equipo e ON e.id = pa.equipo_id
            JOIN fase f ON f.id = pa.fase_id
            WHERE f.torneo_id = :tid AND f.tipo = 'grupo' AND f.nombre NOT ILIKE '%mejores%'
        """),
        {"tid": torneo_id},
    )
    grupos: dict[int, dict] = {}
    grupo_nombres: dict[int, str] = {}
    for row in r_teams.mappings():
        fid = row["fase_id"]
        grupo_nombres[fid] = row["fase_nombre"]
        grupos.setdefault(fid, {})[row["equipo_id"]] = {
            "equipo_id": row["equipo_id"],
            "nombre": row["nombre"],
            "logo_url": row["logo_url"],
            "fifa_ranking": row["fifa_ranking"] or 999,
            "fair_play_pts": 0,
            "pj": 0, "pg": 0, "pe": 0, "pp": 0,
            "gf": 0, "gc": 0, "gd": 0, "pts": 0,
        }

    r_parts = await db.execute(
        text("""
            SELECT f.id AS fase_id,
                   p.equipo_local_id, p.equipo_visitante_id,
                   p.goles_local, p.goles_visitante,
                   COALESCE(p.local_amarillas, 0)     AS local_amarillas,
                   COALESCE(p.visitante_amarillas, 0) AS visitante_amarillas,
                   COALESCE(p.local_rojas, 0)         AS local_rojas,
                   COALESCE(p.visitante_rojas, 0)     AS visitante_rojas
            FROM partido p
            JOIN fase f ON f.id = p.fase_id AND f.tipo = 'grupo' AND f.nombre NOT ILIKE '%mejores%'
            WHERE p.torneo_id = :tid AND p.estado = 'finalizado'
              AND p.goles_local IS NOT NULL AND p.goles_visitante IS NOT NULL
        """),
        {"tid": torneo_id},
    )
    # Guardar resultados por grupo para H2H
    resultados_por_grupo: dict[int, list[dict]] = {}
    for p in r_parts.mappings():
        fid = p["fase_id"]
        lid, vid = p["equipo_local_id"], p["equipo_visitante_id"]
        gl, gv = p["goles_local"], p["goles_visitante"]
        if fid not in grupos or lid not in grupos[fid] or vid not in grupos[fid]:
            continue
        loc, vis = grupos[fid][lid], grupos[fid][vid]
        loc["pj"] += 1; vis["pj"] += 1
        loc["gf"] += gl; vis["gf"] += gv
        loc["gc"] += gv; vis["gc"] += gl
        loc["gd"] += gl - gv; vis["gd"] += gv - gl
        if gl > gv:
            loc["pg"] += 1; loc["pts"] += 3; vis["pp"] += 1
        elif gl == gv:
            loc["pe"] += 1; loc["pts"] += 1; vis["pe"] += 1; vis["pts"] += 1
        else:
            vis["pg"] += 1; vis["pts"] += 3; loc["pp"] += 1
        # Fair play acumulado: amarilla=1pt, roja=3pts (menor = mejor conducta FIFA)
        loc["fair_play_pts"] += p["local_amarillas"]     + p["local_rojas"]     * 3
        vis["fair_play_pts"] += p["visitante_amarillas"] + p["visitante_rojas"] * 3
        # Guardar resultado para desempate H2H
        resultados_por_grupo.setdefault(fid, []).append(
            {"lid": lid, "vid": vid, "gl": gl, "gv": gv}
        )

    result = {}
    for fid, teams in grupos.items():
        # Ordenar usando criterio completo FIFA 2026:
        # pts → DG → GF → H2H pts → H2H DG → H2H GF → fair_play_pts (↓) → FIFA ranking → nombre
        sorted_teams = _sort_grupo_fifa(
            list(teams.values()),
            resultados_por_grupo.get(fid, []),
        )
        nombre = grupo_nombres[fid]
        grupo_letra = nombre.replace("Grupo ", "").strip()
        for idx, eq in enumerate(sorted_teams):
            eq["pos"] = idx + 1
            eq["grupo"] = grupo_letra
        # Estructura {"equipos": [...]} para compatibilidad con seleccionar_mejores_terceros
        result[grupo_letra] = {"fase_id": fid, "grupo": grupo_letra, "equipos": sorted_teams}
    return result


async def _recalc_participacion(db: DBSession, torneo_id: int) -> int:
    """Recalcula la tabla participacion (cuadro de grupos: PJ/PG/PE/PP/GF/GC/Pts/posicion)
    desde los resultados reales en la tabla partido. Idempotente.
    Devuelve el nº de filas actualizadas."""
    standings = await _calc_standings_reales(db, torneo_id)
    actualizadas = 0
    for grupo in standings.values():
        fase_id = grupo.get("fase_id")
        for eq in grupo.get("equipos", []):
            await db.execute(
                text("""
                    UPDATE participacion
                    SET pj=:pj, pg=:pg, pe=:pe, pp=:pp,
                        gf=:gf, gc=:gc, pts=:pts,
                        posicion=:pos, clasifica=:clasifica,
                        fair_play_pts=:fp
                    WHERE fase_id=:fid AND equipo_id=:eid
                """),
                {
                    "pj": eq["pj"], "pg": eq["pg"], "pe": eq["pe"], "pp": eq["pp"],
                    "gf": eq["gf"], "gc": eq["gc"], "pts": eq["pts"],
                    "pos": eq["pos"], "clasifica": eq["pos"] <= 2,
                    "fp":  eq.get("fair_play_pts", 0),
                    "fid": fase_id, "eid": eq["equipo_id"],
                },
            )
            actualizadas += 1
    return actualizadas


async def _ensure_detalle_table(db):
    """Crea la tabla de detalle de puntaje por partido/fase/item si no existe."""
    await db.execute(text("""
        CREATE TABLE IF NOT EXISTS puntaje_detalle (
            id              SERIAL PRIMARY KEY,
            torneo_id       INT NOT NULL,
            fase_id         INT NOT NULL,
            fase_tipo       VARCHAR(30) NOT NULL,
            fase_nombre     VARCHAR(80),
            partido_id      INT NOT NULL,
            apostador_id    INT NOT NULL,
            multiplicador   INT NOT NULL DEFAULT 1,
            pred_local      INT, pred_visitante INT,
            real_local      INT, real_visitante INT,
            pts_marcador_base  INT NOT NULL DEFAULT 0,
            pts_marcador       INT NOT NULL DEFAULT 0,
            pred_minuto     INT, real_minuto INT, gano_minuto BOOLEAN DEFAULT FALSE,
            pts_minuto      INT NOT NULL DEFAULT 0,
            pred_amarillas  INT, real_amarillas INT,
            pts_amarillas   INT NOT NULL DEFAULT 0,
            pred_var        INT, real_var INT,
            pts_var         INT NOT NULL DEFAULT 0,
            pts_bonus       INT NOT NULL DEFAULT 0,
            pts_total       INT NOT NULL DEFAULT 0,
            created_at      TIMESTAMPTZ DEFAULT NOW(),
            UNIQUE (torneo_id, partido_id, apostador_id)
        )
    """))
    # Columnas v2 (se agregan si no existen — idempotente para tablas pre-existentes)
    for _col in [
        "ALTER TABLE puntaje_detalle ADD COLUMN IF NOT EXISTS pts_resultado INT DEFAULT 0",
        "ALTER TABLE puntaje_detalle ADD COLUMN IF NOT EXISTS pts_rojas INT DEFAULT 0",
        "ALTER TABLE puntaje_detalle ADD COLUMN IF NOT EXISTS pts_penales_partido INT DEFAULT 0",
        "ALTER TABLE puntaje_detalle ADD COLUMN IF NOT EXISTS pts_penales_tanda INT DEFAULT 0",
        "ALTER TABLE puntaje_detalle ADD COLUMN IF NOT EXISTS pts_equipo INT DEFAULT 0",
    ]:
        try:
            await db.execute(text(_col))
        except Exception:
            pass


# ── Pronósticos globales A-G ─────────────────────────────────────────────────

@router.post("/apuestas-globales/{torneo_id}", summary="Guardar pronósticos globales A-G del apostador")
async def guardar_apuestas_globales(
    torneo_id: int,
    body: dict,
    current: CurrentUser,
    db: DBSession,
) -> dict:
    """
    Upsert de los pronósticos globales (A-G) del apostador autenticado.
    Campos aceptados: pred_campeon_id, pred_finalista1_id, pred_finalista2_id,
    pred_goleador, pred_peor_equipo_id, pred_goleada_ganador, pred_goleada_perdedor,
    pred_etapa_paraguay, pred_goles_paraguay.
    """
    uid = current.id
    await db.execute(
        text("""
            INSERT INTO apuesta_global
              (torneo_id, apostador_id,
               pred_campeon_id, pred_finalista1_id, pred_finalista2_id,
               pred_goleador, pred_peor_equipo_id,
               pred_goleada_ganador, pred_goleada_perdedor,
               pred_etapa_paraguay, pred_goles_paraguay,
               updated_at)
            VALUES
              (:tid, :uid,
               :campeon, :fin1, :fin2,
               :goleador, :peor,
               :gol_g, :gol_p,
               :etapa, :goles,
               NOW())
            ON CONFLICT (torneo_id, apostador_id) DO UPDATE SET
               pred_campeon_id      = EXCLUDED.pred_campeon_id,
               pred_finalista1_id   = EXCLUDED.pred_finalista1_id,
               pred_finalista2_id   = EXCLUDED.pred_finalista2_id,
               pred_goleador        = EXCLUDED.pred_goleador,
               pred_peor_equipo_id  = EXCLUDED.pred_peor_equipo_id,
               pred_goleada_ganador = EXCLUDED.pred_goleada_ganador,
               pred_goleada_perdedor= EXCLUDED.pred_goleada_perdedor,
               pred_etapa_paraguay  = EXCLUDED.pred_etapa_paraguay,
               pred_goles_paraguay  = EXCLUDED.pred_goles_paraguay,
               updated_at           = NOW()
        """),
        {
            "tid":     torneo_id,
            "uid":     uid,
            "campeon": body.get("pred_campeon_id"),
            "fin1":    body.get("pred_finalista1_id"),
            "fin2":    body.get("pred_finalista2_id"),
            "goleador":body.get("pred_goleador"),
            "peor":    body.get("pred_peor_equipo_id"),
            "gol_g":   body.get("pred_goleada_ganador"),
            "gol_p":   body.get("pred_goleada_perdedor"),
            "etapa":   body.get("pred_etapa_paraguay"),
            "goles":   body.get("pred_goles_paraguay"),
        },
    )
    await db.commit()
    return {"ok": True, "torneo_id": torneo_id, "apostador_id": uid}


@router.get("/apuestas-globales/{torneo_id}", summary="Leer pronósticos globales A-G del apostador")
async def leer_apuestas_globales(
    torneo_id: int,
    current: CurrentUser,
    db: DBSession,
    for_apostador_id: int = None,
) -> dict:
    """
    Devuelve los pronósticos globales del apostador autenticado para el torneo.
    Incluye los puntajes ya calculados (si existen en puntaje_global).
    Admin puede pasar for_apostador_id para ver pronósticos de otro apostador.
    """
    uid = current.id
    if for_apostador_id and await _check_admin(current):
        uid = for_apostador_id
    r = await db.execute(
        text("SELECT * FROM apuesta_global WHERE torneo_id = :tid AND apostador_id = :uid"),
        {"tid": torneo_id, "uid": uid},
    )
    ag = r.mappings().first()

    rp = await db.execute(
        text("SELECT * FROM puntaje_global WHERE torneo_id = :tid AND apostador_id = :uid"),
        {"tid": torneo_id, "uid": uid},
    )
    pg = rp.mappings().first()

    return {
        "apuesta_global":  dict(ag)  if ag else {},
        "puntaje_global":  dict(pg)  if pg else {},
    }


@router.post("/resultados-globales/{torneo_id}",
             summary="Establecer resultados globales C/D para scoring (admin)")
async def set_resultados_globales(
    torneo_id: int,
    body: dict,
    current: CurrentUser,
    db: DBSession,
) -> dict:
    """
    Admin establece los resultados de los ítems que no son auto-computables:
      C: resultado_goleador (nombre del jugador, texto)
      D: resultado_peor_equipo_id (equipo_id del peor equipo en grupos)
    Agrega las columnas al torneo si no existen (idempotente).
    """
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    for col_sql in [
        "ALTER TABLE torneo ADD COLUMN IF NOT EXISTS resultado_goleador VARCHAR(100)",
        "ALTER TABLE torneo ADD COLUMN IF NOT EXISTS resultado_peor_equipo_id INT",
    ]:
        try:
            await db.execute(text(col_sql))
        except Exception:
            await db.rollback()

    updates = []
    params: dict = {"tid": torneo_id}
    if "resultado_goleador" in body:
        updates.append("resultado_goleador = :goleador")
        params["goleador"] = body["resultado_goleador"]
    if "resultado_peor_equipo_id" in body:
        updates.append("resultado_peor_equipo_id = :peor_eq")
        params["peor_eq"] = body["resultado_peor_equipo_id"]

    if updates:
        await db.execute(
            text(f"UPDATE torneo SET {', '.join(updates)} WHERE id = :tid"),
            params,
        )
        await db.commit()

    return {"ok": True, "torneo_id": torneo_id, "campos_actualizados": list(params.keys() - {"tid"})}


@router.post("/importar-globales-apostador/{torneo_id}",
             summary="Importar pronósticos globales A-G por nombre de equipo (admin/importación masiva)")
async def importar_globales_apostador(
    torneo_id: int,
    body: dict,
    current: CurrentUser,
    db: DBSession,
) -> dict:
    """
    Importa pronósticos globales A-G para un apostador identificado por alias/username.
    Acepta nombres de equipos como texto (inglés o español) y los resuelve a equipo.id.

    Body esperado (del importar-apuestas.html consolidado):
      apostador:             alias/username del apostador
      pred_campeon:          nombre equipo campeón
      pred_finalista1:       nombre equipo finalista 1 (mismo que campeón)
      pred_finalista2:       nombre equipo finalista 2 (el otro finalista)
      pred_goleador:         nombre del goleador (texto libre)
      pred_peor_equipo:      nombre equipo peor del mundial
      pred_etapa_paraguay:   etapa (texto: '4tos', 'grupos', 'semis', etc.)
      pred_goles_paraguay:   int
      pred_goleada_ganador:  int
      pred_goleada_perdedor: int
    """
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    alias = str(body.get("apostador") or "").replace("\xa0", "").strip().lower()
    if not alias:
        raise HTTPException(400, "Campo 'apostador' requerido")

    # ── Resolver apostador_id vía app_db (igual que importar_apuestas_grupos) ──
    user_id = None
    for uname_variant in [alias, alias.lstrip("@")]:
        async with _app_engine.connect() as aconn:
            ur = await aconn.execute(
                text("SELECT id FROM users WHERE LOWER(username) = :u AND is_active = TRUE LIMIT 1"),
                {"u": uname_variant}
            )
            row = ur.fetchone()
            if row:
                user_id = row[0]
                break

    if not user_id:
        return {"ok": False, "guardado": False,
                "error": f"Usuario '{alias}' no encontrado en BD"}

    apostador_id = user_id

    # ── Cargar equipos (nombre → id) ──────────────────────────────────────────
    equipos_rows = (await db.execute(
        text("SELECT id, nombre, nombre_es FROM equipo ORDER BY id")
    )).mappings().fetchall()

    equipos_db = len(equipos_rows)

    def _norm(s: str) -> str:
        import unicodedata, re
        s = str(s or "").strip().lower()
        s = unicodedata.normalize("NFD", s)
        s = "".join(c for c in s if unicodedata.category(c) != "Mn")
        s = re.sub(r"[^a-z0-9 ]", " ", s).strip()
        return re.sub(r" +", " ", s)

    # {nombre_norm → equipo_id}
    eq_map: dict[str, int] = {}
    for eq in equipos_rows:
        for col in ("nombre", "nombre_es"):
            v = eq[col]
            if v:
                k = _norm(v)
                if k not in eq_map:
                    eq_map[k] = eq["id"]

    def resolve_team(name_str) -> int | None:
        if not name_str:
            return None
        k = _norm(str(name_str))
        if k in eq_map:
            return eq_map[k]
        # substring match
        for eq_k, eq_id in eq_map.items():
            if k in eq_k or eq_k in k:
                return eq_id
        return None

    # ── Resolver campos ───────────────────────────────────────────────────────
    resueltos: dict = {}
    no_resueltos_input: dict = {}

    campeon_id    = resolve_team(body.get("pred_campeon"))
    fin1_id       = resolve_team(body.get("pred_finalista1") or body.get("pred_campeon"))
    fin2_id       = resolve_team(body.get("pred_finalista2"))
    peor_id       = resolve_team(body.get("pred_peor_equipo"))

    for campo, val, res_id in [
        ("campeon",    body.get("pred_campeon"),    campeon_id),
        ("finalista1", body.get("pred_finalista1"), fin1_id),
        ("finalista2", body.get("pred_finalista2"), fin2_id),
        ("peor_equipo",body.get("pred_peor_equipo"),peor_id),
    ]:
        resueltos[campo] = res_id
        if val and res_id is None:
            no_resueltos_input[campo] = val

    # Normalizar etapa Paraguay
    _ETAPA_MAP = {
        "4tos": "cuartos", "cuartos": "cuartos", "qf": "cuartos",
        "cuartos de final": "cuartos",
        "grupos": "grupo", "grupo": "grupo", "group stage": "grupo",
        "fase de grupos": "grupo", "fase grupos": "grupo",
        "32avos": "ronda32", "ronda32": "ronda32",
        "16avos": "ronda16", "octavos": "ronda16", "ronda16": "ronda16",
        "16avos de final": "ronda16",
        "semis": "semis", "semifinales": "semis", "semifinal": "semis",
        "final": "final", "finalista": "final",
        "tercer puesto": "final", "3er puesto": "final",
        "3rd place": "final", "tercer_puesto": "final",
    }
    etapa_raw = str(body.get("pred_etapa_paraguay") or "").strip()
    etapa_norm = _ETAPA_MAP.get(etapa_raw.lower(), etapa_raw) if etapa_raw else None

    def _to_int(v):
        if v is None or v == "": return None
        try: return int(v)
        except (ValueError, TypeError): return None

    goleador     = str(body.get("pred_goleador") or "").strip() or None
    goles_py     = _to_int(body.get("pred_goles_paraguay"))
    goleada_g    = _to_int(body.get("pred_goleada_ganador"))
    goleada_p    = _to_int(body.get("pred_goleada_perdedor"))

    # ── Upsert apuesta_global ─────────────────────────────────────────────────
    try:
        await db.execute(text("""
            INSERT INTO apuesta_global
                (apostador_id, torneo_id,
                 pred_campeon_id, pred_finalista1_id, pred_finalista2_id,
                 pred_goleador, pred_peor_equipo_id,
                 pred_goleada_ganador, pred_goleada_perdedor,
                 pred_etapa_paraguay, pred_goles_paraguay,
                 created_at, updated_at)
            VALUES
                (:uid, :tid, :camp, :fin1, :fin2,
                 :gol, :peor, :gl_g, :gl_p, :etapa, :goles_py,
                 NOW(), NOW())
            ON CONFLICT (apostador_id, torneo_id) DO UPDATE SET
                 pred_campeon_id      = EXCLUDED.pred_campeon_id,
                 pred_finalista1_id   = EXCLUDED.pred_finalista1_id,
                 pred_finalista2_id   = EXCLUDED.pred_finalista2_id,
                 pred_goleador        = EXCLUDED.pred_goleador,
                 pred_peor_equipo_id  = EXCLUDED.pred_peor_equipo_id,
                 pred_goleada_ganador = EXCLUDED.pred_goleada_ganador,
                 pred_goleada_perdedor= EXCLUDED.pred_goleada_perdedor,
                 pred_etapa_paraguay  = EXCLUDED.pred_etapa_paraguay,
                 pred_goles_paraguay  = EXCLUDED.pred_goles_paraguay,
                 updated_at           = NOW()
        """), {
            "uid":     apostador_id,
            "tid":     torneo_id,
            "camp":    campeon_id,
            "fin1":    fin1_id,
            "fin2":    fin2_id,
            "gol":     goleador,
            "peor":    peor_id,
            "gl_g":    goleada_g,
            "gl_p":    goleada_p,
            "etapa":   etapa_norm,
            "goles_py":goles_py,
        })
        await db.commit()
    except Exception as exc:
        await db.rollback()
        raise HTTPException(500, f"Error al guardar: {exc}")

    return {
        "ok":                 True,
        "guardado":           True,
        "apostador":          alias,
        "apostador_id":       apostador_id,
        "resueltos":          resueltos,
        "no_resueltos_input": no_resueltos_input,
        "equipos_db":         equipos_db,
    }


@router.get("/diagnostico-grupos/{torneo_id}", summary="Diagnóstico completo: NULLs, discrepancias y equipos (admin)")
async def diagnostico_grupos(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """Verifica NULLs en apuesta, discrepancias vs pronosticos_aux, y consistencia de equipo_local/visitante."""
    if not await _check_admin(current):
        raise HTTPException(403, "Admin requerido")

    # 1. Registros con pred_local o pred_visitante NULL
    r_null = await db.execute(text("""
        SELECT a.nombre_apostador, p.numero_fifa,
               el.nombre AS local, ev.nombre AS visitante,
               a.pred_local, a.pred_visitante
        FROM apuesta a
        JOIN partido p  ON p.id  = a.partido_id
        JOIN fase    f  ON f.id  = p.fase_id
        LEFT JOIN equipo el ON el.id = p.equipo_local_id
        LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
        WHERE f.torneo_id = :tid
          AND p.numero_fifa IS NOT NULL AND p.numero_fifa <= 72
          AND (a.pred_local IS NULL OR a.pred_visitante IS NULL)
        ORDER BY p.numero_fifa, a.nombre_apostador
    """), {"tid": torneo_id})
    nulls = [dict(r) for r in r_null.mappings()]

    # 2. Discrepancias pred vs pronosticos_aux
    r_disc = await db.execute(text("""
        SELECT a.nombre_apostador, p.numero_fifa,
               a.pred_local, a.pred_visitante,
               pa.goles_local AS paux_local, pa.goles_visitante AS paux_visitante
        FROM apuesta a
        JOIN partido p ON p.id = a.partido_id
        JOIN fase    f ON f.id = p.fase_id
        JOIN pronosticos_aux pa
          ON pa.numero_partido_fifa = p.numero_fifa
         AND LOWER(TRIM(pa.nombre)) = LOWER(TRIM(a.nombre_apostador))
        WHERE f.torneo_id = :tid
          AND p.numero_fifa <= 72
          AND a.pred_local IS NOT NULL
          AND (a.pred_local != pa.goles_local OR a.pred_visitante != pa.goles_visitante)
        ORDER BY p.numero_fifa, a.nombre_apostador
        LIMIT 20
    """), {"tid": torneo_id})
    discrepancias = [dict(r) for r in r_disc.mappings()]

    # 3. Consistencia de equipos: partido.equipo_local_id vs pronosticos_aux.idequipolocal
    r_eq = await db.execute(text("""
        SELECT DISTINCT pa.id_partido, p.numero_fifa,
               el.nombre AS partido_local, ev.nombre AS partido_visitante,
               pa.equipo_local AS paux_local_nombre, pa.equipo_visitante AS paux_visitante_nombre,
               p.equipo_local_id, pa.idequipolocal,
               p.equipo_visitante_id, pa.idequipovisitante,
               CASE WHEN p.equipo_local_id = pa.idequipolocal
                     AND p.equipo_visitante_id = pa.idequipovisitante
                    THEN 'OK' ELSE 'MISMATCH' END AS match_equipos
        FROM pronosticos_aux pa
        JOIN partido p ON p.numero_fifa = pa.numero_partido_fifa
        JOIN fase    f ON f.id = p.fase_id
        LEFT JOIN equipo el ON el.id = p.equipo_local_id
        LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
        WHERE f.torneo_id = :tid
          AND pa.numero_partido_fifa <= 72
          AND pa.idequipolocal IS NOT NULL
        ORDER BY p.numero_fifa
        LIMIT 80
    """), {"tid": torneo_id})
    equipos = [dict(r) for r in r_eq.mappings()]
    equipos_ok = sum(1 for e in equipos if e["match_equipos"] == "OK")
    equipos_mismatch = [e for e in equipos if e["match_equipos"] == "MISMATCH"]

    return {
        "nulls_en_apuesta": nulls,
        "total_nulls": len(nulls),
        "discrepancias_pred_vs_paux": discrepancias,
        "total_discrepancias": len(discrepancias),
        "equipos_verificados": len(equipos),
        "equipos_ok": equipos_ok,
        "equipos_mismatch": equipos_mismatch[:10],
        "total_mismatch_equipos": len(equipos_mismatch),
    }


@router.get("/verificar-apuestas-grupos/{torneo_id}", summary="Verificar consistencia apuesta vs pronosticos_aux para fase de grupos (admin)")
async def verificar_apuestas_grupos(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """Compara apuesta.pred_local/visitante contra pronosticos_aux para los 72 partidos de grupos."""
    if not await _check_admin(current):
        raise HTTPException(403, "Admin requerido")

    r = await db.execute(text("""
        SELECT
            COUNT(*)                                                                         AS total_pares,
            SUM(CASE WHEN a.pred_local IS NULL OR a.pred_visitante IS NULL THEN 1 ELSE 0 END) AS con_null,
            SUM(CASE WHEN a.pred_local IS NOT NULL
                      AND (a.pred_local != pa.goles_local OR a.pred_visitante != pa.goles_visitante)
                     THEN 1 ELSE 0 END)                                                      AS discrepancias,
            SUM(CASE WHEN a.pred_local = pa.goles_local
                      AND a.pred_visitante = pa.goles_visitante
                     THEN 1 ELSE 0 END)                                                      AS coinciden,
            COUNT(DISTINCT a.apostador_id)                                                   AS apostadores,
            COUNT(DISTINCT p.id)                                                              AS partidos_con_match
        FROM apuesta a
        JOIN partido p  ON p.id  = a.partido_id
        JOIN fase    f  ON f.id  = p.fase_id
        JOIN pronosticos_aux pa
          ON pa.numero_partido_fifa = p.numero_fifa
         AND LOWER(TRIM(pa.nombre)) = LOWER(TRIM(a.nombre_apostador))
        WHERE f.torneo_id = :tid
          AND p.numero_fifa IS NOT NULL
          AND p.numero_fifa <= 72
    """), {"tid": torneo_id})
    row = r.mappings().one()

    # Cuántos registros en apuesta NO tienen match en pronosticos_aux
    r2 = await db.execute(text("""
        SELECT COUNT(*) AS sin_match_paux
        FROM apuesta a
        JOIN partido p ON p.id = a.partido_id
        JOIN fase    f ON f.id = p.fase_id
        LEFT JOIN pronosticos_aux pa
          ON pa.numero_partido_fifa = p.numero_fifa
         AND LOWER(TRIM(pa.nombre)) = LOWER(TRIM(a.nombre_apostador))
        WHERE f.torneo_id = :tid
          AND p.numero_fifa IS NOT NULL AND p.numero_fifa <= 72
          AND pa.id IS NULL
    """), {"tid": torneo_id})
    sin_match = r2.scalar()

    # Cuántos pares en pronosticos_aux no tienen apuesta
    r3 = await db.execute(text("""
        SELECT COUNT(*) AS paux_sin_apuesta
        FROM pronosticos_aux pa
        JOIN partido p ON p.numero_fifa = pa.numero_partido_fifa
        JOIN fase    f ON f.id = p.fase_id
        LEFT JOIN apuesta a
          ON a.partido_id = p.id
         AND LOWER(TRIM(a.nombre_apostador)) = LOWER(TRIM(pa.nombre))
        WHERE f.torneo_id = :tid
          AND pa.numero_partido_fifa <= 72
          AND a.id IS NULL
    """), {"tid": torneo_id})
    paux_sin_apuesta = r3.scalar()

    return {
        "total_pares_matcheados": int(row["total_pares"]),
        "coinciden": int(row["coinciden"]),
        "con_null": int(row["con_null"]),
        "discrepancias": int(row["discrepancias"]),
        "apostadores": int(row["apostadores"]),
        "partidos_con_match": int(row["partidos_con_match"]),
        "apuesta_sin_match_en_paux": int(sin_match),
        "paux_sin_apuesta": int(paux_sin_apuesta),
        "ok": row["discrepancias"] == 0 and row["con_null"] == 0 and paux_sin_apuesta == 0,
    }


@router.post("/reload-apuestas-grupos/{torneo_id}", summary="Eliminar y recargar apuestas de grupos desde pronosticos_aux (admin)")
async def reload_apuestas_grupos(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """Elimina apuestas + puntaje_detalle de grupos y los recarga desde pronosticos_aux."""
    if not await _check_admin(current):
        raise HTTPException(403, "Admin requerido")

    # 1. Construir mapa nombre → apostador_id ANTES de borrar
    r = await db.execute(text("""
        SELECT LOWER(TRIM(nombre_apostador)) AS nombre, apostador_id
        FROM apuesta
        WHERE nombre_apostador IS NOT NULL
        GROUP BY LOWER(TRIM(nombre_apostador)), apostador_id
        ORDER BY COUNT(*) DESC
    """))
    nombre_map: dict[str, int] = {}
    for row in r.mappings():
        k = row["nombre"]
        if k not in nombre_map:
            nombre_map[k] = row["apostador_id"]

    if not nombre_map:
        raise HTTPException(400, "No hay registros en apuesta para construir el mapa nombre→apostador_id")

    # 2. IDs de partidos de grupos en este torneo
    r2 = await db.execute(text("""
        SELECT p.id FROM partido p JOIN fase f ON f.id = p.fase_id
        WHERE f.torneo_id = :tid AND p.numero_fifa IS NOT NULL AND p.numero_fifa <= 72
    """), {"tid": torneo_id})
    partido_ids = [row[0] for row in r2.fetchall()]
    if not partido_ids:
        raise HTTPException(400, "No se encontraron partidos de grupos (numero_fifa 1-72) en el torneo")

    ids_sql = ",".join(str(i) for i in partido_ids)

    # 3. Borrar puntaje_detalle de grupos
    rd = await db.execute(text(f"DELETE FROM puntaje_detalle WHERE partido_id IN ({ids_sql})"))
    det_borrados = rd.rowcount

    # 4. Borrar apuestas de grupos
    ra = await db.execute(text(f"DELETE FROM apuesta WHERE partido_id IN ({ids_sql})"))
    ap_borradas = ra.rowcount

    await db.commit()

    # 5. Insertar desde pronosticos_aux
    # Cargamos todos los datos de pronosticos_aux para los 72 partidos
    r3 = await db.execute(text("""
        SELECT pa.nombre, pa.alias, pa.goles_local, pa.goles_visitante,
               pa.amarillas, pa.rojas, pa.var, pa.penales, pa.primer_gol,
               pa.numero_partido_fifa, p.id AS partido_id
        FROM pronosticos_aux pa
        JOIN partido p ON p.numero_fifa = pa.numero_partido_fifa
        JOIN fase    f ON f.id = p.fase_id
        WHERE f.torneo_id = :tid AND pa.numero_partido_fifa <= 72
        ORDER BY pa.nombre, p.numero_fifa
    """), {"tid": torneo_id})
    paux_rows = r3.mappings().all()

    insertados = 0
    sin_match: list[str] = []

    for row in paux_rows:
        nombre_lower = (row["nombre"] or "").lower().strip()
        alias_lower  = (row["alias"]  or "").lower().strip()
        aid = nombre_map.get(nombre_lower) or nombre_map.get(alias_lower)
        if aid is None:
            if row["nombre"] not in sin_match:
                sin_match.append(row["nombre"])
            continue

        penales_val = row["penales"] or 0
        await db.execute(text("""
            INSERT INTO apuesta (
                apostador_id, partido_id,
                pred_local, pred_visitante,
                pred_amarillas, pred_rojas, pred_var,
                pred_penales_partido, pred_minuto_gol,
                pred_penales, puntos, puntos_bonus,
                nombre_apostador, numero_fifa
            ) VALUES (
                :aid, :pid,
                :pl, :pv,
                :amar, :rojas, :var,
                :pp, :min_gol,
                :pen_bool, 0, 0,
                :nombre, :numfifa
            )
        """), {
            "aid": aid,
            "pid": row["partido_id"],
            "pl":  row["goles_local"],
            "pv":  row["goles_visitante"],
            "amar": row["amarillas"],
            "rojas": row["rojas"],
            "var":  row["var"],
            "pp":   penales_val,
            "min_gol": row["primer_gol"],
            "pen_bool": penales_val > 0,
            "nombre": row["nombre"],
            "numfifa": row["numero_partido_fifa"],
        })
        insertados += 1

    await db.commit()

    return {
        "puntaje_detalle_borrados": det_borrados,
        "apuestas_borradas": ap_borradas,
        "apuestas_insertadas": insertados,
        "sin_match_apostador": sin_match,
        "esperado": len(nombre_map) * 72,
        "ok": insertados > 0 and len(sin_match) == 0,
    }


@router.post("/reload-apuestas-por-equipos/{torneo_id}", summary="Reload apuestas grupos usando team IDs como clave (fix swap numbering)")
async def reload_apuestas_por_equipos(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """
    Igual que reload-apuestas-grupos pero usa idequipolocal/idequipovisitante de pronosticos_aux
    como clave de join con partido, en vez de numero_fifa. Esto corrige los casos donde el
    ordenamiento FIFA en pronosticos_aux no coincide con el de la BD (pares swapped).
    También convierte NULLs en pred_local/pred_visitante a 0.
    """
    if not await _check_admin(current):
        raise HTTPException(403, "Admin requerido")

    # 1. Construir mapa nombre → apostador_id ANTES de borrar
    r = await db.execute(text("""
        SELECT LOWER(TRIM(nombre_apostador)) AS nombre, apostador_id
        FROM apuesta
        WHERE nombre_apostador IS NOT NULL
        GROUP BY LOWER(TRIM(nombre_apostador)), apostador_id
        ORDER BY COUNT(*) DESC
    """))
    nombre_map: dict[str, int] = {}
    for row in r.mappings():
        k = row["nombre"]
        if k not in nombre_map:
            nombre_map[k] = row["apostador_id"]

    if not nombre_map:
        raise HTTPException(400, "No hay registros en apuesta para construir el mapa nombre→apostador_id")

    # 2. IDs de partidos de grupos en este torneo
    r2 = await db.execute(text("""
        SELECT p.id FROM partido p JOIN fase f ON f.id = p.fase_id
        WHERE f.torneo_id = :tid AND p.numero_fifa IS NOT NULL AND p.numero_fifa <= 72
    """), {"tid": torneo_id})
    partido_ids = [row[0] for row in r2.fetchall()]
    if not partido_ids:
        raise HTTPException(400, "No se encontraron partidos de grupos (numero_fifa 1-72) en el torneo")

    ids_sql = ",".join(str(i) for i in partido_ids)

    # 3. Borrar puntaje_detalle de grupos
    rd = await db.execute(text(f"DELETE FROM puntaje_detalle WHERE partido_id IN ({ids_sql})"))
    det_borrados = rd.rowcount

    # 4. Borrar apuestas de grupos
    ra = await db.execute(text(f"DELETE FROM apuesta WHERE partido_id IN ({ids_sql})"))
    ap_borradas = ra.rowcount

    await db.commit()

    # 5. Insertar desde pronosticos_aux usando team IDs como clave primaria
    # Para filas donde idequipolocal IS NULL, fallback a numero_fifa join
    r3 = await db.execute(text("""
        SELECT pa.nombre, pa.alias, pa.goles_local, pa.goles_visitante,
               pa.amarillas, pa.rojas, pa.var, pa.penales, pa.primer_gol,
               pa.numero_partido_fifa,
               pa.idequipolocal, pa.idequipovisitante,
               -- Join por equipos (correcto): partido que REALMENTE tiene esos equipos
               p_eq.id AS partido_id_equipos,
               -- Join por numero_fifa (fallback): para cuando no hay team IDs
               p_num.id AS partido_id_numfifa
        FROM pronosticos_aux pa
        LEFT JOIN partido p_eq ON (
            p_eq.equipo_local_id = pa.idequipolocal
            AND p_eq.equipo_visitante_id = pa.idequipovisitante
            AND pa.idequipolocal IS NOT NULL
        )
        LEFT JOIN fase f_eq ON f_eq.id = p_eq.fase_id AND f_eq.torneo_id = :tid
        LEFT JOIN partido p_num ON p_num.numero_fifa = pa.numero_partido_fifa
        LEFT JOIN fase f_num ON f_num.id = p_num.fase_id AND f_num.torneo_id = :tid
        WHERE pa.numero_partido_fifa <= 72
          AND (f_eq.torneo_id = :tid OR f_num.torneo_id = :tid)
        ORDER BY pa.nombre, pa.numero_partido_fifa
    """), {"tid": torneo_id})
    paux_rows = r3.mappings().all()

    insertados = 0
    sin_match_apostador: list[str] = []
    sin_partido: list[int] = []
    usaron_numfifa: list[int] = []  # partidos donde se usó fallback numero_fifa
    correccion_swap: list[dict] = []

    for row in paux_rows:
        nombre_lower = (row["nombre"] or "").lower().strip()
        alias_lower  = (row["alias"]  or "").lower().strip()
        aid = nombre_map.get(nombre_lower) or nombre_map.get(alias_lower)
        if aid is None:
            if row["nombre"] not in sin_match_apostador:
                sin_match_apostador.append(row["nombre"])
            continue

        # Elegir partido_id: preferir el match por equipos
        partido_id = row["partido_id_equipos"]
        usó_equipos = True
        if partido_id is None:
            partido_id = row["partido_id_numfifa"]
            usó_equipos = False
            if row["numero_partido_fifa"] not in usaron_numfifa:
                usaron_numfifa.append(row["numero_partido_fifa"])

        if partido_id is None:
            if row["numero_partido_fifa"] not in sin_partido:
                sin_partido.append(row["numero_partido_fifa"])
            continue

        # Detectar correcciones de swap (team match != numfifa match)
        if usó_equipos and row["partido_id_numfifa"] and row["partido_id_equipos"] != row["partido_id_numfifa"]:
            entry = {"numfifa": row["numero_partido_fifa"], "partido_id_correcto": partido_id, "partido_id_wrong": row["partido_id_numfifa"]}
            if entry not in correccion_swap:
                correccion_swap.append(entry)

        penales_val = row["penales"] or 0
        # Convertir NULLs a 0
        pred_local = row["goles_local"] if row["goles_local"] is not None else 0
        pred_visit = row["goles_visitante"] if row["goles_visitante"] is not None else 0

        await db.execute(text("""
            INSERT INTO apuesta (
                apostador_id, partido_id,
                pred_local, pred_visitante,
                pred_amarillas, pred_rojas, pred_var,
                pred_penales_partido, pred_minuto_gol,
                pred_penales, puntos, puntos_bonus,
                nombre_apostador, numero_fifa
            ) VALUES (
                :aid, :pid,
                :pl, :pv,
                :amar, :rojas, :var,
                :pp, :min_gol,
                :pen_bool, 0, 0,
                :nombre, :numfifa
            )
        """), {
            "aid": aid,
            "pid": partido_id,
            "pl":  pred_local,
            "pv":  pred_visit,
            "amar": row["amarillas"],
            "rojas": row["rojas"],
            "var":  row["var"],
            "pp":   penales_val,
            "min_gol": row["primer_gol"],
            "pen_bool": penales_val > 0,
            "nombre": row["nombre"],
            "numfifa": row["numero_partido_fifa"],
        })
        insertados += 1

    await db.commit()

    return {
        "puntaje_detalle_borrados": det_borrados,
        "apuestas_borradas": ap_borradas,
        "apuestas_insertadas": insertados,
        "esperado": len(nombre_map) * 72,
        "correcciones_swap": len(correccion_swap),
        "detalle_swap": correccion_swap[:20],  # max 20 en respuesta
        "fallback_numfifa": usaron_numfifa,
        "sin_match_apostador": sin_match_apostador,
        "sin_partido": sin_partido,
        "ok": insertados > 0 and len(sin_match_apostador) == 0 and len(sin_partido) == 0,
    }


@router.post("/calcular-puntajes/{torneo_id}", summary="Calcular puntajes según reglamento oficial por competencia (admin)")
async def calcular_puntajes(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    await _ensure_detalle_table(db)

    # Auto-bloquear fases de grupos completadas antes de recalcular
    _grupos_auto_bloqueadas = await _auto_lock_completed_grupos(db, torneo_id)

    # ── Columnas puntaje_detalle (idempotente, compatibilidad v1 + v2) ───────
    for col_sql in [
        "ALTER TABLE puntaje_detalle ADD COLUMN IF NOT EXISTS teams_match BOOLEAN DEFAULT TRUE",
        "ALTER TABLE puntaje_detalle ADD COLUMN IF NOT EXISTS pred_penales BOOLEAN",
        "ALTER TABLE puntaje_detalle ADD COLUMN IF NOT EXISTS real_penales BOOLEAN",
        "ALTER TABLE puntaje_detalle ADD COLUMN IF NOT EXISTS pts_penales INT NOT NULL DEFAULT 0",
        # Columnas v2 (GRUPO 0)
        "ALTER TABLE puntaje_detalle ADD COLUMN IF NOT EXISTS pts_resultado INT DEFAULT 0",
        "ALTER TABLE puntaje_detalle ADD COLUMN IF NOT EXISTS pts_rojas INT DEFAULT 0",
        "ALTER TABLE puntaje_detalle ADD COLUMN IF NOT EXISTS pts_penales_tanda INT DEFAULT 0",
        "ALTER TABLE puntaje_detalle ADD COLUMN IF NOT EXISTS pts_equipo INT DEFAULT 0",
        "ALTER TABLE puntaje_detalle ADD COLUMN IF NOT EXISTS pts_penales_partido INT DEFAULT 0",
        "ALTER TABLE puntaje_detalle ADD COLUMN IF NOT EXISTS pts_penales_tanda INT DEFAULT 0",
        "ALTER TABLE puntaje_detalle ADD COLUMN IF NOT EXISTS pts_rojas INT DEFAULT 0",
        # Columnas globales en torneo (para scoring C/D — se crean si no existen)
        # IMPORTANTE: deben existir ANTES de que _load_torneo_resultados las consulte,
        # de lo contrario el SELECT falla y deja la transacción en estado aborted.
        "ALTER TABLE torneo ADD COLUMN IF NOT EXISTS resultado_goleador VARCHAR(100)",
        "ALTER TABLE torneo ADD COLUMN IF NOT EXISTS resultado_peor_equipo_id INT",
    ]:
        try:
            await db.execute(text(col_sql))
        except Exception:
            pass  # ADD COLUMN IF NOT EXISTS no debería fallar; no corromper la sesión

    # ── Crear tabla puntaje_item si no existe (idempotente) ──────────────────
    for ddl in [
        """
        CREATE TABLE IF NOT EXISTS puntaje_item (
            id           SERIAL PRIMARY KEY,
            torneo_id    INT         NOT NULL,
            partido_id   INT,
            apostador_id INT         NOT NULL,
            categoria    VARCHAR(10) NOT NULL,
            item         VARCHAR(2)  NOT NULL,
            fase_tipo    VARCHAR(30),
            fase_nombre  VARCHAR(80),
            fecha_partido TIMESTAMPTZ,
            local_nombre VARCHAR(100),
            visit_nombre VARCHAR(100),
            resultado    TEXT,
            apuesta      TEXT,
            puntaje      INT  NOT NULL DEFAULT 0,
            multiplicador INT NOT NULL DEFAULT 1,
            updated_at   TIMESTAMPTZ DEFAULT NOW()
        )
        """,
        """
        CREATE UNIQUE INDEX IF NOT EXISTS uq_puntaje_item_partido
            ON puntaje_item (partido_id, apostador_id, item)
            WHERE partido_id IS NOT NULL
        """,
        """
        CREATE UNIQUE INDEX IF NOT EXISTS uq_puntaje_item_global
            ON puntaje_item (torneo_id, apostador_id, item)
            WHERE partido_id IS NULL
        """,
        "CREATE INDEX IF NOT EXISTS idx_puntaje_item_torneo ON puntaje_item (torneo_id)",
        "CREATE INDEX IF NOT EXISTS idx_puntaje_item_apostador ON puntaje_item (torneo_id, apostador_id)",
    ]:
        try:
            await db.execute(text(ddl))
        except Exception:
            pass

    # ── Resolver engine según competicion.codigo del torneo ──────────────────
    r_torneo = await db.execute(
        text("""
            SELECT t.id, COALESCE(c.codigo, '') AS competicion_codigo
            FROM torneo t
            LEFT JOIN competicion c ON c.id = t.competicion_id
            WHERE t.id = :tid
        """),
        {"tid": torneo_id},
    )
    row_torneo = r_torneo.mappings().first()
    if not row_torneo:
        raise HTTPException(404, f"Torneo {torneo_id} no encontrado.")
    competicion_codigo = row_torneo["competicion_codigo"] or None
    engine = scoring_registry.get_engine(competicion_codigo)

    # ── Delegar puntajes partido a partido ───────────────────────────────────
    calc = ScoringCalculator(db)
    result = await calc.calculate(torneo_id, engine)
    if result is None:
        raise HTTPException(400, "No hay partidos finalizados. Ejecutá 'Simular resultados' primero.")

    plenos   = result["plenos"]
    aciertos = result["aciertos"]
    fallos   = result["fallos"]
    por_fase = result["por_fase"]

    # ── Puntajes globales A-G (si hay apuestas globales) ─────────────────────
    global_result = await calc.calculate_global(torneo_id, engine)
    globales_procesadas = global_result.get("procesadas", 0)

    # ── Item P: clasificados por fase (grupos + KO audit trail) ──────────────
    # Obtenemos IDs de apostadores reales (rol 'apostador') para excluir admins/test
    valid_apostador_ids: set[int] = set()
    try:
        async with _app_engine.connect() as _aconn:
            _ar = await _aconn.execute(text("""
                SELECT u.id FROM users u
                JOIN user_roles ur ON ur.user_id = u.id
                JOIN roles ro ON ro.id = ur.role_id
                WHERE ro.name = 'apostador' AND u.is_active = TRUE
            """))
            valid_apostador_ids = {row[0] for row in _ar.fetchall()}
    except Exception:
        valid_apostador_ids = set()  # Si falla, sin filtro (procesa todos)

    try:
        clas_result = await calc.calculate_clasificados(torneo_id, engine,
                                                        valid_ids=valid_apostador_ids or None)
        grupos_clas_procesados = clas_result.get("grupos_procesados", 0)
        ko_clas_fases = clas_result.get("ko_fases", [])
    except Exception:
        grupos_clas_procesados = 0
        ko_clas_fases = []

    await db.commit()
    await _audit_log("puntajes:calculo", "bets", current=current, method="POST",
                     path=f"/api/v1/bets/calcular-puntajes/{torneo_id}",
                     resource_id=str(torneo_id),
                     details={
                         "evento": "cálculo de puntajes",
                         "engine": competicion_codigo or "default",
                         "plenos": plenos, "aciertos": aciertos, "fallos": fallos,
                         "globales_procesadas": globales_procesadas,
                     })
    return {
        "ok":      True,
        "engine":  competicion_codigo or "default",
        "plenos":  plenos,
        "aciertos": aciertos,
        "fallos":  fallos,
        "por_fase": por_fase,
        "globales_procesadas": globales_procesadas,
        "grupos_auto_bloqueados": _grupos_auto_bloqueadas,
        "clasificados_grupos": grupos_clas_procesados,
        "clasificados_ko_fases": ko_clas_fases,
    }


@router.get("/clasificados/{torneo_id}", summary="Auditoría: equipos clasificados predichos vs reales por fase (admin)")
async def get_clasificados(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """
    Devuelve la auditoría de item P (equipo clasifica) por apostador × fase.
    Incluye equipos predichos, reales, aciertos y puntos obtenidos.
    """
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")
    try:
        r = await db.execute(
            text("""
                SELECT ac.apostador_id,
                       ac.fase_tipo,
                       ac.equipos_pronosticados,
                       ac.equipos_reales,
                       ac.aciertos,
                       ac.pts_por_acierto,
                       ac.pts_obtenidos,
                       ac.calculado_at
                FROM apostador_clasificados ac
                WHERE ac.torneo_id = :tid
                ORDER BY ac.fase_tipo, ac.apostador_id
            """),
            {"tid": torneo_id},
        )
        rows = [dict(row) for row in r.mappings()]
    except Exception as e:
        raise HTTPException(500, f"Error consultando clasificados: {e}")

    # Enriquecer con usernames
    ids = list({row["apostador_id"] for row in rows})
    nombre_map: dict = {}
    if ids:
        try:
            async with _app_engine.connect() as conn:
                nr = await conn.execute(
                    text("SELECT id, COALESCE(nombre, username) AS nombre, username FROM users WHERE id = ANY(:ids)"),
                    {"ids": ids},
                )
                nombre_map = {r[0]: {"nombre": r[1], "username": r[2]} for r in nr.fetchall()}
        except Exception:
            pass

    for row in rows:
        info = nombre_map.get(row["apostador_id"], {})
        row["apostador"] = info.get("username") or f"Apostador {row['apostador_id']}"

    return {"torneo_id": torneo_id, "registros": rows, "total": len(rows)}


@router.post("/sync-resultados/{torneo_id}", summary="Sincronizar resultados desde API-Football (admin)")
async def sync_resultados(
    torneo_id: int,
    current: CurrentUser,
    db: DBSession,
    force: bool = False,
    max_detalle: int = 10,
    reset_resultados: bool = False,
    resync_ayer: bool = False,
    resync_fecha: str | None = None,
) -> dict:
    """
    Cadena completa automática:
      1. (Opcional) reset_resultados=true → limpia scores/puntajes antes de sincronizar.
      2. Auto-mapeo si faltan api_fixture_id (detecta liga, matchea equipos y partidos).
      3. Descarga partidos finalizados de API-Football y actualiza la BD.
      4. Avanza el bracket KO (propaga ganadores a la siguiente fase).
      5. Recalcula todos los puntajes y globales A-G.

    Query params:
      force=true            → re-sincroniza incluso partidos ya marcados 'finalizado'.
      max_detalle=N         → máximo de peticiones individuales por run (default 10).
      reset_resultados=true → resetea scores, puntajes y bracket antes de sincronizar.
                              Útil para borrar datos de prueba al iniciar la competencia real.
      resync_ayer=true      → re-sincroniza solo los partidos de ayer (penales, tarjetas, etc.)
      resync_fecha=YYYY-MM-DD → re-sincroniza partidos de esa fecha específica.
    """
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    # ── Paso 0 (opcional): reset scores/puntajes antes de sincronizar ─────────
    reset_summary: dict = {}
    if reset_resultados:
        if not await _check_superadmin(current):
            raise HTTPException(403, "reset_resultados requiere rol superadmin")
        # a) Reset scores de todos los partidos
        await db.execute(
            text("""
                UPDATE partido
                SET goles_local=NULL, goles_visitante=NULL,
                    minuto_primer_gol=NULL, amarillas=NULL,
                    decisiones_var=NULL, penales_local=NULL, penales_visitante=NULL,
                    estado='programado'
                WHERE torneo_id=:tid
            """),
            {"tid": torneo_id},
        )
        _sopt_r = await db.execute(text("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name='partido' AND table_schema='public'
              AND column_name = ANY(ARRAY['rojas','penales_partido','equipo_clasificado_id'])
        """))
        for _col in {row[0] for row in _sopt_r}:
            await db.execute(text(f"UPDATE partido SET {_col}=NULL WHERE torneo_id=:tid"), {"tid": torneo_id})
        # b) Reset participacion (standings grupos)
        await db.execute(
            text("""UPDATE participacion SET pj=0,pg=0,pe=0,pp=0,gf=0,gc=0,pts=0,clasifica=FALSE
                     WHERE fase_id IN (SELECT id FROM fase WHERE torneo_id=:tid AND tipo='grupo')"""),
            {"tid": torneo_id},
        )
        # c) Reset puntajes calculados
        reset_pts = await _reset_puntajes_todos(db, torneo_id)
        # d) Reset bracket KO a TBD
        try:
            await _resetear_ko_a_tbd(db, torneo_id)
        except Exception as _e:
            log.warning("sync.reset_ko_skip", error=str(_e))
        await db.commit()
        reset_summary = {"reset_resultados": True, **reset_pts}

    from app.services.sync_api_football import sync_torneo
    from datetime import date as _date, timedelta as _td

    # ── Resolver fecha_filtro (resync_ayer / resync_fecha) ────────────────────
    fecha_filtro: "_date | None" = None
    if resync_ayer:
        fecha_filtro = (_date.today() - _td(days=1))
    elif resync_fecha:
        try:
            fecha_filtro = _date.fromisoformat(resync_fecha)
        except ValueError:
            raise HTTPException(400, f"resync_fecha inválido: '{resync_fecha}'. Usar formato YYYY-MM-DD.")
    if fecha_filtro is not None:
        force = True  # siempre forzar cuando se filtra por fecha

    # ── Paso 1: auto-mapeo + sincronizar resultados con API-Football ──────────
    try:
        sync_summary = await sync_torneo(
            db, torneo_id, force=force, max_detalle=max_detalle, fecha_filtro=fecha_filtro
        )
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        # Captura errores de BD (columna faltante, etc.) y los retorna como 400
        # en vez de dejar caer el servidor con 500
        await db.rollback()
        raise HTTPException(400, f"Error en sync: {e}")

    # Commit resultados antes de recalcular standings y avanzar bracket
    await db.commit()

    actualizados  = sync_summary.get("actualizados", 0)

    # ── Paso 1b: recalcular standings de grupos (PJ/PG/PE/PP/GF/GC/Pts) ────────
    # Siempre se recalcula para que el cuadro refleje el estado actual del torneo.
    try:
        await _recalc_participacion(db, torneo_id)
        await db.commit()
    except Exception as e:
        sync_summary["participacion_error"] = str(e)

    # Siempre avanzar bracket y calcular puntajes en cada sync manual.
    # El costo es bajo y garantiza que el estado sea consistente.

    # ── Paso 2: avanzar bracket KO ───────────────────────────────────────────
    bracket_ok = False
    if True:
        try:
            maps = await ko_scoring.build_num_maps(db, torneo_id)
            await _avanzar_bracket(db, torneo_id, maps)
            await db.commit()
            bracket_ok = True
        except Exception as e:
            sync_summary["bracket_error"] = str(e)

    # ── Paso 3: recalcular puntajes ──────────────────────────────────────────
    puntajes_ok = False
    puntajes_summary: dict = {}
    if True:
        try:
            await _ensure_detalle_table(db)
            r_torneo = await db.execute(
                text("""
                    SELECT COALESCE(c.codigo, '') AS competicion_codigo
                    FROM torneo t
                    LEFT JOIN competicion c ON c.id = t.competicion_id
                    WHERE t.id = :tid
                """),
                {"tid": torneo_id},
            )
            row_torneo = r_torneo.mappings().first()
            competicion_codigo = (row_torneo or {}).get("competicion_codigo") or None
            engine = scoring_registry.get_engine(competicion_codigo)

            calc = ScoringCalculator(db)
            result = await calc.calculate(torneo_id, engine)
            if result:
                global_result = await calc.calculate_global(torneo_id, engine)
                await db.commit()
                puntajes_ok = True
                puntajes_summary = {
                    "plenos":              result["plenos"],
                    "aciertos":            result["aciertos"],
                    "fallos":              result["fallos"],
                    "globales_procesadas": global_result.get("procesadas", 0),
                }
        except Exception as e:
            sync_summary["puntajes_error"] = str(e)

    await _audit_log(
        "sync:resultados", "bets",
        current=current, method="POST",
        path=f"/api/v1/bets/sync-resultados/{torneo_id}",
        resource_id=str(torneo_id),
        details={
            "evento":      "sincronización API-Football",
            "force":       force,
            "reset_resultados": reset_resultados,
            "fecha_filtro": str(fecha_filtro) if fecha_filtro else None,
            **reset_summary,
            "actualizados": actualizados,
            "api_calls":   sync_summary.get("api_calls", 0),
            "bracket_ok":  bracket_ok,
            "puntajes_ok": puntajes_ok,
        },
    )

    return {
        "ok":           True,
        "actualizados": actualizados,
        "sync":         sync_summary,
        "bracket_ok":   bracket_ok,
        "puntajes_ok":  puntajes_ok,
        "puntajes":     puntajes_summary,
        "fecha_filtro": str(fecha_filtro) if fecha_filtro else None,
    }


@router.get("/api-mapeo/{torneo_id}", summary="Datos para mapeo API-Football ↔ BD (admin)")
async def api_mapeo_get(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """
    Carga desde API-Football la lista de equipos y fixtures del torneo,
    y los devuelve junto a los registros DB para que el admin realice el mapeo.

    Retorna:
      api_teams:    [{id, name, logo}] — equipos desde API-Football
      api_fixtures: [{id, date, home_id, home_name, away_id, away_name, round, status}]
      db_equipos:   [{id, nombre, api_team_id}]
      db_partidos:  [{id, fecha, equipo_local_id, local_nombre, equipo_visitante_id,
                       visitante_nombre, fase_tipo, api_fixture_id}]
      config:       {api_league_id, api_season} — configuración actual del torneo
    """
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    from app.services.sync_api_football import _headers, API_BASE
    from app.core.config import settings as _settings

    if not _settings.APIFOOTBALL_KEY:
        raise HTTPException(400, "APIFOOTBALL_KEY no configurado en .env")

    # ── Cargar config del torneo ──────────────────────────────────────────────
    r_cfg = await db.execute(
        text("""
            SELECT t.id, t.api_season, c.api_league_id, c.id AS competicion_id
            FROM torneo t
            LEFT JOIN competicion c ON c.id = t.competicion_id
            WHERE t.id = :tid
        """),
        {"tid": torneo_id},
    )
    cfg = r_cfg.mappings().first()
    if not cfg:
        raise HTTPException(404, "Torneo no encontrado")

    api_season    = cfg["api_season"]
    api_league_id = cfg["api_league_id"]

    # ── Cargar equipos DB ─────────────────────────────────────────────────────
    r_eq = await db.execute(
        text("SELECT id, COALESCE(nombre_es, nombre) AS nombre, api_team_id FROM equipo ORDER BY nombre")
    )
    db_equipos = [dict(row) for row in r_eq.mappings()]

    # ── Cargar partidos DB ────────────────────────────────────────────────────
    r_p = await db.execute(
        text("""
            SELECT p.id, p.fecha, p.api_fixture_id, p.equipo_local_id, p.equipo_visitante_id,
                   f.tipo AS fase_tipo, f.nombre AS fase_nombre,
                   COALESCE(el.nombre_es, el.nombre) AS local_nombre,
                   COALESCE(ev.nombre_es, ev.nombre) AS visitante_nombre
            FROM partido p
            JOIN fase f ON f.id = p.fase_id
            LEFT JOIN equipo el ON el.id = p.equipo_local_id
            LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
            WHERE p.torneo_id = :tid
            ORDER BY p.fecha, p.id
        """),
        {"tid": torneo_id},
    )
    db_partidos = []
    for row in r_p.mappings():
        d = dict(row)
        if d.get("fecha"):
            d["fecha"] = d["fecha"].isoformat() if hasattr(d["fecha"], "isoformat") else str(d["fecha"])
        db_partidos.append(d)

    # ── Si no hay api_league_id o api_season, devolver solo BD ───────────────
    if not api_season or not api_league_id:
        return {
            "api_teams":    [],
            "api_fixtures": [],
            "db_equipos":   db_equipos,
            "db_partidos":  db_partidos,
            "config":       {"api_league_id": api_league_id, "api_season": api_season},
            "warning":      "Faltan api_league_id o api_season. Configurá primero y volvé a cargar.",
        }

    # ── Fetch API-Football ────────────────────────────────────────────────────
    api_teams:    list[dict] = []
    api_fixtures: list[dict] = []
    api_error: str | None = None

    import httpx
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            # Equipos del torneo
            r1 = await client.get(
                f"{API_BASE}/teams",
                params={"league": api_league_id, "season": api_season},
                headers=_headers(),
            )
            r1.raise_for_status()
            for t in r1.json().get("response", []):
                api_teams.append({
                    "id":   t["team"]["id"],
                    "name": t["team"]["name"],
                    "logo": t["team"].get("logo", ""),
                })
            api_teams.sort(key=lambda x: x["name"])

            # Fixtures del torneo
            r2 = await client.get(
                f"{API_BASE}/fixtures",
                params={"league": api_league_id, "season": api_season},
                headers=_headers(),
            )
            r2.raise_for_status()
            for fix in r2.json().get("response", []):
                api_fixtures.append({
                    "id":         fix["fixture"]["id"],
                    "date":       fix["fixture"]["date"],
                    "round":      fix["league"].get("round", ""),
                    "status":     fix["fixture"]["status"]["short"],
                    "home_id":    fix["teams"]["home"]["id"],
                    "home_name":  fix["teams"]["home"]["name"],
                    "away_id":    fix["teams"]["away"]["id"],
                    "away_name":  fix["teams"]["away"]["name"],
                })
            api_fixtures.sort(key=lambda x: x["date"])

    except Exception as e:
        api_error = str(e)

    return {
        "api_teams":    api_teams,
        "api_fixtures": api_fixtures,
        "db_equipos":   db_equipos,
        "db_partidos":  db_partidos,
        "config":       {"api_league_id": api_league_id, "api_season": api_season},
        **({"api_error": api_error} if api_error else {}),
    }


class ApiMapeoSave(BaseModel):
    api_league_id: int | None = None
    api_season:    int | None = None
    equipos:  list[dict] = []   # [{id: int, api_team_id: int}]
    partidos: list[dict] = []   # [{id: int, api_fixture_id: int}]


@router.post("/api-mapeo/{torneo_id}/save", summary="Guardar mapeo API-Football → BD (admin)")
async def api_mapeo_save(
    torneo_id: int, body: ApiMapeoSave, current: CurrentUser, db: DBSession
) -> dict:
    """Guarda en BD los IDs de API-Football para liga, season, equipos y partidos."""
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    actualizados_equipos  = 0
    actualizados_partidos = 0

    # ── Liga + Season ─────────────────────────────────────────────────────────
    if body.api_league_id is not None:
        r_comp = await db.execute(
            text("SELECT competicion_id FROM torneo WHERE id = :tid"),
            {"tid": torneo_id},
        )
        row_comp = r_comp.mappings().first()
        if row_comp and row_comp["competicion_id"]:
            await db.execute(
                text("UPDATE competicion SET api_league_id = :lid WHERE id = :cid"),
                {"lid": body.api_league_id, "cid": row_comp["competicion_id"]},
            )

    if body.api_season is not None:
        await db.execute(
            text("UPDATE torneo SET api_season = :s WHERE id = :tid"),
            {"s": body.api_season, "tid": torneo_id},
        )

    # ── Equipos ───────────────────────────────────────────────────────────────
    for eq in body.equipos:
        db_id       = eq.get("id")
        api_team_id = eq.get("api_team_id")
        if db_id and api_team_id is not None:
            await db.execute(
                text("UPDATE equipo SET api_team_id = :atid WHERE id = :eid"),
                {"atid": api_team_id, "eid": db_id},
            )
            actualizados_equipos += 1

    # ── Partidos ──────────────────────────────────────────────────────────────
    for p in body.partidos:
        db_id          = p.get("id")
        api_fixture_id = p.get("api_fixture_id")
        if db_id and api_fixture_id is not None:
            await db.execute(
                text("UPDATE partido SET api_fixture_id = :afid WHERE id = :pid"),
                {"afid": api_fixture_id, "pid": db_id},
            )
            actualizados_partidos += 1

    await db.commit()

    await _audit_log(
        "mapeo:api_football", "bets",
        current=current, method="POST",
        path=f"/api/v1/bets/api-mapeo/{torneo_id}/save",
        resource_id=str(torneo_id),
        details={
            "evento":               "mapeo API-Football guardado",
            "actualizados_equipos":  actualizados_equipos,
            "actualizados_partidos": actualizados_partidos,
        },
    )

    return {
        "ok":                    True,
        "actualizados_equipos":  actualizados_equipos,
        "actualizados_partidos": actualizados_partidos,
    }


@router.post("/api-mapeo/{torneo_id}/auto", summary="Auto-mapear equipos y partidos desde API-Football (admin)")
async def api_mapeo_auto(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """
    Detecta automáticamente y guarda en BD los api_team_id y api_fixture_id
    haciendo match por nombre normalizado y par de equipos.
    También auto-detecta api_league_id y api_season si no están configurados.
    """
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    from app.services.sync_api_football import auto_mapeo_torneo
    from app.core.config import settings as _settings
    import httpx

    if not _settings.APIFOOTBALL_KEY:
        raise HTTPException(400, "APIFOOTBALL_KEY no configurado en .env")

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            result = await auto_mapeo_torneo(db, torneo_id, client)
    except Exception as e:
        raise HTTPException(500, f"Error en auto-mapeo: {e}")

    return result


@router.get("/test-verificacion/{torneo_id}", summary="Datos crudos para verificar el scoring (admin, solo pruebas)")
async def test_verificacion(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """Devuelve datos crudos (sin cálculo) para recalcular el puntaje de forma independiente."""
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    r_part = await db.execute(
        text("""
            SELECT p.id, p.goles_local, p.goles_visitante,
                   p.minuto_primer_gol, p.amarillas, p.decisiones_var
            FROM partido p
            JOIN fase f ON f.id = p.fase_id AND f.tipo = 'grupo' AND f.nombre NOT ILIKE '%mejores%'
            WHERE p.torneo_id = :tid AND p.estado = 'finalizado'
              AND p.goles_local IS NOT NULL AND p.goles_visitante IS NOT NULL
        """),
        {"tid": torneo_id},
    )
    partidos = [dict(row) for row in r_part.mappings()]
    pids = [p["id"] for p in partidos]

    apuestas = []
    if pids:
        r_ap = await db.execute(
            text("""
                SELECT id, apostador_id, partido_id,
                       pred_local, pred_visitante,
                       pred_minuto_gol, pred_amarillas, pred_var,
                       puntos, puntos_bonus
                FROM apuesta WHERE partido_id = ANY(:pids)
            """),
            {"pids": pids},
        )
        apuestas = [dict(row) for row in r_ap.mappings()]

    uids = sorted({a["apostador_id"] for a in apuestas})
    usuarios: dict[int, str] = {}
    if uids:
        async with _app_engine.connect() as conn:
            ur = await conn.execute(
                text("SELECT id, username FROM users WHERE id = ANY(:ids)"),
                {"ids": uids},
            )
            usuarios = {row["id"]: row["username"] for row in ur.mappings()}

    # Tabla de posiciones (standings) derivada de los resultados simulados
    standings: dict = {}
    try:
        st = await _calc_standings_reales(db, torneo_id)
        for letra, grupo in st.items():
            standings[letra] = [
                {
                    "pos": e.get("pos"), "equipo_id": e["equipo_id"], "nombre": e["nombre"],
                    "pj": e["pj"], "pg": e["pg"], "pe": e["pe"], "pp": e["pp"],
                    "gf": e["gf"], "gc": e["gc"], "gd": e["gd"], "pts": e["pts"],
                }
                for e in grupo["equipos"]
            ]
    except Exception as e:
        await db.rollback()
        standings = {"error": str(e)}

    return {
        "partidos":  partidos,
        "apuestas":  apuestas,
        "usuarios":  usuarios,
        "standings": standings,
    }


# ── Verificación de integridad de equipos por competición ────────────────────

@router.get("/verificar-equipos/{torneo_id}", summary="Verifica integridad de equipos del torneo (admin)")
async def verificar_equipos(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """
    Algoritmo de verificación de equipos para un torneo:

    1. COMPETICION_MISMATCH: partidos con equipos de otra competición
    2. EQUIPOS_SIN_COMPETICION: equipos que participan pero no tienen competicion_id asignado
    3. COUNT_CHECK: cantidad de equipos distintos vs num_equipos_esperado de la competición
    4. CROSS_CONTAMINACION: equipos de competición X apareciendo en torneo de competición Y
    5. EQUIPOS_SIN_ISO: selecciones sin codigo_iso (solo aplica si la competición es de selecciones)

    Retorna: { "ok": bool, "errores": [...], "advertencias": [...], "resumen": {...} }
    """
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    errores     = []
    advertencias = []

    # ── Info del torneo y competición ─────────────────────────────────────────
    r_torneo = await db.execute(
        text("""
            SELECT t.id, t.nombre, c.id AS comp_id, c.nombre AS comp_nombre,
                   c.codigo, c.num_equipos_esperado
            FROM torneo t
            JOIN competicion c ON c.id = t.competicion_id
            WHERE t.id = :tid
        """),
        {"tid": torneo_id},
    )
    torneo = r_torneo.mappings().first()
    if not torneo:
        raise HTTPException(404, f"Torneo {torneo_id} no encontrado")

    comp_id         = torneo["comp_id"]
    comp_nombre     = torneo["comp_nombre"]
    num_esperado    = torneo["num_equipos_esperado"]

    # ── 1. Equipos distintos que participan en este torneo ───────────────────
    r_equipos = await db.execute(
        text("""
            SELECT DISTINCT e.id, e.nombre, e.nombre_es, e.competicion_id, e.codigo_iso
            FROM equipo e
            JOIN partido p ON (p.equipo_local_id = e.id OR p.equipo_visitante_id = e.id)
            JOIN fase f ON f.id = p.fase_id
            WHERE f.torneo_id = :tid
              AND e.id IS NOT NULL
            ORDER BY e.nombre
        """),
        {"tid": torneo_id},
    )
    equipos_en_torneo = list(r_equipos.mappings())
    total_equipos = len(equipos_en_torneo)

    # ── 2. CROSS_CONTAMINACION: equipos de otra competición ──────────────────
    cross = [
        e for e in equipos_en_torneo
        if e["competicion_id"] is not None and e["competicion_id"] != comp_id
    ]
    for e in cross:
        errores.append({
            "tipo":     "CROSS_CONTAMINACION",
            "equipo_id": e["id"],
            "equipo":   e["nombre"],
            "msg":      f"El equipo '{e['nombre']}' (id={e['id']}) pertenece a "
                        f"competicion_id={e['competicion_id']} pero está en torneo "
                        f"de competicion_id={comp_id} ({comp_nombre})",
        })

    # ── 3. EQUIPOS_SIN_COMPETICION: participan pero sin competicion_id ────────
    sin_comp = [e for e in equipos_en_torneo if e["competicion_id"] is None]
    for e in sin_comp:
        advertencias.append({
            "tipo":      "SIN_COMPETICION",
            "equipo_id": e["id"],
            "equipo":    e["nombre"],
            "msg":       f"El equipo '{e['nombre']}' (id={e['id']}) no tiene competicion_id asignado",
        })

    # ── 4. COUNT_CHECK: cantidad vs esperado ─────────────────────────────────
    estado_count = "sin_validacion"
    if num_esperado is not None:
        if total_equipos < num_esperado:
            estado_count = "FALTAN_EQUIPOS"
            errores.append({
                "tipo":     "COUNT_FALTAN",
                "msg":      f"Se esperan {num_esperado} equipos para {comp_nombre}, "
                            f"solo hay {total_equipos} en los partidos del torneo.",
                "esperado": num_esperado,
                "actual":   total_equipos,
                "faltan":   num_esperado - total_equipos,
            })
        elif total_equipos > num_esperado:
            estado_count = "EQUIPOS_DE_MAS"
            errores.append({
                "tipo":     "COUNT_DE_MAS",
                "msg":      f"Se esperan {num_esperado} equipos para {comp_nombre}, "
                            f"hay {total_equipos}. Posible duplicado o equipo incorrecto.",
                "esperado": num_esperado,
                "actual":   total_equipos,
            })
        else:
            estado_count = "OK"

    # ── 5. EQUIPOS_SIN_ISO (solo selecciones nacionales — competicion con código) ─
    es_selecciones = torneo["codigo"] and "copa_mundo" in (torneo["codigo"] or "")
    sin_iso = [
        e for e in equipos_en_torneo
        if e["codigo_iso"] is None and e["competicion_id"] == comp_id
    ]
    if es_selecciones and sin_iso:
        for e in sin_iso:
            advertencias.append({
                "tipo":      "SIN_ISO",
                "equipo_id": e["id"],
                "equipo":    e["nombre"],
                "msg":       f"Selección '{e['nombre']}' sin codigo_iso (necesario para banderas y reportes)",
            })

    # ── 6. Partidos con equipo_local_id o equipo_visitante_id NULL ────────────
    r_null = await db.execute(
        text("""
            SELECT COUNT(*) AS total
            FROM partido p
            JOIN fase f ON f.id = p.fase_id
            WHERE f.torneo_id = :tid
              AND (p.equipo_local_id IS NULL OR p.equipo_visitante_id IS NULL)
        """),
        {"tid": torneo_id},
    )
    partidos_sin_equipo = r_null.scalar() or 0
    if partidos_sin_equipo:
        errores.append({
            "tipo":  "PARTIDOS_SIN_EQUIPO",
            "msg":   f"{partidos_sin_equipo} partido(s) sin equipo asignado (equipo_local_id o visitante NULL)",
            "count": partidos_sin_equipo,
        })

    # ── Resumen ───────────────────────────────────────────────────────────────
    ok = len(errores) == 0
    resumen = {
        "torneo_id":         torneo_id,
        "torneo":            torneo["nombre"],
        "competicion":       comp_nombre,
        "competicion_id":    comp_id,
        "competicion_codigo": torneo["codigo"],
        "total_equipos_en_partidos": total_equipos,
        "num_equipos_esperado":      num_esperado,
        "estado_count":      estado_count,
        "cross_contaminacion": len(cross),
        "sin_competicion_id":  len(sin_comp),
        "sin_codigo_iso":      len(sin_iso) if es_selecciones else "N/A",
        "partidos_sin_equipo": partidos_sin_equipo,
        "total_errores":       len(errores),
        "total_advertencias":  len(advertencias),
    }

    return {
        "ok":          ok,
        "errores":     errores,
        "advertencias": advertencias,
        "resumen":     resumen,
        "equipos":     [
            {"id": e["id"], "nombre": e["nombre"], "competicion_id": e["competicion_id"],
             "codigo_iso": e["codigo_iso"]}
            for e in equipos_en_torneo
        ],
    }


# ── Bracket real + simulación secuencial ─────────────────────────────────────

async def _sim_resultado_partido(db, pid: int, es_grupo: bool):
    """Genera un resultado real aleatorio para un partido (con bonus y penales si aplica)."""
    gl = random.choices([0, 1, 2, 3], weights=[18, 38, 30, 14])[0]
    gv = random.choices([0, 1, 2, 3], weights=[18, 38, 30, 14])[0]
    minuto = random.randint(1, 90) if (gl + gv) > 0 else None
    amarillas = random.choices([0, 1, 2, 3, 4, 5, 6], weights=[5, 12, 22, 26, 18, 11, 6])[0]
    var = random.choices([0, 1], weights=[65, 35])[0]
    pen_l = pen_v = None
    if not es_grupo and gl == gv:
        pen_l, pen_v = random.choice([(4, 2), (5, 4), (3, 1), (5, 3), (4, 5), (2, 4), (1, 3)])
    await db.execute(
        text("""
            UPDATE partido
            SET goles_local=:gl, goles_visitante=:gv,
                minuto_primer_gol=:m, amarillas=:a, rojas=:rojas, decisiones_var=:v,
                penales_local=:pl, penales_visitante=:pv, estado='finalizado'
            WHERE id=:pid
        """),
        {"gl": gl, "gv": gv, "m": minuto, "a": amarillas, "rojas": random.choices([0,1,2], weights=[78,18,4])[0], "v": var,
         "pl": pen_l, "pv": pen_v, "pid": pid},
    )


async def _generar_pred_apostador(db, uid: int, pid: int):
    """Inserta/actualiza una predicción aleatoria (marcador + bonus) de un apostador."""
    pl = random.choices([0, 1, 2, 3], weights=[18, 38, 30, 14])[0]
    pv = random.choices([0, 1, 2, 3], weights=[18, 38, 30, 14])[0]
    pmg = random.randint(1, 90)
    pam = random.choices([0, 1, 2, 3, 4, 5, 6], weights=[5, 12, 22, 26, 18, 11, 6])[0]
    pvar = random.choices([0, 1], weights=[60, 40])[0]
    projas = random.choices([0, 1, 2], weights=[78, 18, 4])[0]
    ppp = random.choices([0, 1, 2], weights=[70, 22, 8])[0]
    await db.execute(
        text("""
            INSERT INTO apuesta
                (apostador_id, partido_id, pred_local, pred_visitante,
                 pred_minuto_gol, pred_amarillas, pred_var,
                 pred_rojas, pred_penales_partido)
            VALUES (:uid, :pid, :pl, :pv, :pmg, :pam, :pvar, :projas, :ppp)
            ON CONFLICT (apostador_id, partido_id) DO UPDATE SET
                pred_local=EXCLUDED.pred_local, pred_visitante=EXCLUDED.pred_visitante,
                pred_minuto_gol=EXCLUDED.pred_minuto_gol,
                pred_amarillas=EXCLUDED.pred_amarillas, pred_var=EXCLUDED.pred_var,
                pred_rojas=EXCLUDED.pred_rojas,
                pred_penales_partido=EXCLUDED.pred_penales_partido,
                updated_at=NOW()
        """),
        {"uid": uid, "pid": pid, "pl": pl, "pv": pv, "pmg": pmg, "pam": pam, "pvar": pvar,
         "projas": projas, "ppp": ppp},
    )


async def _grupos_completos(db, torneo_id: int) -> bool:
    """True si TODOS los partidos de fase de grupos están finalizados (al menos 1 existe)."""
    r = await db.execute(
        text("""
            SELECT
                COUNT(*) FILTER (WHERE p.estado != 'finalizado') AS pendientes,
                COUNT(*) AS total
            FROM partido p
            JOIN fase f ON f.id = p.fase_id
            WHERE f.torneo_id = :tid AND f.tipo = 'grupo'
        """),
        {"tid": torneo_id},
    )
    row = r.mappings().first()
    if not row or not row["total"]:
        return False
    return row["pendientes"] == 0


async def _resetear_ko_a_tbd(db, torneo_id: int) -> None:
    """
    Resetea TODOS los partidos KO a TBD (equipo placeholder) y limpia resultados.
    Se llama cuando la fase de grupos no está completa para asegurar que el
    bracket no muestre equipos de corridas anteriores (ej: test_integral).
    """
    tbd = await ko_scoring._tbd_id(db)
    if tbd is None:
        return  # Sin equipo TBD en BD, no hacer nada

    # Todos los partidos KO de este torneo
    r = await db.execute(
        text("""
            SELECT p.id, p.equipo_local_id, p.equipo_visitante_id
            FROM partido p
            JOIN fase f ON f.id = p.fase_id
            WHERE f.torneo_id = :tid AND f.tipo != 'grupo'
        """),
        {"tid": torneo_id},
    )
    partidos_ko = list(r.mappings())

    for p in partidos_ko:
        # Solo resetear si tiene equipos reales (no TBD ya)
        if p["equipo_local_id"] != tbd or p["equipo_visitante_id"] != tbd:
            await db.execute(
                text("""
                    UPDATE partido
                    SET equipo_local_id    = :tbd,
                        equipo_visitante_id = :tbd,
                        estado              = 'programado',
                        goles_local         = NULL,
                        goles_visitante     = NULL,
                        goles_local_prorroga  = NULL,
                        goles_visitante_prorroga = NULL,
                        penales_local       = NULL,
                        penales_visitante   = NULL,
                        minuto_primer_gol   = NULL,
                        amarillas           = NULL,
                        decisiones_var      = NULL,
                        equipo_clasificado_id = NULL
                    WHERE id = :pid
                """),
                {"tbd": tbd, "pid": p["id"]},
            )
            # Limpiar puntajes de apuestas de este partido
            await db.execute(
                text("UPDATE apuesta SET puntos = 0, puntos_bonus = 0 WHERE partido_id = :pid"),
                {"pid": p["id"]},
            )


async def _resetear_ko_post_r32(db, torneo_id: int) -> None:
    """Resetea SOLO las fases KO POSTERIORES a ronda32 (ronda16 en adelante) a TBD.
    Útil cuando se arma la R32 provisional pero el resto aún no está definido."""
    tbd = await ko_scoring._tbd_id(db)
    if tbd is None:
        return

    r = await db.execute(
        text("""
            SELECT p.id
            FROM partido p
            JOIN fase f ON f.id = p.fase_id
            WHERE f.torneo_id = :tid
              AND f.tipo IN ('ronda16', 'cuartos', 'semis', 'tercer_puesto', 'final')
        """),
        {"tid": torneo_id},
    )
    for row in r.mappings():
        await db.execute(
            text("""
                UPDATE partido
                SET equipo_local_id        = :tbd,
                    equipo_visitante_id     = :tbd,
                    estado                  = 'programado',
                    goles_local             = NULL,
                    goles_visitante         = NULL,
                    goles_local_prorroga    = NULL,
                    goles_visitante_prorroga= NULL,
                    penales_local           = NULL,
                    penales_visitante       = NULL,
                    minuto_primer_gol       = NULL,
                    amarillas               = NULL,
                    decisiones_var          = NULL,
                    equipo_clasificado_id   = NULL
                WHERE id = :pid
            """),
            {"tbd": tbd, "pid": row["id"]},
        )
        await db.execute(
            text("UPDATE apuesta SET puntos = 0, puntos_bonus = 0 WHERE partido_id = :pid"),
            {"pid": row["id"]},
        )


async def _avanzar_bracket(db, torneo_id: int, maps: dict, hasta_tipo: str | None = None):
    """Avanza el bracket real asignando equipos a las fases KO según resultados.

    - Si grupos NO están completos: actualiza R32 con los mejores terceros
      provisionales según standings actuales; resetea R16+ a TBD.
    - Si grupos SÍ están completos: arma R32 definitivo y propaga todo el KO.
    - Si hasta_tipo se indica, sólo avanza hasta esa fase (inclusive).
    """
    orden_tipos = ["ronda32", "ronda16", "cuartos", "semis", "tercer_puesto", "final"]

    grupos_ok = await _grupos_completos(db, torneo_id)
    standings  = await _calc_standings_reales(db, torneo_id)

    if standings:
        # fill_incomplete=False: solo usa grupos con TODOS sus partidos finalizados.
        # Terceros de grupos incompletos quedan TBD hasta que terminen.
        # El sync automático recalcula el bracket al finalizar cada partido.
        mejores, _ = seleccionar_mejores_terceros(standings, fill_incomplete=False)
        await ko_scoring.avanzar_ronda32(db, torneo_id, maps["num2pid"], standings, mejores)

    if not grupos_ok:
        # Grupos incompletos → R32 con terceros de grupos completos; R16+ a TBD
        await _resetear_ko_post_r32(db, torneo_id)
        return

    if hasta_tipo == "ronda32":
        return
    for tipo in orden_tipos[1:]:
        await ko_scoring.avanzar_fase_ko(db, torneo_id, tipo, maps)
        if hasta_tipo == tipo:
            break


@router.post("/avanzar-bracket/{torneo_id}", summary="Avanzar el bracket real desde resultados (admin)")
async def avanzar_bracket(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")
    maps = await ko_scoring.build_num_maps(db, torneo_id)
    await _avanzar_bracket(db, torneo_id, maps)
    await db.commit()
    await _audit_log("avance:bracket", "bets", current=current, method="POST",
                     path=f"/api/v1/bets/avanzar-bracket/{torneo_id}",
                     resource_id=str(torneo_id),
                     details={"evento": "avance manual del bracket desde resultados"})
    return {"ok": True, "mensaje": "Bracket real actualizado desde resultados"}


@router.post("/avanzar-bracket-provisional/{torneo_id}",
             summary="Arma R32 provisional desde standings actuales + limpia ronda16+ a TBD (admin)")
async def avanzar_bracket_provisional(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """Avanza el bracket provisionalmente aunque haya partidos de grupos pendientes.
    - Calcula standings actuales (parciales si es necesario)
    - Selecciona mejores terceros provisionales
    - Llena ronda32 con equipos proyectados
    - Resetea ronda16 en adelante a 'a definir' (TBD)
    """
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    maps = await ko_scoring.build_num_maps(db, torneo_id)
    standings = await _calc_standings_reales(db, torneo_id)
    r32_fills = 0
    if standings:
        mejores, _ = seleccionar_mejores_terceros(standings, fill_incomplete=False)
        r32_fills = await ko_scoring.avanzar_ronda32(db, torneo_id, maps["num2pid"], standings, mejores)
    # Reset ronda16+ to TBD regardless
    await _resetear_ko_post_r32(db, torneo_id)
    await db.commit()
    await _audit_log("avance:bracket:provisional", "bets", current=current, method="POST",
                     path=f"/api/v1/bets/avanzar-bracket-provisional/{torneo_id}",
                     resource_id=str(torneo_id),
                     details={"evento": "avance provisional R32 con standings actuales",
                              "r32_fills": r32_fills})
    grupos_completos = await _grupos_completos(db, torneo_id)
    return {
        "ok": True,
        "grupos_completos": grupos_completos,
        "r32_fills": r32_fills,
        "mensaje": f"R32 proyectado ({r32_fills} partidos actualizados). Ronda16+ limpiados a 'a definir'."
                   + ("" if grupos_completos else " ⚠️ Grupos incompletos — proyección provisional.")
    }


@router.post("/simular-secuencial/{torneo_id}/{fase_id}",
             summary="Simular secuencialmente hasta la fase objetivo: avanza bracket, genera pronósticos y resultados, y puntúa (admin)")
async def simular_secuencial(torneo_id: int, fase_id: int, current: CurrentUser, db: DBSession) -> dict:
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    # Fases del torneo ordenadas
    rf = await db.execute(
        text("""SELECT id, nombre, tipo, orden FROM fase
                WHERE torneo_id=:tid ORDER BY orden, nombre"""),
        {"tid": torneo_id},
    )
    fases = [dict(r) for r in rf.mappings()]
    objetivo = next((f for f in fases if f["id"] == fase_id), None)
    if not objetivo:
        raise HTTPException(404, "Fase objetivo no encontrada")
    target_orden = objetivo["orden"]

    # Apostadores del torneo (los que ya tienen apuestas de grupo)
    ra = await db.execute(
        text("""SELECT DISTINCT a.apostador_id
                FROM apuesta a JOIN partido p ON p.id=a.partido_id
                JOIN fase f ON f.id=p.fase_id
                WHERE f.torneo_id=:tid AND f.tipo='grupo'"""),
        {"tid": torneo_id},
    )
    apostadores = [row[0] for row in ra]
    if not apostadores:
        raise HTTPException(400, "No hay apostadores con pronósticos de grupo")

    maps = await ko_scoring.build_num_maps(db, torneo_id)
    pasos: list[dict] = []

    for f in fases:
        if f["orden"] > target_orden:
            break
        tipo = f["tipo"]
        es_grupo = tipo == "grupo"

        # 1) Avanzar bracket (asignar equipos) para fases KO antes de pronosticar
        if not es_grupo:
            await _avanzar_bracket(db, torneo_id, maps, hasta_tipo=tipo)

        # partidos de la fase
        rp = await db.execute(text("SELECT id FROM partido WHERE fase_id=:fid ORDER BY id"),
                              {"fid": f["id"]})
        pids = [row[0] for row in rp]

        # 2) Generar pronósticos KO de cada apostador (los de grupo ya existen)
        preds_creadas = 0
        if not es_grupo:
            for pid in pids:
                for uid in apostadores:
                    await _generar_pred_apostador(db, uid, pid)
                    preds_creadas += 1

        # 3) Simular resultados reales de la fase
        for pid in pids:
            await _sim_resultado_partido(db, pid, es_grupo)

        await db.commit()  # confirmar para que la próxima fase vea resultados
        pasos.append({"fase": f["nombre"], "tipo": tipo,
                      "partidos": len(pids), "pred_generadas": preds_creadas})

    # 4) Avanzar bracket completo final y puntuar todo
    await _avanzar_bracket(db, torneo_id, maps, hasta_tipo=objetivo["tipo"])
    await _recalc_participacion(db, torneo_id)
    await db.commit()
    resumen = await calcular_puntajes(torneo_id, current, db)

    await _audit_log("simulacion:secuencial", "bets", current=current, method="POST",
                     path=f"/api/v1/bets/simular-secuencial/{torneo_id}/{fase_id}",
                     resource_id=str(torneo_id),
                     details={"evento": "simulación secuencial (paso de fase a fase)",
                              "objetivo": objetivo["nombre"],
                              "pasos": pasos, "apostadores": len(apostadores)})
    return {"ok": True, "objetivo": objetivo["nombre"], "pasos": pasos, "puntajes": resumen}


# ── Puntos por fase (análisis) ───────────────────────────────────────────────

async def _puntos_por_fase_data(db, torneo_id: int) -> dict:
    """Lee puntaje_detalle y arma: por_apostador_fase y totales por fase."""
    try:
        r = await db.execute(
            text("""
                SELECT d.apostador_id, d.fase_id, d.fase_tipo, d.fase_nombre,
                       SUM(d.pts_marcador)::int AS marcador,
                       SUM(d.pts_minuto)::int   AS minuto,
                       SUM(d.pts_amarillas)::int AS amarillas,
                       SUM(d.pts_var)::int      AS var,
                       SUM(d.pts_bonus)::int    AS bonus,
                       SUM(d.pts_total)::int    AS total,
                       COUNT(*)::int            AS partidos
                FROM puntaje_detalle d
                WHERE d.torneo_id=:tid
                GROUP BY d.apostador_id, d.fase_id, d.fase_tipo, d.fase_nombre
                ORDER BY d.apostador_id
            """),
            {"tid": torneo_id},
        )
    except Exception:
        await db.rollback()
        return {"filas": [], "usuarios": {}}
    filas = [dict(row) for row in r.mappings()]
    uids = sorted({f["apostador_id"] for f in filas})
    usuarios: dict[int, str] = {}
    if uids:
        async with _app_engine.connect() as conn:
            ur = await conn.execute(
                text("SELECT id, username FROM users WHERE id = ANY(:ids)"), {"ids": uids})
            usuarios = {row["id"]: row["username"] for row in ur.mappings()}
    for f in filas:
        f["nombre"] = usuarios.get(f["apostador_id"], f"Usuario {f['apostador_id']}")
    return {"filas": filas, "usuarios": usuarios}


@router.get("/puntos-por-fase/{torneo_id}", summary="Puntos por fase de cada apostador (análisis)")
async def puntos_por_fase(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")
    data = await _puntos_por_fase_data(db, torneo_id)
    return {"torneo_id": torneo_id, "filas": data["filas"]}


@router.get("/puntos-por-fase/{torneo_id}/excel", summary="Exportar puntos por fase a Excel (admin)")
async def puntos_por_fase_excel(torneo_id: int, current: CurrentUser, db: DBSession):
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")
    try:
        wb, _ = await _build_auditoria_workbook(db, torneo_id)
    except ImportError:
        raise HTTPException(500, "openpyxl no está instalado")

    out_dir = os.path.join(os.path.dirname(__file__), "..", "..", "..", "..", "static", "exports")
    out_dir = os.path.abspath(out_dir)
    os.makedirs(out_dir, exist_ok=True)
    fname = f"puntos_por_fase_torneo_{torneo_id}.xlsx"
    fpath = os.path.join(out_dir, fname)
    wb.save(fpath)
    return FileResponse(fpath, filename=fname,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Auditoría por fase (real vs apostado, puntos por ítem) ───────────────────

@router.get("/auditoria-fase/{torneo_id}", summary="Auditoría por fase: real vs apostado y puntos por ítem")
async def auditoria_fase(torneo_id: int, current: CurrentUser, db: DBSession,
                         apostador_id: int | None = None) -> dict:
    """Devuelve, por fase y por partido, el resultado real vs el apostado y los
    puntos ganados por ítem (marcador/minuto/amarillas/VAR). Si se pasa
    apostador_id, filtra por ese apostador."""
    if not await _check_admin(current):
        # Un apostador puede ver SU propia auditoría
        apostador_id = current.id

    params: dict = {"tid": torneo_id}
    filtro = ""
    if apostador_id is not None:
        filtro = " AND d.apostador_id = :uid"
        params["uid"] = apostador_id

    try:
        r = await db.execute(
            text(f"""
                SELECT d.fase_id, d.fase_tipo, d.fase_nombre, d.partido_id, d.apostador_id,
                       d.multiplicador, d.pred_local, d.pred_visitante,
                       d.real_local, d.real_visitante, d.pts_marcador,
                       d.pred_minuto, d.real_minuto, d.gano_minuto, d.pts_minuto,
                       d.pred_amarillas, d.real_amarillas, d.pts_amarillas,
                       d.pred_var, d.real_var, d.pts_var, d.pts_bonus, d.pts_total
                FROM puntaje_detalle d
                WHERE d.torneo_id=:tid {filtro}
                ORDER BY d.fase_id, d.partido_id, d.apostador_id
            """),
            params,
        )
    except Exception:
        await db.rollback()
        return {"torneo_id": torneo_id, "fases": []}

    rows = [dict(row) for row in r.mappings()]
    uids = sorted({x["apostador_id"] for x in rows})
    usuarios: dict[int, str] = {}
    if uids:
        async with _app_engine.connect() as conn:
            ur = await conn.execute(
                text("SELECT id, username FROM users WHERE id = ANY(:ids)"), {"ids": uids})
            usuarios = {row["id"]: row["username"] for row in ur.mappings()}

    fases: dict[int, dict] = {}
    for x in rows:
        fid = x["fase_id"]
        fase = fases.setdefault(fid, {
            "fase_id": fid, "fase_tipo": x["fase_tipo"],
            "fase_nombre": x["fase_nombre"], "multiplicador": x["multiplicador"],
            "partidos": [],
        })
        fase["partidos"].append({
            "partido_id": x["partido_id"],
            "apostador_id": x["apostador_id"],
            "apostador": usuarios.get(x["apostador_id"], f"Usuario {x['apostador_id']}"),
            "marcador": {"apostado": f"{x['pred_local']}-{x['pred_visitante']}",
                         "real": f"{x['real_local']}-{x['real_visitante']}",
                         "pts": x["pts_marcador"]},
            "minuto": {"apostado": x["pred_minuto"], "real": x["real_minuto"],
                       "gano": x["gano_minuto"], "pts": x["pts_minuto"]},
            "amarillas": {"apostado": x["pred_amarillas"], "real": x["real_amarillas"],
                          "pts": x["pts_amarillas"]},
            "var": {"apostado": x["pred_var"], "real": x["real_var"], "pts": x["pts_var"]},
            "bonus": x["pts_bonus"], "total": x["pts_total"],
        })
    return {"torneo_id": torneo_id, "fases": list(fases.values())}


# ── Transparencia por fase (vista de apostador) ──────────────────────────────

_PHASE_ORDER = ["grupo", "ronda32", "ronda16", "cuartos", "semis", "tercer_puesto", "final"]
_PHASE_LABELS = {
    "grupo": "Grupos", "ronda32": "Ronda 32", "ronda16": "Octavos",
    "cuartos": "Cuartos", "semis": "Semis", "tercer_puesto": "3er puesto", "final": "Final",
}
# Etiquetas completas para títulos de tarjeta (sección "Apuestas y puntajes por fase")
_PHASE_LABELS_FULL = {
    "grupo": "Fase de grupos", "ronda32": "Ronda de 32", "ronda16": "Octavos de final",
    "cuartos": "Cuartos de final", "semis": "Semifinales",
    "tercer_puesto": "Tercer puesto", "final": "Final",
}


async def _build_auditoria_workbook(db, torneo_id: int):
    """Construye el Workbook ÚNICO de auditoría usado por TODAS las salidas de
    export (Auditoría, Transparencia, Puntos por fase).

    Estructura:
      - Hoja 1 'Puntaje general': lista de apostadores con la sumatoria de sus
        puntajes a la fecha (marcador + bonus partido + bonus de terceros),
        rankeada de mayor a menor.
      - Una hoja por fase, empezando por 'Fase de grupos'. Dentro, ordenado por
        grupo y por partido; en cada partido los apostadores se agrupan en
        Pleno (marcador exacto) / Ganador (acertó resultado) / Cero acierto.
        A nivel de apostador se muestra el puntaje del partido (marcador + bonus
        por ítem) sumarizando el Total.

    Devuelve (Workbook, torneo_nombre).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    tr = await db.execute(text("SELECT nombre FROM torneo WHERE id=:tid"), {"tid": torneo_id})
    t = tr.one_or_none()
    torneo_nombre = t[0] if t else f"Torneo {torneo_id}"

    # Orden jerárquico del bracket (izquierda→derecha) por número de partido FIFA.
    # MISMO orden visual que pronosticos/resultado (_renderBracketTree).
    KO_BRACKET_ORDER = {
        "ronda32":       [74, 77, 73, 75, 83, 84, 81, 82, 76, 78, 79, 80, 86, 88, 85, 87],
        "ronda16":       [89, 90, 93, 94, 91, 92, 95, 96],
        "cuartos":       [97, 98, 99, 100],
        "semis":         [101, 102],
        "tercer_puesto": [103],
        "final":         [104],
    }
    try:
        from app.services.ko_scoring import build_num_maps
        _maps = await build_num_maps(db, torneo_id)
        _pid2num = _maps.get("pid2num", {})
    except Exception:
        _pid2num = {}

    def _bracket_pos(tipo, pid):
        order = KO_BRACKET_ORDER.get(tipo)
        if not order:
            return None
        num = _pid2num.get(pid)
        try:
            return order.index(num)
        except (ValueError, TypeError):
            return None

    rm = await db.execute(
        text("""
            SELECT pd.apostador_id, pd.fase_id, pd.fase_tipo, pd.fase_nombre,
                   pd.partido_id, pd.multiplicador,
                   pd.pred_local, pd.pred_visitante, pd.real_local, pd.real_visitante,
                   pd.pts_marcador_base, pd.pts_marcador,
                   COALESCE(pd.pts_resultado, 0)         AS pts_resultado,
                   pd.pts_minuto, pd.pts_amarillas, pd.pts_var,
                   COALESCE(pd.pts_rojas, 0)             AS pts_rojas,
                   COALESCE(pd.pts_penales_partido, 0)   AS pts_penales_partido,
                   COALESCE(pd.pts_penales_tanda, 0)     AS pts_penales_tanda,
                   pd.pts_bonus, pd.pts_total,
                   p.jornada,
                   COALESCE(el.nombre_es, el.nombre) AS local_nombre,
                   COALESCE(ev.nombre_es, ev.nombre) AS visit_nombre
            FROM puntaje_detalle pd
            LEFT JOIN partido p ON p.id = pd.partido_id
            LEFT JOIN equipo el ON el.id = p.equipo_local_id
            LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
            WHERE pd.torneo_id = :tid
            ORDER BY pd.fase_id, p.jornada NULLS LAST, pd.partido_id, pd.apostador_id
        """),
        {"tid": torneo_id},
    )
    detalle = [dict(row) for row in rm.mappings()]

    # Nombres de TODOS los apostadores activos (aunque tengan 0 puntos)
    ids_detalle = {d["apostador_id"] for d in detalle}
    async with _app_engine.connect() as conn:
        ar = await conn.execute(text("""
            SELECT u.id, u.username FROM users u
            JOIN user_roles ur ON ur.user_id = u.id
            JOIN roles ro ON ro.id = ur.role_id
            WHERE ro.name = 'apostador' AND u.is_active = TRUE"""))
        user_map = {row["id"]: row["username"] for row in ar.mappings()}
        extra = [i for i in ids_detalle if i not in user_map]
        if extra:
            ur = await conn.execute(text("SELECT id, username FROM users WHERE id = ANY(:ids)"), {"ids": extra})
            for row in ur.mappings():
                user_map[row["id"]] = row["username"]

    # Puntajes globales A-G por apostador
    pg_r = await db.execute(
        text("SELECT * FROM puntaje_global WHERE torneo_id = :tid"),
        {"tid": torneo_id},
    )
    pts_glob_map: dict = {}
    for row in pg_r.mappings():
        pts_glob_map[row["apostador_id"]] = dict(row)

    # Apuestas globales A-G por apostador (para hoja Globales)
    ag_r = await db.execute(
        text("SELECT * FROM apuesta_global WHERE torneo_id = :tid"),
        {"tid": torneo_id},
    )
    apuesta_glob_map: dict = {}
    for row in ag_r.mappings():
        apuesta_glob_map[row["apostador_id"]] = dict(row)

    # ── Estilos ──
    HDR_FILL    = PatternFill("solid", start_color="1a3a5c")
    GRP_FILL    = PatternFill("solid", start_color="243044")
    ALT_FILL    = PatternFill("solid", start_color="1e2535")
    PART_FILL   = PatternFill("solid", start_color="0f2336")
    GRPHDR_FILL = PatternFill("solid", start_color="33240f")
    PLENO_FILL  = PatternFill("solid", start_color="1e4d2b")
    GANA_FILL   = PatternFill("solid", start_color="3a3410")
    CERO_FILL   = PatternFill("solid", start_color="3a1414")
    TOP_FILL    = PatternFill("solid", start_color="1e4d2b")
    W_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
    N_FONT     = Font(name="Calibri", color="E0E0E0", size=9)
    PLENO_FONT = Font(name="Calibri", color="46d17f", bold=True, size=9)
    GANA_FONT  = Font(name="Calibri", color="fbbf24", bold=True, size=9)
    CERO_FONT  = Font(name="Calibri", color="fc7c7c", bold=True, size=9)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(horizontal="left",   vertical="center")
    thin   = Side(style="thin", color="2e3540")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Hoja 1: Puntaje general ──
    # Columnas: H=resultado, I=marcador exacto, J=amarillas, K=rojas, L=VAR, N=minuto, M=penales partido
    _CAT_KEYS = ["pts_resultado", "pts_marcador", "pts_amarillas", "pts_rojas",
                 "pts_var", "pts_minuto", "pts_penales_partido"]
    sub: dict[int, dict] = defaultdict(lambda: {k: 0 for k in _CAT_KEYS})
    for d in detalle:
        s = sub[d["apostador_id"]]
        for k in _CAT_KEYS:
            s[k] += (d.get(k) or 0)
    gen_cols = ["#", "Apostador",
                "H\nResultado", "I\nExacto", "J\nAmar.", "K\nRojas", "L\nVAR", "N\nMinuto", "M\nPen.P.",
                "Sub", "Glob", "Total"]
    gen_w    = [4, 26, 7, 7, 7, 7, 7, 7, 7, 8, 8, 9]
    TITLE_FILL = PatternFill("solid", start_color="1a2840")
    HDR_CAT_FONT = Font(name="Calibri", color="94a3b8", bold=True, size=8)
    ws_g = wb.create_sheet("Puntaje general")
    ws_g.sheet_view.showGridLines = False
    for i, w in enumerate(gen_w, start=1):
        ws_g.column_dimensions[chr(64 + i)].width = w
    ws_g.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(gen_cols))
    ws_g.cell(1, 1, f"Puntaje general — {torneo_nombre}").font = Font(
        name="Calibri", color="E05020", bold=True, size=13)
    ws_g["A2"] = f"Generado: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M UTC')}"
    ws_g["A2"].font = Font(name="Calibri", color="888888", size=9)
    # Leyenda de columnas fila 3
    leyenda = [
        "", "",
        "H=Resultado (gana/pierde/empata)", "I=Marcador exacto", "J=Amarillas exactas",
        "K=Rojas exactas", "L=Decisiones VAR", "N=Minuto 1er gol", "M=Penales en el partido",
        "Sub=H+I+J+K+L+N+M", "Glob=A-G globales", "Total=Sub+Glob"
    ]
    for col, txt in enumerate(leyenda, start=1):
        c = ws_g.cell(3, col, txt)
        c.font = Font(name="Calibri", color="64748b", italic=True, size=7)
    ws_g.row_dimensions[4].height = 28
    for col, h in enumerate(gen_cols, start=1):
        c = ws_g.cell(4, col, h)
        c.font = HDR_CAT_FONT if col > 2 else W_FONT
        c.fill = HDR_FILL; c.alignment = CENTER; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    filas = []
    for uid, nombre in user_map.items():
        s = sub.get(uid, {k: 0 for k in _CAT_KEYS})
        pg = pts_glob_map.get(uid, {})
        globales = pg.get("pts_total") or 0
        sub_pts = sum(s[k] for k in _CAT_KEYS)
        total = sub_pts + globales
        filas.append((nombre, s, globales, sub_pts, total))
    filas.sort(key=lambda x: (-x[4], x[0].lower()))
    for idx, (nombre, s, globales, sub_pts, total) in enumerate(filas, start=1):
        ri = idx + 4
        vals = [idx, nombre,
                s["pts_resultado"] or "", s["pts_marcador"] or "",
                s["pts_amarillas"] or "", s["pts_rojas"] or "",
                s["pts_var"] or "", s["pts_minuto"] or "",
                s["pts_penales_partido"] or "",
                sub_pts or "", globales or "", total]
        fill = TOP_FILL if idx == 1 and total > 0 else (ALT_FILL if ri % 2 == 0 else GRP_FILL)
        for col, val in enumerate(vals, start=1):
            c = ws_g.cell(ri, col, val)
            c.font = N_FONT; c.fill = fill; c.border = BORDER
            c.alignment = LEFT if col == 2 else CENTER

    # ── Hojas por fase ──
    # por_fase[fase_tipo][grupo/fase_nombre][partido_id] -> filas de apostadores
    por_fase: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for d in detalle:
        por_fase[d["fase_tipo"]][d["fase_nombre"]][d["partido_id"]].append(d)
    fases = [ft for ft in _PHASE_ORDER if ft in por_fase] + \
            [ft for ft in por_fase if ft not in _PHASE_ORDER]

    cols   = ["Apostador", "Pronóstico", "H\nRes.", "I\nExact.", "J\nAmar.", "K\nRojas", "L\nVAR", "N\nMin.", "M\nPen.P.", "O\nP.Tanda", "Total"]
    cols_w = [26, 11, 7, 7, 7, 7, 7, 7, 7, 8, 8]
    cat_meta = {
        3: ("✅ PLENO — marcador exacto",   PLENO_FILL, PLENO_FONT),
        1: ("➕ GANADOR — acertó resultado", GANA_FILL,  GANA_FONT),
        0: ("✗ CERO ACIERTO",               CERO_FILL,  CERO_FONT),
    }

    for ft in fases:
        nombre_fase = _PHASE_LABELS_FULL.get(ft, ft)
        ws = wb.create_sheet(nombre_fase[:31])
        ws.sheet_view.showGridLines = False
        for i, w in enumerate(cols_w, start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        ws.cell(1, 1, f"{nombre_fase} — {torneo_nombre}").font = Font(
            name="Calibri", color="E05020", bold=True, size=13)
        ri = 3
        es_grupo = ft == "grupo"
        for grupo_nombre in sorted(por_fase[ft].keys()):
            partidos = por_fase[ft][grupo_nombre]
            if es_grupo:
                ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=len(cols))
                c = ws.cell(ri, 1, f"▣ {grupo_nombre}")
                c.font = Font(name="Calibri", color="FFD27F", bold=True, size=12)
                c.fill = GRPHDR_FILL; c.alignment = LEFT
                ri += 2

            def _pkey(pid):
                if not es_grupo:
                    bp = _bracket_pos(ft, pid)
                    if bp is not None:
                        return (0, bp)
                r0 = partidos[pid][0]
                return (r0.get("jornada") or 0, pid)

            for pid in sorted(partidos.keys(), key=_pkey):
                rows = partidos[pid]
                d0 = rows[0]
                partido = f"{d0.get('local_nombre') or '?'} vs {d0.get('visit_nombre') or '?'}"
                real = (f"{d0['real_local']}-{d0['real_visitante']}"
                        if d0["real_local"] is not None else "—")
                jor = f"J{d0['jornada']} · " if d0.get("jornada") else ""
                ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=len(cols))
                mult = d0['multiplicador'] or 1
                py_marker = "   🇵🇾 PARAGUAY x2" if mult > 1 else ""
                c = ws.cell(ri, 1, f"⚽ {jor}{partido}   ·   Real: {real}   ·   x{mult}{py_marker}")
                c.font = Font(name="Calibri", color="FFD27F" if mult == 1 else "7EE0A0", bold=True, size=11)
                c.fill = PART_FILL; c.alignment = LEFT
                ri += 1
                ws.row_dimensions[ri].height = 28
                for col, h in enumerate(cols, start=1):
                    c = ws.cell(ri, col, h)
                    c.font = W_FONT; c.fill = HDR_FILL; c.border = BORDER
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ri += 1

                buckets = {3: [], 1: [], 0: []}
                for d in rows:
                    base = d["pts_marcador_base"]
                    buckets[3 if base == 3 else (1 if base == 1 else 0)].append(d)

                for cat in (3, 1, 0):
                    grp = buckets[cat]
                    if not grp:
                        continue
                    label, cfill, cfont = cat_meta[cat]
                    ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=len(cols))
                    c = ws.cell(ri, 1, f"{label}   ({len(grp)})")
                    c.font = cfont; c.fill = cfill; c.alignment = LEFT; c.border = BORDER
                    ri += 1
                    grp.sort(key=lambda d: (-(d["pts_total"] or 0),
                                            user_map.get(d["apostador_id"], "").lower()))
                    for d in grp:
                        nombre = user_map.get(d["apostador_id"], f"Usuario {d['apostador_id']}")
                        pred = (f"{d['pred_local']}-{d['pred_visitante']}"
                                if d["pred_local"] is not None else "—")
                        pts_total_row = (
                            (d.get("pts_resultado") or 0) +
                            (d.get("pts_marcador") or 0) +
                            (d.get("pts_amarillas") or 0) +
                            (d.get("pts_rojas") or 0) +
                            (d.get("pts_var") or 0) +
                            (d.get("pts_minuto") or 0) +
                            (d.get("pts_penales_partido") or 0) +
                            (d.get("pts_penales_tanda") or 0)
                        )
                        vals = [nombre, pred,
                                d.get("pts_resultado") or None,
                                d.get("pts_marcador") or None,
                                d.get("pts_amarillas") or None,
                                d.get("pts_rojas") or None,
                                d.get("pts_var") or None,
                                d.get("pts_minuto") or None,
                                d.get("pts_penales_partido") or None,
                                d.get("pts_penales_tanda") or None,
                                pts_total_row or None]
                        fill = ALT_FILL if ri % 2 == 0 else GRP_FILL
                        for col, val in enumerate(vals, start=1):
                            c = ws.cell(ri, col, val if val is not None else "")
                            c.font = N_FONT; c.fill = fill; c.border = BORDER
                            c.alignment = LEFT if col == 1 else CENTER
                        ri += 1
                ri += 1  # separador entre partidos

    # ── Hoja Globales A-G ──
    GLOB_FILL   = PatternFill("solid", start_color="1a2840")
    GLOB_HDR    = PatternFill("solid", start_color="1e3a5c")
    GLOB_PT_FNT = Font(name="Calibri", color="46d17f", bold=True, size=9)
    glob_cols   = ["Apostador",
                   "A · Campeón", "B · Fin.1", "B · Fin.2", "C · Goleador",
                   "D · Peor eq.", "E · Gol.G", "E · Gol.P",
                   "F · Etapa Py", "G · Goles Py",
                   "Pts A", "Pts B", "Pts C", "Pts D", "Pts E", "Pts F", "Pts G", "Total"]
    glob_w      = [22, 14, 14, 14, 16, 14, 7, 7, 13, 10, 6, 6, 6, 6, 6, 6, 6, 7]
    ws_gl = wb.create_sheet("Globales")
    ws_gl.sheet_view.showGridLines = False
    for i, w in enumerate(glob_w, start=1):
        ws_gl.column_dimensions[chr(64 + i)].width = w
    ws_gl.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(glob_cols))
    ws_gl.cell(1, 1, f"Pronósticos Globales A-G — {torneo_nombre}").font = Font(
        name="Calibri", color="818cf8", bold=True, size=13)
    ws_gl["A2"] = f"Generado: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M UTC')}"
    ws_gl["A2"].font = Font(name="Calibri", color="888888", size=9)
    for col, h in enumerate(glob_cols, start=1):
        c = ws_gl.cell(4, col, h)
        c.font = W_FONT; c.fill = GLOB_HDR; c.alignment = CENTER; c.border = BORDER

    # Mapeo equipo_id -> nombre (para los selectores)
    eq_r = await db.execute(
        text("SELECT id, COALESCE(nombre_es, nombre) AS nombre FROM equipo ORDER BY nombre"),
        {},
    )
    eq_map = {row["id"]: row["nombre"] for row in eq_r.mappings()}

    glob_filas = []
    for uid, nombre in user_map.items():
        ag = apuesta_glob_map.get(uid, {})
        pg = pts_glob_map.get(uid, {})
        if not ag and not pg:
            continue
        def _eq(eid): return eq_map.get(eid, f"#{eid}") if eid else "—"
        def _pt(k): v = pg.get(k); return v if v else ""
        row_vals = [
            nombre,
            _eq(ag.get("pred_campeon_id")),
            _eq(ag.get("pred_finalista1_id")),
            _eq(ag.get("pred_finalista2_id")),
            ag.get("pred_goleador") or "—",
            _eq(ag.get("pred_peor_equipo_id")),
            ag.get("pred_goleada_ganador") if ag.get("pred_goleada_ganador") is not None else "—",
            ag.get("pred_goleada_perdedor") if ag.get("pred_goleada_perdedor") is not None else "—",
            ag.get("pred_etapa_paraguay") or "—",
            ag.get("pred_goles_paraguay") if ag.get("pred_goles_paraguay") is not None else "—",
            _pt("pts_campeon"), _pt("pts_finalistas"), _pt("pts_goleador"), _pt("pts_peor_equipo"),
            _pt("pts_goleada"), _pt("pts_etapa_paraguay"), _pt("pts_goles_paraguay"),
            pg.get("pts_total") or "",
        ]
        glob_filas.append((pg.get("pts_total") or 0, nombre, row_vals))

    glob_filas.sort(key=lambda x: (-x[0], x[1].lower()))
    for idx, (_, _, row_vals) in enumerate(glob_filas, start=1):
        ri = idx + 4
        fill = ALT_FILL if ri % 2 == 0 else GLOB_FILL
        for col, val in enumerate(row_vals, start=1):
            c = ws_gl.cell(ri, col, val)
            is_pts_col = col >= len(glob_cols) - 7  # últimas 8 = pts cols
            c.font = GLOB_PT_FNT if (is_pts_col and val) else N_FONT
            c.fill = fill; c.border = BORDER
            c.alignment = LEFT if col == 1 else CENTER

    if not glob_filas:
        ws_gl.merge_cells(start_row=5, start_column=1, end_row=5, end_column=len(glob_cols))
        c = ws_gl.cell(5, 1, "Sin pronósticos globales registrados")
        c.font = N_FONT; c.fill = GLOB_FILL; c.alignment = CENTER

    if not wb.sheetnames:
        wb.create_sheet("Sin datos")
    return wb, torneo_nombre


@router.get("/transparencia/{torneo_id}",
            summary="Vista de transparencia por fase: resultados/clasificados reales, mis apuestas+puntajes y ranking acumulado con subtotal")
async def transparencia(torneo_id: int, current: CurrentUser, db: DBSession,
                        apostador_id: int | None = None) -> dict:
    # Nombres de equipos
    re_ = await db.execute(text("SELECT id, COALESCE(nombre_es, nombre) AS n FROM equipo"))
    equipos = {row[0]: row[1] for row in re_}

    # 1) Clasificados reales — fase de grupos
    grupos_out: list[dict] = []
    try:
        st = await _calc_standings_reales(db, torneo_id)
        if st:
            mejores, _ = seleccionar_mejores_terceros(st, fill_incomplete=False)
            terceros_ids = {t["equipo_id"] for t in mejores}
            for letra in sorted(st.keys()):
                eqs = st[letra]["equipos"]
                grupos_out.append({"grupo": letra, "equipos": [
                    {"nombre": e["nombre"], "pos": e.get("pos"), "pj": e["pj"],
                     "pts": e["pts"], "gd": e["gd"], "gf": e["gf"],
                     "clasificado": (e.get("pos") in (1, 2)) or (e["equipo_id"] in terceros_ids),
                     "tercero": e["equipo_id"] in terceros_ids}
                    for e in eqs]})
    except Exception:
        await db.rollback()

    # 2) Resultados reales — fases KO (con ganador que avanza)
    rk = await db.execute(
        text("""
            SELECT f.id AS fase_id, f.tipo, f.nombre, f.orden,
                   p.id AS pid, p.equipo_local_id, p.equipo_visitante_id,
                   p.goles_local, p.goles_visitante,
                   p.penales_local, p.penales_visitante, p.estado
            FROM fase f JOIN partido p ON p.fase_id = f.id
            WHERE f.torneo_id = :tid AND f.tipo <> 'grupo'
            ORDER BY f.orden, p.id
        """),
        {"tid": torneo_id},
    )
    ko_map: dict[int, dict] = {}
    for row in rk.mappings():
        d = dict(row)
        fase = ko_map.setdefault(d["fase_id"], {
            "fase_id": d["fase_id"], "tipo": d["tipo"], "nombre": d["nombre"],
            "orden": d["orden"], "partidos": []})
        gana = None
        fin = d["estado"] == "finalizado"
        if fin:
            w, _l = ko_scoring.winner_loser(d)
            gana = equipos.get(w) if w else None
        fase["partidos"].append({
            "local": equipos.get(d["equipo_local_id"], "Por definir"),
            "visitante": equipos.get(d["equipo_visitante_id"], "Por definir"),
            "gl": d["goles_local"], "gv": d["goles_visitante"],
            "pen_l": d["penales_local"], "pen_v": d["penales_visitante"],
            "ganador": gana, "finalizado": fin})
    ko_out = sorted(ko_map.values(), key=lambda x: x["orden"])

    # 3) Ranking acumulado por fase (matriz apostador × fase + subtotales)
    try:
        rr = await db.execute(
            text("""SELECT apostador_id, fase_tipo, SUM(pts_total)::int AS pts
                    FROM puntaje_detalle WHERE torneo_id=:tid
                    GROUP BY apostador_id, fase_tipo"""),
            {"tid": torneo_id},
        )
        bytipo: dict[int, dict] = {}
        for row in rr.mappings():
            bytipo.setdefault(row["apostador_id"], {})[row["fase_tipo"]] = row["pts"]
    except Exception:
        await db.rollback()
        bytipo = {}

    fases_presentes = [t for t in _PHASE_ORDER if any(t in v for v in bytipo.values())]

    # Incluir TODOS los apostadores activos (rol 'apostador') aunque no tengan
    # puntos: pueden sumarse en una fase posterior y deben figurar en cero.
    async with _app_engine.connect() as conn:
        ar = await conn.execute(
            text("""
                SELECT u.id, u.username FROM users u
                JOIN user_roles ur ON ur.user_id = u.id
                JOIN roles ro ON ro.id = ur.role_id
                WHERE ro.name = 'apostador' AND u.is_active = TRUE
            """))
        usuarios: dict[int, str] = {row["id"]: row["username"] for row in ar.mappings()}
        extra_ids = [i for i in bytipo if i not in usuarios]
        if extra_ids:
            ur = await conn.execute(
                text("SELECT id, username FROM users WHERE id = ANY(:ids)"), {"ids": extra_ids})
            for row in ur.mappings():
                usuarios[row["id"]] = row["username"]
    uids = sorted(set(list(usuarios.keys()) + list(bytipo.keys())))

    apostadores = []
    for uid in uids:
        por: dict[str, dict] = {}
        acc = 0
        for t in fases_presentes:
            pts = bytipo.get(uid, {}).get(t, 0)
            acc += pts
            por[t] = {"pts": pts, "acumulado": acc}
        apostadores.append({
            "apostador_id": uid, "nombre": usuarios.get(uid, f"Usuario {uid}"),
            "por_fase": por, "total": acc,
            "es_actual": uid == current.id})
    apostadores.sort(key=lambda x: (-x["total"], (x["nombre"] or "").lower()))

    # 4) Apuestas y puntajes por fase del apostador seleccionado (default: usuario actual)
    #    Incluye nombres y logos de los equipos para mostrar el partido con banderitas.
    target_uid = apostador_id if apostador_id else current.id
    rm = await db.execute(
        text("""
            SELECT pd.fase_id, pd.fase_tipo, pd.fase_nombre, pd.partido_id, pd.multiplicador,
                   pd.pred_local, pd.pred_visitante, pd.real_local, pd.real_visitante, pd.pts_marcador,
                   pd.pred_minuto, pd.real_minuto, pd.gano_minuto, pd.pts_minuto,
                   pd.pred_amarillas, pd.real_amarillas, pd.pts_amarillas,
                   pd.pred_var, pd.real_var, pd.pts_var, pd.pts_bonus, pd.pts_total,
                   COALESCE(el.nombre_es, el.nombre) AS local_nombre,
                   COALESCE(ev.nombre_es, ev.nombre) AS visit_nombre,
                   el.logo_url AS local_logo, ev.logo_url AS visit_logo
            FROM puntaje_detalle pd
            LEFT JOIN partido p ON p.id = pd.partido_id
            LEFT JOIN equipo el ON el.id = p.equipo_local_id
            LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
            WHERE pd.torneo_id=:tid AND pd.apostador_id=:uid
            ORDER BY pd.fase_id, pd.partido_id
        """),
        {"tid": torneo_id, "uid": target_uid},
    )
    mias_map: dict[str, dict] = {}
    for row in rm.mappings():
        d = dict(row)
        f = mias_map.setdefault(d["fase_tipo"], {
            "tipo": d["fase_tipo"],
            "nombre": _PHASE_LABELS_FULL.get(d["fase_tipo"], d["fase_nombre"]),
            "subtotal": 0, "partidos": []})
        f["subtotal"] += d["pts_total"]
        f["partidos"].append(d)
    mias = [mias_map[t] for t in _PHASE_ORDER if t in mias_map]

    return {
        "torneo_id": torneo_id,
        "grupos": grupos_out,
        "ko": ko_out,
        "ranking": {"fases": fases_presentes,
                    "labels": {t: _PHASE_LABELS[t] for t in fases_presentes},
                    "apostadores": apostadores},
        "mias": mias,
        "apostador_sel": target_uid,
        "apostador_sel_nombre": usuarios.get(target_uid, f"Usuario {target_uid}"),
        "apostador_propio": current.id,
    }


@router.get("/transparencia/{torneo_id}/export",
            summary="Exportar a Excel todas las apuestas y puntajes de todos los apostadores (todas las fases)")
async def exportar_transparencia(torneo_id: int, current: CurrentUser, db: DBSession):
    """Genera y descarga el Excel ÚNICO de auditoría (mismo formato que todas las
    salidas): hoja 'Puntaje general' + una hoja por fase con buckets
    Pleno/Ganador/Cero y orden jerárquico de bracket en KO. Cualquier usuario."""
    import io
    wb, _ = await _build_auditoria_workbook(db, torneo_id)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    fname = f"transparencia_torneo{torneo_id}_{ts}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Auditoría ────────────────────────────────────────────────────────────────

@router.get("/auditoria/{torneo_id}", summary="Lista de snapshots de auditoría")
async def listar_auditorias(torneo_id: int, current: CurrentUser, db: DBSession) -> list[dict]:
    try:
        r = await db.execute(
            text("""
                SELECT id, torneo_id, generado_at, descripcion, archivo_path
                FROM auditoria_apuestas
                WHERE torneo_id = :tid
                ORDER BY generado_at DESC
            """),
            {"tid": torneo_id}
        )
        rows = [dict(row) for row in r.mappings()]
    except Exception:
        return []  # Tabla no migrada aún

    for row in rows:
        row["generado_at"] = row["generado_at"].isoformat() if row["generado_at"] else None
    return rows


@router.post("/auditoria/{torneo_id}", summary="Generar Excel de auditoría (admin)")
async def generar_auditoria(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    # Nombre del torneo
    tr = await db.execute(text("SELECT nombre FROM torneo WHERE id=:tid"), {"tid": torneo_id})
    t = tr.one_or_none()
    if not t:
        raise HTTPException(404, "Torneo no encontrado")

    # Conteos para la descripción del snapshot
    rc = await db.execute(
        text("""
            SELECT COUNT(DISTINCT pd.apostador_id) AS aps, COUNT(DISTINCT pd.partido_id) AS pts
            FROM puntaje_detalle pd WHERE pd.torneo_id = :tid
        """),
        {"tid": torneo_id},
    )
    cnt = rc.one_or_none()
    n_aps  = (cnt[0] if cnt else 0) or 0
    n_part = (cnt[1] if cnt else 0) or 0

    # ── Construir Excel ÚNICO (mismo formato que todas las salidas) ──────────
    wb, _ = await _build_auditoria_workbook(db, torneo_id)

    # Guardar en /static/auditorias/
    static_dir = os.path.join(
        os.path.dirname(__file__), "..", "..", "..", "..", "static", "auditorias"
    )
    os.makedirs(static_dir, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"auditoria_torneo{torneo_id}_{ts}.xlsx"
    filepath = os.path.join(static_dir, filename)
    wb.save(filepath)

    # Registrar en BD (si la tabla ya existe)
    try:
        await db.execute(
            text("""
                INSERT INTO auditoria_apuestas
                    (torneo_id, generado_por, archivo_path, descripcion)
                VALUES (:tid, :uid, :path, :desc)
            """),
            {
                "tid":  torneo_id,
                "uid":  current.id,
                "path": f"auditorias/{filename}",
                "desc": (f"Snapshot {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M UTC')}"
                         f" — {n_aps} apostadores, {n_part} partidos"),
            }
        )
        await db.commit()
    except Exception:
        pass  # Tabla no migrada aún; el archivo igual se genera

    return {
        "ok":          True,
        "filename":    filename,
        "apostadores": n_aps,
        "partidos":    n_part,
        "url":         f"/static/auditorias/{filename}",
    }


@router.get("/auditoria/download/{auditoria_id}", summary="Descargar Excel de auditoría")
async def descargar_auditoria(auditoria_id: int, current: CurrentUser, db: DBSession):
    try:
        r = await db.execute(
            text("SELECT archivo_path FROM auditoria_apuestas WHERE id=:aid"),
            {"aid": auditoria_id}
        )
        row = r.one_or_none()
    except Exception:
        raise HTTPException(404, "Tabla de auditoría no disponible")

    if not row:
        raise HTTPException(404, "Auditoría no encontrada")

    static_dir = os.path.join(
        os.path.dirname(__file__), "..", "..", "..", "..", "static"
    )
    filepath = os.path.join(static_dir, row[0])
    if not os.path.exists(filepath):
        raise HTTPException(404, "Archivo no encontrado en servidor")

    return FileResponse(
        filepath,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=os.path.basename(filepath),
    )


# ── Confirmar datos de partido (protege del sync) ────────────────────────────

@router.post("/confirmar-partido-stats/{torneo_id}", summary="Migra columna datos_confirmados y confirma todos los finalizados (admin)")
async def confirmar_partido_stats(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """Agrega la columna datos_confirmados (si no existe) y marca como confirmados
    todos los partidos finalzados del torneo. Los partidos confirmados no son
    sobreescritos por el sync automático de API-Football."""
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")
    # Migración idempotente
    await db.execute(text(
        "ALTER TABLE partido ADD COLUMN IF NOT EXISTS datos_confirmados BOOLEAN DEFAULT FALSE"
    ))
    # Marcar todos los finalizados del torneo
    r = await db.execute(
        text("""
            UPDATE partido SET datos_confirmados = TRUE
            WHERE torneo_id = :tid AND estado = 'finalizado'
        """),
        {"tid": torneo_id},
    )
    confirmados = r.rowcount
    await db.commit()
    return {"ok": True, "confirmados": confirmados, "msg": f"{confirmados} partidos marcados como confirmados — el sync los omitirá"}


@router.patch("/confirmar-partido/{partido_id}", summary="Activar/desactivar protección de un partido (admin)")
async def toggle_partido_confirmado(
    partido_id: int,
    current: CurrentUser,
    db: DBSession,
    confirmar: bool = True,
) -> dict:
    """Activa o desactiva datos_confirmados para un partido individual."""
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")
    await db.execute(
        text("UPDATE partido SET datos_confirmados = :val WHERE id = :pid"),
        {"val": confirmar, "pid": partido_id},
    )
    await db.commit()
    accion = "confirmado (protegido del sync)" if confirmar else "des-confirmado (sync habilitado)"
    return {"ok": True, "partido_id": partido_id, "datos_confirmados": confirmar, "msg": f"Partido {accion}"}


@router.post("/recalc-fair-play/{torneo_id}", summary="Re-extrae tarjetas por equipo de API-Football y recalcula fair play (admin)")
async def recalc_fair_play(
    torneo_id: int,
    current: CurrentUser,
    db: DBSession,
    max_partidos: int = 50,
) -> dict:
    """
    Re-fetcha los eventos de los partidos de grupo desde API-Football para obtener
    las tarjetas por equipo (local_amarillas / visitante_amarillas / local_rojas /
    visitante_rojas). No modifica goles, estado, ni ningún otro campo confirmado.

    Usa cuota de API-Football (1 call por partido + 1 bulk). Con el plan Free
    (100 calls/día) y max_partidos=50, consume hasta 51 calls.

    Luego recalcula fair_play_pts en participacion y por tanto el ranking de
    mejores terceros pasa a aplicar el criterio FIFA de forma correcta.
    """
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    # Asegurar columnas de tarjetas por equipo
    for _col in [
        "ALTER TABLE partido ADD COLUMN IF NOT EXISTS local_amarillas     INT",
        "ALTER TABLE partido ADD COLUMN IF NOT EXISTS visitante_amarillas INT",
        "ALTER TABLE partido ADD COLUMN IF NOT EXISTS local_rojas         INT",
        "ALTER TABLE partido ADD COLUMN IF NOT EXISTS visitante_rojas     INT",
        "ALTER TABLE participacion ADD COLUMN IF NOT EXISTS fair_play_pts INT DEFAULT 0",
    ]:
        try:
            await db.execute(text(_col))
        except Exception:
            pass
    await db.commit()

    # Partidos de grupo finalizados con api_fixture_id
    r_torneo = await db.execute(
        text("""
            SELECT t.api_season, c.api_league_id
            FROM torneo t
            LEFT JOIN competicion c ON c.id = t.competicion_id
            WHERE t.id = :tid
        """),
        {"tid": torneo_id},
    )
    row_t = r_torneo.mappings().first()
    if not row_t:
        raise HTTPException(404, f"Torneo {torneo_id} no encontrado")

    r_parts = await db.execute(
        text("""
            SELECT p.id AS partido_id, p.api_fixture_id,
                   p.equipo_local_id, p.equipo_visitante_id,
                   COALESCE(el.api_team_id, 0) AS local_api_id,
                   COALESCE(ev.api_team_id, 0) AS visitante_api_id
            FROM partido p
            JOIN fase f ON f.id = p.fase_id AND f.tipo = 'grupo' AND f.nombre NOT ILIKE '%mejores%'
            LEFT JOIN equipo el ON el.id = p.equipo_local_id
            LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
            WHERE p.torneo_id = :tid
              AND p.estado = 'finalizado'
              AND p.api_fixture_id IS NOT NULL
            ORDER BY p.numero_fifa ASC NULLS LAST
            LIMIT :lim
        """),
        {"tid": torneo_id, "lim": max_partidos},
    )
    partidos = [dict(r) for r in r_parts.mappings()]

    if not partidos:
        return {"ok": False, "msg": "Sin partidos de grupo finalizados con api_fixture_id mapeado"}

    from app.services.sync_api_football import _headers, API_BASE
    import httpx, asyncio as _aio

    actualizados = 0
    errores: list[dict] = []
    api_calls = 0

    async with httpx.AsyncClient(timeout=30) as client:
        async def _fetch_events(fix_id: int):
            try:
                r = await client.get(
                    f"{API_BASE}/fixtures",
                    params={"id": fix_id},
                    headers=_headers(),
                )
                r.raise_for_status()
                return fix_id, r.json().get("response", [])
            except Exception as e:
                return fix_id, None

        # Procesar en lotes de 10 para no saturar la API en una sola ráfaga
        results = []
        batch_size = 10
        for i in range(0, len(partidos), batch_size):
            batch = partidos[i:i + batch_size]
            batch_results = await _aio.gather(*[_fetch_events(p["api_fixture_id"]) for p in batch])
            results.extend(batch_results)
        api_calls = len(partidos)

    # Procesar resultados y actualizar tarjetas por equipo
    fix_to_partido = {p["api_fixture_id"]: p for p in partidos}
    for fix_id, response in results:
        if not response:
            errores.append({"api_fixture_id": fix_id, "error": "sin respuesta"})
            continue
        fix = response[0]
        p = fix_to_partido[fix_id]
        local_api_id    = p["local_api_id"]
        visitante_api_id = p["visitante_api_id"]

        loc_amar = vis_amar = loc_rojas = vis_rojas = 0
        for ev in fix.get("events", []):
            if ev.get("type") != "Card":
                continue
            detail   = ev.get("detail", "")
            team_id  = ev.get("team", {}).get("id")
            is_local = team_id == local_api_id

            if detail == "Yellow Card":
                if is_local:    loc_amar += 1
                elif team_id == visitante_api_id: vis_amar += 1
            elif detail in ("Red Card", "Second Yellow card"):
                if is_local:    loc_rojas += 1
                elif team_id == visitante_api_id: vis_rojas += 1

        try:
            await db.execute(
                text("""
                    UPDATE partido
                    SET local_amarillas     = :la,
                        visitante_amarillas = :va,
                        local_rojas         = :lr,
                        visitante_rojas     = :vr
                    WHERE id = :pid
                """),
                {"la": loc_amar, "va": vis_amar, "lr": loc_rojas, "vr": vis_rojas,
                 "pid": p["partido_id"]},
            )
            await db.commit()
            actualizados += 1
        except Exception as e:
            await db.rollback()
            errores.append({"partido_id": p["partido_id"], "error": str(e)})

    # Recalcular standings + fair_play_pts en participacion
    fp_ok = False
    try:
        await _recalc_participacion(db, torneo_id)
        await db.commit()
        fp_ok = True
    except Exception as e:
        await db.rollback()
        errores.append({"recalc": str(e)})

    return {
        "ok": True,
        "partidos_procesados": len(partidos),
        "actualizados": actualizados,
        "api_calls": api_calls,
        "fair_play_recalculado": fp_ok,
        "errores": errores,
        "msg": f"Fair play recalculado para {actualizados}/{len(partidos)} partidos de grupo",
    }


# ── Ranking fair play mejores terceros ───────────────────────────────────────

@router.get("/fair-play-terceros/{torneo_id}", summary="Ranking FIFA de los 12 terceros por fair play (admin)")
async def fair_play_terceros(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """
    Devuelve los 12 terceros (uno por grupo) rankeados con el criterio FIFA 2026:
      1. Puntos  2. DG  3. GF  4. Fair Play (amarillas×1 + rojas×3, menor=mejor)
      5. FIFA Ranking  6. Nombre grupo

    Los primeros 8 clasifican a Ronda 32.
    Requiere haber corrido POST /recalc-fair-play/{torneo_id} para tener tarjetas
    por equipo pobladas desde API-Football.
    """
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    standings = await _calc_standings_reales(db, torneo_id)
    if not standings:
        raise HTTPException(404, "Sin grupos disponibles para este torneo")

    mejores, eliminados = seleccionar_mejores_terceros(standings, fill_incomplete=False)

    def _fmt(eq: dict, clasifica: bool, pos: int) -> dict:
        fp = eq.get("fair_play_pts", 0)
        return {
            "pos":             pos,
            "clasifica":       clasifica,
            "grupo":           eq.get("grupo", "?"),
            "equipo":          eq.get("nombre", "?"),
            "pts":             eq.get("pts", 0),
            "dg":              eq.get("gd", 0),
            "gf":              eq.get("gf", 0),
            "pj":              eq.get("pj", 0),
            "pg":              eq.get("pg", 0),
            "pe":              eq.get("pe", 0),
            "pp":              eq.get("pp", 0),
            "fair_play_pts":   fp,
            "amarillas_acum":  eq.get("amarillas_acum"),   # solo si disponible
            "rojas_acum":      eq.get("rojas_acum"),
            "fifa_ranking":    eq.get("fifa_ranking"),
            # Desglose del criterio FIFA para transparencia
            "criterio": {
                "1_pts":        eq.get("pts", 0),
                "2_dg":         eq.get("gd", 0),
                "3_gf":         eq.get("gf", 0),
                "4_fair_play":  fp,
                "5_fifa_rank":  eq.get("fifa_ranking") or 999,
                "6_grupo":      eq.get("grupo", "?"),
            },
        }

    ranking = (
        [_fmt(e, True, i + 1) for i, e in enumerate(mejores)] +
        [_fmt(e, False, i + 9) for i, e in enumerate(eliminados)]
    )

    # Estadísticas agregadas de tarjetas para diagnóstico
    total_fp_data = sum(1 for e in mejores + eliminados
                        if e.get("fair_play_pts", 0) > 0)

    return {
        "torneo_id": torneo_id,
        "clasifican": [r["equipo"] for r in ranking if r["clasifica"]],
        "eliminados": [r["equipo"] for r in ranking if not r["clasifica"]],
        "ranking":    ranking,
        "datos_tarjetas_ok": total_fp_data,
        "aviso": (
            "⚠️ fair_play_pts = 0 para todos — correr POST /recalc-fair-play para poblar tarjetas"
            if total_fp_data == 0 else None
        ),
    }


# ── Fix stats desde Excel (temporal) ─────────────────────────────────────────

@router.post("/fix-stats-excel/{torneo_id}", summary="Aplica correcciones de minuto_gol y var desde Excel oficial (admin)")
async def fix_stats_excel(torneo_id: int, current: CurrentUser, db: DBSession) -> dict:
    """Endpoint temporal: actualiza minuto_primer_gol y decisiones_var para los
    44 partidos de grupo confirmados en el Excel oficial 20260623check.xlsx."""
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    # decisiones_var fixes (5 partidos)
    var_fixes = [
        (146, 2), (150, 0), (148, 0), (169, 2), (180, 1),
    ]
    # minuto_primer_gol updates (44 partidos; None = sin primer gol / 0-0)
    minuto_fixes = [
        (143,10),(144,59),(145,21),(146,7),(149,28),(150,27),(148,21),(147,17),
        (153,90),(151,6),(152,51),(154,7),(157,41),(158,7),(156,20),(159,66),
        (160,29),(161,17),(162,21),(165,95),(164,12),(163,6),(166,40),(167,6),
        (168,74),(169,16),(170,50),(173,23),(172,2),(174,2),(171,11),(176,30),
        (177,None),(175,5),(178,4),(181,21),(180,None),(182,15),(185,43),(184,14),
        (183,38),(186,36),(155,None),(179,10),
    ]

    var_ok = 0
    for pid, val in var_fixes:
        await db.execute(text("UPDATE partido SET decisiones_var = :v WHERE id = :id"),
                         {"v": val, "id": pid})
        var_ok += 1

    min_ok = 0
    for pid, val in minuto_fixes:
        await db.execute(text("UPDATE partido SET minuto_primer_gol = :v WHERE id = :id"),
                         {"v": val, "id": pid})
        min_ok += 1

    await db.commit()

    # Recalcular puntajes
    from app.services.scoring import registry as scoring_registry
    from app.services.scoring.calculator import ScoringCalculator
    comp_r = await db.execute(
        text("SELECT c.codigo FROM torneo t JOIN competicion c ON c.id=t.competicion_id WHERE t.id=:tid"),
        {"tid": torneo_id})
    comp_row = comp_r.one_or_none()
    engine = scoring_registry.get_engine(comp_row[0] if comp_row else None)
    calc_result = await ScoringCalculator(db).calculate(torneo_id, engine)
    await ScoringCalculator(db).calculate_global(torneo_id, engine)
    await db.commit()

    return {
        "ok": True,
        "var_actualizados": var_ok,
        "minuto_actualizados": min_ok,
        "partidos_procesados": calc_result.get("partidos_procesados", 0),
        "plenos": calc_result.get("plenos", 0),
        "aciertos": calc_result.get("aciertos", 0),
    }


# ── Mensajes del administrador ───────────────────────────────────────────────

@router.get("/mensajes/{torneo_id}", summary="Mensajes del administrador para los apostadores")
async def get_mensajes(torneo_id: int, db: DBSession) -> list[dict]:
    """Devuelve los mensajes activos del torneo ordenados del más reciente al más antiguo."""
    try:
        r = await db.execute(
            text("""
                SELECT id, numero, titulo, contenido, autor_nombre,
                       created_at AT TIME ZONE 'UTC' AS created_at
                FROM mensaje_admin
                WHERE torneo_id = :tid AND es_activo = TRUE
                ORDER BY numero DESC
            """),
            {"tid": torneo_id},
        )
        rows = r.fetchall()
    except Exception:
        # Si la tabla no existe (migración pendiente) retorna lista vacía
        return []

    return [
        {
            "id":           row.id,
            "numero":       row.numero,
            "titulo":       row.titulo,
            "contenido":    row.contenido,
            "autor_nombre": row.autor_nombre or "Admin",
            "created_at":   row.created_at.isoformat() if row.created_at else None,
        }
        for row in rows
    ]


@router.post("/mensajes/{torneo_id}", summary="Crear mensaje (solo admin)")
async def create_mensaje(
    torneo_id: int,
    body: MensajeIn,
    db: DBSession,
    current_user: CurrentUser,
) -> dict:
    """Crea un mensaje nuevo. Solo accesible para admin y superadmin."""
    # Verificar rol admin
    async with _app_engine.connect() as app_db:
        r = await app_db.execute(
            text("""
                SELECT r.name FROM user_roles ur
                JOIN roles r ON r.id = ur.role_id
                WHERE ur.user_id = :uid AND r.name IN ('admin','superadmin')
                LIMIT 1
            """),
            {"uid": current_user.id},
        )
        if not r.one_or_none():
            raise HTTPException(403, "Solo administradores pueden crear mensajes")

        # Nombre del autor
        ur = await app_db.execute(
            text("SELECT username, email FROM users WHERE id = :uid"),
            {"uid": current_user.id},
        )
        user_row = ur.one_or_none()
        autor_nombre = (user_row.username or user_row.email) if user_row else "Admin"

    # Siguiente número secuencial para el torneo
    r_num = await db.execute(
        text("SELECT COALESCE(MAX(numero), 0) + 1 FROM mensaje_admin WHERE torneo_id = :tid"),
        {"tid": torneo_id},
    )
    next_num = r_num.scalar() or 1

    r_ins = await db.execute(
        text("""
            INSERT INTO mensaje_admin (numero, torneo_id, titulo, contenido, autor_id, autor_nombre)
            VALUES (:num, :tid, :titulo, :contenido, :uid, :autor)
            RETURNING id, numero, titulo, contenido, autor_nombre,
                      created_at AT TIME ZONE 'UTC' AS created_at
        """),
        {
            "num":      next_num,
            "tid":      torneo_id,
            "titulo":   body.titulo.strip(),
            "contenido": body.contenido.strip(),
            "uid":      current_user.id,
            "autor":    autor_nombre,
        },
    )
    await db.commit()
    row = r_ins.one()
    return {
        "id":           row.id,
        "numero":       row.numero,
        "titulo":       row.titulo,
        "contenido":    row.contenido,
        "autor_nombre": row.autor_nombre,
        "created_at":   row.created_at.isoformat() if row.created_at else None,
    }


@router.delete("/mensajes/{torneo_id}/{mensaje_id}", summary="Eliminar mensaje (solo admin)")
async def delete_mensaje(
    torneo_id: int,
    mensaje_id: int,
    db: DBSession,
    current_user: CurrentUser,
) -> dict:
    """Desactiva un mensaje (soft delete). Solo admin / superadmin."""
    async with _app_engine.connect() as app_db:
        r = await app_db.execute(
            text("""
                SELECT r.name FROM user_roles ur
                JOIN roles r ON r.id = ur.role_id
                WHERE ur.user_id = :uid AND r.name IN ('admin','superadmin')
                LIMIT 1
            """),
            {"uid": current_user.id},
        )
        if not r.one_or_none():
            raise HTTPException(403, "Solo administradores pueden eliminar mensajes")

    await db.execute(
        text("""
            UPDATE mensaje_admin
            SET es_activo = FALSE
            WHERE id = :mid AND torneo_id = :tid
        """),
        {"mid": mensaje_id, "tid": torneo_id},
    )
    await db.commit()
    return {"ok": True, "id": mensaje_id}


# ─────────────────────────────────────────────────────────────────────────────
# IMPORTAR APUESTAS — FASE DE GRUPOS
# POST /bets/importar-apuestas-grupos/{torneo_id}
# ─────────────────────────────────────────────────────────────────────────────

class ImportRowIn(BaseModel):
    apostador:            str
    nombre:               str | None = None
    email:                str | None = None
    telefono:             str | None = None
    partido_num:          int | None = None
    equipo_local:         str | None = None
    equipo_visitante:     str | None = None
    goles_local:          int
    goles_visitante:      int
    pred_minuto_gol:      int | None = None
    pred_amarillas:       int | None = None
    pred_var:             int | None = None
    pred_rojas:           int | None = None
    pred_penales_partido: int | None = None




@router.post("/equipo-alias", summary="Guardar alias de equipo en BD (admin)")
async def guardar_equipo_alias(
    equipo_id: int,
    nombre_es: str,
    current: CurrentUser,
    db: DBSession,
) -> dict:
    """
    Actualiza equipo.nombre_es con el alias en español.
    Después de esto el import lo resuelve directamente sin necesidad de _EQ_ALIASES.
    """
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")
    r = await db.execute(
        text("UPDATE equipo SET nombre_es = :alias WHERE id = :eid RETURNING id, nombre, nombre_es"),
        {"alias": nombre_es.strip(), "eid": equipo_id},
    )
    row = r.one_or_none()
    if not row:
        raise HTTPException(404, f"Equipo {equipo_id} no encontrado")
    await db.commit()
    return {"ok": True, "equipo_id": row[0], "nombre": row[1], "nombre_es": row[2]}

@router.post("/importar-apuestas-grupos/{torneo_id}")
async def importar_apuestas_grupos(
    torneo_id: int,
    rows: list[ImportRowIn],
    current: CurrentUser,
    db: DBSession,
    crear_usuarios: bool = False,
):
    """Importa apuestas de fase de grupos desde Excel (admin only)."""
    # Solo admin
    # Admin check usa app_db (user_roles/roles NO existen en becbuc)
    if not await _check_admin(current):
        raise HTTPException(403, "Solo administradores pueden importar apuestas")

    # Verificar torneo
    torneo_r = await db.execute(text("SELECT id FROM torneo WHERE id=:id"), {"id": torneo_id})
    if not torneo_r.one_or_none():
        raise HTTPException(404, f"Torneo {torneo_id} no encontrado")

    # Pre-cargar partidos de fase grupos
    import unicodedata as _ucd

    def _norm_eq(s: str) -> str:
        """Normaliza nombre de equipo: sin tildes, lowercase, sin puntuación extra."""
        s = (s or "").lower().strip()
        s = _ucd.normalize("NFD", s)
        return "".join(c for c in s if _ucd.category(c) != "Mn")

    # Aliases: nombre_normalizado_excel → nombre_normalizado_bd
    # Cubre nombres en español del Excel vs nombres en inglés de la BD
    # Aliases: nombre_normalizado_excel → nombre_normalizado_bd (nombre EN en BD).
    # Las comparaciones son case-insensitive porque _norm_eq() lowercasea todo.
    _EQ_ALIASES: dict[str, str] = {
        # Congo
        "congo":                    "congo dr",
        "rep congo":                "congo dr",
        "rd congo":                 "congo dr",
        "republica del congo":      "congo dr",
        # Turquía
        "turquia":                  "turkiye",
        "turkey":                   "turkiye",
        # Costa de Marfil — nombre en BD: "Ivory Coast"
        "costa marfil":             "ivory coast",
        "costa de marfil":          "ivory coast",
        "cote d'ivoire":            "ivory coast",
        "cote divoire":             "ivory coast",
        # Países en español → nombre EN en BD
        "noruega":                  "norway",
        "croacia":                  "croatia",
        "argelia":                  "algeria",
        "corea del sur":            "south korea",
        "corea":                    "south korea",
        "corea sur":                "south korea",
        "korea republic":           "south korea",
        "republic of korea":        "south korea",
        "corea del norte":          "korea dpr",
        "chequia":                  "czech republic",
        "rep checa":                "czech republic",
        "sudafrica":                "south africa",
        "paises bajos":             "netherlands",
        "holanda":                  "netherlands",
        "nueva zelanda":            "new zealand",
        "estados unidos":           "united states",
        "usa":                      "united states",
        "arabia saudita":           "saudi arabia",
        "japon":                    "japan",
        "belgica":                  "belgium",
        "dinamarca":                "denmark",
        "polonia":                  "poland",
        "suecia":                   "sweden",
        "suiza":                    "switzerland",
        "egipto":                   "egypt",
        "marruecos":                "morocco",
        "iran":                     "ir iran",
        "irak":                     "iraq",
        "jordania":                 "jordan",
        "curazao":                  "curacao",
        "cabo verde":               "cape verde islands",
        "cabo verde islands":       "cape verde islands",
        "haiti":                    "haiti",
        "escocia":                  "scotland",
        "gales":                    "wales",
        "uzbekistan":               "uzbekistan",
        "catar":                    "qatar",
        "ghana":                    "ghana",
        "senegal":                  "senegal",
        # Sesion 29+
        "brasil":                   "brazil",
        "alemania":                 "germany",
        "tunez":                    "tunisia",
        "francia":                  "france",
        "inglaterra":               "england",
        "espana":                   "spain",
        "españa":                   "spain",
        # Bosnia
        "bosnia":                   "bosnia & herzegovina",
        "bosnia herzegovina":       "bosnia & herzegovina",
        "bosnia y herzegovina":     "bosnia & herzegovina",
        "bosnia i hercegovina":     "bosnia & herzegovina",
    }

    # ── 1. Índice de equipos: nombre_normalizado → equipo_id ─────────────────
    # Indexa nombre (EN) y nombre_es (ES) + todos los aliases conocidos.
    eq_r = await db.execute(
        text("""
            SELECT DISTINCT e.id, e.nombre, e.nombre_es
            FROM equipo e
            WHERE e.id IN (
                SELECT equipo_local_id    FROM partido WHERE torneo_id = :tid AND equipo_local_id    IS NOT NULL
                UNION
                SELECT equipo_visitante_id FROM partido WHERE torneo_id = :tid AND equipo_visitante_id IS NOT NULL
            )
        """),
        {"tid": torneo_id},
    )
    equipo_name_map: dict[str, int] = {}  # nombre_norm → equipo_id
    eq_list = eq_r.mappings().all()
    for eq in eq_list:
        for col in (eq["nombre"], eq["nombre_es"]):
            if col:
                equipo_name_map[_norm_eq(col)] = eq["id"]

    # Resolver aliases: si el alias apunta a un nombre que existe en el mapa, agregar
    for alias_key, target_raw in _EQ_ALIASES.items():
        # Normalizar alias_key y target para comparación case-insensitive
        alias_norm = _norm_eq(alias_key)
        target_norm = _norm_eq(target_raw)
        if target_norm in equipo_name_map and alias_norm not in equipo_name_map:
            equipo_name_map[alias_norm] = equipo_name_map[target_norm]

    # Fallback por prefijo: "congo" → encuentra "congo dr" si no hubo match exacto
    _eq_norm_keys = list(equipo_name_map.keys())

    # Índice inverso equipo_id → nombre legible (para sugerencias)
    _eq_id_to_nombre: dict[int, str] = {
        eq["id"]: eq["nombre"] for eq in eq_list if eq["nombre"]
    }

    import difflib as _difflib

    def _sugerir_equipo(nombre_excel: str) -> list[dict]:
        """Busca los nombres más cercanos en equipo_name_map para un nombre no resuelto."""
        n = _norm_eq(nombre_excel)
        if not n:
            return []
        matches = _difflib.get_close_matches(n, _eq_norm_keys, n=3, cutoff=0.5)
        result = []
        for m in matches:
            eid = equipo_name_map[m]
            result.append({
                "nombre_bd": _eq_id_to_nombre.get(eid, m),
                "nombre_norm_bd": m,
                "equipo_id": eid,
            })
        return result

    # Colectar nombres de equipos del Excel que no resuelven (para reportar al final)
    _unresolved_eq: dict[str, dict] = {}  # nombre_excel → {count, sugerencias}

    # ── 2. Partidos de grupos: (equipo_local_id, equipo_visitante_id) → partido_id ──
    partidos_r = await db.execute(
        text("""
            SELECT p.id,
                   COALESCE(p.numero_fifa, 0)                       AS numero_fifa,
                   ROW_NUMBER() OVER (ORDER BY f.orden, p.id)::int  AS num_seq,
                   p.equipo_local_id,
                   p.equipo_visitante_id
            FROM partido p
            JOIN fase f ON f.id = p.fase_id
            WHERE p.torneo_id = :tid AND f.tipo = 'grupo'
            ORDER BY f.orden, p.id
        """),
        {"tid": torneo_id},
    )
    partido_rows = partidos_r.all()
    # Índice principal: numero_fifa → partido_id (referencia oficial FIFA P1-P104)
    partido_map_fifa: dict[int, int] = {
        row.numero_fifa: row.id
        for row in partido_rows
        if row.numero_fifa
    }
    # Fallback legado: num_seq → partido_id (orden BD)
    partido_map_seq: dict[int, int] = {row.num_seq: row.id for row in partido_rows}
    # Inverso: partido_id → numero_fifa (para grabar en apuesta al insertar)
    partido_map_fifa_inv: dict[int, int] = {
        row.id: row.numero_fifa
        for row in partido_rows
        if row.numero_fifa
    }
    # Índice por par de equipo_id (confiable)
    partido_map_ids: dict[tuple[int, int], int] = {
        (row.equipo_local_id, row.equipo_visitante_id): row.id
        for row in partido_rows
        if row.equipo_local_id and row.equipo_visitante_id
    }
    total_partidos_grupos = len(partido_map_seq)

    def _resolve_equipo(name: str) -> int | None:
        """Resuelve equipo_id desde nombre del Excel.
        1. Match exacto (cubre nombre EN, nombre_es ES y aliases).
        2. Fallback: busca claves de BD que empiecen con el nombre del Excel
           (ej. 'congo' matchea 'congo dr').
        Retorna None y registra en _unresolved_eq si no encuentra.
        """
        n = _norm_eq(name)
        if not n:
            return None
        # 1. Exacto
        eid = equipo_name_map.get(n)
        if eid:
            return eid
        # 2. Prefijo: n es prefijo de alguna clave conocida
        for key in _eq_norm_keys:
            if key.startswith(n + " ") or key == n:
                return equipo_name_map[key]
        # No resuelto: registrar para el reporte
        nombre_orig = (name or "").strip()
        if nombre_orig and nombre_orig not in _unresolved_eq:
            _unresolved_eq[nombre_orig] = {
                "nombre_excel": nombre_orig,
                "nombre_norm": n,
                "count": 0,
                "sugerencias": _sugerir_equipo(nombre_orig),
            }
        if nombre_orig:
            _unresolved_eq[nombre_orig]["count"] += 1
        return None

    def _find_partido(row: "ImportRowIn") -> tuple[int | None, str]:
        """Resuelve partido_id desde la fila del Excel.

        Prioridad:
          1. Inferir por nombres de equipo local y visitante → partido_map_ids
             (más confiable: usa los IDs reales de la BD, garantiza numero_fifa correcto)
          2. partido_num del Excel → numero_fifa en BD (número oficial FIFA P1-P104)
          3. partido_num del Excel → num_seq en BD (fallback legado — evitar)

        Returns: (partido_id | None, método: "equipo" | "numero_fifa" | "num_seq" | "none")
        """
        # 1. PRIORIDAD: inferir por nombres de equipo (PRIMARY — más robusto)
        if row.equipo_local and row.equipo_visitante:
            lid = _resolve_equipo(row.equipo_local)
            vid = _resolve_equipo(row.equipo_visitante)
            if lid and vid:
                pid = partido_map_ids.get((lid, vid))
                if pid:
                    return pid, "equipo"
        # 2. partido_num del Excel → numero_fifa en BD
        if row.partido_num:
            pid = partido_map_fifa.get(int(row.partido_num))
            if pid:
                return pid, "numero_fifa"
        # 3. FALLBACK LEGADO: número de partido del Excel → num_seq en BD
        if row.partido_num:
            pid = partido_map_seq.get(int(row.partido_num))
            if pid:
                return pid, "num_seq"
        return None, "none"

    # Pre-cargar usuarios existentes (username -> id, nombre) — desde app_db
    async with _app_engine.connect() as aconn:
        users_r = await aconn.execute(
            text("SELECT id, username, email, COALESCE(nombre, username) AS nombre FROM users WHERE is_active = TRUE")
        )
        user_rows = users_r.all()
        # strip + lower en username para comparacion robusta (espacios, \xa0, mayusculas)
        user_map = {row.username.strip().replace("\xa0", "").lower(): row.id for row in user_rows}
        # mapa id → nombre para usar en INSERT
        user_nombre_map: dict[int, str] = {row.id: row.nombre for row in user_rows}

    def _norm_alias(s: str) -> str:
        """Normaliza alias para comparacion: sin espacios extremos, sin \xa0, lowercase."""
        return (s or "").replace("\xa0", "").strip().lower()

    errores = []
    # Filas donde la inferencia de equipos no fue posible (resueltas por fallback o sin resolver)
    sin_inferencia: list[dict] = []
    skipped_ko = 0  # filas de P073+ ignoradas silenciosamente

    # ── PASO 1: Pre-resolver apostadores y partidos (sin tocar BD aún) ────────
    resolved: list[tuple[int, int, "ImportRowIn", str]] = []  # (partido_id, user_id, row, resolucion)
    # key = (user_id, partido_id) → índice en resolved; para deduplicar filas repetidas
    _resolved_key: dict[tuple[int, int], int] = {}

    for i, row in enumerate(rows, start=2):
        fila = i

        # Filtro: solo partidos de fase grupos (P001-P072).
        # Filas con partido_num > total_partidos_grupos son KO y se omiten silenciosamente.
        if row.partido_num is not None:
            try:
                _pnum = int(row.partido_num)
            except (ValueError, TypeError):
                _pnum = 0
            if _pnum > total_partidos_grupos:
                skipped_ko += 1
                continue

        # Resolver partido: por nombres de equipo (preferido) o por num_seq (fallback)
        partido_id, resolucion = _find_partido(row)
        if not partido_id:
            desc = f"equipo='{row.equipo_local} vs {row.equipo_visitante}'" if row.equipo_local else f"num={row.partido_num}"
            errores.append({
                "fila": fila,
                "motivo": f"Partido no encontrado ({desc}) en fase grupos del torneo {torneo_id} "
                          f"(hay {total_partidos_grupos} partidos)",
                "sin_mapeo": True,
            })
            continue

        # Resolver apostador — comparacion normalizada
        alias = _norm_alias(row.apostador)
        user_id = user_map.get(alias)

        if not user_id:
            if row.email:
                async with _app_engine.connect() as aconn:
                    eu = await aconn.execute(
                        text("SELECT id FROM users WHERE email=:e"), {"e": row.email.strip()}
                    )
                    eu_row = eu.one_or_none()
                if eu_row:
                    user_id = eu_row[0]
                    user_map[alias] = user_id

            if not user_id:
                if not crear_usuarios:
                    errores.append({
                        "fila": fila,
                        "motivo": f"Usuario '{row.apostador}' no existe en el sistema",
                        "usuario_faltante": True,
                        "apostador": row.apostador.strip(),
                        "email": row.email or "",
                        "nombre": row.nombre or "",
                        "telefono": row.telefono or "",
                    })
                    continue

                if not row.email:
                    errores.append({"fila": fila, "motivo": f"Usuario '{row.apostador}' no existe y no se proporcionó email para crearlo"})
                    continue
                pwd = "mundial2026"
                from app.core.security import hash_password as _hp
                async with _app_engine.begin() as aconn:
                    new_u = await aconn.execute(
                        text("""
                            INSERT INTO users (username, email, nombre, telefono, password_hash, is_active, must_change_password)
                            VALUES (:u, :e, :n, :t, :ph, TRUE, TRUE)
                            ON CONFLICT (username) DO UPDATE SET email=EXCLUDED.email
                            RETURNING id
                        """),
                        {
                            "u": row.apostador.strip(),
                            "e": row.email.strip(),
                            "n": row.nombre,
                            "t": row.telefono,
                            "ph": _hp(pwd),
                        },
                    )
                    user_id = new_u.scalar_one()
                user_map[alias] = user_id

        # Registrar si la resolución no fue por inferencia de equipos
        if resolucion != "equipo":
            sin_inferencia.append({
                "fila": fila,
                "apostador": row.apostador.strip() if row.apostador else "",
                "partido_num": row.partido_num,
                "equipo_local": row.equipo_local or "",
                "equipo_visitante": row.equipo_visitante or "",
                "resolucion": resolucion,
                "partido_id": partido_id,
            })

        key = (user_id, partido_id)
        if key in _resolved_key:
            # Reemplazar fila anterior con la más reciente (deduplicar)
            resolved[_resolved_key[key]] = (partido_id, user_id, row, resolucion)
        else:
            _resolved_key[key] = len(resolved)
            resolved.append((partido_id, user_id, row, resolucion))

    # ── PASO 2: DELETE masivo — borra TODAS las apuestas de la fase grupos
    #            del torneo completo (sin filtrar por apostador)
    #            Garantiza tabla limpia antes del INSERT fresco ───────────────────
    all_partido_ids  = list(partido_map_seq.values())

    # Borrar TODAS las apuestas de fase grupos del torneo (todos los apostadores)
    eliminadas = 0
    if all_partido_ids:
        ids_p = ",".join(str(x) for x in all_partido_ids)
        del_r = await db.execute(
            text(f"DELETE FROM apuesta WHERE partido_id IN ({ids_p})")
        )
        eliminadas = del_r.rowcount

    # ── PASO 3: INSERT limpio (sin ON CONFLICT — el DELETE garantiza tabla limpia) ──
    creadas = 0
    for partido_id, user_id, row, resolucion in resolved:
        # Nombre: del Excel si viene, sino del mapa de usuarios de app_db
        nombre_apost = (row.nombre or "").strip() or user_nombre_map.get(user_id) or row.apostador.strip()
        # numero_fifa SOLO cuando la resolución fue por inferencia de equipos:
        # partidos.equipo_local_id = lid AND partidos.equipo_visitante_id = vid
        # → el partido_id que viene de partido_map_ids[(lid, vid)] garantiza la correspondencia.
        # Para rutas "numero_fifa" o "num_seq" no se confirma por equipos → NULL.
        _nfifa = partido_map_fifa_inv.get(partido_id) if resolucion == "equipo" else None
        await db.execute(
            text("""
                INSERT INTO apuesta (apostador_id, partido_id, nombre_apostador, id_partido_ok,
                    numero_fifa,
                    equipo_local_excel, equipo_visitante_excel,
                    pred_local, pred_visitante, pred_minuto_gol,
                    pred_amarillas, pred_var, pred_rojas, pred_penales_partido)
                VALUES (:uid, :pid, :nombre, :pid_ok, :nfifa, :eq_l, :eq_v, :l, :v, :mg, :am, :var, :ro, :pp)
            """),
            {
                "uid": user_id, "pid": partido_id, "nombre": nombre_apost, "pid_ok": partido_id,
                "nfifa": _nfifa,
                "eq_l": row.equipo_local or None,
                "eq_v": row.equipo_visitante or None,
                "l": row.goles_local, "v": row.goles_visitante,
                "mg": row.pred_minuto_gol, "am": row.pred_amarillas,
                "var": row.pred_var, "ro": row.pred_rojas,
                "pp": row.pred_penales_partido,
            },
        )
        creadas += 1

    # ── PASO 3b: Completar partidos faltantes con 0-0 ────────────────────────
    all_partido_ids_set = set(partido_map_seq.values())
    seq_of_pid: dict[int, int] = {v: k for k, v in partido_map_seq.items()}

    # Partidos cubiertos por apostador (desde resolved)
    covered_by_user: dict[int, set[int]] = {}
    for pid, uid, _, _res in resolved:
        covered_by_user.setdefault(uid, set()).add(pid)

    completados_0_0 = 0
    faltantes_por_apostador: list[dict] = []

    for uid, covered in covered_by_user.items():
        missing = sorted(all_partido_ids_set - covered, key=lambda p: seq_of_pid.get(p, 9999))
        if not missing:
            continue
        nombre_apost = user_nombre_map.get(uid) or ""
        _completados_this = []
        for pid in missing:
            _nfifa = partido_map_fifa_inv.get(pid)
            await db.execute(
                text("""
                    INSERT INTO apuesta (apostador_id, partido_id, nombre_apostador, id_partido_ok,
                        numero_fifa, pred_local, pred_visitante)
                    VALUES (:uid, :pid, :nombre, :pid_ok, :nfifa, 0, 0)
                """),
                {"uid": uid, "pid": pid, "nombre": nombre_apost, "pid_ok": pid, "nfifa": _nfifa},
            )
            completados_0_0 += 1
            creadas += 1
            _completados_this.append({
                "partido_id": pid,
                "num_seq": seq_of_pid.get(pid),
            })
        faltantes_por_apostador.append({
            "apostador_id": uid,
            "completados_0_0": len(_completados_this),
            "partidos": _completados_this,
        })

    await db.commit()

    # ── PASO 3c: Contar registros reales en BD y verificar cobertura ──────────
    num_apostadores = len(covered_by_user)
    esperados = num_apostadores * total_partidos_grupos
    count_r = await db.execute(
        text(f"SELECT COUNT(*) FROM apuesta WHERE partido_id IN ({','.join(str(x) for x in all_partido_ids)})")
    )
    creadas_real = count_r.scalar() or 0
    cobertura_ok = (creadas_real == esperados)
    cobertura_advertencia = (
        f"Se esperaban {esperados} registros ({num_apostadores} apostadores × {total_partidos_grupos} partidos) "
        f"pero la tabla tiene {creadas_real}. Diferencia: {creadas_real - esperados:+d}."
        if not cobertura_ok else None
    )

    # ── PASO 4: Recalcular puntajes automáticamente ───────────────────────────
    puntajes_ok = False
    puntajes_procesados = 0
    try:
        comp_r = await db.execute(
            text("""
                SELECT c.codigo FROM torneo t
                JOIN competicion c ON c.id = t.competicion_id
                WHERE t.id = :tid
            """),
            {"tid": torneo_id},
        )
        comp_row = comp_r.one_or_none()
        competicion_codigo = comp_row[0] if comp_row else None
        engine = scoring_registry.get_engine(competicion_codigo)
        calc = ScoringCalculator(db)
        result = await calc.calculate(torneo_id, engine)
        await calc.calculate_global(torneo_id, engine)
        puntajes_procesados = result.get("procesados", 0)
        puntajes_ok = True
    except Exception as _e:
        pass  # no bloquear el import si el cálculo falla

    # Resumen de métodos de resolución
    metodos = {}
    for _, _, _, res in resolved:
        metodos[res] = metodos.get(res, 0) + 1

    return {
        "ok": len(errores) == 0 and cobertura_ok,
        "eliminadas": eliminadas,
        "creadas": creadas_real,            # COUNT real post-commit
        "skipped_ko": skipped_ko,
        "completados_0_0": completados_0_0,
        "errores": errores,
        "total_partidos_grupos": total_partidos_grupos,
        "num_apostadores": num_apostadores,
        "esperados": esperados,
        "faltantes_por_apostador": faltantes_por_apostador,
        "cobertura_completa": cobertura_ok,
        "cobertura_advertencia": cobertura_advertencia,
        "puntajes_ok": puntajes_ok,
        "puntajes_procesados": puntajes_procesados,
        "resolucion_metodos": metodos,
        "sin_inferencia_equipos": sin_inferencia,
        "equipos_sin_resolver": sorted(_unresolved_eq.values(), key=lambda x: -x["count"]),
    }


# ── Sync histórico: importar todos los partidos jugados ──────────────────────

@router.post("/sync-historico/{torneo_id}", summary="Importar resultados históricos (admin)")
async def sync_historico(
    torneo_id: int,
    current: CurrentUser,
    db: DBSession,
    max_detalle: int = 50,
) -> dict:
    """
    Importa todos los resultados finalizados desde API-Football (partidos jugados
    hasta hoy). A diferencia del sync normal:

    - force=True siempre: re-procesa aunque esté marcado 'finalizado'.
    - max_detalle=50: más llamadas para traer estadísticas completas.
    - Incluye auto-mapeo si los api_fixture_id no están configurados.
    - Retorna lista detallada de cada partido importado para verificación.

    Cadena: auto-mapeo → sync force → standings → bracket → puntajes.
    """
    if not await _check_admin(current):
        raise HTTPException(403, "Se requiere rol admin o superadmin")

    from app.services.sync_api_football import sync_torneo

    try:
        sync_summary = await sync_torneo(db, torneo_id, force=True, max_detalle=max_detalle)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        await db.rollback()
        raise HTTPException(400, f"Error en sync histórico: {e}")

    await db.commit()

    actualizados = sync_summary.get("actualizados", 0)

    # Standings
    try:
        await _recalc_participacion(db, torneo_id)
        await db.commit()
    except Exception as e:
        await db.rollback()
        sync_summary["participacion_error"] = str(e)

    # Bracket
    bracket_ok = False
    try:
        maps = await ko_scoring.build_num_maps(db, torneo_id)
        await _avanzar_bracket(db, torneo_id, maps)
        await db.commit()
        bracket_ok = True
    except Exception as e:
        await db.rollback()
        sync_summary["bracket_error"] = str(e)

    # Puntajes
    puntajes_ok = False
    puntajes_summary: dict = {}
    try:
        await _ensure_detalle_table(db)
        r_torneo = await db.execute(
            text("""
                SELECT COALESCE(c.codigo, '') AS competicion_codigo
                FROM torneo t
                LEFT JOIN competicion c ON c.id = t.competicion_id
                WHERE t.id = :tid
            """),
            {"tid": torneo_id},
        )
        row_torneo = r_torneo.mappings().first()
        competicion_codigo = (row_torneo or {}).get("competicion_codigo") or None
        engine = scoring_registry.get_engine(competicion_codigo)
        calc = ScoringCalculator(db)
        result = await calc.calculate(torneo_id, engine)
        if result:
            global_result = await calc.calculate_global(torneo_id, engine)
            await db.commit()
            puntajes_ok = True
            puntajes_summary = {
                "plenos":              result["plenos"],
                "aciertos":            result["aciertos"],
                "fallos":              result["fallos"],
                "globales_procesadas": global_result.get("procesadas", 0),
            }
    except Exception as e:
        await db.rollback()
        sync_summary["puntajes_error"] = str(e)

    # Detalle de cada partido importado (para verificación en UI)
    partidos_importados: list[dict] = []
    ids_act = sync_summary.get("ids_actualizados") or []
    if ids_act:
        try:
            ids_sql = ",".join(str(i) for i in ids_act)
            r_p = await db.execute(
                text(f"""
                    SELECT p.id,
                           COALESCE(el.nombre_es, el.nombre, '?') AS local,
                           COALESCE(ev.nombre_es, ev.nombre, '?') AS visitante,
                           p.goles_local, p.goles_visitante, p.estado,
                           p.amarillas, p.rojas, p.decisiones_var,
                           p.penales_local, p.penales_visitante, p.minuto_primer_gol,
                           p.api_fixture_id
                    FROM partido p
                    LEFT JOIN equipo el ON el.id = p.equipo_local_id
                    LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
                    WHERE p.id IN ({ids_sql})
                    ORDER BY p.fecha
                """)
            )
            for row in r_p.mappings():
                gl = row["goles_local"]
                gv = row["goles_visitante"]
                pl = row["penales_local"]
                pv = row["penales_visitante"]
                partidos_importados.append({
                    "partido_id":     row["id"],
                    "nombre":         f"{row['local']} vs {row['visitante']}",
                    "resultado":      f"{gl if gl is not None else '?'}-{gv if gv is not None else '?'}",
                    "penales":        f"({pl}-{pv})" if pl is not None else None,
                    "estado":         row["estado"],
                    "amarillas":      row["amarillas"],
                    "rojas":          row["rojas"],
                    "var":            row["decisiones_var"],
                    "minuto_gol":     row["minuto_primer_gol"],
                    "api_fixture_id": row["api_fixture_id"],
                })
        except Exception as e:
            await db.rollback()
            sync_summary["partidos_importados_error"] = str(e)

    await _audit_log(
        "sync:historico", "bets",
        current=current, method="POST",
        path=f"/api/v1/bets/sync-historico/{torneo_id}",
        resource_id=str(torneo_id),
        details={
            "evento":       "importación histórica API-Football",
            "max_detalle":  max_detalle,
            "actualizados": actualizados,
            "api_calls":    sync_summary.get("api_calls", 0),
            "bracket_ok":   bracket_ok,
            "puntajes_ok":  puntajes_ok,
        },
    )

    return {
        "ok":                  True,
        "actualizados":        actualizados,
        "partidos_importados": partidos_importados,
          "sync":                sync_summary,
        "bracket_ok":          bracket_ok,
        "puntajes_ok":         puntajes_ok,
        "puntajes":            puntajes_summary,
    }


@router.get("/verificar-importacion/{torneo_id}", summary="Verificar datos importados (admin)")
async def verificar_importacion(
    torneo_id: int,
    current: CurrentUser,
    db: DBSession,
    limite: int = 100,
) -> dict:
    if not await _check_admin(current):
        raise HTTPException(status_code=403, detail="Solo admin")
    r = await db.execute(
        text("""
            SELECT p.id, p.api_fixture_id,
                COALESCE(el.nombre_es, el.nombre) AS local,
                COALESCE(ev.nombre_es, ev.nombre) AS visitante,
                p.goles_local, p.goles_visitante, p.estado,
                p.fecha AT TIME ZONE 'UTC' AS fecha,
                COALESCE(p.amarillas, 0) AS amarillas,
                COALESCE(p.rojas, 0) AS rojas,
                COALESCE(p.decisiones_var, 0) AS var_decisiones,
                p.penales_local, p.penales_visitante, p.penales_partido,
                p.minuto_primer_gol,
                (p.api_fixture_id IS NOT NULL) AS mapeado
            FROM partido p
            LEFT JOIN equipo el ON el.id = p.equipo_local_id
            LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
            WHERE p.torneo_id = :tid
              AND p.estado IN ('finalizado', 'en_juego')
            ORDER BY p.fecha DESC NULLS LAST
            LIMIT :lim
        """),
        {"tid": torneo_id, "lim": limite},
    )
    rows = r.mappings().fetchall()
    return {
        "ok": True,
        "partidos": [dict(row) for row in rows],
    }


# ── Live panel: próximo partido + predicciones (público con token) ────────────

@router.get("/live-panel/{torneo_id}", summary="Panel en vivo para becbuc-live.html")
async def live_panel(
    torneo_id: int,
    current: CurrentUser,
    db: DBSession,
    numero_fifa: int | None = None,
) -> dict:
    """
    Devuelve el próximo partido pendiente (o el indicado por numero_fifa) junto con
    las predicciones de todos los apostadores y sus puntos totales actuales.
    Las predicciones se buscan por numero_fifa en apuesta (robusto ante partido_id incorrecto).
    """
    # 1. Buscar partido: por numero_fifa si se provee, sino auto-detección
    _partido_sql = """
        SELECT
            p.id,
            COALESCE(p.numero_fifa, 0)            AS numero_fifa,
            COALESCE(el.nombre_es, el.nombre)    AS equipo_local,
            COALESCE(ev.nombre_es, ev.nombre)    AS equipo_visitante,
            el.codigo_iso                         AS bandera_local,
            ev.codigo_iso                         AS bandera_visitante,
            p.goles_local, p.goles_visitante,
            p.estado, p.fecha,
            p.minuto_actual, p.minuto_primer_gol,
            COALESCE(p.amarillas, 0)              AS amarillas,
            COALESCE(p.rojas, 0)                  AS rojas,
            COALESCE(p.decisiones_var, 0)         AS decisiones_var,
            p.penales_partido,
            p.penales_local                       AS penales_tanda_local,
            p.penales_visitante                   AS penales_tanda_visitante,
            p.equipo_clasificado_id,
            COALESCE(p.local_amarillas, 0)        AS local_amarillas,
            COALESCE(p.visitante_amarillas, 0)    AS visitante_amarillas,
            COALESCE(p.local_rojas, 0)            AS local_rojas,
            COALESCE(p.visitante_rojas, 0)        AS visitante_rojas,
            f.nombre                              AS fase_nombre,
            f.tipo                                AS fase_tipo,
            el.logo_url                           AS logo_local,
            ev.logo_url                           AS logo_visitante,
            p.eventos_api::text                   AS eventos_api_raw,
            (el.nombre ILIKE '%paraguay%'
             OR el.nombre_es ILIKE '%paraguay%'
             OR ev.nombre ILIKE '%paraguay%'
             OR ev.nombre_es ILIKE '%paraguay%')  AS es_paraguay
        FROM partido p
        JOIN fase f ON f.id = p.fase_id
        LEFT JOIN equipo el ON el.id = p.equipo_local_id
        LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
        WHERE p.torneo_id = :tid
    """

    if numero_fifa is not None:
        # Partido específico solicitado por numero_fifa
        partido_r = await db.execute(
            text(_partido_sql + " AND p.numero_fifa = :nfifa LIMIT 1"),
            {"tid": torneo_id, "nfifa": numero_fifa},
        )
        partido_row = partido_r.mappings().fetchone()
    else:
        # Auto: en_juego primero, luego el próximo pendiente/programado por fecha
        partido_r = await db.execute(
            text(_partido_sql + """
                  AND p.estado IN ('programado', 'pendiente', 'en_juego')
                ORDER BY
                    CASE p.estado WHEN 'en_juego' THEN 0 ELSE 1 END,
                    p.fecha ASC NULLS LAST
                LIMIT 1
            """),
            {"tid": torneo_id},
        )
        partido_row = partido_r.mappings().fetchone()
        if not partido_row:
            partido_r2 = await db.execute(
                text(_partido_sql + " AND p.estado = 'finalizado' ORDER BY p.fecha DESC NULLS LAST LIMIT 1"),
                {"tid": torneo_id},
            )
            partido_row = partido_r2.mappings().fetchone()

    if not partido_row:
        return {"partido": None, "apostadores": [], "numeros_fifa": [], "partidos_en_juego": []}

    partido      = dict(partido_row)
    # Serializar fecha con "Z" (UTC) para que JS calcule countdown correcto en CR (UTC-6)
    from datetime import datetime as _dtp
    if partido.get("fecha") and isinstance(partido["fecha"], _dtp):
        partido["fecha"] = partido["fecha"].strftime("%Y-%m-%dT%H:%M:%SZ")
    # Parsear eventos_api (JSON almacenado por sync)
    import json as _json_lp
    raw_ev = partido.pop("eventos_api_raw", None)
    try:
        partido["eventos_api"] = _json_lp.loads(raw_ev) if raw_ev else []
    except Exception:
        partido["eventos_api"] = []

    partido_id   = partido_row["id"]
    partido_nfifa = partido_row["numero_fifa"] or 0

    # 2. Lista de numeros_fifa disponibles para navegación
    nfifa_r = await db.execute(
        text("""
            SELECT DISTINCT p.numero_fifa
            FROM partido p
            JOIN fase f ON f.id = p.fase_id
            WHERE p.torneo_id = :tid AND p.numero_fifa IS NOT NULL
            ORDER BY p.numero_fifa
        """),
        {"tid": torneo_id},
    )
    numeros_fifa = [r[0] for r in nfifa_r.fetchall()]

    # 2b. Todos los partidos en_juego simultáneos (para la barra de selección)
    ej_r = await db.execute(
        text("""
            SELECT
                p.id, COALESCE(p.numero_fifa, 0) AS numero_fifa,
                COALESCE(el.nombre_es, el.nombre) AS equipo_local,
                COALESCE(ev.nombre_es, ev.nombre) AS equipo_visitante,
                el.codigo_iso AS bandera_local,
                ev.codigo_iso AS bandera_visitante,
                p.goles_local, p.goles_visitante, p.minuto_actual
            FROM partido p
            JOIN fase f ON f.id = p.fase_id
            LEFT JOIN equipo el ON el.id = p.equipo_local_id
            LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
            WHERE p.torneo_id = :tid AND p.estado = 'en_juego'
            ORDER BY p.numero_fifa
        """),
        {"tid": torneo_id},
    )
    partidos_en_juego = [dict(r) for r in ej_r.mappings().fetchall()]

    # 3. Predicciones por partido_id (fuente de verdad tras reload-apuestas-por-equipos)
    # NOTA: no usar a.numero_fifa porque el reload guarda numero_partido_fifa del paux,
    # que puede diferir del numero_fifa real del partido cuando hubo swap de pares (P049↔P050).
    apuestas_r = await db.execute(
        text("""
            SELECT
                a.apostador_id,
                a.pred_local, a.pred_visitante,
                a.pred_minuto_gol, a.pred_amarillas,
                a.pred_var, a.pred_rojas, a.pred_penales_partido,
                a.pred_penales_local_tanda, a.pred_penales_visitante_tanda,
                a.pred_equipo_clasifica
            FROM apuesta a
            WHERE a.partido_id = :pid
        """),
        {"pid": partido_id},
    )
    apuestas_map = {r["apostador_id"]: dict(r)
                    for r in apuestas_r.mappings().fetchall()}

    # 4. Ranking acumulado desde v_copamundial_puntajes (vista ya calculada, incluye globales)
    vista_map: dict[int, dict] = {}
    _live_pd_error: str | None = None
    try:
        vr = await db.execute(
            text("""
                SELECT apostador_id, apostador,
                       "H_resultado", "I_marcador", "J_amarillas", "K_rojas",
                       "L_var", "M_penales_partido", "N_minuto_gol", "O_penales_tanda",
                       "A_campeon", "B_finalistas", "C_goleador", "D_peor_equipo",
                       "E_mayor_goleada", "F_etapa_paraguay", "G_goles_paraguay",
                       subtotal_globales, total_puntos
                FROM v_copamundial_puntajes
            """)
        )
        for row in vr.mappings().fetchall():
            aid = int(row["apostador_id"])
            vista_map[aid] = {
                "nombre":          row["apostador"] or f"Apostador {aid}",
                "pts_H": int(row["H_resultado"]    or 0),
                "pts_I": int(row["I_marcador"]     or 0),
                "pts_J": int(row["J_amarillas"]    or 0),
                "pts_K": int(row["K_rojas"]        or 0),
                "pts_L": int(row["L_var"]          or 0),
                "pts_M": int(row["M_penales_partido"] or 0),
                "pts_N": int(row["N_minuto_gol"]   or 0),
                "pts_O": int(row["O_penales_tanda"] or 0),
                "pts_A": int(row["A_campeon"]      or 0),
                "pts_B": int(row["B_finalistas"]   or 0),
                "pts_C": int(row["C_goleador"]     or 0),
                "pts_D": int(row["D_peor_equipo"]  or 0),
                "pts_E": int(row["E_mayor_goleada"] or 0),
                "pts_F": int(row["F_etapa_paraguay"] or 0),
                "pts_G": int(row["G_goles_paraguay"] or 0),
                "subtotal_globales": int(row["subtotal_globales"] or 0),
                "total": int(row["total_puntos"]   or 0),
            }
    except Exception as _e:
        _live_pd_error = str(_e)
        print(f"[live_panel] ERROR v_copamundial_puntajes: {_e}")
        await db.rollback()
        # Fallback: leer puntaje_detalle + puntaje_global directamente
        try:
            fb_r = await db.execute(text("""
                SELECT d.apostador_id,
                       COALESCE(SUM(d.pts_resultado),0)        AS pts_H,
                       COALESCE(SUM(d.pts_marcador),0)         AS pts_I,
                       COALESCE(SUM(d.pts_amarillas),0)        AS pts_J,
                       COALESCE(SUM(d.pts_rojas),0)            AS pts_K,
                       COALESCE(SUM(d.pts_var),0)              AS pts_L,
                       COALESCE(SUM(d.pts_penales_partido),0)  AS pts_M,
                       COALESCE(SUM(d.pts_minuto),0)           AS pts_N,
                       COALESCE(SUM(d.pts_penales_tanda),0)    AS pts_O,
                       COALESCE(SUM(d.pts_resultado)+SUM(d.pts_marcador)+SUM(d.pts_amarillas)
                         +SUM(d.pts_rojas)+SUM(d.pts_var)+SUM(d.pts_penales_partido)
                         +SUM(d.pts_minuto)+SUM(d.pts_penales_tanda),0) AS total_partidos
                FROM puntaje_detalle d
                JOIN partido p ON p.id = d.partido_id
                JOIN fase    f ON f.id = p.fase_id
                WHERE f.torneo_id = :tid
                GROUP BY d.apostador_id
            """), {"tid": torneo_id})
            for row in fb_r.mappings().fetchall():
                aid = int(row["apostador_id"])
                vista_map[aid] = {
                    "nombre": f"Apostador {aid}",
                    "pts_H": int(row["pts_H"] or 0), "pts_I": int(row["pts_I"] or 0),
                    "pts_J": int(row["pts_J"] or 0), "pts_K": int(row["pts_K"] or 0),
                    "pts_L": int(row["pts_L"] or 0), "pts_M": int(row["pts_M"] or 0),
                    "pts_N": int(row["pts_N"] or 0), "pts_O": int(row["pts_O"] or 0),
                    "pts_A": 0, "pts_B": 0, "pts_C": 0, "pts_D": 0,
                    "pts_E": 0, "pts_F": 0, "pts_G": 0,
                    "subtotal_globales": 0,
                    "total": int(row["total_partidos"] or 0),
                }
        except Exception as _e2:
            print(f"[live_panel] ERROR fallback puntaje_detalle: {_e2}")
            vista_map = {}

    # 4b. Puntos de ESTE partido en puntaje_detalle (para excluirlos del acumulado "antes del partido")
    #     Si el partido ya está finalizado y se recalcularon puntajes, total_puntos LO INCLUYE.
    #     Restamos esos pts para que: acum_antes + live_calc = total_proyectado (sin doble conteo).
    try:
        det_r = await db.execute(
            text("""
                SELECT apostador_id,
                       COALESCE(SUM(
                           COALESCE(pts_resultado,0) + COALESCE(pts_marcador,0)
                           + COALESCE(pts_amarillas,0) + COALESCE(pts_rojas,0)
                           + COALESCE(pts_var,0) + COALESCE(pts_penales_partido,0)
                           + COALESCE(pts_minuto,0) + COALESCE(pts_penales_tanda,0)
                           + COALESCE(pts_equipo,0)
                       ), 0) AS pts_este_partido
                FROM puntaje_detalle
                WHERE partido_id = :pid
                GROUP BY apostador_id
            """),
            {"pid": partido_id},
        )
        pts_este_partido: dict[int, int] = {
            r["apostador_id"]: int(r["pts_este_partido"] or 0)
            for r in det_r.mappings().fetchall()
        }
    except Exception:
        await db.rollback()
        pts_este_partido = {}

    # 4c. Puntos KO por ítem (solo fases no-grupo, excluyendo partido actual)
    # Permite mostrar el acumulado KO por ítem en el panel live.
    try:
        ko_r = await db.execute(
            text("""
                SELECT d.apostador_id,
                       COALESCE(SUM(d.pts_resultado),       0) AS ko_h,
                       COALESCE(SUM(d.pts_marcador),        0) AS ko_i,
                       COALESCE(SUM(d.pts_amarillas),       0) AS ko_j,
                       COALESCE(SUM(d.pts_rojas),           0) AS ko_k,
                       COALESCE(SUM(d.pts_var),             0) AS ko_l,
                       COALESCE(SUM(d.pts_penales_partido), 0) AS ko_m,
                       COALESCE(SUM(d.pts_minuto),          0) AS ko_n,
                       COALESCE(SUM(d.pts_penales_tanda),   0) AS ko_o,
                       COALESCE(SUM(d.pts_equipo),          0) AS ko_p
                FROM puntaje_detalle d
                JOIN partido p2 ON p2.id = d.partido_id
                JOIN fase    f2 ON f2.id = p2.fase_id
                WHERE d.torneo_id = :tid
                  AND f2.tipo NOT ILIKE 'grupo%%'
                  AND d.partido_id != :pid
                GROUP BY d.apostador_id
            """),
            {"tid": torneo_id, "pid": partido_id},
        )
        ko_pts: dict[int, dict] = {
            int(r["apostador_id"]): {
                "H": int(r["ko_h"] or 0), "I": int(r["ko_i"] or 0),
                "J": int(r["ko_j"] or 0), "K": int(r["ko_k"] or 0),
                "L": int(r["ko_l"] or 0), "M": int(r["ko_m"] or 0),
                "N": int(r["ko_n"] or 0), "O": int(r["ko_o"] or 0),
                "P": int(r["ko_p"] or 0),
            }
            for r in ko_r.mappings().fetchall()
        }
    except Exception:
        await db.rollback()
        ko_pts = {}

    # 4d. Extra-bonus grupos P (clasificados a R32 — calculado una vez al cerrar grupos)
    try:
        gp_r = await db.execute(
            text("""
                SELECT apostador_id,
                       aciertos, pts_obtenidos AS pts_grupos_p,
                       equipos_pronosticados, equipos_reales
                FROM apostador_clasificados
                WHERE torneo_id = :tid AND fase_tipo = 'grupo'
            """),
            {"tid": torneo_id},
        )
        grupos_p_map: dict[int, dict] = {
            int(r["apostador_id"]): {
                "aciertos": int(r["aciertos"] or 0),
                "pts": int(r["pts_grupos_p"] or 0),
                "pronosticados": list(r["equipos_pronosticados"] or []),
                "reales": list(r["equipos_reales"] or []),
            }
            for r in gp_r.mappings().fetchall()
        }
    except Exception:
        await db.rollback()
        grupos_p_map = {}

    # 4e. Nombres de equipos reales en R32 (para mostrar lista de 32 con bonus info)
    equipos_r32_nombres: dict[int, str] = {}
    try:
        r32_reales = set()
        for gp_data in grupos_p_map.values():
            r32_reales.update(gp_data.get("reales", []))
        if r32_reales:
            ids_sql_r32 = ",".join(str(x) for x in r32_reales)
            eq_r = await db.execute(
                text(f"SELECT id, COALESCE(nombre_es, nombre) AS nombre FROM equipo WHERE id IN ({ids_sql_r32})")
            )
            equipos_r32_nombres = {int(r["id"]): r["nombre"] for r in eq_r.mappings().fetchall()}
    except Exception:
        pass

    # 5. Usernames desde app_db — para todos (ranking usa alias, no nombre completo)
    names_map: dict = {}
    all_ids = list(set(list(apuestas_map.keys()) + list(vista_map.keys())))
    if all_ids:
        try:
            async with _app_engine.connect() as aconn:
                ids_sql = ",".join(str(x) for x in all_ids)
                names_r = await aconn.execute(
                    text(f"SELECT id, COALESCE(nombre, username) AS nombre, username FROM users WHERE id IN ({ids_sql})")
                )
                names_map = {r[0]: {"nombre": r[1], "username": r[2]} for r in names_r.fetchall()}
        except Exception:
            pass

    # 5b. Filtrar solo apostadores reales (excluir admins/test como 'jose')
    valid_apostador_ids_lp: set[int] = set()
    try:
        async with _app_engine.connect() as _aconn_lp:
            _ar_lp = await _aconn_lp.execute(text("""
                SELECT u.id FROM users u
                JOIN user_roles ur ON ur.user_id = u.id
                JOIN roles ro ON ro.id = ur.role_id
                WHERE ro.name = 'apostador' AND u.is_active = TRUE
            """))
            valid_apostador_ids_lp = {row[0] for row in _ar_lp.fetchall()}
    except Exception:
        valid_apostador_ids_lp = set()

    # 6. Ensamblar lista de predicciones de este partido
    # Iteramos sobre vista_map (todos los apostadores del torneo) para que,
    # aunque el partido actual no tenga apuestas enlazadas, el panel igualmente
    # muestre a todos con sus puntos acumulados (pred_* quedan null → 0 pts live).
    apostadores = []
    all_base_ids = list(vista_map.keys()) if vista_map else list(apuestas_map.keys())
    base_ids = [uid for uid in all_base_ids if not valid_apostador_ids_lp or uid in valid_apostador_ids_lp]
    for uid in base_ids:
        ap = apuestas_map.get(uid, {})
        info = names_map.get(uid, {})
        v    = vista_map.get(uid, {"pts_H":0,"pts_I":0,"pts_J":0,"pts_K":0,
                                    "pts_L":0,"pts_M":0,"pts_N":0,"pts_O":0,
                                    "pts_A":0,"pts_B":0,"pts_C":0,"pts_D":0,
                                    "pts_E":0,"pts_F":0,"pts_G":0,
                                    "nombre": f"Usuario {uid}",
                                    "subtotal_globales":0,"total":0})
        pts_ya_calc  = pts_este_partido.get(uid, 0)
        puntos_antes = max(0, v["total"] - pts_ya_calc)
        ko_a = ko_pts.get(uid, {"H":0,"I":0,"J":0,"K":0,"L":0,"M":0,"N":0,"O":0,"P":0})
        gp   = grupos_p_map.get(uid, {"aciertos": 0, "pts": 0, "pronosticados": [], "reales": []})
        apostadores.append({
            "apostador_id": uid,
            "nombre":       info.get("nombre") or v.get("nombre") or f"Usuario {uid}",
            "username":     info.get("username", f"user{uid}"),
            "puntos_total": puntos_antes,
            "acum": {k: v[k] for k in ["pts_H","pts_I","pts_J","pts_K","pts_L","pts_M","pts_N","pts_O",
                                        "pts_A","pts_B","pts_C","pts_D","pts_E","pts_F","pts_G",
                                        "subtotal_globales"]},
            "ko_acum": ko_a,
            "grupos_p": gp,
            **{k: ap.get(k) for k in [
                "pred_local", "pred_visitante",
                "pred_minuto_gol", "pred_amarillas", "pred_var",
                "pred_rojas", "pred_penales_partido",
                "pred_penales_local_tanda", "pred_penales_visitante_tanda",
                "pred_equipo_clasifica",
            ]},
        })

    apostadores.sort(key=lambda x: x["nombre"])

    # 7. Ranking completo desde vista_map — usa username (alias) como nombre de display
    ranking_vista = []
    for aid, v in vista_map.items():
        if valid_apostador_ids_lp and aid not in valid_apostador_ids_lp:
            continue
        pts_ya_calc  = pts_este_partido.get(aid, 0)
        puntos_antes = max(0, v["total"] - pts_ya_calc)
        info = names_map.get(aid, {})
        ko_a = ko_pts.get(aid, {"H":0,"I":0,"J":0,"K":0,"L":0,"M":0,"N":0,"O":0,"P":0})
        gp   = grupos_p_map.get(aid, {"aciertos": 0, "pts": 0, "pronosticados": [], "reales": []})
        ranking_vista.append({
            "apostador_id": aid,
            "nombre":       v.get("nombre") or f"Apostador {aid}",
            "username":     info.get("username") or v.get("nombre") or f"Apostador {aid}",
            "puntos_antes": puntos_antes,
            "acum": {k: v[k] for k in ["pts_H","pts_I","pts_J","pts_K","pts_L","pts_M","pts_N","pts_O",
                                        "pts_A","pts_B","pts_C","pts_D","pts_E","pts_F","pts_G",
                                        "subtotal_globales"]},
            "ko_acum": ko_a,
            "grupos_p": gp,
        })
    ranking_vista.sort(key=lambda x: -x["puntos_antes"])

    return {
        "partido":          partido,
        "apostadores":      apostadores,
        "numeros_fifa":     numeros_fifa,
        "partidos_en_juego": partidos_en_juego,
        "ranking_vista":    ranking_vista,
        "equipos_r32":      equipos_r32_nombres,
        "_debug":           {"pd_error": _live_pd_error} if _live_pd_error else None,
    }


@router.patch("/partido-stats/{partido_id}", summary="Corrección manual de stats J/K/L/M (admin)")
async def patch_partido_stats(
    partido_id: int,
    current: CurrentAdmin,
    db: DBSession,
    amarillas:        int | None = None,
    rojas:            int | None = None,
    decisiones_var:   int | None = None,
    penales_partido:  int | None = None,
) -> dict:
    """
    Permite al admin corregir manualmente amarillas (J), rojas (K),
    VAR (L) y penales_partido (M) de un partido específico.
    Solo actualiza los campos que se envíen (no nulos).
    Recalcula puntajes automáticamente.
    """
    updates: dict = {}
    if amarillas       is not None: updates["amarillas"]       = amarillas
    if rojas           is not None: updates["rojas"]           = rojas
    if decisiones_var  is not None: updates["decisiones_var"]  = decisiones_var
    if penales_partido is not None: updates["penales_partido"] = penales_partido

    if not updates:
        raise HTTPException(400, "No se enviaron campos a actualizar")

    # Verificar que el partido existe
    pq = await db.execute(
        text("""SELECT p.id, p.estado,
                       COALESCE(el.nombre_es, el.nombre) AS local_nombre,
                       COALESCE(ev.nombre_es, ev.nombre) AS visit_nombre,
                       f.torneo_id
                FROM partido p
                JOIN fase f ON f.id = p.fase_id
                JOIN equipo el ON el.id = p.equipo_local_id
                JOIN equipo ev ON ev.id = p.equipo_visitante_id
                WHERE p.id = :pid"""),
        {"pid": partido_id},
    )
    row = pq.mappings().first()
    if not row:
        raise HTTPException(404, f"Partido {partido_id} no encontrado")

    set_clause = ", ".join(f"{k} = :{k}" for k in updates)
    await db.execute(
        text(f"UPDATE partido SET {set_clause} WHERE id = :pid"),
        {**updates, "pid": partido_id},
    )
    await db.commit()

    # Recalcular puntajes
    torneo_id = row["torneo_id"]
    puntajes_ok = False
    try:
        from app.services.scoring.registry import get_engine as _get_engine
        from app.services.scoring.calculator import ScoringCalculator as _SC
        r_comp = await db.execute(
            text("""SELECT c.codigo FROM torneo t
                    JOIN competicion c ON c.id = t.competicion_id
                    WHERE t.id = :tid"""),
            {"tid": torneo_id},
        )
        row_comp = r_comp.mappings().first()
        codigo   = (row_comp["codigo"] if row_comp else None) or "default"
        engine   = _get_engine(codigo)
        await _SC(db).calculate(torneo_id, engine)
        await db.commit()
        puntajes_ok = True
    except Exception:
        pass

    return {
        "ok": True,
        "partido_id": partido_id,
        "local": row["local_nombre"],
        "visitante": row["visit_nombre"],
        "updates": updates,
        "puntajes_recalculados": puntajes_ok,
    }


@router.get("/espn-verify/{partido_id}", summary="Verifica stats con ESPN (live page)")
async def espn_verify_live(
    partido_id: int,
    current: CurrentAdmin,
    db: DBSession,
) -> dict:
    from app.services.sync_api_football import _espn_verify_and_patch
    import httpx as _httpx
    pq = await db.execute(
        text("""
            SELECT p.id, p.estado, p.fecha,
                   COALESCE(p.datos_confirmados, FALSE) AS datos_confirmados,
                   COALESCE(el.nombre_es, el.nombre) AS local_nombre,
                   COALESCE(ev.nombre_es, ev.nombre) AS visit_nombre
            FROM partido p
            LEFT JOIN equipo el ON el.id = p.equipo_local_id
            LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
            WHERE p.id = :pid
        """),
        {"pid": partido_id},
    )
    row = pq.mappings().fetchone()
    if not row:
        raise HTTPException(404, f"Partido {partido_id} no encontrado")
    if row["datos_confirmados"]:
        return {
            "ok":         True,
            "partido_id": partido_id,
            "estado":     row["estado"],
            "correcciones": {},
            "msg": "Partido confirmado — datos protegidos, ESPN verify omitido",
        }
    async with _httpx.AsyncClient(timeout=20) as client:
        correcciones = await _espn_verify_and_patch(db, client, dict(row), partido_id, {})
    await db.commit()
    return {
        "ok":           True,
        "partido_id":   partido_id,
        "estado":       row["estado"],
        "correcciones": correcciones,
    }


@router.post("/sofascore-verify/{torneo_id}", summary="Corrige J/K/L/M con SofaScore (sin cuota API-Football)")
async def sofascore_verify(
    torneo_id: int,
    current: CurrentAdmin,
    db: DBSession,
    fecha: str | None = None,   # YYYY-MM-DD; None = ayer
) -> dict:
    """
    Corre verificación SofaScore sobre partidos finalizados del torneo.
    No consume cuota de API-Football. Corrige amarillas (J), rojas (K),
    VAR (L) y penales_partido (M). Luego recalcula puntajes.

    Params:
      fecha: YYYY-MM-DD (default: ayer)
             "hoy"  → partidos de hoy
             "all"  → todos los finalizados del torneo
    """
    from app.services.sync_api_football import (
        _sofascore_scoreboard, _sofascore_find_event,
        _sofascore_get_incidents, _sofascore_extract_stats,
    )
    import httpx as _httpx
    from datetime import date as _date, timedelta as _td, datetime as _dt

    # ── Resolver fecha ────────────────────────────────────────────────────────
    if fecha is None or fecha == "ayer":
        fecha_filter = str(_date.today() - _td(days=1))
    elif fecha == "hoy":
        fecha_filter = str(_date.today())
    elif fecha == "all":
        fecha_filter = None
    else:
        try:
            _dt.strptime(fecha, "%Y-%m-%d")
            fecha_filter = fecha
        except ValueError:
            raise HTTPException(400, f"Formato de fecha inválido: {fecha}. Usar YYYY-MM-DD, 'hoy', 'ayer' o 'all'")

    # ── Cargar partidos finalizados ───────────────────────────────────────────
    if fecha_filter:
        pq = await db.execute(
            text("""
                SELECT p.id, p.fecha,
                       COALESCE(el.nombre_es, el.nombre) AS local_nombre,
                       COALESCE(ev.nombre_es, ev.nombre) AS visit_nombre,
                       p.amarillas, p.rojas, p.decisiones_var, p.penales_partido
                FROM partido p
                JOIN fase f ON f.id = p.fase_id
                JOIN equipo el ON el.id = p.equipo_local_id
                JOIN equipo ev ON ev.id = p.equipo_visitante_id
                WHERE f.torneo_id = :tid AND p.estado = 'finalizado'
                  AND DATE(p.fecha) = :fd
                ORDER BY p.fecha
            """),
            {"tid": torneo_id, "fd": fecha_filter},
        )
    else:
        pq = await db.execute(
            text("""
                SELECT p.id, p.fecha,
                       COALESCE(el.nombre_es, el.nombre) AS local_nombre,
                       COALESCE(ev.nombre_es, ev.nombre) AS visit_nombre,
                       p.amarillas, p.rojas, p.decisiones_var, p.penales_partido
                FROM partido p
                JOIN fase f ON f.id = p.fase_id
                JOIN equipo el ON el.id = p.equipo_local_id
                JOIN equipo ev ON ev.id = p.equipo_visitante_id
                WHERE f.torneo_id = :tid AND p.estado = 'finalizado'
                ORDER BY p.fecha DESC
            """),
            {"tid": torneo_id},
        )

    partidos = [dict(r) for r in pq.mappings()]
    if not partidos:
        return {"ok": True, "msg": "Sin partidos finalizados para esa fecha", "correcciones": []}

    # ── Verificar con SofaScore ───────────────────────────────────────────────
    ss_cache: dict = {}
    correcciones = []
    sin_match = []

    async with _httpx.AsyncClient(timeout=15) as client:
        for p in partidos:
            partido_id = p["id"]
            local      = p["local_nombre"]
            visitante  = p["visit_nombre"]
            fecha_p    = p["fecha"]
            fecha_str  = fecha_p.strftime("%Y-%m-%d") if hasattr(fecha_p, "strftime") else str(fecha_p)[:10]

            if fecha_str not in ss_cache:
                ss_cache[fecha_str] = await _sofascore_scoreboard(client, fecha_p)

            events   = ss_cache.get(fecha_str, [])
            event_id = _sofascore_find_event(events, local, visitante)
            if not event_id:
                sin_match.append(f"{local} vs {visitante}")
                continue

            incidents = await _sofascore_get_incidents(client, event_id)
            if not incidents:
                continue

            ss = _sofascore_extract_stats(incidents)
            updates: dict = {}
            # SofaScore corrige solo si da un valor MAYOR que el actual (API+ESPN)
            if ss["amarillas"]       > (p["amarillas"]       or 0): updates["amarillas"]       = ss["amarillas"]
            if ss["rojas"]           > (p["rojas"]           or 0): updates["rojas"]           = ss["rojas"]
            if ss["decisiones_var"]  > (p["decisiones_var"]  or 0): updates["decisiones_var"]  = ss["decisiones_var"]
            if ss["penales_partido"] > (p["penales_partido"] or 0): updates["penales_partido"] = ss["penales_partido"]

            if updates:
                set_clause = ", ".join(f"{k} = :{k}" for k in updates)
                await db.execute(
                    text(f"UPDATE partido SET {set_clause} WHERE id = :pid"),
                    {**updates, "pid": partido_id},
                )
                await db.commit()
                correcciones.append({
                    "partido_id": partido_id,
                    "local": local, "visitante": visitante,
                    "updates": updates,
                    "sofascore_event_id": event_id,
                })

    # ── Recalcular puntajes si hubo correcciones ──────────────────────────────
    puntajes_ok = False
    if correcciones:
        try:
            from app.api.v1.endpoints.apostador_bets import calcular_puntajes as _calc
        except Exception:
            _calc = None
        try:
            calc_result = await db.execute(
                text("SELECT 1")  # placeholder; se llama la función directamente abajo
            )
        except Exception:
            pass
        # Llamar calcular_puntajes internamente (reusa la lógica del endpoint)
        from app.services.scoring.registry import get_engine as _get_engine
        from app.services.scoring.calculator import ScoringCalculator as _SC
        try:
            r_comp = await db.execute(
                text("""SELECT c.codigo FROM torneo t
                        JOIN competicion c ON c.id = t.competicion_id
                        WHERE t.id = :tid"""),
                {"tid": torneo_id},
            )
            row_comp = r_comp.mappings().first()
            codigo   = (row_comp["codigo"] if row_comp else None) or "default"
            engine   = _get_engine(codigo)
            result   = await _SC(db).calculate(torneo_id, engine)
            await _SC(db).calculate_global(torneo_id, engine)
            await db.commit()
            puntajes_ok = True
        except Exception as e:
            puntajes_ok = False

    return {
        "ok": True,
        "fecha": fecha_filter or "all",
        "partidos_revisados": len(partidos),
        "correcciones": correcciones,
        "sin_match_sofascore": sin_match,
        "puntajes_recalculados": puntajes_ok,
    }


@router.post(
    "/populate-stats-fuentes/{torneo_id}",
    summary="Pobla tabla partido_stats_fuentes para todos los partidos finalizados (admin)",
)
async def populate_stats_fuentes(
    torneo_id: int,
    current: CurrentAdmin,
    db: DBSession,
) -> dict:
    """
    Llama ESPN + SofaScore para TODOS los partidos finalizados del torneo y
    almacena los valores crudos de cada fuente en partido_stats_fuentes.

    No usa cuota de API-Football. Aplica lógica "máximo gana": SofaScore solo
    corrige el partido si su valor es mayor que el actual (API+ESPN). ESPN solo
    se usa para lectura sin aplicar (ya fue aplicado en sincronizaciones previas).

    Después de corregir, recalcula puntajes si hubo algún cambio.
    """
    import httpx as _httpx
    from app.services.sync_api_football import populate_stats_fuentes_all

    async with _httpx.AsyncClient(timeout=20) as client:
        result = await populate_stats_fuentes_all(db, torneo_id, client)

    # Recalcular puntajes si algún partido fue corregido
    puntajes_ok = False
    if result.get("corregidos_ss", 0) > 0:
        try:
            from app.services.scoring.registry import get_engine as _get_engine
            from app.services.scoring.calculator import ScoringCalculator as _SC
            r_comp = await db.execute(
                text("""SELECT c.codigo FROM torneo t
                        JOIN competicion c ON c.id = t.competicion_id
                        WHERE t.id = :tid"""),
                {"tid": torneo_id},
            )
            row_comp = r_comp.mappings().first()
            codigo   = (row_comp["codigo"] if row_comp else None) or "default"
            engine   = _get_engine(codigo)
            await _SC(db).calculate(torneo_id, engine)
            await _SC(db).calculate_global(torneo_id, engine)
            await db.commit()
            puntajes_ok = True
        except Exception as e:
            puntajes_ok = False

    return {
        "ok": True,
        **result,
        "puntajes_recalculados": puntajes_ok,
    }


@router.get(
    "/stats-fuentes/{torneo_id}",
    summary="Ver tabla de comparación de fuentes por partido (admin)",
)
async def get_stats_fuentes(
    torneo_id: int,
    current: CurrentAdmin,
    db: DBSession,
    solo_diferencias: bool = False,
    estado: str | None = None,   # filtrar por 'pendiente'/'live'/'finalizado'
) -> dict:
    """
    Devuelve el contenido de partido_stats_fuentes para el torneo.

    Params:
      solo_diferencias=true: solo partidos donde las fuentes difieren entre sí.
      estado: filtrar por estado ('pendiente', 'live', 'finalizado').
    """
    base_q = """
        SELECT
            sf.partido_id,
            sf.fecha,
            sf.local,
            sf.visitante,
            sf.estado,
            sf.ultimo_minuto,
            sf.fuentes_run_at,
            sf.api_amarillas,    sf.api_rojas,    sf.api_var,    sf.api_penales,
            sf.espn_amarillas,   sf.espn_rojas,   sf.espn_var,   sf.espn_penales,
            sf.ss_amarillas,     sf.ss_rojas,      sf.ss_var,     sf.ss_penales,
            sf.final_amarillas,  sf.final_rojas,   sf.final_var,  sf.final_penales,
            sf.fuente_amarillas, sf.fuente_rojas,  sf.fuente_var, sf.fuente_penales,
            sf.synced_at
        FROM partido_stats_fuentes sf
        WHERE sf.torneo_id = :tid
    """
    params: dict = {"tid": torneo_id}

    if estado:
        base_q += " AND sf.estado = :estado"
        params["estado"] = estado

    if solo_diferencias:
        base_q += """
          AND (
               COALESCE(sf.api_amarillas,0)  != COALESCE(sf.ss_amarillas,0)
            OR COALESCE(sf.api_rojas,0)      != COALESCE(sf.ss_rojas,0)
            OR COALESCE(sf.api_var,0)        != COALESCE(sf.ss_var,0)
            OR COALESCE(sf.api_penales,0)    != COALESCE(sf.ss_penales,0)
            OR COALESCE(sf.espn_amarillas,0) != COALESCE(sf.ss_amarillas,0)
            OR COALESCE(sf.espn_rojas,0)     != COALESCE(sf.ss_rojas,0)
          )
        """
    base_q += " ORDER BY sf.fecha, sf.partido_id"

    rq = await db.execute(text(base_q), params)
    rows = [dict(r) for r in rq.mappings().all()]

    # Serializar fechas y timestamps
    for r in rows:
        for col in ("fecha", "fuentes_run_at", "synced_at"):
            if r.get(col) and hasattr(r[col], "isoformat"):
                r[col] = r[col].isoformat()

    # Resumen por estado
    from collections import Counter
    estado_counts = Counter(r.get("estado", "?") for r in rows)

    return {
        "ok": True,
        "torneo_id": torneo_id,
        "total": len(rows),
        "solo_diferencias": solo_diferencias,
        "resumen": dict(estado_counts),
        "partidos": rows,
    }
