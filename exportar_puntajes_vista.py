"""
exportar_puntajes_vista.py
Genera un Excel con dos hojas desde las vistas de becbuc:
  1. Resumen  <- v_copamundial_puntajes
  2. Detalle  <- v_copamundial_puntajes_det
Ejecutar: python exportar_puntajes_vista.py
"""
import subprocess, json, sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB      = "becbuc"
USER    = "app_user"
CONT    = "core-postgres"
OUTFILE = r"C:\proyecto FAST API\puntajes_copa_mundial.xlsx"

# ─── helpers ─────────────────────────────────────────────────────────────────
def psql(sql):
    result = subprocess.run(
        ["docker", "exec", "-i", CONT, "psql", "-U", USER, "-d", DB,
         "-t", "-A", "-F", "\t", "-c", sql],
        capture_output=True, text=True, encoding="utf-8"
    )
    if result.returncode != 0:
        print("ERROR psql:", result.stderr); sys.exit(1)
    rows = []
    for line in result.stdout.strip().splitlines():
        rows.append(line.split("\t"))
    return rows

def col_names(view):
    rows = psql(f"""
        SELECT column_name FROM information_schema.columns
        WHERE table_name = '{view}'
        ORDER BY ordinal_position
    """)
    return [r[0] for r in rows]

def fetch(view):
    cols  = col_names(view)
    data  = psql(f"SELECT * FROM {view}")
    return cols, data

# ─── estilos ─────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
ALT_FILL   = PatternFill("solid", fgColor="DCE6F1")
NORM_FILL  = PatternFill("solid", fgColor="FFFFFF")
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FONT = Font(bold=True, name="Calibri", size=10)
CELL_FONT  = Font(name="Calibri", size=10)
CENTER     = Alignment(horizontal="center", vertical="center")
LEFT       = Alignment(horizontal="left",   vertical="center")
thin       = Side(style="thin", color="BBBBBB")
BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)

# columnas de puntos (contienen "_pts" o "total")
def is_pts(col): return col.endswith("_pts") or col == "total_partido" or col == "total_puntos"
def is_total(col): return col in ("total_partido", "total_puntos", "subtotal_partidos", "subtotal_globales")

PTS_FILL   = PatternFill("solid", fgColor="E2EFDA")
PTS_FONT   = Font(name="Calibri", size=10, color="375623")

def write_sheet(ws, cols, rows, title):
    ws.title = title

    # fila de cabecera
    for c, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=col.replace("_", " ").title())
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = CENTER
        cell.border    = BORDER

    # filas de datos
    for r, row in enumerate(rows, 2):
        alt = (r % 2 == 0)
        for c, (col, val) in enumerate(zip(cols, row), 1):
            # convertir números
            try:    val = int(val)   if val and val != "" else (None if val == "" else val)
            except: pass
            cell = ws.cell(row=r, column=c, value=val)
            cell.border    = BORDER
            cell.font      = TOTAL_FONT if is_total(col) else CELL_FONT
            cell.alignment = CENTER if (isinstance(val, int) or col in ("goles_local","goles_visitante")) else LEFT
            if is_total(col):
                cell.fill = TOTAL_FILL
            elif is_pts(col):
                cell.fill = PTS_FILL
                cell.font = PTS_FONT
            else:
                cell.fill = ALT_FILL if alt else NORM_FILL

    # ancho automático
    for c, col in enumerate(cols, 1):
        max_len = max(
            len(str(col)),
            max((len(str(r[c-1])) for r in rows if c-1 < len(r)), default=0)
        )
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 3, 35)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

# ─── main ────────────────────────────────────────────────────────────────────
print("Consultando v_copamundial_puntajes …")
cols1, rows1 = fetch("v_copamundial_puntajes")

print("Consultando v_copamundial_puntajes_det …")
cols2, rows2 = fetch("v_copamundial_puntajes_det")

wb = Workbook()
ws1 = wb.active
write_sheet(ws1, cols1, rows1, "Resumen")

ws2 = wb.create_sheet()
write_sheet(ws2, cols2, rows2, "Detalle")

wb.save(OUTFILE)
print(f"\n✅ Guardado en: {OUTFILE}")
print(f"   Resumen : {len(rows1)} apostadores")
print(f"   Detalle : {len(rows2)} filas")
