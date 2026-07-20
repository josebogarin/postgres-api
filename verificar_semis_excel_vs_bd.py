# -*- coding: utf-8 -*-
"""
verificar_semis_excel_vs_bd.py
Compara ITEM POR ITEM las apuestas de SEMIFINAL (P101-P102) de cada apostador
entre el Excel (hoja '50- TBL MASTER', fase '50- SEMIFINAL') y la BD (tabla apuesta).

Reporta las diferencias por apostador. Si no hay ninguna: "TODO IDENTICO".
Solo lectura: NO modifica la BD.

Uso:
  backend\\.venv\\Scripts\\python.exe verificar_semis_excel_vs_bd.py
"""
import sys, os

BASE = os.path.dirname(os.path.abspath(__file__))

EXCEL_FILE = None
for f in os.listdir(BASE):
    if 'SEMIFINAL' in f.upper() and f.endswith('.xlsx'):
        EXCEL_FILE = os.path.join(BASE, f); break
if not EXCEL_FILE:
    sys.exit(f"ERROR: no se encontro Excel *SEMIFINAL*.xlsx en {BASE}")
print(f"Excel: {os.path.basename(EXCEL_FILE)}")

try:
    import openpyxl
except ImportError:
    os.system(f'"{sys.executable}" -m pip install openpyxl --quiet'); import openpyxl
try:
    import psycopg2, psycopg2.extras
except ImportError:
    os.system(f'"{sys.executable}" -m pip install psycopg2-binary --quiet'); import psycopg2, psycopg2.extras

CONN_BEC = "host=localhost port=5432 dbname=becbuc user=app_user password=superpassword"
CONN_APP = "host=localhost port=5432 dbname=app_db  user=app_user password=superpassword"
TORNEO_ID = 2
SHEET = '50- TBL MASTER'
FASE_TXT = '50- SEMIFINAL'
NF_MIN, NF_MAX = 101, 102

try:
    conn_bec = psycopg2.connect(CONN_BEC); conn_app = psycopg2.connect(CONN_APP)
except Exception as e:
    sys.exit(f"ERROR conexion: {e}\nDocker corriendo? docker start core-postgres")
cur_bec = conn_bec.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
cur_app = conn_app.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# -- Apostadores ------------------------------------------------------------
cur_app.execute("SELECT id, username FROM users WHERE is_active=TRUE ORDER BY id")
all_users = cur_app.fetchall()
cur_bec.execute("""
    SELECT DISTINCT a.apostador_id FROM apuesta a
    JOIN partido p ON p.id = a.partido_id
    JOIN fase f ON f.id = p.fase_id
    WHERE f.torneo_id = %s
""", (TORNEO_ID,))
bec_ids = {r['apostador_id'] for r in cur_bec.fetchall()}
bd_apostadores = [u for u in all_users if u['id'] in bec_ids]
apostador_to_id = {u['username'].lower(): u['id'] for u in bd_apostadores}
username_by_id = {u['id']: u['username'] for u in bd_apostadores}

# -- Equipos ----------------------------------------------------------------
cur_bec.execute("SELECT id, nombre, nombre_es FROM equipo")
equipo_id_by_nombre = {}
equipo_nombre_by_id = {}
for eq in cur_bec.fetchall():
    equipo_nombre_by_id[eq['id']] = eq['nombre'] or eq['nombre_es']
    if eq['nombre']:    equipo_id_by_nombre[eq['nombre'].upper().strip()] = eq['id']
    if eq['nombre_es']: equipo_id_by_nombre[eq['nombre_es'].upper().strip()] = eq['id']
EQUIPO_ALIAS = {
    'FRANCIA':'France','ESPAÑA':'Spain','ESPANA':'Spain',
    'INGLATERRA':'England','ARGENTINA':'Argentina',
    'MARRUECOS':'Morocco','BELGICA':'Belgium','NORUEGA':'Norway','SUIZA':'Switzerland',
}
def find_equipo_id(nombre_excel):
    if not nombre_excel: return None
    key = str(nombre_excel).upper().strip()
    if key in equipo_id_by_nombre: return equipo_id_by_nombre[key]
    alt = EQUIPO_ALIAS.get(key)
    if alt and alt.upper().strip() in equipo_id_by_nombre:
        return equipo_id_by_nombre[alt.upper().strip()]
    for k, v in equipo_id_by_nombre.items():
        if key in k or k in key: return v
    return None

# -- Partidos semifinal -----------------------------------------------------
cur_bec.execute("""
    SELECT p.id, p.numero_fifa FROM partido p JOIN fase f ON f.id=p.fase_id
    WHERE f.torneo_id=%s AND p.numero_fifa BETWEEN %s AND %s
""", (TORNEO_ID, NF_MIN, NF_MAX))
pid_by_key = {f"P{r['numero_fifa']:03d}": r['id'] for r in cur_bec.fetchall()}
key_by_pid = {v: k for k, v in pid_by_key.items()}

# -- apuestas en BD ---------------------------------------------------------
cur_bec.execute("""
    SELECT a.apostador_id, a.partido_id, a.pred_local, a.pred_visitante,
           a.pred_amarillas, a.pred_rojas, a.pred_var, a.pred_penales_partido, a.pred_minuto_gol,
           a.pred_penales_local_tanda, a.pred_penales_visitante_tanda, a.pred_equipo_clasifica
    FROM apuesta a
    WHERE a.partido_id = ANY(%s)
""", (list(pid_by_key.values()),))
db_map = {(r['apostador_id'], r['partido_id']): dict(r) for r in cur_bec.fetchall()}

# -- Helpers ----------------------------------------------------------------
def clean_alias(s):
    if not s: return ''
    return str(s).replace('\xa0','').strip().upper().lstrip('@').replace('  ',' ')
def find_apostador_id(alias_excel):
    a = clean_alias(alias_excel)
    for k, v in apostador_to_id.items():
        if k.upper().lstrip('@') == a: return v
    for k, v in apostador_to_id.items():
        if a and (a in k.upper() or k.upper() in a): return v
    return None
def to_int(v):
    try:
        s = str(v).strip()
        if s in ('', '-', 'None'): return None
        return int(float(s))
    except: return None

# columnas 1-based del Excel -> campo
FIELDS = [
    ('pred_local', 13, 'goles'),
    ('pred_visitante', 15, 'goles'),
    ('pred_amarillas', 25, 'int'),
    ('pred_rojas', 26, 'int'),
    ('pred_var', 27, 'int'),
    ('pred_penales_partido', 28, 'int'),
    ('pred_minuto_gol', 29, 'int'),
    ('pred_penales_local_tanda', 30, 'int'),
    ('pred_penales_visitante_tanda', 31, 'int'),
    ('pred_equipo_clasifica', 32, 'equipo'),
]

# -- Leer Excel y comparar --------------------------------------------------
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb[SHEET]

diffs = []           # (username, pid, campo, excel, bd)
sin_alias = {}
sin_bd = []          # apuesta en excel pero no en BD
filas_excel = 0
comparados = 0

for r in range(2, ws.max_row + 1):
    if str(ws.cell(r, 7).value) != FASE_TXT:
        continue
    filas_excel += 1
    pid = str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value else ''
    alias = ws.cell(r, 10).value
    uid = find_apostador_id(alias)
    if not uid:
        k = clean_alias(alias); sin_alias[k] = sin_alias.get(k, 0) + 1; continue
    partido_id = pid_by_key.get(pid)
    if not partido_id:
        continue
    dbrow = db_map.get((uid, partido_id))
    uname = username_by_id.get(uid, str(uid))
    if not dbrow:
        sin_bd.append((uname, pid)); continue
    comparados += 1
    for campo, col, tipo in FIELDS:
        raw = ws.cell(r, col).value
        if tipo == 'equipo':
            exc = find_equipo_id(raw)
            db = dbrow[campo]
            if exc != db:
                diffs.append((uname, pid, campo,
                              equipo_nombre_by_id.get(exc, exc),
                              equipo_nombre_by_id.get(db, db)))
        else:
            exc = to_int(raw)
            if tipo == 'goles' and exc is None:
                exc = 0   # el import guarda 0 cuando el goles esta vacio
            db = dbrow[campo]
            if exc != db:
                diffs.append((uname, pid, campo, exc, db))

# -- Reporte ----------------------------------------------------------------
print("\n" + "="*66)
print("VERIFICACION SEMIFINAL: Excel vs BD (item por item)")
print("="*66)
print(f"Partidos BD: {sorted(key_by_pid.values())}")
print(f"Filas semifinal en Excel: {filas_excel}")
print(f"Apuestas comparadas (Excel con match en BD): {comparados}  (esperado 88)")
if sin_alias:
    print(f"Aliases del Excel sin match: {dict(sin_alias)}")
if sin_bd:
    print(f"En Excel pero SIN apuesta en BD ({len(sin_bd)}): {sin_bd}")

if not diffs:
    print("\nRESULTADO: TODO IDENTICO - el Excel coincide 100% con la BD para cada jugador.")
else:
    print(f"\nRESULTADO: {len(diffs)} DIFERENCIA(S) encontradas:")
    print(f"{'APOSTADOR':<22}{'PART':<6}{'CAMPO':<28}{'EXCEL':<16}{'BD':<16}")
    print("-"*88)
    for uname, pid, campo, exc, db in diffs:
        print(f"{uname:<22}{pid:<6}{campo:<28}{str(exc):<16}{str(db):<16}")
    # resumen por apostador
    from collections import Counter
    porap = Counter(d[0] for d in diffs)
    print("\nDiferencias por apostador:")
    for u, n in porap.most_common():
        print(f"  {u}: {n}")

conn_bec.close(); conn_app.close()
