"""
comparar_quiroga.py
====================
Compara puntajes de QUIROGA partido a partido:
  - Excel TBL MASTER: predicciones raw (pred_local, pred_visitante, amarillas, etc.)
  - Sistema BD: puntaje_detalle calculado
  - Recalculo Python: aplica reglamento BEC BUC sobre predicciones Excel + resultados BD

Uso: python comparar_quiroga.py
"""
import os as _osp
_BASE = _osp.path.dirname(_osp.path.abspath(__file__))

import re
import psycopg2
import openpyxl

# ---------- CONFIGURACION ----------
DB_HOST = "localhost"
DB_PORT = 5432
DB_USER = "app_user"
DB_NAME = "becbuc"
DB_APP = "app_db"   # BD de usuarios
TORNEO_ID = 2

EXCEL_PATH = _osp.path.join(_BASE, '20260611_2000- TBL CONSOLIDADA PRONOSTICOS ok.xlsx')
ALIAS_BUSCAR = "QUIROGA"  # case-insensitive

# Puntos por fase (reglamento BEC BUC)
FASE_PTS = {
    "grupo":         {"H": 4,  "I": 8},
    "ronda32":       {"H": 6,  "I": 12},
    "ronda16":       {"H": 8,  "I": 16},
    "cuartos":       {"H": 10, "I": 20},
    "semis":         {"H": 12, "I": 24},
    "tercer_puesto": {"H": 14, "I": 28},
    "final":         {"H": 20, "I": 40},
}

FASE_EXCEL_MAP = {
    "10- grupos": "grupo",
    "20- ronda de 32": "ronda32",
    "30- ronda de 16": "ronda16",
    "40- cuartos": "cuartos",
    "50- semifinales": "semis",
    "60- tercer puesto": "tercer_puesto",
    "70- final": "final",
}


def wdl(l, v):
    """Gana/Empata/Pierde local."""
    if l > v:
        return "G"
    elif l == v:
        return "E"
    return "P"


def calc_excel_pts(pred_l, pred_v, real_l, real_v, fase_tipo,
                   pred_amar, real_amar, pred_rojas, real_rojas,
                   pred_var, real_var, pred_pp, real_pp,
                   es_paraguay=False):
    """Calcula puntaje según reglamento BEC BUC (mismo algoritmo del engine Python)."""
    mult = 2 if es_paraguay else 1
    cfg = FASE_PTS.get(fase_tipo, {"H": 4, "I": 8})

    pts_h, pts_i, pts_j, pts_k, pts_l, pts_m = 0, 0, 0, 0, 0, 0

    if None in (pred_l, pred_v, real_l, real_v):
        return 0, 0, 0, 0, 0, 0

    # H - Resultado
    if wdl(pred_l, pred_v) == wdl(real_l, real_v):
        pts_h = cfg["H"] * mult

    # I - Exacto (sólo si H también acertó)
    if pred_l == real_l and pred_v == real_v:
        pts_h = 0  # I reemplaza a H
        pts_i = cfg["I"] * mult

    # J - Amarillas (NULL → 0)
    if (pred_amar or 0) == (real_amar or 0):
        pts_j = 1 * mult

    # K - Rojas (NULL → 0)
    if (pred_rojas or 0) == (real_rojas or 0):
        pts_k = 1 * mult

    # L - VAR (NULL → 0)
    if (pred_var or 0) == (real_var or 0):
        pts_l = 1 * mult

    # M - Penales partido (NULL → 0)
    if (pred_pp or 0) == (real_pp or 0):
        pts_m = 1 * mult

    return pts_h, pts_i, pts_j, pts_k, pts_l, pts_m


def main():
    # ---- 1. Leer Excel: predicciones de Quiroga ----
    print("Leyendo Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["50- TBL MASTER"]

    quiroga_rows = {}  # partido_num -> dict con predicciones
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        alias = str(row[9] or "").strip()
        if ALIAS_BUSCAR.lower() not in alias.lower():
            continue

        id_partido_str = str(row[1] or "").strip()   # col B: 'P001'
        # Extraer numero de los primeros 4 chars: P001 -> 1
        m = re.match(r"P0*(\d+)", id_partido_str)
        if not m:
            continue
        partido_num = int(m.group(1))

        fase_str = str(row[6] or "").strip().lower()  # col G: FASE

        quiroga_rows[partido_num] = {
            "id_partido_str": id_partido_str,
            "fase_str":       fase_str,
            "pred_local":     row[12],   # col M
            "pred_visitante":  row[14],  # col O
            "pred_amar":      row[23],   # col X
            "pred_rojas":     row[24],   # col Y
            "pred_var":       row[25],   # col Z
            "pred_pp":        row[26],   # col [ (index 26)
            "pred_minuto":    row[27],   # col \ (index 27)
            "v_text":         str(row[21] or ""),  # col V
        }

    print(f"  Quiroga: {len(quiroga_rows)} partidos en Excel")

    # ---- 2. Conectar a BD ----
    conn = psycopg2.connect(host=DB_HOST, port=DB_PORT, user=DB_USER, dbname=DB_NAME)
    cur = conn.cursor()

    # Mapeo partido_num (secuencia) → partido_id BD
    cur.execute("""
        SELECT
            ROW_NUMBER() OVER (ORDER BY f.orden, p.id) AS num_seq,
            p.id,
            p.goles_local, p.goles_visitante, p.estado,
            p.amarillas, p.rojas, p.decisiones_var, p.penales_partido,
            p.minuto_primer_gol,
            p.penales_local, p.penales_visitante,
            f.tipo AS fase_tipo,
            e1.nombre AS equipo_local,
            e2.nombre AS equipo_visitante
        FROM partido p
        JOIN fase f ON f.id = p.fase_id
        JOIN equipo e1 ON e1.id = p.equipo_local_id
        JOIN equipo e2 ON e2.id = p.equipo_visitante_id
        WHERE f.torneo_id = %s
        ORDER BY f.orden, p.id
    """, (TORNEO_ID,))

    partidos_db = {}
    for row in cur.fetchall():
        partidos_db[row[0]] = {  # key = num_seq
            "partido_id": row[1],
            "goles_local": row[2],
            "goles_visitante": row[3],
            "estado": row[4],
            "amarillas": row[5],
            "rojas": row[6],
            "decisiones_var": row[7],
            "penales_partido": row[8],
            "minuto_primer_gol": row[9],
            "penales_local": row[10],
            "penales_visitante": row[11],
            "fase_tipo": row[12],
            "local": row[13],
            "visitante": row[14],
        }

    print(f"  BD: {len(partidos_db)} partidos totales")

    # Obtener apostador_id de Quiroga desde app_db (usuarios)
    conn_app = psycopg2.connect(host=DB_HOST, port=DB_PORT, user=DB_USER, dbname=DB_APP)
    cur_app = conn_app.cursor()
    cur_app.execute("""
        SELECT id, username
        FROM users
        WHERE LOWER(username) LIKE %s
        LIMIT 5
    """, (f"%{ALIAS_BUSCAR.lower()}%",))
    usuarios = cur_app.fetchall()
    cur_app.close()
    conn_app.close()

    if not usuarios:
        print("❌ No se encontró usuario Quiroga en app_db")
        return

    print(f"  Usuario(s) encontrado(s): {usuarios}")
    apostador_id = usuarios[0][0]

    # Obtener puntaje_detalle de Quiroga
    cur.execute("""
        SELECT
            d.partido_id,
            COALESCE(d.pts_resultado, 0),
            COALESCE(d.pts_marcador, 0),
            COALESCE(d.pts_amarillas, 0),
            COALESCE(d.pts_rojas, 0),
            COALESCE(d.pts_var, 0),
            COALESCE(d.pts_penales_partido, 0),
            COALESCE(d.pts_minuto, 0),
            COALESCE(d.pts_penales_tanda, 0),
            COALESCE(d.pts_total, 0)
        FROM puntaje_detalle d
        WHERE d.apostador_id = %s
    """, (apostador_id,))
    detalle_db = {}
    for row in cur.fetchall():
        detalle_db[row[0]] = {
            "d_h": row[1], "d_i": row[2], "d_j": row[3],
            "d_k": row[4], "d_l": row[5], "d_m": row[6],
            "d_n": row[7], "d_o": row[8], "d_total": row[9],
        }
    print(f"  puntaje_detalle: {len(detalle_db)} filas para Quiroga")

    # ---- 3. Comparar ----
    SEP = "=" * 130
    print("\n" + SEP)
    hdr = (f"{'#':<5} {'PARTIDO':<28} {'FASE':<10}  "
           f"{'PRED':>5} {'REAL':>5}  "
           f"{'EXCEL':^23}  {'SISTEMA':^23}  "
           f"{'DIFF':>5}")
    sub = (f"{'':5} {'':28} {'':10}  "
           f"{'':5} {'':5}  "
           f"{'H':>3} {'I':>3} {'J':>3} {'K':>3} {'L':>3} {'M':>3} {'TOT':>5}  "
           f"{'H':>3} {'I':>3} {'J':>3} {'K':>3} {'L':>3} {'M':>3} {'TOT':>5}  "
           f"{'':5}")
    print(hdr)
    print(sub)
    print("-" * 130)

    total_excel = 0
    total_sistema = 0
    diffs = []

    for num in sorted(quiroga_rows.keys()):
        ex = quiroga_rows[num]
        p = partidos_db.get(num)
        if not p:
            continue

        # Solo partidos finalizados
        if p["estado"] != "finalizado" and p.get("goles_local") is None:
            continue

        rl = p["goles_local"]
        rv = p["goles_visitante"]
        if rl is None:
            continue

        fase_tipo = p["fase_tipo"]
        es_paraguay = ("paraguay" in p["local"].lower() or "paraguay" in p["visitante"].lower())

        pts_h, pts_i, pts_j, pts_k, pts_l, pts_m = calc_excel_pts(
            ex["pred_local"], ex["pred_visitante"], rl, rv, fase_tipo,
            ex["pred_amar"], p["amarillas"],
            ex["pred_rojas"], p["rojas"],
            ex["pred_var"], p["decisiones_var"],
            ex["pred_pp"], p["penales_partido"],
            es_paraguay
        )
        ex_tot = pts_h + pts_i + pts_j + pts_k + pts_l + pts_m

        pid = p["partido_id"]
        d = detalle_db.get(pid, {})
        sis_h   = d.get("d_h", 0)
        sis_i   = d.get("d_i", 0)
        sis_j   = d.get("d_j", 0)
        sis_k   = d.get("d_k", 0)
        sis_l   = d.get("d_l", 0)
        sis_m   = d.get("d_m", 0)
        sis_n   = d.get("d_n", 0)
        sis_o   = d.get("d_o", 0)
        sis_tot = d.get("d_total", 0)

        diff = ex_tot - sis_tot
        total_excel   += ex_tot
        total_sistema += sis_tot

        pl = ex["pred_local"]
        pv = ex["pred_visitante"]
        partido_str = f"{p['local'][:13]} vs {p['visitante'][:13]}"

        # Marcar items con diferencia con asterisco
        def c(ex_v, sis_v):
            return f"{ex_v:>3}{'*' if ex_v != sis_v else ' '}"

        diff_str = f"{'+' if diff>0 else ''}{diff}" if diff != 0 else "  ok"
        py_mark = " 🇵🇾x2" if es_paraguay else ""
        no_det  = " ⚠nodet" if pid not in detalle_db else ""

        print(f"P{num:03d}  {partido_str:<28} {fase_tipo:<10}  "
              f"{str(pl)+'-'+str(pv):>5} {str(rl)+'-'+str(rv):>5}  "
              f"{c(pts_h,sis_h)} {c(pts_i,sis_i)} {c(pts_j,sis_j)} {c(pts_k,sis_k)} {c(pts_l,sis_l)} {c(pts_m,sis_m)} {ex_tot:>5}  "
              f"{sis_h:>3}  {sis_i:>3}  {sis_j:>3}  {sis_k:>3}  {sis_l:>3}  {sis_m:>3}  {sis_tot:>5}  "
              f"{diff_str:>5}{py_mark}{no_det}")

        # Detalle de predicciones en línea adicional si hay diferencia
        if diff != 0 or pid not in detalle_db:
            pred_j = ex["pred_amar"]; real_j = p["amarillas"]
            pred_k = ex["pred_rojas"]; real_k = p["rojas"]
            pred_l = ex["pred_var"];  real_l = p["decisiones_var"]
            pred_m = ex["pred_pp"];   real_m = p["penales_partido"]
            print(f"     → preds: J={pred_j}|{real_j}  K={pred_k}|{real_k}  L={pred_l}|{real_l}  M={pred_m}|{real_m}  "
                  f"{'N(sis)='+str(sis_n) if sis_n else ''}{'  O(sis)='+str(sis_o) if sis_o else ''}")
            diffs.append((num, p["local"], p["visitante"], diff, ex_tot, sis_tot))

    print(SEP)
    print(f"\n{'TOTAL EXCEL (recalculado)':<35}: {total_excel:>4} pts")
    print(f"{'TOTAL SISTEMA (puntaje_detalle)':<35}: {total_sistema:>4} pts")
    print(f"{'DIFERENCIA':<35}: {total_excel-total_sistema:>+4} pts")
    if sis_n or sis_o:
        print("  (Sistema incluye N+O; Excel no tiene N/O en este cálculo)")
    print(f"\nPartidos con diferencia: {len(diffs)}")
    for d in diffs:
        print(f"  P{d[0]:03d} {d[1]} vs {d[2]}: Excel={d[4]} Sis={d[5]} diff={d[3]:+d}")

    cur.close()
    conn.close()


if __name__ == "__main__":
    main()
