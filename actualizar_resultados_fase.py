# -*- coding: utf-8 -*-
"""
actualizar_resultados_fase.py <fase> [--apply]
Actualiza los ITEMS DE RESULTADO OFICIAL de los partidos en la BD (tabla partido)
para que sean IDENTICOS al Excel (hoja '40- RESULTADOS OFICIALES').

fase = grupos | r32 | octavos | cuartos | semis | todas   (default: semis)

Escribe SIEMPRE: goles_local, goles_visitante, amarillas(J), rojas(K),
                 decisiones_var(L), penales_partido(M), minuto_primer_gol(N),
                 estado='finalizado'.
Solo en KO (r32..semis) ademas: penales_local/visitante (tanda O) y equipo_clasificado_id (P).
En grupos NO toca tanda ni equipo_clasificado_id (el pase es por tabla, no por partido).

Reglas: tanda 99 = sin tanda -> NULL ; N=0 (sin gol) -> NULL.

SIN --apply: DRY RUN (muestra que cambiaria, no escribe).
CON --apply: escribe en la BD.

NOTA: no modifica el item P (escala por fase, decision de reglamento) ni recalcula
      puntajes. Tras --apply, correr run_recalc_hasta_semis.bat para re-puntuar.

Uso:
  backend\.venv\Scripts\python.exe actualizar_resultados_fase.py todas
  backend\.venv\Scripts\python.exe actualizar_resultados_fase.py todas --apply
"""
import sys, os

PHASES = {
    'grupos':  ('10- GRUPOS',        1,  72, False),
    'r32':     ('20- DIECISEISAVOS', 73,  88, True),
    'octavos': ('30- OCTAVOS',       89,  96, True),
    'cuartos': ('40- CUARTOS',       97, 100, True),
    'semis':   ('50- SEMIFINAL',    101, 102, True),
}
ORDER = ['grupos', 'r32', 'octavos', 'cuartos', 'semis']
args = [a.lower() for a in sys.argv[1:]]
DO_APPLY = '--apply' in args
sel = next((a for a in args if a in PHASES or a in ('todas','all','todo')), 'semis')
fases_run = ORDER if sel in ('todas','all','todo') else [sel]
print(f"Fases: {fases_run}   {'[APPLY]' if DO_APPLY else '[DRY RUN]'}")

BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = None
for f in os.listdir(BASE):
    if 'SEMIFINAL' in f.upper() and f.endswith('.xlsx'):
        EXCEL_FILE = os.path.join(BASE, f); break
if not EXCEL_FILE:
    sys.exit(f"ERROR: no se encontro Excel *SEMIFINAL*.xlsx en {BASE}")
print(f"Excel: {os.path.basename(EXCEL_FILE)}")

try:
    import openpyxl
except ImportError:
    os.system(f'"{sys.executable}" -m pip install openpyxl --quiet'); import openpyxl
try:
    import psycopg2, psycopg2.extras
except ImportError:
    os.system(f'"{sys.executable}" -m pip install psycopg2-binary --quiet'); import psycopg2, psycopg2.extras

CONN_BEC = "host=localhost port=5432 dbname=becbuc user=app_user password=superpassword"
TORNEO_ID = 2
SHEET = '40- RESULTADOS OFICIALES'

conn = psycopg2.connect(CONN_BEC); conn.autocommit = False
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

cur.execute("SELECT id, nombre, nombre_es FROM equipo")
eid = {}; ename = {}
for eq in cur.fetchall():
    ename[eq['id']] = eq['nombre'] or eq['nombre_es']
    if eq['nombre']:    eid[eq['nombre'].upper().strip()] = eq['id']
    if eq['nombre_es']: eid[eq['nombre_es'].upper().strip()] = eq['id']
ALIAS = {'FRANCIA':'France','ESPANA':'Spain','INGLATERRA':'England','ARGENTINA':'Argentina',
    'MARRUECOS':'Morocco','BELGICA':'Belgium','NORUEGA':'Norway','SUIZA':'Switzerland',
    'COLOMBIA':'Colombia','MEXICO':'Mexico','BRASIL':'Brazil','PORTUGAL':'Portugal',
    'PARAGUAY':'Paraguay','CANADA':'Canada','EGIPTO':'Egypt','ESTADOS UNIDOS':'USA',
    'EE UU':'USA','EEUU':'USA','ALEMANIA':'Germany'}
def find_eid(nom):
    if not nom: return None
    k = str(nom).upper().strip()
    if k in eid: return eid[k]
    a = ALIAS.get(k)
    if a and a.upper() in eid: return eid[a.upper()]
    for kk, vv in eid.items():
        if k in kk or kk in k: return vv
    return None
def to_int(v):
    try:
        s = str(v).strip()
        if s in ('','-','None'): return None
        return int(float(s))
    except: return None
def tanda(v):
    x = to_int(v); return None if x == 99 else x
def minuto(v):
    x = to_int(v); return None if x == 0 else x

COL = dict(pid=1, gl=12, gv=14, J=30, K=31, L=32, M=33, N=34, tl=35, tv=36, pasa=37)
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb[SHEET]
excel = {}
for r in range(2, ws.max_row + 1):
    pid = ws.cell(r, COL['pid']).value
    if not (isinstance(pid, str) and pid.startswith('P')):
        continue
    excel[pid] = {
        'goles_local': to_int(ws.cell(r, COL['gl']).value),
        'goles_visitante': to_int(ws.cell(r, COL['gv']).value),
        'amarillas': to_int(ws.cell(r, COL['J']).value),
        'rojas': to_int(ws.cell(r, COL['K']).value),
        'decisiones_var': to_int(ws.cell(r, COL['L']).value),
        'penales_partido': to_int(ws.cell(r, COL['M']).value),
        'minuto_primer_gol': minuto(ws.cell(r, COL['N']).value),
        'penales_local': tanda(ws.cell(r, COL['tl']).value),
        'penales_visitante': tanda(ws.cell(r, COL['tv']).value),
        'equipo_clasificado_id': find_eid(ws.cell(r, COL['pasa']).value),
        '_pasa_txt': ws.cell(r, COL['pasa']).value,
    }

CAMPOS_BASE = ['goles_local','goles_visitante','amarillas','rojas','decisiones_var',
               'penales_partido','minuto_primer_gol']
CAMPOS_KO   = ['penales_local','penales_visitante','equipo_clasificado_id']

print("\n" + "="*72)
print("ACTUALIZAR RESULTADOS OFICIALES en BD desde Excel")
print("="*72)

to_update = []      # (pid, partido_id, campos, valores)
cambios_por_campo = {}
total_cambios = 0

for fase in fases_run:
    fase_txt, nf_min, nf_max, es_ko = PHASES[fase]
    campos = CAMPOS_BASE + (CAMPOS_KO if es_ko else [])
    cur.execute("""
        SELECT p.numero_fifa, p.id, p.estado,
               p.goles_local, p.goles_visitante, p.amarillas, p.rojas, p.decisiones_var,
               p.penales_partido, p.minuto_primer_gol, p.penales_local, p.penales_visitante,
               p.equipo_clasificado_id
        FROM partido p JOIN fase f ON f.id=p.fase_id
        WHERE f.torneo_id=%s AND p.numero_fifa BETWEEN %s AND %s
        ORDER BY p.numero_fifa
    """, (TORNEO_ID, nf_min, nf_max))
    bd = {f"P{r['numero_fifa']:03d}": dict(r) for r in cur.fetchall()}
    print(f"\n--- {fase.upper()} ({fase_txt}) ---")
    fase_cambios = 0
    for n in range(nf_min, nf_max + 1):
        pid = f"P{n:03d}"
        ex = excel.get(pid); db = bd.get(pid)
        if not ex or not db:
            continue
        if ex['goles_local'] is None and ex['goles_visitante'] is None:
            continue  # sin resultado en Excel, no tocar
        difs = [(c, db[c], ex[c]) for c in campos if db[c] != ex[c]]
        estado_cambia = db['estado'] != 'finalizado'
        if not difs and not estado_cambia:
            continue
        for c, old, new in difs:
            extra = ''
            if c == 'equipo_clasificado_id':
                extra = f"  ({ename.get(old,old)} -> {ex['_pasa_txt']})"
            print(f"  {pid} {c:<22} {str(old):<8} -> {str(new):<8}{extra}")
            cambios_por_campo[c] = cambios_por_campo.get(c, 0) + 1
            total_cambios += 1; fase_cambios += 1
        if estado_cambia:
            print(f"  {pid} {'estado':<22} {db['estado']:<8} -> finalizado")
        to_update.append((pid, db['id'], campos, {c: ex[c] for c in campos}))
    if fase_cambios == 0:
        print("  (sin cambios)")

print("\n" + "-"*72)
print(f"Partidos a actualizar: {len(to_update)}   Campos que cambian: {total_cambios}")
if cambios_por_campo:
    print("Cambios por campo:")
    for c, k in sorted(cambios_por_campo.items(), key=lambda x:-x[1]):
        print(f"   {c}: {k}")

if not DO_APPLY:
    print("\n[DRY RUN] No se escribio nada. Para aplicar:")
    print(f"   actualizar_resultados_fase.py {sel} --apply")
    conn.close(); sys.exit(0)

print("\nAPLICANDO cambios...")
for pid, partido_id, campos, vals in to_update:
    set_sql = ", ".join(f"{c}=%({c})s" for c in campos) + ", estado='finalizado'"
    cur.execute(f"UPDATE partido SET {set_sql} WHERE id=%(id)s", {**vals, 'id': partido_id})
conn.commit()
print(f"OK. {len(to_update)} partidos actualizados.")
print("\nAHORA recalcular puntajes: run_recalc_hasta_semis.bat")
print("(las fases estan bloqueadas; ese script las desbloquea, recalcula y restaura el bloqueo)")
conn.close()
