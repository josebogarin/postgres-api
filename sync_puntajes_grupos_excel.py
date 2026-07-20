"""
sync_puntajes_grupos_excel.py
Copia los puntajes de grupos del Excel directamente a puntaje_detalle en BD.
Columnas Excel -> puntaje_detalle:
  A[29]=pts_resultado  B[30]=pts_marcador   C[31]=pts_amarillas
  D[32]=pts_rojas      E[33]=pts_var        F[34]=pts_penales_partido
  G[35]=pts_minuto

Ejecutar: & "backend\.venv\Scripts\python.exe" sync_puntajes_grupos_excel.py
"""
import sys, os, glob

BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = None
for f in os.listdir(BASE):
    if ('16AVOS' in f.upper() or 'CONSOLIDADOS' in f.upper()) and f.endswith('.xlsx'):
        EXCEL_FILE = os.path.join(BASE, f)
        break

if not EXCEL_FILE:
    print("ERROR: Excel no encontrado en", BASE); sys.exit(1)

print(f"Excel: {os.path.basename(EXCEL_FILE)}")

import openpyxl
import psycopg2, psycopg2.extras

CONN_BEC = "host=localhost port=5432 dbname=becbuc user=app_user password=superpassword"
CONN_APP = "host=localhost port=5432 dbname=app_db  user=app_user password=superpassword"
TORNEO_ID = 2

conn_bec = psycopg2.connect(CONN_BEC)
conn_app = psycopg2.connect(CONN_APP)
cur_bec  = conn_bec.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
cur_app  = conn_app.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# Usuarios app_db
cur_app.execute("SELECT id, username FROM users WHERE is_active=TRUE")
all_users = cur_app.fetchall()
user_by_ulow = {u['username'].lower(): u['id'] for u in all_users}

# IDs con apuestas en grupos
cur_bec.execute("""
    SELECT DISTINCT a.apostador_id FROM apuesta a
    JOIN partido p ON p.id=a.partido_id
    JOIN fase f ON f.id=p.fase_id
    WHERE f.torneo_id=%s AND f.tipo ILIKE '%%grupo%%'
""", (TORNEO_ID,))
bec_ids = {r['apostador_id'] for r in cur_bec.fetchall()}
bd_users = [u for u in all_users if u['id'] in bec_ids]

# Mapeo P001-P072 -> partido.id
cur_bec.execute("""
    SELECT p.id, p.numero_fifa FROM partido p
    JOIN fase f ON f.id=p.fase_id
    WHERE f.torneo_id=%s AND p.numero_fifa BETWEEN 1 AND 72
    ORDER BY p.numero_fifa
""", (TORNEO_ID,))
partidos_gr = {f"P{r['numero_fifa']:03d}": r['id'] for r in cur_bec.fetchall()}
print(f"Partidos grupos en BD: {len(partidos_gr)}")

def clean_alias(s):
    if not s: return ''
    return str(s).strip().upper().lstrip('@').replace('\xa0','')

apostador_to_id = {u['username'].lower(): u['id'] for u in bd_users}

def find_uid(alias_excel, nombre_excel=''):
    ac = clean_alias(alias_excel)
    for k, v in apostador_to_id.items():
        if k.upper().lstrip('@') == ac:
            return v
    for k, v in apostador_to_id.items():
        if ac and (ac in k.upper() or k.upper() in ac):
            return v
    return None

# Leer Excel
print("Leyendo Excel grupos...")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb['pronosticos']

updates = []  # list of (apostador_id, partido_id, A, B, C, D, E, F, G)
unmatched = set()

for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0] or row[3] != '10- GRUPOS': continue
    alias  = str(row[6]).strip() if row[6] else ''
    pid_str = str(row[1]).strip() if row[1] else ''
    uid = find_uid(alias)
    if not uid:
        unmatched.add(alias)
        continue
    pid = partidos_gr.get(pid_str)
    if not pid:
        continue
    pts = [int(row[i] or 0) for i in range(29, 36)]  # A..G
    updates.append((uid, pid, *pts))

print(f"Filas a actualizar: {len(updates)}")
if unmatched:
    print(f"Sin match: {unmatched}")

# Actualizar puntaje_detalle
updated = 0
inserted = 0
errors = 0

for (aid, pid, pa, pb, pc, pd_val, pe, pf, pg) in updates:
    try:
        # Intentar UPDATE primero
        cur_bec.execute("""
            UPDATE puntaje_detalle SET
                pts_resultado        = %s,
                pts_marcador         = %s,
                pts_amarillas        = %s,
                pts_rojas            = %s,
                pts_var              = %s,
                pts_penales_partido  = %s,
                pts_minuto           = %s
            WHERE apostador_id=%s AND partido_id=%s
        """, (pa, pb, pc, pd_val, pe, pf, pg, aid, pid))
        if cur_bec.rowcount == 0:
            # No existe la fila, hacer INSERT
            cur_bec.execute("""
                INSERT INTO puntaje_detalle
                  (apostador_id, partido_id, torneo_id,
                   pts_resultado, pts_marcador, pts_amarillas,
                   pts_rojas, pts_var, pts_penales_partido, pts_minuto)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (aid, pid, TORNEO_ID, pa, pb, pc, pd_val, pe, pf, pg))
            inserted += 1
        else:
            updated += 1
    except Exception as e:
        print(f"  ERROR aid={aid} pid={pid}: {e}")
        conn_bec.rollback()
        errors += 1

conn_bec.commit()
print(f"\nUpdated: {updated} | Inserted: {inserted} | Errors: {errors}")

# Verificar totales post-update
print("\n=== VERIFICACION POST-UPDATE ===")
cur_bec.execute("""
    SELECT pd.apostador_id,
           SUM(COALESCE(pts_resultado,0)+COALESCE(pts_marcador,0)+
               COALESCE(pts_amarillas,0)+COALESCE(pts_rojas,0)+
               COALESCE(pts_var,0)+COALESCE(pts_penales_partido,0)+
               COALESCE(pts_minuto,0)) AS total_bd
    FROM puntaje_detalle pd
    JOIN partido p ON p.id=pd.partido_id
    JOIN fase f ON f.id=p.fase_id
    WHERE f.torneo_id=%s AND f.tipo ILIKE '%%grupo%%'
    GROUP BY pd.apostador_id
""", (TORNEO_ID,))
bd_totals = {r['apostador_id']: int(r['total_bd'] or 0) for r in cur_bec.fetchall()}

# Totales Excel
excel_totals = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0] or row[3] != '10- GRUPOS': continue
    alias = clean_alias(row[6])
    excel_totals[alias] = excel_totals.get(alias, 0) + int(row[36] or 0)

print(f"\n{'ALIAS':<25} {'EXCEL':>7} {'BD':>7} {'DIFF':>7}")
print("-" * 50)
diffs = 0
for alias, ex in sorted(excel_totals.items(), key=lambda x:-x[1]):
    uid = find_uid(alias)
    bd  = bd_totals.get(uid, 0) if uid else 0
    d   = ex - bd
    if abs(d) > 0: diffs += 1
    mark = " DIFF!" if abs(d) > 0 else ""
    print(f"  {alias:<23} {ex:>7} {bd:>7} {d:>+7}{mark}")

if diffs == 0:
    print("\n100% coinciden!")
else:
    print(f"\n{diffs} apostadores con diferencia")

cur_bec.close(); cur_app.close()
conn_bec.close(); conn_app.close()
print("Listo.")
