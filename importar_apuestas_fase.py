# -*- coding: utf-8 -*-
"""
importar_apuestas_fase.py <fase> [--import]
Importa las apuestas (predicciones) por apostador de una fase KO desde el Excel
(hoja '50- TBL MASTER'), haciendo upsert en la tabla apuesta.

fase = semis | final3p    (default: final3p)
  semis    -> '50- SEMIFINAL'   P101-P102
  final3p  -> '60- FINAL Y 3P'  P103 (3er puesto) + P104 (final)

Estructura hoja '50- TBL MASTER' (1-based):
  col 2  = ID PARTIDO   col 7  = FASE   col 9 = NOMBRE   col 10 = ALIAS
  col 13 = pred goles local   col 15 = pred goles visitante
  col 25 = J  col 26 = K  col 27 = L  col 28 = M(pen.juego)  col 29 = N(1er gol)
  col 30 = O-tanda EQ1(local)  col 31 = P-tanda EQ2(visit)   col 32 = Q-quien clasifica

Filas con marcador vacio (apostador no cargo) se saltan.

Uso:
  backend\.venv\Scripts\python.exe importar_apuestas_fase.py final3p            <- DRY RUN
  backend\.venv\Scripts\python.exe importar_apuestas_fase.py final3p --import   <- ESCRIBE
"""
import sys, os

PHASES = {
    'semis':   ('50- SEMIFINAL',   101, 102),
    'final3p': ('60- FINAL Y 3P',  103, 104),
}
args = [a.lower() for a in sys.argv[1:]]
DO_IMPORT = '--import' in args
FASE = next((a for a in args if a in PHASES), 'final3p')
FASE_TXT, NF_MIN, NF_MAX = PHASES[FASE]
print(f"Fase: {FASE}  ({FASE_TXT}, P{NF_MIN:03d}-P{NF_MAX:03d})  {'[IMPORT]' if DO_IMPORT else '[DRY RUN]'}")

BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = None
for f in os.listdir(BASE):
    if 'SEMIFINAL' in f.upper() and f.endswith('.xlsx'):
        EXCEL_FILE = os.path.join(BASE, f); break
if not EXCEL_FILE:
    sys.exit(f"ERROR: no se encontro Excel *SEMIFINAL*.xlsx en {BASE}")
print(f"Excel: {os.path.basename(EXCEL_FILE)}")

try:
    import openpyxl
except ImportError:
    os.system(f'"{sys.executable}" -m pip install openpyxl --quiet'); import openpyxl
try:
    import psycopg2, psycopg2.extras
except ImportError:
    os.system(f'"{sys.executable}" -m pip install psycopg2-binary --quiet'); import psycopg2, psycopg2.extras

CONN_BEC = "host=localhost port=5432 dbname=becbuc user=app_user password=superpassword"
CONN_APP = "host=localhost port=5432 dbname=app_db  user=app_user password=superpassword"
TORNEO_ID = 2
SHEET = '50- TBL MASTER'

print("Conectando a BD...")
try:
    conn_bec = psycopg2.connect(CONN_BEC); conn_app = psycopg2.connect(CONN_APP)
except Exception as e:
    sys.exit(f"ERROR conexion: {e}\nDocker corriendo? docker start core-postgres")
cur_bec = conn_bec.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
cur_app = conn_app.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

cur_app.execute("SELECT id, username FROM users WHERE is_active=TRUE ORDER BY id")
all_users = cur_app.fetchall()
cur_bec.execute("""
    SELECT DISTINCT a.apostador_id FROM apuesta a
    JOIN partido p ON p.id = a.partido_id
    JOIN fase f ON f.id = p.fase_id
    WHERE f.torneo_id = %s
""", (TORNEO_ID,))
bec_ids = {r['apostador_id'] for r in cur_bec.fetchall()}
bd_apostadores = [u for u in all_users if u['id'] in bec_ids]
apostador_to_id = {u['username'].lower(): u['id'] for u in bd_apostadores}
print(f"Apostadores activos en BD: {len(bd_apostadores)}")

cur_bec.execute("SELECT id, nombre, nombre_es FROM equipo")
equipo_id_by_nombre = {}
for eq in cur_bec.fetchall():
    if eq['nombre']:    equipo_id_by_nombre[eq['nombre'].upper().strip()] = eq['id']
    if eq['nombre_es']: equipo_id_by_nombre[eq['nombre_es'].upper().strip()] = eq['id']
EQUIPO_ALIAS = {
    'FRANCIA':'France','ESPANA':'Spain','INGLATERRA':'England','ARGENTINA':'Argentina',
    'MARRUECOS':'Morocco','BELGICA':'Belgium','NORUEGA':'Norway','SUIZA':'Switzerland',
}
def find_equipo_id(nombre_excel):
    if not nombre_excel: return None
    key = str(nombre_excel).upper().strip()
    if key in equipo_id_by_nombre: return equipo_id_by_nombre[key]
    alt = EQUIPO_ALIAS.get(key)
    if alt and alt.upper().strip() in equipo_id_by_nombre:
        return equipo_id_by_nombre[alt.upper().strip()]
    for k, v in equipo_id_by_nombre.items():
        if key in k or k in key: return v
    return None

cur_bec.execute("""
    SELECT p.id, p.numero_fifa, el.nombre AS local, ev.nombre AS visit, p.estado
    FROM partido p JOIN fase f ON f.id = p.fase_id
    LEFT JOIN equipo el ON el.id = p.equipo_local_id
    LEFT JOIN equipo ev ON ev.id = p.equipo_visitante_id
    WHERE f.torneo_id = %s AND p.numero_fifa BETWEEN %s AND %s
    ORDER BY p.numero_fifa
""", (TORNEO_ID, NF_MIN, NF_MAX))
partidos_c = {f"P{r['numero_fifa']:03d}": dict(r) for r in cur_bec.fetchall()}
print(f"\nPartidos {FASE} en BD ({len(partidos_c)}):")
for k, v in sorted(partidos_c.items()):
    print(f"  {k} id={v['id']}: {v['local']} vs {v['visit']} [{v['estado']}]")

def clean_alias(s):
    if not s: return ''
    return str(s).replace('\xa0','').strip().upper().lstrip('@').replace('  ',' ')
def find_apostador_id(alias_excel):
    a = clean_alias(alias_excel)
    for k, v in apostador_to_id.items():
        if k.upper().lstrip('@') == a: return v
    for k, v in apostador_to_id.items():
        if a and (a in k.upper() or k.upper() in a): return v
    return None
def to_int(v):
    try:
        s = str(v).strip()
        if s in ('', '-', 'None'): return None
        return int(float(s))
    except: return None

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb[SHEET]
preds, unmatched, partidos_sin_match, sin_preds, clasifica_sin_match = [], {}, set(), [], set()

for r in range(2, ws.max_row + 1):
    if str(ws.cell(r, 7).value) != FASE_TXT:
        continue
    pid   = str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value else ''
    alias = ws.cell(r, 10).value
    uid = find_apostador_id(alias)
    if not uid:
        k = clean_alias(alias); unmatched[k] = unmatched.get(k, 0) + 1; continue
    pdb = partidos_c.get(pid)
    if not pdb:
        partidos_sin_match.add(pid); continue
    pred_l = to_int(ws.cell(r, 13).value)
    pred_v = to_int(ws.cell(r, 15).value)
    if pred_l is None and pred_v is None:
        sin_preds.append((clean_alias(alias), pid)); continue
    clasifica_nombre = ws.cell(r, 32).value
    clasifica_id = find_equipo_id(clasifica_nombre) if clasifica_nombre else None
    if clasifica_nombre and clasifica_id is None and str(clasifica_nombre).strip() not in ('-', ''):
        clasifica_sin_match.add(str(clasifica_nombre))
    preds.append({
        'apostador_id': uid, 'partido_id': pdb['id'],
        'pred_local': pred_l if pred_l is not None else 0,
        'pred_visitante': pred_v if pred_v is not None else 0,
        'pred_amarillas': to_int(ws.cell(r, 25).value),
        'pred_rojas': to_int(ws.cell(r, 26).value),
        'pred_var': to_int(ws.cell(r, 27).value),
        'pred_penales_partido': to_int(ws.cell(r, 28).value),
        'pred_minuto_gol': to_int(ws.cell(r, 29).value),
        'pred_penales_local_tanda': to_int(ws.cell(r, 30).value),
        'pred_penales_visitante_tanda': to_int(ws.cell(r, 31).value),
        'pred_equipo_clasifica': clasifica_id,
        '_alias': clean_alias(alias), '_partido': pid,
    })

esperado = 44 * (NF_MAX - NF_MIN + 1)
print(f"\nPronosticos resueltos: {len(preds)}  (maximo 44 x {NF_MAX-NF_MIN+1} = {esperado})")
print(f"Sin prediccion / no cargados (skip): {len(sin_preds)}")
print(f"Aliases sin match: {list(unmatched.keys()) if unmatched else 'ninguno OK'}")
if partidos_sin_match: print(f"Partidos no hallados en BD: {sorted(partidos_sin_match)}")
if clasifica_sin_match: print(f"'Quien clasifica' sin match equipo: {sorted(clasifica_sin_match)}")

if not DO_IMPORT:
    from collections import Counter
    print("\n[DRY RUN] Muestra (12 primeras):")
    for p in preds[:12]:
        alias_bd = next((u['username'] for u in bd_apostadores if u['id']==p['apostador_id']), str(p['apostador_id']))
        print(f"  {alias_bd:<20} {p['_partido']} {p['pred_local']}-{p['pred_visitante']}"
              f" J={p['pred_amarillas']} K={p['pred_rojas']} L={p['pred_var']} M={p['pred_penales_partido']}"
              f" N={p['pred_minuto_gol']} TL={p['pred_penales_local_tanda']} TV={p['pred_penales_visitante_tanda']}"
              f" Cls={p['pred_equipo_clasifica']}")
    porp = Counter(p['_partido'] for p in preds)
    print(f"\nPor partido: {dict(sorted(porp.items()))}")
    print(f"\nApostadores que NO cargaron ({len(sin_preds)}): {sorted(set(a for a,_ in sin_preds))}")
    print(f"\nPara ESCRIBIR en BD: importar_apuestas_fase.py {FASE} --import")
    sys.exit(0)

print(f"\nIMPORTANDO {len(preds)} apuestas de {FASE}...")
up, err = 0, 0
for pred in preds:
    try:
        cur_bec.execute("""
            INSERT INTO apuesta (
                apostador_id, partido_id, pred_local, pred_visitante,
                pred_amarillas, pred_rojas, pred_var, pred_penales_partido, pred_minuto_gol,
                pred_penales_local_tanda, pred_penales_visitante_tanda, pred_equipo_clasifica
            ) VALUES (
                %(apostador_id)s, %(partido_id)s, %(pred_local)s, %(pred_visitante)s,
                %(pred_amarillas)s, %(pred_rojas)s, %(pred_var)s, %(pred_penales_partido)s, %(pred_minuto_gol)s,
                %(pred_penales_local_tanda)s, %(pred_penales_visitante_tanda)s, %(pred_equipo_clasifica)s
            )
            ON CONFLICT (apostador_id, partido_id) DO UPDATE SET
                pred_local=EXCLUDED.pred_local, pred_visitante=EXCLUDED.pred_visitante,
                pred_amarillas=EXCLUDED.pred_amarillas, pred_rojas=EXCLUDED.pred_rojas,
                pred_var=EXCLUDED.pred_var, pred_penales_partido=EXCLUDED.pred_penales_partido,
                pred_minuto_gol=EXCLUDED.pred_minuto_gol,
                pred_penales_local_tanda=EXCLUDED.pred_penales_local_tanda,
                pred_penales_visitante_tanda=EXCLUDED.pred_penales_visitante_tanda,
                pred_equipo_clasifica=EXCLUDED.pred_equipo_clasifica
        """, pred)
        up += 1
    except Exception as e:
        err += 1; conn_bec.rollback()
        print(f"  ERROR {pred['_alias']} {pred['_partido']}: {e}")
conn_bec.commit()
print(f"\nImportadas: {up}   Errores: {err}")

cur_bec.execute("""
    SELECT COUNT(*) AS total, COUNT(DISTINCT a.apostador_id) AS aps, COUNT(DISTINCT a.partido_id) AS parts
    FROM apuesta a JOIN partido p ON p.id=a.partido_id
    WHERE p.numero_fifa BETWEEN %s AND %s
""", (NF_MIN, NF_MAX))
v = cur_bec.fetchone()
print(f"Verificacion BD {FASE}: {v['total']} apuestas ({v['aps']} apostadores x {v['parts']} partidos)")
print("\nListo. Recalcular cuando finalicen: POST /calcular-puntajes/2.")
conn_bec.close(); conn_app.close()
