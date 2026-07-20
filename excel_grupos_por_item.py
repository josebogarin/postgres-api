"""
Genera Excel: Puntaje por apostador x ítem — Fase de Grupos
Incluye: H I J K L M N O P (partidos) + D (Peor equipo, global)
Ejecutar con: python excel_grupos_por_item.py
"""
import sys, datetime
import psycopg2
import psycopg2.extras
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB = "postgresql://app_user:superpassword@localhost:5432/becbuc"

# ─── Colores ─────────────────────────────────────────────────────────────
HDR_BG   = "1e3a5f"   # azul oscuro — cabecera principal
HDR_FG   = "FFFFFF"
ITEM_BG  = "0f2840"   # azul más oscuro — fila de ítems
ITEM_FG  = "93c5fd"
AP_BG    = "0d1b2a"   # filas de apostadores (alternado)
AP_BG2   = "111f30"
POS_FG   = "fbbf24"   # dorado — pos/total
GLOB_BG  = "1a2e1a"   # verde oscuro — columna global D
GLOB_FG  = "4ade80"
TOTAL_BG = "292524"   # gris — total
TOTAL_FG = "fde68a"
ZERO     = "374151"   # texto pts = 0 (gris)

def col_hex(hex_str):
    return PatternFill("solid", fgColor=hex_str)

def thin():
    s = Side(style="thin", color="1e293b")
    return Border(left=s, right=s, top=s, bottom=s)

def main():
    conn = psycopg2.connect(DB.replace("postgresql://", "postgresql://"))
    cur  = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # ── 1. Apostadores con rol 'apostador' desde app_db ──────────────────
    # (se usa becbuc directamente; alias desde apuesta.nombre_apostador)
    # Primero sacamos todos los apostadores distintos que tienen apuestas en grupos
    cur.execute("""
        SELECT DISTINCT
            pd.apostador_id,
            COALESCE(a.nombre_apostador, a.apostador::text) AS alias
        FROM puntaje_detalle pd
        JOIN apuesta a ON a.id = pd.apuesta_id
        JOIN partido p ON p.id = pd.partido_id
        JOIN fase f ON f.id = p.fase_id
        WHERE f.torneo_id = 2
          AND f.tipo ILIKE 'grupo%'
        ORDER BY alias
    """)
    apostadores = cur.fetchall()

    if not apostadores:
        print("No se encontraron registros en fase de grupos.")
        conn.close()
        sys.exit(1)

    ap_ids = [r["apostador_id"] for r in apostadores]
    alias_map = {r["apostador_id"]: r["alias"] for r in apostadores}

    # ── 2. Puntaje por ítem agregado por apostador (solo grupos) ─────────
    id_list = ",".join(str(i) for i in ap_ids)
    cur.execute(f"""
        SELECT
            pd.apostador_id,
            COALESCE(SUM(pd.pts_resultado),0)              AS h,
            COALESCE(SUM(pd.pts_marcador),0)               AS i,
            COALESCE(SUM(pd.pts_amarillas),0)              AS j,
            COALESCE(SUM(COALESCE(pd.pts_rojas,0)),0)      AS k,
            COALESCE(SUM(pd.pts_var),0)                    AS l,
            COALESCE(SUM(COALESCE(pd.pts_penales_partido,0)),0) AS m,
            COALESCE(SUM(pd.pts_minuto),0)                 AS n,
            COALESCE(SUM(COALESCE(pd.pts_penales_tanda,0)),0)   AS o,
            COALESCE(SUM(COALESCE(pd.pts_equipo,0)),0)     AS p_ko,
            COUNT(DISTINCT pd.partido_id)                  AS partidos
        FROM puntaje_detalle pd
        JOIN partido pt ON pt.id = pd.partido_id
        JOIN fase f ON f.id = pt.fase_id
        WHERE f.torneo_id = 2
          AND f.tipo ILIKE 'grupo%'
          AND pd.apostador_id IN ({id_list})
        GROUP BY pd.apostador_id
    """)
    pts_rows = {r["apostador_id"]: dict(r) for r in cur.fetchall()}

    # ── 3. pts_grupos_p desde apostador_clasificados ─────────────────────
    cur.execute(f"""
        SELECT apostador_id,
               COALESCE(aciertos, 0) AS aciertos
        FROM apostador_clasificados
        WHERE torneo_id = 2
          AND fase_tipo = 'grupo'
          AND apostador_id IN ({id_list})
    """)
    grp_p = {r["apostador_id"]: r["aciertos"] for r in cur.fetchall()}

    # ── 4. Peor equipo (D) desde puntaje_global ───────────────────────────
    cur.execute(f"""
        SELECT apostador_id,
               COALESCE(pts_peor_equipo, 0) AS d
        FROM puntaje_global
        WHERE torneo_id = 2
          AND apostador_id IN ({id_list})
    """)
    glob_d = {r["apostador_id"]: r["d"] for r in cur.fetchall()}

    conn.close()

    # ── 5. Construir filas ────────────────────────────────────────────────
    ITEMS = ["H", "I", "J", "K", "L", "M", "N", "O", "P"]
    LABELS = {
        "H": "Resultado", "I": "Marcador exacto",
        "J": "Amarillas", "K": "Rojas", "L": "VAR",
        "M": "Penales juego", "N": "Minuto gol",
        "O": "Penales tanda", "P": "Clasifica R32"
    }

    def get_row(aid):
        p = pts_rows.get(aid, {})
        gp = grp_p.get(aid, 0)
        return {
            "H": p.get("h", 0),
            "I": p.get("i", 0),
            "J": p.get("j", 0),
            "K": p.get("k", 0),
            "L": p.get("l", 0),
            "M": p.get("m", 0),
            "N": p.get("n", 0),
            "O": p.get("o", 0),
            "P": p.get("p_ko", 0) + gp,
            "D": glob_d.get(aid, 0),
        }

    rows_data = []
    for aid in ap_ids:
        r = get_row(aid)
        r["alias"] = alias_map[aid]
        r["aid"]   = aid
        r["total_grupos"] = sum(r[k] for k in ITEMS)
        r["total_con_d"]  = r["total_grupos"] + r["D"]
        rows_data.append(r)

    # Ordenar por total_con_d desc
    rows_data.sort(key=lambda x: -x["total_con_d"])

    # ── 6. Excel ──────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Puntaje Grupos"
    ws.sheet_view.showGridLines = False

    # Fondo general oscuro
    ws.sheet_format.defaultRowHeight = 18

    # ── Cabecera fila 1 ───────────────────────────────────────────────────
    headers = ["#", "Apostador"] + ITEMS + ["🌐 D\nPeor equipo", "Total\nGrupos", "TOTAL\n(con D)"]
    col_map = {}  # header label → col index (1-based)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill      = col_hex(HDR_BG)
        cell.font      = Font(color=HDR_FG, bold=True, size=10, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin()
        col_map[h] = ci

    ws.row_dimensions[1].height = 30

    # Columnas D / Total grupos / TOTAL con D — colores especiales
    d_col     = headers.index("🌐 D\nPeor equipo") + 1
    tg_col    = headers.index("Total\nGrupos") + 1
    total_col = headers.index("TOTAL\n(con D)") + 1

    ws.cell(row=1, column=d_col).fill = col_hex(GLOB_BG)
    ws.cell(row=1, column=d_col).font = Font(color=GLOB_FG, bold=True, size=10, name="Calibri")
    ws.cell(row=1, column=total_col).fill = col_hex(TOTAL_BG)
    ws.cell(row=1, column=total_col).font = Font(color=TOTAL_FG, bold=True, size=11, name="Calibri")

    # ── Fila 2: sub-labels de ítems ───────────────────────────────────────
    sub = ["", ""] + [LABELS[k] for k in ITEMS] + ["", "", ""]
    for ci, s in enumerate(sub, 1):
        cell = ws.cell(row=2, column=ci, value=s)
        cell.fill      = col_hex(ITEM_BG)
        cell.font      = Font(color=ITEM_FG, size=8, italic=True, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin()
    ws.row_dimensions[2].height = 14

    # ── Filas de datos ────────────────────────────────────────────────────
    for ri, row in enumerate(rows_data, 3):
        bg = AP_BG if (ri % 2 == 1) else AP_BG2

        def wc(ci, value, *, bold=False, color="C8D3E0", bg_ovr=None, align="center", size=10):
            c = ws.cell(row=ri, column=ci, value=value)
            c.fill      = col_hex(bg_ovr or bg)
            c.font      = Font(color=color if value != 0 else ZERO,
                               bold=bold, size=size, name="Calibri")
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border    = thin()
            return c

        # Pos
        wc(1, ri - 2, bold=True, color=POS_FG)
        # Alias
        wc(2, row["alias"], align="left", bold=True, color="e2e8f0")

        # Ítems H-P
        for ki, k in enumerate(ITEMS, 3):
            v = row[k]
            clr = "4ade80" if v > 0 else ZERO
            if k in ("H", "I"):
                clr = "38bdf8" if v > 0 else ZERO   # azul para resultado/marcador
            wc(ki, v, color=clr)

        # D (Peor equipo)
        v_d = row["D"]
        wc(d_col, v_d, bg_ovr=GLOB_BG, color=GLOB_FG if v_d > 0 else ZERO, bold=(v_d > 0))

        # Total grupos (sin D)
        wc(tg_col, row["total_grupos"], color="94a3b8", bold=False)

        # TOTAL con D
        wc(total_col, row["total_con_d"], bg_ovr=TOTAL_BG,
           color=TOTAL_FG, bold=True, size=11)

        ws.row_dimensions[ri].height = 18

    # ── Fila de totales (suma por columna) ───────────────────────────────
    tot_row = len(rows_data) + 3
    ws.cell(row=tot_row, column=1, value="∑")
    ws.cell(row=tot_row, column=2, value="TOTAL TODOS").font = Font(color="fbbf24", bold=True, name="Calibri")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=tot_row, column=c)
        cell.fill   = col_hex("1c1917")
        cell.font   = Font(color="fbbf24", bold=True, size=10, name="Calibri")
        cell.border = thin()
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for ki, k in enumerate(ITEMS, 3):
        ws.cell(row=tot_row, column=ki,
                value=sum(r[k] for r in rows_data))

    ws.cell(row=tot_row, column=d_col,
            value=sum(r["D"] for r in rows_data))
    ws.cell(row=tot_row, column=tg_col,
            value=sum(r["total_grupos"] for r in rows_data))
    ws.cell(row=tot_row, column=total_col,
            value=sum(r["total_con_d"] for r in rows_data))
    ws.cell(row=tot_row, column=2, value="TOTAL TODOS")

    # ── Anchos de columna ─────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 22
    item_cols = [get_column_letter(i) for i in range(3, len(ITEMS) + 3)]
    for lc in item_cols:
        ws.column_dimensions[lc].width = 7
    ws.column_dimensions[get_column_letter(d_col)].width = 9
    ws.column_dimensions[get_column_letter(tg_col)].width = 9
    ws.column_dimensions[get_column_letter(total_col)].width = 10

    # ── Guardar ───────────────────────────────────────────────────────────
    ts  = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    out = f"C:/proyecto FAST API/puntaje_grupos_{ts}.xlsx"
    wb.save(out)
    print(f"\n✅ Excel generado: {out}")
    print(f"   Apostadores: {len(rows_data)}")
    print(f"   Columnas:    # | Apostador | H I J K L M N O P | D (Peor equipo) | Total Grupos | TOTAL")

if __name__ == "__main__":
    main()
