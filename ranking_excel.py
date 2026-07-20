"""
ranking_excel.py
================
Genera el Excel de ranking general de todos los apostadores BECBUC.
Formato: RANK | ALIAS | A | B | C | D | E | F | G | TOTAL PUNTOS
Colores: verde top tercio, amarillo medio, rojo bajo.

Uso:
    python ranking_excel.py [torneo_id] [output.xlsx]
"""

import sys, os, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

TORNEO_ID   = int(sys.argv[1]) if len(sys.argv) > 1 else 2
OUTPUT_PATH = sys.argv[2] if len(sys.argv) > 2 else None
if OUTPUT_PATH is None:
    script_dir  = os.path.dirname(os.path.abspath(__file__))
    OUTPUT_PATH = os.path.join(script_dir, "ranking_auditoria.xlsx")

PG_BEC = dict(host="localhost", port=5432, user="app_user",
              password="superpassword", dbname="becbuc")
PG_APP = dict(host="localhost", port=5432, user="app_user",
              password="superpassword", dbname="app_db")

try:
    import psycopg2, psycopg2.extras
except ImportError:
    sys.exit("ERROR: psycopg2 no disponible. Usar el venv del backend.")
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: openpyxl no disponible.")

def connect(cfg):
    return psycopg2.connect(**cfg, cursor_factory=psycopg2.extras.RealDictCursor)

def fetchall(conn, sql, params=None):
    with conn.cursor() as cur:
        cur.execute(sql, params or ())
        return cur.fetchall()

def fill(hex_color):
    return PatternFill("solid", fgColor="FF" + hex_color)

def bd(color="BFBFBF", thick=False):
    s = Side(style="medium" if thick else "thin", color="FF" + color)
    return Border(left=s, right=s, top=s, bottom=s)

AL = Alignment(horizontal="left",   vertical="center")
AC = Alignment(horizontal="center", vertical="center")

C_HEADER  = fill("1F3864")
C_GREEN   = fill("C6EFCE")
C_YELLOW  = fill("FFEB9C")
C_RED     = fill("FFC7CE")
C_TOTAL   = fill("1F3864")

F_WHITE_B = Font(bold=True,  color="FFFFFF", size=9)
F_DARK_B  = Font(bold=True,  color="1F3864", size=9)
F_DARK    = Font(bold=False, color="1F3864", size=9)
F_TOTAL   = Font(bold=True,  color="FFFFFF", size=10)

DASH = "-"

def val(v):
    return v if (v and v > 0) else DASH

# ── Query ranking ─────────────────────────────────────────────────────────────
SQL = """
WITH pd_agg AS (
    SELECT
        pd.apostador_id,
        SUM(COALESCE(pd.pts_resultado,       0))::int AS pts_h,
        SUM(COALESCE(pd.pts_marcador,        0))::int AS pts_i,
        SUM(COALESCE(pd.pts_amarillas,       0))::int AS pts_j,
        SUM(COALESCE(pd.pts_rojas,           0))::int AS pts_k,
        SUM(COALESCE(pd.pts_var,             0))::int AS pts_l,
        SUM(COALESCE(pd.pts_penales_partido, 0) +
            COALESCE(pd.pts_penales_tanda,   0))::int AS pts_f,
        SUM(COALESCE(pd.pts_minuto,          0))::int AS pts_n,
        SUM(COALESCE(pd.pts_total,           0))::int AS pts_part
    FROM puntaje_detalle pd
    JOIN partido p ON p.id = pd.partido_id
    JOIN fase    f ON f.id = p.fase_id
    WHERE f.torneo_id = %(tid)s
    GROUP BY pd.apostador_id
),
pg_agg AS (
    SELECT apostador_id, COALESCE(SUM(pts_total), 0)::int AS pts_glob
    FROM puntaje_global
    WHERE torneo_id = %(tid)s
    GROUP BY apostador_id
),
nombres AS (
    SELECT apostador_id, MAX(nombre_apostador) AS alias
    FROM apuesta a
    JOIN partido p ON p.id = a.partido_id
    JOIN fase    f ON f.id = p.fase_id
    WHERE f.torneo_id = %(tid)s
    GROUP BY apostador_id
)
SELECT
    COALESCE(n.alias, pd.apostador_id::text) AS alias,
    COALESCE(pd.pts_h,    0) AS pts_h,
    COALESCE(pd.pts_i,    0) AS pts_i,
    COALESCE(pd.pts_j,    0) AS pts_j,
    COALESCE(pd.pts_k,    0) AS pts_k,
    COALESCE(pd.pts_l,    0) AS pts_l,
    COALESCE(pd.pts_f,    0) AS pts_f,
    COALESCE(pd.pts_n,    0) AS pts_n,
    COALESCE(pd.pts_part, 0) + COALESCE(pg.pts_glob, 0) AS total
FROM pd_agg pd
LEFT JOIN pg_agg  pg ON pg.apostador_id = pd.apostador_id
LEFT JOIN nombres n  ON n.apostador_id  = pd.apostador_id
ORDER BY total DESC
"""

print("Consultando ranking …")
with connect(PG_BEC) as conn:
    ranking = fetchall(conn, SQL, {"tid": TORNEO_ID})

# Completar nombres vacíos desde app_db
uid_sin_nombre = [r for r in ranking if not r.get("alias") or str(r["alias"]).isdigit()]
if uid_sin_nombre:
    try:
        uid_ids = [int(r["alias"]) for r in uid_sin_nombre if str(r.get("alias","")).isdigit()]
        if uid_ids:
            with connect(PG_APP) as ca:
                name_rows = fetchall(ca,
                    f"SELECT id, COALESCE(nombre, username) AS alias FROM users "
                    f"WHERE id IN ({','.join(str(x) for x in uid_ids)})")
            name_map = {r["id"]: r["alias"] for r in name_rows}
            ranking = [{**r, "alias": name_map.get(int(r["alias"]), r["alias"])
                        if str(r.get("alias","")).isdigit() else r["alias"]}
                       for r in ranking]
    except Exception:
        pass

ranking = [dict(r, alias=str(r.get("alias") or "").upper()) for r in ranking]
print(f"  → {len(ranking)} apostadores encontrados")

# ── Excel ─────────────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Ranking general"
ws.freeze_panes = "A2"
ws.sheet_view.showGridLines = False

# Anchos
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 26
for c in range(3, 11):
    ws.column_dimensions[get_column_letter(c)].width = 13

HEADERS = [
    "RANK", "ALIAS",
    "A- GANA-EMPATA-\nPIERDE",
    "B- RESULTADO\nEXACTO",
    "C-\nAMARILLAS",
    "D- ROJAS",
    "E- VAR",
    "F- PENALES",
    "G- 1ER GOL",
    "TOTAL\nPUNTOS",
]

ws.row_dimensions[1].height = 40
for col, hdr in enumerate(HEADERS, 1):
    c = ws.cell(1, col, hdr)
    c.fill      = C_HEADER
    c.font      = F_WHITE_B
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = bd("FFFFFF", thick=True)

n = len(ranking)
top_cut = max(1, n // 3)
bot_cut = max(1, n - n // 3)

for row_idx, r in enumerate(ranking, 2):
    rank = row_idx - 1
    ws.row_dimensions[row_idx].height = 18

    bg = C_GREEN if rank <= top_cut else (C_RED if rank > bot_cut else C_YELLOW)

    row_vals = [
        rank,
        r["alias"],
        val(r["pts_h"]),
        val(r["pts_i"]),
        val(r["pts_j"]),
        val(r["pts_k"]),
        val(r["pts_l"]),
        val(r["pts_f"]),
        val(r["pts_n"]),
        r["total"] if r["total"] else DASH,
    ]
    for col, v in enumerate(row_vals, 1):
        c = ws.cell(row_idx, col, v)
        c.fill   = bg
        c.border = bd()
        c.font   = F_DARK_B if col in (1, 2, 10) else F_DARK
        c.alignment = AL if col == 2 else AC

# Fila totales por columna
tr = n + 2
ws.row_dimensions[tr].height = 20
tot_labels = ["", "TOTAL GENERAL",
              sum(r["pts_h"] for r in ranking),
              sum(r["pts_i"] for r in ranking),
              sum(r["pts_j"] for r in ranking),
              sum(r["pts_k"] for r in ranking),
              sum(r["pts_l"] for r in ranking),
              sum(r["pts_f"] for r in ranking),
              sum(r["pts_n"] for r in ranking),
              sum(r["total"] for r in ranking),
              ]
for col, v in enumerate(tot_labels, 1):
    c = ws.cell(tr, col, v)
    c.fill   = C_TOTAL
    c.font   = F_TOTAL
    c.border = bd("FFFFFF", thick=True)
    c.alignment = AL if col == 2 else AC

wb.save(OUTPUT_PATH)
print(f"\nExcel guardado en: {OUTPUT_PATH}")

# Imprimir top 5
print("\nTOP 5:")
for i, r in enumerate(ranking[:5], 1):
    print(f"  {i}. {r['alias']:<30} {r['total']:>4} pts")
