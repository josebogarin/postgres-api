# -*- coding: utf-8 -*-
r"""
diag_tanda_semis.py  (solo lectura)

Re-analiza el Excel de cierre para la TANDA de penales en las SEMIS (P101, P102)
y la Final/3P (P103, P104). Vuelca celda por celda:
  - Excel: prediccion tanda (col 30/31) y TODA la banda de puntos (col 40..49) con
    sus encabezados, para ubicar exactamente donde aparece el "4".
  - BD: resultado real de tanda (partido.penales_local/visitante + estado),
    prediccion tanda del apostador y pts_penales_tanda de puntaje_detalle.

Sirve para confirmar si el Excel esta otorgando puntos de tanda donde NO hubo
definicion por penales, y si mi lectura de columnas es correcta.

Uso:
  backend\.venv\Scripts\python.exe diag_tanda_semis.py
"""
import sys, os
BASE = os.path.dirname(os.path.abspath(__file__))
_OK = ('TORNEO CERRADO', 'FIN DE TORNEO', '20260720'); _BAD = ('CORRECCIONES', 'SEMIFINAL')
_c = [os.path.join(BASE, f) for f in os.listdir(BASE)
      if f.lower().endswith('.xlsx') and not f.startswith('~')
      and any(k in f.upper() for k in _OK) and not any(b in f.upper() for b in _BAD)]
if not _c:
    sys.exit("ERROR: no encontre el Excel de cierre en la raiz.")
EXCEL = sorted(_c, key=lambda p: os.path.getmtime(p), reverse=True)[0]
print(f"Excel: {os.path.basename(EXCEL)}")

try:
    import openpyxl
except ImportError:
    os.system(f'"{sys.executable}" -m pip install openpyxl --quiet'); import openpyxl
try:
    import psycopg2, psycopg2.extras
except ImportError:
    os.system(f'"{sys.executable}" -m pip install psycopg2-binary --quiet'); import psycopg2, psycopg2.extras

CONN_BEC = "host=localhost port=5432 dbname=becbuc user=app_user password=superpassword"
CONN_APP = "host=localhost port=5432 dbname=app_db  user=app_user password=superpassword"
TID = 2
FOCO = {'CHEREM', 'COTO', 'SONI'}
PARTIDOS = ['P101', 'P102', 'P103', 'P104']

def clean_alias(s):
    if not s: return ''
    return str(s).replace('\xa0', '').strip().upper().lstrip('@').replace('  ', ' ')

conn = psycopg2.connect(CONN_BEC); capp = psycopg2.connect(CONN_APP)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
cua = capp.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# BD: resultado real de tanda por partido
cur.execute("""SELECT p.numero_fifa, p.estado, p.goles_local, p.goles_visitante,
                      p.penales_local, p.penales_visitante
               FROM partido p JOIN fase f ON f.id=p.fase_id
               WHERE f.torneo_id=%s AND p.numero_fifa IN (101,102,103,104)
               ORDER BY p.numero_fifa""", (TID,))
real = {r['numero_fifa']: r for r in cur.fetchall()}
print("\n== BD: resultado real de esos partidos ==")
for nf in (101, 102, 103, 104):
    r = real.get(nf)
    if r:
        print(f"   P{nf}: estado={r['estado']}  marcador={r['goles_local']}-{r['goles_visitante']}  "
              f"tanda(penales_local/visit)={r['penales_local']}/{r['penales_visitante']}")

# BD: pred tanda + pts_penales_tanda por apostador (foco)
cua.execute("SELECT id, username FROM users WHERE is_active=TRUE")
uname = {u['username'].lower(): u['id'] for u in cua.fetchall()}
foco_ids = {u: i for u, i in uname.items() if u.upper() in FOCO}
print(f"\n== BD: apostadores foco -> ids {foco_ids} ==")
cur.execute("""SELECT a.apostador_id, p.numero_fifa,
                      a.pred_penales_local_tanda AS pl, a.pred_penales_visitante_tanda AS pv,
                      COALESCE(pd.pts_penales_tanda,0) AS pts
               FROM apuesta a JOIN partido p ON p.id=a.partido_id
               LEFT JOIN puntaje_detalle pd ON pd.partido_id=p.id AND pd.apostador_id=a.apostador_id
               WHERE p.numero_fifa IN (101,102,103,104)
                 AND a.apostador_id = ANY(%s)""", (list(foco_ids.values()),))
bd_foco = {(r['apostador_id'], r['numero_fifa']): r for r in cur.fetchall()}
id2u = {i: u for u, i in uname.items()}
for (aid, nf), r in sorted(bd_foco.items(), key=lambda x: (x[0][1], id2u.get(x[0][0], ''))):
    print(f"   {id2u.get(aid, aid):<10} P{nf}: pred_tanda={r['pl']}/{r['pv']}  pts_penales_tanda(BD)={r['pts']}")

# Excel: encabezados de la banda de puntos + celdas de foco
wb = openpyxl.load_workbook(EXCEL, data_only=True)
SHEET = next((s for s in wb.sheetnames if 'MASTER' in s.upper()), None)
ws = wb[SHEET]
print(f"\n== Excel hoja '{SHEET}': encabezados columnas 25..52 ==")
for c in range(25, 53):
    h = ws.cell(1, c).value
    if h is not None:
        print(f"   col {c:>2}: {h}")

print("\n== Excel: filas de P101..P104 para CHEREM/COTO/SONI (cols 28..49) ==")
print(f"   {'PART':<6}{'ALIAS':<10}" + "".join(f"c{c:>3}" for c in range(28, 50)))
hits = 0
for r in range(2, ws.max_row + 1):
    raw = ws.cell(r, 2).value
    pid = str(raw).strip() if raw is not None else ''
    if pid not in PARTIDOS:
        continue
    al = clean_alias(ws.cell(r, 10).value)
    if al not in FOCO:
        continue
    hits += 1
    vals = "".join(f"{('' if ws.cell(r,c).value is None else ws.cell(r,c).value)!s:>4}" for c in range(28, 50))
    print(f"   {pid:<6}{al:<10}{vals}")
if not hits:
    print("   (no encontre filas de foco; revisar alias/columna ID PARTIDO)")

# Excel: quien tiene puntos de tanda (>0) en P101/P102 segun col 47 y 48
print("\n== Excel: TODOS los apostadores con puntos de tanda (col47+col48) > 0 en P101/P102 ==")
def sc(v):
    try:
        s = str(v).strip()
        if s in ('', '-', 'None'): return 0
        return int(float(s))
    except Exception:
        return 0
cnt = 0
for r in range(2, ws.max_row + 1):
    raw = ws.cell(r, 2).value
    pid = str(raw).strip() if raw is not None else ''
    if pid not in ('P101', 'P102'):
        continue
    o1, o2 = sc(ws.cell(r, 47).value), sc(ws.cell(r, 48).value)
    if o1 + o2 > 0:
        cnt += 1
        print(f"   {pid}  {clean_alias(ws.cell(r,10).value):<12} col47={o1} col48={o2}  (pred col30/31="
              f"{ws.cell(r,30).value}/{ws.cell(r,31).value})")
if cnt == 0:
    print("   NINGUNO: el Excel de cierre NO otorga puntos de tanda en P101/P102 (col47/48 = 0).")

print("\nConclusion esperada: en P101/P102 la tanda real es None/None (sin definicion por")
print("penales) -> lo correcto es 0 pts de tanda. Si arriba aparece col47/48=4, el Excel")
print("esta mal ahi; la BD (pts_penales_tanda=0) es la correcta.")
conn.close(); capp.close()
