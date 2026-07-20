# -*- coding: utf-8 -*-
"""
comparar_items_resultados.py <fase>
Compara ITEM POR ITEM el RESULTADO OFICIAL de los partidos entre el Excel
(hoja '40- RESULTADOS OFICIALES') y la BD (tabla partido).

fase = grupos | r32 | octavos | cuartos | semis | todas   (default: semis)
  'todas' = grupos + r32 + octavos + cuartos + semis (modo compacto: solo diferencias)

Items: goles_local, goles_visitante, J-amarillas, K-rojas, L-VAR,
       M-penales(juego), N-minuto_1er_gol, y (solo KO) tanda EQ1/EQ2 + equipo que pasa.
Reglas: tanda 99 = sin tanda -> NULL ; N=0 (sin gol) -> NULL.
Solo lectura.

Uso:
  backend\.venv\Scripts\python.exe comparar_items_resultados.py todas
"""
import sys, os

# fase -> (texto_fase_excel, nf_min, nf_max, es_KO)
PHASES = {
    'grupos':  ('10- GRUPOS',        1,  72, False),
    'r32':     ('20- DIECISEISAVOS', 73,  88, True),
    'octavos': ('30- OCTAVOS',       89,  96, True),
    'cuartos': ('40- CUARTOS',       97, 100, True),
    'semis':   ('50- SEMIFINAL',    101, 102, True),
}
ORDER = ['grupos', 'r32', 'octavos', 'cuartos', 'semis']

arg = (sys.argv[1].lower() if len(sys.argv) > 1 else 'semis')
if arg in ('todas', 'all', 'todo'):
    fases_run = ORDER; COMPACT = True
elif arg in PHASES:
    fases_run = [arg]; COMPACT = False
else:
    sys.exit(f"fase invalida '{arg}'. Opciones: {', '.join(ORDER)}, todas")

BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = None
for f in os.listdir(BASE):
    if 'SEMIFINAL' in f.upper() and f.endswith('.xlsx'):
        EXCEL_FILE = os.path.join(BASE, f); break
if not EXCEL_FILE:
    sys.exit(f"ERROR: no se encontro Excel *SEMIFINAL*.xlsx en {BASE}")
print(f"Excel: {os.path.basename(EXCEL_FILE)}")

try:
    import openpyxl
except ImportError:
    os.system(f'"{sys.executable}" -m pip install openpyxl --quiet'); import openpyxl
try:
    import psycopg2, psycopg2.extras
except ImportError:
    os.system(f'"{sys.executable}" -m pip install psycopg2-binary --quiet'); import psycopg2, psycopg2.extras

CONN_BEC = "host=localhost port=5432 dbname=becbuc user=app_user password=superpassword"
TORNEO_ID = 2
SHEET = '40- RESULTADOS OFICIALES'

conn = psycopg2.connect(CONN_BEC)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

cur.execute("SELECT id, nombre, nombre_es FROM equipo")
equipo_id_by_nombre = {}; equipo_nombre_by_id = {}
for eq in cur.fetchall():
    equipo_nombre_by_id[eq['id']] = eq['nombre'] or eq['nombre_es']
    if eq['nombre']:    equipo_id_by_nombre[eq['nombre'].upper().strip()] = eq['id']
    if eq['nombre_es']: equipo_id_by_nombre[eq['nombre_es'].upper().strip()] = eq['id']
EQUIPO_ALIAS = {
    'FRANCIA':'France','ESPANA':'Spain','INGLATERRA':'England',
    'ARGENTINA':'Argentina','MARRUECOS':'Morocco','BELGICA':'Belgium','NORUEGA':'Norway',
    'SUIZA':'Switzerland','COLOMBIA':'Colombia','MEXICO':'Mexico','BRASIL':'Brazil',
    'PORTUGAL':'Portugal','PARAGUAY':'Paraguay','CANADA':'Canada','EGIPTO':'Egypt',
    'ESTADOS UNIDOS':'USA','EE UU':'USA','EEUU':'USA','ALEMANIA':'Germany',
}
def find_equipo_id(nombre_excel):
    if not nombre_excel: return None
    key = str(nombre_excel).upper().strip()
    if key in equipo_id_by_nombre: return equipo_id_by_nombre[key]
    alt = EQUIPO_ALIAS.get(key)
    if alt and alt.upper().strip() in equipo_id_by_nombre:
        return equipo_id_by_nombre[alt.upper().strip()]
    for k, v in equipo_id_by_nombre.items():
        if key in k or k in key: return v
    return None

def to_int(v):
    try:
        s = str(v).strip()
        if s in ('', '-', 'None'): return None
        return int(float(s))
    except: return None
def tanda(v):
    x = to_int(v); return None if x == 99 else x
def minuto(v):
    x = to_int(v); return None if x == 0 else x

COL = dict(pid=1, gl=12, gv=14, J=30, K=31, L=32, M=33, N=34, tl=35, tv=36, pasa=37)

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb[SHEET]
excel_all = {}
for r in range(2, ws.max_row + 1):
    pid = ws.cell(r, COL['pid']).value
    if not (isinstance(pid, str) and pid.startswith('P')):
        continue
    excel_all[pid] = {
        'goles_local': to_int(ws.cell(r, COL['gl']).value),
        'goles_visitante': to_int(ws.cell(r, COL['gv']).value),
        'amarillas': to_int(ws.cell(r, COL['J']).value),
        'rojas': to_int(ws.cell(r, COL['K']).value),
        'decisiones_var': to_int(ws.cell(r, COL['L']).value),
        'penales_partido': to_int(ws.cell(r, COL['M']).value),
        'minuto_primer_gol': minuto(ws.cell(r, COL['N']).value),
        'penales_local': tanda(ws.cell(r, COL['tl']).value),
        'penales_visitante': tanda(ws.cell(r, COL['tv']).value),
        'equipo_clasificado_id': find_equipo_id(ws.cell(r, COL['pasa']).value),
        '_pasa_txt': ws.cell(r, COL['pasa']).value,
    }

LABELS_BASE = [
    ('goles_local', 'Goles local (H)'),
    ('goles_visitante', 'Goles visit (H)'),
    ('amarillas', 'J- Amarillas'),
    ('rojas', 'K- Rojas'),
    ('decisiones_var', 'L- VAR'),
    ('penales_partido', 'M- Penales juego'),
    ('minuto_primer_gol', 'N- Minuto 1er gol'),
]
LABELS_KO = [
    ('penales_local', 'O- Tanda local'),
    ('penales_visitante', 'O- Tanda visit'),
    ('equipo_clasificado_id', 'P- Equipo que pasa'),
]

resumen = []
gran_total = 0

for fase in fases_run:
    fase_txt, nf_min, nf_max, es_ko = PHASES[fase]
    labels = LABELS_BASE + (LABELS_KO if es_ko else [])

    cur.execute("""
        SELECT p.numero_fifa, el.nombre AS local, ev.nombre AS visit, p.estado,
               p.goles_local, p.goles_visitante,
               p.amarillas, p.rojas, p.decisiones_var, p.penales_partido,
               p.minuto_primer_gol, p.penales_local, p.penales_visitante,
               p.equipo_clasificado_id
        FROM partido p JOIN fase f ON f.id=p.fase_id
        LEFT JOIN equipo el ON el.id=p.equipo_local_id
        LEFT JOIN equipo ev ON ev.id=p.equipo_visitante_id
        WHERE f.torneo_id=%s AND p.numero_fifa BETWEEN %s AND %s
        ORDER BY p.numero_fifa
    """, (TORNEO_ID, nf_min, nf_max))
    bd = {f"P{r['numero_fifa']:03d}": dict(r) for r in cur.fetchall()}

    print("\n" + "="*74)
    print(f"FASE {fase.upper()}  ({fase_txt}, P{nf_min:03d}-P{nf_max:03d})")
    print("="*74)

    fase_diffs = 0; partidos_con_diff = 0; n_part = 0
    for n in range(nf_min, nf_max + 1):
        pid = f"P{n:03d}"
        ex = excel_all.get(pid); db = bd.get(pid)
        if not ex or not db:
            if ex or db:
                print(f"  {pid}: {'falta en BD' if ex else 'falta en Excel'}")
            continue
        n_part += 1
        difs = []
        for campo, label in labels:
            ve = ex[campo]; vb = db[campo]
            if ve != vb:
                if campo == 'equipo_clasificado_id':
                    difs.append((label, ex['_pasa_txt'], equipo_nombre_by_id.get(vb, vb)))
                else:
                    difs.append((label, ve, vb))

        if not COMPACT:
            print(f"\n  {pid}: {db['local']} {db['goles_local']}-{db['goles_visitante']} {db['visit']} [{db['estado']}]")
            print(f"    {'ITEM':<20}{'EXCEL':<16}{'BD':<16}{'ESTADO'}")
            print("    " + "-"*56)
            for campo, label in labels:
                ve = ex[campo]; vb = db[campo]
                if campo == 'equipo_clasificado_id':
                    se = ex['_pasa_txt']; sb = equipo_nombre_by_id.get(vb, vb)
                    ok = (ve == vb)
                    print(f"    {label:<20}{str(se):<16}{str(sb):<16}{'OK' if ok else '<-- DIF'}")
                else:
                    ok = (ve == vb)
                    print(f"    {label:<20}{str(ve):<16}{str(vb):<16}{'OK' if ok else '<-- DIF'}")
        else:
            if difs:
                print(f"  {pid}: {db['local']} {db['goles_local']}-{db['goles_visitante']} {db['visit']}")
                for label, ve, vb in difs:
                    print(f"      {label:<20} excel={str(ve):<14} bd={str(vb)}")

        if difs:
            fase_diffs += len(difs); partidos_con_diff += 1

    if COMPACT and fase_diffs == 0:
        print("  (sin diferencias)")
    print(f"  --> {n_part} partidos | {fase_diffs} diferencia(s) en {partidos_con_diff} partido(s)")
    resumen.append((fase, n_part, fase_diffs, partidos_con_diff))
    gran_total += fase_diffs

print("\n" + "#"*74)
print("RESUMEN GENERAL (Excel vs BD)")
print("#"*74)
print(f"  {'FASE':<10}{'PARTIDOS':>10}{'DIFERENCIAS':>14}{'PART. C/DIF':>14}")
for fase, n_part, d, pc in resumen:
    print(f"  {fase:<10}{n_part:>10}{d:>14}{pc:>14}")
print("  " + "-"*46)
print(f"  {'TOTAL':<10}{sum(x[1] for x in resumen):>10}{gran_total:>14}{sum(x[3] for x in resumen):>14}")
if gran_total == 0:
    print("\n  RESULTADO: TODO IDENTICO. El Excel coincide 100% con la BD en todas las fases.")
else:
    print(f"\n  RESULTADO: {gran_total} diferencia(s) en total. Revisar el detalle arriba.")
conn.close()
